import streamlit as st
import sqlite3
import re
import io
import threading
import unicodedata
import json
import hashlib
import os
import pickle
import time
from datetime import datetime
from urllib.parse import quote
from openpyxl import load_workbook, Workbook

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(page_title="Equivalencias El Chavo", page_icon="🔧", layout="wide")


# ============================================================
# MODO DE VISTA (celular / computadora)
# ============================================================
# En vez de mantener dos archivos separados (que habría que corregir dos veces cada vez que
# se cambia algo, y terminarían desincronizándose), es la misma app con dos modos de vista.
# Cada uno acomoda la pantalla distinto: el celular apila las cosas en vertical y usa un
# selector compacto; la computadora aprovecha el ancho con columnas y pestañas en fila.
def es_celular():
    return st.session_state.get("modo_vista", "📱 Celular") == "📱 Celular"


def cols(pesos, apilar_en_celular=True):
    """Devuelve columnas como st.columns, pero en modo celular apila los elementos en vertical
    (uno abajo del otro, a ancho completo) en vez de aplastarlos de costado. Para filas de 2
    elementos simples se puede pasar apilar_en_celular=False y dejarlas lado a lado."""
    cantidad = len(pesos) if isinstance(pesos, (list, tuple)) else pesos
    if es_celular() and apilar_en_celular:
        return [st.container() for _ in range(cantidad)]
    return st.columns(pesos)


CSS_CUSTOM = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');

:root {
  --bg: #14171B;
  --bg-panel: #1D2126;
  --bg-panel-2: #242830;
  --border: #2D3239;
  --text: #ECEEF0;
  --text-muted: #9BA3AC;
  --accent: #E8A33D;
  --accent-hover: #F2B658;
  --accent-2: #4C8BF5;
}

html, body, [class*="css"] { font-family: 'Inter', -apple-system, sans-serif; }

[data-testid="stAppViewContainer"] {
  background: radial-gradient(circle at 15% 0%, #1A1E24 0%, var(--bg) 45%) fixed;
  color: var(--text);
}
[data-testid="stHeader"] { background: transparent; }

/* Encabezado tipo ficha/etiqueta de repuesto */
.app-header {
  display: flex; flex-direction: column; gap: 2px;
  padding: 18px 22px; margin-bottom: 10px;
  background: linear-gradient(135deg, var(--bg-panel) 0%, var(--bg-panel-2) 100%);
  border: 1px solid var(--border); border-left: 3px solid var(--accent);
  border-radius: 10px;
}
.app-header__eyebrow {
  font-family: 'IBM Plex Mono', monospace; font-size: 0.68rem;
  letter-spacing: 0.14em; color: var(--accent); text-transform: uppercase; margin: 0 0 4px 0;
}
.app-header h1 {
  font-family: 'Space Grotesk', sans-serif; font-weight: 700; font-size: 1.9rem;
  margin: 0; color: var(--text); letter-spacing: -0.01em;
}
.app-header p { margin: 4px 0 0 0; color: var(--text-muted); font-size: 0.92rem; }

/* Pestañas (nivel principal) */
.stTabs [data-baseweb="tab-list"] { gap: 4px; border-bottom: 1px solid var(--border); }
.stTabs [data-baseweb="tab"] {
  font-family: 'Space Grotesk', sans-serif; font-weight: 600; font-size: 0.86rem;
  color: var(--text-muted); background: transparent; border-radius: 8px 8px 0 0; padding: 10px 14px;
}
.stTabs [aria-selected="true"] { color: var(--accent) !important; border-bottom: 2px solid var(--accent) !important; }

/* Sub-pestañas anidadas (ej: dentro de Administrar o Modo Mecánico) — más chicas y sutiles,
   para que se note la jerarquía: esto es una subdivisión de la pestaña principal, no otra más. */
.stTabs .stTabs [data-baseweb="tab-list"] {
  border-bottom: 1px solid var(--border); gap: 2px; margin-top: 4px; margin-bottom: 8px;
}
.stTabs .stTabs [data-baseweb="tab"] {
  font-size: 0.78rem; padding: 7px 11px; color: var(--text-muted); opacity: 0.85;
}
.stTabs .stTabs [aria-selected="true"] { opacity: 1; }

/* Botones */
.stButton > button, .stDownloadButton > button, .stLinkButton > a, .stFormSubmitButton > button {
  border-radius: 8px; border: 1px solid var(--border); font-family: 'Inter', sans-serif;
  font-weight: 600; transition: all 0.15s ease;
}
.stButton > button[kind="primary"], .stFormSubmitButton > button[kind="primary"] {
  background: var(--accent); color: #1A1300; border: none;
}
.stButton > button[kind="primary"]:hover, .stFormSubmitButton > button[kind="primary"]:hover { background: var(--accent-hover); }
.stButton > button:hover { border-color: var(--accent); color: var(--accent); }
.stButton > button:disabled { opacity: 0.4; }

/* Radios usados como selector de modo (ej: Código/Descripción, Foto real/IA) —
   look de pastillas en vez del radio suelto por defecto, para que se sienta como
   un selector de vista, no como un formulario más. */
.stRadio [role="radiogroup"] { gap: 6px; flex-wrap: wrap; }
.stRadio label {
  background: var(--bg-panel); border: 1px solid var(--border); border-radius: 999px;
  padding: 5px 14px 5px 10px !important; transition: all 0.15s ease;
}
.stRadio label:has(input:checked) { border-color: var(--accent); background: rgba(232, 163, 61, 0.12); }

/* Inputs */
.stTextInput input, .stNumberInput input, .stTextArea textarea,
.stSelectbox div[data-baseweb="select"] > div {
  background: var(--bg-panel) !important; border: 1px solid var(--border) !important;
  border-radius: 7px !important; color: var(--text) !important;
}
.stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
  border-color: var(--accent) !important; box-shadow: 0 0 0 1px var(--accent) !important;
}
.stCheckbox input:checked, .stCheckbox [data-baseweb="checkbox"] svg { accent-color: var(--accent); }

/* Subida de archivos */
[data-testid="stFileUploaderDropzone"] {
  background: var(--bg-panel) !important; border: 1px dashed var(--border) !important; border-radius: 8px !important;
}

/* Expanders */
[data-testid="stExpander"] { border: 1px solid var(--border) !important; border-radius: 8px !important; margin-bottom: 4px; }
.streamlit-expanderHeader, [data-testid="stExpander"] summary {
  background: var(--bg-panel) !important; border-radius: 8px !important; font-weight: 600;
}

/* Métricas */
[data-testid="stMetric"] {
  background: var(--bg-panel); border: 1px solid var(--border); border-radius: 10px; padding: 10px 14px;
}
[data-testid="stMetricValue"] { font-family: 'IBM Plex Mono', monospace; color: var(--accent); }

/* Alertas y tablas */
[data-testid="stAlert"] { border-radius: 8px; border: 1px solid var(--border); }
[data-testid="stDataFrame"] { border: 1px solid var(--border); border-radius: 8px; overflow: hidden; }

code { font-family: 'IBM Plex Mono', monospace; color: var(--accent-2); }
hr { border-color: var(--border) !important; margin: 1.1rem 0 !important; }

/* Los gráficos (st.bar_chart / Vega-Lite) muestran un tooltip flotante al tocar una barra.
   En celular no hay evento que lo "suelte" al hacer scroll con el dedo, y queda pegado en
   pantalla tapando el contenido de abajo. Se desactiva: el dato ya se ve en las barras. */
#vg-tooltip-element, .vg-tooltip { display: none !important; }
</style>
"""
st.markdown(CSS_CUSTOM, unsafe_allow_html=True)

# Ajustes según el modo de vista elegido. En celular se agranda lo que hay que tocar con el
# dedo y se achica el texto de las tablas para que entre; en computadora se aprovecha el ancho.
if es_celular():
    st.markdown("""
    <style>
    .block-container { padding: 0.8rem 0.7rem 3rem 0.7rem !important; max-width: 100% !important; }
    .stButton > button, .stDownloadButton > button, .stLinkButton > a, .stFormSubmitButton > button {
      min-height: 2.7rem; width: 100%;
    }
    [data-testid="stDataFrame"] { font-size: 0.78rem; }
    .app-header h1 { font-size: 1.45rem !important; }
    .app-header p { font-size: 0.8rem !important; }
    [data-testid="stExpander"] summary { padding: 0.65rem 0.5rem !important; }
    h3 { font-size: 1.1rem !important; }
    /* Navegación en pastillas: compactas para que las 8 secciones entren en pocas filas
       sin comerse la pantalla, y con buen tamaño para tocar con el dedo. */
    .stRadio [role="radiogroup"] { gap: 4px; }
    .stRadio label { padding: 4px 10px 4px 6px !important; font-size: 0.82rem; }
    .stRadio label p { font-size: 0.82rem !important; }
    </style>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
    <style>
    .block-container { padding: 1.6rem 3rem 4rem 3rem !important; max-width: 1500px !important; }
    [data-testid="stDataFrame"] { font-size: 0.88rem; }
    </style>
    """, unsafe_allow_html=True)

DB_PATH = "equivalencias_app.db"

# La conexión a la base se comparte entre todas las personas que usan la app al mismo
# tiempo (así funciona el hosting gratuito). Este candado evita que dos operaciones
# (por ejemplo una importación larga y una búsqueda de otra persona) se pisen y
# dejen todo trabado.
db_lock = threading.Lock()

# Subir este número cuando se agreguen WMI nuevos: hace que la lista se vuelva a aplicar una vez
# sobre las bases que ya existen, sin pisar lo que el usuario haya corregido a mano.
SEMILLA_WMI_VERSION = "2"


def es_admin():
    return st.session_state.get("nivel_usuario") == "admin"


def es_operador_o_admin():
    """Para acciones que un empleado de confianza puede hacer (usar funciones de IA, agregar
    piezas a un esquema) sin necesitar la contraseña completa de administrador, que además
    desbloquea borrados y configuración sensible."""
    return st.session_state.get("nivel_usuario") in ("admin", "operador")


def hash_password(password, salt=None):
    """Nunca guardamos la contraseña en texto plano — se guarda un hash junto con una sal
    aleatoria distinta por usuario, para que ni siquiera dos personas con la misma clave
    tengan el mismo hash guardado."""
    if salt is None:
        salt = os.urandom(16).hex()
    h = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
    return h, salt


def crear_usuario(nombre, password, rol="operador"):
    h, salt = hash_password(password)
    with db_lock:
        c.execute("INSERT INTO usuarios (nombre, password_hash, salt, rol) VALUES (?, ?, ?, ?)",
                   (nombre.strip(), h, salt, rol))
        conn.commit()


def listar_usuarios():
    c.execute("""SELECT id AS "ID", nombre AS "Nombre", rol AS "Rol",
                 CASE WHEN activo=1 THEN 'Sí' ELSE 'No' END AS "Activo", creado_en AS "Creado"
                 FROM usuarios ORDER BY nombre""")
    return filas_a_listas(c)


def validar_password_usuario(password):
    c.execute("SELECT nombre, password_hash, salt, rol FROM usuarios WHERE activo = 1")
    for fila in c.fetchall():
        h, _ = hash_password(password, fila["salt"])
        if h == fila["password_hash"]:
            return fila["nombre"], fila["rol"]
    return None, None


def cambiar_password_usuario(usuario_id, nueva_password):
    h, salt = hash_password(nueva_password)
    with db_lock:
        c.execute("UPDATE usuarios SET password_hash=?, salt=? WHERE id=?", (h, salt, usuario_id))
        conn.commit()


def activar_desactivar_usuario(usuario_id, activo):
    with db_lock:
        c.execute("UPDATE usuarios SET activo=? WHERE id=?", (1 if activo else 0, usuario_id))
        conn.commit()


def eliminar_usuario(usuario_id):
    with db_lock:
        c.execute("DELETE FROM usuarios WHERE id=?", (usuario_id,))
        conn.commit()


def crear_mecanico(nombre, password):
    h, salt = hash_password(password)
    with db_lock:
        c.execute("INSERT INTO mecanicos (nombre, password_hash, salt) VALUES (?, ?, ?)",
                   (nombre.strip(), h, salt))
        conn.commit()


def listar_mecanicos():
    c.execute("""SELECT id AS "ID", nombre AS "Nombre",
                 CASE WHEN activo=1 THEN 'Sí' ELSE 'No' END AS "Activo", creado_en AS "Creado"
                 FROM mecanicos ORDER BY nombre""")
    return filas_a_listas(c)


def validar_password_mecanico(password):
    c.execute("SELECT id, nombre, password_hash, salt FROM mecanicos WHERE activo = 1")
    for fila in c.fetchall():
        h, _ = hash_password(password, fila["salt"])
        if h == fila["password_hash"]:
            return fila["id"], fila["nombre"]
    return None, None


def activar_desactivar_mecanico(mecanico_id, activo):
    with db_lock:
        c.execute("UPDATE mecanicos SET activo=? WHERE id=?", (1 if activo else 0, mecanico_id))
        conn.commit()


def eliminar_mecanico(mecanico_id):
    with db_lock:
        c.execute("DELETE FROM mecanicos WHERE id=?", (mecanico_id,))
        conn.commit()


def validar_password(clave):
    """Chequea la contraseña contra los secrets de admin/operador Y contra las cuentas creadas
    desde la propia app (tabla usuarios). Devuelve (nombre, nivel, error) — nivel es 'admin',
    'operador', o None si no matcheó ninguna. Los mecánicos externos se validan aparte, con
    validar_password_mecanico(), porque tienen su propio portal separado."""
    secretos = st.secrets if hasattr(st, "secrets") else {}
    # [admin_passwords] / [operador_passwords] en Streamlit Secrets, cada una con nombre:clave.
    # También soporta la forma anterior de una sola clave (admin_password) por compatibilidad.
    admin_passwords = dict(secretos.get("admin_passwords", {}))
    clave_unica = secretos.get("admin_password")
    if clave_unica:
        admin_passwords.setdefault("admin", clave_unica)
    operador_passwords = dict(secretos.get("operador_passwords", {}))

    nombre_admin = next((n for n, p in admin_passwords.items() if p == clave), None)
    if nombre_admin:
        return nombre_admin, "admin", None
    nombre_operador = next((n for n, p in operador_passwords.items() if p == clave), None)
    if nombre_operador:
        return nombre_operador, "operador", None

    # Cuentas creadas desde la propia app (además de las de Secrets)
    nombre_db, rol_db = validar_password_usuario(clave)
    if nombre_db:
        return nombre_db, rol_db, None

    if not admin_passwords and not operador_passwords:
        c.execute("SELECT COUNT(*) FROM usuarios")
        if c.fetchone()[0] == 0:
            return None, None, (
                "No configuraste todavía ninguna contraseña en Streamlit Cloud (Settings → Secrets) "
                "ni creaste ningún usuario desde la app. Sin eso, nadie puede entrar a las secciones protegidas."
            )
    return None, None, None


def pedir_password_admin(motivo=""):
    """Muestra un formulario de contraseña de ADMINISTRADOR COMPLETO. Devuelve True si ya está
    autenticado como admin — para borrados y configuración sensible, un 'operador' no alcanza."""
    if es_admin():
        return True

    st.warning(f"🔒 Esta sección está protegida{(' — ' + motivo) if motivo else ''}.")
    with st.form(f"login_admin_{motivo}"):
        clave = st.text_input("Contraseña de administrador:", type="password")
        entrar = st.form_submit_button("Ingresar")

    if entrar:
        nombre, nivel, error = validar_password(clave)
        if error:
            st.error(error)
        elif nivel == "admin":
            st.session_state.nivel_usuario = "admin"
            st.session_state.admin_nombre = nombre
            st.rerun()
        elif nivel == "operador":
            st.error("Esa es una contraseña de operador — para esto hace falta la de administrador completo.")
        else:
            st.error("Contraseña incorrecta.")
    return False


def mostrar_login_inicial():
    """Pide la contraseña apenas se abre la app, con opción de seguir sin loguearse para
    quien solo quiera buscar/consultar. Las acciones destructivas van a seguir pidiendo la
    contraseña de administrador completo aparte, esto es solo la pantalla de entrada."""
    st.markdown("### 👋 ¿Quién sos?")
    st.caption(
        "Poné tu nombre para que tus búsquedas recientes queden separadas de las de tus "
        "compañeros — el resto de la información (catálogo, esquemas, etc.) la ven todos igual. "
        "Es opcional, si lo dejás vacío vas a figurar como 'Invitado'."
    )
    with st.form("login_inicial"):
        nombre_usuario = st.text_input("Tu nombre:", placeholder="Ej: Matías", key="login_inicial_nombre")
        st.markdown("---")
        st.caption(
            "Si tenés contraseña (de administrador completo o de operador), ingresala acá. "
            "Un operador puede usar las funciones de IA y cargar cosas, pero no borrar ni configurar."
        )
        clave = st.text_input("Contraseña (opcional):", type="password", key="login_inicial_clave")
        col_a, col_b = st.columns(2)
        entrar = col_a.form_submit_button("🔓 Ingresar con contraseña", type="primary", use_container_width=True)
        seguir = col_b.form_submit_button("➡️ Continuar", use_container_width=True)

    if entrar:
        nombre, nivel, error = validar_password(clave)
        if nivel:
            st.session_state.nivel_usuario = nivel
            st.session_state.admin_nombre = nombre
            st.session_state.saltar_login = True
            st.rerun()
        else:
            mecanico_id, nombre_mecanico = validar_password_mecanico(clave)
            if mecanico_id:
                st.session_state.nivel_usuario = "mecanico"
                st.session_state.admin_nombre = nombre_mecanico
                st.session_state.mecanico_id = mecanico_id
                st.session_state.saltar_login = True
                st.rerun()
            elif error:
                st.error(error)
            else:
                st.error("Contraseña incorrecta.")
    if seguir:
        st.session_state.usuario_nombre = nombre_usuario.strip() or "Invitado"
        st.session_state.saltar_login = True
        st.rerun()


# ============================================================
# CONEXIÓN Y ESQUEMA
# ============================================================
ARCHIVO_SEMILLA = "datos_iniciales.db"


def _restaurar_desde_semilla(conexion):
    """Streamlit Cloud borra el disco de la app cada vez que se redespliega o se reinicia, así
    que la base de datos se pierde. Los archivos del REPOSITORIO, en cambio, sí sobreviven
    (son parte del despliegue). Entonces: si la base está vacía y en el repositorio hay una
    copia llamada 'datos_iniciales.db', se restaura sola al arrancar.
    Para actualizar esa copia: bajar el backup desde Estadísticas → Backup y config, y subir
    ese archivo al repositorio de GitHub con el nombre 'datos_iniciales.db'."""
    if not os.path.exists(ARCHIVO_SEMILLA):
        return False
    try:
        cur = conexion.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='productos'")
        if cur.fetchone():
            cur.execute("SELECT COUNT(*) FROM productos")
            if cur.fetchone()[0] > 0:
                return False  # ya hay datos cargados: no se toca nada
        origen = sqlite3.connect(ARCHIVO_SEMILLA)
        origen.backup(conexion)
        origen.close()
        return True
    except Exception:
        return False


@st.cache_resource
def get_connection():
    """Conexión única y persistente entre reruns de Streamlit."""
    # isolation_level=None => autocommit: cada sentencia se confirma sola.
    #
    # Es importante y no es un detalle técnico. La conexión es UNA SOLA compartida por todos los
    # que usan la app al mismo tiempo (así la deja @st.cache_resource). Con el modo por defecto,
    # Python abre una transacción implícita y la deja abierta hasta el commit, así que las
    # operaciones de dos personas se mezclan en la MISMA transacción:
    #   · si uno confirma, confirma también la importación a medio hacer del otro;
    #   · y peor: si uno cancela, se pierde el trabajo que el otro ya había guardado.
    # Con autocommit eso no puede pasar. Donde hace falta que varias sentencias sean una sola
    # cosa, se usa el bloque transaccion() de abajo.
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, isolation_level=None)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")  # mejor concurrencia / menos bloqueos
    conn.execute("PRAGMA busy_timeout = 8000")  # esperar en vez de fallar si otro está escribiendo
    # No hace falta agrupar las importaciones en una transacción explícita para compensar: se
    # midió la importación real de 10.000 filas en los dos modos y tarda lo mismo (0,33 s contra
    # 0,36 s), porque en modo WAL confirmar es barato.
    # Si el disco se borró (pasa al redesplegar), recuperar desde la copia del repositorio.
    # Va antes de crear las tablas: después las migraciones ponen al día el esquema.
    st.session_state["_restaurado_de_semilla"] = _restaurar_desde_semilla(conn)
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS marcas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE NOT NULL,
        tipo TEXT NOT NULL DEFAULT 'PROVEEDOR',
        url_ficha_template TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )""")
    columnas_marcas = [f[1] for f in c.execute("PRAGMA table_info(marcas)").fetchall()]
    if "url_ficha_template" not in columnas_marcas:
        c.execute("ALTER TABLE marcas ADD COLUMN url_ficha_template TEXT")

    c.execute("""CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_raw TEXT NOT NULL,
        codigo_clean TEXT NOT NULL,
        descripcion TEXT,
        marca_id INTEGER NOT NULL REFERENCES marcas(id) ON DELETE CASCADE,
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(codigo_clean, marca_id)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS equivalencias (
        producto_a_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        producto_b_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        created_at TEXT DEFAULT (datetime('now')),
        PRIMARY KEY (producto_a_id, producto_b_id)
    )""")

    # Qué repuesto le va a cada auto, según el catálogo del fabricante del repuesto.
    # Stock apartado al cotizar. Con varios atendiendo a la vez, vender dos veces la misma
    # pieza es cuestión de tiempo: uno cotiza 4 pastillas, el otro ve stock 4 y las vende.
    c.execute("""CREATE TABLE IF NOT EXISTS reservas_stock (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        cantidad INTEGER NOT NULL,
        cliente TEXT,
        nota TEXT,
        reservado_por TEXT,
        estado TEXT DEFAULT 'activa',
        fecha TEXT DEFAULT (datetime('now'))
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_reservas_prod ON reservas_stock(producto_id, estado)")

    c.execute("""CREATE TABLE IF NOT EXISTS aplicaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        marca_auto TEXT NOT NULL,
        modelo_auto TEXT NOT NULL,
        motor TEXT,
        combustible TEXT,
        anio_desde INTEGER,
        anio_hasta INTEGER,
        codigo TEXT NOT NULL,
        codigo_clean TEXT,
        marca_repuesto TEXT,
        tipo_pieza TEXT,
        origen TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE (marca_auto, modelo_auto, motor, anio_desde, anio_hasta, codigo)
    )""")
    _cols_aplic = [f[1] for f in c.execute("PRAGMA table_info(aplicaciones)").fetchall()]
    if "tipo_pieza" not in _cols_aplic:
        c.execute("ALTER TABLE aplicaciones ADD COLUMN tipo_pieza TEXT")
    c.execute("CREATE INDEX IF NOT EXISTS idx_aplic_auto ON aplicaciones(marca_auto, modelo_auto)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_aplic_codigo ON aplicaciones(codigo_clean)")

    c.execute("""CREATE TABLE IF NOT EXISTS importaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        marca TEXT,
        archivo TEXT,
        filas_cargadas INTEGER,
        filas_omitidas INTEGER,
        fecha TEXT DEFAULT (datetime('now'))
    )""")
    # Huella del archivo importado, para reconocer la MISMA planilla aunque le hayan cambiado el
    # nombre. Reimportar la misma lista sin darse cuenta duplica el trabajo de revisión y puede
    # revivir precios viejos encima de los actualizados a mano.
    _cols_imp = [f[1] for f in c.execute("PRAGMA table_info(importaciones)").fetchall()]
    if "huella" not in _cols_imp:
        c.execute("ALTER TABLE importaciones ADD COLUMN huella TEXT")
    if "lote" not in _cols_imp:
        c.execute("ALTER TABLE importaciones ADD COLUMN lote TEXT")
    c.execute("CREATE INDEX IF NOT EXISTS idx_importaciones_huella ON importaciones(huella)")

    c.execute("""CREATE TABLE IF NOT EXISTS catalogos_externos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE NOT NULL,
        url TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now'))
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS historial_busquedas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        termino TEXT NOT NULL,
        usuario TEXT,
        fecha TEXT DEFAULT (datetime('now'))
    )""")
    columnas_historial = [f[1] for f in c.execute("PRAGMA table_info(historial_busquedas)").fetchall()]
    if "usuario" not in columnas_historial:
        c.execute("ALTER TABLE historial_busquedas ADD COLUMN usuario TEXT")

    # Migraciones: agregar columnas nuevas si no existen todavía
    columnas_productos = [f[1] for f in c.execute("PRAGMA table_info(productos)").fetchall()]
    if "precio" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN precio REAL")
    if "stock" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN stock INTEGER")
    if "favorito" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN favorito INTEGER DEFAULT 0")
    if "imagen_url" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN imagen_url TEXT")
    if "imagen_orb_blob" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN imagen_orb_blob BLOB")
    # Estado del procesamiento visual de la foto: NULL = sin intentar, 'ok' = tiene descriptores,
    # 'sin_detalle' = se intentó y la foto no da (pieza lisa, borrosa), 'error' = falló el proceso.
    # Sin esto, una foto que no sirve quedaba "pendiente" para siempre: el botón de procesar
    # mostraba pendientes, se tocaba, no cambiaba nada, y volvía a mostrar lo mismo.
    if "imagen_orb_estado" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN imagen_orb_estado TEXT")
    # Recuerda que a ese código ya se le buscó foto en la ficha del proveedor y no había. Sin
    # esto, cada tanda volvía a golpear los mismos miles de códigos sin foto y nunca avanzaba.
    if "foto_busqueda_estado" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN foto_busqueda_estado TEXT")

    # Varias fotos por producto: la del catálogo del proveedor, la que sacaste vos, la de otra
    # marca del mismo repuesto. Al buscar se compara contra todas y se queda con la mejor. Es lo
    # único que resuelve de verdad el cambio de ángulo: ninguna comparación reconoce una pieza
    # fotografiada de frente en una foto sacada de costado.
    c.execute("""CREATE TABLE IF NOT EXISTS producto_fotos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        imagen_data TEXT,
        firma_blob BLOB,
        estado TEXT,
        origen TEXT,
        fuente TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_producto_fotos_prod ON producto_fotos(producto_id)")
    # Con qué versión del comparador se calculó cada firma. Cuando el comparador mejora, las
    # firmas viejas quedan sin los datos nuevos y la mejora no se aplica al catálogo que ya
    # tenías cargado — que es justo donde hace falta. Con esto se detectan y se recalculan.
    _cols_fotos = [f[1] for f in c.execute("PRAGMA table_info(producto_fotos)").fetchall()]
    if "firma_version" not in _cols_fotos:
        c.execute("ALTER TABLE producto_fotos ADD COLUMN firma_version INTEGER")
    # Miniatura chica aparte: la foto normal pesa bastante y la búsqueda la traía entera por
    # cada resultado, aunque en la tabla se vea en chiquito. Acá se guarda una versión liviana
    # solo para esas listas; la grande se sigue usando al ver el producto en Administrar.
    if "imagen_thumb" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN imagen_thumb TEXT")
    if "diametro_interno" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN diametro_interno REAL")
    if "diametro_externo" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN diametro_externo REAL")
    if "diametro_interno_cara_b" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN diametro_interno_cara_b REAL")
    if "diametro_externo_cara_b" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN diametro_externo_cara_b REAL")
    if "diametro_rosca_homocinetica" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN diametro_rosca_homocinetica REAL")
    if "diametro_copa" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN diametro_copa REAL")
    # La copa de una homocinética es cónica: mide distinto en la base que en la boca. Con un
    # solo diámetro no alcanzaba para distinguir dos copas que arrancan igual y terminan
    # distinto, y son justo las que no se pueden intercambiar.
    # Precio de costo, aparte del de venta. Con un solo precio no se puede saber el margen, y
    # sin margen la decisión de qué ofrecer entre tres equivalentes se toma a ojo: no siempre
    # conviene el más barato para el cliente.
    if "precio_costo" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN precio_costo REAL")

    if "diametro_copa_superior" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN diametro_copa_superior REAL")
    # El largo total es la medida que primero descarta: dos homocinéticas con las mismas estrías
    # y la misma copa pero distinto largo no entran en el mismo auto.
    if "largo_total" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN largo_total REAL")
    if "ancho" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN ancho REAL")
    if "paso_rosca" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN paso_rosca TEXT")
    if "cantidad_estrias" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN cantidad_estrias INTEGER")
    if "estrias_internas" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN estrias_internas INTEGER")
    if "estrias_externas" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN estrias_externas INTEGER")
    if "posicion_seguro" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN posicion_seguro TEXT")
    if "tiene_abs" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN tiene_abs INTEGER")
    if "ubicacion" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN ubicacion TEXT")
    if "veces_buscado" not in columnas_productos:
        c.execute("ALTER TABLE productos ADD COLUMN veces_buscado INTEGER DEFAULT 0")

    # Vehículos y ficha digital ("mellizo digital") para historial de piezas por patente
    c.execute("""CREATE TABLE IF NOT EXISTS vehiculos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patente TEXT UNIQUE NOT NULL,
        cliente_nombre TEXT,
        cliente_telefono TEXT,
        marca_auto TEXT,
        modelo_auto TEXT,
        anio TEXT,
        motorizacion TEXT,
        km_registro INTEGER,
        km_actual INTEGER,
        km_actualizado_fecha TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )""")
    columnas_vehiculos_extra = [f[1] for f in c.execute("PRAGMA table_info(vehiculos)").fetchall()]
    if "anio" not in columnas_vehiculos_extra:
        c.execute("ALTER TABLE vehiculos ADD COLUMN anio TEXT")
    if "motorizacion" not in columnas_vehiculos_extra:
        c.execute("ALTER TABLE vehiculos ADD COLUMN motorizacion TEXT")
    # VIN en la ficha: permite entrar por el número de chasis además de por la patente, y hace
    # que el decodificador pueda devolver datos REALES del auto (no estimados) cuando ya pasó
    # por el mostrador. De paso, cada ficha con VIN + modelo le enseña el patrón a la app.
    if "vin" not in columnas_vehiculos_extra:
        c.execute("ALTER TABLE vehiculos ADD COLUMN vin TEXT")
    c.execute("CREATE INDEX IF NOT EXISTS idx_vehiculos_vin ON vehiculos(vin)")

    # Migración: instalaciones existentes que no tenían km_registro (km de cuando se cargó
    # el vehículo por primera vez, fijo, para poder calcular km recorridos).
    columnas_vehiculos = [f[1] for f in c.execute("PRAGMA table_info(vehiculos)").fetchall()]
    if "km_registro" not in columnas_vehiculos:
        c.execute("ALTER TABLE vehiculos ADD COLUMN km_registro INTEGER")
        # Para los vehículos que ya existían, se usa el km_actual que tengan como punto de partida
        # (es lo mejor que se puede hacer sin el dato original; a partir de ahora queda fijo).
        c.execute("UPDATE vehiculos SET km_registro = km_actual WHERE km_registro IS NULL")

    c.execute("""CREATE TABLE IF NOT EXISTS historial_piezas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehiculo_id INTEGER NOT NULL REFERENCES vehiculos(id) ON DELETE CASCADE,
        producto_id INTEGER REFERENCES productos(id) ON DELETE SET NULL,
        descripcion_pieza TEXT NOT NULL,
        marca_pieza TEXT,
        codigo_pieza TEXT,
        km_instalacion INTEGER,
        fecha_instalacion TEXT DEFAULT (datetime('now')),
        vida_util_km INTEGER,
        nota TEXT
    )""")

    # Auditoría diaria de stock por muestreo aleatorio
    c.execute("""CREATE TABLE IF NOT EXISTS auditoria_diaria (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        stock_sistema INTEGER,
        stock_contado INTEGER,
        diferencia INTEGER,
        resuelto INTEGER DEFAULT 0,
        UNIQUE(fecha, producto_id)
    )""")

    # ---- Modo Mecánico ----
    # fabricante = '' significa código genérico (estándar OBD-II, válido para cualquier auto).
    # Un mismo código (ej. P1105) puede repetirse con distinto fabricante, porque en los
    # códigos específicos de marca el mismo número significa cosas distintas según el auto.
    c.execute("""CREATE TABLE IF NOT EXISTS codigos_dtc (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT NOT NULL,
        fabricante TEXT NOT NULL DEFAULT '',
        descripcion TEXT NOT NULL,
        sistema TEXT,
        causas_posibles TEXT,
        UNIQUE(codigo, fabricante)
    )""")

    # Migración: las instalaciones que ya tenían la tabla vieja (sin columna fabricante,
    # con UNIQUE solo en codigo) se convierten al esquema nuevo sin perder datos cargados.
    columnas_dtc = [f[1] for f in c.execute("PRAGMA table_info(codigos_dtc)").fetchall()]
    if "fabricante" not in columnas_dtc:
        c.execute("ALTER TABLE codigos_dtc RENAME TO codigos_dtc_old")
        c.execute("""CREATE TABLE codigos_dtc (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL,
            fabricante TEXT NOT NULL DEFAULT '',
            descripcion TEXT NOT NULL,
            sistema TEXT,
            causas_posibles TEXT,
            UNIQUE(codigo, fabricante)
        )""")
        c.execute("""INSERT INTO codigos_dtc (codigo, fabricante, descripcion, sistema, causas_posibles)
                     SELECT codigo, '', descripcion, sistema, causas_posibles FROM codigos_dtc_old""")
        c.execute("DROP TABLE codigos_dtc_old")

    c.execute("""CREATE TABLE IF NOT EXISTS fabricantes_vin (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        wmi TEXT UNIQUE NOT NULL,
        fabricante TEXT NOT NULL,
        pais TEXT
    )""")

    # Modelos por patrón VDS (posiciones 4 a 8 del VIN).
    # A diferencia del WMI, que es un registro internacional y siempre significa lo mismo, las
    # posiciones 4-8 las define CADA fabricante como quiere: no hay forma de deducir el modelo
    # de un VIN sin una base licenciada (TecDoc y similares, que son pagas). Lo que sí se puede
    # es que la app APRENDA: la primera vez lo cargás vos, y de ahí en más todo VIN con el mismo
    # patrón se autocompleta solo. Con el tiempo cubre los autos que realmente atendés.
    c.execute("""CREATE TABLE IF NOT EXISTS modelos_vin (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        wmi TEXT NOT NULL,
        vds TEXT NOT NULL,
        modelo TEXT NOT NULL,
        notas TEXT,
        veces INTEGER DEFAULT 1,
        UNIQUE (wmi, vds)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_modelos_vin ON modelos_vin(wmi, vds)")
    _cols_modelos_vin = [f[1] for f in c.execute("PRAGMA table_info(modelos_vin)").fetchall()]
    if "motor" not in _cols_modelos_vin:
        c.execute("ALTER TABLE modelos_vin ADD COLUMN motor TEXT")

    # El motor por la 8ª posición del VIN.
    # En los VIN de Norteamérica esa posición es, POR NORMA, el código de motor. Fuera de
    # Norteamérica no hay norma, pero casi todos los fabricantes la usan igual para eso.
    # Se guarda aparte del modelo porque generaliza distinto: un mismo código de motor aparece
    # en varios modelos de la misma marca, así que aprendido una vez sirve para todos.
    c.execute("""CREATE TABLE IF NOT EXISTS motores_vin (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        wmi TEXT NOT NULL,
        codigo TEXT NOT NULL,
        motor TEXT NOT NULL,
        notas TEXT,
        veces INTEGER DEFAULT 1,
        UNIQUE (wmi, codigo)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_motores_vin ON motores_vin(wmi, codigo)")

    c.execute("""CREATE TABLE IF NOT EXISTS esquemas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT NOT NULL,
        marca_auto TEXT,
        modelo_auto TEXT,
        sistema TEXT,
        descripcion TEXT,
        imagen_blob BLOB,
        imagen_nombre TEXT,
        generado_ia INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    )""")
    columnas_esquemas = [f[1] for f in c.execute("PRAGMA table_info(esquemas)").fetchall()]
    if "generado_ia" not in columnas_esquemas:
        c.execute("ALTER TABLE esquemas ADD COLUMN generado_ia INTEGER DEFAULT 0")

    # Catálogo de marca/vehículo para "Explorar por categoría", separado de los esquemas en sí:
    # permite precargar la estructura (Volkswagen > Gol Trend) sin necesidad de subir ya una imagen.
    c.execute("""CREATE TABLE IF NOT EXISTS esquemas_catalogo (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        marca TEXT NOT NULL,
        modelo TEXT NOT NULL,
        UNIQUE(marca, modelo)
    )""")

    # Piezas marcadas dentro de un esquema (número/nombre + código), para poder buscarlas
    # directamente en el catálogo desde el diagrama — es lo que le da función de "despiece".
    c.execute("""CREATE TABLE IF NOT EXISTS esquema_puntos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        esquema_id INTEGER NOT NULL REFERENCES esquemas(id) ON DELETE CASCADE,
        numero TEXT,
        nombre_pieza TEXT NOT NULL,
        codigo TEXT,
        producto_id INTEGER REFERENCES productos(id) ON DELETE SET NULL,
        pos_x REAL,
        pos_y REAL,
        orden INTEGER DEFAULT 0
    )""")

    # Alias/CBU para el QR de transferencia en las cotizaciones. Se pueden cargar varios
    # (Mercado Pago, distintos bancos, etc.) y elegir cuál usar en cada cotización puntual.
    c.execute("""CREATE TABLE IF NOT EXISTS alias_transferencia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        alias TEXT,
        cbu TEXT,
        titular TEXT,
        qr_real_blob BLOB
    )""")
    columnas_alias = [f[1] for f in c.execute("PRAGMA table_info(alias_transferencia)").fetchall()]
    if "qr_real_blob" not in columnas_alias:
        c.execute("ALTER TABLE alias_transferencia ADD COLUMN qr_real_blob BLOB")

    # Configuración simple de clave/valor (ej: encabezado/pie del mensaje de WhatsApp).
    c.execute("""CREATE TABLE IF NOT EXISTS configuracion (
        clave TEXT PRIMARY KEY,
        valor TEXT
    )""")

    # Contador de uso de las funciones de IA — para ver de un vistazo cuánto se usa cada una
    # y anticipar si alguna se está acercando a los límites gratuitos.
    c.execute("""CREATE TABLE IF NOT EXISTS uso_ia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        funcion TEXT NOT NULL,
        usuario TEXT,
        exito INTEGER,
        fecha TEXT DEFAULT (datetime('now'))
    )""")

    # Papelera: guarda una copia de lo que se borra (marcas, productos, combos, alias) para
    # poder restaurarlo si fue un error. No reemplaza el backup completo, es para el día a día.
    c.execute("""CREATE TABLE IF NOT EXISTS papelera (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT NOT NULL,
        datos_json TEXT NOT NULL,
        eliminado_por TEXT,
        eliminado_en TEXT DEFAULT (datetime('now'))
    )""")

    # Cuando un empleado busca algo y no hay stock (o le falta), lo marca acá para que el dueño
    # lo revise después y decida qué pedirle a cada proveedor. Un mismo producto pedido varias
    # veces por distintos empleados suma en "veces_solicitado" en vez de duplicar filas.
    c.execute("""CREATE TABLE IF NOT EXISTS pedidos_reposicion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        veces_solicitado INTEGER DEFAULT 1,
        ultimo_solicitado_por TEXT,
        ultima_fecha TEXT DEFAULT (datetime('now')),
        estado TEXT DEFAULT 'pendiente',
        UNIQUE(producto_id)
    )""")

    # Historial de precios: cada vez que se cambia el precio de un producto queda un registro,
    # para poder ver cómo fue variando en el tiempo.
    c.execute("""CREATE TABLE IF NOT EXISTS historial_precios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        precio REAL,
        fecha TEXT DEFAULT (datetime('now'))
    )""")

    # Ventas registradas desde el mostrador: qué se llevó el cliente y qué había pedido.
    # Es la materia prima para descubrir equivalencias solas: si alguien pide el código A y
    # termina llevándose el B, eso es una equivalencia que pasó en la vida real, aunque no
    # figure en ningún catálogo.
    c.execute("""CREATE TABLE IF NOT EXISTS ventas_registradas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
        termino_pedido TEXT,
        codigo_pedido_clean TEXT,
        usuario TEXT,
        fecha TEXT DEFAULT (datetime('now'))
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ventas_fecha ON ventas_registradas(fecha)")
    # Sin este, "cuándo se vendió por última vez este producto" recorría la tabla de ventas
    # ENTERA una vez por producto. Con el estante lleno eso es minutos de espera.
    c.execute("CREATE INDEX IF NOT EXISTS idx_ventas_prod ON ventas_registradas(producto_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ventas_codigo ON ventas_registradas(codigo_pedido_clean)")

    # Mapeo de columnas recordado por proveedor. Cada vez que se importa una lista hay que
    # volver a indicar qué columna es el código, cuál el precio, etc. Un error ahí es lo que
    # mete basura en la base (un producto con código "1" colgado de decenas de equivalencias),
    # así que conviene que la próxima vez venga preseleccionado como la vez que funcionó.
    c.execute("""CREATE TABLE IF NOT EXISTS mapeo_columnas (
        proveedor TEXT PRIMARY KEY,
        idx_prov INTEGER, idx_oem INTEGER, idx_desc INTEGER,
        idx_precio INTEGER, idx_stock INTEGER,
        buscar_oem_en_desc INTEGER DEFAULT 0,
        prov_es_oem INTEGER DEFAULT 0,
        fecha TEXT DEFAULT (datetime('now'))
    )""")

    # Decisiones ya tomadas sobre un vínculo puntual. Sirve para dos cosas: que lo revisado no
    # vuelva a aparecer en la auditoría, y que lo rechazado no se vuelva a crear si más adelante
    # se importa de nuevo la misma lista del proveedor.
    c.execute("""CREATE TABLE IF NOT EXISTS equivalencias_revisadas (
        producto_a_id INTEGER NOT NULL,
        producto_b_id INTEGER NOT NULL,
        decision TEXT NOT NULL,
        revisado_por TEXT,
        fecha TEXT DEFAULT (datetime('now')),
        PRIMARY KEY (producto_a_id, producto_b_id)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_revisadas_decision ON equivalencias_revisadas(decision)")

    # Vínculos que llegaron de una lista de proveedor y esperan revisión. Una importación puede
    # generar miles de vínculos de una: si se cargaran solos, un error en la columna de código de
    # fábrica te ensucia la base entera sin que nadie se entere.
    c.execute("""CREATE TABLE IF NOT EXISTS equivalencias_pendientes (
        producto_a_id INTEGER NOT NULL,
        producto_b_id INTEGER NOT NULL,
        origen TEXT,
        lote TEXT,
        fecha TEXT DEFAULT (datetime('now')),
        PRIMARY KEY (producto_a_id, producto_b_id)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_pendientes_lote ON equivalencias_pendientes(lote)")

    # Evidencia que respalda cada equivalencia sugerida. La idea es NO cargar nada solo:
    # cada sugerencia llega al panel con el detalle de en qué se basa, para poder decidir
    # con fundamento en vez de a ciegas.
    c.execute("""CREATE TABLE IF NOT EXISTS evidencia_equivalencia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_clean TEXT NOT NULL,
        producto_id INTEGER NOT NULL,
        tipo TEXT NOT NULL,
        detalle TEXT,
        fecha TEXT DEFAULT (datetime('now')),
        UNIQUE(codigo_clean, producto_id, tipo)
    )""")

    # Sugerencias que el dueño ya miró y descartó, para no volver a proponérselas.
    c.execute("""CREATE TABLE IF NOT EXISTS equivalencias_descartadas (
        codigo_clean TEXT NOT NULL,
        producto_id INTEGER NOT NULL,
        descartado_por TEXT,
        fecha TEXT DEFAULT (datetime('now')),
        PRIMARY KEY (codigo_clean, producto_id)
    )""")

    # Cuentas de empleados creadas desde la propia app (además de las que se pueden cargar en
    # Streamlit Secrets) — así el dueño no depende de tocar la configuración de Streamlit Cloud
    # cada vez que entra o se va alguien del equipo. La contraseña nunca se guarda en texto plano.
    c.execute("""CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        rol TEXT NOT NULL DEFAULT 'operador',
        activo INTEGER DEFAULT 1,
        creado_en TEXT DEFAULT (datetime('now'))
    )""")

    # Mecánicos externos (no son empleados del local): tienen su propio login, pero solo ven
    # el portal de armar presupuestos — nunca las secciones internas del negocio.
    c.execute("""CREATE TABLE IF NOT EXISTS mecanicos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        activo INTEGER DEFAULT 1,
        creado_en TEXT DEFAULT (datetime('now'))
    )""")

    # Presupuestos armados por mecánicos externos: repuestos elegidos + su propia mano de obra.
    c.execute("""CREATE TABLE IF NOT EXISTS presupuestos_mecanico (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mecanico_id INTEGER NOT NULL REFERENCES mecanicos(id) ON DELETE CASCADE,
        cliente_nombre TEXT,
        items_json TEXT NOT NULL,
        mano_obra REAL DEFAULT 0,
        total REAL,
        creado_en TEXT DEFAULT (datetime('now'))
    )""")

    # Combos de repuestos que suelen cambiarse juntos (ej: correa de distribución -> kit + tensor + bomba de agua).
    # "disparador" es la palabra/frase que se busca dentro de la descripción del producto encontrado.
    c.execute("""CREATE TABLE IF NOT EXISTS combos_sugeridos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        disparador TEXT NOT NULL,
        item TEXT NOT NULL
    )""")
    c.execute("SELECT COUNT(*) FROM combos_sugeridos")
    if c.fetchone()[0] == 0:
        c.executemany(
            "INSERT INTO combos_sugeridos (disparador, item) VALUES (?, ?)",
            [
                ("correa de distribucion", "Kit de distribución"),
                ("correa de distribucion", "Tensor de distribución"),
                ("correa de distribucion", "Bomba de agua"),
            ]
        )

    # Semilla inicial de códigos DTC genéricos (estándar OBD-II / SAE J2012, no específicos de
    # marca), verificados contra fuentes de referencia. Es un punto de partida — sumá o corregí
    # los que necesites desde la app. Los códigos P1xxx específicos de fabricante se cargan
    # aparte indicando la marca (ver Modo Mecánico → Códigos DTC).
    c.execute("SELECT COUNT(*) FROM codigos_dtc")
    if c.fetchone()[0] == 0:
        seed_dtc = [
            ("P0010","Falla eléctrica en el actuador de posición A del árbol de levas, banco 1","Motor - Sensores/Admisión","Cableado cortado o en corto, conector sucio/flojo, sensor o actuador dañado"),
            ("P0011","Avance excesivo o mal desempeño en la posición A del árbol de levas, banco 1","Motor - Sensores/Admisión","Sensor descalibrado, obstrucción física, fuga, componente mecánico desgastado"),
            ("P0012","Retardo excesivo en la posición A del árbol de levas, banco 1","Motor - Sensores/Admisión","Sensor descalibrado, obstrucción física, fuga, componente mecánico desgastado"),
            ("P0013","Falla eléctrica en el actuador de posición B del árbol de levas, banco 1","Motor - Sensores/Admisión","Cableado cortado o en corto, conector sucio/flojo, sensor o actuador dañado"),
            ("P0014","Avance excesivo o mal desempeño en la posición B del árbol de levas, banco 1","Motor - Sensores/Admisión","Sensor descalibrado, obstrucción física, fuga, componente mecánico desgastado"),
            ("P0015","Retardo excesivo en la posición B del árbol de levas, banco 1","Motor - Sensores/Admisión","Sensor descalibrado, obstrucción física, fuga, componente mecánico desgastado"),
            ("P0020","Falla eléctrica en el actuador de posición A del árbol de levas, banco 2","Motor - Sensores/Admisión","Cableado cortado o en corto, conector sucio/flojo, sensor o actuador dañado"),
            ("P0021","Avance excesivo o mal desempeño en la posición A del árbol de levas, banco 2","Motor - Sensores/Admisión","Sensor descalibrado, obstrucción física, fuga, componente mecánico desgastado"),
            ("P0022","Retardo excesivo en la posición A del árbol de levas, banco 2","Motor - Sensores/Admisión","Sensor descalibrado, obstrucción física, fuga, componente mecánico desgastado"),
            ("P0030","Falla eléctrica en el calefactor del sensor de oxígeno, banco 1 sensor 1","Emisiones","Cableado cortado o en corto, conector sucio/flojo, sensor o actuador dañado"),
            ("P0031","Señal baja en el calefactor del sensor de oxígeno, banco 1 sensor 1","Emisiones","Cortocircuito a masa, sensor en mal estado, cableado dañado"),
            ("P0032","Señal alta en el calefactor del sensor de oxígeno, banco 1 sensor 1","Emisiones","Circuito abierto, cortocircuito a positivo, sensor en mal estado"),
            ("P0036","Falla eléctrica en el calefactor del sensor de oxígeno, banco 1 sensor 2","Emisiones","Cableado cortado o en corto, conector sucio/flojo, sensor o actuador dañado"),
            ("P0037","Señal baja en el calefactor del sensor de oxígeno, banco 1 sensor 2","Emisiones","Cortocircuito a masa, sensor en mal estado, cableado dañado"),
            ("P0038","Señal alta en el calefactor del sensor de oxígeno, banco 1 sensor 2","Emisiones","Circuito abierto, cortocircuito a positivo, sensor en mal estado"),
            ("P0050","Falla eléctrica en el calefactor del sensor de oxígeno, banco 2 sensor 1","Emisiones","Cableado cortado o en corto, conector sucio/flojo, sensor o actuador dañado"),
            ("P0051","Señal baja en el calefactor del sensor de oxígeno, banco 2 sensor 1","Emisiones","Cortocircuito a masa, sensor en mal estado, cableado dañado"),
            ("P0052","Señal alta en el calefactor del sensor de oxígeno, banco 2 sensor 1","Emisiones","Circuito abierto, cortocircuito a positivo, sensor en mal estado"),
            ("P0056","Falla eléctrica en el calefactor del sensor de oxígeno, banco 2 sensor 2","Emisiones","Cableado cortado o en corto, conector sucio/flojo, sensor o actuador dañado"),
            ("P0057","Señal baja en el calefactor del sensor de oxígeno, banco 2 sensor 2","Emisiones","Cortocircuito a masa, sensor en mal estado, cableado dañado"),
            ("P0058","Señal alta en el calefactor del sensor de oxígeno, banco 2 sensor 2","Emisiones","Circuito abierto, cortocircuito a positivo, sensor en mal estado"),
            ("P0070","Falla eléctrica en el sensor de temperatura de aire ambiente","Motor - Sensores/Admisión","Cableado cortado o en corto, conector sucio/flojo, sensor dañado"),
            ("P0071","Sensor de temperatura de aire ambiente fuera de rango","Motor - Sensores/Admisión","Sensor descalibrado, cableado dañado"),
            ("P0072","Señal baja en el sensor de temperatura de aire ambiente","Motor - Sensores/Admisión","Cortocircuito a masa, sensor en mal estado"),
            ("P0073","Señal alta en el sensor de temperatura de aire ambiente","Motor - Sensores/Admisión","Circuito abierto, sensor en mal estado"),
            ("P0074","Señal intermitente en el sensor de temperatura de aire ambiente","Motor - Sensores/Admisión","Conector flojo u oxidado, falso contacto"),
            ("P0100","Falla eléctrica en el medidor de caudal de aire (MAF)","Motor - Sensores/Admisión","Sensor sucio, cableado, conector"),
            ("P0101","Medidor de caudal de aire (MAF) fuera de rango","Motor - Sensores/Admisión","Filtro de aire sucio, fugas de vacío, sensor sucio"),
            ("P0102","Señal baja en el medidor de caudal de aire (MAF)","Motor - Sensores/Admisión","Cortocircuito a masa, sensor sucio o dañado"),
            ("P0103","Señal alta en el medidor de caudal de aire (MAF)","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0104","Señal intermitente en el medidor de caudal de aire (MAF)","Motor - Sensores/Admisión","Conector flojo u oxidado, falso contacto"),
            ("P0105","Falla eléctrica en el sensor de presión absoluta de múltiple / barométrica (MAP)","Motor - Sensores/Admisión","Manguera de vacío rota, sensor o cableado dañado"),
            ("P0106","Sensor MAP fuera de rango","Motor - Sensores/Admisión","Fuga de vacío, sensor descalibrado"),
            ("P0107","Señal baja en el sensor MAP","Motor - Sensores/Admisión","Cortocircuito a masa, sensor dañado"),
            ("P0108","Señal alta en el sensor MAP","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0109","Señal intermitente en el sensor MAP","Motor - Sensores/Admisión","Conector flojo u oxidado, falso contacto"),
            ("P0110","Falla eléctrica en el sensor de temperatura del aire de admisión (IAT)","Motor - Sensores/Admisión","Sensor o cableado en mal estado"),
            ("P0111","Sensor IAT fuera de rango","Motor - Sensores/Admisión","Sensor descalibrado, cableado dañado"),
            ("P0112","Señal baja en el sensor IAT","Motor - Sensores/Admisión","Cortocircuito a masa, sensor dañado"),
            ("P0113","Señal alta en el sensor IAT","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0114","Señal intermitente en el sensor IAT","Motor - Sensores/Admisión","Conector flojo u oxidado, falso contacto"),
            ("P0115","Falla eléctrica en el sensor de temperatura del refrigerante (ECT)","Motor - Sensores/Admisión","Sensor, conector, cableado"),
            ("P0116","Sensor ECT fuera de rango","Motor - Sensores/Admisión","Sensor descalibrado, nivel de refrigerante bajo"),
            ("P0117","Señal baja en el sensor ECT","Motor - Sensores/Admisión","Cortocircuito a masa, sensor dañado"),
            ("P0118","Señal alta en el sensor ECT","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0119","Señal intermitente en el sensor ECT","Motor - Sensores/Admisión","Conector flojo u oxidado, falso contacto"),
            ("P0120","Falla eléctrica en el sensor de posición del acelerador/pedal A (TPS)","Motor - Sensores/Admisión","Sensor TPS, cableado"),
            ("P0121","Sensor TPS A fuera de rango","Motor - Sensores/Admisión","Sensor descalibrado, cableado dañado"),
            ("P0122","Señal baja en el sensor TPS A","Motor - Sensores/Admisión","Cortocircuito a masa, sensor dañado"),
            ("P0123","Señal alta en el sensor TPS A","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0124","Señal intermitente en el sensor TPS A","Motor - Sensores/Admisión","Conector flojo u oxidado, falso contacto"),
            ("P0125","El refrigerante no llega a la temperatura necesaria para el lazo cerrado de combustible","Motor - Sensores/Admisión","Termostato pegado en abierto, sensor ECT"),
            ("P0128","Termostato: el refrigerante no alcanza la temperatura de regulación","Motor - Sensores/Admisión","Termostato pegado en abierto"),
            ("P0130","Falla eléctrica en el sensor de oxígeno, banco 1 sensor 1","Emisiones","Sonda lambda, cableado"),
            ("P0131","Voltaje bajo en el sensor de oxígeno, banco 1 sensor 1","Emisiones","Cortocircuito a masa, sonda en mal estado"),
            ("P0132","Voltaje alto en el sensor de oxígeno, banco 1 sensor 1","Emisiones","Circuito abierto, sonda en mal estado"),
            ("P0133","Respuesta lenta en el sensor de oxígeno, banco 1 sensor 1","Emisiones","Sonda envejecida o contaminada"),
            ("P0134","Sin actividad detectada en el sensor de oxígeno, banco 1 sensor 1","Emisiones","Sonda desconectada o sin actividad, cableado"),
            ("P0135","Falla eléctrica en el calefactor del sensor de oxígeno, banco 1 sensor 1","Emisiones","Calefactor de la sonda dañado, fusible, cableado"),
            ("P0136","Falla eléctrica en el sensor de oxígeno, banco 1 sensor 2","Emisiones","Sonda lambda, cableado"),
            ("P0137","Voltaje bajo en el sensor de oxígeno, banco 1 sensor 2","Emisiones","Cortocircuito a masa, sonda en mal estado"),
            ("P0138","Voltaje alto en el sensor de oxígeno, banco 1 sensor 2","Emisiones","Circuito abierto, sonda en mal estado"),
            ("P0140","Sin actividad detectada en el sensor de oxígeno, banco 1 sensor 2","Emisiones","Sonda desconectada o sin actividad, cableado"),
            ("P0141","Falla eléctrica en el calefactor del sensor de oxígeno, banco 1 sensor 2","Emisiones","Calefactor de la sonda dañado, fusible, cableado"),
            ("P0150","Falla eléctrica en el sensor de oxígeno, banco 2 sensor 1","Emisiones","Sonda lambda, cableado"),
            ("P0155","Falla eléctrica en el calefactor del sensor de oxígeno, banco 2 sensor 1","Emisiones","Calefactor de la sonda dañado, fusible, cableado"),
            ("P0170","Ajuste de mezcla fuera de rango, banco 1","Motor - Sensores/Admisión","Sonda O2, inyectores, fugas de vacío"),
            ("P0171","Mezcla demasiado pobre, banco 1","Motor - Sensores/Admisión","Fuga de vacío, inyector, sensor MAF"),
            ("P0172","Mezcla demasiado rica, banco 1","Motor - Sensores/Admisión","Inyector, presión de combustible, sensor O2"),
            ("P0173","Ajuste de mezcla fuera de rango, banco 2","Motor - Sensores/Admisión","Sonda O2, inyectores, fugas de vacío"),
            ("P0174","Mezcla demasiado pobre, banco 2","Motor - Sensores/Admisión","Fuga de vacío, inyector, sensor MAF"),
            ("P0175","Mezcla demasiado rica, banco 2","Motor - Sensores/Admisión","Inyector, presión de combustible, sensor O2"),
            ("P0200","Falla eléctrica general en el circuito de inyectores","Motor - Inyectores/Combustible","Inyector, cableado, módulo"),
            ("P0201","Falla eléctrica en el inyector del cilindro 1","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0202","Falla eléctrica en el inyector del cilindro 2","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0203","Falla eléctrica en el inyector del cilindro 3","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0204","Falla eléctrica en el inyector del cilindro 4","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0217","Sobretemperatura del motor","Motor - Encendido/Combustión","Refrigerante, bomba de agua, termostato"),
            ("P0230","Falla eléctrica en el circuito primario de la bomba de combustible","Motor - Inyectores/Combustible","Bomba, relé, cableado"),
            ("P0300","Fallos de encendido detectados en varios cilindros o aleatorios","Motor - Encendido/Combustión","Bujías, bobinas, compresión"),
            ("P0301","Fallo de encendido en el cilindro 1","Motor - Encendido/Combustión","Bujía, bobina, inyector de ese cilindro"),
            ("P0302","Fallo de encendido en el cilindro 2","Motor - Encendido/Combustión","Bujía, bobina, inyector de ese cilindro"),
            ("P0303","Fallo de encendido en el cilindro 3","Motor - Encendido/Combustión","Bujía, bobina, inyector de ese cilindro"),
            ("P0304","Fallo de encendido en el cilindro 4","Motor - Encendido/Combustión","Bujía, bobina, inyector de ese cilindro"),
            ("P0325","Falla eléctrica en el sensor de detonación (knock sensor), banco 1 o único","Motor - Encendido/Combustión","Sensor, cableado"),
            ("P0326","Sensor de detonación 1 fuera de rango","Motor - Encendido/Combustión","Sensor descalibrado, cableado"),
            ("P0327","Señal baja en el sensor de detonación 1","Motor - Encendido/Combustión","Cortocircuito a masa, sensor dañado"),
            ("P0328","Señal alta en el sensor de detonación 1","Motor - Encendido/Combustión","Circuito abierto, sensor dañado"),
            ("P0335","Falla eléctrica en el sensor de posición del cigüeñal (CKP)","Motor - Encendido/Combustión","Sensor CKP, cableado, tone wheel"),
            ("P0336","Sensor CKP fuera de rango","Motor - Encendido/Combustión","Sensor descalibrado, rueda fónica dañada"),
            ("P0340","Falla eléctrica en el sensor de posición del árbol de levas (CMP)","Motor - Encendido/Combustión","Sensor CMP, cableado"),
            ("P0341","Sensor CMP fuera de rango","Motor - Encendido/Combustión","Sensor descalibrado, cableado"),
            ("P0400","Falla en el caudal de recirculación de gases de escape (EGR)","Emisiones","Válvula EGR, conductos obstruidos"),
            ("P0401","Caudal insuficiente de EGR","Emisiones","Válvula EGR trabada cerrada, conducto obstruido"),
            ("P0402","Caudal excesivo de EGR","Emisiones","Válvula EGR trabada abierta"),
            ("P0420","Eficiencia del catalizador por debajo del umbral, banco 1","Emisiones","Catalizador, sonda lambda"),
            ("P0430","Eficiencia del catalizador por debajo del umbral, banco 2","Emisiones","Catalizador, sonda lambda"),
            ("P0440","Falla general en el sistema de control de emisiones evaporativas (EVAP)","Emisiones","Tapa de nafta, válvula, mangueras"),
            ("P0441","Caudal de purga EVAP incorrecto","Emisiones","Válvula de purga, mangueras obstruidas"),
            ("P0442","Fuga pequeña detectada en el sistema EVAP","Emisiones","Tapa de nafta floja, manguera con fisura"),
            ("P0446","Falla eléctrica en la válvula de ventilación del sistema EVAP","Emisiones","Válvula de venteo, cableado"),
            ("P0447","Circuito de ventilación EVAP abierto","Emisiones","Cableado cortado, válvula desconectada"),
            ("P0448","Circuito de ventilación EVAP en corto","Emisiones","Cableado en corto, válvula dañada"),
            ("P0451","Falla eléctrica en el sensor de presión del sistema EVAP","Emisiones","Sensor de presión, cableado"),
            ("P0452","Señal baja en el sensor de presión del sistema EVAP","Emisiones","Cortocircuito a masa, sensor dañado"),
            ("P0453","Señal alta en el sensor de presión del sistema EVAP","Emisiones","Circuito abierto, sensor dañado"),
            ("P0455","Fuga grande detectada en el sistema EVAP","Emisiones","Tapa de nafta, manguera desconectada"),
            ("P0456","Fuga muy pequeña detectada en el sistema EVAP","Emisiones","Tapa de nafta, fisura muy pequeña"),
            ("P0457","Fuga detectada, posible tapa de combustible floja o mal cerrada","Emisiones","Tapa de nafta floja, dañada o mal puesta"),
            ("P0461","Sensor de nivel de combustible fuera de rango","Motor - Inyectores/Combustible","Sensor de nivel, flotante"),
            ("P0462","Señal baja en el sensor de nivel de combustible","Motor - Inyectores/Combustible","Cortocircuito a masa, sensor dañado"),
            ("P0463","Señal alta en el sensor de nivel de combustible","Motor - Inyectores/Combustible","Circuito abierto, sensor dañado"),
            ("P0500","Falla eléctrica en el sensor de velocidad del vehículo (VSS)","Transmisión","Sensor VSS, cableado"),
            ("P0501","Sensor VSS fuera de rango","Transmisión","Sensor descalibrado, cableado"),
            ("P0505","Falla en el sistema de control de marcha lenta (IAC)","Motor - Sensores/Admisión","Válvula IAC, cuerpo de aceleración sucio"),
            ("P0506","RPM de marcha lenta por debajo de lo esperado","Motor - Sensores/Admisión","Válvula IAC, fuga de vacío"),
            ("P0507","RPM de marcha lenta por encima de lo esperado","Motor - Sensores/Admisión","Válvula IAC trabada, fuga de vacío grande"),
            ("P0600","Falla en el enlace serial de comunicaciones del módulo","Módulo de control / Eléctrico","Cableado del bus de datos, módulo"),
            ("P0601","Error de suma de verificación en la memoria del módulo de control","Módulo de control / Eléctrico","Módulo de control con falla interna"),
            ("P0700","Avería general en el sistema de control de la transmisión","Transmisión","Ver códigos específicos de la TCM"),
            ("P0701","Sistema de control de la transmisión fuera de rango","Transmisión","Sensor o solenoide de la transmisión"),
            ("P0705","Falla eléctrica en el sensor de rango de la transmisión (PRNDL)","Transmisión","Sensor, cableado"),
            ("P0710","Falla eléctrica en el sensor de temperatura del fluido de la transmisión","Transmisión","Sensor, cableado"),
            ("P0715","Falla eléctrica en el sensor de velocidad de entrada / turbina","Transmisión","Sensor, cableado"),
            ("P0720","Falla eléctrica en el sensor de velocidad de salida de la transmisión","Transmisión","Sensor, cableado"),
            ("P0730","Relación de engranes incorrecta","Transmisión","Solenoides de cambio, fluido bajo o degradado"),
            ("P0740","Falla en el circuito del embrague del convertidor de par","Transmisión","Solenoide TCC, cableado"),
            ("P0750","Falla en el solenoide de cambios A","Transmisión","Solenoide, cableado"),
            ("P0755","Falla en el solenoide de cambios B","Transmisión","Solenoide, cableado"),
            ("P0205","Falla eléctrica en el inyector del cilindro 5","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0206","Falla eléctrica en el inyector del cilindro 6","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0207","Falla eléctrica en el inyector del cilindro 7","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0208","Falla eléctrica en el inyector del cilindro 8","Motor - Inyectores/Combustible","Inyector, cableado de ese cilindro"),
            ("P0221","Sensor TPS B fuera de rango","Motor - Sensores/Admisión","Sensor descalibrado, cableado dañado"),
            ("P0222","Señal baja en el sensor TPS B","Motor - Sensores/Admisión","Cortocircuito a masa, sensor dañado"),
            ("P0223","Señal alta en el sensor TPS B","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0224","Señal intermitente en el sensor TPS B","Motor - Sensores/Admisión","Conector flojo u oxidado, falso contacto"),
            ("P0231","Señal baja en el circuito secundario de la bomba de combustible","Motor - Inyectores/Combustible","Cortocircuito a masa, bomba, relé"),
            ("P0232","Señal alta en el circuito secundario de la bomba de combustible","Motor - Inyectores/Combustible","Circuito abierto, bomba, relé"),
            ("P0234","Sobrepresión de sobrealimentación (turbo)","Motor - Sensores/Admisión","Válvula wastegate, actuador del turbo"),
            ("P0261","Señal baja en el inyector del cilindro 1","Motor - Inyectores/Combustible","Cortocircuito a masa, inyector dañado"),
            ("P0262","Señal alta en el inyector del cilindro 1","Motor - Inyectores/Combustible","Circuito abierto, inyector dañado"),
            ("P0330","Falla eléctrica en el sensor de detonación 2, banco 2","Motor - Encendido/Combustión","Sensor, cableado"),
            ("P0331","Sensor de detonación 2 fuera de rango","Motor - Encendido/Combustión","Sensor descalibrado, cableado"),
            ("P0339","Señal intermitente en el sensor CKP","Motor - Encendido/Combustión","Conector flojo u oxidado, falso contacto"),
            ("P0343","Señal alta en el sensor CMP","Motor - Encendido/Combustión","Circuito abierto, sensor dañado"),
            ("P0344","Señal intermitente en el sensor CMP","Motor - Encendido/Combustión","Conector flojo u oxidado, falso contacto"),
            ("P0350","Falla general en el circuito primario/secundario de bobina de encendido","Motor - Encendido/Combustión","Bobina, cableado, módulo"),
            ("P0351","Falla en la bobina de encendido A","Motor - Encendido/Combustión","Bobina, cableado"),
            ("P0352","Falla en la bobina de encendido B","Motor - Encendido/Combustión","Bobina, cableado"),
            ("P0353","Falla en la bobina de encendido C","Motor - Encendido/Combustión","Bobina, cableado"),
            ("P0354","Falla en la bobina de encendido D","Motor - Encendido/Combustión","Bobina, cableado"),
            ("P0370","Falla en la señal de referencia de sincronización de alta resolución A","Motor - Encendido/Combustión","Sensor, cableado, rueda fónica"),
            ("P0380","Falla en la bujía/circuito calefactor (motores diésel)","Motor - Encendido/Combustión","Bujía de precalentamiento, relé, cableado"),
            ("P0410","Falla en el sistema de inyección de aire secundario","Emisiones","Bomba de aire secundario, válvulas, mangueras"),
            ("P0411","Caudal incorrecto en la inyección de aire secundario","Emisiones","Bomba de aire secundario, fugas"),
            ("P0480","Falla eléctrica en el circuito de control del ventilador de enfriamiento 1","Motor - Sensores/Admisión","Relé, motor del ventilador, cableado"),
            ("P0481","Falla eléctrica en el circuito de control del ventilador de enfriamiento 2","Motor - Sensores/Admisión","Relé, motor del ventilador, cableado"),
            ("P0510","Falla en el interruptor de mariposa en posición cerrada","Motor - Sensores/Admisión","Interruptor, cableado"),
            ("P0520","Falla eléctrica en el circuito de presión de aceite del motor","Motor - Sensores/Admisión","Sensor, cableado"),
            ("P0521","Presión de aceite del motor fuera de rango","Motor - Sensores/Admisión","Sensor descalibrado, nivel de aceite"),
            ("P0522","Voltaje bajo en la señal de presión de aceite del motor","Motor - Sensores/Admisión","Cortocircuito a masa, sensor dañado"),
            ("P0523","Voltaje alto en la señal de presión de aceite del motor","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0530","Falla eléctrica en el sensor de presión del refrigerante de A/C","Motor - Sensores/Admisión","Sensor, cableado"),
            ("P0532","Voltaje bajo en el sensor de presión del refrigerante de A/C","Motor - Sensores/Admisión","Cortocircuito a masa, sensor dañado"),
            ("P0533","Voltaje alto en el sensor de presión del refrigerante de A/C","Motor - Sensores/Admisión","Circuito abierto, sensor dañado"),
            ("P0534","Pérdida de carga de refrigerante del A/C","Motor - Sensores/Admisión","Fuga en el circuito de A/C"),
            ("P0560","Falla en el voltaje del sistema","Módulo de control / Eléctrico","Batería, alternador, cableado"),
            ("P0562","Voltaje del sistema bajo","Módulo de control / Eléctrico","Batería descargada, alternador"),
            ("P0563","Voltaje del sistema alto","Módulo de control / Eléctrico","Regulador de tensión, alternador"),
            ("P0602","Módulo de control sin programar","Módulo de control / Eléctrico","Requiere programación con equipo de diagnóstico"),
            ("P0603","Falla en la memoria KAM (no borrable) del módulo de control","Módulo de control / Eléctrico","Módulo de control con falla interna"),
            ("P0604","Falla en la memoria RAM del módulo de control","Módulo de control / Eléctrico","Módulo de control con falla interna"),
            ("P0605","Falla en la memoria ROM del módulo de control","Módulo de control / Eléctrico","Módulo de control con falla interna"),
            ("P0606","Falla en el procesador del módulo de control (PCM)","Módulo de control / Eléctrico","Módulo de control con falla interna"),
            ("P0620","Falla eléctrica en el circuito de control del generador/alternador","Módulo de control / Eléctrico","Alternador, cableado, regulador"),
            ("P0630","VIN no programado o no coincide con el ECM/PCM","Módulo de control / Eléctrico","Requiere reprogramación con equipo de diagnóstico"),
            ("P0650","Falla eléctrica en el circuito de la luz indicadora de fallas (MIL)","Módulo de control / Eléctrico","Bombilla, cableado, módulo"),
            ("P0703","Falla en el circuito del interruptor de freno / convertidor de par B","Transmisión","Interruptor de freno, cableado"),
            ("P0706","Sensor de rango de la transmisión fuera de rango","Transmisión","Sensor PRNDL descalibrado, cableado"),
            ("P0725","Falla en el circuito de entrada de velocidad del motor","Transmisión","Sensor, cableado"),
            ("P0731","Relación de engranes incorrecta en primera marcha","Transmisión","Solenoides de cambio, fluido bajo o degradado"),
            ("P0732","Relación de engranes incorrecta en segunda marcha","Transmisión","Solenoides de cambio, fluido bajo o degradado"),
            ("P0733","Relación de engranes incorrecta en tercera marcha","Transmisión","Solenoides de cambio, fluido bajo o degradado"),
            ("P0734","Relación de engranes incorrecta en cuarta marcha","Transmisión","Solenoides de cambio, fluido bajo o degradado"),
            ("P0743","Problema eléctrico en el embrague del convertidor de par","Transmisión","Solenoide TCC, cableado"),
            ("P0745","Falla en el solenoide de control de presión de la transmisión","Transmisión","Solenoide, cableado, fluido"),
            ("P0760","Falla en el solenoide de cambios C","Transmisión","Solenoide, cableado"),
            ("P0765","Falla en el solenoide de cambios D","Transmisión","Solenoide, cableado"),
            ("P0770","Falla en el solenoide de cambios E","Transmisión","Solenoide, cableado"),
            ("P0850","Falla en el interruptor de posición de estacionamiento/neutro","Transmisión","Interruptor, cableado"),
        ]
        c.executemany(
            "INSERT OR IGNORE INTO codigos_dtc (codigo, fabricante, descripcion, sistema, causas_posibles) "
            "VALUES (?, '', ?, ?, ?)",
            seed_dtc
        )

    # Fabricantes por WMI (los 3 primeros caracteres del VIN) precargados, para no tener que
    # ir cargándolos de a uno. Están los que circulan en Argentina: fabricación nacional,
    # importados de Brasil (mayoría del parque) y las marcas más comunes de otros orígenes.
    # Son editables desde la app: si alguno no coincide, se corrige ahí.
    # Antes esto solo corría con la tabla vacía, así que quien ya tenía la app nunca recibía los
    # WMI nuevos. Ahora se aplica una vez por versión de lista, con INSERT OR IGNORE: lo que vos
    # hayas cargado o corregido a mano NO se pisa nunca.
    c.execute("SELECT valor FROM configuracion WHERE clave = 'semilla_wmi_version'")
    _fila_sem = c.fetchone()
    if (_fila_sem["valor"] if _fila_sem else None) != SEMILLA_WMI_VERSION:
        seed_wmi = [
            # Argentina
            ('8AC', 'Mercedes-Benz Argentina', 'Argentina'),
            ('8AD', 'Peugeot Argentina', 'Argentina'),
            ('8AF', 'Ford Argentina', 'Argentina'),
            ('8AG', 'General Motors / Chevrolet Argentina', 'Argentina'),
            ('8AJ', 'Toyota Argentina', 'Argentina'),
            ('8AK', 'Suzuki Argentina', 'Argentina'),
            ('8AP', 'Fiat Argentina', 'Argentina'),
            ('8AW', 'Volkswagen Argentina', 'Argentina'),
            ('8A1', 'Renault Argentina', 'Argentina'),
            # Brasil
            ('935', 'Citroën Brasil', 'Brasil'),
            ('936', 'Peugeot Brasil', 'Brasil'),
            ('93H', 'Honda Brasil', 'Brasil'),
            ('93R', 'Toyota Brasil', 'Brasil'),
            ('93U', 'Audi Brasil', 'Brasil'),
            ('93V', 'Audi Brasil', 'Brasil'),
            ('93X', 'Mitsubishi Brasil', 'Brasil'),
            ('93Y', 'Renault Brasil', 'Brasil'),
            ('94D', 'Nissan Brasil', 'Brasil'),
            ('9BD', 'Fiat Brasil', 'Brasil'),
            ('9BF', 'Ford Brasil', 'Brasil'),
            ('9BG', 'General Motors / Chevrolet Brasil', 'Brasil'),
            ('9BM', 'Mercedes-Benz Brasil', 'Brasil'),
            ('9BR', 'Toyota Brasil', 'Brasil'),
            ('9BS', 'Scania Brasil', 'Brasil'),
            ('9BW', 'Volkswagen Brasil', 'Brasil'),
            # Chile
            ('8GD', 'Peugeot Chile', 'Chile'),
            ('8GG', 'Chevrolet Chile', 'Chile'),
            # Colombia
            ('9FB', 'Renault Colombia', 'Colombia'),
            # México
            ('3C4', 'Chrysler México', 'México'),
            ('3D3', 'Dodge México', 'México'),
            ('3FA', 'Ford México', 'México'),
            ('3FE', 'Ford México', 'México'),
            ('3G', 'General Motors México', 'México'),
            ('3H', 'Honda México', 'México'),
            ('3MZ', 'Mazda México', 'México'),
            ('3N', 'Nissan México', 'México'),
            ('3P3', 'Plymouth México', 'México'),
            ('3VW', 'Volkswagen México', 'México'),
            # Estados Unidos
            ('1B3', 'Dodge', 'Estados Unidos'),
            ('1C3', 'Chrysler', 'Estados Unidos'),
            ('1C6', 'Chrysler', 'Estados Unidos'),
            ('1D3', 'Dodge', 'Estados Unidos'),
            ('1FA', 'Ford', 'Estados Unidos'),
            ('1FB', 'Ford', 'Estados Unidos'),
            ('1FC', 'Ford', 'Estados Unidos'),
            ('1FD', 'Ford', 'Estados Unidos'),
            ('1FM', 'Ford (SUV)', 'Estados Unidos'),
            ('1FT', 'Ford (camionetas)', 'Estados Unidos'),
            ('1FU', 'Freightliner', 'Estados Unidos'),
            ('1FV', 'Freightliner', 'Estados Unidos'),
            ('1G', 'General Motors', 'Estados Unidos'),
            ('1GC', 'Chevrolet (camionetas)', 'Estados Unidos'),
            ('1GT', 'GMC (camionetas)', 'Estados Unidos'),
            ('1G1', 'Chevrolet', 'Estados Unidos'),
            ('1G2', 'Pontiac', 'Estados Unidos'),
            ('1G3', 'Oldsmobile', 'Estados Unidos'),
            ('1G4', 'Buick', 'Estados Unidos'),
            ('1G6', 'Cadillac', 'Estados Unidos'),
            ('1G8', 'Saturn', 'Estados Unidos'),
            ('1GM', 'Pontiac', 'Estados Unidos'),
            ('1GY', 'Cadillac', 'Estados Unidos'),
            ('1H', 'Honda', 'Estados Unidos'),
            ('1HD', 'Harley-Davidson (motos)', 'Estados Unidos'),
            ('1J4', 'Jeep', 'Estados Unidos'),
            ('1L', 'Lincoln', 'Estados Unidos'),
            ('1ME', 'Mercury', 'Estados Unidos'),
            ('1M1', 'Mack (camiones)', 'Estados Unidos'),
            ('1M2', 'Mack (camiones)', 'Estados Unidos'),
            ('1M3', 'Mack (camiones)', 'Estados Unidos'),
            ('1M4', 'Mack (camiones)', 'Estados Unidos'),
            ('1N', 'Nissan', 'Estados Unidos'),
            ('1NX', 'NUMMI (Toyota/GM)', 'Estados Unidos'),
            ('1P3', 'Plymouth', 'Estados Unidos'),
            ('1VW', 'Volkswagen', 'Estados Unidos'),
            ('1XK', 'Kenworth (camiones)', 'Estados Unidos'),
            ('1XP', 'Peterbilt (camiones)', 'Estados Unidos'),
            ('1YV', 'Mazda (AutoAlliance)', 'Estados Unidos'),
            ('1ZV', 'Ford (AutoAlliance)', 'Estados Unidos'),
            ('4F', 'Mazda', 'Estados Unidos'),
            ('4JG', 'Mercedes-Benz', 'Estados Unidos'),
            ('4M', 'Mercury', 'Estados Unidos'),
            ('4S', 'Subaru-Isuzu', 'Estados Unidos'),
            ('4T', 'Toyota', 'Estados Unidos'),
            ('4US', 'BMW', 'Estados Unidos'),
            ('4V1', 'Volvo (camiones)', 'Estados Unidos'),
            ('4V2', 'Volvo (camiones)', 'Estados Unidos'),
            ('4V4', 'Volvo (camiones)', 'Estados Unidos'),
            ('5F', 'Honda (Alabama)', 'Estados Unidos'),
            ('5L', 'Lincoln', 'Estados Unidos'),
            ('5N1', 'Nissan', 'Estados Unidos'),
            ('5NP', 'Hyundai', 'Estados Unidos'),
            ('5T', 'Toyota (camionetas)', 'Estados Unidos'),
            ('5YJ', 'Tesla', 'Estados Unidos'),
            ('538', 'Zero Motorcycles (motos)', 'Estados Unidos'),
            # Canadá
            ('2A4', 'Chrysler Canadá', 'Canadá'),
            ('2B3', 'Dodge Canadá', 'Canadá'),
            ('2B7', 'Dodge Canadá', 'Canadá'),
            ('2C3', 'Chrysler Canadá', 'Canadá'),
            ('2CN', 'CAMI (GM/Suzuki)', 'Canadá'),
            ('2D3', 'Dodge Canadá', 'Canadá'),
            ('2FA', 'Ford Canadá', 'Canadá'),
            ('2FB', 'Ford Canadá', 'Canadá'),
            ('2FC', 'Ford Canadá', 'Canadá'),
            ('2FM', 'Ford Canadá', 'Canadá'),
            ('2FT', 'Ford Canadá (camionetas)', 'Canadá'),
            ('2G', 'General Motors Canadá', 'Canadá'),
            ('2G1', 'Chevrolet Canadá', 'Canadá'),
            ('2G2', 'Pontiac Canadá', 'Canadá'),
            ('2G3', 'Oldsmobile Canadá', 'Canadá'),
            ('2G4', 'Buick Canadá', 'Canadá'),
            ('2HG', 'Honda Canadá', 'Canadá'),
            ('2HK', 'Honda Canadá', 'Canadá'),
            ('2HJ', 'Honda Canadá', 'Canadá'),
            ('2HM', 'Hyundai Canadá', 'Canadá'),
            ('2M', 'Mercury Canadá', 'Canadá'),
            ('2T', 'Toyota Canadá', 'Canadá'),
            ('2V4', 'Volkswagen Canadá', 'Canadá'),
            ('2V8', 'Volkswagen Canadá', 'Canadá'),
            # Alemania
            ('WAG', 'Neoplan (ómnibus)', 'Alemania'),
            ('WAU', 'Audi', 'Alemania'),
            ('WA1', 'Audi (SUV)', 'Alemania'),
            ('WBA', 'BMW', 'Alemania'),
            ('WBS', 'BMW M', 'Alemania'),
            ('WDA', 'Daimler', 'Alemania'),
            ('WDB', 'Mercedes-Benz', 'Alemania'),
            ('WDC', 'DaimlerChrysler', 'Alemania'),
            ('WDD', 'Mercedes-Benz', 'Alemania'),
            ('WDF', 'Mercedes-Benz (comerciales)', 'Alemania'),
            ('WEB', 'Evobus (ómnibus Mercedes)', 'Alemania'),
            ('WJM', 'Iveco Magirus', 'Alemania'),
            ('WF0', 'Ford Alemania', 'Alemania'),
            ('WMA', 'MAN (camiones)', 'Alemania'),
            ('WME', 'smart', 'Alemania'),
            ('WMW', 'MINI', 'Alemania'),
            ('WMX', 'Mercedes-AMG', 'Alemania'),
            ('WP0', 'Porsche', 'Alemania'),
            ('WP1', 'Porsche (SUV)', 'Alemania'),
            ('W0L', 'Opel', 'Alemania'),
            ('WUA', 'quattro GmbH (Audi)', 'Alemania'),
            ('WVG', 'Volkswagen (SUV/monovolumen)', 'Alemania'),
            ('WVW', 'Volkswagen', 'Alemania'),
            ('WV1', 'Volkswagen Comerciales', 'Alemania'),
            ('WV2', 'Volkswagen (furgones)', 'Alemania'),
            ('WV3', 'Volkswagen (camiones)', 'Alemania'),
            # Francia
            ('VF1', 'Renault', 'Francia'),
            ('VF2', 'Renault', 'Francia'),
            ('VF3', 'Peugeot', 'Francia'),
            ('VF4', 'Talbot', 'Francia'),
            ('VF6', 'Renault (camiones y ómnibus)', 'Francia'),
            ('VF7', 'Citroën', 'Francia'),
            ('VF8', 'Matra', 'Francia'),
            ('VLU', 'Scania Francia', 'Francia'),
            ('VN1', 'SOVAB (Renault)', 'Francia'),
            ('VNE', 'Irisbus', 'Francia'),
            ('VNK', 'Toyota Francia', 'Francia'),
            ('VNV', 'Renault-Nissan', 'Francia'),
            # España
            ('VSA', 'Mercedes-Benz España', 'España'),
            ('VSE', 'Suzuki España (Santana)', 'España'),
            ('VSK', 'Nissan España', 'España'),
            ('VSS', 'SEAT', 'España'),
            ('VSX', 'Opel España', 'España'),
            ('VS6', 'Ford España', 'España'),
            ('VS7', 'Citroën España', 'España'),
            ('VWA', 'Nissan España', 'España'),
            ('VWV', 'Volkswagen España', 'España'),
            # Italia
            ('ZAM', 'Maserati', 'Italia'),
            ('ZAP', 'Piaggio / Vespa / Gilera (motos)', 'Italia'),
            ('ZAR', 'Alfa Romeo', 'Italia'),
            ('ZCF', 'Iveco', 'Italia'),
            ('ZCG', 'Cagiva / MV Agusta (motos)', 'Italia'),
            ('ZDM', 'Ducati (motos)', 'Italia'),
            ('ZD4', 'Aprilia (motos)', 'Italia'),
            ('ZFA', 'Fiat', 'Italia'),
            ('ZFC', 'Fiat Veicoli Industriali', 'Italia'),
            ('ZFF', 'Ferrari', 'Italia'),
            ('ZGU', 'Moto Guzzi (motos)', 'Italia'),
            ('ZHW', 'Lamborghini', 'Italia'),
            ('ZLA', 'Lancia', 'Italia'),
            # Reino Unido
            ('SAL', 'Land Rover', 'Reino Unido'),
            ('SAJ', 'Jaguar', 'Reino Unido'),
            ('SAR', 'Rover', 'Reino Unido'),
            ('SB1', 'Toyota Reino Unido', 'Reino Unido'),
            ('SBM', 'McLaren', 'Reino Unido'),
            ('SCA', 'Rolls-Royce', 'Reino Unido'),
            ('SCB', 'Bentley', 'Reino Unido'),
            ('SCC', 'Lotus', 'Reino Unido'),
            ('SCF', 'Aston Martin', 'Reino Unido'),
            ('SDB', 'Peugeot Reino Unido', 'Reino Unido'),
            ('SFA', 'Ford Reino Unido', 'Reino Unido'),
            ('SHH', 'Honda Reino Unido', 'Reino Unido'),
            ('SHS', 'Honda Reino Unido', 'Reino Unido'),
            ('SJN', 'Nissan Reino Unido', 'Reino Unido'),
            ('SKF', 'Vauxhall', 'Reino Unido'),
            ('SMT', 'Triumph (motos)', 'Reino Unido'),
            # República Checa
            ('TMA', 'Hyundai República Checa', 'República Checa'),
            ('TMB', 'Škoda', 'República Checa'),
            ('TMT', 'Tatra (camiones)', 'República Checa'),
            # Hungría
            ('TRU', 'Audi Hungría', 'Hungría'),
            ('TSM', 'Suzuki Hungría', 'Hungría'),
            # Portugal
            ('TW1', 'Toyota Caetano', 'Portugal'),
            # Polonia
            ('SUF', 'Fiat Polonia', 'Polonia'),
            ('SUP', 'FSO-Daewoo', 'Polonia'),
            # Rumania
            ('UU1', 'Renault Dacia', 'Rumania'),
            # Eslovaquia
            ('U5Y', 'Kia Eslovaquia', 'Eslovaquia'),
            ('U6Y', 'Kia Eslovaquia', 'Eslovaquia'),
            # Austria
            ('VAG', 'Magna Steyr Puch', 'Austria'),
            ('VAN', 'MAN Austria', 'Austria'),
            ('VBK', 'KTM (motos)', 'Austria'),
            # Países Bajos
            ('XLB', 'Volvo (NedCar)', 'Países Bajos'),
            ('XLE', 'Scania Países Bajos', 'Países Bajos'),
            ('XLR', 'DAF (camiones)', 'Países Bajos'),
            ('XMC', 'Mitsubishi (NedCar)', 'Países Bajos'),
            # Bélgica
            ('YBW', 'Volkswagen Bélgica', 'Bélgica'),
            ('YCM', 'Mazda Bélgica', 'Bélgica'),
            # Suecia
            ('YS2', 'Scania (camiones)', 'Suecia'),
            ('YS3', 'Saab', 'Suecia'),
            ('YS4', 'Scania (ómnibus)', 'Suecia'),
            ('YV1', 'Volvo', 'Suecia'),
            ('YV4', 'Volvo', 'Suecia'),
            ('YV2', 'Volvo (camiones)', 'Suecia'),
            ('YV3', 'Volvo (ómnibus)', 'Suecia'),
            # Rusia
            ('XTA', 'Lada / AvtoVAZ', 'Rusia'),
            ('XTT', 'UAZ', 'Rusia'),
            ('X7L', 'Renault Rusia', 'Rusia'),
            # Serbia
            ('VX1', 'Zastava / Yugo', 'Serbia'),
            # Turquía
            ('NM0', 'Ford Turquía', 'Turquía'),
            ('NM4', 'Tofaş (Fiat Turquía)', 'Turquía'),
            ('NMT', 'Toyota Turquía', 'Turquía'),
            ('NLH', 'Hyundai Turquía', 'Turquía'),
            ('NLE', 'Mercedes-Benz Turquía (camiones)', 'Turquía'),
            # Japón
            ('JA', 'Isuzu', 'Japón'),
            ('JA3', 'Mitsubishi', 'Japón'),
            ('JA4', 'Mitsubishi', 'Japón'),
            ('JD', 'Daihatsu', 'Japón'),
            ('JF', 'Subaru', 'Japón'),
            ('JH', 'Honda', 'Japón'),
            ('JK', 'Kawasaki (motos)', 'Japón'),
            ('JL5', 'Mitsubishi Fuso (camiones)', 'Japón'),
            ('JMB', 'Mitsubishi', 'Japón'),
            ('JMY', 'Mitsubishi', 'Japón'),
            ('JMZ', 'Mazda', 'Japón'),
            ('JN', 'Nissan', 'Japón'),
            ('JS', 'Suzuki', 'Japón'),
            ('JT', 'Toyota', 'Japón'),
            ('JY', 'Yamaha (motos)', 'Japón'),
            # Corea del Sur
            ('KL', 'Daewoo / GM Corea', 'Corea del Sur'),
            ('KM', 'Hyundai', 'Corea del Sur'),
            ('KN', 'Kia', 'Corea del Sur'),
            ('KNM', 'Renault Samsung', 'Corea del Sur'),
            ('KPA', 'SsangYong', 'Corea del Sur'),
            ('KPT', 'SsangYong', 'Corea del Sur'),
            ('KM1', 'Hyosung (motos)', 'Corea del Sur'),
            ('KMY', 'Daelim (motos)', 'Corea del Sur'),
            # China
            ('LBE', 'Beijing Hyundai', 'China'),
            ('LDC', 'Dongfeng Peugeot Citroën', 'China'),
            ('LE4', 'Beijing Benz', 'China'),
            ('LFP', 'FAW', 'China'),
            ('LFV', 'FAW-Volkswagen', 'China'),
            ('LGB', 'Dongfeng', 'China'),
            ('LGX', 'BYD', 'China'),
            ('LJC', 'JAC', 'China'),
            ('LJ1', 'JAC', 'China'),
            ('LSG', 'Shanghai General Motors', 'China'),
            ('LSJ', 'MG / SAIC', 'China'),
            ('LSV', 'Shanghai Volkswagen', 'China'),
            ('LSY', 'Brilliance', 'China'),
            ('LTV', 'Toyota Tianjin', 'China'),
            ('LUC', 'GAC Honda', 'China'),
            ('LVS', 'Ford Chang An', 'China'),
            ('LVV', 'Chery', 'China'),
            ('LVZ', 'DFSK (Dongfeng Sokon)', 'China'),
            ('LZM', 'MAN China', 'China'),
            ('LZE', 'Isuzu Guangzhou', 'China'),
            ('LZG', 'Shaanxi (camiones)', 'China'),
            ('LZY', 'Yutong (ómnibus)', 'China'),
            ('LBB', 'Zhejiang Qianjiang / Keeway (motos)', 'China'),
            ('LCE', 'CFMOTO (motos)', 'China'),
            # India
            ('MAB', 'Mahindra', 'India'),
            ('MAC', 'Mahindra', 'India'),
            ('MA1', 'Mahindra', 'India'),
            ('MAJ', 'Ford India', 'India'),
            ('MAK', 'Honda India', 'India'),
            ('MAL', 'Hyundai India', 'India'),
            ('MAT', 'Tata', 'India'),
            ('MA3', 'Suzuki India (Maruti)', 'India'),
            ('MBH', 'Suzuki India (Maruti)', 'India'),
            ('MBJ', 'Toyota India', 'India'),
            ('MBR', 'Mercedes-Benz India', 'India'),
            ('MB1', 'Ashok Leyland', 'India'),
            ('MCA', 'Fiat India', 'India'),
            ('MDH', 'Nissan India', 'India'),
            ('MD2', 'Bajaj (motos)', 'India'),
            ('MEE', 'Renault India', 'India'),
            ('MEX', 'Volkswagen India', 'India'),
            # Indonesia
            ('MHF', 'Toyota Indonesia', 'Indonesia'),
            ('MHR', 'Honda Indonesia', 'Indonesia'),
            # Tailandia
            ('MLC', 'Suzuki Tailandia', 'Tailandia'),
            ('MLH', 'Honda Tailandia', 'Tailandia'),
            ('MMB', 'Mitsubishi Tailandia', 'Tailandia'),
            ('MMC', 'Mitsubishi Tailandia', 'Tailandia'),
            ('MMM', 'Chevrolet Tailandia', 'Tailandia'),
            ('MMT', 'Mitsubishi Tailandia', 'Tailandia'),
            ('MM8', 'Mazda Tailandia', 'Tailandia'),
            ('MNB', 'Ford Tailandia', 'Tailandia'),
            ('MNT', 'Nissan Tailandia', 'Tailandia'),
            ('MPA', 'Isuzu Tailandia', 'Tailandia'),
            ('MP1', 'Isuzu Tailandia', 'Tailandia'),
            ('MRH', 'Honda Tailandia', 'Tailandia'),
            ('MR0', 'Toyota Tailandia', 'Tailandia'),
            # Malasia
            ('PL1', 'Proton', 'Malasia'),
            # Filipinas
            ('PE1', 'Ford Filipinas', 'Filipinas'),
            ('PE3', 'Mazda Filipinas', 'Filipinas'),
            # Taiwán
            ('RFB', 'Kymco (motos)', 'Taiwán'),
            ('RFG', 'SYM (motos)', 'Taiwán'),
            # Sudáfrica
            ('AAV', 'Volkswagen Sudáfrica', 'Sudáfrica'),
            ('AC5', 'Hyundai Sudáfrica', 'Sudáfrica'),
            ('ADD', 'Hyundai Sudáfrica', 'Sudáfrica'),
            ('AFA', 'Ford Sudáfrica', 'Sudáfrica'),
            ('AHT', 'Toyota Sudáfrica', 'Sudáfrica'),
            # Australia
            ('6AB', 'MAN Australia', 'Australia'),
            ('6F4', 'Nissan Australia', 'Australia'),
            ('6F5', 'Kenworth Australia', 'Australia'),
            ('6FP', 'Ford Australia', 'Australia'),
            ('6G1', 'Holden (GM)', 'Australia'),
            ('6G2', 'Pontiac Australia', 'Australia'),
            ('6H8', 'Holden (GM)', 'Australia'),
            ('6MM', 'Mitsubishi Australia', 'Australia'),
            ('6T1', 'Toyota Australia', 'Australia'),
        ]
        c.executemany(
            "INSERT OR IGNORE INTO fabricantes_vin (wmi, fabricante, pais) VALUES (?, ?, ?)",
            seed_wmi
        )
        c.execute("INSERT INTO configuracion (clave, valor) VALUES ('semilla_wmi_version', ?) "
                  "ON CONFLICT(clave) DO UPDATE SET valor = excluded.valor", (SEMILLA_WMI_VERSION,))

    columnas_equiv = [f[1] for f in c.execute("PRAGMA table_info(equivalencias)").fetchall()]
    if "verificada" not in columnas_equiv:
        c.execute("ALTER TABLE equivalencias ADD COLUMN verificada INTEGER DEFAULT 0")
    if "nivel" not in columnas_equiv:
        c.execute("ALTER TABLE equivalencias ADD COLUMN nivel TEXT DEFAULT 'Exacta'")
    if "nota" not in columnas_equiv:
        c.execute("ALTER TABLE equivalencias ADD COLUMN nota TEXT")
    # De qué importación salió cada vínculo. Antes se perdía: el lote solo existía mientras la
    # equivalencia estaba pendiente, y al aprobarla quedaba huérfana. Así no había manera de
    # responder "¿de dónde salió esto?" ni de deshacer una lista que resultó estar mal cargada
    # — la única salida era borrar de a uno, a mano, entre miles.
    if "lote" not in columnas_equiv:
        c.execute("ALTER TABLE equivalencias ADD COLUMN lote TEXT")
    c.execute("CREATE INDEX IF NOT EXISTS idx_eq_lote ON equivalencias(lote)")
    # Confianza guardada de cada vínculo. Se calcula una vez y queda, para que el buscador pueda
    # arrastrarla por la cadena sin recalcular nada en cada búsqueda.
    if "confianza" not in columnas_equiv:
        c.execute("ALTER TABLE equivalencias ADD COLUMN confianza INTEGER")

    # Códigos con muchos vínculos que YA revisaste y están bien. Hay repuestos que legítimamente
    # equivalen a decenas: un filtro común, una bujía que va en media gama. Sin esta lista, esos
    # códigos aparecían como problema en cada revisión y no había forma de sacarlos del aviso.
    # Códigos que el fabricante reemplazó por otros. NO es una equivalencia común: tiene
    # dirección. El viejo se deja de fabricar y el nuevo lo reemplaza, pero no al revés.
    c.execute("""CREATE TABLE IF NOT EXISTS reemplazos_codigo (
        codigo_viejo TEXT NOT NULL,
        codigo_viejo_clean TEXT NOT NULL,
        codigo_nuevo TEXT NOT NULL,
        codigo_nuevo_clean TEXT NOT NULL,
        marca TEXT,
        nota TEXT,
        cargado_por TEXT,
        fecha TEXT DEFAULT (datetime('now')),
        PRIMARY KEY (codigo_viejo_clean, codigo_nuevo_clean)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_reemp_viejo ON reemplazos_codigo(codigo_viejo_clean)")

    c.execute("""CREATE TABLE IF NOT EXISTS puentes_aprobados (
        producto_id INTEGER PRIMARY KEY REFERENCES productos(id) ON DELETE CASCADE,
        aprobado_por TEXT,
        nota TEXT,
        fecha TEXT DEFAULT (datetime('now'))
    )""")

    columnas_historial = [f[1] for f in c.execute("PRAGMA table_info(historial_busquedas)").fetchall()]
    if "sin_resultado" not in columnas_historial:
        c.execute("ALTER TABLE historial_busquedas ADD COLUMN sin_resultado INTEGER DEFAULT 0")

    c.execute("CREATE INDEX IF NOT EXISTS idx_codigo_clean ON productos(codigo_clean)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_marca_id ON productos(marca_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_eq_a ON equivalencias(producto_a_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_eq_b ON equivalencias(producto_b_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_vehiculo_patente ON vehiculos(patente)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_historial_vehiculo ON historial_piezas(vehiculo_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_fecha ON auditoria_diaria(fecha)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_dtc_codigo ON codigos_dtc(codigo)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_vin_wmi ON fabricantes_vin(wmi)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_esquema_puntos ON esquema_puntos(esquema_id)")
    # Índices de las tablas más nuevas — se consultan seguido y no los tenían
    c.execute("CREATE INDEX IF NOT EXISTS idx_hist_busq_usuario ON historial_busquedas(usuario)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_hist_precios_prod ON historial_precios(producto_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_pedidos_repo_estado ON pedidos_reposicion(estado)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_uso_ia_fecha ON uso_ia(fecha)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_presup_mecanico ON presupuestos_mecanico(mecanico_id)")
    conn.commit()
    return conn


conn = get_connection()


class _CursorPorHilo:
    """Un cursor propio para cada persona que esté usando la app en ese momento.

    Antes había UNO SOLO compartido por todos. La conexión sí se puede compartir —SQLite la
    serializa por dentro—, pero el cursor no: es el que guarda el resultado pendiente. Streamlit
    atiende a cada persona en su propio hilo, así que dos consultas simultáneas se pisaban en el
    mismo cursor y una podía llevarse las filas de la otra. No es un riesgo teórico: en una
    prueba con dos hilos haciendo 400 consultas cada uno sobre un cursor compartido, 748 de las
    800 contestaron mal o directamente reventaron.

    Lo que se ve desde afuera no cambia: se sigue escribiendo c.execute(...), c.fetchall(),
    c.rowcount. Por eso el arreglo es este envoltorio y no tocar los cientos de lugares que lo
    usan. Y como cada hilo tiene el suyo, rowcount y lastrowid siguen siendo los de la consulta
    que uno mismo acaba de hacer, que es justamente lo que antes no estaba garantizado.

    El db_lock sigue haciendo falta igual: es el que evita que dos escrituras se mezclen."""

    def __init__(self, conexion):
        self._conexion = conexion
        self._propio = threading.local()

    @property
    def _cursor(self):
        cursor = getattr(self._propio, "cursor", None)
        if cursor is None:
            cursor = self._propio.cursor = self._conexion.cursor()
        return cursor

    def __getattr__(self, nombre):
        # Solo llega acá lo que no es atributo propio: execute, fetchone, fetchall,
        # executemany, rowcount, lastrowid, fetchmany.
        return getattr(self._cursor, nombre)

    def __iter__(self):
        return iter(self._cursor)


c = _CursorPorHilo(conn)




# ============================================================
# UTILIDADES
# ============================================================
def es_fecha_disfrazada(valor):
    """¿Esta celda es una fecha que en realidad era un código?

    Pasa todo el tiempo y en silencio: el proveedor abre la lista en Excel, y códigos como
    '12-15', '3/8' o '8-10' se convierten solos en fechas al guardar. Después llegan acá como
    una fecha de verdad, y limpiarlas normalmente da '20261215000000' — un código que no existe
    y que no va a coincidir con nada, sin que nadie se entere de por qué."""
    import datetime as _dt
    return isinstance(valor, (_dt.datetime, _dt.date))


def sanitizar(codigo):
    """Limpia un código dejando solo letras y números en mayúscula.

    Ojo con el '.0' del final: Excel guarda los códigos numéricos como número, así que
    '2776400' llega como '2776400.0'. Si no se saca antes de limpiar, queda '27764000'
    (un cero de más) y ese producto se vuelve imposible de encontrar por su código real.

    Y ojo también con la notación científica: un código largo que Excel muestra como
    '1.09E+11' limpiado a lo bruto queda '109E11', que no es ningún código."""
    if codigo is None:
        return ""
    if es_fecha_disfrazada(codigo):
        # Se devuelve vacío a propósito: es preferible que la fila se saltee y quede contada
        # como problema, antes que cargar un código inventado de 14 dígitos que nunca va a
        # coincidir con nada. El aviso al importar explica cómo arreglar el archivo.
        return ""
    codigo = str(codigo).strip()
    if codigo == "" or codigo.lower() == "nan":
        return ""
    if re.fullmatch(r"\d+\.0+", codigo):
        codigo = codigo.split(".")[0]
    # Notación científica: se pasa al número entero que representa
    if re.fullmatch(r"\d+(\.\d+)?[Ee][+-]?\d+", codigo):
        try:
            entero = int(float(codigo))
            if abs(float(codigo) - entero) < 1e-6:
                codigo = str(entero)
        except (ValueError, OverflowError):
            pass
    return re.sub(r'[^A-Z0-9]', '', codigo.upper())


LARGO_MINIMO_NUMERICO = 3  # un código que es SOLO números tiene que tener al menos esta cantidad


def es_codigo_util(texto):
    """¿Vale la pena cargar esto como código? Descarta lo que es solo 1 o 2 dígitos y nada más
    ('1', '12', '07', '3.0', '  2  ').

    Por qué: en muchas listas la columna de código trae metida la cantidad, el número de orden
    o la cantidad por bulto. Cada uno de esos números entraba como si fuera un producto, y como
    la app vincula todo lo que aparece en la misma fila, el '1' de una fila terminaba siendo
    "equivalente" al '1' de otra — y con eso, dos repuestos que no tienen nada que ver quedaban
    linkeados entre sí. Un solo número basura arrastra decenas de equivalencias falsas.

    Un código con al menos una letra se acepta aunque sea corto (A1, B2 existen de verdad)."""
    limpio = sanitizar(texto)
    if not limpio:
        return False
    if limpio.isdigit() and len(limpio) < LARGO_MINIMO_NUMERICO:
        return False
    # Tampoco entran los de DOS caracteres, tengan letra o no ("1S", "A1", "2S").
    # Esto estaba incoherente: la importación los dejaba pasar, y después la revisión marcaba
    # cada vínculo suyo con «es demasiado corto para ser un código». Con un solo "1S" de
    # IMPERIAL eso son cientos de pendientes con la misma alarma, uno por cada fila donde
    # apareció. Si el propio análisis dice que no es un código, no tiene por qué entrar.
    if len(limpio) < 3:
        return False
    return True


def _partir_por_barra(trozo):
    """La barra es el caso jodido: a veces separa dos códigos y a veces es PARTE del código.
    'W712/94' y 'WK842/2' son códigos Mann enteros, un filtro solo — partirlos ahí generaba dos
    productos falsos ('W712' y '94') y encima los dejaba vinculados entre sí como si fueran
    equivalentes. Pero '1109AN/1109AB' sí son dos códigos.

    La diferencia práctica: cuando la barra separa de verdad, los dos lados son códigos completos.
    Cuando es parte del código, del otro lado queda un numerito corto (el sufijo de la variante).
    Con espacios alrededor ('ABC / DEF') es separador seguro."""
    if "/" not in trozo:
        return [trozo]
    if re.search(r'\s/|/\s', trozo):          # 'ABC / DEF' → separador
        return [p for p in re.split(r'\s*/\s*', trozo) if p]
    partes = [p for p in trozo.split("/") if p]
    for parte in partes:
        limpio = sanitizar(parte)
        if limpio.isdigit() and len(limpio) <= 3:   # sufijo de variante: es un solo código
            return [trozo]
    return partes


def dividir_codigos(celda):
    """Separa una celda que puede traer varios códigos juntos (coma, punto y coma, barra, salto
    de línea). De paso descarta los pedazos que no son un código (ver es_codigo_util)."""
    if celda is None:
        return []
    texto = str(celda).strip()
    if texto == "" or texto.lower() == "nan":
        return []
    salida = []
    for trozo in re.split(r'[,;\n|]+', texto):
        trozo = trozo.strip()
        if not trozo:
            continue
        for parte in _partir_por_barra(trozo):
            parte = parte.strip()
            if parte and es_codigo_util(parte):
                salida.append(parte)
    return salida


def extraer_codigos_de_texto(texto, minimo=6):
    """Busca códigos de fábrica escondidos dentro de una descripción.
    Muchas listas de proveedor no traen una columna de OEM aparte, pero lo meten en el texto
    ('ROTULA VW GOL - ORIG 6Q0407365'). Esto lo saca de ahí.
    Es a propósito conservador — ante la duda, no lo toma — porque un código inventado ensucia
    la base con equivalencias falsas, que es peor que no tener la equivalencia:
      - descarta palabras sin números (ROTULA, DERECHA, DELANTERO)
      - descarta números solos cortos: años, medidas, cilindradas (2005, 1.6, 16V)
      - pide un largo mínimo, porque los códigos de fábrica son largos
    """
    if not texto:
        return []
    ruido = {"16V", "8V", "12V", "24V", "4X4", "4X2", "2WD", "4WD", "TDI", "TSI", "CRDI",
             "16valv", "MM", "CM", "KG"}

    # Formas que NO son códigos de repuesto y aparecen todo el tiempo en las descripciones.
    # Se aplican SOLO acá, cuando el código se está adivinando de un texto. Si el proveedor lo
    # puso en la columna del código, es un código y se respeta: la exigencia va donde estamos
    # suponiendo, no donde nos lo dijeron.
    formas_prohibidas = (
        # Códigos de MOTOR: letras, números y letras al final. MR20DE, B4204S, Z18XER, X20XEV,
        # DV6DTED, MT560B. Describen la motorización del auto, no la pieza — y como se repiten
        # en decenas de filas, cada uno vincula entre sí todo lo que lo menciona.
        # Se midió sobre 4.340 códigos extraídos de una lista real: descarta 36, y los 36 son
        # códigos de motor. El 1% de pérdida vale, porque cada uno de esos generaba decenas de
        # equivalencias falsas.
        re.compile(r'^[A-Z]{1,3}\d{1,5}[A-Z]{1,4}$'),
        # Medidas: 14X20X1, 7,5X12X5
        re.compile(r'^\d+([.,]\d+)?X\d+([.,]\d+)?(X\d+([.,]\d+)?)?$'),
        # Cilindradas y potencias sueltas: 1.6, 2.0TDI, 110CV
        re.compile(r'^\d[.,]\d[A-Z]*$'),
        re.compile(r'^\d+(CV|HP|KW|CC)$'),
    )
    encontrados = []
    for token in re.split(r'[\s,;/|()\[\]]+', str(texto)):
        limpio = token.strip().strip(".-_")
        if len(limpio) < minimo:
            continue
        if limpio.upper() in ruido:
            continue
        if any(p.match(limpio.upper()) for p in formas_prohibidas):
            continue
        if not any(ch.isdigit() for ch in limpio):
            continue
        tiene_letra = any(ch.isalpha() for ch in limpio)
        # Números sueltos: solo se aceptan si son largos (un código de barras o de fábrica),
        # así no se cuelan años ni medidas
        if not tiene_letra and len(re.sub(r'\D', '', limpio)) < 7:
            continue
        # Descartar cosas tipo "1.6" o "2.0TDI" que empiezan con cilindrada
        if re.match(r'^\d\.\d', limpio):
            continue
        if sanitizar(limpio):
            encontrados.append(limpio)
    # sin repetidos, conservando el orden
    vistos, salida = set(), []
    for cod in encontrados:
        clave = sanitizar(cod)
        if clave not in vistos:
            vistos.add(clave)
            salida.append(cod)
    return salida


def valor_o_vacio(valor):
    """Devuelve el valor de una celda como texto, o '' si está vacía.

    Le saca el '.0' que Excel le agrega a los códigos numéricos. Esto no es cosmético: el
    código que se guarda acá es el que después se MUESTRA y se copia en un presupuesto o en un
    mensaje de WhatsApp. Guardarlo como '2776400.0' significa mandarle al cliente un código que
    no existe. La búsqueda igual funcionaba porque sanitizar() lo limpiaba, así que el problema
    pasaba desapercibido hasta que alguien copiaba el código de la pantalla."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    if re.fullmatch(r"\d+\.0+", texto):
        return texto.split(".")[0]
    return texto


def valor_codigo(valor):
    """Lee una celda que TIENE que ser un código. Devuelve '' si es una fecha.

    Va aparte de valor_o_vacio a propósito. La detección de fechas tiene que hacerse sobre la
    celda cruda, ANTES de pasarla a texto: una vez convertida ya es '2026-12-15 00:00:00' y no
    hay forma de distinguirla de un código raro. Ese era el agujero — el chequeo estaba puesto
    más adelante en la cadena, cuando el dato ya se había perdido."""
    if es_fecha_disfrazada(valor):
        return ""
    return valor_o_vacio(valor)


def get_or_create_marca(nombre, tipo="PROVEEDOR"):
    nombre = nombre.strip().upper()
    c.execute("INSERT OR IGNORE INTO marcas (nombre, tipo) VALUES (?, ?)", (nombre, tipo))
    c.execute("SELECT id FROM marcas WHERE nombre = ?", (nombre,))
    return c.fetchone()[0]


def listar_catalogos_externos():
    c.execute("SELECT id, nombre, url FROM catalogos_externos ORDER BY nombre")
    return [dict(row) for row in c.fetchall()]


def agregar_catalogo_externo(nombre, url):
    nombre = nombre.strip()
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    with db_lock:
        c.execute("INSERT OR REPLACE INTO catalogos_externos (nombre, url) VALUES (?, ?)", (nombre, url))
        conn.commit()


def eliminar_catalogo_externo(catalogo_id):
    with db_lock:
        c.execute("DELETE FROM catalogos_externos WHERE id = ?", (catalogo_id,))
        conn.commit()


def depurar_huerfanos():
    """Borra productos que no tienen ninguna equivalencia vinculada (quedaron sueltos)."""
    with db_lock:
        c.execute("""
            DELETE FROM productos
            WHERE id NOT IN (SELECT DISTINCT producto_a_id FROM equivalencias)
              AND id NOT IN (SELECT DISTINCT producto_b_id FROM equivalencias)
        """)
        borrados = c.rowcount
        conn.commit()
    return borrados


def chequear_integridad_bd():
    """Revisa la base en busca de datos rotos o inconsistentes — sobre todo útil para detectar
    algo que se haya colado antes de que ciertas protecciones existieran, o algo que se rompió
    a mano editando la base fuera de la app."""
    resultados = []

    c.execute("""SELECT COUNT(*) FROM productos p
                 WHERE p.marca_id NOT IN (SELECT id FROM marcas)""")
    n = c.fetchone()[0]
    resultados.append({"Chequeo": "Productos con una marca que ya no existe", "Problemas": n})

    c.execute("""SELECT COUNT(*) FROM equivalencias e
                 WHERE e.producto_a_id NOT IN (SELECT id FROM productos)
                    OR e.producto_b_id NOT IN (SELECT id FROM productos)""")
    n = c.fetchone()[0]
    resultados.append({"Chequeo": "Equivalencias que apuntan a un producto que ya no existe", "Problemas": n})

    c.execute("""SELECT COUNT(*) FROM productos
                 WHERE codigo_raw IS NULL OR TRIM(codigo_raw) = ''
                    OR codigo_clean IS NULL OR TRIM(codigo_clean) = ''""")
    n = c.fetchone()[0]
    resultados.append({"Chequeo": "Productos con código vacío", "Problemas": n})

    c.execute("SELECT COUNT(*) FROM productos WHERE precio IS NOT NULL AND precio < 0")
    n = c.fetchone()[0]
    resultados.append({"Chequeo": "Productos con precio negativo", "Problemas": n})

    c.execute("SELECT COUNT(*) FROM productos WHERE stock IS NOT NULL AND stock < 0")
    n = c.fetchone()[0]
    resultados.append({"Chequeo": "Productos con stock negativo", "Problemas": n})

    c.execute("""SELECT codigo_clean, marca_id, COUNT(*) AS repetidos FROM productos
                 GROUP BY codigo_clean, marca_id HAVING COUNT(*) > 1""")
    duplicados = c.fetchall()
    resultados.append({"Chequeo": "Códigos duplicados dentro de la misma marca", "Problemas": len(duplicados)})

    c.execute("""SELECT COUNT(*) FROM vehiculos
                 WHERE km_registro IS NOT NULL AND km_actual IS NOT NULL AND km_actual < km_registro""")
    n = c.fetchone()[0]
    resultados.append({"Chequeo": "Vehículos con km actual menor al km de registro", "Problemas": n})

    c.execute("""SELECT COUNT(*) FROM historial_piezas h
                 WHERE h.vehiculo_id NOT IN (SELECT id FROM vehiculos)""")
    n = c.fetchone()[0]
    resultados.append({"Chequeo": "Piezas de historial que apuntan a un vehículo que ya no existe", "Problemas": n})

    return resultados


def listar_productos_sin_equivalencias(marca_filtro="Todas", limite=500):
    """Devuelve productos que no tienen ninguna equivalencia vinculada, sin borrarlos."""
    query = """
        SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
               m.nombre AS "Marca", m.tipo AS "Tipo"
        FROM productos p JOIN marcas m ON m.id = p.marca_id
        WHERE p.id NOT IN (SELECT DISTINCT producto_a_id FROM equivalencias)
          AND p.id NOT IN (SELECT DISTINCT producto_b_id FROM equivalencias)
    """
    params = []
    if marca_filtro and marca_filtro != "Todas":
        query += " AND UPPER(m.nombre) = ?"
        params.append(marca_filtro.upper())
    query += " ORDER BY m.nombre, p.codigo_raw LIMIT ?"
    params.append(limite)
    c.execute(query, params)
    return filas_a_listas(c)


def restaurar_backup(archivo_subido):
    """Reemplaza la base de datos actual por un archivo .db subido, de forma segura."""
    with db_lock:
        conn.commit()
        contenido = archivo_subido.read()
        conn.close()
        with open(DB_PATH, "wb") as f:
            f.write(contenido)
        # Muy importante: la conexión estaba cacheada por Streamlit. Si no limpiamos el
        # caché, la próxima vez que se pida se devolvería esta misma conexión ya cerrada.
        get_connection.clear()


def listar_marcas_con_conteo():
    c.execute("""SELECT m.id, m.nombre, m.tipo, COUNT(p.id) AS productos
                 FROM marcas m LEFT JOIN productos p ON p.marca_id = m.id
                 GROUP BY m.id ORDER BY m.nombre""")
    return c.fetchall()


def config_github():
    """Los datos para subir el backup solo. Devuelve None si no están configurados."""
    try:
        secretos = st.secrets if hasattr(st, "secrets") else {}
        token = secretos.get("github_token")
        repo = secretos.get("github_repo")      # formato: "usuario/repositorio"
    except Exception:
        return None
    if not token or not repo or "/" not in str(repo):
        return None
    return {"token": str(token), "repo": str(repo),
            "rama": str(secretos.get("github_rama", "main")),
            "archivo": str(secretos.get("github_archivo", ARCHIVO_SEMILLA))}


def subir_backup_a_github(datos_db, mensaje=""):
    """Sube la copia de la base al repositorio, que es lo único que sobrevive a un reinicio.

    Esta es la solución de fondo al problema que veníamos midiendo. Hasta ahora la app podía
    decirte «se perderían 180 productos», pero arreglarlo era un trámite manual: bajar el
    backup, entrar a GitHub, subirlo con el nombre exacto. Cuando algo depende de que una
    persona se acuerde de hacer un trámite, tarde o temprano no se hace.

    Necesita dos datos en los secretos de Streamlit (Settings → Secrets):
        github_token = "ghp_..."         (un token con permiso de escritura en el repo)
        github_repo  = "usuario/repo"

    Si no están configurados no hace nada y lo dice: no es obligatorio, es una mejora.
    Devuelve (ok, mensaje)."""
    cfg = config_github()
    if not cfg:
        return False, ("Falta configurar `github_token` y `github_repo` en los secretos de "
                       "Streamlit. Sin eso el backup hay que subirlo a mano.")
    import base64
    url = f"https://api.github.com/repos/{cfg['repo']}/contents/{cfg['archivo']}"
    cabeceras = {"Authorization": f"Bearer {cfg['token']}",
                 "Accept": "application/vnd.github+json"}
    try:
        # GitHub exige el sha del archivo que se reemplaza; si no existe todavía, se crea
        actual = requests.get(url, headers=cabeceras, params={"ref": cfg["rama"]}, timeout=20)
        sha = actual.json().get("sha") if actual.status_code == 200 else None

        cuerpo = {
            "message": mensaje or f"Backup automático {datetime.now():%Y-%m-%d %H:%M}",
            "content": base64.b64encode(datos_db).decode(),
            "branch": cfg["rama"],
        }
        if sha:
            cuerpo["sha"] = sha
        r = requests.put(url, headers=cabeceras, json=cuerpo, timeout=90)
        if r.status_code in (200, 201):
            guardar_config("ultimo_backup_github",
                           datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            marcar_backup_hecho()
            return True, f"Backup subido a {cfg['repo']} como `{cfg['archivo']}`."
        detalle = ""
        try:
            detalle = r.json().get("message", "")
        except Exception:
            pass
        if r.status_code in (401, 403):
            return False, ("GitHub rechazó el token. Revisá que tenga permiso de escritura "
                           f"sobre {cfg['repo']}. ({detalle})")
        if r.status_code == 404:
            return False, (f"GitHub no encuentra {cfg['repo']} o la rama «{cfg['rama']}». "
                           "Revisá el nombre. Si el repo es privado, el token tiene que "
                           "tener acceso.")
        return False, f"GitHub respondió {r.status_code}. {detalle}"
    except Exception as e:
        return False, f"No se pudo conectar con GitHub: {type(e).__name__}: {e}"


def cuanto_perderias_si_reinicia():
    """Qué se pierde si el servidor reinicia ahora mismo. Devuelve None si no aplica.

    Este es el riesgo más grande de todos y hasta ahora se contaba en abstracto: «bajate un
    backup». Pero un aviso genérico se ignora; un número no.

    Cómo funciona el mecanismo: el servidor borra el disco al reiniciar, y la app se restaura
    sola desde el archivo `datos_iniciales.db` que está en el repositorio. Ese archivo NO se
    actualiza solo — hay que bajar el backup y subirlo a GitHub a mano. Todo lo cargado desde
    la última vez que se hizo eso está viviendo únicamente en un disco que se borra."""
    try:
        c.execute("SELECT COUNT(*) FROM productos")
        ahora_total = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM equivalencias")
        eq_total = c.fetchone()[0]
    except sqlite3.OperationalError:
        return None

    if not os.path.exists(ARCHIVO_SEMILLA):
        # Sin copia en el repositorio se pierde TODO, no cero. Devolver ceros acá escondía
        # justamente el caso más grave.
        return {"hay_semilla": False, "productos_ahora": ahora_total, "productos_semilla": 0,
                "en_riesgo": ahora_total, "equivalencias_en_riesgo": eq_total,
                "fecha_semilla": ""}
    try:
        c.execute("SELECT COUNT(*) FROM productos")
        ahora = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM equivalencias")
        eq_ahora = c.fetchone()[0]

        origen = sqlite3.connect(f"file:{ARCHIVO_SEMILLA}?mode=ro", uri=True)
        cur = origen.cursor()
        cur.execute("SELECT COUNT(*) FROM productos")
        en_semilla = cur.fetchone()[0]
        try:
            cur.execute("SELECT COUNT(*) FROM equivalencias")
            eq_semilla = cur.fetchone()[0]
        except sqlite3.OperationalError:
            eq_semilla = 0
        origen.close()
    except Exception:
        return None

    return {
        "hay_semilla": True,
        "productos_ahora": ahora, "productos_semilla": en_semilla,
        "en_riesgo": max(ahora - en_semilla, 0),
        "equivalencias_en_riesgo": max(eq_ahora - eq_semilla, 0),
        "fecha_semilla": datetime.fromtimestamp(
            os.path.getmtime(ARCHIVO_SEMILLA)).strftime("%d/%m/%Y"),
    }


def tareas_automaticas_del_dia(presupuesto_segundos=6):
    """El mantenimiento que nadie tiene por qué acordarse de hacer.

    Todo lo que veníamos llamando «automático» corría solo cuando alguien abría esa pantalla en
    particular. Si nadie entra a Mantenimiento en dos semanas, los vínculos nuevos quedan sin
    puntuar y las fotos sin procesar. Esto lo hace solo, una vez por día, en la primera visita
    de cualquiera.

    Dos cosas la hacen segura:
      · Corre UNA vez por día. Queda anotado en la configuración, así que no se repite aunque
        entren veinte personas.
      · Tiene presupuesto de tiempo. Se revisa el reloj entre tarea y tarea y se corta al
        llegar al límite: lo que quedó pendiente se hace mañana. Nadie va a esperar diez
        segundos por un mantenimiento que no pidió.
    """
    hoy = datetime.now().strftime("%Y-%m-%d")
    if obtener_config("ultimas_tareas_dia", "") == hoy:
        return None

    arranque = time.time()
    hecho = []

    def queda_tiempo():
        return time.time() - arranque < presupuesto_segundos

    # 1. Puntuar los vínculos que quedaron sin nota: es lo que usa el buscador
    if queda_tiempo():
        try:
            c.execute("SELECT COUNT(*) FROM equivalencias WHERE confianza IS NULL")
            if c.fetchone()[0]:
                n = recalcular_confianzas(limite=3000)
                if n:
                    hecho.append(f"{n:,} vínculo(s) puntuados")
        except Exception:
            pass

    # 2. Procesar fotos pendientes, de a poco
    if queda_tiempo():
        try:
            r = migrar_imagenes_pendientes(limite=20)
            if r.get("listas"):
                hecho.append(f"{r['listas']} foto(s) procesadas")
        except Exception:
            pass

    # 3. Aprender modelos y motores de las fichas de vehículo cargadas
    if queda_tiempo():
        try:
            n_mod, n_mot = aprender_modelos_de_fichas_existentes()
            if n_mod or n_mot:
                hecho.append(f"{n_mod} modelo(s) y {n_mot} motor(es) de VIN aprendidos")
        except Exception:
            pass

    # 4. Las equivalencias que el mostrador ya confirmó: si vendiste el mismo reemplazo varias
    # veces, no tiene sentido esperar a que alguien entre a Mantenimiento a descubrirlo.
    if queda_tiempo():
        try:
            nuevas = [x for x in sustituciones_reales() if not x["_ya"]]
            if nuevas:
                n = guardar_equivalencias_derivadas(
                    [(x["_a"], x["_b"]) for x in nuevas],
                    f"CONFIRMADAS EN EL MOSTRADOR · {hoy}")
                if n:
                    hecho.append(f"{n} equivalencia(s) confirmadas por venta, a revisión")
        except Exception:
            pass

    # 5. Traer fotos de las fichas del proveedor, de a poco. Es la tarea más tediosa que queda
    # —son miles de productos— y la que mejor se presta a avanzar sola: una tanda chica por día
    # va completando el catálogo sin que nadie se siente a esperar. Se hace solo si está
    # activado, porque sale a internet y consume tiempo del proveedor.
    if queda_tiempo() and obtener_config("fotos_automaticas", "0") == "1":
        try:
            c.execute("""SELECT p.marca_id, COUNT(*) AS faltan FROM productos p
                         WHERE (p.imagen_url IS NULL OR p.imagen_url = '')
                           AND COALESCE(p.foto_busqueda_estado, '') = ''
                           AND COALESCE(p.stock, 0) > 0
                         GROUP BY p.marca_id ORDER BY faltan DESC LIMIT 1""")
            fila = c.fetchone()
            if fila:
                # OJO: devuelve una TUPLA (bajadas, fallidas, sin_foto), no un diccionario.
                # Tratarla como dict daba siempre cero y la tarea parecía no hacer nada.
                traidas, _fallidas, _sin_foto = bajar_fotos_desde_catalogo(
                    fila["marca_id"], limite=15, filtro="con_stock")
                if traidas:
                    hecho.append(f"{traidas} foto(s) traídas del proveedor")
        except Exception:
            pass

    # 6. Subir el backup al repositorio, si está configurado. Es lo único que sobrevive a un
    # reinicio del servidor, y depender de que alguien se acuerde de hacerlo a mano es
    # exactamente cómo se pierden las bases de datos.
    if queda_tiempo() and config_github():
        try:
            c.execute("SELECT COUNT(*) FROM productos")
            hay = c.fetchone()[0]
            riesgo = cuanto_perderias_si_reinicia()
            # Solo si hay algo nuevo que proteger: subir por subir gasta tiempo y llena el
            # historial del repositorio de commits iguales.
            if hay and riesgo and riesgo.get("en_riesgo", 0) > 0:
                ok, _ = subir_backup_a_github(
                    generar_backup_sin_fotos(),
                    f"Backup automático — {hay:,} productos")
                if ok:
                    hecho.append("backup subido al repositorio")
        except Exception:
            pass

    guardar_config("ultimas_tareas_dia", hoy)
    guardar_config("ultimas_tareas_detalle", " · ".join(hecho) if hecho else "nada pendiente")
    return hecho


def informe_post_importacion(lote, nombre_prov, cargados):
    """Qué pasó realmente con la lista que se acaba de importar, sin salir a buscarlo.

    Hasta ahora la app decía «se importaron 6.900 filas» y ahí terminaba. Los problemas —los
    vínculos malos, los precios raros, los productos que el proveedor dejó de mandar— había que
    salir a buscarlos por Mantenimiento, y en la práctica nadie lo hace hasta que algo falla en
    el mostrador.

    Esto corre los mismos controles que ya existen, pero acotados a lo que ACABA de entrar, y
    junta el resultado en un solo lugar mientras uno todavía tiene la lista fresca."""
    informe = {"puntos": [], "vinculos_nuevos": 0, "rojos": 0}

    try:
        c.execute("SELECT COUNT(*) FROM equivalencias_pendientes WHERE lote = ?", (lote,))
        pendientes = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM equivalencias WHERE lote = ?", (lote,))
        cargadas = c.fetchone()[0]
        informe["vinculos_nuevos"] = pendientes + cargadas
    except sqlite3.OperationalError:
        pendientes = cargadas = 0

    # Los vínculos que entraron, evaluados con el mismo análisis de siempre
    if pendientes:
        try:
            limpias, sospechosas = analizar_lote_pendiente(lote, limite=600)
            rojos = [x for x in (limpias + sospechosas) if x.get("confianza", 50) < 30]
            informe["rojos"] = len(rojos)
            if rojos:
                informe["puntos"].append((
                    "alto",
                    f"{len(rojos)} de los {pendientes} vínculos nuevos están «casi seguro mal»",
                    "Se pueden descartar todos juntos, sin mirarlos de a uno.",
                    "Estadísticas → 🔗 Equivalencias sugeridas",
                ))
        except Exception:
            pass

    # Códigos que no parecen códigos y entraron igual
    try:
        c.execute("""SELECT COUNT(*) FROM productos p JOIN marcas m ON m.id = p.marca_id
                     WHERE UPPER(m.nombre) = UPPER(?) AND LENGTH(p.codigo_clean) <= 3""",
                  (nombre_prov,))
        cortos = c.fetchone()[0]
        if cortos:
            informe["puntos"].append((
                "medio", f"{cortos} código(s) de 3 caracteres o menos en {nombre_prov.upper()}",
                "Suelen ser cantidades o números de orden que se colaron en la columna del código.",
                "Estadísticas → Mantenimiento → 🧹 Limpiar vínculos",
            ))
    except sqlite3.OperationalError:
        pass

    # Productos de esta marca que dejaron de venir en las últimas listas
    try:
        c.execute("SELECT id FROM marcas WHERE UPPER(nombre) = UPPER(?)", (nombre_prov,))
        fila_marca = c.fetchone()
        if fila_marca:
            discont, _ = productos_probablemente_discontinuados(marca_id=fila_marca["id"])
            con_stock = [x for x in discont if (x.get("_stock") or 0) > 0]
            if con_stock:
                informe["puntos"].append((
                    "medio",
                    f"{len(con_stock)} producto(s) con stock no vinieron en esta lista",
                    "Si el proveedor dejó de mandarlos, es mercadería que conviene liquidar.",
                    "Estadísticas → Reposición → Puede que ya no se fabriquen",
                ))
    except Exception:
        pass

    return informe


DIAS_VENCIMIENTO_RESERVA = 7


def vencer_reservas_viejas():
    """Libera las reservas que quedaron colgadas. Devuelve cuántas."""
    with db_lock:
        c.execute("""UPDATE reservas_stock SET estado = 'vencida'
                     WHERE estado = 'activa'
                       AND fecha < datetime('now', ?)""",
                  (f"-{DIAS_VENCIMIENTO_RESERVA} days",))
        n = c.rowcount
        conn.commit()
    return n


def reservar_stock(producto_id, cantidad, cliente="", nota=""):
    """Aparta unidades para un presupuesto. Devuelve (ok, mensaje).

    No deja reservar más de lo que hay libre: eso convertiría la reserva en otra forma de
    prometer lo que no tenés, que es justo lo que se quiere evitar."""
    cantidad = int(cantidad or 0)
    if cantidad <= 0:
        return False, "La cantidad tiene que ser mayor que cero."
    libre = stock_libre(producto_id)
    if libre is None:
        return False, "No encontré ese producto."
    if cantidad > libre:
        return False, (f"Solo quedan {libre} sin reservar. Si igual querés apartarlas, "
                       "primero liberá alguna reserva.")
    with db_lock:
        c.execute("""INSERT INTO reservas_stock (producto_id, cantidad, cliente, nota, reservado_por)
                     VALUES (?, ?, ?, ?, ?)""",
                  (producto_id, cantidad, (cliente or "").strip() or None,
                   (nota or "").strip() or None, obtener_usuario_actual()))
        conn.commit()
    return True, f"Se apartaron {cantidad} unidad(es)."


def stock_libre(producto_id):
    """Lo que queda realmente disponible: el stock menos lo reservado. None si no existe."""
    c.execute("SELECT stock FROM productos WHERE id = ?", (producto_id,))
    fila = c.fetchone()
    if not fila:
        return None
    try:
        c.execute("""SELECT COALESCE(SUM(cantidad), 0) FROM reservas_stock
                     WHERE producto_id = ? AND estado = 'activa'""", (producto_id,))
        reservado = c.fetchone()[0]
    except sqlite3.OperationalError:
        reservado = 0
    return max((fila["stock"] or 0) - reservado, 0)


def stock_libre_de_varios(ids):
    """Lo mismo que stock_libre() pero para muchos productos, con dos consultas en total.

    De a uno son dos consultas por producto. En una búsqueda que trae 50 equivalencias eso son
    200 idas a la base en CADA refresco de pantalla —o sea, cada vez que se toca cualquier botón
    de la sección—. Así son dos, sin importar cuántos resultados haya."""
    ids = [int(i) for i in ids if i is not None]
    if not ids:
        return {}
    marcas = ",".join("?" * len(ids))
    c.execute(f"SELECT id, COALESCE(stock, 0) AS stock FROM productos WHERE id IN ({marcas})", ids)
    total = {f["id"]: f["stock"] for f in c.fetchall()}
    try:
        c.execute(f"""SELECT producto_id, COALESCE(SUM(cantidad), 0) AS reservado
                      FROM reservas_stock
                      WHERE estado = 'activa' AND producto_id IN ({marcas})
                      GROUP BY producto_id""", ids)
        reservado = {f["producto_id"]: f["reservado"] for f in c.fetchall()}
    except sqlite3.OperationalError:
        reservado = {}
    return {i: max(v - reservado.get(i, 0), 0) for i, v in total.items()}


def reservas_activas(limite=200):
    """Lo que está apartado ahora mismo."""
    try:
        c.execute("""SELECT r.id AS "_id", p.codigo_raw AS "Código", m.nombre AS "Marca",
                            p.descripcion AS "Descripción", r.cantidad AS "Apartadas",
                            p.stock AS "Stock total", r.cliente AS "Cliente",
                            r.reservado_por AS "Quién", substr(r.fecha, 1, 16) AS "Desde",
                            CAST(julianday('now') - julianday(r.fecha) AS INTEGER) AS "Días"
                     FROM reservas_stock r
                     JOIN productos p ON p.id = r.producto_id
                     JOIN marcas m ON m.id = p.marca_id
                     WHERE r.estado = 'activa'
                     ORDER BY r.fecha DESC LIMIT ?""", (limite,))
        return filas_a_listas(c)
    except sqlite3.OperationalError:
        return []


def cerrar_reserva(reserva_id, vendida):
    """Se concretó la venta (descuenta del stock) o se cayó (solo libera)."""
    with db_lock:
        c.execute("SELECT producto_id, cantidad FROM reservas_stock WHERE id = ? AND estado='activa'",
                  (reserva_id,))
        fila = c.fetchone()
        if not fila:
            conn.commit()
            return False
        if vendida:
            c.execute("UPDATE productos SET stock = MAX(COALESCE(stock,0) - ?, 0) WHERE id = ?",
                      (fila["cantidad"], fila["producto_id"]))
        c.execute("UPDATE reservas_stock SET estado = ? WHERE id = ?",
                  ("vendida" if vendida else "cancelada", reserva_id))
        conn.commit()
    return True


def margen_de(precio_venta, precio_costo):
    """Cuánto queda de margen. Devuelve (porcentaje, pesos) o (None, None) si falta un dato.

    Se calcula sobre el PRECIO DE VENTA, que es como se mira el margen en el mostrador
    («de cada 100 que entran, me quedan 35»). Calcularlo sobre el costo da un número más
    grande y más lindo, pero no es el que sirve para decidir."""
    if not precio_venta or not precio_costo or precio_venta <= 0 or precio_costo <= 0:
        return None, None
    ganancia = precio_venta - precio_costo
    return (ganancia / precio_venta) * 100, ganancia


def agregar_margen(filas):
    """Le suma a cada resultado su margen, si están los dos precios.

    NO borra el costo de la fila, aunque el costo no se muestre. Antes lo borraba, y estas filas
    son las mismas que quedan guardadas en session_state entre refrescos: al segundo toque de
    cualquier botón ya no había costo, así que la columna Margen aparecía vacía y el aviso de
    «mejor margen» desaparecía. Sin error y sin aviso: se veía una vez y no volvía más.
    De que el costo no llegue a la pantalla se encarga quien la dibuja, sacando las claves que
    empiezan con guión bajo."""
    for f in filas:
        pct, pesos = margen_de(f.get("Precio"), f.get("_costo"))
        f["Margen"] = f"{pct:.0f}% (${pesos:,.0f})" if pct is not None else ""
    return filas


def mejor_margen_entre_equivalentes(res):
    """De los equivalentes con stock, cuál te deja más ganancia.

    No es lo mismo que el más barato: entre dos que sirven igual, el que te deja más margen
    puede ser el más caro para el cliente o el más barato. Este dato hoy no existía y la
    decisión se tomaba a ojo."""
    candidatos = [f for f in res
                  if (f.get("Stock") or 0) > 0 and f.get("Precio") and f.get("_costo")]
    if len(candidatos) < 2:
        return None
    def ganancia(f):
        return (f["Precio"] or 0) - (f["_costo"] or 0)
    mejor = max(candidatos, key=ganancia)
    peor = min(candidatos, key=ganancia)
    if ganancia(mejor) <= ganancia(peor):
        return None
    return {"codigo": mejor["Codigo"], "marca": mejor["Marca"],
            "ganancia": ganancia(mejor), "diferencia": ganancia(mejor) - ganancia(peor)}


def marcas_probablemente_duplicadas(limite=40):
    """Marcas que son la misma cargada dos veces. Devuelve pares para revisar.

    Pasa todo el tiempo: una lista viene como «MAHLE» y la siguiente como «MAHLE FILTER», o
    alguien la tipeó distinto. Quedan como dos proveedores separados, y ahí el buscador deja de
    encontrar equivalencias que en realidad existen — el mismo repuesto figura en dos lados sin
    ninguna relación entre sí.

    Se detecta por dos caminos, y ambos importan:
      · el NOMBRE se parece (una contiene a la otra, o difieren en pocas letras)
      · comparten CÓDIGOS: si dos marcas tienen los mismos códigos, son el mismo proveedor
        aunque se llamen distinto — que es el caso que ningún chequeo por nombre agarra."""
    c.execute("""SELECT m.id, m.nombre, COUNT(p.id) AS productos
                 FROM marcas m LEFT JOIN productos p ON p.marca_id = m.id
                 GROUP BY m.id HAVING COUNT(p.id) > 0""")
    marcas = [dict(r) for r in c.fetchall()]
    if len(marcas) < 2:
        return []

    salida = []
    for i, a in enumerate(marcas):
        for b in marcas[i + 1:]:
            na, nb = a["nombre"].upper().strip(), b["nombre"].upper().strip()
            motivo = None
            # Nombre: una contenida en la otra, o a un par de letras de distancia
            if na in nb or nb in na:
                motivo = "una contiene a la otra"
            elif _distancia_edicion(na.replace(" ", ""), nb.replace(" ", ""), tope=2) <= 2:
                motivo = "se escriben casi igual"

            # Códigos compartidos: la señal más fuerte, y la única que agarra los nombres
            # que no se parecen en nada
            c.execute("""SELECT COUNT(*) FROM productos pa
                         JOIN productos pb ON pa.codigo_clean = pb.codigo_clean
                         WHERE pa.marca_id = ? AND pb.marca_id = ?""", (a["id"], b["id"]))
            compartidos = c.fetchone()[0]
            menor = min(a["productos"], b["productos"])
            proporcion = compartidos / menor if menor else 0
            if proporcion >= 0.6 and compartidos >= 5:
                motivo = (f"comparten {compartidos} código(s): el "
                          f"{proporcion * 100:.0f}% de la más chica")
            if not motivo:
                continue
            salida.append({
                "Marca A": a["nombre"], "Productos A": a["productos"],
                "Marca B": b["nombre"], "Productos B": b["productos"],
                "Por qué": motivo, "Códigos en común": compartidos,
                "_ida": a["id"], "_idb": b["id"],
                "_peso": compartidos * 10 + (100 - abs(a["productos"] - b["productos"]) / 100),
            })
    salida.sort(key=lambda x: -x["_peso"])
    return salida[:limite]


def fusionar_productos(id_perdedor, id_ganador):
    """Junta dos productos que son el mismo en uno solo. Devuelve True si se fusionó.

    Hace falta porque productos tiene UNIQUE(codigo_clean, marca_id): si dos productos terminan
    con el mismo código en la misma marca, la base lo rechaza. Mover a la fuerza uno encima del
    otro tira IntegrityError y la operación se cae entera.

    Lo que NO se pierde: las equivalencias, los pendientes, las fotos y el precio y el stock que
    tuviera el que se va y le falte al que queda. Se descarta el registro duplicado, no los datos."""
    if id_perdedor == id_ganador:
        return False
    # Las equivalencias del que se va pasan al que queda, sin duplicar ni auto-vincular
    c.execute("""SELECT CASE WHEN producto_a_id = ? THEN producto_b_id ELSE producto_a_id END AS otro,
                        lote
                 FROM equivalencias WHERE producto_a_id = ? OR producto_b_id = ?""",
              (id_perdedor, id_perdedor, id_perdedor))
    for fila in c.fetchall():
        otro = fila["otro"]
        if otro == id_ganador:
            continue
        c.execute("INSERT OR IGNORE INTO equivalencias (producto_a_id, producto_b_id, lote) "
                  "VALUES (?, ?, ?)", (min(otro, id_ganador), max(otro, id_ganador), fila["lote"]))
    c.execute("DELETE FROM equivalencias WHERE producto_a_id = ? OR producto_b_id = ?",
              (id_perdedor, id_perdedor))

    for tabla, col_a, col_b in (("equivalencias_pendientes", "producto_a_id", "producto_b_id"),):
        try:
            c.execute(f"DELETE FROM {tabla} WHERE {col_a} = ? OR {col_b} = ?",
                      (id_perdedor, id_perdedor))
        except sqlite3.OperationalError:
            pass

    # Datos que el que queda podría no tener
    for tabla, columna in (("producto_fotos", "producto_id"), ("historial_precios", "producto_id")):
        try:
            c.execute(f"UPDATE {tabla} SET {columna} = ? WHERE {columna} = ?",
                      (id_ganador, id_perdedor))
        except sqlite3.OperationalError:
            pass
    c.execute("""UPDATE productos SET
                    precio = COALESCE(precio, (SELECT precio FROM productos WHERE id = ?)),
                    stock = COALESCE(stock, (SELECT stock FROM productos WHERE id = ?)),
                    descripcion = COALESCE(descripcion, (SELECT descripcion FROM productos WHERE id = ?))
                 WHERE id = ?""", (id_perdedor, id_perdedor, id_perdedor, id_ganador))
    c.execute("DELETE FROM productos WHERE id = ?", (id_perdedor,))
    return True


def fusionar_marcas(marca_origen_id, marca_destino_id):
    """Mueve todos los productos de una marca a otra y borra la marca origen.

    Devuelve (movidos, fusionados). Los que ya existían en la marca destino con el mismo código
    no se pueden mover —la base no admite dos veces el mismo código en una marca— así que se
    fusionan con el que ya estaba, conservando sus equivalencias. Antes esto no se contemplaba
    y la operación entera fallaba con IntegrityError, sin mover nada."""
    movidos = fusionados = 0
    with db_lock:
        c.execute("SELECT id, codigo_clean FROM productos WHERE marca_id = ?", (marca_origen_id,))
        productos_origen = [(r["id"], r["codigo_clean"]) for r in c.fetchall()]
        for pid, clean in productos_origen:
            c.execute("SELECT id FROM productos WHERE marca_id = ? AND codigo_clean = ?",
                      (marca_destino_id, clean))
            existente = c.fetchone()
            if existente:
                fusionar_productos(pid, existente["id"])
                fusionados += 1
            else:
                c.execute("UPDATE productos SET marca_id = ? WHERE id = ?", (marca_destino_id, pid))
                movidos += 1
        c.execute("DELETE FROM marcas WHERE id = ?", (marca_origen_id,))
        conn.commit()
    return movidos, fusionados


def aumentar_precios_por_marca(marca_id, porcentaje):
    """Sube (o baja, con porcentaje negativo) todos los precios cargados de una marca."""
    with db_lock:
        c.execute(
            "UPDATE productos SET precio = ROUND(precio * (1 + ? / 100.0), 2) "
            "WHERE marca_id = ? AND precio IS NOT NULL",
            (porcentaje, marca_id)
        )
        afectados = c.rowcount
        conn.commit()
    return afectados


def listar_favoritos_stock_bajo(umbral=2):
    c.execute("""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
                 m.nombre AS "Marca", p.precio AS "Precio", p.stock AS "Stock"
                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                 WHERE p.favorito = 1 AND (p.stock IS NULL OR p.stock <= ?)
                 ORDER BY p.stock ASC, p.codigo_raw""", (umbral,))
    return filas_a_listas(c)


def registrar_busqueda_sin_resultado(termino):
    with db_lock:
        c.execute("INSERT INTO historial_busquedas (termino, usuario, sin_resultado) VALUES (?, ?, 1)",
                   (termino, obtener_usuario_actual()))
        conn.commit()


# ============================================================
# EQUIVALENCIAS DESCUBIERTAS DESDE LAS VENTAS
# ============================================================
# La idea: si un cliente pide el código A y termina llevándose el B, eso es una equivalencia
# que pasó de verdad en el mostrador — aunque no figure en ningún catálogo. Cruzando lo que se
# pidió contra lo que se vendió, el sistema propone equivalencias candidatas para que el dueño
# las confirme. No inventa nada solo: propone, y una persona decide.

def registrar_venta(producto_id, termino_pedido=""):
    """Anota que un producto se vendió, y qué había pedido el cliente cuando lo pidió."""
    with db_lock:
        c.execute(
            "INSERT INTO ventas_registradas (producto_id, termino_pedido, codigo_pedido_clean, usuario) "
            "VALUES (?, ?, ?, ?)",
            (producto_id, termino_pedido.strip(), sanitizar(termino_pedido), obtener_usuario_actual())
        )
        conn.commit()


def red_de_equivalencias(codigo_clean):
    """IDs de todos los productos que ya están vinculados a ese código (directa o
    indirectamente). Sirve para no proponer como novedad algo que ya está cargado."""
    if not codigo_clean:
        return set()
    c.execute("""
    WITH RECURSIVE Red(id) AS (
        SELECT id FROM productos WHERE codigo_clean = ?
        UNION
        SELECT CASE WHEN eq.producto_a_id = re.id THEN eq.producto_b_id ELSE eq.producto_a_id END
        FROM equivalencias eq JOIN Red re ON (eq.producto_a_id = re.id OR eq.producto_b_id = re.id)
    )
    SELECT id FROM Red""", (codigo_clean,))
    return {r["id"] for r in c.fetchall()}


def descubrir_equivalencias_candidatas(min_veces=2, dias=180, minutos_ventana=20):
    """Devuelve pares (código pedido → producto vendido) que se repitieron y que todavía NO
    están cargados como equivalentes. Se arma de dos fuentes:
      1) Directa: el empleado marcó "se llevó este" sobre el resultado de una búsqueda.
      2) Deducida: se vendió algo justo después de que el mismo empleado buscara un código
         que no dio resultado — el caso típico de "no lo tengo, pero le doy este que sirve".
    Es solo una sugerencia: siempre la confirma una persona antes de que quede cargada."""
    conteos = {}

    def sumar(codigo_clean, termino, producto_id, veces, origen):
        if not codigo_clean or not producto_id:
            return
        clave = (codigo_clean, producto_id)
        actual = conteos.get(clave)
        if actual:
            actual["veces"] += veces
            actual["origenes"].add(origen)
        else:
            conteos[clave] = {"termino": termino, "veces": veces, "origenes": {origen}}

    # Fuente 1: marcado explícito en el mostrador
    c.execute("""SELECT codigo_pedido_clean AS cod, MAX(termino_pedido) AS termino,
                        producto_id AS pid, COUNT(*) AS veces
                 FROM ventas_registradas
                 WHERE codigo_pedido_clean IS NOT NULL AND codigo_pedido_clean <> ''
                   AND fecha >= datetime('now', ?)
                 GROUP BY codigo_pedido_clean, producto_id""", (f"-{dias} days",))
    for f in c.fetchall():
        sumar(f["cod"], f["termino"], f["pid"], f["veces"], "mostrador")

    # Fuente 2: venta poco después de una búsqueda sin resultado del mismo empleado
    c.execute("""SELECT h.termino AS termino, v.producto_id AS pid, COUNT(*) AS veces
                 FROM ventas_registradas v
                 JOIN historial_busquedas h
                   ON h.usuario = v.usuario
                  AND h.sin_resultado = 1
                  AND h.fecha <= v.fecha
                  AND h.fecha >= datetime(v.fecha, ?)
                 WHERE v.fecha >= datetime('now', ?)
                 GROUP BY h.termino, v.producto_id""",
              (f"-{minutos_ventana} minutes", f"-{dias} days"))
    for f in c.fetchall():
        sumar(sanitizar(f["termino"]), f["termino"], f["pid"], f["veces"], "deducida")

    if not conteos:
        return []

    descartadas = set()
    c.execute("SELECT codigo_clean, producto_id FROM equivalencias_descartadas")
    for f in c.fetchall():
        descartadas.add((f["codigo_clean"], f["producto_id"]))

    redes = {}
    candidatas = []
    for (codigo_clean, producto_id), datos in conteos.items():
        if datos["veces"] < min_veces or (codigo_clean, producto_id) in descartadas:
            continue
        if codigo_clean not in redes:
            redes[codigo_clean] = red_de_equivalencias(codigo_clean)
        if producto_id in redes[codigo_clean]:
            continue  # ya está cargado como equivalente, no es novedad

        c.execute("""SELECT p.codigo_raw, p.codigo_clean, p.descripcion, m.nombre AS marca
                     FROM productos p JOIN marcas m ON m.id = p.marca_id WHERE p.id = ?""", (producto_id,))
        prod = c.fetchone()
        if not prod or prod["codigo_clean"] == codigo_clean:
            continue  # se llevó exactamente lo que pidió, no hay nada nuevo

        c.execute("SELECT id FROM productos WHERE codigo_clean = ? LIMIT 1", (codigo_clean,))
        existente = c.fetchone()

        # Evidencia local que se puede calcular sola: si los dos productos tienen medidas
        # cargadas, compararlas es prueba física, no una suposición.
        if existente:
            coinciden, detalle_medidas = comparar_medidas_productos(existente["id"], producto_id)
            if coinciden is True:
                guardar_evidencia(codigo_clean, producto_id, "medidas", detalle_medidas)
            elif coinciden is False:
                guardar_evidencia(codigo_clean, producto_id, "medidas_no_coinciden", detalle_medidas)

        guardar_evidencia(codigo_clean, producto_id, "mostrador",
                           f"Se repitió {datos['veces']} vez/veces en el mostrador")
        evidencias = listar_evidencia(codigo_clean, producto_id)
        etiqueta_confianza, puntaje = nivel_por_evidencias(evidencias)

        candidatas.append({
            "codigo_pedido": datos["termino"] or codigo_clean,
            "codigo_clean": codigo_clean,
            "producto_id": producto_id,
            "codigo_vendido": prod["codigo_raw"],
            "marca_vendida": prod["marca"],
            "descripcion": prod["descripcion"] or "",
            "veces": datos["veces"],
            "origen": "mostrador" if "mostrador" in datos["origenes"] else "deducida",
            "pedido_ya_cargado": existente is not None,
            "id_pedido": existente["id"] if existente else None,
            "evidencias": evidencias,
            "confianza": etiqueta_confianza,
            "puntaje": puntaje,
        })

    # Primero lo mejor respaldado, y dentro de eso lo que más se repitió
    candidatas.sort(key=lambda x: (-x["puntaje"], -x["veces"]))
    return candidatas


# ---- Evidencia que respalda (o descarta) una equivalencia sugerida ----------------
# Regla de oro: NADA se carga solo. Cada sugerencia llega al panel con el detalle de en qué
# se basa, y una persona decide. Las fuentes están ordenadas de más a menos confiable.

PESOS_EVIDENCIA = {
    "catalogo_oficial": 100,   # el código aparece escrito en la ficha oficial del proveedor
    "lista_proveedor": 60,     # vinieron relacionados en un Excel del propio proveedor
    "medidas": 45,             # las medidas mecánicas cargadas coinciden
    "mostrador": 20,           # se repitió en ventas reales
}


def guardar_evidencia(codigo_clean, producto_id, tipo, detalle=""):
    with db_lock:
        c.execute("INSERT OR REPLACE INTO evidencia_equivalencia "
                   "(codigo_clean, producto_id, tipo, detalle) VALUES (?, ?, ?, ?)",
                   (codigo_clean, producto_id, tipo, detalle))
        conn.commit()


def listar_evidencia(codigo_clean, producto_id):
    c.execute("""SELECT tipo, detalle, fecha FROM evidencia_equivalencia
                 WHERE codigo_clean = ? AND producto_id = ?""", (codigo_clean, producto_id))
    return [dict(r) for r in c.fetchall()]


CAMPOS_MEDIDAS = [
    ("diametro_interno", "diám. interno"), ("diametro_externo", "diám. externo"),
    ("diametro_interno_cara_b", "diám. interno cara B"),
    ("diametro_externo_cara_b", "diám. externo cara B"),
    ("diametro_rosca_homocinetica", "diám. rosca"), ("diametro_copa", "diám. copa (base)"),
    ("diametro_copa_superior", "diám. copa (boca)"), ("largo_total", "largo total"),
    ("ancho", "ancho"),
]
COLUMNAS_MEDIDAS = ", ".join(cn for cn, _ in CAMPOS_MEDIDAS) + ", paso_rosca, cantidad_estrias"


def cargar_medidas_de_varios(ids):
    """Trae las medidas de muchos productos de una sola consulta.

    Por qué: analizar un lote pendiente hacía DOS consultas por cada par para comparar medidas.
    Con 400 pares son 800 consultas — por eso el análisis estaba topeado en 400. Precargando
    todo de una, revisar 5.000 pares cuesta casi lo mismo que revisar 400."""
    medidas = {}
    ids = list({int(i) for i in ids})
    for inicio in range(0, len(ids), 800):   # SQLite tiene tope de parámetros por consulta
        tanda = ids[inicio:inicio + 800]
        marcadores = ",".join("?" * len(tanda))
        c.execute(f"SELECT id, {COLUMNAS_MEDIDAS} FROM productos WHERE id IN ({marcadores})", tanda)
        for fila in c.fetchall():
            medidas[fila["id"]] = dict(fila)
    return medidas


def comparar_medidas(a, b, tolerancia_pct=3):
    """El núcleo de la comparación, sobre dos diccionarios de medidas ya leídos."""
    campos = CAMPOS_MEDIDAS
    if not a or not b:
        return None, "No se pudo leer alguno de los dos productos."

    comparadas, diferencias = [], []
    for campo, etiqueta in campos:
        va, vb = a[campo], b[campo]
        if va is None or vb is None:
            continue
        comparadas.append(etiqueta)
        if va == 0 or vb == 0:
            continue
        diferencia = abs(va - vb) / max(va, vb) * 100
        if diferencia > tolerancia_pct:
            diferencias.append(f"{etiqueta}: {va} vs {vb}")

    for campo, etiqueta in (("paso_rosca", "paso de rosca"), ("cantidad_estrias", "estrías")):
        va, vb = a[campo], b[campo]
        if va in (None, "") or vb in (None, ""):
            continue
        comparadas.append(etiqueta)
        if str(va).strip().upper() != str(vb).strip().upper():
            diferencias.append(f"{etiqueta}: {va} vs {vb}")

    if not comparadas:
        return None, "Ninguno de los dos tiene medidas cargadas todavía."
    if diferencias:
        return False, "NO coinciden: " + "; ".join(diferencias)
    return True, "Coinciden en " + ", ".join(comparadas)


def comparar_medidas_productos(id_a, id_b, tolerancia_pct=3):
    """Compara las medidas mecánicas cargadas de dos productos. Devuelve (coinciden, detalle).
    Es evidencia física real, no una suposición: si un retén mide distinto, no entra, punto.
    Si a alguno le faltan medidas cargadas, no dice nada — no es prueba ni a favor ni en contra."""
    c.execute(f"SELECT id, {COLUMNAS_MEDIDAS} FROM productos WHERE id = ?", (id_a,))
    a = c.fetchone()
    c.execute(f"SELECT id, {COLUMNAS_MEDIDAS} FROM productos WHERE id = ?", (id_b,))
    b = c.fetchone()
    return comparar_medidas(dict(a) if a else None, dict(b) if b else None, tolerancia_pct)


def verificar_en_catalogo_oficial(codigo_a_buscar, url_ficha, tiempo_maximo=8):
    """Abre la ficha oficial del proveedor y fija si el otro código aparece escrito ahí.
    Es la evidencia más fuerte que se puede conseguir sin que nadie opine: o el proveedor
    lo lista en su propia página, o no. Compara sin guiones ni espacios, porque cada catálogo
    los escribe distinto (6Q0 407 365 / 6Q0407365 / 6Q0-407-365).
    Devuelve (encontrado, detalle). 'encontrado' es None si no se pudo consultar la página."""
    import requests
    if not url_ficha:
        return None, "Esa marca no tiene cargada la dirección de su catálogo."
    objetivo = sanitizar(codigo_a_buscar)
    if not objetivo:
        return None, "Código no válido."
    try:
        respuesta = requests.get(
            url_ficha, timeout=tiempo_maximo,
            headers={"User-Agent": "Mozilla/5.0 (compatible; EquivalenciasElChavo/1.0)"}
        )
        if respuesta.status_code != 200:
            return None, f"La página respondió {respuesta.status_code} — no se pudo verificar."
        texto_plano = sanitizar(re.sub(r"<[^>]+>", " ", respuesta.text))
        if objetivo in texto_plano:
            return True, f"El código {codigo_a_buscar} aparece en la ficha oficial del proveedor."
        return False, (f"El código {codigo_a_buscar} NO aparece en esa ficha. Puede que el "
                        "proveedor no lo liste, o que la página cargue los datos aparte.")
    except Exception as e:
        return None, f"No se pudo consultar la página ({type(e).__name__})."


def nivel_por_evidencias(evidencias):
    """Traduce la evidencia acumulada a algo legible. Nunca da 'confirmada': eso lo decide
    una persona. Si hay evidencia EN CONTRA (ej: las medidas no coinciden), lo marca.

    Se llamaba nivel_de_confianza, igual que la de más abajo que recibe un PUNTAJE. Dos def con
    el mismo nombre no son un error de Python: la segunda simplemente pisa a la primera. Así que
    esta acá nunca llegaba a correr — la llamada de descubrir_equivalencias_candidatas() le
    pasaba la lista de evidencias a la que espera un número y reventaba con
    "'>=' not supported between instances of 'list' and 'int'", tirando abajo toda la pantalla
    de equivalencias sugeridas apenas había una candidata con evidencia."""
    tipos = {e["tipo"] for e in evidencias}
    if "medidas_no_coinciden" in tipos or "catalogo_no_lo_lista" in tipos:
        return "⛔ Con evidencia en contra", 0
    puntaje = sum(PESOS_EVIDENCIA.get(t, 0) for t in tipos)
    if puntaje >= 100:
        return "🟢 Respaldo fuerte", puntaje
    if puntaje >= 45:
        return "🟡 Respaldo medio", puntaje
    return "🔴 Solo por repetición", puntaje


def marcar_revision(pares, decision):
    """Recuerda la decisión tomada sobre un vínculo, en los dos sentidos. 'ok' = ya lo miré y
    está bien (no volver a marcarlo en la auditoría). 'rechazada' = no es equivalente (además
    de borrarlo, no se vuelve a crear aunque se reimporte la lista del proveedor)."""
    if not pares:
        return
    filas = []
    for a, b in pares:
        filas.append((a, b, decision, obtener_usuario_actual()))
        filas.append((b, a, decision, obtener_usuario_actual()))
    with db_lock:
        c.executemany(
            "INSERT OR REPLACE INTO equivalencias_revisadas "
            "(producto_a_id, producto_b_id, decision, revisado_por) VALUES (?, ?, ?, ?)", filas
        )
        conn.commit()


def pares_rechazados():
    c.execute("SELECT producto_a_id, producto_b_id FROM equivalencias_revisadas WHERE decision = 'rechazada'")
    return {(r["producto_a_id"], r["producto_b_id"]) for r in c.fetchall()}


def eliminar_equivalencia(par_a, par_b, recordar_rechazo=True):
    """Borra el vínculo en los dos sentidos. Los productos quedan; solo se corta la relación."""
    with db_lock:
        c.execute("DELETE FROM equivalencias WHERE (producto_a_id = ? AND producto_b_id = ?) "
                   "OR (producto_a_id = ? AND producto_b_id = ?)", (par_a, par_b, par_b, par_a))
        conn.commit()
    if recordar_rechazo:
        marcar_revision([(par_a, par_b)], "rechazada")


def auditar_equivalencias_existentes(limite=20000):
    """Pasa las alarmas por las equivalencias YA cargadas y las devuelve AGRUPADAS por el
    conflicto, no de a pares sueltos. La diferencia importa: si una importación mal mapeada
    creó un producto basura (código '1', por ejemplo) vinculado a cientos de códigos de
    fábrica, verlo de a pares es imposible de resolver — agrupado se ve de una y se corta.
    Lo ya revisado como correcto no vuelve a aparecer."""
    campos_medida = ["diametro_interno", "diametro_externo", "diametro_interno_cara_b",
                      "diametro_externo_cara_b", "diametro_rosca_homocinetica", "diametro_copa",
                      "diametro_copa_superior", "largo_total", "ancho"]
    etiquetas = {"diametro_interno": "diám. interno", "diametro_externo": "diám. externo",
                  "diametro_interno_cara_b": "diám. interno cara B",
                  "diametro_externo_cara_b": "diám. externo cara B",
                  "diametro_rosca_homocinetica": "diám. rosca", "diametro_copa": "diám. copa (base)",
                  "diametro_copa_superior": "diám. copa (boca)", "largo_total": "largo total",
                  "ancho": "ancho"}
    sel_a = ", ".join(f"pa.{campo} AS a_{campo}" for campo in campos_medida)
    sel_b = ", ".join(f"pb.{campo} AS b_{campo}" for campo in campos_medida)

    c.execute(f"""SELECT e.producto_a_id AS a, e.producto_b_id AS b,
                         pa.codigo_raw AS cod_a, pa.descripcion AS desc_a, ma.nombre AS marca_a, ma.tipo AS tipo_a,
                         pb.codigo_raw AS cod_b, pb.descripcion AS desc_b, mb.nombre AS marca_b, mb.tipo AS tipo_b,
                         pa.paso_rosca AS a_paso, pb.paso_rosca AS b_paso,
                         pa.cantidad_estrias AS a_estrias, pb.cantidad_estrias AS b_estrias,
                         {sel_a}, {sel_b}
                  FROM equivalencias e
                  JOIN productos pa ON pa.id = e.producto_a_id
                  JOIN productos pb ON pb.id = e.producto_b_id
                  JOIN marcas ma ON ma.id = pa.marca_id
                  JOIN marcas mb ON mb.id = pb.marca_id
                  WHERE e.producto_a_id < e.producto_b_id
                  LIMIT ?""", (limite,))
    filas = [dict(r) for r in c.fetchall()]

    c.execute("SELECT producto_a_id, producto_b_id FROM equivalencias_revisadas WHERE decision = 'ok'")
    ya_ok = {(r["producto_a_id"], r["producto_b_id"]) for r in c.fetchall()}

    # Cuántos vínculos tiene cada producto: sirve para detectar productos basura, que
    # terminan colgados de decenas o cientos de códigos de fábrica.
    vinculos_por_producto = {}
    for f in filas:
        for lado in ("a", "b"):
            pid = f[lado]
            info = vinculos_por_producto.setdefault(pid, {
                "id": pid, "codigo": f[f"cod_{lado}"], "descripcion": f[f"desc_{lado}"],
                "marca": f[f"marca_{lado}"], "tipo": f[f"tipo_{lado}"], "cantidad": 0
            })
            info["cantidad"] += 1

    # Conflicto: un código de fábrica que apunta a varios productos del MISMO proveedor
    grupos = {}
    for f in filas:
        if f["tipo_a"] == "OEM" and f["tipo_b"] != "OEM":
            oem_cod, oem_desc, otro_lado = f["cod_a"], f["desc_a"], "b"
        elif f["tipo_b"] == "OEM" and f["tipo_a"] != "OEM":
            oem_cod, oem_desc, otro_lado = f["cod_b"], f["desc_b"], "a"
        else:
            continue
        clave = (sanitizar(oem_cod), f[f"marca_{otro_lado}"])
        grupo = grupos.setdefault(clave, {
            "codigo_oem": oem_cod, "descripcion_oem": oem_desc or "",
            "marca_proveedor": f[f"marca_{otro_lado}"], "productos": {}
        })
        pid = f[otro_lado]
        grupo["productos"][pid] = {
            "id": pid, "codigo": f[f"cod_{otro_lado}"], "descripcion": f[f"desc_{otro_lado}"] or "",
            "par": (f["a"], f["b"]),
            "revisado_ok": (f["a"], f["b"]) in ya_ok,
            "vinculos_totales": vinculos_por_producto.get(pid, {}).get("cantidad", 0),
        }

    conflictos = []
    for clave, g in grupos.items():
        pendientes = [p for p in g["productos"].values() if not p["revisado_ok"]]
        if len(g["productos"]) > 1 and pendientes:
            g["productos"] = sorted(g["productos"].values(), key=lambda p: -p["vinculos_totales"])
            conflictos.append(g)
    conflictos.sort(key=lambda g: -len(g["productos"]))

    # Medidas contradictorias (esto sí tiene sentido de a pares)
    # Códigos que no parecen códigos: lo que deja una importación mal mapeada
    codigos_malos = {}
    for f in filas:
        if (f["a"], f["b"]) in ya_ok:
            continue
        for lado in ("a", "b"):
            pid = f[lado]
            if pid in codigos_malos:
                continue
            malo, motivo = codigo_sospechoso(f[f"cod_{lado}"], f.get(f"desc_{lado}", ""))
            if malo:
                codigos_malos[pid] = {
                    "id": pid, "codigo": f[f"cod_{lado}"], "marca": f[f"marca_{lado}"],
                    "motivo": motivo,
                    "vinculos": vinculos_por_producto.get(pid, {}).get("cantidad", 0),
                }

    por_medidas = []
    for f in filas:
        if (f["a"], f["b"]) in ya_ok:
            continue
        diferencias = []
        for campo in campos_medida:
            va, vb = f[f"a_{campo}"], f[f"b_{campo}"]
            if not va or not vb:
                continue
            if abs(va - vb) / max(va, vb) * 100 > 3:
                diferencias.append(f"{etiquetas[campo]}: {va} vs {vb}")
        for clave_c, etiqueta in (("paso", "paso de rosca"), ("estrias", "estrías")):
            va, vb = f[f"a_{clave_c}"], f[f"b_{clave_c}"]
            if va in (None, "") or vb in (None, ""):
                continue
            if str(va).strip().upper() != str(vb).strip().upper():
                diferencias.append(f"{etiqueta}: {va} vs {vb}")
        if diferencias:
            por_medidas.append({
                "a": f["a"], "b": f["b"], "cod_a": f["cod_a"], "desc_a": f["desc_a"] or "",
                "marca_a": f["marca_a"], "cod_b": f["cod_b"], "desc_b": f["desc_b"] or "",
                "marca_b": f["marca_b"], "detalle": "; ".join(diferencias),
            })

    # Productos con una cantidad de vínculos fuera de lo normal: candidatos a basura
    sospechosos = sorted(
        [v for v in vinculos_por_producto.values() if v["cantidad"] >= 10 and v["tipo"] != "OEM"],
        key=lambda v: -v["cantidad"]
    )

    # Si se llegó al tope, se revisó solo una parte: hay que decirlo, porque si no queda la
    # falsa sensación de que está todo limpio cuando ni siquiera se miró la mitad.
    c.execute("SELECT COUNT(*) FROM equivalencias WHERE producto_a_id < producto_b_id")
    total_en_base = c.fetchone()[0]

    return {
        "total_revisados": len(filas),
        "total_en_base": total_en_base,
        "quedo_corta": len(filas) >= limite and total_en_base > len(filas),
        "ya_revisados_ok": len(ya_ok),
        "conflictos": conflictos,
        "por_medidas": por_medidas,
        "codigos_malos": sorted(codigos_malos.values(), key=lambda x: -x["vinculos"])[:50],
        "productos_sospechosos": sospechosos[:30],
    }


def cb_auditoria_eliminar(par_a, par_b):
    """Borra el vínculo Y vuelve a calcular la auditoría. Sin esto último el panel seguía
    mostrando el resultado viejo (guardado desde que se tocó 'Auditar'), y parecía que el
    botón no hacía nada aunque el vínculo sí se hubiera borrado."""
    eliminar_equivalencia(par_a, par_b)
    st.session_state["resultado_auditoria"] = auditar_equivalencias_existentes()


def cb_auditoria_dejar(pares):
    marcar_revision(pares, "ok")
    st.session_state["resultado_auditoria"] = auditar_equivalencias_existentes()


def cb_auditoria_cortar_todos(producto_id):
    cortar_todos_los_vinculos(producto_id)
    st.session_state["resultado_auditoria"] = auditar_equivalencias_existentes()


def contar_vinculos_producto(producto_id):
    c.execute("SELECT COUNT(*) FROM equivalencias WHERE producto_a_id = ? OR producto_b_id = ?",
               (producto_id, producto_id))
    return c.fetchone()[0]


def cortar_todos_los_vinculos(producto_id, recordar_rechazo=True):
    """Corta TODAS las equivalencias de un producto de una sola vez. Para cuando quedó un
    producto basura de una importación mal mapeada colgado de decenas de códigos."""
    c.execute("""SELECT producto_a_id, producto_b_id FROM equivalencias
                 WHERE producto_a_id = ? OR producto_b_id = ?""", (producto_id, producto_id))
    pares = [(r["producto_a_id"], r["producto_b_id"]) for r in c.fetchall()]
    with db_lock:
        c.execute("DELETE FROM equivalencias WHERE producto_a_id = ? OR producto_b_id = ?",
                   (producto_id, producto_id))
        conn.commit()
    if recordar_rechazo and pares:
        marcar_revision(pares, "rechazada")
    return len(pares)


def guardar_equivalencias_pendientes(pares, origen, lote):
    """Guarda vínculos para revisar en vez de cargarlos directo."""
    if not pares:
        return 0
    with db_lock:
        c.executemany(
            "INSERT OR IGNORE INTO equivalencias_pendientes "
            "(producto_a_id, producto_b_id, origen, lote) VALUES (?, ?, ?, ?)",
            [(a, b, origen, lote) for a, b in pares]
        )
        conn.commit()
    return len(pares)


def resumen_lotes_pendientes():
    c.execute("""SELECT lote, origen, COUNT(*) AS cantidad, MIN(fecha) AS fecha
                 FROM equivalencias_pendientes GROUP BY lote, origen ORDER BY MIN(fecha) DESC""")
    return [dict(r) for r in c.fetchall()]


def contar_pendientes_del_lote(lote):
    c.execute("""SELECT COUNT(*) FROM equivalencias_pendientes
                 WHERE lote = ? AND producto_a_id < producto_b_id""", (lote,))
    return c.fetchone()[0]


MINIMO_PARA_APRENDER = 12      # decisiones necesarias antes de confiar en un patrón


def sustituciones_reales(minimo_veces=2, limite=300):
    """Qué código pediste y qué terminaste vendiendo. Es la evidencia más fuerte que hay.

    La app venía guardando esto en cada venta y no se usaba para nada. Y no es poca cosa: si un
    cliente pidió el código A y se le vendió el B, alguien del mostrador decidió que el B servía,
    el cliente se lo llevó y no volvió a reclamar. Eso pesa más que cualquier lista de
    proveedor — una lista dice lo que el proveedor cree; esto es lo que pasó de verdad.

    Se pide que haya ocurrido varias veces: una sola puede ser un error de tipeo o una venta
    por descarte."""
    c.execute("""SELECT v.codigo_pedido_clean AS pedido, v.producto_id AS vendido,
                        COUNT(*) AS veces, MAX(v.fecha) AS ultima
                 FROM ventas_registradas v
                 JOIN productos p ON p.id = v.producto_id
                 WHERE v.codigo_pedido_clean IS NOT NULL
                   AND v.codigo_pedido_clean <> ''
                   AND v.codigo_pedido_clean <> p.codigo_clean
                 GROUP BY v.codigo_pedido_clean, v.producto_id
                 HAVING COUNT(*) >= ?
                 ORDER BY COUNT(*) DESC LIMIT ?""", (minimo_veces, limite))
    filas = filas_a_listas(c)

    salida = []
    for f in filas:
        c.execute("""SELECT p.id, p.codigo_raw, p.descripcion, m.nombre AS marca
                     FROM productos p JOIN marcas m ON m.id = p.marca_id
                     WHERE p.codigo_clean = ? LIMIT 1""", (f["pedido"],))
        pedido = c.fetchone()
        c.execute("""SELECT p.id, p.codigo_raw, p.descripcion, m.nombre AS marca
                     FROM productos p JOIN marcas m ON m.id = p.marca_id
                     WHERE p.id = ?""", (f["vendido"],))
        vendido = c.fetchone()
        if not pedido or not vendido or pedido["id"] == vendido["id"]:
            continue
        c.execute("""SELECT 1 FROM equivalencias
                     WHERE producto_a_id = ? AND producto_b_id = ?""",
                  (min(pedido["id"], vendido["id"]), max(pedido["id"], vendido["id"])))
        ya_esta = c.fetchone() is not None
        salida.append({
            "Te pidieron": pedido["codigo_raw"], "Marca pedida": pedido["marca"],
            "Vendiste": vendido["codigo_raw"], "Marca vendida": vendido["marca"],
            "Veces": f["veces"], "Última": (f["ultima"] or "")[:10],
            "¿Ya vinculado?": "sí" if ya_esta else "no",
            "_a": pedido["id"], "_b": vendido["id"], "_ya": ya_esta,
        })
    return salida


def pares_confirmados_por_ventas(minimo_veces=2):
    """Los pares que la venta real confirmó, para usarlos como señal de confianza."""
    try:
        c.execute("""SELECT p1.id AS a, v.producto_id AS b, COUNT(*) AS veces
                     FROM ventas_registradas v
                     JOIN productos p ON p.id = v.producto_id
                     JOIN productos p1 ON p1.codigo_clean = v.codigo_pedido_clean
                     WHERE v.codigo_pedido_clean IS NOT NULL
                       AND v.codigo_pedido_clean <> ''
                       AND p1.id <> v.producto_id
                     GROUP BY p1.id, v.producto_id
                     HAVING COUNT(*) >= ?""", (minimo_veces,))
        return {(min(r["a"], r["b"]), max(r["a"], r["b"])): r["veces"] for r in c.fetchall()}
    except sqlite3.OperationalError:
        return {}


def aprender_de_las_decisiones():
    """Mira lo que fuiste aprobando y descartando, y saca patrones para usarlos después.

    Es la parte que faltaba para que el sistema mejore con el uso. Cada vez que aprobás o
    descartás un vínculo esa decisión se guardaba, pero no servía para nada. Y ahí hay
    información concreta sobre TU base:

      · combinaciones de marcas que casi siempre resultan bien (MAHLE ↔ OEM) o casi siempre mal
      · rubros que en tu catálogo sí se cruzan de verdad, aunque parezcan distintos

    Se exige un mínimo de decisiones antes de creerle a un patrón: con tres casos no se puede
    concluir nada, y una regla sacada de poca evidencia es peor que no tener regla."""
    patrones = {"marcas": {}, "total": 0}
    try:
        c.execute("""SELECT ma.nombre AS marca_a, mb.nombre AS marca_b, r.decision,
                            COUNT(*) AS n
                     FROM equivalencias_revisadas r
                     JOIN productos pa ON pa.id = r.producto_a_id
                     JOIN productos pb ON pb.id = r.producto_b_id
                     JOIN marcas ma ON ma.id = pa.marca_id
                     JOIN marcas mb ON mb.id = pb.marca_id
                     WHERE r.producto_a_id < r.producto_b_id
                     GROUP BY ma.nombre, mb.nombre, r.decision""")
        crudo = c.fetchall()
    except sqlite3.OperationalError:
        return patrones

    acumulado = {}
    for fila in crudo:
        clave = tuple(sorted((fila["marca_a"], fila["marca_b"])))
        d = acumulado.setdefault(clave, {"ok": 0, "rechazada": 0})
        d[fila["decision"]] = d.get(fila["decision"], 0) + fila["n"]

    for clave, d in acumulado.items():
        total = d["ok"] + d["rechazada"]
        patrones["total"] += total
        if total >= MINIMO_PARA_APRENDER:
            patrones["marcas"][clave] = {
                "tasa_ok": d["ok"] / total, "decisiones": total,
                "ok": d["ok"], "rechazadas": d["rechazada"],
            }
    return patrones


def senal_aprendida(marca_a, marca_b, patrones):
    """Cuánto suma o resta esta combinación de marcas, según cómo te fue antes con ella."""
    if not patrones or not patrones.get("marcas"):
        return 0, None
    dato = patrones["marcas"].get(tuple(sorted((marca_a or "", marca_b or ""))))
    if not dato:
        return 0, None
    tasa, n = dato["tasa_ok"], dato["decisiones"]
    if tasa >= 0.85:
        return 15, ("bien", f"📚 De {n} vínculos {marca_a}↔{marca_b} que revisaste, "
                             f"aprobaste el {tasa*100:.0f}%")
    if tasa <= 0.20:
        return -30, ("mal", f"📚 De {n} vínculos {marca_a}↔{marca_b} que revisaste, "
                             f"descartaste el {(1-tasa)*100:.0f}%")
    return 0, None


def evaluar_equivalencia(desc_a, desc_b, medidas_a=None, medidas_b=None,
                         precio_a=None, precio_b=None, veces_confirmada=1,
                         respaldo_fabricante=False, marca_a="", marca_b="", patrones=None,
                         vendido_como_reemplazo=0):
    """Pesa toda la evidencia disponible sobre un vínculo. Devuelve (puntaje 0-100, señales).

    La diferencia con lo que había antes: las alarmas eran una lista plana, así que 397 vínculos
    con alarma se veían todos igual de mal y había que mirarlos de a uno. Pero no valen lo mismo.
    Un vínculo entre un filtro y una pastilla de freno es basura segura; uno entre dos filtros
    con precios parecidos y confirmado por dos listas distintas es casi seguro bueno.

    Las señales, de la más fuerte a la más débil:
      · las medidas se contradicen  -> prueba física en contra, no hay vuelta
      · son de rubros distintos     -> un filtro no equivale a una pastilla
      · lo dice el fabricante       -> el catálogo de aplicaciones lo respalda
      · lo confirman varias listas  -> dos proveedores independientes coinciden
      · los precios no cierran      -> sospechoso, pero puede ser una diferencia de marca
    """
    puntaje = 50.0
    senales = []

    if medidas_a and medidas_b:
        coinciden, detalle = comparar_medidas(medidas_a, medidas_b)
        if coinciden is False:
            puntaje -= 45
            senales.append(("mal", f"📐 {detalle}"))
        elif coinciden is True:
            puntaje += 30
            senales.append(("bien", f"📐 {detalle}"))

    # Rubro: es la señal más barata y una de las que más basura caza. Si las descripciones
    # hablan de piezas de familias distintas, el vínculo no puede ser correcto.
    fam_a = clasificar_repuesto(desc_a) if desc_a else "Sin clasificar"
    fam_b = clasificar_repuesto(desc_b) if desc_b else "Sin clasificar"
    if fam_a != "Sin clasificar" and fam_b != "Sin clasificar":
        if fam_a != fam_b:
            puntaje -= 40
            senales.append(("mal", f"🧩 Son de rubros distintos: «{fam_a}» y «{fam_b}»"))
        else:
            puntaje += 15
            senales.append(("bien", f"🧩 Los dos son de «{fam_a}»"))

    # La venta real es la señal más fuerte de todas: no es lo que alguien cree que sirve, es
    # lo que efectivamente se vendió en su lugar y el cliente se llevó.
    if vendido_como_reemplazo >= 2:
        puntaje += 35
        senales.append(("bien", f"🧾 Ya lo vendiste como reemplazo {vendido_como_reemplazo} "
                                 "vez(ces): funcionó en el mostrador"))

    if respaldo_fabricante:
        puntaje += 25
        senales.append(("bien", "🏭 El catálogo del fabricante respalda este vínculo"))

    if veces_confirmada >= 2:
        puntaje += 20
        senales.append(("bien", f"📋 Lo confirman {veces_confirmada} listas distintas"))

    # Lo aprendido de tus propias decisiones anteriores
    ajuste, senal = senal_aprendida(marca_a, marca_b, patrones)
    if senal:
        puntaje += ajuste
        senales.append(senal)

    if precio_a and precio_b and precio_a > 0 and precio_b > 0:
        razon = max(precio_a, precio_b) / min(precio_a, precio_b)
        if razon >= 8:
            puntaje -= 25
            senales.append(("mal", f"💲 Los precios se diferencian {razon:.0f} veces"))
        elif razon <= 2:
            puntaje += 10
            senales.append(("bien", "💲 Los precios son parecidos"))

    return max(0.0, min(100.0, puntaje)), senales


def nivel_de_confianza(puntaje):
    """Traduce el puntaje a algo accionable, sin prometer de más."""
    if puntaje >= 75:
        return "🟢 Muy probable", "se puede aprobar sin mirar"
    if puntaje >= 55:
        return "🟡 Probable", "razonable, pero conviene una mirada"
    if puntaje >= 30:
        return "🟠 Dudosa", "revisala"
    return "🔴 Casi seguro mal", "descartala salvo que sepas que está bien"


def analizar_lote_pendiente(lote, limite=400, desde=0):
    """Revisa los vínculos de una importación y marca los sospechosos. Dos alarmas:
      - Las medidas mecánicas cargadas se contradicen (prueba física en contra).
      - Un mismo código de fábrica termina apuntando a dos productos distintos del MISMO
        proveedor: uno de los dos está mal, porque un proveedor no tiene dos piezas
        distintas para el mismo código original.
    Lo que no dispara ninguna alarma se considera limpio."""
    # Se traen también las descripciones: hacen falta para saber si un código que "parece una
    # medida" en realidad es un retén o un o'ring, donde la medida ES el código.
    c.execute("""SELECT ep.producto_a_id AS a, ep.producto_b_id AS b,
                        pa.codigo_raw AS cod_a, ma.nombre AS marca_a, ma.tipo AS tipo_a,
                        pa.descripcion AS desc_a, pa.precio AS precio_a,
                        pb.codigo_raw AS cod_b, mb.nombre AS marca_b, mb.tipo AS tipo_b,
                        pb.descripcion AS desc_b, pb.precio AS precio_b
                 FROM equivalencias_pendientes ep
                 JOIN productos pa ON pa.id = ep.producto_a_id
                 JOIN productos pb ON pb.id = ep.producto_b_id
                 JOIN marcas ma ON ma.id = pa.marca_id
                 JOIN marcas mb ON mb.id = pb.marca_id
                 WHERE ep.lote = ? AND ep.producto_a_id < ep.producto_b_id
                 ORDER BY ep.producto_a_id, ep.producto_b_id
                 LIMIT ? OFFSET ?""", (lote, limite, desde))
    filas = [dict(r) for r in c.fetchall()]

    # Todas las medidas de una sola vez, en vez de dos consultas por par
    medidas = cargar_medidas_de_varios([f["a"] for f in filas] + [f["b"] for f in filas])

    # Detectar códigos de fábrica que apuntan a varios productos del mismo proveedor
    apuntados = {}
    for f in filas:
        oem, otro, marca_otro = ((f["cod_a"], f["b"], f["marca_b"]) if f["tipo_a"] == "OEM"
                                  else (f["cod_b"], f["a"], f["marca_a"]))
        apuntados.setdefault((sanitizar(oem), marca_otro), set()).add(otro)
    ambiguos = {k for k, v in apuntados.items() if len(v) > 1}

    # ¿Este par aparece en más de una lista? Que dos proveedores independientes digan lo mismo
    # es la mejor confirmación que se puede tener sin mirar la pieza.
    ids = list({f["a"] for f in filas} | {f["b"] for f in filas})
    confirmaciones = {}
    codigos_con_respaldo = set()
    if ids:
        marcadores = ",".join("?" * len(ids))
        c.execute(f"""SELECT producto_a_id, producto_b_id, COUNT(DISTINCT lote) AS n
                      FROM equivalencias_pendientes
                      WHERE producto_a_id IN ({marcadores}) AND producto_b_id IN ({marcadores})
                      GROUP BY producto_a_id, producto_b_id""", ids * 2)
        for r in c.fetchall():
            confirmaciones[(r["producto_a_id"], r["producto_b_id"])] = r["n"]
        try:
            c.execute(f"""SELECT DISTINCT p.codigo_clean FROM productos p
                          JOIN aplicaciones ap ON ap.codigo_clean = p.codigo_clean
                          WHERE p.id IN ({marcadores})""", ids)
            codigos_con_respaldo = {r["codigo_clean"] for r in c.fetchall()}
        except sqlite3.OperationalError:
            codigos_con_respaldo = set()

    # Lo que se aprendió de las revisiones anteriores. Se calcula una vez para todo el lote.
    patrones_aprendidos = aprender_de_las_decisiones()
    ventas_confirman = pares_confirmados_por_ventas()

    limpias, sospechosas = [], []
    for f in filas:
        alarmas = []
        # ¿Los códigos parecen códigos? Esto caza las importaciones mal mapeadas, donde la
        # columna que se tomó como código en realidad tenía medidas o descripciones.
        for lado, etiqueta in (("a", "Código A"), ("b", "Código B")):
            malo, motivo = codigo_sospechoso(f[f"cod_{lado}"], f.get(f"desc_{lado}", ""))
            if malo:
                alarmas.append(f"🚫 {etiqueta}: {motivo}")
        coinciden, detalle = comparar_medidas(medidas.get(f["a"]), medidas.get(f["b"]))
        if coinciden is False:
            alarmas.append(f"📐 {detalle}")
        oem, marca_otro = ((f["cod_a"], f["marca_b"]) if f["tipo_a"] == "OEM"
                            else (f["cod_b"], f["marca_a"]))
        if (sanitizar(oem), marca_otro) in ambiguos:
            alarmas.append(f"⚠️ El código {oem} apunta a más de un producto de {marca_otro} — "
                            "alguno de los dos está mal cargado")
        # Puntaje de confianza: junta toda la evidencia en un número, para poder ordenar por
        # lo peor primero en vez de mirar cientos de alarmas planas.
        puntaje, senales = evaluar_equivalencia(
            f.get("desc_a", ""), f.get("desc_b", ""),
            medidas.get(f["a"]), medidas.get(f["b"]),
            f.get("precio_a"), f.get("precio_b"),
            veces_confirmada=confirmaciones.get((f["a"], f["b"]), 1),
            respaldo_fabricante=(sanitizar(f["cod_a"]) in codigos_con_respaldo
                                  or sanitizar(f["cod_b"]) in codigos_con_respaldo),
            marca_a=f.get("marca_a", ""), marca_b=f.get("marca_b", ""),
            patrones=patrones_aprendidos,
            vendido_como_reemplazo=ventas_confirman.get((min(f["a"], f["b"]),
                                                          max(f["a"], f["b"])), 0),
        )
        for tipo, texto in senales:
            if tipo == "mal" and texto not in alarmas:
                alarmas.append(texto)

        # Las alarmas estructurales (el código no parece un código, un OEM que apunta a dos
        # productos) descuentan fuerte: son problemas de carga, no matices.
        puntaje -= 35 * len([a for a in alarmas if a.startswith(("🚫", "⚠️"))])
        puntaje = max(0.0, min(100.0, puntaje))

        f["alarmas"] = alarmas
        f["confianza"] = puntaje
        f["senales"] = senales
        # El corte lo decide el PUNTAJE, no si hay alguna alarma. Antes bastaba una alarma
        # menor para mandar a revisión un vínculo con toda la evidencia a favor, y así se
        # juntaban cientos de casos que no hacía falta mirar mezclados con los que sí.
        (limpias if puntaje >= 55 else sospechosas).append(f)

    # Lo más dudoso primero: si hay que revisar 400, que los peores estén arriba
    sospechosas.sort(key=lambda x: x["confianza"])
    limpias.sort(key=lambda x: -x["confianza"])
    return limpias, sospechosas


def productos_que_mas_ensucian(lote, limite=15):
    """Qué productos son la causa de más vínculos pendientes marcados como problemáticos.

    Nace de un caso real: un producto con código «1S» generaba decenas de pendientes, todos con
    la misma alarma («es demasiado corto para ser un código»), y había que resolverlos de a uno.
    Cuando el problema es UN producto, corresponde resolverlo una vez y no cincuenta."""
    c.execute("""SELECT p.id AS "_id", p.codigo_raw AS "Código", m.nombre AS "Marca",
                        p.descripcion AS "Descripción",
                        COUNT(*) AS "Pendientes que genera"
                 FROM equivalencias_pendientes ep
                 JOIN productos p ON p.id IN (ep.producto_a_id, ep.producto_b_id)
                 JOIN marcas m ON m.id = p.marca_id
                 WHERE ep.lote = ?
                 GROUP BY p.id
                 HAVING COUNT(*) >= 3
                 ORDER BY COUNT(*) DESC LIMIT ?""", (lote, limite))
    filas = filas_a_listas(c)
    # Solo interesan los que además tienen un código dudoso: un producto legítimo con muchas
    # equivalencias no es un problema, es un repuesto que sirve para muchos autos.
    salida = []
    for f in filas:
        malo, motivo = codigo_sospechoso(f["Código"], f.get("Descripción") or "")
        if malo:
            f["Problema"] = motivo
            salida.append(f)
    return salida


def rechazar_pendientes_de_producto(producto_id, lote=None):
    """Descarta de una todos los pendientes que involucran a un producto. Devuelve cuántos."""
    with db_lock:
        if lote:
            c.execute("""SELECT producto_a_id, producto_b_id FROM equivalencias_pendientes
                         WHERE lote = ? AND (producto_a_id = ? OR producto_b_id = ?)""",
                      (lote, producto_id, producto_id))
        else:
            c.execute("""SELECT producto_a_id, producto_b_id FROM equivalencias_pendientes
                         WHERE producto_a_id = ? OR producto_b_id = ?""",
                      (producto_id, producto_id))
        pares = [(r["producto_a_id"], r["producto_b_id"]) for r in c.fetchall()]
        if lote:
            c.execute("""DELETE FROM equivalencias_pendientes
                         WHERE lote = ? AND (producto_a_id = ? OR producto_b_id = ?)""",
                      (lote, producto_id, producto_id))
        else:
            c.execute("""DELETE FROM equivalencias_pendientes
                         WHERE producto_a_id = ? OR producto_b_id = ?""",
                      (producto_id, producto_id))
        borrados = c.rowcount
        conn.commit()
    if pares:
        # Quedan registrados como rechazados para que una reimportación no los reviva
        marcar_revision(pares, "rechazada")
    return borrados


def aprobar_pendientes(lote, solo_estos_pares=None):
    """Pasa los vínculos pendientes a equivalencias reales."""
    with db_lock:
        if solo_estos_pares is None:
            c.execute("SELECT producto_a_id, producto_b_id FROM equivalencias_pendientes WHERE lote = ?", (lote,))
            pares = [(r["producto_a_id"], r["producto_b_id"]) for r in c.fetchall()]
        else:
            pares = list(solo_estos_pares)
        if not pares:
            return 0
        # El lote viaja con el vínculo: es lo que después permite deshacer toda una lista.
        # Y el par va siempre ordenado (menor primero), para no dejar la misma equivalencia
        # guardada dos veces en direcciones opuestas.
        c.executemany(
            "INSERT OR IGNORE INTO equivalencias (producto_a_id, producto_b_id, created_at, lote) "
            "VALUES (?, ?, datetime('now'), ?)",
            [(min(a, b), max(a, b), lote) for a, b in pares if a != b]
        )
        c.executemany("DELETE FROM equivalencias_pendientes WHERE producto_a_id = ? AND producto_b_id = ?", pares)
        conn.commit()
    # Queda registrado que ya se revisó, así la auditoría de lo existente no lo vuelve a marcar
    marcar_revision(pares, "ok")
    return len(pares)


def rechazar_pendientes(lote, solo_estos_pares=None):
    with db_lock:
        if solo_estos_pares is None:
            # Hay que leer los pares ANTES de borrarlos, si no queda sin registrar el rechazo
            c.execute("SELECT producto_a_id, producto_b_id FROM equivalencias_pendientes WHERE lote = ?", (lote,))
            marcar_para_recordar = [(r["producto_a_id"], r["producto_b_id"]) for r in c.fetchall()]
            c.execute("DELETE FROM equivalencias_pendientes WHERE lote = ?", (lote,))
            borrados = c.rowcount
        else:
            pares = list(solo_estos_pares)
            c.executemany("DELETE FROM equivalencias_pendientes WHERE producto_a_id = ? AND producto_b_id = ?", pares)
            borrados = len(pares)
            marcar_para_recordar = pares
        conn.commit()
    # Se recuerda el rechazo para que no vuelva a aparecer si se reimporta la misma lista
    marcar_revision(marcar_para_recordar, "rechazada")
    return borrados


def descartar_candidata(codigo_clean, producto_id):
    with db_lock:
        c.execute("INSERT OR REPLACE INTO equivalencias_descartadas "
                   "(codigo_clean, producto_id, descartado_por) VALUES (?, ?, ?)",
                   (codigo_clean, producto_id, obtener_usuario_actual()))
        conn.commit()


    with db_lock:
        c.execute("INSERT OR REPLACE INTO equivalencias_descartadas "
                   "(codigo_clean, producto_id, descartado_por) VALUES (?, ?, ?)",
                   (codigo_clean, producto_id, obtener_usuario_actual()))
        conn.commit()


def confirmar_candidata(codigo_clean, producto_id, codigo_original, marca_para_nuevo=None):
    """Convierte una sugerencia en una equivalencia real. Si el código que pedía el cliente
    todavía no existe como producto (caso típico: un código de fábrica que no tenés cargado),
    lo crea con la marca elegida y recién ahí los vincula."""
    with db_lock:
        c.execute("SELECT id FROM productos WHERE codigo_clean = ? LIMIT 1", (codigo_clean,))
        fila = c.fetchone()
        if fila:
            id_pedido = fila["id"]
        else:
            if not marca_para_nuevo:
                return False, "Elegí con qué marca cargar el código que pedía el cliente."
            marca_id = get_or_create_marca(marca_para_nuevo)
            c.execute("SELECT descripcion FROM productos WHERE id = ?", (producto_id,))
            desc = (c.fetchone() or {"descripcion": ""})["descripcion"] or ""
            id_pedido = get_or_create_producto(codigo_original.strip(), codigo_clean, desc, marca_id)

        if id_pedido == producto_id:
            return False, "Los dos códigos son el mismo producto."

        for a, b in ((id_pedido, producto_id), (producto_id, id_pedido)):
            c.execute(
                "INSERT OR REPLACE INTO equivalencias "
                "(producto_a_id, producto_b_id, created_at, verificada, nivel, nota) "
                "VALUES (?, ?, datetime('now'), 1, ?, ?)",
                (a, b, "Exacta", "Descubierta desde las ventas del mostrador")
            )
        c.execute("DELETE FROM equivalencias_descartadas WHERE codigo_clean = ? AND producto_id = ?",
                   (codigo_clean, producto_id))
        conn.commit()
    return True, None


def listar_busquedas_sin_resultado(limite=50):
    c.execute("""SELECT termino AS "Buscado", COUNT(*) AS "Veces", MAX(fecha) AS "Última vez"
                 FROM historial_busquedas WHERE sin_resultado = 1
                 GROUP BY termino ORDER BY COUNT(*) DESC, MAX(fecha) DESC LIMIT ?""", (limite,))
    return filas_a_listas(c)


def contar_productos_sin_equivalencias():
    c.execute("""
        SELECT COUNT(*) FROM productos
        WHERE id NOT IN (SELECT DISTINCT producto_a_id FROM equivalencias)
          AND id NOT IN (SELECT DISTINCT producto_b_id FROM equivalencias)
    """)
    return c.fetchone()[0]


# ============================================================
# COSAS QUE SE RESUELVEN SOLAS AL IMPORTAR
# ============================================================

# Palabras que suelen titular cada columna en las listas de proveedor. Se buscan en el
# encabezado para sugerir el mapeo sin que haya que elegirlo a mano cada vez.
PISTAS_COLUMNAS = {
    "prov":   ["COD", "ART", "REF", "NRO", "N°", "NUMERO", "PIEZA", "PART", "ITEM", "SKU",
                "INVENTORY"],
    # El código de barras va aparte y NUNCA puede quedar como código principal. En el mostrador
    # se pide el número de parte ("150000-R"), no el EAN: si el EAN ocupa el lugar del código,
    # el número real del repuesto no queda cargado en ningún lado y no se puede buscar.
    "ean":    ["EAN", "BARRA", "BARCODE", "GTIN", "UPC"],
    "oem":    ["OEM", "ORIG", "EQUIV", "CRUCE", "FABRICA", "FÁBRICA", "APLIC"],
    "desc":   ["DESC", "DETALLE", "PROD", "ARTICULO", "ARTÍCULO", "NOMBRE", "RUBRO"],
    "precio": ["PRECIO", "P.VENTA", "PVENTA", "P. VENTA", "IMPORTE", "VALOR", "LISTA",
                "COSTO", "NETO", "UNITARIO", "$"],
    "stock":  ["STOCK", "EXIST", "CANT", "DISPON", "SALDO", "DEPOSITO", "DEPÓSITO"],
}


def adivinar_columnas(encabezado):
    """Sugiere qué columna es cada cosa mirando los títulos. Devuelve un dict con los índices.

    Gana la pista MÁS LARGA que coincida, y una columna no puede tener dos roles.
    Sin eso, un título como «ARTICULO» quedaba como código (por contener «ART») y a la vez como
    descripción (por «ARTICULO»), y terminaba importando la descripción como si fuera el código.
    Comparando el largo, «ARTICULO» le gana a «ART» y cada columna cae donde corresponde.

    Para precio y stock gana la PRIMERA columna que coincide: las listas suelen traer varias
    (costo, lista, con IVA, con descuento) y la primera es casi siempre la que corresponde."""
    titulos = [str(x).upper().strip() if x else "" for x in encabezado]

    # Para cada columna, su mejor rol: el de la pista más larga que aparezca en el título
    mejor_rol = {}
    for i, titulo in enumerate(titulos):
        if not titulo:
            continue
        candidatos = []
        for clave, pistas in PISTAS_COLUMNAS.items():
            for pista in pistas:
                if pista in titulo:
                    # "EAN" pesa más que "COD" aunque midan lo mismo: un título como
                    # "CODIGO_EAN" tiene las dos, y sin esta prioridad el empate se resolvía a
                    # favor de "COD" y el código de barras terminaba ocupando el lugar del
                    # número de parte.
                    prioridad = 1 if clave == "ean" else 0
                    candidatos.append((prioridad, len(pista), clave))
        if candidatos:
            mejor_rol[i] = max(candidatos)[2]

    hallado = {"prov": None, "oem": None, "desc": None, "precio": None, "stock": None,
               "ean": None}
    for i, clave in mejor_rol.items():
        if clave in ("precio", "stock"):
            if hallado[clave] is None:      # la primera manda
                hallado[clave] = i
        else:
            hallado[clave] = i              # la última manda

    # El código de barras se carga como un código MÁS del producto, no como el principal. Así
    # escanear la caja encuentra el repuesto, y el número de parte sigue siendo el que se busca
    # y se muestra.
    if hallado["ean"] is not None:
        if hallado["prov"] == hallado["ean"]:
            hallado["prov"] = None
        if hallado["oem"] is None:
            hallado["oem"] = hallado["ean"]

    # El código de proveedor siempre tiene que apuntar a algo: si ningún título se reconoció,
    # la primera columna libre es la apuesta razonable.
    if hallado["prov"] is None:
        ocupadas = {v for k, v in hallado.items() if v is not None}
        hallado["prov"] = next((i for i in range(max(len(titulos), 1)) if i not in ocupadas), 0)
    if hallado["oem"] == hallado["prov"]:
        hallado["oem"] = None
    return hallado


def diagnosticar_lista(filas, header_row, idx_prov, idx_oem, idx_desc, muestra=300):
    """Simula la importación sobre las primeras filas y cuenta qué va a pasar con cada una.

    Es la respuesta a «se carga mal y no sé por qué». Antes uno mapeaba las columnas, importaba,
    y recién después descubría que el 99% había salido con alarmas — sin ninguna pista de en qué
    paso se rompió. Esto muestra ANTES lo que la app entendió de cada columna, con ejemplos
    reales del archivo, para poder darse cuenta de un vistazo si el mapeo apunta a donde debe."""
    datos = filas[header_row + 1:header_row + 1 + muestra]
    total = len(datos)
    r = {
        "total": total, "vacias": 0, "sin_codigo": 0, "codigo_basura": 0,
        "ok": 0, "con_oem": 0, "sospechosas": [], "fechas": 0, "ejemplos_fechas": [],
        "ejemplos_prov": [], "ejemplos_oem": [], "ejemplos_desc": [],
        "columnas_desiguales": 0,
    }
    if not total:
        return r

    ancho_encabezado = len([x for x in filas[header_row] if x is not None and str(x).strip()])

    for fila in datos:
        celdas = [x for x in fila if x is not None and str(x).strip() != ""]
        if not celdas:
            r["vacias"] += 1
            continue
        if ancho_encabezado and abs(len(celdas) - ancho_encabezado) > 2:
            r["columnas_desiguales"] += 1

        def celda(i, es_codigo=False):
            if i is None or i >= len(fila) or fila[i] is None:
                return ""
            if es_codigo:
                return valor_codigo(fila[i])
            return str(fila[i]).strip()

        # Fechas donde debería haber un código: Excel se las comió al guardar
        for i in (idx_prov, idx_oem):
            if i is not None and i < len(fila) and es_fecha_disfrazada(fila[i]):
                r["fechas"] += 1
                if len(r["ejemplos_fechas"]) < 5:
                    r["ejemplos_fechas"].append(str(fila[i])[:10])

        crudo_prov = celda(idx_prov, es_codigo=True)
        crudo_oem = celda(idx_oem, es_codigo=True)
        crudo_desc = celda(idx_desc)
        if len(r["ejemplos_prov"]) < 5 and crudo_prov:
            r["ejemplos_prov"].append(crudo_prov)
        if len(r["ejemplos_oem"]) < 5 and crudo_oem:
            r["ejemplos_oem"].append(crudo_oem)
        if len(r["ejemplos_desc"]) < 5 and crudo_desc:
            r["ejemplos_desc"].append(crudo_desc)

        if not crudo_prov:
            r["sin_codigo"] += 1
            continue
        codigos = dividir_codigos(crudo_prov)
        if not codigos:
            r["codigo_basura"] += 1
            if len(r["sospechosas"]) < 20:
                r["sospechosas"].append({
                    "Fila": filas.index(fila) + 1 if fila in filas else "?",
                    "Código leído": crudo_prov[:30],
                    "Motivo": "es un número suelto de 1 o 2 dígitos, o está vacío",
                })
            continue
        r["ok"] += 1
        if crudo_oem and dividir_codigos(crudo_oem):
            r["con_oem"] += 1
    return r


def _pinta_columna(ejemplos, esperado):
    """¿Los valores de esta columna se parecen a lo que se espera de ella?"""
    if not ejemplos:
        return "⚠️ vacía", "no trajo ningún valor en las filas de muestra"
    if esperado == "codigo":
        limpios = [sanitizar(e) for e in ejemplos]
        if all(l.isdigit() and len(l) <= 2 for l in limpios if l):
            return "❌ no son códigos", "son números sueltos (¿cantidad? ¿número de orden?)"
        # Lo que separa un código de una descripción no es el largo sino los ESPACIOS:
        # 'W712/94' y '036115561G' no tienen ninguno, 'FILTRO DE ACEITE FORD' tiene varios.
        # Mirando solo el largo, una descripción de 32 caracteres pasaba como código.
        con_frases = sum(1 for e in ejemplos if e.count(" ") >= 2)
        if con_frases >= len(ejemplos) / 2:
            return "❌ parece descripción", "los valores tienen varias palabras, no son códigos"
        largos = sum(1 for e in ejemplos if len(e) > 40)
        if largos >= len(ejemplos) / 2:
            return "❌ parece descripción", "los valores son frases largas, no códigos"
        if all(e.replace(".", "").replace(",", "").isdigit() and len(sanitizar(e)) <= 3
               for e in ejemplos if e):
            return "❌ no son códigos", "son números cortos (¿cantidad? ¿bulto?)"
        return "✅ parecen códigos", ""
    if esperado == "descripcion":
        # Una descripción real tiene más de una palabra
        sin_espacios = sum(1 for e in ejemplos if " " not in e)
        if sin_espacios >= len(ejemplos) / 2:
            return "⚠️ parecen códigos", "son una sola palabra, no descripciones"
        if all(len(e) <= 12 for e in ejemplos):
            return "⚠️ muy cortos", "parecen códigos, no descripciones"
        return "✅ parece descripción", ""
    return "✅", ""


def _perfil_de_columna(valores):
    """Mide una columna para poder adivinar qué es, sin depender del título."""
    llenos = [str(v).strip() for v in valores if v is not None and str(v).strip() != ""]
    if not llenos:
        return None
    distintos = len(set(llenos))
    numericos = sum(1 for v in llenos
                    if str(v).replace(".", "").replace(",", "").replace("-", "").isdigit())
    return {
        "llenado": len(llenos) / max(len(valores), 1),
        "unicidad": distintos / len(llenos),        # los códigos casi no se repiten
        "largo": sum(len(v) for v in llenos) / len(llenos),
        "espacios": sum(1 for v in llenos if " " in v) / len(llenos),
        "numerico": numericos / len(llenos),
        "muestra": llenos[:3],
    }


def adivinar_columnas_por_datos(filas_datos, ancho):
    """Deduce qué es cada columna mirando los VALORES, cuando la lista no trae títulos.

    Hace falta porque muchísimas listas de proveedor no tienen encabezado: arrancan directo en
    el primer producto. Sin títulos, la detección por palabras clave no tiene de dónde agarrarse
    y caía siempre en la columna 0 — que en estas listas está vacía, así que no entraba ni un
    producto.

    Lo que distingue a cada columna:
      descripción → es la de texto más largo y con espacios
      código      → casi no se repite (cada fila trae uno distinto)
      precio      → es numérica y SÍ se repite mucho (muchos productos valen lo mismo)
    Esa diferencia de repetición es la clave para separar el código del precio cuando los dos
    son números, que es el caso más difícil."""
    perfiles = {}
    for i in range(ancho):
        p = _perfil_de_columna([f[i] if i < len(f) else None for f in filas_datos])
        if p and p["llenado"] >= 0.3:
            perfiles[i] = p
    hallado = {"prov": None, "oem": None, "desc": None, "precio": None, "stock": None}
    if not perfiles:
        return hallado

    # Descripción: texto largo y con espacios
    candidatas_desc = {i: p for i, p in perfiles.items()
                       if p["largo"] >= 12 and p["espacios"] >= 0.5}
    if candidatas_desc:
        hallado["desc"] = max(candidatas_desc, key=lambda i: perfiles[i]["largo"])

    libres = [i for i in perfiles if i != hallado["desc"]]

    # Código: el más único de los que quedan
    if libres:
        hallado["prov"] = max(libres, key=lambda i: (perfiles[i]["unicidad"], -perfiles[i]["numerico"]))

    # Precio ANTES que el código de fábrica, y a propósito: es la columna numérica que queda.
    # Al revés, una lista de precios ordenada de menor a mayor tiene precios casi todos
    # distintos, así que pasaba por "código de fábrica" — y eso es lo peor que puede salir mal:
    # vincularía entre sí todos los productos que valen lo mismo.
    numericas = [i for i in libres if i != hallado["prov"] and perfiles[i]["numerico"] >= 0.9]
    if numericas:
        hallado["precio"] = min(numericas, key=lambda i: perfiles[i]["unicidad"])

    # Código de fábrica: solo si queda una columna muy única y que NO sea numérica pura.
    # Ante la duda se deja vacío: que falte es un inconveniente, que esté mal genera
    # equivalencias falsas, y eso ensucia la base para siempre.
    for i in libres:
        if i in (hallado["prov"], hallado["precio"]):
            continue
        p = perfiles[i]
        if p["unicidad"] > 0.8 and p["espacios"] < 0.5 and p["numerico"] < 0.9:
            hallado["oem"] = i
            break
    return hallado


def adivinar_proveedor(nombre_archivo, marcas_conocidas=()):
    """Saca el nombre del proveedor del nombre del archivo.

    Las listas llegan como 'ILLINOIS 17 07 2026.xlsx' o 'Lista_MAHLE_agosto.xlsx': el proveedor
    está ahí escrito y no hay razón para volver a tipearlo cada vez. Primero se busca alguna de
    las marcas que ya existen en la base (lo más confiable); si no aparece ninguna, se limpian
    fechas, números y palabras de relleno y se toma lo que queda."""
    base = re.sub(r'\.[A-Za-z0-9]{2,5}$', '', str(nombre_archivo or "").strip())
    if not base:
        return ""
    texto = re.sub(r'[_\-]+', ' ', base).upper()

    for marca in sorted(marcas_conocidas, key=len, reverse=True):
        if marca and len(marca) >= 3 and marca.upper() in texto:
            return marca.upper()

    relleno = {"LISTA", "LISTAS", "PRECIOS", "PRECIO", "CATALOGO", "CATÁLOGO", "ACTUALIZADA",
               "ACTUALIZADO", "NUEVA", "NUEVO", "COPIA", "FINAL", "VIGENTE", "DE", "DEL", "LA",
               "EL", "Y", "CON", "SIN", "VENTA", "MAYORISTA", "XLSX", "XLS", "CSV", "PDF"}
    meses = {"ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO",
             "SEPTIEMBRE", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"}
    palabras = [p for p in re.split(r'[^A-ZÁÉÍÓÚÑ0-9]+', texto) if p]
    utiles = [p for p in palabras
              if not p.isdigit() and p not in relleno and p not in meses and len(p) >= 3]
    return " ".join(utiles[:2]) if utiles else ""


def estado_del_backup():
    """Qué se cargó desde el último backup. Devuelve un dict con lo que haga falta para avisar.

    Esto importa más acá que en un sistema normal: el hosting borra el disco cuando reinicia y
    la app se restaura desde la última copia del repositorio. Todo lo cargado después de esa
    copia se pierde, y uno se entera recién cuando busca un código y no aparece."""
    c.execute("SELECT COUNT(*) FROM productos")
    total = c.fetchone()[0]
    marca = obtener_config("ultimo_backup_fecha", "")

    if not marca:
        return {"hay_backup": False, "productos_nuevos": total, "importaciones": None,
                "dias": None, "urgente": total > 0}

    desde = int(obtener_config("ultimo_backup_productos", "0") or 0)
    c.execute("SELECT COUNT(*) FROM importaciones WHERE fecha > ?", (marca,))
    importaciones = c.fetchone()[0]
    try:
        dias = (datetime.now() - datetime.strptime(marca[:19], "%Y-%m-%d %H:%M:%S")).days
    except Exception:
        dias = None

    nuevos = max(total - desde, 0)
    # Se avisa por cualquiera de las tres: productos nuevos, listas importadas, o tiempo.
    # Una lista importada puede cambiar miles de precios sin sumar un solo producto, así que
    # contar solo los productos nuevos dejaría pasar justo el caso que más duele perder.
    urgente = nuevos >= 50 or importaciones >= 1 or (dias is not None and dias >= 14)
    return {"hay_backup": True, "productos_nuevos": nuevos, "importaciones": importaciones,
            "dias": dias, "urgente": urgente}


def marcar_backup_hecho():
    """Se llama al descargar un backup: deja la marca para poder avisar cuando se atrase."""
    c.execute("SELECT COUNT(*) FROM productos")
    guardar_config("ultimo_backup_productos", str(c.fetchone()[0]))
    guardar_config("ultimo_backup_fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    invalidar_salud()


# Archivos que genera la propia app. Reimportar uno de estos es un error que no da ningún
# aviso y ensucia la base: subir el filas_omitidas.xlsx crea una marca llamada "FILAS OMITIDAS"
# con los productos que justamente se habían descartado.
ARCHIVOS_QUE_GENERA_LA_APP = {
    "filas_omitidas": "las filas que se descartaron en una importación anterior",
    "codigos_basura": "los códigos basura que marcaste para borrar",
    "precios_frenados": "los precios que se frenaron por parecer un error",
    "precios_incoherentes": "el listado de precios que no cerraban",
    "vinculos_dudosos": "los vínculos que el análisis marcó como dudosos",
    "productos_sin_equivalencias": "el listado de productos sin vínculos",
    "datos_iniciales": "el backup de la base",
}


def es_archivo_generado_por_la_app(nombre_archivo):
    """¿Este archivo lo generó la app? Devuelve (sí/no, qué es).

    No es una restricción caprichosa: estos exports son listados de diagnóstico, no listas de
    proveedor. Importarlos vuelve a meter en la base justo lo que se había separado."""
    base = re.sub(r'\.[A-Za-z0-9]{2,5}$', '', str(nombre_archivo or "")).strip().lower()
    base = re.sub(r'[\s\-]+', '_', base)
    base = re.sub(r'_?\d{6,8}$', '', base)      # el sufijo de fecha que llevan algunos
    for clave, que_es in ARCHIVOS_QUE_GENERA_LA_APP.items():
        if base == clave or base.startswith(clave):
            return True, que_es
    return False, ""


def huella_de_archivo(datos):
    """Huella del contenido del archivo. Reconoce la misma planilla aunque le cambien el nombre."""
    if not datos:
        return None
    return hashlib.sha256(datos).hexdigest()[:32]


def importacion_previa(huella):
    """Si esta misma planilla ya se importó, devuelve cuándo y con qué marca."""
    if not huella:
        return None
    c.execute("""SELECT marca, archivo, fecha, filas_cargadas FROM importaciones
                 WHERE huella = ? ORDER BY fecha DESC LIMIT 1""", (huella,))
    fila = c.fetchone()
    return dict(fila) if fila else None


def leer_numero(valor):
    """Lee un número de una celda de Excel, aguantando cómo lo escribe cada proveedor.

    El problema real: en Argentina el punto separa miles y la coma decimales ("1.234,56"), pero
    muchas listas vienen exportadas al revés ("1,234.56"), y otras traen el símbolo de peso,
    espacios o texto pegado ("$850.-"). Leerlo mal convierte $1.234,56 en $123.456: cien veces
    de más.

    Las reglas, en orden:
      1. Si están los DOS separadores, el que está más a la derecha es el decimal.
      2. Si hay uno solo y aparece VARIAS veces, es separador de miles ("12.345.678").
      3. Si hay uno solo y aparece una vez, decide cuántos dígitos lo siguen: tres significa
         miles ("1.234" son mil doscientos treinta y cuatro, no uno con doscientos treinta y
         cuatro), cualquier otra cantidad significa decimales ("1.50", "0,5").
    """
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return None
    negativo = texto.lstrip().startswith("-")
    texto = re.sub(r'[^\d,.]', '', texto)
    texto = texto.strip(",.")               # se come el "$850.-" y el "1.234,-"
    if not texto or not any(ch.isdigit() for ch in texto):
        return None

    comas, puntos = texto.count(","), texto.count(".")

    if comas and puntos:
        # Regla 1: manda el de más a la derecha
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif comas or puntos:
        sep = "," if comas else "."
        if (comas + puntos) > 1:
            texto = texto.replace(sep, "")                    # Regla 2: miles
        else:
            decimales = len(texto.split(sep)[1])
            if decimales == 3:
                texto = texto.replace(sep, "")                # Regla 3: miles
            else:
                texto = texto.replace(sep, ".")               # Regla 3: decimales

    try:
        numero = float(texto)
    except ValueError:
        return None
    return -numero if negativo else numero


def salto_de_precio_sospechoso(precio_viejo, precio_nuevo, tope_pct=200):
    """¿Este cambio de precio parece un error de la lista y no un aumento? (sospechoso, motivo).

    No mira solo el porcentaje: un salto de exactamente 100 o 1000 veces casi siempre es el
    separador de decimales mal leído, y conviene decirlo con esas palabras para que se entienda
    qué revisar."""
    if not tope_pct or precio_viejo is None or precio_nuevo is None:
        return False, None
    if precio_viejo <= 0 or precio_nuevo <= 0:
        return False, None
    razon = precio_nuevo / precio_viejo
    for factor, texto in ((1000, "mil"), (100, "cien"), (10, "diez")):
        if abs(razon - factor) / factor < 0.02:
            return True, f"quedó {texto} veces más caro — suele ser el separador de decimales"
        if abs(razon - 1 / factor) * factor < 0.02:
            return True, f"quedó {texto} veces más barato — suele ser el separador de decimales"
    # Se compara la RAZÓN, no el porcentaje, y a propósito.
    # Un aumento no tiene techo (+500%, +2000%), pero una baja no puede pasar de -100%: con un
    # límite del 200% en porcentaje, NINGUNA baja se frenaría nunca, ni siquiera un precio que
    # se divide por veinte. Mirando la razón, multiplicar por 3 y dividir por 3 pesan igual.
    factor = 1 + tope_pct / 100.0
    if razon > factor:
        return True, f"queda {razon:.1f} veces más caro (+{(razon - 1) * 100:.0f}%)"
    if razon < 1 / factor:
        return True, f"queda {1 / razon:.1f} veces más barato ({(razon - 1) * 100:.0f}%)"
    return False, None


def get_or_create_producto(raw, clean, desc, marca_id, imagen_url=None):
    c.execute(
        "INSERT OR IGNORE INTO productos (codigo_raw, codigo_clean, descripcion, marca_id) VALUES (?, ?, ?, ?)",
        (raw, clean, desc, marca_id)
    )
    if desc:
        c.execute(
            "UPDATE productos SET descripcion = ?, codigo_raw = ? "
            "WHERE codigo_clean = ? AND marca_id = ? AND (descripcion IS NULL OR descripcion = '')",
            (desc, raw, clean, marca_id)
        )
    if imagen_url:
        c.execute(
            "UPDATE productos SET imagen_url = ? WHERE codigo_clean = ? AND marca_id = ?",
            (imagen_url, clean, marca_id)
        )
    c.execute("SELECT id FROM productos WHERE codigo_clean = ? AND marca_id = ?", (clean, marca_id))
    return c.fetchone()[0]


def filas_a_listas(cursor):
    """Convierte el resultado de un cursor (sqlite3.Row) en una lista de diccionarios."""
    return [dict(row) for row in cursor.fetchall()]


def buscar_por_codigo(clean_code, marca_filtro="Todas", max_saltos=None):
    """Busca un código y todo lo que esté encadenado con él.

    La búsqueda es TRANSITIVA: si la lista A dice que el 1 equivale al 2, y la lista B dice que
    el 2 equivale al 3, buscar el 1 también trae el 3. Eso es lo que la hace potente, y también
    lo que la hace frágil: un solo vínculo mal cargado fusiona dos familias de repuestos que no
    tienen nada que ver, y a partir de ahí la búsqueda devuelve cosas que no entran.

    Por eso ahora se cuenta a cuántos SALTOS está cada resultado del código buscado. Un salto es
    un vínculo directo: alguien lo puso en la misma fila. Cinco saltos es una cadena larga donde
    cualquier eslabón puede estar mal. Con max_saltos se corta la cadena."""
    tope = int(max_saltos) if max_saltos else 99
    # La consulta arrastra, además de los saltos, la confianza del ESLABÓN MÁS DÉBIL del camino.
    # Es lo que faltaba para poder confiar en un resultado: no alcanza con saber que está a dos
    # saltos, hace falta saber si esos dos saltos son sólidos. Una cadena vale lo que su eslabón
    # más flojo, así que se va guardando el mínimo. Los vínculos viejos, todavía sin puntuar,
    # cuentan como 50 (ni a favor ni en contra) para no ensuciar el resultado.
    query = '''
    WITH RECURSIVE Red(id, saltos, peor) AS (
        SELECT id, 0, 100 FROM productos WHERE codigo_clean = ?
        UNION
        SELECT CASE WHEN eq.producto_a_id = re.id THEN eq.producto_b_id ELSE eq.producto_a_id END,
               re.saltos + 1,
               MIN(re.peor, COALESCE(eq.confianza, 50))
        FROM equivalencias eq JOIN Red re ON (eq.producto_a_id = re.id OR eq.producto_b_id = re.id)
        WHERE re.saltos < ?
    )
    SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
           m.nombre AS "Marca", m.tipo AS "Tipo", p.precio AS "Precio", p.stock AS "Stock",
           p.favorito AS "Favorito", COALESCE(p.imagen_thumb, p.imagen_url) AS "Imagen",
           p.precio_costo AS "_costo",
           m.url_ficha_template AS "_template", MIN(r.saltos) AS "_saltos",
           MAX(r.peor) AS "_peor"
    FROM Red r JOIN productos p ON p.id = r.id JOIN marcas m ON m.id = p.marca_id
    '''
    params = [clean_code, tope]
    if marca_filtro and marca_filtro != "Todas":
        query += " WHERE UPPER(m.nombre) = ?"
        params.append(marca_filtro.upper())
    # El GROUP BY reemplaza al DISTINCT de antes: hace falta para quedarse con el camino MÁS
    # CORTO hasta cada producto, que es el que dice cuánto confiar.
    # Tope de resultados. Una red sana tiene entre 2 y 20 códigos; si devuelve cientos, es que
    # un código puente fusionó familias que no tienen relación, y mostrar 1.800 filas no ayuda
    # a nadie — solo tarda y tapa lo bueno. Se ordena por saltos, así lo primero es lo cercano.
    query += ' GROUP BY p.id ORDER BY MIN(r.saltos), m.tipo, m.nombre LIMIT 400;'

    with db_lock:
        c.execute(query, params)
        res = filas_a_listas(c)
        if not res:
            return res

        # Marca qué filas están verificadas con un link directo hacia el producto buscado,
        # y trae el nivel/nota de esa relación. Una sola consulta para todo el lote.
        c.execute("SELECT id FROM productos WHERE codigo_clean = ?", (clean_code,))
        origenes = [r["id"] for r in c.fetchall()]
        verificados_set = set()
        info_relacion = {}  # producto_id -> {"nivel": ..., "nota": ...}
        if origenes:
            result_ids = [f["ID"] for f in res]
            placeholders_o = ",".join("?" * len(origenes))
            placeholders_r = ",".join("?" * len(result_ids))
            c.execute(
                f"""SELECT producto_a_id, producto_b_id, verificada, nivel, nota FROM equivalencias
                    WHERE ((producto_a_id IN ({placeholders_o}) AND producto_b_id IN ({placeholders_r}))
                        OR (producto_b_id IN ({placeholders_o}) AND producto_a_id IN ({placeholders_r})))""",
                origenes + result_ids + origenes + result_ids
            )
            for a, b, verif, nivel, nota in c.fetchall():
                if verif:
                    verificados_set.add(a)
                    verificados_set.add(b)
                otro_id = b if a in origenes else a
                if nivel or nota:
                    info_relacion[otro_id] = {"nivel": nivel, "nota": nota}

    for fila in res:
        saltos = fila.pop("_saltos", 0) or 0
        peor = fila.pop("_peor", None)
        fila["Cadena"] = ("— el buscado" if saltos == 0 else
                          "🟢 directo" if saltos == 1 else
                          f"🟡 {saltos} saltos" if saltos <= 3 else
                          f"🔴 {saltos} saltos")
        # Lo que vale el camino entero: su eslabón más flojo. Un resultado a dos saltos por
        # vínculos sólidos es más confiable que uno directo colgado de un vínculo malo.
        if saltos and peor is not None:
            fila["Confianza"] = ("🟢 sólida" if peor >= 70 else
                                 "🟡 razonable" if peor >= 50 else
                                 "🟠 floja" if peor >= 30 else
                                 "🔴 muy débil")
        else:
            fila["Confianza"] = ""
        fila["Verificada"] = "✅" if fila["ID"] in verificados_set else ""
        rel = info_relacion.get(fila["ID"], {})
        fila["Nivel"] = rel.get("nivel") or ("Exacta" if fila["ID"] in verificados_set else "")
        fila["Nota"] = rel.get("nota") or ""
        template = fila.pop("_template", None)
        fila["Ficha"] = template.replace("{codigo}", quote(fila["Codigo"], safe="")) if template else ""
    return res


def incrementar_veces_buscado(clean_code):
    """Suma 1 al contador de búsquedas de un código (usado para la matriz ABC).
    Se llama únicamente desde el buscador público, no desde búsquedas internas de administración."""
    with db_lock:
        c.execute(
            "UPDATE productos SET veces_buscado = COALESCE(veces_buscado, 0) + 1 WHERE codigo_clean = ?",
            (clean_code,)
        )
        conn.commit()


def armar_lista_picking(codigos_texto):
    """Busca varios códigos a la vez y devuelve el resultado ordenado por ubicación en el
    depósito, para que el que arma el pedido camine en un solo recorrido en vez de ir y volver."""
    codigos = [sanitizar(x) for x in codigos_texto.split(",")]
    codigos = [x for x in codigos if x]
    if not codigos:
        return []
    placeholders = ",".join("?" * len(codigos))
    c.execute(f'''SELECT p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion", m.nombre AS "Marca",
                  p.ubicacion AS "Ubicación", p.stock AS "Stock"
                  FROM productos p JOIN marcas m ON m.id = p.marca_id
                  WHERE p.codigo_clean IN ({placeholders})''', codigos)
    resultado = filas_a_listas(c)
    resultado.sort(key=lambda r: (not r["Ubicación"], r["Ubicación"] or ""))
    return resultado


def _sql_sin_acentos(columna):
    """Arma una expresión SQL que le saca los acentos a una columna (funciona con mayúscula
    y minúscula, porque SQLite no toca letras acentuadas al hacer UPPER())."""
    reemplazos = [("á", "A"), ("Á", "A"), ("é", "E"), ("É", "E"), ("í", "I"), ("Í", "I"),
                  ("ó", "O"), ("Ó", "O"), ("ú", "U"), ("Ú", "U"), ("ñ", "N"), ("Ñ", "N")]
    expr = f"UPPER({columna})"
    for viejo, nuevo in reemplazos:
        expr = f"REPLACE({expr},'{viejo}','{nuevo}')"
    return expr


def contar_codigos_con_decimal():
    c.execute(r"SELECT COUNT(*) FROM productos WHERE codigo_raw LIKE '%.0' AND codigo_raw GLOB '[0-9]*'")
    return c.fetchone()[0]


def reparar_codigos_con_decimal():
    """Arregla los códigos que quedaron con '.0' del final por venir de una celda numérica de
    Excel. Además de verse feo, los volvía imposibles de encontrar: '2776400.0' se limpiaba
    como '27764000' (con un cero de más) y nunca coincidía con el código real."""
    c.execute(r"""SELECT id, codigo_raw FROM productos
                  WHERE codigo_raw LIKE '%.0' AND codigo_raw GLOB '[0-9]*'""")
    filas = [(r["id"], r["codigo_raw"]) for r in c.fetchall()]
    arreglados = 0
    with db_lock:
        for pid, raw in filas:
            if not re.fullmatch(r"\d+\.0+", str(raw).strip()):
                continue
            nuevo_raw = str(raw).strip().split(".")[0]
            nuevo_clean = sanitizar(nuevo_raw)
            try:
                c.execute("UPDATE productos SET codigo_raw = ?, codigo_clean = ? WHERE id = ?",
                           (nuevo_raw, nuevo_clean, pid))
                arreglados += 1
            except sqlite3.IntegrityError:
                # En teoría no debería pasar: '2776400.0' y '2776400' se limpian al mismo
                # codigo_clean, así que la base nunca deja que existan los dos en la misma
                # marca. Si igual llegara a ocurrir, se fusionan en vez de dejar el '.0'.
                c.execute("SELECT marca_id FROM productos WHERE id = ?", (pid,))
                fila_marca = c.fetchone()
                if not fila_marca:
                    continue
                c.execute("SELECT id FROM productos WHERE marca_id = ? AND codigo_clean = ? AND id <> ?",
                           (fila_marca["marca_id"], nuevo_clean, pid))
                bueno = c.fetchone()
                if bueno and fusionar_productos(pid, bueno["id"]):
                    arreglados += 1
        conn.commit()
    return arreglados


def listar_importaciones_deshacibles(limite=40):
    """Importaciones que dejaron vínculos rastreables, con cuántos quedan vivos."""
    c.execute("""SELECT i.id AS "_id", i.marca AS "Marca", i.archivo AS "Archivo",
                        substr(i.fecha, 1, 16) AS "Fecha", i.filas_cargadas AS "Filas",
                        i.lote AS "_lote",
                        (SELECT COUNT(*) FROM equivalencias e WHERE e.lote = i.lote) AS "Vínculos vivos",
                        (SELECT COUNT(*) FROM equivalencias_pendientes ep WHERE ep.lote = i.lote)
                            AS "Sin revisar"
                 FROM importaciones i
                 WHERE i.lote IS NOT NULL
                 ORDER BY i.fecha DESC LIMIT ?""", (limite,))
    return filas_a_listas(c)


def previsualizar_deshacer(lote):
    """Qué se va a borrar si se deshace esta importación. Se mira ANTES de tocar nada.

    Los productos se cuentan aparte de los vínculos porque son cosas distintas: un producto que
    solo trajo esta lista se puede sacar sin problema, pero uno que además tiene precio, stock,
    ubicación o historial de ventas NO se toca aunque haya entrado con esta lista — borrarlo
    perdería datos que no vinieron de acá."""
    c.execute("SELECT COUNT(*) FROM equivalencias WHERE lote = ?", (lote,))
    vinculos = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM equivalencias_pendientes WHERE lote = ?", (lote,))
    pendientes = c.fetchone()[0]
    return {"vinculos": vinculos, "pendientes": pendientes, "lote": lote}


def deshacer_importacion(lote, borrar_pendientes=True):
    """Saca los vínculos que dejó una importación. Devuelve cuántos se borraron.

    Solo toca las equivalencias, nunca los productos: los precios, el stock, la ubicación y el
    historial que hayas cargado sobre esos productos se quedan donde están. Lo que se deshace
    es la parte peligrosa —los vínculos falsos— y eso es reversible sin perder trabajo."""
    if not lote:
        return 0, 0
    with db_lock:
        c.execute("SELECT producto_a_id, producto_b_id FROM equivalencias WHERE lote = ?", (lote,))
        pares = [(r["producto_a_id"], r["producto_b_id"]) for r in c.fetchall()]
        c.execute("DELETE FROM equivalencias WHERE lote = ?", (lote,))
        borrados = c.rowcount
        pend = 0
        if borrar_pendientes:
            c.execute("DELETE FROM equivalencias_pendientes WHERE lote = ?", (lote,))
            pend = c.rowcount
        conn.commit()
    # Queda registrado el rechazo para que una reimportación de la misma lista no los reviva.
    # OJO con el valor: tiene que ser exactamente "rechazada" — es el que busca pares_rechazados().
    # Escrito de cualquier otra forma se guarda igual y no filtra nada, y encima en silencio.
    if pares:
        marcar_revision(pares, "rechazada")
    return borrados, pend



def origenes_de_los_vinculos_directos(producto_id, ids_resultado):
    """De qué lista salió cada vínculo directo del código buscado. {id_producto: lote}.

    Sirve en el momento de decidir: si un resultado raro salió de una lista que ya sabés que
    vino mal mapeada, no hace falta pensarlo más. Y si esa lista es un desastre entero, se
    deshace de una desde Mantenimiento en vez de ir corrigiendo de a un vínculo."""
    if not ids_resultado:
        return {}
    marcadores = ",".join("?" * len(ids_resultado))
    c.execute(f"""SELECT CASE WHEN producto_a_id = ? THEN producto_b_id ELSE producto_a_id END AS otro,
                         lote
                  FROM equivalencias
                  WHERE (producto_a_id = ? OR producto_b_id = ?)
                    AND lote IS NOT NULL
                    AND (producto_a_id IN ({marcadores}) OR producto_b_id IN ({marcadores}))""",
              [producto_id, producto_id, producto_id] + list(ids_resultado) * 2)
    return {r["otro"]: r["lote"] for r in c.fetchall()}


def auditar_equivalencias_cargadas(limite=300, tope_confianza=35, revisar=8000):
    """Pasa el mismo análisis de confianza por las equivalencias YA cargadas.

    Es la herramienta que faltaba. El análisis de confianza solo miraba los vínculos pendientes
    de revisión, pero el problema grande está en los que YA entraron: miles cargados por
    importaciones viejas que nadie revisó. Sin esto, la única forma de encontrarlos era
    tropezarse con uno buscando un código.

    Devuelve los peores primero, con el motivo escrito."""
    c.execute("""SELECT e.producto_a_id AS a, e.producto_b_id AS b, e.lote,
                        pa.codigo_raw AS cod_a, pa.descripcion AS desc_a, pa.precio AS precio_a,
                        ma.nombre AS marca_a,
                        pb.codigo_raw AS cod_b, pb.descripcion AS desc_b, pb.precio AS precio_b,
                        mb.nombre AS marca_b
                 FROM equivalencias e
                 JOIN productos pa ON pa.id = e.producto_a_id
                 JOIN productos pb ON pb.id = e.producto_b_id
                 JOIN marcas ma ON ma.id = pa.marca_id
                 JOIN marcas mb ON mb.id = pb.marca_id
                 LIMIT ?""", (revisar,))
    filas = [dict(r) for r in c.fetchall()]
    if not filas:
        return [], 0

    medidas = cargar_medidas_de_varios([f["a"] for f in filas] + [f["b"] for f in filas])
    aprobados = puentes_aprobados_ids()
    patrones = aprender_de_las_decisiones()
    ventas_confirman = pares_confirmados_por_ventas()

    dudosas = []
    for f in filas:
        # Un vínculo de un código puente ya aprobado a mano no se vuelve a cuestionar
        if f["a"] in aprobados or f["b"] in aprobados:
            continue
        puntaje, senales = evaluar_equivalencia(
            f["desc_a"], f["desc_b"], medidas.get(f["a"]), medidas.get(f["b"]),
            f["precio_a"], f["precio_b"],
            marca_a=f["marca_a"], marca_b=f["marca_b"], patrones=patrones,
            vendido_como_reemplazo=ventas_confirman.get((min(f["a"], f["b"]),
                                                          max(f["a"], f["b"])), 0)
        )
        for lado in ("a", "b"):
            malo, _ = codigo_sospechoso(f[f"cod_{lado}"], f.get(f"desc_{lado}") or "")
            if malo:
                puntaje -= 35
        puntaje = max(0.0, puntaje)
        if puntaje <= tope_confianza:
            dudosas.append({
                "Confianza": round(puntaje),
                "Código A": f["cod_a"], "Marca A": f["marca_a"],
                "Código B": f["cod_b"], "Marca B": f["marca_b"],
                "Por qué": " · ".join(t for tipo, t in senales if tipo == "mal") or
                            "el código no parece un código",
                "Vino de": (f["lote"] or "").split(" · ")[0],
                "_a": f["a"], "_b": f["b"],
            })
    dudosas.sort(key=lambda x: x["Confianza"])
    return dudosas[:limite], len(filas)


def recalcular_confianzas(limite=20000, progreso=None):
    """Calcula y guarda la confianza de cada vínculo cargado.

    Se guarda en vez de calcularse al vuelo porque el buscador la necesita en CADA búsqueda:
    hacer el análisis completo ahí lo volvería lento. Así se hace una vez y queda."""
    c.execute("""SELECT e.producto_a_id AS a, e.producto_b_id AS b,
                        pa.codigo_raw AS cod_a, pa.descripcion AS desc_a, pa.precio AS precio_a,
                        ma.nombre AS marca_a,
                        pb.codigo_raw AS cod_b, pb.descripcion AS desc_b, pb.precio AS precio_b,
                        mb.nombre AS marca_b
                 FROM equivalencias e
                 JOIN productos pa ON pa.id = e.producto_a_id
                 JOIN productos pb ON pb.id = e.producto_b_id
                 JOIN marcas ma ON ma.id = pa.marca_id
                 JOIN marcas mb ON mb.id = pb.marca_id
                 LIMIT ?""", (limite,))
    filas = [dict(r) for r in c.fetchall()]
    if not filas:
        return 0
    medidas = cargar_medidas_de_varios([f["a"] for f in filas] + [f["b"] for f in filas])
    patrones = aprender_de_las_decisiones()
    aprobados = puentes_aprobados_ids()
    ventas_confirman = pares_confirmados_por_ventas()

    valores = []
    for i, f in enumerate(filas):
        puntaje, _ = evaluar_equivalencia(
            f["desc_a"], f["desc_b"], medidas.get(f["a"]), medidas.get(f["b"]),
            f["precio_a"], f["precio_b"],
            marca_a=f["marca_a"], marca_b=f["marca_b"], patrones=patrones,
            vendido_como_reemplazo=ventas_confirman.get((min(f["a"], f["b"]),
                                                          max(f["a"], f["b"])), 0)
        )
        if f["a"] not in aprobados and f["b"] not in aprobados:
            for lado in ("a", "b"):
                malo, _ = codigo_sospechoso(f[f"cod_{lado}"], f.get(f"desc_{lado}") or "")
                if malo:
                    puntaje -= 35
        valores.append((max(0, min(100, round(puntaje))), f["a"], f["b"]))
        if progreso and i % 500 == 0:
            progreso(i, len(filas))

    with db_lock:
        c.executemany("UPDATE equivalencias SET confianza = ? "
                      "WHERE producto_a_id = ? AND producto_b_id = ?", valores)
        conn.commit()
    return len(valores)


def borrar_equivalencias_dudosas(pares):
    """Corta los vínculos elegidos y los deja anotados como rechazados."""
    if not pares:
        return 0
    with db_lock:
        c.executemany("DELETE FROM equivalencias WHERE producto_a_id = ? AND producto_b_id = ?",
                      [(min(a, b), max(a, b)) for a, b in pares])
        borrados = c.rowcount
        conn.commit()
    marcar_revision(list(pares), "rechazada")
    return borrados


def contar_equivalencias_espejadas():
    """Cuántas equivalencias están guardadas dos veces, una en cada dirección."""
    c.execute("""SELECT COUNT(*) FROM equivalencias e
                 WHERE e.producto_a_id > e.producto_b_id
                   AND EXISTS (SELECT 1 FROM equivalencias e2
                               WHERE e2.producto_a_id = e.producto_b_id
                                 AND e2.producto_b_id = e.producto_a_id)""")
    return c.fetchone()[0]


def unificar_equivalencias_espejadas():
    """Deja una sola fila por equivalencia, siempre con el id menor primero.

    No cambia ninguna equivalencia: son exactamente las mismas, solo que estaban anotadas dos
    veces. La búsqueda nunca lo notó porque consulta las dos columnas con OR, pero los conteos
    sí: un código con 100 equivalencias reales figuraba con 200."""
    with db_lock:
        # Primero las que están espejadas (existe la pareja al revés): se borra la invertida
        c.execute("""DELETE FROM equivalencias
                     WHERE producto_a_id > producto_b_id
                       AND EXISTS (SELECT 1 FROM equivalencias e2
                                   WHERE e2.producto_a_id = equivalencias.producto_b_id
                                     AND e2.producto_b_id = equivalencias.producto_a_id)""")
        borradas = c.rowcount
        # Y las que quedaron sueltas al revés se dan vuelta, para que todas queden igual
        c.execute("""UPDATE equivalencias
                     SET producto_a_id = producto_b_id, producto_b_id = producto_a_id
                     WHERE producto_a_id > producto_b_id""")
        dadas_vuelta = c.rowcount
        c.execute("DELETE FROM equivalencias WHERE producto_a_id = producto_b_id")
        conn.commit()
    return borradas, dadas_vuelta


def aprobar_puente(producto_id, nota=""):
    """Marca un código con muchos vínculos como revisado y correcto."""
    with db_lock:
        c.execute("""INSERT INTO puentes_aprobados (producto_id, aprobado_por, nota)
                     VALUES (?, ?, ?)
                     ON CONFLICT(producto_id) DO UPDATE SET nota = excluded.nota,
                        aprobado_por = excluded.aprobado_por, fecha = datetime('now')""",
                  (producto_id, obtener_usuario_actual(), (nota or "").strip() or None))
        conn.commit()
    return True


def desaprobar_puente(producto_id):
    with db_lock:
        c.execute("DELETE FROM puentes_aprobados WHERE producto_id = ?", (producto_id,))
        conn.commit()
    return c.rowcount


def puentes_aprobados_ids():
    try:
        c.execute("SELECT producto_id FROM puentes_aprobados")
        return {r["producto_id"] for r in c.fetchall()}
    except sqlite3.OperationalError:
        return set()


def contar_codigos_puente(minimo=30):
    """Solo el conteo, sin traer los datos. Es lo que usa el chequeo de salud, que corre seguido:
    la consulta completa arma dos subconsultas por producto y no hace falta para un número."""
    c.execute("""SELECT COUNT(*) FROM productos p
                 WHERE (SELECT COUNT(*) FROM equivalencias e
                        WHERE e.producto_a_id = p.id OR e.producto_b_id = p.id) >= ?
                   AND p.id NOT IN (SELECT producto_id FROM puentes_aprobados)""", (minimo,))
    return c.fetchone()[0]


def invalidar_salud():
    """Fuerza a recalcular el chequeo de salud en el próximo refresco.

    Se llama después de cada acción que cambia los números: importar, unificar duplicadas,
    cortar vínculos, bajar un backup. Sin esto el aviso de arriba seguiría mostrando el
    problema durante tres minutos después de haberlo arreglado, y uno no sabría si funcionó."""
    try:
        st.session_state.pop("_salud_cache", None)
    except Exception:
        pass


def diagnostico_de_salud():
    """Corre todos los controles de mantenimiento de una y devuelve solo lo que necesita atención.

    Por qué: los controles se fueron sumando de a uno y quedaron repartidos en distintas
    pantallas de Mantenimiento. Ninguno avisa solo, así que hay que acordarse de entrar y mirar
    los siete — y en la práctica nadie lo hace hasta que algo ya salió mal. Esto los junta en un
    solo lugar y aparece cuando hay algo para revisar.

    Cada punto trae a dónde ir y qué pasa si no se toca, porque un número suelto no dice nada."""
    problemas = []

    def sumar(nivel, titulo, detalle, donde):
        problemas.append({"nivel": nivel, "titulo": titulo, "detalle": detalle, "donde": donde})

    try:
        puentes = contar_codigos_puente(30)
        if puentes:
            sumar("alto", f"{puentes} código(s) puente",
                  "Están vinculados a decenas de repuestos y fusionan familias que no tienen "
                  "relación. Es lo que hace que el buscador devuelva cosas que no entran.",
                  "Estadísticas → Mantenimiento → Códigos puente")
    except Exception:
        pass

    try:
        espejadas = contar_equivalencias_espejadas()
        if espejadas:
            sumar("medio", f"{espejadas:,} equivalencias anotadas dos veces",
                  "La misma relación guardada en las dos direcciones. No cambia lo que encuentra "
                  "el buscador, pero duplica todos los conteos.",
                  "Estadísticas → Mantenimiento → Equivalencias anotadas dos veces")
    except Exception:
        pass

    try:
        basura = contar_codigos_basura()
        if basura:
            sumar("alto", f"{basura} código(s) que son un número suelto",
                  "Entraron cantidades o números de orden en la columna del código. Cada uno "
                  "vincula entre sí repuestos que no tienen nada que ver.",
                  "Estadísticas → Mantenimiento → Códigos que son solo un número suelto")
    except Exception:
        pass

    try:
        con_punto = contar_codigos_con_decimal()
        if con_punto:
            sumar("medio", f"{con_punto} código(s) terminados en '.0'",
                  "Excel los guardó como número. Se encuentran igual, pero el código que se "
                  "muestra y se copia en un presupuesto está mal.",
                  "Estadísticas → Mantenimiento → Códigos que quedaron con '.0'")
    except Exception:
        pass

    try:
        c.execute("""SELECT COUNT(*) FROM equivalencias e
                     JOIN productos pa ON pa.id = e.producto_a_id
                     JOIN productos pb ON pb.id = e.producto_b_id
                     WHERE pa.precio > 0 AND pb.precio > 0
                       AND MAX(pa.precio, pb.precio) / MIN(pa.precio, pb.precio) >= 8""")
        precios = c.fetchone()[0]
        if precios:
            sumar("alto", f"{precios} par(es) de equivalentes con precios muy distintos",
                  "O el precio está mal cargado, o no son la misma pieza. Cualquiera de las dos "
                  "cuesta plata: o cotizás mal, o vendés algo que no entra.",
                  "Estadísticas → Mantenimiento → Precios que no cierran")
    except Exception:
        pass

    try:
        c.execute("SELECT COUNT(*) FROM equivalencias_pendientes")
        pendientes = c.fetchone()[0]
        if pendientes > 500:
            sumar("medio", f"{pendientes:,} vínculos esperando revisión",
                  "Mientras no se revisen no están cargados, así que el buscador no los usa.",
                  "Estadísticas → Vínculos de listas esperando revisión")
    except Exception:
        pass

    try:
        c.execute("SELECT COUNT(*) FROM equivalencias WHERE confianza IS NULL")
        sin_punt = c.fetchone()[0]
        if sin_punt > 200:
            sumar("medio", f"{sin_punt:,} vínculos sin puntuar",
                  "El buscador no puede decirte qué tan sólido es el camino de cada resultado "
                  "hasta que se calculen. Es un solo botón.",
                  "Estadísticas → Mantenimiento → Puntuar los vínculos")
    except sqlite3.OperationalError:
        pass

    try:
        c.execute("SELECT COUNT(*) FROM aplicaciones")
        if c.fetchone()[0] == 0:
            c.execute("SELECT COUNT(*) FROM productos")
            if c.fetchone()[0] > 500:
                sumar("medio", "Sin catálogos de aplicaciones cargados",
                      "Son los que dicen qué repuesto le va a cada auto. Varios fabricantes los "
                      "publican gratis (NGK, Bosch, Mann, SKF). Sin ellos, la búsqueda por "
                      "vehículo tiene que adivinar desde las descripciones del proveedor.",
                      "Estadísticas → Mantenimiento → Catálogo de aplicaciones")
    except sqlite3.OperationalError:
        pass

    try:
        quiebres = productos_por_quebrar(dias_aviso=14)
        en_cero = [x for x in quiebres if x["Stock"] <= 0]
        if en_cero:
            sumar("alto", f"{len(en_cero)} producto(s) que se venden seguido están sin stock",
                  "Se venden todos los meses y están en cero. Un cliente los va a pedir y no "
                  "van a estar.",
                  "Estadísticas → Reposición → Lo que se va a acabar")
        elif quiebres:
            sumar("medio", f"{len(quiebres)} producto(s) se acaban en menos de 2 semanas",
                  "Según el ritmo con que se vienen vendiendo y el stock que queda.",
                  "Estadísticas → Reposición → Lo que se va a acabar")
    except sqlite3.OperationalError:
        pass

    # El número concreto de lo que se perdería. Un aviso genérico se ignora; «perdés 3.412
    # productos» no.
    try:
        riesgo = cuanto_perderias_si_reinicia()
        if riesgo and not riesgo["hay_semilla"] and riesgo.get("productos_ahora", 0) > 100:
            sumar("alto", "No hay copia en el repositorio",
                  "El servidor borra el disco al reiniciar y se restaura desde "
                  "`datos_iniciales.db`, que no está. Hoy un reinicio borra TODO.",
                  "Estadísticas → Backup y config")
        elif riesgo and riesgo["en_riesgo"] > 200:
            sumar("alto", f"{riesgo['en_riesgo']:,} productos viven solo en el disco",
                  f"La copia del repositorio es del {riesgo['fecha_semilla']} y tiene "
                  f"{riesgo['productos_semilla']:,}; hoy tenés {riesgo['productos_ahora']:,}. "
                  "Si el servidor reinicia, la diferencia se pierde.",
                  "Estadísticas → Backup y config")
    except Exception:
        pass

    bk = estado_del_backup()
    if bk["urgente"]:
        if not bk["hay_backup"]:
            sumar("alto", "Nunca se bajó un backup",
                  "El servidor borra el disco al reiniciar y restaura desde la última copia. "
                  "Hoy podrías perder todo lo cargado.",
                  "Estadísticas → Backup y config")
        else:
            partes = []
            if bk["productos_nuevos"]:
                partes.append(f"{bk['productos_nuevos']:,} productos nuevos")
            if bk["importaciones"]:
                partes.append(f"{bk['importaciones']} lista(s) importadas")
            if bk["dias"]:
                partes.append(f"{bk['dias']} día(s)")
            sumar("alto", "Backup atrasado",
                  "Desde el último hay " + ", ".join(partes) +
                  ". Si el servidor reinicia ahora, eso se pierde.",
                  "Estadísticas → Backup y config")

    orden = {"alto": 0, "medio": 1}
    problemas.sort(key=lambda x: orden.get(x["nivel"], 2))
    return problemas


def camino_entre(origen_id, destino_id, tope_nodos=3000):
    """Por qué cadena de vínculos este resultado llegó hasta acá. Devuelve la lista de pasos.

    Es lo que faltaba para que todo lo demás sirva. El buscador ya avisaba «el camino es débil»,
    pero no decía DÓNDE: había que salir a buscar el eslabón malo a mano. Ahora muestra la
    cadena completa, con el puntaje y la lista de origen de cada paso, así se ve de una cuál
    cortar.

    Se busca el camino MÁS CONFIABLE, no el más corto: si hay dos maneras de llegar, la que
    pasa por vínculos sólidos es la que hay que mostrar — es la que decide si el resultado
    sirve o no."""
    if origen_id == destino_id:
        return []
    import heapq
    # Dijkstra maximizando el peor eslabón: el costo de un camino es su vínculo más flojo
    mejor = {origen_id: 100}
    previo = {}
    monton = [(-100, origen_id)]
    visitados = set()
    while monton and len(visitados) < tope_nodos:
        peor_neg, nodo = heapq.heappop(monton)
        if nodo in visitados:
            continue
        visitados.add(nodo)
        if nodo == destino_id:
            break
        c.execute("""SELECT CASE WHEN producto_a_id = ? THEN producto_b_id ELSE producto_a_id END
                            AS otro, COALESCE(confianza, 50) AS conf, lote
                     FROM equivalencias
                     WHERE producto_a_id = ? OR producto_b_id = ?""", (nodo, nodo, nodo))
        for fila in c.fetchall():
            otro, conf = fila["otro"], fila["conf"]
            if otro in visitados:
                continue
            nuevo_peor = min(-peor_neg, conf)
            if nuevo_peor > mejor.get(otro, -1):
                mejor[otro] = nuevo_peor
                previo[otro] = (nodo, conf, fila["lote"])
                heapq.heappush(monton, (-nuevo_peor, otro))

    if destino_id not in previo and destino_id != origen_id:
        return []

    # Se reconstruye el camino desde el final
    pasos = []
    actual = destino_id
    while actual != origen_id:
        anterior, conf, lote = previo[actual]
        pasos.append((anterior, actual, conf, lote))
        actual = anterior
    pasos.reverse()

    ids = {x for paso in pasos for x in paso[:2]}
    marcadores = ",".join("?" * len(ids))
    c.execute(f"""SELECT p.id, p.codigo_raw, p.descripcion, m.nombre AS marca
                  FROM productos p JOIN marcas m ON m.id = p.marca_id
                  WHERE p.id IN ({marcadores})""", list(ids))
    info = {r["id"]: r for r in c.fetchall()}

    salida = []
    for a, b, conf, lote in pasos:
        if a not in info or b not in info:
            continue
        salida.append({
            "Paso": f"{info[a]['codigo_raw']} ({info[a]['marca']}) → "
                     f"{info[b]['codigo_raw']} ({info[b]['marca']})",
            "Confianza": conf,
            "Vino de": (lote or "—").split(" · ")[0],
            "_a": a, "_b": b,
        })
    return salida


def _red_de(producto_id, tope=2000):
    """Trae la red completa de equivalencias conectada a un producto: nodos y aristas."""
    c.execute("""WITH RECURSIVE Red(id) AS (
                     SELECT ?
                     UNION
                     SELECT CASE WHEN e.producto_a_id = r.id THEN e.producto_b_id
                                 ELSE e.producto_a_id END
                     FROM equivalencias e JOIN Red r
                       ON (e.producto_a_id = r.id OR e.producto_b_id = r.id)
                 )
                 SELECT id FROM Red LIMIT ?""", (producto_id, tope))
    nodos = [r["id"] for r in c.fetchall()]
    if len(nodos) < 3:
        return nodos, []
    marcadores = ",".join("?" * len(nodos))
    c.execute(f"""SELECT producto_a_id AS a, producto_b_id AS b, COALESCE(confianza, 50) AS conf
                  FROM equivalencias
                  WHERE producto_a_id IN ({marcadores}) AND producto_b_id IN ({marcadores})""",
              nodos * 2)
    aristas = [(r["a"], r["b"], r["conf"]) for r in c.fetchall()]
    return nodos, aristas


def _vinculos_que_parten_la_red(nodos, aristas):
    """Encuentra los vínculos que, si se cortan, parten la red en dos.

    Es el caso que ni los códigos puente ni el puntaje individual detectan: dos familias de
    repuestos perfectamente legítimas —los filtros por un lado, los frenos por el otro— unidas
    por UN solo vínculo mal cargado. Ningún código tiene muchos enlaces, así que no aparece como
    puente; y el vínculo malo puede tener un puntaje mediano, así que tampoco salta solo.
    Pero es el único que sostiene la unión: cortándolo, las dos familias se separan.

    Se usa el algoritmo de Tarjan, que los encuentra todos en una sola pasada. Buscarlos
    probando de a uno —cortar y ver si se desconecta— sería inviable en una red grande."""
    vecinos = {n: [] for n in nodos}
    for a, b, conf in aristas:
        if a in vecinos and b in vecinos:
            vecinos[a].append((b, conf))
            vecinos[b].append((a, conf))

    orden, bajo = {}, {}
    contador = [0]
    puentes = []

    for raiz in nodos:
        if raiz in orden:
            continue
        # Recorrido sin recursión: una red grande haría explotar la pila
        pila = [(raiz, None, iter(vecinos[raiz]))]
        orden[raiz] = bajo[raiz] = contador[0]
        contador[0] += 1
        while pila:
            nodo, padre, iterador = pila[-1]
            avanzo = False
            for vecino, conf in iterador:
                if vecino == padre:
                    continue
                if vecino not in orden:
                    orden[vecino] = bajo[vecino] = contador[0]
                    contador[0] += 1
                    pila.append((vecino, nodo, iter(vecinos[vecino])))
                    avanzo = True
                    break
                bajo[nodo] = min(bajo[nodo], orden[vecino])
            if not avanzo:
                pila.pop()
                if pila:
                    arriba = pila[-1][0]
                    bajo[arriba] = min(bajo[arriba], bajo[nodo])
                    if bajo[nodo] > orden[arriba]:
                        conf = next((cf for v, cf in vecinos[nodo] if v == arriba), 50)
                        puentes.append((arriba, nodo, conf))
    return puentes


def _tamano_del_lado(inicio, excluido, vecinos, tope=5000):
    """Cuántos nodos quedan de un lado si se corta un vínculo."""
    visto = {inicio}
    pila = [inicio]
    while pila and len(visto) < tope:
        n = pila.pop()
        for v, _ in vecinos.get(n, []):
            if (n, v) == excluido or (v, n) == excluido or v in visto:
                continue
            visto.add(v)
            pila.append(v)
    return len(visto)


def vinculos_que_unen_familias(producto_id, minimo_lado=3):
    """Los vínculos que están uniendo dos grupos grandes que quizá no tengan relación.

    Solo interesan los que dejan grupos GRANDES de los dos lados: si al cortar queda un producto
    suelto de un lado, eso es normal (una punta de la cadena). Si quedan 20 y 25, ese vínculo
    está sosteniendo la unión de dos familias enteras — y vale la pena mirarlo."""
    nodos, aristas = _red_de(producto_id)
    if len(nodos) < 6:
        return []
    puentes = _vinculos_que_parten_la_red(nodos, aristas)
    if not puentes:
        return []

    vecinos = {n: [] for n in nodos}
    for a, b, conf in aristas:
        if a in vecinos and b in vecinos:
            vecinos[a].append((b, conf))
            vecinos[b].append((a, conf))

    salida = []
    for a, b, conf in puentes:
        lado_a = _tamano_del_lado(a, (a, b), vecinos)
        lado_b = len(nodos) - lado_a
        if min(lado_a, lado_b) < minimo_lado:
            continue
        c.execute("""SELECT p.id, p.codigo_raw, p.descripcion, m.nombre AS marca
                     FROM productos p JOIN marcas m ON m.id = p.marca_id WHERE p.id IN (?, ?)""",
                  (a, b))
        info = {r["id"]: r for r in c.fetchall()}
        if a not in info or b not in info:
            continue
        salida.append({
            "Código A": info[a]["codigo_raw"], "Marca A": info[a]["marca"],
            "Código B": info[b]["codigo_raw"], "Marca B": info[b]["marca"],
            "Separa": f"{lado_a} y {lado_b} productos",
            "Confianza del vínculo": conf,
            "_equilibrio": min(lado_a, lado_b),
            "_a": a, "_b": b,
        })
    # Primero los más equilibrados y de menor confianza: son los más sospechosos
    salida.sort(key=lambda x: (-x["_equilibrio"], x["Confianza del vínculo"]))
    return salida


def codigos_puente(minimo=15, limite=100):
    """Códigos vinculados a demasiadas cosas. Son los que rompen la búsqueda.

    Como la búsqueda es transitiva, un código mal vinculado no ensucia solo su fila: fusiona
    todas las familias que toca. Un filtro de aceite legítimo puede tener 10 o 15 equivalencias
    entre marcas; si aparece con 200, casi seguro es un código que se cargó mal (una cantidad,
    un número de orden, o una columna corrida) y quedó de puente entre repuestos que no tienen
    nada que ver. Cortar UNO de estos limpia miles de resultados falsos de una."""
    c.execute("""SELECT p.id AS "ID", p.codigo_raw AS "Código", p.descripcion AS "Descripción",
                        m.nombre AS "Marca",
                        (SELECT COUNT(*) FROM equivalencias e
                          WHERE e.producto_a_id = p.id OR e.producto_b_id = p.id) AS "Vínculos",
                        (SELECT COUNT(DISTINCT m2.nombre) FROM equivalencias e
                          JOIN productos p2 ON p2.id = CASE WHEN e.producto_a_id = p.id
                                                            THEN e.producto_b_id ELSE e.producto_a_id END
                          JOIN marcas m2 ON m2.id = p2.marca_id
                          WHERE e.producto_a_id = p.id OR e.producto_b_id = p.id) AS "Marcas distintas"
                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                 WHERE "Vínculos" >= ?
                   AND p.id NOT IN (SELECT producto_id FROM puentes_aprobados)
                 ORDER BY "Vínculos" DESC LIMIT ?""", (minimo, limite))
    return filas_a_listas(c)


def puentes_en_el_resultado(ids_resultado, umbral=15):
    """De los productos que trajo una búsqueda, cuáles son códigos puente sospechosos.

    Hace falta además del control de saltos, y no en lugar de él: un código puente NO está lejos,
    está a un salto de todo. Al vincularse con cientos de cosas crea atajos, así que dos repuestos
    que no tienen nada que ver quedan a dos saltos uno del otro. Limitar la distancia sirve para
    cadenas largas, pero contra un puente no alcanza — hay que verlo y cortarlo."""
    if not ids_resultado:
        return []
    marcadores = ",".join("?" * len(ids_resultado))
    c.execute(f"""SELECT p.id AS "ID", p.codigo_raw AS "Código", p.descripcion AS "Descripción",
                         (SELECT COUNT(*) FROM equivalencias e
                           WHERE e.producto_a_id = p.id OR e.producto_b_id = p.id) AS "Vínculos"
                  FROM productos p WHERE p.id IN ({marcadores})
                    AND "Vínculos" >= ?
                    AND p.id NOT IN (SELECT producto_id FROM puentes_aprobados)
                  ORDER BY "Vínculos" DESC""",
              list(ids_resultado) + [umbral])
    return filas_a_listas(c)


def tamano_de_la_red(producto_id, tope=500):
    """Cuántos productos quedan encadenados a este. Si da cientos, la red está contaminada."""
    c.execute("""WITH RECURSIVE Red(id) AS (
                     SELECT ?
                     UNION
                     SELECT CASE WHEN eq.producto_a_id = re.id THEN eq.producto_b_id
                                 ELSE eq.producto_a_id END
                     FROM equivalencias eq JOIN Red re
                       ON (eq.producto_a_id = re.id OR eq.producto_b_id = re.id)
                 )
                 SELECT COUNT(*) FROM (SELECT id FROM Red LIMIT ?)""", (producto_id, tope))
    return c.fetchone()[0]


def cortar_vinculos_de(producto_id):
    """Corta TODAS las equivalencias de un código, sin borrar el producto."""
    with db_lock:
        c.execute("DELETE FROM equivalencias WHERE producto_a_id = ? OR producto_b_id = ?",
                  (producto_id, producto_id))
        borrados = c.rowcount
        conn.commit()
    return borrados


def listar_codigos_basura(limite=200):
    """Productos ya cargados cuyo código es solo 1 o 2 dígitos ('1', '12', '07'). Entraron con
    las importaciones viejas, antes del filtro, y son los que arrastran equivalencias falsas:
    todos los '1' de todas las listas terminaron vinculados entre sí."""
    # Se incluyen TODOS los de 1 o 2 caracteres, tengan letra o no. Antes solo se limpiaban los
    # puramente numéricos, así que un "1S" quedaba en la base generando cientos de pendientes.
    c.execute("""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
                 m.nombre AS "Marca",
                 (SELECT COUNT(*) FROM equivalencias e
                   WHERE e.producto_a_id = p.id OR e.producto_b_id = p.id) AS "Vinculos"
                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                 WHERE LENGTH(p.codigo_clean) <= 2
                 ORDER BY "Vinculos" DESC LIMIT ?""", (limite,))
    return [dict(r) for r in c.fetchall()]


def contar_codigos_basura():
    c.execute("SELECT COUNT(*) FROM productos WHERE LENGTH(codigo_clean) <= 2")
    return c.fetchone()[0]


def borrar_codigos_basura():
    """Borra esos productos. Las equivalencias falsas que colgaban de ellos se van solas por el
    ON DELETE CASCADE — que es justamente el punto de la limpieza.
    No van a la papelera a propósito: restaurarlos sería volver a meter la basura, y la lista de
    lo que se borra se puede bajar en Excel desde la pantalla antes de confirmar."""
    items = listar_codigos_basura(limite=100000)
    if not items:
        return 0
    with db_lock:
        c.executemany("DELETE FROM productos WHERE id = ?", [(i["ID"],) for i in items])
        conn.commit()
    return len(items)


def descargar_imagen(url, tiempo_maximo=12, tamano_maximo_mb=8):
    """Baja una imagen de una dirección web. Devuelve (bytes, error)."""
    import requests
    try:
        respuesta = requests.get(url, timeout=tiempo_maximo, stream=True,
                                  headers={"User-Agent": "Mozilla/5.0 (compatible; EquivalenciasElChavo/1.0)"})
        if respuesta.status_code != 200:
            return None, f"respondió {respuesta.status_code}"
        tipo = respuesta.headers.get("Content-Type", "")
        if "image" not in tipo.lower():
            return None, "la dirección no devuelve una imagen"
        datos = b""
        for bloque in respuesta.iter_content(65536):
            datos += bloque
            if len(datos) > tamano_maximo_mb * 1024 * 1024:
                return None, "la imagen pesa demasiado"
        return datos, None
    except Exception as e:
        return None, type(e).__name__


def buscar_imagen_en_ficha(url_ficha, tiempo_maximo=12):
    """Busca la foto del producto dentro de la ficha oficial del proveedor. Es la fuente más
    confiable que hay gratis: es la foto de ESE código, puesta por el propio proveedor.
    No existe ninguna base pública y gratuita de fotos por número de parte — la del rubro
    (TecDoc) es un servicio pago con licencia."""
    import requests
    from urllib.parse import urljoin
    try:
        respuesta = requests.get(url_ficha, timeout=tiempo_maximo,
                                  headers={"User-Agent": "Mozilla/5.0 (compatible; EquivalenciasElChavo/1.0)"})
        if respuesta.status_code != 200:
            return None, f"la ficha respondió {respuesta.status_code}"
        html = respuesta.text
        candidatas = []
        for m in re.finditer(r'<img[^>]+src=["\']([^"\']+)["\']', html, re.I):
            src = m.group(1)
            if any(x in src.lower() for x in ("logo", "icon", "sprite", "banner", "pixel", ".svg")):
                continue
            candidatas.append(urljoin(url_ficha, src))
        # Las fichas suelen tener la foto del producto en las primeras imágenes útiles
        for url_img in candidatas[:6]:
            datos, error = descargar_imagen(url_img)
            if datos and len(datos) > 8000:   # descarta íconos chiquitos
                return datos, None
        return None, "no encontré una foto de producto en esa ficha"
    except Exception as e:
        return None, type(e).__name__


def contar_fotos_por_bajar():
    """Productos cuya foto es un link externo (cargado desde Excel) y todavía no se bajó."""
    c.execute("""SELECT COUNT(*) FROM productos
                 WHERE imagen_url IS NOT NULL AND imagen_url LIKE 'http%'""")
    return c.fetchone()[0]


def _bajar_en_paralelo(tareas, funcion, hilos=6, progreso=None, cancelado=None):
    """Corre las bajadas en varios hilos a la vez y devuelve los resultados en orden de llegada.

    Por qué: cada foto es una espera de red de 1 a 3 segundos, y en fila india eso son 100 fotos
    en 3-5 minutos. La espera de una no impide empezar la otra, así que de a 6 en paralelo la
    misma tanda baja en menos de un minuto. Solo la parte de red va en hilos: escribir en la
    base se hace después, en el hilo principal, para no pelearse por el archivo.

    6 hilos y no 50 a propósito: es el sitio del proveedor el que atiende, y no corresponde
    martillarlo — además muchos cortan por exceso de pedidos y ahí no baja ninguna."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    resultados = []
    total = len(tareas)
    with ThreadPoolExecutor(max_workers=hilos) as pool:
        futuros = {pool.submit(funcion, t): t for t in tareas}
        for hechos, futuro in enumerate(as_completed(futuros), 1):
            tarea = futuros[futuro]
            try:
                datos, error = futuro.result()
            except Exception as e:
                datos, error = None, type(e).__name__
            resultados.append((tarea, datos, error))
            if progreso:
                progreso(hechos, total)
            if cancelado and cancelado():
                for f in futuros:
                    f.cancel()
                break
    return resultados


def bajar_fotos_pendientes(limite=200, progreso=None, hilos=6, liviano=True):
    """Baja las fotos que están como link externo y las guarda, con miniatura y firma visual."""
    c.execute("""SELECT id, imagen_url FROM productos
                 WHERE imagen_url IS NOT NULL AND imagen_url LIKE 'http%' LIMIT ?""", (limite,))
    pendientes = [(r["id"], r["imagen_url"]) for r in c.fetchall()]
    if not pendientes:
        return 0, []

    resultados = _bajar_en_paralelo(
        pendientes, lambda t: descargar_imagen(t[1]), hilos=hilos, progreso=progreso
    )

    bajadas, fallidas = 0, []
    for (pid, url), datos, error in resultados:
        if datos:
            try:
                actualizar_imagen_producto(pid, datos, origen="link", fuente=url, liviano=liviano)
                bajadas += 1
            except Exception as e:
                fallidas.append((url, type(e).__name__))
        else:
            fallidas.append((url, error))
    return bajadas, fallidas


FILTROS_FOTOS = {
    "todos": ("", "todos los códigos"),
    "stock": (" AND stock > 0", "solo los que tienen stock"),
    "precio": (" AND precio IS NOT NULL AND precio > 0", "solo los que tienen precio cargado"),
}


def _condicion_filtro_fotos(filtro):
    return FILTROS_FOTOS.get(filtro, FILTROS_FOTOS["todos"])[0]


def contar_fotos_por_traer_de_catalogo(marca_id, filtro="todos"):
    """Cuántos códigos de esa marca faltan probar. No cuenta los que ya se probaron y no tenían
    foto: esos ya no se reintentan, si no la tanda nunca avanzaba."""
    c.execute(f"""SELECT COUNT(*) FROM productos
                  WHERE marca_id = ? AND imagen_url IS NULL
                    AND (foto_busqueda_estado IS NULL OR foto_busqueda_estado = 'error')
                    {_condicion_filtro_fotos(filtro)}""", (marca_id,))
    return c.fetchone()[0]


def reintentar_codigos_sin_foto(marca_id=None):
    """Vuelve a habilitar los códigos marcados como 'sin foto en la ficha', por si el proveedor
    la subió después o se cayó el sitio justo esa vez."""
    with db_lock:
        if marca_id:
            c.execute("UPDATE productos SET foto_busqueda_estado = NULL WHERE marca_id = ?", (marca_id,))
        else:
            c.execute("UPDATE productos SET foto_busqueda_estado = NULL")
        cambiados = c.rowcount
        conn.commit()
    return cambiados


def bajar_fotos_desde_catalogo(marca_id, limite=100, progreso=None, hilos=6, liviano=True,
                               cancelado=None, filtro="todos"):
    """Para los productos de una marca sin foto: entra a la ficha oficial del proveedor de cada
    código y trae la foto de ahí. Los códigos cuya ficha no tiene foto quedan marcados para no
    volver a consultarlos en cada tanda."""
    c.execute("SELECT url_ficha_template FROM marcas WHERE id = ?", (marca_id,))
    fila = c.fetchone()
    if not fila or not fila["url_ficha_template"]:
        return 0, [("", "esa marca no tiene cargada la dirección de su catálogo")], 0
    plantilla = fila["url_ficha_template"]

    c.execute(f"""SELECT id, codigo_raw FROM productos
                  WHERE marca_id = ? AND imagen_url IS NULL
                    AND (foto_busqueda_estado IS NULL OR foto_busqueda_estado = 'error')
                    {_condicion_filtro_fotos(filtro)}
                  ORDER BY (stock > 0) DESC, id
                  LIMIT ?""", (marca_id, limite))
    pendientes = [(r["id"], r["codigo_raw"]) for r in c.fetchall()]
    if not pendientes:
        return 0, [], 0

    def traer(tarea):
        _, codigo = tarea
        url_ficha = plantilla.replace("{codigo}", quote(str(codigo), safe=""))
        return buscar_imagen_en_ficha(url_ficha)

    resultados = _bajar_en_paralelo(pendientes, traer, hilos=hilos, progreso=progreso,
                                    cancelado=cancelado)

    bajadas, fallidas, sin_foto = 0, [], 0
    for (pid, codigo), datos, error in resultados:
        url_ficha = plantilla.replace("{codigo}", quote(str(codigo), safe=""))
        if datos:
            try:
                actualizar_imagen_producto(pid, datos, origen="ficha", fuente=url_ficha,
                                            liviano=liviano)
                bajadas += 1
                continue
            except Exception as e:
                error = type(e).__name__
        fallidas.append((codigo, error))
        # "no encontré una foto" es definitivo para ese código; un error de red no lo es
        definitivo = error and "no encontré" in str(error)
        with db_lock:
            c.execute("UPDATE productos SET foto_busqueda_estado = ? WHERE id = ?",
                      ("sin_foto" if definitivo else "error", pid))
            conn.commit()
        if definitivo:
            sin_foto += 1
    return bajadas, fallidas, sin_foto


def peso_estimado_por_foto(liviano=True):
    """KB aproximados que ocupa cada foto en la base, para poder avisar antes de llenarla.
    Medido sobre fotos de producto reales: en liviano son la firma visual (~20 KB) más la
    miniatura (~2 KB); en completo se suma la imagen de 500px, que es la que pesa."""
    return 22 if liviano else 120


def generar_backup_sin_fotos():
    """Copia de la base SIN las fotos. Las fotos son lo que más pesa: con unos 1.000 productos
    con foto el archivo pasa los 100 MB que acepta GitHub, y ahí se pierde la copia de
    seguridad del repositorio justo cuando más datos hay para proteger.
    Todo lo demás va completo — catálogo, precios, equivalencias, vehículos, historial. Las
    fotos se vuelven a traer con los botones de Mantenimiento."""
    import tempfile
    ruta_temporal = os.path.join(tempfile.gettempdir(), "backup_sin_fotos.db")
    if os.path.exists(ruta_temporal):
        os.remove(ruta_temporal)
    destino = sqlite3.connect(ruta_temporal)
    with db_lock:
        conn.backup(destino)
    destino.execute("UPDATE productos SET imagen_url = NULL, imagen_thumb = NULL, imagen_orb_blob = NULL, "
                     "imagen_orb_estado = NULL")
    try:
        destino.execute("DELETE FROM producto_fotos")
    except sqlite3.OperationalError:
        pass
    try:
        destino.execute("UPDATE esquemas SET imagen_blob = NULL")
    except sqlite3.OperationalError:
        pass
    try:
        destino.execute("UPDATE alias_transferencia SET qr_real_blob = NULL")
    except sqlite3.OperationalError:
        pass
    destino.commit()
    destino.execute("VACUUM")   # sin esto el archivo sigue pesando lo mismo
    destino.commit()
    destino.close()
    with open(ruta_temporal, "rb") as f:
        datos = f.read()
    os.remove(ruta_temporal)
    return datos


def peso_de_las_fotos():
    """Cuánto de la base ocupan las fotos, para poder avisar antes de que sea un problema."""
    c.execute("""SELECT COUNT(*) AS con_foto,
                        COALESCE(SUM(LENGTH(COALESCE(imagen_url,'')) +
                                     LENGTH(COALESCE(imagen_thumb,''))), 0) AS bytes
                 FROM productos WHERE imagen_url IS NOT NULL""")
    fila = c.fetchone()
    total = fila["bytes"]
    try:
        c.execute("""SELECT COALESCE(SUM(LENGTH(COALESCE(imagen_data,'')) +
                                         LENGTH(COALESCE(firma_blob, X''))), 0) FROM producto_fotos""")
        total += c.fetchone()[0]
    except sqlite3.OperationalError:
        pass
    return fila["con_foto"], total / (1024 * 1024)


def guardar_mapeo_columnas(proveedor, idx_prov, idx_oem, idx_desc, idx_precio, idx_stock,
                            buscar_oem_en_desc, prov_es_oem):
    """Recuerda cómo se mapearon las columnas de este proveedor, para que la próxima vez venga
    preseleccionado igual y no haya que acertarle de nuevo."""
    if not proveedor or not proveedor.strip():
        return
    with db_lock:
        c.execute("""INSERT OR REPLACE INTO mapeo_columnas
                     (proveedor, idx_prov, idx_oem, idx_desc, idx_precio, idx_stock,
                      buscar_oem_en_desc, prov_es_oem, fecha)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))""",
                  (proveedor.strip().upper(), idx_prov, idx_oem, idx_desc, idx_precio, idx_stock,
                   1 if buscar_oem_en_desc else 0, 1 if prov_es_oem else 0))
        conn.commit()


def leer_mapeo_columnas(proveedor):
    if not proveedor or not proveedor.strip():
        return None
    c.execute("SELECT * FROM mapeo_columnas WHERE proveedor = ?", (proveedor.strip().upper(),))
    fila = c.fetchone()
    return dict(fila) if fila else None


# Piezas que se identifican POR SU MEDIDA: en retenes, o'rings, rulemanes y bujes, la medida
# ES el código. Un retén Taranto se pide como "35x52x7" y así figura en el catálogo.
_RE_PIEZA_POR_MEDIDA = re.compile(
    r'\b(RETEN|RETENES|O.?RING|ORING|ANILLO|JUNTA\s+TORICA|SELLO|BUJE|BUJES|'
    r'RULEMAN|RODAMIENTO|ARANDELA|ESPACIADOR|SEPARADOR)\b', re.I)


def codigo_sospechoso(codigo, descripcion=""):
    """¿Esto parece un código de repuesto de verdad? Devuelve (es_sospechoso, motivo).
    Sirve para cazar importaciones mal mapeadas: cuando la columna que se tomó como código
    en realidad tenía medidas, cantidades o pedazos de la descripción, quedan cargados como
    productos códigos tipo '0', '1200cc' o '20x2.50x180' — que después ensucian todo.

    OJO con las medidas: en retenes, o'rings, rulemanes y bujes la medida ES el código. Un
    retén Taranto se pide como «35x52x7» y así figura en el catálogo. Marcarlos como
    sospechosos hacía que TODOS los retenes cargados por medida aparecieran con alarma, que es
    justo lo contrario de lo que sirve. Por eso se mira la descripción antes de descartar."""
    es_por_medida = bool(descripcion and _RE_PIEZA_POR_MEDIDA.search(str(descripcion)))
    if codigo is None:
        return True, "está vacío"
    texto = str(codigo).strip()
    if not texto:
        return True, "está vacío"

    limpio = sanitizar(texto)
    if len(limpio) <= 2:
        return True, f"«{texto}» es demasiado corto para ser un código"
    if re.fullmatch(r"\d{1,3}", limpio):
        return True, f"«{texto}» es solo un número chico, no parece un código"
    if re.search(r"[Ee][+\-]\d+", texto):
        return True, f"«{texto}» quedó en notación científica de Excel (el número real se perdió)"
    if not es_por_medida:
        if re.search(r"\d\s*[xX]\s*\d+[.,]?\d*\s*[xX]?\s*\d*\s*(MM|mm)?$", texto) and "x" in texto.lower():
            return True, f"«{texto}» parece una medida, no un código"
        if re.search(r"\d+\s*(MM|CM|CC|ML|KG|GR|LTS?|V|W)\b", texto, re.I):
            return True, f"«{texto}» parece una medida o especificación"
        if "Ø" in texto or '"' in texto or "″" in texto:
            return True, f"«{texto}» tiene símbolos de medida (Ø o pulgadas)"
    if re.search(r"\b(DIESEL|NAFTA|SECTOR|CANAL|JUEGO|ARO|CHAPA|TIPO|MEDIDA)\b", texto, re.I):
        return True, f"«{texto}» parece un pedazo de la descripción"
    return False, None


def cb_ver_equivalencias(codigo_raw):
    """Callback para que cualquier código listado en pantalla sea clickeable: al tocarlo,
    busca sus equivalencias y las deja mostradas arriba, sin tener que copiar el código a mano
    y volver a buscarlo."""
    clean = sanitizar(codigo_raw)
    res = buscar_por_codigo(clean) if clean else []
    if res:
        incrementar_veces_buscado(clean)
    guardar_busqueda(codigo_raw)
    st.session_state["ultima_busqueda_codigo"] = [
        {"codigo_individual": codigo_raw, "clean": clean, "res": res}
    ]
    st.session_state["sugerencia_busqueda"] = codigo_raw
    st.session_state["modo_busqueda"] = "Código"   # por si se tocó desde la búsqueda por texto


def mostrar_lista_clickeable(filas, prefijo_key, limite=15, nota=None):
    """Muestra resultados con el código como botón: al tocarlo se abren sus equivalencias.
    Antes esto era una tabla y había que ir copiando los códigos de a uno para buscarlos."""
    if nota:
        st.caption(nota)
    for f in filas[:limite]:
        col_cod, col_desc = st.columns([1.2, 3])
        col_cod.button(f"🔎 {f['Codigo']}", key=f"{prefijo_key}_{f['ID']}",
                        on_click=cb_ver_equivalencias, args=(f["Codigo"],),
                        help="Ver sus equivalencias")
        descripcion = (f.get("Descripcion") or "")[:90]
        precio = f"${f['Precio']:,.0f}" if f.get("Precio") else ""
        stock = f" · stock {f['Stock']}" if f.get("Stock") is not None else ""
        col_desc.caption(f"**{f.get('Marca', '')}** {descripcion}  \n{precio}{stock}")
    if len(filas) > limite:
        st.caption(f"(mostrando {limite} de {len(filas)})")


def buscar_codigos_parecidos(clean_code, limite=30):
    """Cuando el código exacto no aparece, busca códigos que EMPIECEN igual o que lo contengan.
    Es el caso típico de las familias: pedís 'TC-421' y en la base están 'TC-421-15' y
    'TC-421-20' (mismo repuesto, distinto espesor/variante). Antes eso no aparecía por ningún
    lado, porque la búsqueda por código exige coincidencia exacta."""
    if not clean_code or len(clean_code) < 3:
        return []
    c.execute("""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
                        m.nombre AS "Marca", m.tipo AS "Tipo", p.precio AS "Precio",
                        p.stock AS "Stock", p.codigo_clean AS "_clean"
                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                 WHERE p.codigo_clean LIKE ? OR p.codigo_clean LIKE ?
                 ORDER BY CASE WHEN p.codigo_clean LIKE ? THEN 0 ELSE 1 END,
                          LENGTH(p.codigo_clean), p.codigo_clean
                 LIMIT ?""",
              (f"{clean_code}%", f"%{clean_code}%", f"{clean_code}%", limite))
    return filas_a_listas(c)


def _distancia_edicion(a, b, tope=2):
    """Cuántos cambios de un carácter hacen falta para pasar de un código al otro.
    Corta apenas supera el tope: no interesa saber si son 7 u 8, solo si están cerca."""
    if abs(len(a) - len(b)) > tope:
        return tope + 1
    anterior = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        actual = [i]
        for j, cb in enumerate(b, 1):
            actual.append(min(anterior[j] + 1, actual[j - 1] + 1,
                              anterior[j - 1] + (ca != cb)))
        if min(actual) > tope:
            return tope + 1
        anterior = actual
    return anterior[-1]


def codigos_por_tipeo(clean_code, limite=10):
    """Códigos que se escriben casi igual al que buscaste. Es para el error de tipeo.

    Lo que ya existía busca códigos que EMPIECEN igual o que lo CONTENGAN, y eso no sirve
    cuando el error está en el medio: pidiendo 'W71249' en vez de 'W71294' no aparece nada,
    aunque sea el mismo filtro con dos dígitos cambiados de lugar. Acá se mide cuántos
    caracteres hay que cambiar para pasar de uno al otro, así que agarra los dígitos dados
    vuelta, la letra de más y la que falta.

    El truco para que no tarde: el código se parte en tres pedazos y se le pide a la base que
    el candidato contenga AL MENOS UNO igual. Con dos errores como mucho se pueden arruinar dos
    pedazos, así que el tercero sobrevive sí o sí — no se pierde ningún candidato y la base
    devuelve un puñado en vez de todo el catálogo. Sin esto, con 40.000 productos tardaba casi
    un segundo en cada búsqueda fallida."""
    if not clean_code or len(clean_code) < 4:
        return []
    largo = len(clean_code)
    tercio = max(largo // 3, 1)
    # Los tres pedazos van SIEMPRE, aunque alguno quede de un solo carácter. Descartar los
    # cortos rompía la garantía: con 'OC9O' quedaba solo el pedazo '9O', y el código correcto
    # 'OC90' no lo contiene, así que se perdía. Son exactamente tres para que dos errores no
    # puedan arruinarlos a todos.
    pedazos = [clean_code[:tercio], clean_code[tercio:tercio * 2], clean_code[tercio * 2:]]
    pedazos = [p for p in pedazos if p]

    condiciones = " OR ".join("p.codigo_clean LIKE ?" for _ in pedazos)
    params = [f"%{p}%" for p in pedazos] + [largo - 2, largo + 2, clean_code]
    c.execute(f"""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
                         m.nombre AS "Marca", m.tipo AS "Tipo", p.precio AS "Precio",
                         p.stock AS "Stock", p.codigo_clean AS "_clean"
                  FROM productos p JOIN marcas m ON m.id = p.marca_id
                  WHERE ({condiciones})
                    AND LENGTH(p.codigo_clean) BETWEEN ? AND ?
                    AND p.codigo_clean <> ?""", params)

    candidatos = []
    for fila in filas_a_listas(c):
        d = _distancia_edicion(clean_code, fila["_clean"])
        if d <= 2:
            fila["_dist"] = d
            candidatos.append(fila)
    # Primero los de un solo carácter de diferencia, y dentro de esos los que tienen stock
    candidatos.sort(key=lambda f: (f["_dist"], -(f["Stock"] or 0)))
    return candidatos[:limite]


def precios_incoherentes_entre_equivalentes(factor=8, limite=200):
    """Pares de productos marcados como equivalentes cuyos precios no se parecen en nada.

    Dos repuestos que hacen lo mismo pueden costar distinto según la marca, pero no ocho veces
    distinto. Cuando pasa eso, una de dos cosas está mal, y las dos importan:
      - el precio (se importó una columna equivocada o el separador de decimales al revés), o
      - la equivalencia (los vinculó una lista mal cargada y no son la misma pieza).
    Cualquiera de las dos que sea, es plata: o cotizás mal, o vendés lo que no entra."""
    c.execute("""SELECT pa.codigo_raw AS "Código A", ma.nombre AS "Marca A", pa.precio AS "Precio A",
                        pb.codigo_raw AS "Código B", mb.nombre AS "Marca B", pb.precio AS "Precio B",
                        ROUND(MAX(pa.precio, pb.precio) / MIN(pa.precio, pb.precio), 1) AS "Veces",
                        pa.descripcion AS "Descripción",
                        pa.id AS "_ida", pb.id AS "_idb"
                 FROM equivalencias e
                 JOIN productos pa ON pa.id = e.producto_a_id
                 JOIN productos pb ON pb.id = e.producto_b_id
                 JOIN marcas ma ON ma.id = pa.marca_id
                 JOIN marcas mb ON mb.id = pb.marca_id
                 WHERE pa.precio > 0 AND pb.precio > 0
                   AND MAX(pa.precio, pb.precio) / MIN(pa.precio, pb.precio) >= ?
                 ORDER BY "Veces" DESC LIMIT ?""", (factor, limite))
    return filas_a_listas(c)


# Se ordena de más largo a más corto para que gane la coincidencia más específica:
# "LAND ROVER" tiene que ganarle a "ROVER", y "MERCEDES BENZ" a "MERCEDES".
# No se agregan marcas de 2 letras (MG, DS) porque en una descripción de repuesto casi siempre
# son otra cosa (medidas, siglas) y ensuciarían más de lo que aportan.
MARCAS_VEHICULO = sorted(set([
    # --- Las que ya estaban ---
    "MERCEDES BENZ", "M.BENZ", "MERCEDES", "VOLKSWAGEN", "CHEVROLET", "MITSUBISHI", "LAND ROVER",
    "MASSEY FERGUSON", "M. FERGUSON", "AGCO SISU POWER", "JOHN DEERE", "NEW HOLLAND", "ALFA ROMEO",
    "CITROEN", "PEUGEOT", "RENAULT", "CHRYSLER", "CUMMINS", "PERKINS", "ILLINOIS", "TOYOTA",
    "NISSAN", "HYUNDAI", "DAEWOO", "SUZUKI", "SCANIA", "IVECO", "AGRALE", "DEUTZ", "VOLVO",
    "HONDA", "DODGE", "BURMOR", "M.W.M.", "M.W.M", "MWM", "KNORR", "VARGA", "VALTRA", "ZANELLO",
    "PAUNY", "FIAT", "FORD", "JEEP", "AUDI", "SEAT", "BMW", "KIA", "MAN", "DAF", "HINO",
    "ISUZU", "CASE", "GM",
    # --- Chinas, que hoy son parte del parque argentino ---
    "GREAT WALL", "DONGFENG", "SHINERAY", "CHANGAN", "BAIC", "CHERY", "GEELY", "HAVAL",
    "FOTON", "LIFAN", "JINBEI", "MAXUS", "JETOUR", "SOUEAST", "BYD", "JAC",
    # --- Coreanas, japonesas y del resto de Asia ---
    "SSANGYONG", "MAHINDRA", "SUBARU", "DAIHATSU", "LEXUS", "INFINITI", "ACURA", "GENESIS",
    "MAZDA", "TATA", "PROTON",
    # --- Europeas que faltaban ---
    "VAUXHALL", "LAMBORGHINI", "ROLLS ROYCE", "ASTON MARTIN", "MASERATI", "PORSCHE", "FERRARI",
    "BENTLEY", "SKODA", "LANCIA", "JAGUAR", "DACIA", "OPEL", "SMART", "MINI", "LADA", "ROVER",
    "ZASTAVA", "YUGO", "TALBOT", "SAAB", "LOTUS", "MCLAREN",
    # --- Norteamericanas ---
    "INTERNATIONAL", "FREIGHTLINER", "WESTERN STAR", "OLDSMOBILE", "KENWORTH", "PETERBILT",
    "PLYMOUTH", "CADILLAC", "PONTIAC", "LINCOLN", "MERCURY", "SATURN", "NAVISTAR", "BUICK",
    "TESLA", "MACK", "GMC", "RAM",
    # --- Camiones, ómnibus, agro e industria ---
    "MERCEDES-BENZ", "LANDINI", "MARCOPOLO", "METALPAR", "RASTROJERO", "CATERPILLAR", "YANMAR",
    "KUBOTA", "CLAAS", "FENDT", "JCB", "VETRA", "APACHE", "SEVEL", "SIAM", "DI TELLA",
    # --- Motos ---
    "ROYAL ENFIELD", "HARLEY DAVIDSON", "MOTO GUZZI", "MV AGUSTA", "KAWASAKI", "HUSQVARNA",
    "MOTOMEL", "GUERRERO", "ZANELLA", "KYMCO", "APRILIA", "YAMAHA", "DUCATI", "PIAGGIO",
    "BENELLI", "CORVEN", "KELLER", "GILERA", "MONDIAL", "BAJAJ", "VESPA", "KTM", "SYM",
    "HERO", "TVS", "BETA",
]), key=len, reverse=True)


def extraer_anios(descripcion):
    """Saca el rango de años de una descripción. Las listas los escriben de varias formas:
    '1969/78' (1969 a 1978), '1998/...' (1998 en adelante), '2005' (solo ese año).
    Devuelve (desde, hasta) — hasta=None significa 'en adelante'."""
    if not descripcion:
        return None, None
    texto = str(descripcion)
    # Rango con dos años: 1969/78, 1974/1981, 2005/09
    m = re.search(r'\b(19\d{2}|20\d{2})\s*/\s*(\d{2}|\d{4})\b', texto)
    if m:
        desde = int(m.group(1))
        fin = m.group(2)
        hasta = int(fin) if len(fin) == 4 else int(str(desde)[:2] + fin)
        if hasta < desde:            # 1998/02 significa 1998 a 2002
            hasta += 100
        return desde, hasta
    # Año en adelante: 1998/... o 1998/…
    m = re.search(r'\b(19\d{2}|20\d{2})\s*/\s*\.{2,}', texto)
    if m:
        return int(m.group(1)), None
    # Un año suelto
    m = re.search(r'\b(19\d{2}|20\d{2})\b', texto)
    if m:
        return int(m.group(1)), int(m.group(1))
    return None, None


def sirve_para_anio(descripcion, anio):
    """¿Este repuesto aplica a un vehículo de ese año? Si la descripción no dice nada de años,
    devuelve None: no se puede afirmar ni descartar, y es mejor mostrarlo que ocultarlo."""
    desde, hasta = extraer_anios(descripcion)
    if desde is None:
        return None
    if hasta is None:
        return anio >= desde
    return desde <= anio <= hasta


PALABRAS_NO_MODELO = {
    "JUNTA", "JUNTAS", "JUEGO", "DESPIECE", "TAPA", "CILINDROS", "VALVULAS", "CARTER", "BOMBA",
    "ACEITE", "AGUA", "COMBUSTIBLE", "NAFTA", "TERMOSTATO", "RETEN", "ARO", "AROS", "PISTON",
    "CIL", "CILINDRO", "MOTOR", "SERIE", "PICK", "UP", "BUS", "CAMION", "TRACTOR", "DIESEL",
    "TURBO", "INY", "INYECCION", "CID", "DOHC", "SOHC", "MM", "CC", "STD", "COMPLETO", "SIN",
    "CON", "PARA", "DE", "DEL", "LA", "EL", "Y", "O", "REPARACION", "ADMISION", "ESCAPE",
    "CARBURADOR", "DESCARBONIZACION", "LATERAL", "SUPLEMENTO", "CAPERUZA", "BOLILLEROS",
    "DIRECCION", "MECANICA", "AGRICOLA", "CARGO", "GRAND", "SEMI", "ORING", "ARANDELA",
    "ALUMINIO", "CLAVITO", "BANCADA", "CAPUCHON", "BUJIA", "BRIDA", "CAÑO", "CALEFACCION",
    "ARBOL", "LEVAS", "SALIDA", "TAPON", "VALVULA", "MARIPOSA", "BASE", "DISTRIBUIDOR",
    "CHUPADOR", "INTERMEDIA", "V", "L", "S", "R", "AX", "DD", "F",
}


@st.cache_data(show_spinner=False, max_entries=20)
def modelos_de_marca(marca_vehiculo, _version, minimo=2):
    """Arma la lista de modelos de una marca leyendo el catálogo.

    Cómo distingue un modelo de una palabra de repuesto: las palabras de repuesto (JUNTA,
    BOMBA, ORING) aparecen en MUCHAS marcas distintas, mientras que un modelo aparece casi
    solo en la suya — un ASTRA es de Chevrolet y de nadie más. Con eso se filtra sin tener
    que cargar una lista de modelos a mano, y crece solo con cada lista nueva que importás."""
    c.execute("""SELECT p.descripcion FROM productos p
                 WHERE p.descripcion IS NOT NULL AND UPPER(p.descripcion) LIKE ?
                 LIMIT 4000""", (f"%{marca_vehiculo}%",))
    propias = [r["descripcion"] for r in c.fetchall()]

    from collections import Counter
    cuenta_propia = Counter()
    for desc in propias:
        _, marca_det, resto = separar_por_marca_vehiculo(desc)
        if marca_det != marca_vehiculo or not resto:
            continue
        for token in re.findall(r"[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ0-9\-]{2,}", resto.upper()):
            if token in PALABRAS_NO_MODELO or token in MARCAS_VEHICULO:
                continue
            if re.fullmatch(r"[\d\-]+", token):
                continue
            cuenta_propia[token] += 1

    if not cuenta_propia:
        return []

    # Cuántas veces aparece cada palabra en el catálogo entero (para descartar las genéricas)
    candidatos = [t for t, n in cuenta_propia.items() if n >= minimo]
    modelos = []
    for token in candidatos:
        c.execute("""SELECT COUNT(*) FROM productos
                     WHERE descripcion IS NOT NULL AND UPPER(descripcion) LIKE ?""",
                  (f"%{token}%",))
        total_catalogo = c.fetchone()[0] or 1
        # Si la mayoría de las veces que aparece es dentro de esta marca, es un modelo suyo
        if cuenta_propia[token] / total_catalogo >= 0.6:
            modelos.append((token, cuenta_propia[token]))
    return sorted(modelos, key=lambda x: -x[1])


def separar_por_marca_vehiculo(descripcion):
    """Parte una descripción en (categoría, marca del vehículo, resto).
    Ejemplo: 'Junta Tapa de Cilindros FORD TAUNUS COUPE'
             -> ('Junta Tapa de Cilindros', 'FORD', 'TAUNUS COUPE')
    Es lo que permite armar el catálogo por vehículo sin cargar nada a mano: la relación
    pieza-vehículo ya venía en las listas de los proveedores, solo hay que leerla."""
    if not descripcion:
        return None, None, None
    texto = separar_texto_pegado(str(descripcion))
    for marca in MARCAS_VEHICULO:
        patron = re.compile(r'(?<![A-Za-zÁÉÍÓÚÑ])' + re.escape(marca) + r'(?![A-Za-zÁÉÍÓÚÑ])')
        m = patron.search(texto)
        if m:
            categoria = texto[:m.start()].strip(" -,/")
            resto = texto[m.end():].strip(" -,/")
            return (categoria or None), marca, (resto or None)
    return texto.strip() or None, None, None


@st.cache_data(show_spinner=False, max_entries=30)
def catalogo_por_vehiculo(marca_vehiculo, _version):
    """Todos los productos cuya descripción menciona esa marca de vehículo, agrupados por
    categoría. '_version' solo sirve para que el caché se refresque cuando cambia el catálogo."""
    c.execute("""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
                        m.nombre AS "Marca", p.precio AS "Precio", p.stock AS "Stock"
                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                 WHERE p.descripcion IS NOT NULL AND UPPER(p.descripcion) LIKE ?
                 LIMIT 4000""", (f"%{marca_vehiculo}%",))
    filas = filas_a_listas(c)
    por_categoria = {}
    for f in filas:
        categoria, marca_detectada, resto = separar_por_marca_vehiculo(f["Descripcion"])
        if marca_detectada != marca_vehiculo:
            continue
        f["_categoria"] = categoria or "Sin categoría"
        f["_aplicacion"] = resto or ""
        por_categoria.setdefault(f["_categoria"], []).append(f)
    return por_categoria


@st.cache_data(show_spinner=False, max_entries=5)
def marcas_vehiculo_disponibles(_version):
    """Qué marcas de vehículo aparecen realmente en el catálogo cargado."""
    disponibles = []
    for marca in MARCAS_VEHICULO:
        c.execute("""SELECT COUNT(*) FROM productos
                     WHERE descripcion IS NOT NULL AND UPPER(descripcion) LIKE ?""",
                  (f"%{marca}%",))
        cantidad = c.fetchone()[0]
        if cantidad:
            disponibles.append((marca, cantidad))
    return sorted(disponibles, key=lambda x: -x[1])


# Familias de repuestos. Gana la palabra clave MÁS LARGA que aparezca en la descripción, así
# "BOMBA DE AGUA" (refrigeración) le gana a "BOMBA" y "BOMBA DE ACEITE" no cae en el mismo lado.
FAMILIAS_REPUESTO = {
    "Filtros": [
        "FILTRO DE ACEITE", "FILTRO DE AIRE", "FILTRO DE COMBUSTIBLE", "FILTRO DE NAFTA",
        "FILTRO DE GASOIL", "FILTRO DE HABITACULO", "FILTRO DE CABINA", "FILTRO SECADOR",
        "FILTRO", "SEPARADOR DE AGUA",
    ],
    "Frenos": [
        "PASTILLA DE FRENO", "PASTILLAS DE FRENO", "DISCO DE FRENO", "CAMPANA DE FRENO",
        "CILINDRO DE FRENO", "BOMBA DE FRENO", "ZAPATA DE FRENO", "CABLE DE FRENO",
        "LATIGUILLO", "SERVOFRENO", "PASTILLA", "PASTILLAS", "ZAPATA", "ZAPATAS", "CALIPER",
        "MORDAZA", "CAMPANA", "FRENO", "FRENOS", "ABS",
    ],
    "Suspensión": [
        "AMORTIGUADOR", "ESPIRAL", "ELASTICO", "ROTULA", "BIELETA", "BARRA ESTABILIZADORA",
        "BRAZO SUSPENSION", "PARRILLA", "SOPORTE DE AMORTIGUADOR", "KIT DE SUSPENSION",
        "BUJE DE PARRILLA", "TREN DELANTERO", "TOPE DE SUSPENSION", "FUELLE DE AMORTIGUADOR",
        "SUSPENSION", "MUNON",
    ],
    "Dirección": [
        "CREMALLERA", "EXTREMO DE DIRECCION", "AXIAL DE DIRECCION", "BOMBA DE DIRECCION",
        "BRAZO PITMAN", "BIELETA DE DIRECCION", "COLUMNA DE DIRECCION", "EXTREMO", "AXIAL",
        "DIRECCION",
    ],
    "Palier y transmisión": [
        "HOMOCINETICA", "PUNTA DE EJE", "SEMIEJE", "PALIER", "CRUCETA", "CARDAN",
        "FUELLE DE PALIER", "FUELLE DE HOMOCINETICA", "JUNTA HOMOCINETICA", "TRIPOIDE",
    ],
    "Embrague": [
        "DISCO DE EMBRAGUE", "PLATO DE EMBRAGUE", "PLACA DE EMBRAGUE", "KIT DE EMBRAGUE",
        "COLLARIN", "RULEMAN DE EMPUJE", "CILINDRO DE EMBRAGUE", "BOMBA DE EMBRAGUE",
        "CABLE DE EMBRAGUE", "EMBRAGUE", "VOLANTE MOTOR",
    ],
    "Caja y diferencial": [
        "CAJA DE VELOCIDAD", "CORONA Y PINON", "SINCRONIZADO", "DIFERENCIAL", "SATELITE",
        "CAJA DE CAMBIOS", "PALANCA DE CAMBIOS",
    ],
    "Distribución": [
        "CORREA DE DISTRIBUCION", "KIT DE DISTRIBUCION", "CADENA DE DISTRIBUCION",
        "TENSOR DE DISTRIBUCION", "CORREA POLY V", "CORREA DENTADA", "DISTRIBUCION",
        "TENSOR", "POLEA", "CORREA",
    ],
    "Refrigeración": [
        "BOMBA DE AGUA", "RADIADOR", "TERMOSTATO", "ELECTROVENTILADOR", "TAPA DE RADIADOR",
        "MANGUERA DE RADIADOR", "INTERCOOLER", "DEPOSITO DE AGUA", "REFRIGERACION",
        "VENTILADOR",
    ],
    "Lubricación": [
        "BOMBA DE ACEITE", "CARTER", "ENFRIADOR DE ACEITE", "VARILLA DE ACEITE",
        "TAPA DE VALVULAS", "MALLA DE ACEITE",
    ],
    "Motor - interno": [
        "PISTON", "PISTONES", "ARO DE PISTON", "AROS", "COJINETE", "BIELA", "CIGUENAL",
        "ARBOL DE LEVAS", "VALVULA DE ADMISION", "VALVULA DE ESCAPE", "GUIA DE VALVULA",
        "BUJE DE BIELA", "CAMISA", "CULATA", "TAPA DE CILINDRO", "BANCADA", "PERNO",
        "VALVULA", "VALVULAS", "RESORTE DE VALVULA",
    ],
    "Juntas y retenes": [
        # Abreviaturas que usan los proveedores en sus listas: sin ellas, «Jgo de motor» y
        # «JTA T.C.» no caían en ninguna familia y se colaban en cualquier búsqueda.
        "JGO DE MOTOR", "JUEGO DE MOTOR", "JGO MOTOR", "JGO DE JUNTAS", "JGO JUNTAS",
        "JTA T C", "JTA TC", "JTA DE TAPA", "JTA TAPA", "JTA",
        "JUNTA DE TAPA", "JUNTA TAPA", "JUEGO DE JUNTAS", "JUNTA HOMOCINETICA", "RETEN",
        "RETENES", "JUNTA", "JUNTAS", "ORING", "O-RING", "EMPAQUETADURA", "SELLO",
    ],
    "Combustible": [
        "BOMBA DE NAFTA", "BOMBA DE COMBUSTIBLE", "INYECTOR", "CARBURADOR", "RIEL DE INYECCION",
        "REGULADOR DE PRESION", "TANQUE DE COMBUSTIBLE", "AFORADOR", "INYECCION",
    ],
    "Escape": [
        "CANO DE ESCAPE", "SILENCIADOR", "CATALIZADOR", "SONDA LAMBDA", "MULTIPLE DE ESCAPE",
        "ESCAPE",
    ],
    "Eléctrico y encendido": [
        "CABLE DE BUJIA", "BUJIA", "BUJIAS", "BOBINA DE ENCENDIDO", "ALTERNADOR",
        "MOTOR DE ARRANQUE", "BURRO DE ARRANQUE", "BATERIA", "REGULADOR DE VOLTAJE",
        "DISTRIBUIDOR", "PLATINO", "SENSOR", "MODULO", "RELE", "FUSIBLE", "BOBINA",
        "CAPUCHON DE BUJIA",
    ],
    "Rodamientos y mazas": [
        "RULEMAN DE RUEDA", "MAZA DE RUEDA", "CUBO DE RUEDA", "RODAMIENTO", "RULEMAN",
        "BOLILLERO",
    ],
    "Climatización": [
        "COMPRESOR DE AIRE", "CONDENSADOR", "EVAPORADOR", "AIRE ACONDICIONADO",
        "FILTRO DE POLEN", "CALEFACCION",
    ],
    "Soportes y bujes": [
        "SOPORTE DE MOTOR", "SOPORTE DE CAJA", "BUJE", "BUJES", "TACO DE MOTOR", "SOPORTE",
    ],
    "Cables y comandos": [
        "CABLE DE ACELERADOR", "CABLE DE VELOCIMETRO", "CABLE DE CAPOT", "GUAYA", "CABLE",
    ],
    "Carrocería y accesorios": [
        "OPTICA", "FARO", "ESPEJO", "PARAGOLPE", "MANIJA", "CERRADURA", "BURLETE",
        "ESCOBILLA", "PARABRISAS", "GUARDABARRO", "CAPOT", "PARRILLA DE RADIADOR",
        "PLUMA", "CRIQUE",
    ],
}


def _normalizar_desc(texto):
    """Mayúsculas, sin acentos y con espacios simples, para poder comparar contra las claves."""
    return " " + " ".join(normalizar_texto(str(texto or "")).replace("-", " ").split()) + " "


def clasificar_repuesto(descripcion):
    """Devuelve a qué familia pertenece un repuesto, mirando su descripción.

    Por qué hace falta: antes la 'categoría' era literalmente el texto que venía antes de la
    marca del auto en la descripción. Como cada proveedor la escribe distinto ('JUNTA TAPA DE
    CILINDROS', 'JUNTA DE TAPA CIL.', 'JUEGO JUNTA TAPA'), salían cientos de categorías casi
    iguales repetidas, y el filtro no servía para encontrar nada.

    Gana la palabra clave que aparece PRIMERO; entre las que empiezan en el mismo lugar, la
    más larga. Ese orden no es un capricho: en las listas de repuestos el nombre de la pieza va
    al principio y lo que sigue es dónde va o de qué auto es. Con solo mirar el largo,
    'RETEN DELANTERO CIGUENAL' caía en Motor por 'CIGUENAL' en vez de en Retenes, que es lo que
    la pieza realmente es. Y el desempate por largo resuelve el otro caso: 'BOMBA DE AGUA' cae
    en Refrigeración y no en la misma bolsa que 'BOMBA DE ACEITE' o 'BOMBA DE FRENO'."""
    texto = _normalizar_desc(descripcion)
    if not texto.strip():
        return "Sin clasificar"
    mejor = None   # (posición, -largo, familia)
    for familia, claves in FAMILIAS_REPUESTO.items():
        for clave in claves:
            pos = texto.find(f" {clave} ")
            if pos >= 0:
                candidato = (pos, -len(clave), familia)
                if mejor is None or candidato[:2] < mejor[:2]:
                    mejor = candidato
    return mejor[2] if mejor else "Sin clasificar"


# ============================================================
# CATÁLOGOS DE APLICACIONES (qué repuesto le va a cada auto)
# ============================================================
# Es OTRA cosa que una lista de precios. Una lista dice "este código cuesta tanto"; un catálogo
# de aplicaciones dice "a este auto le va este código". Es justamente el dato que no existe
# gratis de forma general —para eso están las bases licenciadas tipo TecDoc—, pero varios
# fabricantes publican el suyo: NGK, Bosch, Mann, SKF. Cargando esos catálogos, la búsqueda por
# VIN o por vehículo deja de adivinar desde las descripciones y pasa a tener el dato real.

# Fila que es solo el nombre de la marca: una celda con texto y el resto vacío
_RE_CONTINUACION = re.compile(r'\s*-?\s*(continua[çc][ãa]o|continuaci[óo]n|cont\.?)\s*$', re.I)


def _vacia(v):
    return v is None or str(v).strip() == ""


def _limpiar(v):
    return "" if _vacia(v) else " ".join(str(v).split())


def parsear_anios(texto):
    """Saca (desde, hasta) de las formas en que se escriben los años en estos catálogos:
    '2013 a 2020', '05/1999 a 08/2000', 'Desde 2000', 'Até 2005', '2011'."""
    t = _limpiar(texto)
    if not t:
        return None, None
    anios = [int(a) for a in re.findall(r'(19\d{2}|20\d{2})', t)]
    if not anios:
        return None, None
    bajo, alto = min(anios), max(anios)
    if re.search(r'desde|a partir', t, re.I):
        return bajo, None
    if re.search(r'\bat[ée]\b|hasta', t, re.I):
        return None, alto
    return bajo, (alto if alto != bajo else None)


def es_fila_de_marca(fila):
    """¿Es el título de una marca? Texto solo en la primera celda y nada más en la fila."""
    if _vacia(fila[0]):
        return False
    if any(not _vacia(x) for x in fila[1:]):
        return False
    texto = _limpiar(fila[0])
    # Los títulos de marca son cortos y en mayúsculas; "FIAT - Continuação" también cuenta
    base = _RE_CONTINUACION.sub("", texto).strip()
    if not base or len(base) > 30:
        return False
    letras = [ch for ch in base if ch.isalpha()]
    return bool(letras) and sum(1 for ch in letras if ch.isupper()) / len(letras) > 0.7


_RE_ANIO = re.compile(r'(19\d{2}|20\d{2})')


def parece_catalogo_de_aplicaciones(filas, muestra=200):
    """¿Este archivo es un catálogo de aplicaciones y no una lista de precios?

    Existe para evitar un error caro y silencioso: los dos se abren igual, y si un catálogo de
    aplicaciones entra por el importador de listas, carga los MODELOS DE AUTO como si fueran
    códigos de repuesto ('A4', 'Q3', 'S3') y los años como si fueran precios.

    Las señales, que no aparecen nunca en una lista de precios:
      · filas con una sola celda, en mayúsculas, que son el nombre de una marca de auto
      · una columna llena de rangos de años ('2013 a 2020', 'Desde 2000')
      · la primera columna que se repite mucho, porque el modelo se escribe una sola vez y
        las motorizaciones de abajo la dejan vacía
    Devuelve (True/False, motivo)."""
    datos = [f for f in filas[:muestra] if f and any(x is not None and str(x).strip() for x in f)]
    if len(datos) < 10:
        return False, ""

    marcas_solas = 0
    for fila in datos:
        try:
            if es_fila_de_marca(fila) and _limpiar(fila[0]).upper() in MARCAS_VEHICULO:
                marcas_solas += 1
        except Exception:
            continue

    ancho = max(len(f) for f in datos)
    col_con_anios = 0
    for i in range(ancho):
        con_rango = sum(1 for f in datos
                        if i < len(f) and f[i] is not None
                        and re.search(r'(19|20)\d{2}\s*(a|-|hasta|até)\s*(19|20)\d{2}'
                                      r'|desde\s*(19|20)\d{2}|at[ée]\s*(19|20)\d{2}',
                                      str(f[i]), re.I))
        if con_rango >= len(datos) * 0.3:
            col_con_anios += 1

    razones = []
    if marcas_solas >= 2:
        razones.append(f"{marcas_solas} fila(s) son solo el nombre de una marca de auto")
    if col_con_anios:
        razones.append("hay una columna de rangos de años")
    # Hacen falta las dos señales: una sola se puede dar en una lista de precios común
    if marcas_solas >= 2 and col_con_anios:
        return True, " y ".join(razones)
    return False, ""


def detectar_columnas_aplicaciones(filas):
    """Encuentra en qué columna está cada cosa, mirando los valores.

    Hace falta porque un mismo catálogo mezcla formatos: el de NGK tiene tablas de 5 columnas y
    otras de 13, con los años y el código en posiciones distintas. Leer las de 13 con los
    índices de las de 5 hace que la columna de años se cargue como si fuera el código — y ahí
    terminan entrando 'Até 1991' o 'Desde 2005' como si fueran números de repuesto."""
    if not filas:
        return {"modelo": 0, "motor": 1, "comb": 2, "anios": 3, "codigo": 4}
    ancho = max(len(f) for f in filas)
    puntajes_anio = [0] * ancho
    puntajes_cod = [0] * ancho
    for fila in filas:
        for i in range(min(len(fila), ancho)):
            v = _limpiar(fila[i])
            if not v:
                continue
            if _RE_ANIO.search(v) or re.search(r'desde|at[ée]|hasta', v, re.I):
                puntajes_anio[i] += 1
            # Un código: tiene letras Y números, sin espacios de por medio, y no es un año
            elif re.fullmatch(r'[A-Z0-9][A-Z0-9\-./ ]{2,28}', v.upper()) and \
                    any(ch.isdigit() for ch in v) and any(ch.isalpha() for ch in v):
                puntajes_cod[i] += 1

    col_anios = max(range(ancho), key=lambda i: puntajes_anio[i]) if any(puntajes_anio) else 3
    # El código va DESPUÉS de los años en estos catálogos, y nunca es la columna del modelo
    candidatas = [i for i in range(ancho) if i not in (0, 1, col_anios)]
    col_codigo = max(candidatas, key=lambda i: puntajes_cod[i]) if candidatas else ancho - 1
    return {"modelo": 0, "motor": 1, "comb": min(2, ancho - 1),
            "anios": col_anios, "codigo": col_codigo}


def parsear_catalogo_aplicaciones(filas, col_modelo=0, col_motor=1, col_comb=2,
                                  col_anios=3, col_codigo=4):
    """Devuelve una lista de aplicaciones: qué código le corresponde a qué auto."""
    apps = []
    marca_actual = ""
    modelo_actual = ""

    for fila in filas:
        if not fila or len(fila) <= col_codigo:
            continue
        if all(_vacia(x) for x in fila):
            continue

        if es_fila_de_marca(fila):
            marca_actual = _RE_CONTINUACION.sub("", _limpiar(fila[col_modelo])).strip().upper()
            modelo_actual = ""      # al cambiar de marca se olvida el modelo anterior
            continue

        codigo = _limpiar(fila[col_codigo])
        if not codigo:
            continue

        # Si la primera celda trae algo, es un modelo nuevo. Si está vacía, sigue el anterior:
        # el catálogo no repite el nombre del modelo en cada motorización.
        if not _vacia(fila[col_modelo]):
            modelo_actual = _limpiar(fila[col_modelo])
        if not marca_actual or not modelo_actual:
            continue

        desde, hasta = parsear_anios(fila[col_anios] if col_anios < len(fila) else "")

        # Una celda puede traer VARIOS códigos: "BKR6EKPA / PMR7A" son dos bujías distintas
        # (la común y la de platino) que le van al mismo auto. Guardadas como un solo texto no
        # coinciden con nada; separadas, cada una queda buscable por su cuenta.
        codigos = dividir_codigos(codigo) or [codigo]
        for cod in codigos:
            # En algunas páginas las columnas se corren y lo que cae en el lugar del código es
            # el tipo de combustible ('G' de gasolina, 'B' de flex). Un código de repuesto de
            # una sola letra no existe: si entra, después aparece como si NGK recomendara la
            # pieza «B» para un Ford, que es una recomendación inventada.
            limpio_cod = sanitizar(cod)
            if len(limpio_cod) <= 2 or not any(ch.isdigit() for ch in limpio_cod):
                continue
            apps.append({
                "marca_auto": marca_actual,
                "modelo_auto": modelo_actual.upper(),
                "motor": _limpiar(fila[col_motor]) if col_motor < len(fila) else "",
                "combustible": _limpiar(fila[col_comb]) if col_comb < len(fila) else "",
                "anio_desde": desde,
                "anio_hasta": hasta,
                "codigo": cod,
            })
    return apps


def tablas_de_archivo(archivo):
    """Devuelve las tablas de un PDF o de una planilla, para leer catálogos de aplicaciones."""
    nombre = archivo if isinstance(archivo, str) else getattr(archivo, "name", "")
    if nombre.lower().endswith(".pdf"):
        import pdfplumber
        if not isinstance(archivo, str):
            archivo.seek(0)
        tablas = []
        with pdfplumber.open(archivo) as pdf:
            for pagina in pdf.pages:
                tablas.extend(t for t in pagina.extract_tables() if t)
        return tablas
    return [leer_excel(archivo)]


def guardar_aplicaciones(apps, marca_repuesto, origen="", tipo_pieza=""):
    """Guarda las aplicaciones leídas de un catálogo. Devuelve cuántas se cargaron."""
    if not apps:
        return 0
    with db_lock:
        c.executemany("""INSERT OR IGNORE INTO aplicaciones
                         (marca_auto, modelo_auto, motor, combustible, anio_desde, anio_hasta,
                          codigo, codigo_clean, marca_repuesto, tipo_pieza, origen)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                      [(a["marca_auto"], a["modelo_auto"], a["motor"], a["combustible"],
                        a["anio_desde"], a["anio_hasta"], a["codigo"], sanitizar(a["codigo"]),
                        marca_repuesto.strip().upper(), (tipo_pieza or "").strip(), origen)
                       for a in apps])
        conn.commit()
    return len(apps)


def derivar_equivalencias_de_aplicaciones(limite=500, minimo_autos=2):
    """Deduce equivalencias cruzando los catálogos de aplicaciones de distintos fabricantes.

    El razonamiento: si NGK dice que su bujía U2003 va en un Palio 1.0 2003-2006, y Bosch dice
    que la suya va en EXACTAMENTE el mismo auto, las dos hacen el mismo trabajo. Son
    equivalentes, y ninguna lista de proveedor te lo iba a decir — es información que sale de
    cruzar dos catálogos que ya tenés.

    Tres condiciones, y las tres importan:
      · MISMO tipo de pieza. Sin esto se cruzaría una bujía con un filtro por ir al mismo auto,
        que es exactamente el error que venimos limpiando.
      · Fabricantes DISTINTOS. Dos códigos de la misma marca para el mismo auto suelen ser
        variantes (la común y la de platino), no equivalentes entre sí.
      · Coincidir en varios autos, no en uno. Una coincidencia suelta puede ser casualidad;
        que dos códigos vayan juntos en varios modelos ya es un patrón.

    No las carga: las deja como pendientes para que pasen por la misma revisión que el resto."""
    c.execute("""SELECT a.codigo_clean AS cod_a, a.marca_repuesto AS marca_a,
                        b.codigo_clean AS cod_b, b.marca_repuesto AS marca_b,
                        COUNT(DISTINCT a.marca_auto || '|' || a.modelo_auto || '|' || a.motor) AS autos
                 FROM aplicaciones a
                 JOIN aplicaciones b
                   ON a.marca_auto = b.marca_auto
                  AND a.modelo_auto = b.modelo_auto
                  AND a.motor = b.motor
                  AND COALESCE(a.tipo_pieza,'') = COALESCE(b.tipo_pieza,'')
                  AND a.marca_repuesto <> b.marca_repuesto
                  AND a.codigo_clean < b.codigo_clean
                 WHERE COALESCE(a.tipo_pieza,'') <> ''
                 GROUP BY a.codigo_clean, b.codigo_clean
                 HAVING autos >= ?
                 ORDER BY autos DESC LIMIT ?""", (minimo_autos, limite))
    candidatos = filas_a_listas(c)

    # Solo sirven los que además existen en el catálogo propio: proponer una equivalencia entre
    # dos códigos que no tenés cargados no le sirve a nadie.
    salida = []
    for x in candidatos:
        c.execute("""SELECT p.id, p.codigo_raw, m.nombre AS marca FROM productos p
                     JOIN marcas m ON m.id = p.marca_id WHERE p.codigo_clean = ? LIMIT 1""",
                  (x["cod_a"],))
        pa = c.fetchone()
        c.execute("""SELECT p.id, p.codigo_raw, m.nombre AS marca FROM productos p
                     JOIN marcas m ON m.id = p.marca_id WHERE p.codigo_clean = ? LIMIT 1""",
                  (x["cod_b"],))
        pb = c.fetchone()
        if not pa or not pb or pa["id"] == pb["id"]:
            continue
        salida.append({
            "Código A": pa["codigo_raw"], "Marca A": pa["marca"],
            "Código B": pb["codigo_raw"], "Marca B": pb["marca"],
            "Coinciden en": f"{x['autos']} auto(s)",
            "Según": f"{x['marca_a']} y {x['marca_b']}",
            "_a": pa["id"], "_b": pb["id"],
        })
    return salida


def guardar_equivalencias_derivadas(pares, lote):
    """Deja las equivalencias deducidas como PENDIENTES, no cargadas.

    A propósito: por más buena que sea la deducción, sigue siendo una deducción. Pasa por la
    misma revisión que todo lo demás, y ahí el sistema de confianza la evalúa como a cualquier
    otra."""
    if not pares:
        return 0
    rechazados = pares_rechazados()
    nuevos = [(min(a, b), max(a, b)) for a, b in pares
              if (min(a, b), max(a, b)) not in rechazados]
    if not nuevos:
        return 0
    with db_lock:
        c.executemany("""INSERT OR IGNORE INTO equivalencias_pendientes
                         (producto_a_id, producto_b_id, origen, lote)
                         VALUES (?, ?, 'catalogos_fabricante', ?)""",
                      [(a, b, lote) for a, b in nuevos])
        conn.commit()
    return len(nuevos)


def buscar_aplicaciones(marca_auto, modelo="", anio=None, limite=200):
    """Los códigos que un catálogo de aplicaciones dice que le van a este auto.

    A diferencia de buscar por la descripción del proveedor, acá el dato lo puso el fabricante
    del repuesto: si NGK dice que a un Fiat Argo 1.3 le va la U5443, le va."""
    condiciones = ["UPPER(marca_auto) LIKE ?"]
    params = [f"%{(marca_auto or '').strip().upper()}%"]
    if modelo:
        condiciones.append("UPPER(modelo_auto) LIKE ?")
        params.append(f"%{modelo.strip().upper()}%")
    if anio:
        # Se descarta solo lo que declara OTRO rango; lo que no aclara años se deja pasar
        condiciones.append("(anio_desde IS NULL OR anio_desde <= ?)")
        params.append(int(anio))
        condiciones.append("(anio_hasta IS NULL OR anio_hasta >= ?)")
        params.append(int(anio))
    c.execute(f"""SELECT a.codigo AS "Código", a.marca_repuesto AS "Marca del repuesto",
                         a.marca_auto AS "Marca auto", a.modelo_auto AS "Modelo",
                         a.motor AS "Motor",
                         COALESCE(a.anio_desde, '') || '-' || COALESCE(a.anio_hasta, '') AS "Años",
                         (SELECT COUNT(*) FROM productos p
                           WHERE p.codigo_clean = a.codigo_clean) AS "En tu catálogo"
                  FROM aplicaciones a
                  WHERE {" AND ".join(condiciones)}
                  ORDER BY a.marca_repuesto, a.modelo_auto LIMIT ?""", params + [limite])
    return filas_a_listas(c)


def repuestos_de_este_auto(marca_auto, modelo="", anio=None, motor="", vin="", limite=500):
    """Los códigos que le corresponden a ESTE auto. Es lo que hace falta en el mostrador: el VIN
    dice qué auto es, pero lo que se vende son códigos.

    IMPORTANTE sobre de dónde sale cada cosa. No existe una tabla pública que diga qué repuesto
    lleva cada auto — eso es una base de aplicaciones licenciada (TecDoc y similares) y se paga.
    Lo que sí hay es información propia del negocio, y se usa en este orden de confianza:

      1. Lo que YA se le puso a ESTE auto (por VIN o patente). Certeza total: alguien lo instaló.
      2. Lo que se le puso a OTROS autos del MISMO modelo. Evidencia real del mostrador.
      3. Lo que dicen las descripciones del catálogo del proveedor ('... FORD FIESTA 1.6 2010/15').
         Es lo más amplio y lo menos seguro: depende de cómo escriba cada proveedor.

    Devuelve un dict con las tres listas por separado, a propósito: mezclarlas escondería que
    una es un hecho y la otra una coincidencia de texto."""
    marca_auto = (marca_auto or "").strip().upper()
    modelo = (modelo or "").strip().upper()
    motor = (motor or "").strip().upper()
    vin = re.sub(r'\s', '', (vin or "").strip().upper())
    resultado = {"de_este_auto": [], "de_otros_iguales": [], "del_catalogo": [],
                 "del_fabricante": [], "modelo_usado": modelo, "sin_datos": True}

    # Fuente nueva y la más confiable de las que no son propias: el catálogo de aplicaciones
    # del fabricante del repuesto. No es una coincidencia de texto, es el fabricante diciendo
    # a qué auto le va su pieza.
    try:
        resultado["del_fabricante"] = buscar_aplicaciones(marca_auto, modelo, anio)
    except sqlite3.OperationalError:
        resultado["del_fabricante"] = []

    # --- 1. Lo que ya se le puso a este auto ---
    vehiculo_id = None
    if len(vin) == 17:
        c.execute("SELECT id FROM vehiculos WHERE vin = ?", (vin,))
        f = c.fetchone()
        vehiculo_id = f["id"] if f else None
    if vehiculo_id:
        c.execute("""SELECT h.descripcion_pieza AS "Pieza", h.codigo_pieza AS "Código",
                            h.marca_pieza AS "Marca", h.km_instalacion AS "Km",
                            substr(h.fecha_instalacion, 1, 10) AS "Fecha", h.producto_id AS "_pid"
                     FROM historial_piezas h WHERE h.vehiculo_id = ?
                     ORDER BY h.fecha_instalacion DESC""", (vehiculo_id,))
        resultado["de_este_auto"] = filas_a_listas(c)

    # --- 2. Lo que se le puso a otros autos del mismo modelo ---
    if modelo:
        c.execute("""SELECT h.descripcion_pieza AS "Pieza", h.codigo_pieza AS "Código",
                            h.marca_pieza AS "Marca", COUNT(*) AS "Veces",
                            COUNT(DISTINCT v.id) AS "Autos"
                     FROM historial_piezas h JOIN vehiculos v ON v.id = h.vehiculo_id
                     WHERE UPPER(COALESCE(v.modelo_auto,'')) LIKE ?
                       AND (? = '' OR UPPER(COALESCE(v.marca_auto,'')) LIKE ?)
                       AND (v.id IS NOT ?)
                       AND h.codigo_pieza IS NOT NULL AND h.codigo_pieza <> ''
                     GROUP BY UPPER(h.codigo_pieza)
                     ORDER BY "Autos" DESC, "Veces" DESC LIMIT 100""",
                  (f"%{modelo}%", marca_auto, f"%{marca_auto}%", vehiculo_id))
        resultado["de_otros_iguales"] = filas_a_listas(c)

    # --- 3. El catálogo, por lo que dicen las descripciones ---
    if marca_auto:
        condiciones = ["p.descripcion IS NOT NULL", "UPPER(p.descripcion) LIKE ?"]
        params = [f"%{marca_auto}%"]
        if modelo:
            condiciones.append("UPPER(p.descripcion) LIKE ?")
            params.append(f"%{modelo}%")
        c.execute(f"""SELECT p.id AS "ID", p.codigo_raw AS "Código", p.descripcion AS "Descripcion",
                             m.nombre AS "Marca", p.precio AS "Precio", p.stock AS "Stock"
                      FROM productos p JOIN marcas m ON m.id = p.marca_id
                      WHERE {" AND ".join(condiciones)}
                      LIMIT 4000""", params)
        candidatos = filas_a_listas(c)

        # La cilindrada es lo que más afina ("1.6", "2.0"): si el motor la trae, se usa para
        # filtrar, pero solo descartando lo que declara OTRA cilindrada — lo que no dice nada
        # se deja pasar, porque la mayoría de las listas no la aclaran.
        cilindrada = None
        if motor:
            m_cil = re.search(r'\d[.,]\d', motor)
            cilindrada = m_cil.group(0).replace(",", ".") if m_cil else None

        filtrados = []
        for f in candidatos:
            desc = (f["Descripcion"] or "").upper()
            if anio:
                sirve = sirve_para_anio(f["Descripcion"], int(anio))
                if sirve is False:
                    continue
                f["Año"] = "✅ coincide" if sirve else "— no aclara"
            if cilindrada:
                otras = set(re.findall(r'\d[.,]\d', desc))
                if otras and cilindrada not in {o.replace(",", ".") for o in otras}:
                    continue
                f["Motor"] = "✅ coincide" if cilindrada in desc else "— no aclara"
            f["Categoría"] = clasificar_repuesto(f["Descripcion"])
            filtrados.append(f)

        filtrados.sort(key=lambda x: (x.get("Año") != "✅ coincide",
                                       x.get("Motor") != "✅ coincide",
                                       x["Categoría"]))
        resultado["del_catalogo"] = filtrados[:limite]
        resultado["total_catalogo"] = len(filtrados)

    resultado["sin_datos"] = not (resultado["de_este_auto"] or resultado["de_otros_iguales"]
                                   or resultado["del_catalogo"] or resultado["del_fabricante"])
    return resultado


def panel_vin(clave="vin", mostrar_ensenar=True):
    """La ÚNICA pantalla de VIN de la app. Antes había dos (una en el Buscador y otra en Modo
    Mecánico) que hacían casi lo mismo y ninguna terminaba en lo que hace falta: los códigos.
    Esta identifica el auto y va derecho a qué repuestos lleva."""
    explicar(
        "Pegá el número de chasis y te dice qué repuestos lleva ese auto.",
        "Del VIN salen con certeza el fabricante, el país y el año. El modelo y el motor no "
        "están normalizados: los aprende la app de tus propias fichas, o se los enseñás una "
        "vez."
    )
    vin_txt = st.text_input("VIN / número de chasis (17 caracteres):",
                             placeholder="Ej: 9BWZZZ377VT004251", key=f"{clave}_texto").strip().upper()
    if not vin_txt:
        return None

    d = decodificar_vin(vin_txt)
    if not d["valido"]:
        st.error(d["error"])
        return None

    # --- Qué auto es ---
    fab = d.get("fabricante") or "(fabricante no cargado)"
    if d.get("fabricante_por_prefijo"):
        fab += " (por familia de WMI)"
    linea = f"**{fab}** · {d['pais']}"
    if d.get("anio_estimado"):
        linea += f" · año {d['anio_estimado']}" if d.get("anio_preciso") else f" · año ~{d['anio_estimado']}"
    if d.get("modelo"):
        linea += f" · **{d['modelo']}**"
    if d.get("motor"):
        linea += f" · motor {d['motor']}"
    st.success(linea)

    if d.get("digito_verificador") is False:
        st.error("❌ El dígito verificador no da: revisá si hay algún carácter mal tipeado. "
                  "Buscar repuestos con un VIN mal copiado es buscar los del auto equivocado.")

    ficha = d.get("vehiculo")
    if ficha:
        st.info(f"🎯 Este chasis ya está en tus fichas: patente **{ficha.get('patente')}**"
                 + (f" — cliente {ficha['cliente_nombre']}" if ficha.get("cliente_nombre") else ""))

    anio_usar = d.get("anio_estimado")
    if d.get("anio_alternativo"):
        # El código de año se repite cada 30. Elegir mal acá filtra los repuestos correctos.
        anio_usar = st.radio(
            "El VIN no permite saber cuál de los dos años es — elegí (miralo en la cédula):",
            sorted({d["anio_estimado"], d["anio_alternativo"]}), horizontal=True,
            key=f"{clave}_anio"
        )

    # --- Lo que importa: los códigos ---
    marca_para_buscar = None
    if d.get("fabricante"):
        marca_para_buscar = next(
            (mv for mv in MARCAS_VEHICULO if mv in d["fabricante"].upper()), None
        )

    if not marca_para_buscar:
        st.warning(
            "No puedo buscar repuestos porque no sé de qué marca es este WMI. "
            "Cargalo abajo en «Enseñar» y la próxima vez sale solo."
        )
    else:
        # OJO con value= junto a key=: Streamlit ignora el value y usa lo que haya en la sesión,
        # que arranca vacío. Por eso el modelo que el VIN reconocía ("Fiesta") aparecía en el
        # cartel verde pero el campo quedaba en blanco, y la búsqueda terminaba trayendo TODOS
        # los repuestos de la marca en vez de los del modelo.
        # Se siembra en session_state, y solo cuando cambia el VIN, para no pisar una corrección.
        clave_modelo = f"{clave}_modelo_manual"
        if st.session_state.get(f"{clave}_vin_previo") != vin_txt:
            st.session_state[clave_modelo] = d.get("modelo") or ""
            st.session_state[f"{clave}_vin_previo"] = vin_txt
        st.session_state.setdefault(clave_modelo, d.get("modelo") or "")

        modelo_manual = st.text_input(
            "Modelo (corregilo si hace falta):", key=clave_modelo,
            help="Se completa solo con lo que reconoce del VIN. Corregilo si no acertó; si lo "
                 "dejás vacío, busca todos los repuestos de la marca."
        ).strip()

        motor_reconocido = d.get("motor") or ""
        if motor_reconocido:
            st.caption(f"⚙️ Filtrando además por motor **{motor_reconocido}** "
                        "(se usa la cilindrada para afinar el catálogo).")

        r = repuestos_de_este_auto(
            marca_para_buscar, modelo_manual, anio_usar, d.get("motor") or "", vin_txt
        )
        _mostrar_repuestos_del_auto(r, marca_para_buscar, modelo_manual, anio_usar, clave)

    if mostrar_ensenar:
        with st.expander("✏️ Enseñarle a la app este auto (para que la próxima salga solo)"):
            _formulario_ensenar_vin(d, clave)
    return d


def _mostrar_repuestos_del_auto(r, marca, modelo, anio, clave):
    """Muestra los códigos separados por fuente. Van separados a propósito: uno es un hecho
    (se lo pusiste a este auto) y el otro es una coincidencia de texto en una descripción.
    Mezclarlos haría parecer que todos valen lo mismo, y no es así."""
    if r["de_este_auto"]:
        st.markdown("#### 🎯 Ya se le puso a ESTE auto")
        st.caption("Certeza total: alguien lo instaló y quedó registrado en la ficha.")
        st.dataframe([{k: v for k, v in f.items() if not k.startswith("_")}
                       for f in r["de_este_auto"]], use_container_width=True, hide_index=True)

    if r["de_otros_iguales"]:
        st.markdown("#### 🔁 Se le puso a otros autos del mismo modelo")
        st.caption(
            "Evidencia real de tu mostrador: estos códigos se instalaron en autos iguales. "
            "«Autos» es en cuántos distintos — mientras más, más confiable."
        )
        st.dataframe(r["de_otros_iguales"], use_container_width=True, hide_index=True)

    if r.get("del_fabricante"):
        st.markdown(f"#### 🏭 Según el catálogo del fabricante ({len(r['del_fabricante'])})")
        explicar(
            "Esto no es una coincidencia de texto: es el fabricante del repuesto diciendo a qué "
            "auto le va su pieza.",
            "Es lo más confiable después de lo que ya le pusiste vos. La columna «En tu catálogo» "
            "dice si ese código está cargado en tus listas."
        )
        st.dataframe(r["del_fabricante"], use_container_width=True, hide_index=True)

    if r["del_catalogo"]:
        total = r.get("total_catalogo", len(r["del_catalogo"]))
        st.markdown(f"#### 📚 Del catálogo, según las descripciones ({total} código(s))")
        st.caption(
            "Sale de que la descripción del proveedor menciona esta marca y modelo. Es lo más "
            "amplio y lo menos seguro: depende de cómo escriba cada proveedor. "
            + (f"Se filtró por año {anio} descartando lo que declara otro rango. " if anio else "")
            + "Confirmá por código antes de vender."
        )
        # Antes acá el filtro listaba el texto crudo previo a la marca en cada descripción, así
        # que con varios proveedores salían cientos de opciones casi iguales ("JUNTA TAPA DE
        # CILINDROS", "JUNTA DE TAPA CIL.", "JUEGO JUNTA TAPA"...) y no servía para encontrar
        # nada. Ahora son ~20 familias fijas, con la cantidad al lado.
        conteo = {}
        for f in r["del_catalogo"]:
            conteo[f["Categoría"]] = conteo.get(f["Categoría"], 0) + 1
        orden_familias = sorted(conteo, key=lambda k: (-conteo[k], k))
        etiquetas = {f"{k} ({conteo[k]})": k for k in orden_familias}

        cf1, cf2 = st.columns([2, 1])
        with cf1:
            elegidas_lbl = st.multiselect("Tipo de pieza:", list(etiquetas.keys()),
                                           key=f"{clave}_cats",
                                           placeholder="Todas — o elegí una o varias")
        with cf2:
            texto_filtro = st.text_input("Buscar en estos resultados:", key=f"{clave}_txt",
                                          placeholder="Ej: delantero, 1.6, kit").strip()

        elegidas = {etiquetas[e] for e in elegidas_lbl}
        filas = [f for f in r["del_catalogo"] if not elegidas or f["Categoría"] in elegidas]
        if texto_filtro:
            # Todas las palabras tienen que estar, sin importar el orden ni los acentos: así
            # "kit delantero" encuentra "Kit de rodamiento delantero" igual.
            palabras = [normalizar_texto(x) for x in texto_filtro.split() if x.strip()]
            filas = [f for f in filas
                     if all(pal in normalizar_texto(f"{f['Código']} {f['Descripcion']}")
                            for pal in palabras)]
        st.caption(f"Mostrando {len(filas)} de {len(r['del_catalogo'])}.")
        st.dataframe(quitar_id(filas), use_container_width=True, hide_index=True)
        st.download_button(
            "⬇️ Bajar estos códigos en Excel",
            data=to_excel_bytes(quitar_id(filas)),
            file_name=f"repuestos_{marca}_{modelo or 'todos'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{clave}_dl"
        )
        if total > len(r["del_catalogo"]):
            st.caption(f"(mostrando {len(r['del_catalogo'])} de {total} — afiná el modelo para ver menos)")

    if r["sin_datos"]:
        st.warning(
            f"No encontré repuestos para **{marca} {modelo}**. Puede ser que en tu catálogo las "
            "descripciones lo escriban distinto (probá con otra forma del modelo, o dejá el modelo "
            "vacío para ver todo lo de la marca), o que todavía no tengas cargada esa lista."
        )


def _formulario_ensenar_vin(d, clave):
    """Cargar el modelo y el motor de este patrón, una sola vez."""
    st.caption(
        f"Patrón `{d['wmi']}`-`{d['vds']}` · 8ª posición `{d['codigo_motor']}`. "
        "El modelo se guarda por el patrón; el motor por la 8ª posición, que vale para toda la marca."
    )
    if not d.get("fabricante"):
        cw1, cw2 = st.columns(2)
        nuevo_fab = cw1.text_input("Fabricante de este WMI:", key=f"{clave}_fab")
        nuevo_pais = cw2.text_input("País:", value=d["pais"], key=f"{clave}_pais")
        if st.button("💾 Guardar fabricante", key=f"{clave}_guardar_fab"):
            if nuevo_fab.strip():
                agregar_fabricante_vin(d["wmi"], nuevo_fab, nuevo_pais)
                avisar("success", f"WMI {d['wmi']} guardado.")
                st.rerun()
            else:
                st.warning("Escribí el fabricante.")

    with st.form(f"{clave}_form_ensenar", clear_on_submit=True):
        ce1, ce2 = st.columns(2)
        mod_in = ce1.text_input("Modelo", value=d.get("modelo") or "", placeholder="Ej: FIESTA")
        mot_in = ce2.text_input("Motor", value=d.get("motor") or "", placeholder="Ej: 1.6 16V nafta")
        nota_in = st.text_input("Nota (opcional)", placeholder="Ej: 5 puertas")
        alcance = st.radio("El MODELO, ¿para qué VIN vale?",
                           ["Este patrón exacto (5 caracteres)", "Toda la familia (3 caracteres)"],
                           horizontal=True)
        if st.form_submit_button("💾 Guardar", type="primary"):
            hechos = []
            if mod_in.strip():
                largo = 5 if alcance.startswith("Este") else 3
                enseniar_modelo_vin(d["wmi"], d["vds"][:largo], mod_in, nota_in)
                hechos.append(f"modelo {mod_in.strip()}")
            if mot_in.strip():
                enseniar_motor_vin(d["wmi"], d["codigo_motor"], mot_in, nota_in)
                hechos.append(f"motor {mot_in.strip()}")
            if hechos:
                avisar("success", "Guardado: " + " y ".join(hechos) + ".")
                st.rerun()
            else:
                st.warning("Escribí al menos el modelo o el motor.")


def esquemas_de_vehiculo(marca_vehiculo):
    c.execute("""SELECT id, titulo, marca_auto, modelo_auto, sistema FROM esquemas
                 WHERE UPPER(marca_auto) LIKE ? ORDER BY titulo""", (f"%{marca_vehiculo}%",))
    return [dict(r) for r in c.fetchall()]


# Marcas que se usan para despegar descripciones. Se dejan solo las de 5 letras o más y se
# excluyen las que además son palabras comunes del rubro: separar por "MAN" partiría MANGUERA
# en "MAN GUERA", y por "RAM" partiría RAMAL. Con las largas el riesgo desaparece y son
# justamente las que aparecen pegadas en las listas (VOLKSWAGEN, CHEVROLET, MITSUBISHI...).
_MARCAS_RIESGOSAS = {"BETA", "CASE", "HINO", "SEAT", "LADA", "TATA", "MINI", "HERO", "TVS"}
MARCAS_PARA_DESPEGAR = [m for m in MARCAS_VEHICULO
                        if len(m) >= 4 and m not in _MARCAS_RIESGOSAS and " " not in m]

# Las expresiones se arman UNA vez, al arrancar. Antes se compilaban las 164 de nuevo en cada
# fila: 332 µs por descripción, o sea 3,3 s en una lista de 10.000 filas — diez veces más que
# toda la importación junta.
_RE_PEGADO_MAYUS = re.compile(r'(?<=[a-záéíóúñ])(?=[A-ZÁÉÍÓÚÑ])')
# Se rodea la marca de espacios y después se colapsan los sobrantes. Es más simple y más
# robusto que exigir letra pegada de un lado o del otro: con "DespieceCHEVROLETCAPUCHON" hay
# que separar por IZQUIERDA y por DERECHA a la vez, y una condición sola nunca cubre las dos.
# El orden es de marca más larga a más corta, así "VOLKSWAGEN" gana antes de que pruebe "VW".
_RE_MARCAS_PEGADAS = re.compile(
    "(" + "|".join(re.escape(m) for m in sorted(MARCAS_PARA_DESPEGAR, key=len, reverse=True)) + ")"
) if MARCAS_PARA_DESPEGAR else None
_RE_ESPACIOS = re.compile(r'\s{2,}')


def separar_texto_pegado(texto):
    """Algunas listas de proveedor exportan varias columnas pegadas sin espacio en el medio:
    'Junta Tapa de CilindrosFORDTAUNUS COUPE' o 'PASTILLAS FRENOVOLKSWAGENGOL'.
    Esto las vuelve legibles separando en dos puntos:
      1) donde una minúscula toca una MAYÚSCULA (ahí se pegaron dos campos), y
      2) donde una marca de vehículo quedó pegada a la descripción.

    El punto 2 antes solo funcionaba si la descripción tenía minúsculas: pedía que el carácter
    anterior a la marca NO fuera mayúscula, así que en una lista escrita toda en mayúsculas
    —que son la mayoría— no separaba nada."""
    if not texto:
        return texto
    t = str(texto).strip()
    t = _RE_PEGADO_MAYUS.sub(' ', t)
    if _RE_MARCAS_PEGADAS is not None:
        t = _RE_MARCAS_PEGADAS.sub(r' \1 ', t)
    return _RE_ESPACIOS.sub(' ', t).strip()


def contar_descripciones_pegadas(limite_muestra=3000):
    c.execute("SELECT id, descripcion FROM productos WHERE descripcion IS NOT NULL LIMIT ?",
               (limite_muestra,))
    return sum(1 for r in c.fetchall() if separar_texto_pegado(r["descripcion"]) != r["descripcion"])


def reparar_descripciones_pegadas():
    c.execute("SELECT id, descripcion FROM productos WHERE descripcion IS NOT NULL")
    cambios = [(separar_texto_pegado(r["descripcion"]), r["id"]) for r in c.fetchall()
               if separar_texto_pegado(r["descripcion"]) != r["descripcion"]]
    with db_lock:
        c.executemany("UPDATE productos SET descripcion = ? WHERE id = ?", cambios)
        conn.commit()
    return len(cambios)


def buscar_por_texto(texto):
    """Busca por descripción de forma flexible: cada palabra tiene que aparecer en algún lado
    (descripción o código), sin importar el orden ni las tildes. Así 'ruleman delantero gol'
    encuentra 'Gol 1.6 - Ruleman de rueda delantero', y 'rótula' encuentra 'ROTULA' aunque el
    catálogo la tenga cargada sin tilde (frecuente en listas de proveedores)."""
    palabras = [normalizar_texto(p.strip()) for p in texto.upper().split() if p.strip()]
    if not palabras:
        return []
    # Compara contra la descripción/código sin tildes de ningún lado, para que no importe si
    # la búsqueda o el dato cargado tienen o no acentos.
    desc_sin_acentos = _sql_sin_acentos("p.descripcion")
    codigo_sin_acentos = _sql_sin_acentos("p.codigo_raw")
    # En vez de exigir que estén TODAS las palabras, se cuenta cuántas coinciden y se ordena
    # por eso. Así "junta tapa cilindro ford taunus" igual encuentra la que dice
    # "Junta Tapa de Cilindros FORD TAUNUS COUPE" aunque no diga exactamente lo mismo, y las
    # que más se parecen quedan arriba. Antes, si fallaba una sola palabra, no aparecía nada.
    puntajes = []
    params = []
    for palabra in palabras:
        # También se compara contra el código sin guiones ni espacios: si alguien escribe
        # "TC421" o "tc-421", tiene que encontrar igual el producto cargado como "TC-421-15".
        puntajes.append(f"(CASE WHEN {desc_sin_acentos} LIKE ? OR {codigo_sin_acentos} LIKE ? "
                         f"OR p.codigo_clean LIKE ? THEN 1 ELSE 0 END)")
        like = f"%{palabra}%"
        params.extend([like, like, f"%{sanitizar(palabra)}%"])
    suma = " + ".join(puntajes)

    # Con una o dos palabras se piden todas (si no, aparece cualquier cosa). Con tres o más
    # alcanza con que coincida la mayoría: es lo que permite "interpretar" y no fallar por una.
    minimo = len(palabras) if len(palabras) <= 2 else max(2, (len(palabras) * 2) // 3)

    query = f'''
    SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
           m.nombre AS "Marca", m.tipo AS "Tipo", p.precio AS "Precio", p.stock AS "Stock",
           p.favorito AS "Favorito", ({suma}) AS _coincidencias
    FROM productos p JOIN marcas m ON m.id = p.marca_id
    WHERE ({suma}) >= ?
    ORDER BY _coincidencias DESC, LENGTH(p.descripcion), m.nombre LIMIT 200;
    '''
    with db_lock:
        c.execute(query, params + params + [minimo])
        filas = filas_a_listas(c)

    # Filtro por RUBRO. Contar palabras coincidentes no alcanza: buscando «bujía golf 1.4 tsi»
    # aparecían juntas de tapa y juegos de motor, porque coinciden en «golf», «1.4» y «tsi» —
    # o sea, en el AUTO, no en la pieza. Y el auto es lo de menos: nadie que pide una bujía se
    # lleva una junta porque va al mismo Golf.
    #
    # Si en el pedido se reconoce un rubro, se dejan solo los resultados de ESE rubro. Los que
    # no se pudieron clasificar se conservan: descartar lo que no se entiende es peor que
    # mostrarlo, porque las descripciones de proveedor son un desastre y muchas piezas legítimas
    # no caen en ninguna familia.
    familia_pedida = clasificar_repuesto(texto)
    if familia_pedida != "Sin clasificar":
        del_rubro, sin_clasificar, de_otro_rubro = [], [], []
        for f in filas:
            fam = clasificar_repuesto(f.get("Descripcion") or "")
            if fam == familia_pedida:
                del_rubro.append(f)
            elif fam == "Sin clasificar":
                sin_clasificar.append(f)
            else:
                de_otro_rubro.append(f)
        # Solo se descarta lo de otro rubro si quedó algo del rubro pedido; si no, es mejor
        # mostrar todo que dejar la pantalla vacía.
        if del_rubro:
            filas = del_rubro + sin_clasificar

    for f in filas:
        f.pop("_coincidencias", None)
    return filas


# ============================================================
# COMBOS DE REPUESTOS RELACIONADOS (ej: correa de distribución -> kit + tensor + bomba de agua)
# ============================================================
def normalizar_texto(texto):
    """Mayúsculas y sin acentos, para poder comparar 'distribución' con 'distribucion'."""
    texto = texto or ""
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto.upper().strip()


def buscar_combos_para_descripcion(descripcion):
    """Devuelve {disparador: [items]} para los disparadores que aparecen dentro de la descripción dada."""
    desc_norm = normalizar_texto(descripcion)
    c.execute("SELECT disparador, item FROM combos_sugeridos ORDER BY disparador, item")
    resultado = {}
    for row in c.fetchall():
        if normalizar_texto(row["disparador"]) in desc_norm:
            resultado.setdefault(row["disparador"], []).append(row["item"])
    return resultado


def listar_combos():
    c.execute("SELECT DISTINCT disparador FROM combos_sugeridos ORDER BY disparador")
    disparadores = [r["disparador"] for r in c.fetchall()]
    resultado = []
    for d in disparadores:
        c.execute("SELECT item FROM combos_sugeridos WHERE disparador = ? ORDER BY item", (d,))
        resultado.append({"disparador": d, "items": [r["item"] for r in c.fetchall()]})
    return resultado


def guardar_combo(disparador, items_lista):
    disparador = disparador.strip().lower()
    with db_lock:
        c.execute("DELETE FROM combos_sugeridos WHERE disparador = ?", (disparador,))
        c.executemany(
            "INSERT INTO combos_sugeridos (disparador, item) VALUES (?, ?)",
            [(disparador, item.strip()) for item in items_lista if item.strip()]
        )
        conn.commit()


def eliminar_combo(disparador):
    disparador = disparador.strip().lower()
    c.execute("SELECT item FROM combos_sugeridos WHERE disparador = ?", (disparador,))
    items = [r["item"] for r in c.fetchall()]
    if items:
        mover_a_papelera("combo", {"disparador": disparador, "items": items})
    with db_lock:
        c.execute("DELETE FROM combos_sugeridos WHERE disparador = ?", (disparador,))
        conn.commit()


def identificar_pieza_por_foto(imagen_bytes):
    """Le manda una foto a Gemini y le pide que identifique la pieza, extrayendo el código
    de forma estructurada (no solo texto libre) para poder buscarlo directo en el catálogo."""
    from google import genai
    from google.genai import types

    api_key = st.secrets.get("gemini_api_key") if hasattr(st, "secrets") else None
    if not api_key:
        return None, "No configuraste 'gemini_api_key' en Streamlit Cloud (Settings → Secrets)."

    try:
        client = genai.Client(api_key=api_key)
        prompt = (
            "Esta es una foto de un repuesto de auto tomada en un taller o local de repuestos. Tu "
            "tarea principal es ENCONTRAR EL CÓDIGO — mirá con mucha atención toda la superficie de "
            "la pieza: suelen estar grabados en bajorrelieve sobre el metal (a veces se ven mejor con "
            "el contraste de la luz, poco legibles a simple vista), impresos en una etiqueta pegada, "
            "moldeados en el plástico/goma, o troquelados en el borde. Es una combinación de letras y "
            "números, a veces con guiones, barras o puntos. Revisá TODOS los lados de la pieza que se "
            "vean en la foto antes de rendirte. Devolvé ÚNICAMENTE un JSON válido (sin texto extra, "
            'sin markdown), con esta forma exacta: {"codigo": "...", "marca_visible": "...", '
            '"tipo_pieza": "...", "confianza": "alta/media/baja"}. Si después de mirar con atención en '
            'serio no hay ningún código legible, dejá "codigo" como null — no inventes ni completes un '
            "código que no se vea con claridad."
        )
        response = client.models.generate_content(
            model="gemini-flash-latest",
            contents=[
                prompt,
                types.Part.from_bytes(data=imagen_bytes, mime_type="image/jpeg"),
            ],
        )
        texto = response.text.strip()
        if texto.startswith("```"):
            texto = texto.split("```")[1]
            texto = texto[4:] if texto.lower().startswith("json") else texto
        datos = json.loads(texto)
        registrar_uso_ia("Identificar pieza por foto", True)
        return datos, None
    except json.JSONDecodeError:
        registrar_uso_ia("Identificar pieza por foto", False)
        return None, "No pude interpretar la respuesta — probá con una foto más clara y de más cerca."
    except Exception as e:
        registrar_uso_ia("Identificar pieza por foto", False)
        return None, traducir_error_gemini(e)


def extraer_datos_cedula(imagen_bytes):
    """Lee una foto de cédula verde/azul o título del auto y extrae patente, marca, modelo, año
    y motorización con Gemini. SIEMPRE hay que revisar antes de guardar — el OCR puede confundir
    caracteres parecidos (0/O, 1/I) y en la patente o el VIN eso es grave."""
    from google import genai
    from google.genai import types
    import json

    api_key = st.secrets.get("gemini_api_key") if hasattr(st, "secrets") else None
    if not api_key:
        return None, "No configuraste 'gemini_api_key' en Streamlit Cloud (Settings → Secrets)."

    try:
        client = genai.Client(api_key=api_key)
        prompt = (
            "Esta es una foto de una cédula verde/azul o título de un vehículo argentino. Extraé "
            "ÚNICAMENTE un JSON válido (sin texto extra, sin markdown), con esta forma exacta: "
            '{"patente": "...", "marca": "...", "modelo": "...", "anio": "...", "motorizacion": "..."}. '
            "Si no podés leer algún campo con claridad, dejalo como null en vez de adivinar. No inventes datos."
        )
        response = client.models.generate_content(
            model="gemini-flash-latest",
            contents=[prompt, types.Part.from_bytes(data=imagen_bytes, mime_type="image/jpeg")],
        )
        texto = response.text.strip()
        if texto.startswith("```"):
            texto = texto.split("```")[1]
            texto = texto[4:] if texto.lower().startswith("json") else texto
        datos = json.loads(texto)
        registrar_uso_ia("Leer cédula/título", True)
        return datos, None
    except json.JSONDecodeError:
        registrar_uso_ia("Leer cédula/título", False)
        return None, "No pude interpretar la respuesta como datos del vehículo — probá con una foto más clara."
    except Exception as e:
        registrar_uso_ia("Leer cédula/título", False)
        return None, traducir_error_gemini(e)


def transcribir_audio(audio_bytes, mime_type="audio/wav"):
    """Transcribe un audio a texto con Gemini — esto es solo 'hablar en vez de tipear', no un
    asistente conversacional: el texto transcripto se busca con el buscador normal de siempre."""
    from google import genai
    from google.genai import types

    api_key = st.secrets.get("gemini_api_key") if hasattr(st, "secrets") else None
    if not api_key:
        return None, "No configuraste 'gemini_api_key' en Streamlit Cloud (Settings → Secrets)."

    try:
        client = genai.Client(api_key=api_key)
        prompt = "Transcribí exactamente lo que se dice en este audio, en español. Devolvé solo el texto transcripto, nada más — sin comillas, sin comentarios."
        response = client.models.generate_content(
            model="gemini-flash-latest",
            contents=[prompt, types.Part.from_bytes(data=audio_bytes, mime_type=mime_type)],
        )
        registrar_uso_ia("Búsqueda por voz", True)
        return response.text.strip(), None
    except Exception as e:
        registrar_uso_ia("Búsqueda por voz", False)
        return None, traducir_error_gemini(e)


def leer_remito_por_foto(imagen_bytes):
    """Le pide a Gemini que lea un remito/factura de proveedor y devuelva los ítems en JSON.
    Devuelve (lista_items, error) — lista_items siempre queda para revisión manual antes de
    tocar el stock, la IA nunca actualiza nada por sí sola."""
    from google import genai
    from google.genai import types
    import json

    api_key = st.secrets.get("gemini_api_key") if hasattr(st, "secrets") else None
    if not api_key:
        return None, "No configuraste 'gemini_api_key' en Streamlit Cloud (Settings → Secrets)."

    try:
        client = genai.Client(api_key=api_key)
        prompt = (
            "Esta es una foto de un remito o factura de un proveedor de repuestos. Extraé cada ítem "
            "listado y devolvé ÚNICAMENTE un JSON válido (sin texto extra, sin markdown), con esta forma "
            'exacta: [{"codigo": "...", "descripcion": "...", "cantidad": 0, "costo_unitario": 0.0}, ...]. '
            "Si no podés leer algún campo con claridad, dejalo como null. No inventes datos que no estén "
            "visibles en la imagen."
        )
        response = client.models.generate_content(
            model="gemini-flash-latest",
            contents=[prompt, types.Part.from_bytes(data=imagen_bytes, mime_type="image/jpeg")],
        )
        texto = response.text.strip()
        if texto.startswith("```"):
            texto = texto.split("```")[1]
            texto = texto[4:] if texto.lower().startswith("json") else texto
        items = json.loads(texto)
        if not isinstance(items, list):
            registrar_uso_ia("Leer remito por foto", False)
            return None, "Gemini no devolvió una lista de ítems reconocible."
        registrar_uso_ia("Leer remito por foto", True)
        return items, None
    except json.JSONDecodeError:
        registrar_uso_ia("Leer remito por foto", False)
        return None, "No pude interpretar la respuesta como una lista de ítems — probá con una foto más clara."
    except Exception as e:
        registrar_uso_ia("Leer remito por foto", False)
        return None, traducir_error_gemini(e)


def cotejar_items_remito(items):
    """Para cada ítem leído del remito, busca si el código ya existe en el catálogo."""
    resultado = []
    for item in items:
        codigo = (item.get("codigo") or "").strip()
        clean = sanitizar(codigo) if codigo else ""
        producto_id, marca_actual, stock_actual = None, None, None
        if clean:
            c.execute("""SELECT p.id, m.nombre AS marca, p.stock FROM productos p
                         JOIN marcas m ON m.id = p.marca_id WHERE p.codigo_clean = ? LIMIT 1""", (clean,))
            fila = c.fetchone()
            if fila:
                producto_id, marca_actual, stock_actual = fila["id"], fila["marca"], fila["stock"]
        resultado.append({
            "Código": codigo or "(sin leer)",
            "Descripción": item.get("descripcion") or "",
            "Cantidad": item.get("cantidad") if item.get("cantidad") is not None else 0,
            "Costo unitario": item.get("costo_unitario") if item.get("costo_unitario") is not None else 0.0,
            "_producto_id": producto_id,
            "Coincide con": f"{marca_actual} (stock actual: {stock_actual})" if producto_id else "❌ No está en el catálogo",
        })
    return resultado


def aplicar_carga_remito(items_cotejados):
    """Suma la cantidad recibida al stock de los ítems que sí coinciden con un producto ya cargado."""
    actualizados = 0
    with db_lock:
        for item in items_cotejados:
            if item.get("_producto_id"):
                cantidad = item.get("Cantidad") or 0
                c.execute("UPDATE productos SET stock = COALESCE(stock, 0) + ? WHERE id = ?",
                          (cantidad, item["_producto_id"]))
                actualizados += 1
        conn.commit()
    return actualizados


def actualizar_precio_stock(producto_id, precio, stock, costo=None):
    """Guarda precio, stock y —si se pasa— el precio de costo.

    El costo va como parámetro opcional para que las llamadas viejas sigan funcionando: hay
    varias en la app y cambiarlas todas de golpe es pedir un error tonto."""
    with db_lock:
        c.execute("SELECT precio FROM productos WHERE id = ?", (producto_id,))
        fila = c.fetchone()
        precio_anterior = fila["precio"] if fila else None
        if costo is not None:
            c.execute("UPDATE productos SET precio_costo = ? WHERE id = ?",
                      (costo or None, producto_id))
        c.execute("UPDATE productos SET precio = ?, stock = ? WHERE id = ?", (precio, stock, producto_id))
        # Solo se guarda un registro nuevo en el historial si el precio realmente cambió
        # (evita ensuciar el historial cada vez que se toca el stock sin tocar el precio).
        if precio_anterior != precio:
            c.execute("INSERT INTO historial_precios (producto_id, precio) VALUES (?, ?)", (producto_id, precio))
        conn.commit()


def historial_precio_producto(producto_id, limite=50):
    c.execute("""SELECT precio AS "Precio", fecha AS "Fecha" FROM historial_precios
                 WHERE producto_id = ? ORDER BY fecha DESC LIMIT ?""", (producto_id, limite))
    return filas_a_listas(c)


# Palabras que aparecen en cualquier pedido y no aportan: si se toman por modelo de auto,
# «necesito algo para el auto» terminaría buscando productos que digan «NECESITO».
_PALABRAS_DE_RELLENO = {
    "PARA", "DEL", "LOS", "LAS", "CON", "SIN", "UNA", "UNO", "UNAS", "UNOS", "QUE", "POR",
    "NECESITO", "QUIERO", "BUSCO", "TENES", "TIENE", "HAY", "DAME", "PASAME", "AUTO", "COCHE",
    "CAMIONETA", "VEHICULO", "REPUESTO", "PIEZA", "ALGO", "ESTE", "ESTA", "ESE", "ESA",
}


def interpretar_pedido_hablado(texto):
    """Saca de una frase suelta qué pieza y qué auto está pidiendo el cliente.

    En el mostrador nadie dice un código: dice «pastillas de freno para un Gol 1.6». Cuando la
    búsqueda por texto no encuentra nada, hoy se muere ahí. Pero de esa frase se puede sacar
    bastante: qué RUBRO es (con el clasificador que ya usamos) y qué AUTO (con la lista de
    marcas de vehículo). Con esas dos cosas se puede ofrecer algo en vez de nada."""
    if not texto or not texto.strip():
        return None
    limpio = normalizar_texto(texto)
    familia = clasificar_repuesto(texto)
    marca_auto = next((mv for mv in MARCAS_VEHICULO
                       if f" {mv} " in f" {limpio} "), None)
    # El cliente casi nunca dice la marca: dice «para el Gol», «para un Palio». Se busca el
    # MODELO directamente en las descripciones del catálogo, que es donde están los que
    # realmente vendés. Sin esto, «bujías para el gol» perdía el auto entero.
    modelo_suelto = None
    if not marca_auto:
        # Se descartan también las palabras que ya se usaron para reconocer el RUBRO: en
        # «pastillas de freno para un gol», «freno» aparece en muchas descripciones y se
        # tomaba por modelo de auto. No rompía el resultado, pero decía una cosa por otra —
        # y un cartel que dice «para un FRENO» hace desconfiar de todo lo demás.
        del_rubro = set()
        if familia and familia in FAMILIAS_REPUESTO:
            for clave in FAMILIAS_REPUESTO[familia]:
                del_rubro.update(clave.split())
        palabras_utiles = [w for w in limpio.split()
                           if len(w) >= 3 and w not in _PALABRAS_DE_RELLENO
                           and w not in del_rubro
                           and not w.replace(".", "").replace(",", "").isdigit()]
        for w in palabras_utiles:
            try:
                c.execute(f"""SELECT COUNT(*) FROM productos p
                              WHERE {_sql_sin_acentos('p.descripcion')} LIKE ?""", (f"% {w} %",))
                if c.fetchone()[0] >= 2:
                    modelo_suelto = w
                    break
            except Exception:
                break
    cil = re.search(r'\b(\d[.,]\d)\b', limpio)
    anio = re.search(r'\b(19\d{2}|20\d{2})\b', limpio)
    # El modelo: lo que queda después de sacar la marca, el rubro y los números
    modelo = modelo_suelto
    if marca_auto:
        resto = limpio.split(marca_auto, 1)[1].strip() if marca_auto in limpio else ""
        candidatos = [w for w in resto.split()
                      if len(w) >= 3 and not w.replace(".", "").replace(",", "").isdigit()
                      and w not in ("PARA", "DEL", "LOS", "LAS", "CON", "SIN")]
        modelo = candidatos[0] if candidatos else modelo_suelto
    if familia == "Sin clasificar" and not marca_auto and not modelo_suelto:
        return None
    return {"familia": familia if familia != "Sin clasificar" else None,
            "marca_auto": marca_auto, "modelo": modelo,
            "cilindrada": cil.group(1).replace(",", ".") if cil else None,
            "anio": int(anio.group(1)) if anio else None}


def buscar_por_pieza_y_auto(familia=None, marca_auto=None, modelo=None,
                            cilindrada=None, limite=60):
    """Lo que tenés de ese rubro para ese auto, aunque no coincida ni una palabra del pedido."""
    condiciones, params = [], []
    if marca_auto:
        condiciones.append(f"{_sql_sin_acentos('p.descripcion')} LIKE ?")
        params.append(f"%{normalizar_texto(marca_auto)}%")
    if modelo:
        condiciones.append(f"{_sql_sin_acentos('p.descripcion')} LIKE ?")
        params.append(f"%{normalizar_texto(modelo)}%")
    if cilindrada:
        condiciones.append("p.descripcion LIKE ?")
        params.append(f"%{cilindrada}%")
    if not condiciones:
        return []
    c.execute(f"""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
                         m.nombre AS "Marca", p.precio AS "Precio", p.stock AS "Stock"
                  FROM productos p JOIN marcas m ON m.id = p.marca_id
                  WHERE {" AND ".join(condiciones)}
                  LIMIT 800""", params)
    filas = filas_a_listas(c)
    if familia:
        filas = [f for f in filas if clasificar_repuesto(f["Descripcion"]) == familia]
    filas.sort(key=lambda f: (-(f["Stock"] or 0), f["Codigo"]))
    return filas[:limite]


def autos_de_un_codigo(clean_code, limite=200):
    """A qué autos le va este código, según los catálogos de fabricante que cargaste."""
    if not clean_code:
        return []
    try:
        c.execute("""SELECT DISTINCT marca_auto AS "Marca", modelo_auto AS "Modelo",
                            motor AS "Motor",
                            COALESCE(anio_desde, '') || '-' || COALESCE(anio_hasta, '') AS "Años",
                            marca_repuesto AS "Según"
                     FROM aplicaciones WHERE codigo_clean = ?
                     ORDER BY marca_auto, modelo_auto LIMIT ?""", (clean_code, limite))
        return filas_a_listas(c)
    except sqlite3.OperationalError:
        return []


def le_sirve_a_este_auto(clean_code, marca_auto, modelo="", anio=None):
    """¿Este código le va a ese auto? Devuelve (respuesta, motivo).

    Es la pregunta que más se hace en el mostrador después de «¿lo tenés?», y la que más caro
    sale contestar mal: una pieza vendida para el auto equivocado vuelve, y con ella se va la
    confianza del cliente.

    Se responde SOLO con lo que dice el catálogo del fabricante. Cuando no hay dato, se dice que
    no hay dato — no se estira una respuesta que no se tiene."""
    aplic = autos_de_un_codigo(clean_code)
    if not aplic:
        return None, ("Ese código no figura en ningún catálogo de aplicaciones cargado, así que "
                      "no puedo confirmarlo por acá. Verificalo por medidas o con el proveedor.")

    marca_auto = (marca_auto or "").strip().upper()
    modelo = (modelo or "").strip().upper()
    coinciden = [a for a in aplic if marca_auto and marca_auto in (a["Marca"] or "").upper()]
    if modelo:
        coinciden = [a for a in coinciden if modelo in (a["Modelo"] or "").upper()]

    if not coinciden:
        marcas_que_si = sorted({a["Marca"] for a in aplic})[:6]
        return False, ("Según el catálogo, ese código **no** figura para ese auto. Sí figura "
                       f"para: {', '.join(marcas_que_si)}.")

    if anio:
        en_anio = []
        for a in coinciden:
            desde, hasta = (a["Años"].split("-") + [""])[:2]
            d = int(desde) if desde.isdigit() else None
            h = int(hasta) if hasta.isdigit() else None
            if (d is None or d <= int(anio)) and (h is None or h >= int(anio)):
                en_anio.append(a)
        if not en_anio:
            rangos = ", ".join(sorted({a["Años"] for a in coinciden}))[:80]
            return False, (f"Figura para ese modelo, pero **no para el año {anio}**. "
                           f"El catálogo lo da para: {rangos}.")
        coinciden = en_anio

    motores = sorted({a["Motor"] for a in coinciden if a["Motor"]})[:5]
    fuente = sorted({a["Según"] for a in coinciden})
    detalle = f"Lo dice el catálogo de {', '.join(fuente)}."
    if motores:
        detalle += f" Motorizaciones: {', '.join(motores)}."
    return True, detalle


def identificar_codigo_ajeno(clean_code):
    """Un código que NO tenés cargado, pero que aparece en algún catálogo de fabricante.

    Es una venta que hoy se pierde. Te piden un código, no lo tenés, y la app dice «no hay
    ningún producto con ese código» — punto. Pero puede estar en un catálogo de aplicaciones
    que ya cargaste: ahí figura qué pieza es, de qué marca y a qué autos le va.

    Con eso se puede ofrecer un equivalente en vez de perder al cliente."""
    if not clean_code:
        return None
    try:
        c.execute("""SELECT codigo, marca_repuesto, tipo_pieza,
                            COUNT(*) AS aplicaciones,
                            COUNT(DISTINCT marca_auto || '|' || modelo_auto) AS autos
                     FROM aplicaciones WHERE codigo_clean = ?
                     GROUP BY codigo, marca_repuesto, tipo_pieza LIMIT 1""", (clean_code,))
        info = c.fetchone()
    except sqlite3.OperationalError:
        return None
    if not info:
        return None

    c.execute("""SELECT DISTINCT marca_auto, modelo_auto, motor, anio_desde, anio_hasta
                 FROM aplicaciones WHERE codigo_clean = ?
                 ORDER BY marca_auto, modelo_auto LIMIT 40""", (clean_code,))
    autos = [dict(r) for r in c.fetchall()]
    return {"codigo": info["codigo"], "marca": info["marca_repuesto"],
            "tipo": info["tipo_pieza"] or "", "autos": autos,
            "cantidad_autos": info["autos"]}


def equivalentes_para_los_mismos_autos(clean_code, limite=30):
    """Qué SÍ tenés que le sirva a los mismos autos que ese código que no tenés.

    Es el paso que convierte «no lo tengo» en una venta. Si el código que te piden va a un
    Palio 1.0 y vos tenés otra marca que, según su propio catálogo, también va a ese Palio 1.0,
    eso es lo que hay que ofrecer.

    Se exige el mismo tipo de pieza y otro fabricante, igual que al deducir equivalencias: dos
    piezas que van al mismo auto no son intercambiables si una es una bujía y la otra un filtro."""
    try:
        c.execute("""SELECT DISTINCT p.codigo_raw AS "Código", m.nombre AS "Marca",
                            p.descripcion AS "Descripción", p.precio AS "Precio",
                            p.stock AS "Stock", b.marca_repuesto AS "Según el catálogo de",
                            COUNT(DISTINCT b.marca_auto || b.modelo_auto || b.motor) AS "Autos en común"
                     FROM aplicaciones a
                     JOIN aplicaciones b
                       ON a.marca_auto = b.marca_auto AND a.modelo_auto = b.modelo_auto
                      AND a.motor = b.motor
                      AND COALESCE(a.tipo_pieza,'') = COALESCE(b.tipo_pieza,'')
                      AND a.marca_repuesto <> b.marca_repuesto
                     JOIN productos p ON p.codigo_clean = b.codigo_clean
                     JOIN marcas m ON m.id = p.marca_id
                     WHERE a.codigo_clean = ?
                       AND COALESCE(a.tipo_pieza,'') <> ''
                     GROUP BY p.id
                     ORDER BY (p.stock > 0) DESC, "Autos en común" DESC LIMIT ?""",
                  (clean_code, limite))
        return filas_a_listas(c)
    except sqlite3.OperationalError:
        return []


def quien_conviene_por_rubro(limite=40, minimo_comparaciones=5):
    """En los repuestos que tenés de varias marcas, cuál sale más barata más seguido.

    Solo se comparan productos que son EQUIVALENTES entre sí: comparar el precio promedio de dos
    marcas sin eso no dice nada, porque una puede vender frenos caros y la otra filtros baratos.
    Acá cada comparación es entre dos códigos que hacen el mismo trabajo."""
    c.execute("""SELECT ma.nombre AS marca_a, mb.nombre AS marca_b,
                        pa.precio AS precio_a, pb.precio AS precio_b
                 FROM equivalencias e
                 JOIN productos pa ON pa.id = e.producto_a_id
                 JOIN productos pb ON pb.id = e.producto_b_id
                 JOIN marcas ma ON ma.id = pa.marca_id
                 JOIN marcas mb ON mb.id = pb.marca_id
                 WHERE pa.precio > 0 AND pb.precio > 0
                   AND ma.nombre <> mb.nombre
                   AND COALESCE(e.confianza, 50) >= 50
                   AND MAX(pa.precio, pb.precio) / MIN(pa.precio, pb.precio) < 8""")
    marcador = {}
    for r in c.fetchall():
        clave = tuple(sorted((r["marca_a"], r["marca_b"])))
        d = marcador.setdefault(clave, {"total": 0, clave[0]: 0, clave[1]: 0, "dif": []})
        d["total"] += 1
        barata = r["marca_a"] if r["precio_a"] < r["precio_b"] else r["marca_b"]
        d[barata] = d.get(barata, 0) + 1
        d["dif"].append(max(r["precio_a"], r["precio_b"]) / min(r["precio_a"], r["precio_b"]))

    salida = []
    for (ma, mb), d in marcador.items():
        if d["total"] < minimo_comparaciones:
            continue
        gana, pierde = (ma, mb) if d[ma] >= d[mb] else (mb, ma)
        proporcion = d[gana] / d["total"]
        d["dif"].sort()
        mediana_dif = d["dif"][len(d["dif"]) // 2]
        salida.append({
            "Comparación": f"{ma} vs {mb}",
            "Sale más barata": gana,
            "En": f"{proporcion * 100:.0f}% de los casos",
            "Diferencia típica": f"{(mediana_dif - 1) * 100:.0f}%",
            "Repuestos comparados": d["total"],
            "_prop": proporcion, "_total": d["total"],
        })
    salida.sort(key=lambda x: (-x["_prop"], -x["_total"]))
    return salida[:limite]


def variacion_de_precios_por_marca(meses=6, minimo_productos=10):
    """Cuánto aumentó cada proveedor, medido sobre tu propio historial.

    La app viene guardando cada cambio de precio y solo lo mostraba producto por producto. Pero
    la pregunta que importa no es cuánto aumentó UN filtro, sino cuánto aumentó el proveedor:
    con eso se decide a quién comprarle y con quién hay que hablar.

    Se usa la MEDIANA, no el promedio. Un solo precio mal cargado —de esos que llegan con el
    separador de decimales al revés y quedan cien veces más caros— alcanza para inflar un
    promedio y hacer parecer que un proveedor aumentó 400%. La mediana lo ignora."""
    c.execute("""SELECT m.nombre AS marca, p.id AS pid,
                        (SELECT hp.precio FROM historial_precios hp
                          WHERE hp.producto_id = p.id AND hp.fecha >= datetime('now', ?)
                          ORDER BY hp.fecha ASC LIMIT 1) AS antes,
                        (SELECT hp.precio FROM historial_precios hp
                          WHERE hp.producto_id = p.id
                          ORDER BY hp.fecha DESC LIMIT 1) AS ahora
                 FROM productos p JOIN marcas m ON m.id = p.marca_id""",
              (f"-{int(meses) * 30} days",))
    por_marca = {}
    for r in c.fetchall():
        antes, ahora = r["antes"], r["ahora"]
        if not antes or not ahora or antes <= 0 or ahora <= 0 or antes == ahora:
            continue
        razon = ahora / antes
        # Se descartan los saltos absurdos: son errores de carga, no aumentos
        if razon > 20 or razon < 0.05:
            continue
        por_marca.setdefault(r["marca"], []).append(razon)

    salida = []
    for marca, razones in por_marca.items():
        if len(razones) < minimo_productos:
            continue
        razones.sort()
        n = len(razones)
        mediana = razones[n // 2] if n % 2 else (razones[n // 2 - 1] + razones[n // 2]) / 2
        subieron = sum(1 for x in razones if x > 1.01)
        salida.append({
            "Marca": marca,
            "Aumento típico": f"{(mediana - 1) * 100:+.0f}%",
            "Productos medidos": n,
            "Subieron": f"{subieron / n * 100:.0f}%",
            "El que más subió": f"{(razones[-1] - 1) * 100:+.0f}%",
            "El que menos": f"{(razones[0] - 1) * 100:+.0f}%",
            "_mediana": mediana,
        })
    salida.sort(key=lambda x: -x["_mediana"])
    return salida


def productos_probablemente_discontinuados(marca_id=None, listas_seguidas=2, limite=300):
    """Productos que el proveedor dejó de mandar en sus últimas listas.

    Cuando llega una lista nueva de una marca, los productos que vienen se actualizan y los que
    NO vienen quedan intactos. Si un código faltó en las últimas dos o tres listas, casi seguro
    el proveedor lo discontinuó o le cambió el número — pero en la base sigue figurando igual,
    con su precio viejo, como si nada.

    Eso cuesta plata de dos maneras: se cotiza algo que ya no existe y el cliente se va cuando
    no llega, y el stock muerto queda ocupando lugar sin que nadie lo note.

    No se borra nada: son candidatos para que los mires. Un producto puede faltar en una lista
    simplemente porque el proveedor lo mandó aparte."""
    c.execute("""SELECT m.id, m.nombre, COUNT(*) AS listas, MAX(i.fecha) AS ultima
                 FROM importaciones i JOIN marcas m ON UPPER(m.nombre) = UPPER(i.marca)
                 WHERE (? IS NULL OR m.id = ?)
                 GROUP BY m.id HAVING COUNT(*) >= ?""",
              (marca_id, marca_id, listas_seguidas))
    marcas = [dict(r) for r in c.fetchall()]
    if not marcas:
        return [], 0

    salida = []
    revisados = 0
    for mk in marcas:
        # La fecha de corte: el comienzo de las últimas N listas de esta marca
        c.execute("""SELECT fecha FROM importaciones
                     WHERE UPPER(marca) = UPPER(?) ORDER BY fecha DESC LIMIT ?""",
                  (mk["nombre"], listas_seguidas))
        fechas = [r["fecha"] for r in c.fetchall()]
        if len(fechas) < listas_seguidas:
            continue
        corte = fechas[-1]

        c.execute("""SELECT p.id AS "_id", p.codigo_raw AS "Código", p.descripcion AS "Descripción",
                            p.precio AS "Precio", p.stock AS "Stock",
                            (SELECT MAX(hp.fecha) FROM historial_precios hp
                              WHERE hp.producto_id = p.id) AS "_ultimo_precio",
                            (SELECT COUNT(*) FROM ventas_registradas v
                              WHERE v.producto_id = p.id
                                AND v.fecha >= datetime('now','-365 days')) AS "Ventas del año"
                     FROM productos p WHERE p.marca_id = ?""", (mk["id"],))
        for f in filas_a_listas(c):
            revisados += 1
            ultimo = f.pop("_ultimo_precio", None)
            # Sin precio nunca cargado no se puede concluir nada: puede que esa marca no traiga
            # precios en sus listas.
            if not ultimo or ultimo >= corte:
                continue
            f["Marca"] = mk["nombre"]
            f["Sin aparecer desde"] = str(ultimo)[:10]
            f["_stock"] = f["Stock"] or 0
            salida.append(f)

    # Primero los que tenés en stock: es plata parada en algo que quizá ya no se pide
    salida.sort(key=lambda x: (-(x["_stock"] or 0), -(x["Ventas del año"] or 0)))
    return salida[:limite], revisados


def guardar_reemplazo(codigo_viejo, codigo_nuevo, marca="", nota=""):
    """Anota que un código fue reemplazado por otro. Devuelve (ok, mensaje)."""
    v, n = sanitizar(codigo_viejo), sanitizar(codigo_nuevo)
    if not v or not n:
        return False, "Faltan los dos códigos."
    if v == n:
        return False, "Son el mismo código."
    # Si el nuevo ya lleva de vuelta al viejo, esto arma un círculo (A→B→A). Buscar no se cuelga
    # —la cadena corta al repetirse— pero la respuesta pasa a depender de por dónde se entre, que
    # es peor que no tener el dato: se ve creíble y está mal.
    if any(paso["clean"] == v for paso in cadena_de_reemplazos(n)):
        return False, (f"No se puede: {codigo_nuevo} ya lleva de vuelta a {codigo_viejo}. "
                       "Revisá cuál de los dos es el vigente.")
    # Un código viejo tiene UN reemplazo vigente, no varios. La clave de la tabla es el par
    # (viejo, nuevo), así que cargar A→B y después A→C dejaba las dos filas, y la búsqueda seguía
    # la que SQLite devolviera primero: la misma consulta podía contestar B o C según cómo
    # estuvieran guardadas las filas. Se reemplaza el anterior y se avisa cuál se pisó.
    anterior = None
    try:
        c.execute("""SELECT codigo_nuevo FROM reemplazos_codigo
                     WHERE codigo_viejo_clean = ? AND codigo_nuevo_clean != ?""", (v, n))
        fila_previa = c.fetchone()
        anterior = fila_previa["codigo_nuevo"] if fila_previa else None
    except sqlite3.OperationalError:
        anterior = None
    with db_lock:
        c.execute("DELETE FROM reemplazos_codigo WHERE codigo_viejo_clean = ?", (v,))
        c.execute("""INSERT OR REPLACE INTO reemplazos_codigo
                     (codigo_viejo, codigo_viejo_clean, codigo_nuevo, codigo_nuevo_clean,
                      marca, nota, cargado_por)
                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                  (codigo_viejo.strip(), v, codigo_nuevo.strip(), n,
                   (marca or "").strip().upper() or None, (nota or "").strip() or None,
                   obtener_usuario_actual()))
        conn.commit()
    if anterior:
        return True, (f"{codigo_viejo} → {codigo_nuevo} anotado. "
                      f"Antes decía que lo reemplazaba {anterior}; quedó este.")
    return True, f"{codigo_viejo} → {codigo_nuevo} anotado."


def cadena_de_reemplazos(clean_code, tope=6):
    """Sigue la cadena de reemplazos hasta el código vigente.

    Los fabricantes discontinúan un código y lo reemplazan por otro, que a su vez puede volver
    a reemplazarse. Sin esto, alguien busca el código viejo, no aparece, y se rechaza una venta
    creyendo que la pieza no existe — cuando en realidad existe con otro número.

    Tiene tope de saltos porque un dato mal cargado puede armar un círculo (A→B→A) y dejar la
    búsqueda dando vueltas para siempre."""
    if not clean_code:
        return []
    cadena, visto, actual = [], {clean_code}, clean_code
    for _ in range(tope):
        try:
            # Por si quedaron filas viejas de antes de que se guardara uno solo por código:
            # gana el más reciente, para que la respuesta no dependa del orden de la tabla.
            c.execute("""SELECT codigo_nuevo, codigo_nuevo_clean, marca, nota
                         FROM reemplazos_codigo WHERE codigo_viejo_clean = ?
                         ORDER BY fecha DESC LIMIT 1""", (actual,))
            fila = c.fetchone()
        except sqlite3.OperationalError:
            return []
        if not fila or fila["codigo_nuevo_clean"] in visto:
            break
        cadena.append({"codigo": fila["codigo_nuevo"], "clean": fila["codigo_nuevo_clean"],
                       "marca": fila["marca"] or "", "nota": fila["nota"] or ""})
        visto.add(fila["codigo_nuevo_clean"])
        actual = fila["codigo_nuevo_clean"]
    return cadena


def productos_estancados(dias_sin_vender=180, minimo_stock=1, limite=100):
    """Lo que tenés en el estante y no se mueve. Es capital dormido.

    Se cruza el stock con la última venta. Lo que nunca se vendió cuenta como estancado solo si
    hace rato que está cargado: un producto que entró la semana pasada todavía no tuvo chance.

    Para saber desde cuándo está se usa el primer cambio de precio y, si no tuvo ninguno, la
    fecha en que se cargó. Antes se miraba SOLO el historial de precios, y ese historial recién
    se escribe cuando un precio CAMBIA: un producto importado una vez y nunca tocado no tiene
    ninguna fila ahí. O sea que quedaban afuera justamente los que nunca se movieron —los clavos
    más clavos, que son los que esta pantalla existe para encontrar."""
    try:
        # Las fechas se sacan con dos tablas ya resumidas, no con una subconsulta por
        # producto: así se leen las ventas una sola vez en total y no una vez por artículo.
        # Tampoco se puede unir directo contra las tablas crudas, porque un producto con 30
        # ventas y 10 cambios de precio saldría 300 veces antes de agrupar.
        c.execute("""SELECT p.id AS "_id", p.codigo_raw AS "Código", m.nombre AS "Marca",
                            p.descripcion AS "Descripción", p.stock AS "Stock",
                            p.precio AS "Precio",
                            v.ultima AS "_ultima_venta",
                            COALESCE(h.primera, p.created_at) AS "_desde"
                     FROM productos p
                     JOIN marcas m ON m.id = p.marca_id
                     LEFT JOIN (SELECT producto_id, MAX(fecha) AS ultima
                                  FROM ventas_registradas GROUP BY producto_id) v
                            ON v.producto_id = p.id
                     LEFT JOIN (SELECT producto_id, MIN(fecha) AS primera
                                  FROM historial_precios GROUP BY producto_id) h
                            ON h.producto_id = p.id
                     WHERE COALESCE(p.stock, 0) >= ?""", (minimo_stock,))
        filas = filas_a_listas(c)
    except sqlite3.OperationalError:
        return []

    from datetime import datetime as _dt
    hoy = _dt.now()
    salida = []
    for f in filas:
        ultima = f.pop("_ultima_venta", None)
        desde = f.pop("_desde", None)
        if ultima:
            try:
                dias = (hoy - _dt.strptime(str(ultima)[:10], "%Y-%m-%d")).days
            except Exception:
                continue
            detalle = f"hace {dias} día(s)"
        else:
            # Nunca vendido: solo cuenta si hace rato que está en la base
            if not desde:
                continue
            try:
                dias = (hoy - _dt.strptime(str(desde)[:10], "%Y-%m-%d")).days
            except Exception:
                continue
            detalle = "nunca se vendió"
        if dias < dias_sin_vender:
            continue
        f["Última venta"] = detalle
        f["Plata parada"] = (f["Precio"] or 0) * (f["Stock"] or 0)
        f["_dias"] = dias
        salida.append(f)
    # Primero lo que más plata tiene inmovilizada
    salida.sort(key=lambda x: -(x["Plata parada"] or 0))
    return salida[:limite]


def productos_por_quebrar(dias_aviso=21, minimo_ventas=3, limite=100):
    """Qué se va a acabar antes de que alguien se dé cuenta.

    La reposición hoy es 100% manual: alguien tiene que acordarse de tocar «Pedir». Y de lo que
    uno no se acuerda es justamente de lo que se vende parejo todos los días — el filtro común
    que nadie mira hasta que un cliente lo pide y no está.

    Se calcula el ritmo real de venta de cada producto y se estima en cuántos días se termina.
    Se ignora lo que se vendió una o dos veces: con eso no se puede calcular un ritmo, solo
    ruido."""
    c.execute("""SELECT p.id AS "_id", p.codigo_raw AS "Código", m.nombre AS "Marca",
                        p.descripcion AS "Descripción", p.stock AS "Stock",
                        COUNT(v.id) AS "Vendidos",
                        julianday('now') - julianday(MIN(v.fecha)) AS "_dias_historia"
                 FROM productos p
                 JOIN marcas m ON m.id = p.marca_id
                 JOIN ventas_registradas v ON v.producto_id = p.id
                 WHERE v.fecha >= datetime('now', '-180 days')
                 GROUP BY p.id
                 HAVING COUNT(v.id) >= ?
                 ORDER BY COUNT(v.id) DESC LIMIT 400""", (minimo_ventas,))
    filas = filas_a_listas(c)

    salida = []
    for f in filas:
        dias = max(f["_dias_historia"] or 0, 7)      # con menos de una semana no hay ritmo
        por_dia = f["Vendidos"] / dias
        if por_dia <= 0:
            continue
        stock = f["Stock"] or 0
        dias_restantes = stock / por_dia
        if dias_restantes > dias_aviso:
            continue
        salida.append({
            "Código": f["Código"], "Marca": f["Marca"],
            "Descripción": (f["Descripción"] or "")[:48],
            "Stock": stock,
            "Se vende": f"{por_dia * 30:.1f} por mes",
            "Se acaba en": ("ya sin stock" if stock <= 0
                             else f"{dias_restantes:.0f} día(s)"),
            "_dias": dias_restantes, "_id": f["_id"],
        })
    salida.sort(key=lambda x: x["_dias"])
    return salida[:limite]


def solicitar_reposicion(producto_id):
    with db_lock:
        c.execute(
            "INSERT INTO pedidos_reposicion (producto_id, veces_solicitado, ultimo_solicitado_por, ultima_fecha, estado) "
            "VALUES (?, 1, ?, datetime('now'), 'pendiente') "
            "ON CONFLICT(producto_id) DO UPDATE SET veces_solicitado = veces_solicitado + 1, "
            "ultimo_solicitado_por = excluded.ultimo_solicitado_por, ultima_fecha = excluded.ultima_fecha, "
            "estado = 'pendiente'",
            (producto_id, obtener_usuario_actual())
        )
        conn.commit()


def listar_pedidos_reposicion(estado="pendiente"):
    c.execute("""SELECT pr.id AS "ID", p.id AS "ProductoID", p.codigo_raw AS "Codigo",
                 p.descripcion AS "Descripcion", m.nombre AS "Marca", p.stock AS "Stock actual",
                 pr.veces_solicitado AS "Veces pedido", pr.ultimo_solicitado_por AS "Último en pedirlo",
                 pr.ultima_fecha AS "Fecha"
                 FROM pedidos_reposicion pr
                 JOIN productos p ON p.id = pr.producto_id
                 JOIN marcas m ON m.id = p.marca_id
                 WHERE pr.estado = ?
                 ORDER BY pr.veces_solicitado DESC, pr.ultima_fecha DESC""", (estado,))
    return filas_a_listas(c)


def marcar_pedido_resuelto(pedido_id):
    with db_lock:
        c.execute("UPDATE pedidos_reposicion SET estado = 'resuelto' WHERE id = ?", (pedido_id,))
        conn.commit()


def descartar_pedido_reposicion(pedido_id):
    with db_lock:
        c.execute("DELETE FROM pedidos_reposicion WHERE id = ?", (pedido_id,))
        conn.commit()


def alternar_favorito(producto_id, valor):
    with db_lock:
        c.execute("UPDATE productos SET favorito = ? WHERE id = ?", (1 if valor else 0, producto_id))
        conn.commit()


def listar_favoritos():
    c.execute("""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion",
                 m.nombre AS "Marca", p.precio AS "Precio", p.stock AS "Stock"
                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                 WHERE p.favorito = 1 ORDER BY p.codigo_raw""")
    return filas_a_listas(c)


def obtener_config(clave, default=""):
    c.execute("SELECT valor FROM configuracion WHERE clave = ?", (clave,))
    fila = c.fetchone()
    return fila["valor"] if fila and fila["valor"] is not None else default


def guardar_config(clave, valor):
    with db_lock:
        c.execute(
            "INSERT INTO configuracion (clave, valor) VALUES (?, ?) "
            "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor",
            (clave, valor)
        )
        conn.commit()


def obtener_usuario_actual():
    """Nombre que identifica a la persona para su historial personal: si está logueada como
    admin usa ese nombre, si no usa el que puso al entrar, o 'Invitado' si no puso nada."""
    return st.session_state.get("admin_nombre") or st.session_state.get("usuario_nombre") or "Invitado"


def registrar_uso_ia(funcion, exito):
    with db_lock:
        c.execute("INSERT INTO uso_ia (funcion, usuario, exito) VALUES (?, ?, ?)",
                   (funcion, obtener_usuario_actual(), 1 if exito else 0))
        conn.commit()


def traducir_error_gemini(e):
    """Convierte el JSON crudo de error de la API de Gemini en un mensaje legible en español.
    Estas funciones tienen un límite de consultas por minuto compartido entre todos los
    empleados — chocarse con ese límite es lo más común que puede pasar."""
    texto_error = str(e)
    if "RESOURCE_EXHAUSTED" in texto_error or "429" in texto_error or "quota" in texto_error.lower():
        return (
            "⏳ Se alcanzó el límite de consultas de la IA por ahora (es un límite por minuto, "
            "compartido entre todos los que usan la app). Esperá un minuto y probá de nuevo — "
            "no es un error, es solo que hay que esperar a que se libere cupo."
        )
    return f"Error consultando a Gemini: {texto_error}"


def resumen_uso_ia(dias=30):
    c.execute("""SELECT funcion, COUNT(*) AS total, SUM(exito) AS exitosos
                 FROM uso_ia WHERE fecha >= datetime('now', ?) GROUP BY funcion ORDER BY total DESC""",
              (f"-{dias} days",))
    return [{"Función": r["funcion"], "Usos": r["total"], "Exitosos": r["exitosos"],
              "Con error": r["total"] - r["exitosos"]} for r in c.fetchall()]


def mover_a_papelera(tipo, datos_dict):
    with db_lock:
        c.execute("INSERT INTO papelera (tipo, datos_json, eliminado_por) VALUES (?, ?, ?)",
                   (tipo, json.dumps(datos_dict, ensure_ascii=False), obtener_usuario_actual()))
        conn.commit()


def eliminar_marca_con_papelera(nombre_marca):
    """Guarda la marca completa (con todos sus productos y las equivalencias que los tocan)
    en la papelera antes de borrarla — es la operación más destructiva de la app, así que
    ahora también tiene red de seguridad."""
    c.execute("SELECT * FROM marcas WHERE nombre = ?", (nombre_marca,))
    marca_row = c.fetchone()
    if not marca_row:
        return False
    marca_id = marca_row["id"]
    c.execute("SELECT * FROM productos WHERE marca_id = ?", (marca_id,))
    productos_rows = [dict(r) for r in c.fetchall()]
    producto_ids = [p["id"] for p in productos_rows]
    equivalencias_rows = []
    if producto_ids:
        placeholders = ",".join("?" * len(producto_ids))
        c.execute(
            f"SELECT * FROM equivalencias WHERE producto_a_id IN ({placeholders}) "
            f"OR producto_b_id IN ({placeholders})",
            producto_ids + producto_ids
        )
        equivalencias_rows = [dict(r) for r in c.fetchall()]

    snapshot = {"marca": dict(marca_row), "productos": productos_rows, "equivalencias": equivalencias_rows}
    mover_a_papelera("marca", snapshot)

    with db_lock:
        c.execute("DELETE FROM marcas WHERE id = ?", (marca_id,))
        conn.commit()
    return True


def listar_papelera():
    c.execute("""SELECT id AS "ID", tipo AS "Tipo", datos_json, eliminado_por AS "Eliminado por",
                 eliminado_en AS "Fecha" FROM papelera ORDER BY id DESC LIMIT 100""")
    filas = []
    for row in c.fetchall():
        detalle = ""
        if row["Tipo"] == "marca":
            datos = json.loads(row["datos_json"])
            detalle = f"{datos['marca']['nombre']} ({len(datos['productos'])} producto(s))"
        elif row["Tipo"] == "producto":
            datos = json.loads(row["datos_json"])
            detalle = datos.get("codigo_raw", "")
        elif row["Tipo"] == "combo":
            datos = json.loads(row["datos_json"])
            detalle = datos.get("disparador", "")
        elif row["Tipo"] == "alias":
            datos = json.loads(row["datos_json"])
            detalle = datos.get("nombre", "")
        filas.append({"ID": row["ID"], "Tipo": row["Tipo"], "Detalle": detalle,
                       "Eliminado por": row["Eliminado por"], "Fecha": row["Fecha"]})
    return filas


def quitar_item_lista_sesion(nombre_lista, indice):
    """Saca un ítem de una lista guardada en la sesión (presupuesto del mecánico, tanda de
    equivalencias). Va como callback para no tener que forzar un refresco de página."""
    lista = st.session_state.get(nombre_lista)
    if lista and 0 <= indice < len(lista):
        lista.pop(indice)


def borrar_papelera_definitivo(item_id):
    with db_lock:
        c.execute("DELETE FROM papelera WHERE id = ?", (item_id,))
        conn.commit()


def cb_restaurar_papelera(item_id):
    """Callback para el botón de restaurar. Guarda el resultado en session_state para poder
    mostrarlo después del refresco, ya que un callback corre antes de dibujar la pantalla."""
    ok, error = restaurar_de_papelera(item_id)
    st.session_state["resultado_papelera"] = ("ok", "Restaurado.") if ok else ("error", error)


def vaciar_papelera_antigua(dias=30):
    """Borra en forma permanente lo que ya lleva más de `dias` en la papelera."""
    with db_lock:
        c.execute("DELETE FROM papelera WHERE eliminado_en < datetime('now', ?)", (f"-{dias} days",))
        conn.commit()


def restaurar_de_papelera(item_id):
    c.execute("SELECT tipo, datos_json FROM papelera WHERE id = ?", (item_id,))
    row = c.fetchone()
    if not row:
        return False, "No se encontró ese ítem en la papelera (puede que ya se haya restaurado)."
    tipo, datos = row["tipo"], json.loads(row["datos_json"])
    with db_lock:
        try:
            if tipo == "combo":
                for item in datos["items"]:
                    c.execute("INSERT INTO combos_sugeridos (disparador, item) VALUES (?, ?)",
                              (datos["disparador"], item))
            elif tipo == "alias":
                c.execute(
                    "INSERT INTO alias_transferencia (nombre, alias, cbu, titular) VALUES (?, ?, ?, ?)",
                    (datos["nombre"], datos["alias"], datos["cbu"], datos["titular"])
                )
            elif tipo == "producto":
                columnas = ", ".join(datos.keys())
                placeholders = ", ".join("?" * len(datos))
                c.execute(f"INSERT INTO productos ({columnas}) VALUES ({placeholders})", list(datos.values()))
            elif tipo == "marca":
                marca = datos["marca"]
                columnas_marca = ", ".join(marca.keys())
                placeholders_marca = ", ".join("?" * len(marca))
                c.execute(f"INSERT INTO marcas ({columnas_marca}) VALUES ({placeholders_marca})",
                          list(marca.values()))
                for producto in datos["productos"]:
                    columnas_p = ", ".join(producto.keys())
                    placeholders_p = ", ".join("?" * len(producto))
                    c.execute(f"INSERT INTO productos ({columnas_p}) VALUES ({placeholders_p})",
                              list(producto.values()))
                for equiv in datos["equivalencias"]:
                    columnas_e = ", ".join(equiv.keys())
                    placeholders_e = ", ".join("?" * len(equiv))
                    c.execute(f"INSERT INTO equivalencias ({columnas_e}) VALUES ({placeholders_e})",
                              list(equiv.values()))
            else:
                return False, f"No sé cómo restaurar el tipo '{tipo}'."
            c.execute("DELETE FROM papelera WHERE id = ?", (item_id,))
            conn.commit()
            return True, None
        except Exception as e:
            conn.rollback()
            return False, f"No se pudo restaurar: {e}"


def guardar_busqueda(termino):
    with db_lock:
        c.execute("INSERT INTO historial_busquedas (termino, usuario) VALUES (?, ?)",
                   (termino, obtener_usuario_actual()))
        conn.commit()


def historial_reciente(limite=10):
    """Solo las búsquedas de la persona actual — antes mezclaba las de todos los empleados."""
    c.execute("""SELECT DISTINCT termino FROM historial_busquedas WHERE usuario = ?
                 ORDER BY id DESC LIMIT ?""", (obtener_usuario_actual(), limite))
    return [r["termino"] for r in c.fetchall()]


def similitud(a, b):
    """Similitud simple entre dos strings (0 a 1) usando coincidencia de secuencia."""
    import difflib
    return difflib.SequenceMatcher(None, a, b).ratio()


def detectar_posibles_duplicados(marca_id, umbral=0.87, limite_productos=1500):
    """Busca códigos parecidos pero no idénticos dentro de la misma marca (posibles errores de tipeo).
    Es una comparación O(n²), así que por seguridad no corre si la marca tiene demasiados productos."""
    c.execute("SELECT id, codigo_raw, codigo_clean FROM productos WHERE marca_id = ?", (marca_id,))
    productos = c.fetchall()
    if len(productos) > limite_productos:
        return None  # catálogo muy grande: se omite para no colgar la app
    sospechosos = []
    vistos = set()
    for i in range(len(productos)):
        for j in range(i + 1, len(productos)):
            a, b = productos[i], productos[j]
            if a["codigo_clean"] == b["codigo_clean"]:
                continue
            par = tuple(sorted([a["id"], b["id"]]))
            if par in vistos:
                continue
            if similitud(a["codigo_clean"], b["codigo_clean"]) >= umbral:
                sospechosos.append({"Código 1": a["codigo_raw"], "Código 2": b["codigo_raw"]})
                vistos.add(par)
    return sospechosos


def quitar_id(filas):
    """Quita la clave ID de cada diccionario para mostrar en pantalla."""
    return [{k: v for k, v in f.items() if k != "ID"} for f in filas]


def to_excel_bytes(filas, columnas=None):
    """Genera un archivo .xlsx en memoria a partir de una lista de diccionarios."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    if not filas:
        wb.save(buf := io.BytesIO())
        return buf.getvalue()
    columnas = columnas or list(filas[0].keys())
    ws.append(columnas)
    for fila in filas:
        ws.append([fila.get(col, "") for col in columnas])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def listar_alias_transferencia():
    c.execute("""SELECT id AS "ID", nombre AS "Nombre", alias AS "Alias",
                 cbu AS "CBU", titular AS "Titular",
                 CASE WHEN qr_real_blob IS NOT NULL THEN 1 ELSE 0 END AS "TieneQrReal"
                 FROM alias_transferencia ORDER BY nombre""")
    return filas_a_listas(c)


def guardar_alias_transferencia(nombre, alias, cbu, titular, alias_id=None, qr_real_bytes=None):
    with db_lock:
        if alias_id:
            if qr_real_bytes is not None:
                c.execute(
                    "UPDATE alias_transferencia SET nombre=?, alias=?, cbu=?, titular=?, qr_real_blob=? WHERE id=?",
                    (nombre.strip(), alias.strip(), cbu.strip(), titular.strip(), qr_real_bytes, alias_id)
                )
            else:
                c.execute(
                    "UPDATE alias_transferencia SET nombre=?, alias=?, cbu=?, titular=? WHERE id=?",
                    (nombre.strip(), alias.strip(), cbu.strip(), titular.strip(), alias_id)
                )
        else:
            c.execute(
                "INSERT INTO alias_transferencia (nombre, alias, cbu, titular, qr_real_blob) VALUES (?, ?, ?, ?, ?)",
                (nombre.strip(), alias.strip(), cbu.strip(), titular.strip(), qr_real_bytes)
            )
        conn.commit()


def obtener_qr_real(alias_id):
    c.execute("SELECT qr_real_blob FROM alias_transferencia WHERE id = ?", (alias_id,))
    fila = c.fetchone()
    return fila["qr_real_blob"] if fila else None


def eliminar_qr_real(alias_id):
    with db_lock:
        c.execute("UPDATE alias_transferencia SET qr_real_blob = NULL WHERE id = ?", (alias_id,))
        conn.commit()


def eliminar_alias_transferencia(alias_id):
    c.execute("SELECT nombre, alias, cbu, titular, qr_real_blob FROM alias_transferencia WHERE id = ?", (alias_id,))
    fila = c.fetchone()
    if fila:
        mover_a_papelera("alias", {
            "nombre": fila["nombre"], "alias": fila["alias"], "cbu": fila["cbu"], "titular": fila["titular"]
        })
        # El QR real (si tenía) no se puede guardar en la papelera como texto — si restaurás este
        # alias vas a tener que volver a subirlo.
    with db_lock:
        c.execute("DELETE FROM alias_transferencia WHERE id = ?", (alias_id,))
        conn.commit()



def generar_qr_bytes(texto):
    """Genera una imagen QR (PNG) con el texto dado — el alias/CBU/titular como texto plano.
    No es un pago directo por QR (eso requiere ser comercio adherido a un sistema de cobro real):
    al escanearlo, la mayoría de las apps de billetera muestran ese texto para que el
    cliente confirme la transferencia, en vez de tener que tipear el alias a mano."""
    import qrcode
    img = qrcode.make(texto)
    salida = io.BytesIO()
    img.save(salida, format="PNG")
    return salida.getvalue()


def seccion_plegable(titulo, key, abierto=False):
    """Una sección que se abre y se cierra, PERO que se puede usar adentro de un expander.

    Streamlit no permite un st.expander dentro de otro: tira
    «Expanders may not be nested inside other expanders» y se corta el renderizado ahí mismo,
    dejando media pantalla sin dibujar. Los resultados de la búsqueda ya van dentro de un
    expander por código, así que todo lo de adentro tiene que usar esto en vez de otro expander.
    """
    return st.toggle(titulo, key=key, value=abierto)


def explicar(resumen, detalle, abierto=False, en_expander=False):
    """Una línea corta siempre visible, y el porqué largo a un toque de distancia.

    Las explicaciones largas sirven la primera vez y estorban las otras cien: en el celular
    empujan los botones fuera de la pantalla y hay que scrollear para llegar a lo que uno vino
    a hacer. Pero borrarlas tampoco sirve — sin ellas nadie entiende para qué es cada cosa.
    Así queda el resumen a la vista y el detalle disponible para quien lo necesite.

    Ojo con dónde se la llama. Esta función abre un expander, así que llamarla adentro de otro
    deja un expander dentro de un expander: en el celular quedan dos cajas anidadas y hay que
    tocar dos veces para leer tres renglones. Las versiones viejas de Streamlit ni siquiera lo
    permitían —tiraban excepción y cortaban el renderizado ahí, con lo cual la mitad de abajo de
    la pantalla no se dibujaba—; las nuevas lo dejan pasar pero sigue quedando mal.

    Para eso está en_expander: adentro de un expander el detalle va en un popover, que se abre
    encima y no agrega otro nivel. Y por las dudas, si algo de eso no está disponible en la
    versión de Streamlit que haya instalada, el detalle se muestra directamente: peor queda un
    poco de texto de más que media pantalla sin dibujar."""
    st.caption(resumen)
    caja = None
    if not en_expander:
        try:
            caja = st.expander("¿Por qué? / ¿Cómo funciona?", expanded=abierto)
        except Exception:
            caja = None
    if caja is None:
        try:
            caja = st.popover("¿Por qué? / ¿Cómo funciona?")
        except Exception:
            st.caption(detalle)
            return
    with caja:
        st.markdown(detalle)


def avisar(tipo, texto):
    """Guarda un mensaje para mostrarlo DESPUÉS del refresco de pantalla.

    El problema que resuelve: en toda la app había mensajes escritos justo antes de un
    st.rerun(). El rerun redibuja la pantalla de cero, así que ese "Guardado" se borraba antes
    de que nadie llegara a leerlo. Uno tocaba el botón, la pantalla parpadeaba, y quedaba sin
    saber si había funcionado o no — y de paso volvía a tocarlo por las dudas."""
    st.session_state.setdefault("_avisos", []).append((tipo, texto))


def mostrar_avisos_pendientes():
    """Muestra lo que quedó guardado por avisar() antes del último refresco."""
    for tipo, texto in st.session_state.pop("_avisos", []):
        getattr(st, tipo, st.info)(texto)


def recordar_archivo(archivo, clave):
    """Guarda el archivo subido en la sesión y devuelve siempre esa copia.

    Por qué: en el celular, con conexión lenta, la subida a veces se pierde (un refresco de
    pantalla en el medio, la conexión que se corta) y el widget vuelve a quedar vacío — de ahí
    lo de 'tengo que subirlo 2 o 3 veces'. Con esto, la PRIMERA subida que llegue bien queda
    guardada, y aunque el widget se vacíe después, el archivo sigue disponible."""
    import io as _io

    if archivo is not None:
        try:
            datos = archivo.getvalue()
            if datos:
                st.session_state[f"_archivo_{clave}"] = {
                    "nombre": getattr(archivo, "name", "archivo"),
                    "datos": datos,
                }
        except Exception:
            pass

    guardado = st.session_state.get(f"_archivo_{clave}")
    if not guardado:
        return None

    copia = _io.BytesIO(guardado["datos"])
    copia.name = guardado["nombre"]
    copia.size = len(guardado["datos"])
    return copia


def _clave_widget_archivo(clave):
    """La 'key' que usa el widget de subida hoy. Le va pegado un número que se incrementa cada
    vez que se pide vaciarlo: cambiarle la key es la única forma de que Streamlit lo trate como
    un widget nuevo y arranque en blanco."""
    return f"_up_{clave}_{st.session_state.get(f'_nonce_{clave}', 0)}"


def olvidar_archivo(clave):
    """Borra la copia guardada Y vacía el widget de subida de verdad.

    Antes solo se borraba la copia de la sesión. El widget seguía teniendo el archivo adentro,
    así que en el refresco siguiente lo devolvía igual, se volvía a guardar solo, y el botón
    'Usar otra foto / otro archivo' quedaba sin efecto: no había forma de cambiar el archivo
    sin recargar la página entera. Ese era el problema de la carga que seguía apareciendo."""
    st.session_state.pop(f"_archivo_{clave}", None)
    viejo = _clave_widget_archivo(clave)
    st.session_state.pop(viejo, None)  # soltar los bytes del widget viejo, si no queda ocupando RAM
    st.session_state[f"_nonce_{clave}"] = st.session_state.get(f"_nonce_{clave}", 0) + 1


def subir_archivo(etiqueta, tipos, clave, **kwargs):
    """Subida de archivo estándar de la app: widget + memoria + reset. Usar SIEMPRE esta en vez
    de st.file_uploader directo, así todos los puntos de carga se comportan igual."""
    subido = st.file_uploader(etiqueta, type=tipos, key=_clave_widget_archivo(clave), **kwargs)
    return recordar_archivo(subido, clave)


def boton_otro_archivo(clave, etiqueta="🗑️ Usar otro archivo", key=None):
    """Botón para descartar lo subido y empezar de nuevo. Devuelve True si se tocó."""
    if st.button(etiqueta, key=key or f"btn_otro_{clave}"):
        olvidar_archivo(clave)
        st.rerun()
    return False


def archivo_listo(archivo, etiqueta="archivo"):
    """Muestra si el archivo terminó de subir. Antes los botones directamente no aparecían hasta
    que el archivo estaba, así que con una conexión lenta parecía que no había pasado nada y la
    gente volvía a tocar 2 o 3 veces pensando que había fallado. Ahora el botón está siempre,
    apagado hasta que el archivo llega, y acá abajo se ve el estado."""
    if archivo is None:
        st.caption(
            f"⏳ Todavía no llegó ningún {etiqueta}. Si ya lo elegiste y la conexión está lenta, "
            "esperá unos segundos sin volver a tocar: cuando termine de subir se avisa acá."
        )
        return False
    tamano = getattr(archivo, "size", None)
    detalle = f" ({tamano/1024:,.0f} KB)" if tamano else ""
    st.caption(f"✅ Recibido: {getattr(archivo, 'name', etiqueta)}{detalle}")
    return True


def pdf_con_cache(nombre, generador, *args):
    """El botón de descarga de Streamlit necesita el archivo listo de antemano, así que el PDF
    se arma en CADA refresco de pantalla aunque nadie lo descargue. Esto guarda el último
    generado y lo reusa mientras los datos no cambien, en vez de rehacerlo cada vez.
    Se usa un caché propio (y no st.cache_data) para no depender de cómo Streamlit compara
    listas y diccionarios: acá la clave se calcula de forma explícita y predecible."""
    try:
        firma = json.dumps(args, default=str, sort_keys=True)
    except Exception:
        return generador(*args)  # si algo no se puede resumir, se genera sin cachear
    clave = nombre + ":" + hashlib.md5(firma.encode("utf-8")).hexdigest()
    guardado = st.session_state.get("_cache_pdf")
    if guardado and guardado[0] == clave:
        return guardado[1]
    resultado = generador(*args)
    st.session_state["_cache_pdf"] = (clave, resultado)
    return resultado


def generar_pdf_cotizacion(lista_productos, incluir_precio=True, incluir_stock=False, alias_qr=None, qr_real_bytes=None):
    """Genera un PDF simple de cotización a partir de la lista armada para WhatsApp.
    Si se pasa alias_qr (un dict con nombre/alias/cbu/titular), agrega un QR con esos datos
    para transferencia — el cliente lo escanea y ve el alias/CBU listo para pegar, sin tipear."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Equivalencias El Chavo - Cotizacion", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Fecha: {datetime.now():%d/%m/%Y %H:%M}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    def limpiar(texto):
        # fpdf2 con fuentes básicas no soporta todo unicode; reemplazamos lo problemático
        return str(texto).encode("latin-1", "replace").decode("latin-1")

    for item in lista_productos:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, limpiar(f"Codigo buscado: {item['codigo_buscado']}"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for fila in item["resultados"]:
            linea = f"  - {fila['Marca']}: {fila['Codigo']}"
            if fila.get("Descripcion"):
                linea += f" - {fila['Descripcion']}"
            extras = []
            if incluir_precio and fila.get("Precio"):
                extras.append(f"${fila['Precio']:,.0f}")
            if incluir_stock and fila.get("Stock") is not None:
                extras.append(f"Stock: {fila['Stock']}")
            if extras:
                linea += " (" + " / ".join(extras) + ")"
            pdf.multi_cell(0, 6, limpiar(linea), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    if alias_qr:
        qr_bytes = qr_real_bytes if qr_real_bytes else generar_qr_bytes(
            f"Alias: {alias_qr['Alias']}\nCBU: {alias_qr['CBU']}\nTitular: {alias_qr['Titular']}"
        )
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, limpiar(f"Transferir a: {alias_qr['Nombre']}"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(0, 5, limpiar(f"Alias: {alias_qr['Alias']}  /  CBU: {alias_qr['CBU']}  /  Titular: {alias_qr['Titular']}"),
                       new_x="LMARGIN", new_y="NEXT")
        pdf.image(io.BytesIO(qr_bytes), w=35)
        pdf.set_font("Helvetica", "I", 8)
        if qr_real_bytes:
            pdf.multi_cell(0, 4, "Escaneá el QR con tu app de Mercado Pago/MODO/banco para transferir.",
                           new_x="LMARGIN", new_y="NEXT")
        else:
            pdf.multi_cell(0, 4, "Escaneá el QR para ver el alias/CBU y transferir desde tu banco o billetera virtual.",
                           new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


def generar_pdf_ficha_vehiculo(vehiculo, km_calc, alertas, proyeccion, historial):
    """Genera un PDF con el resumen de la ficha digital del vehículo, para entregarle al cliente."""
    from fpdf import FPDF

    def limpiar(texto):
        return str(texto).encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Equivalencias El Chavo - Ficha del vehiculo", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Fecha: {datetime.now():%d/%m/%Y %H:%M}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, limpiar(f"Patente: {vehiculo['patente']}"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    nombre_auto = f"{vehiculo.get('marca_auto') or ''} {vehiculo.get('modelo_auto') or ''}".strip()
    if nombre_auto:
        pdf.cell(0, 6, limpiar(nombre_auto), new_x="LMARGIN", new_y="NEXT")
    if vehiculo.get("cliente_nombre"):
        pdf.cell(0, 6, limpiar(f"Cliente: {vehiculo['cliente_nombre']}"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, "Kilometraje", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    km_reg = vehiculo.get("km_registro")
    km_act = vehiculo.get("km_actual")
    pdf.cell(0, 6, limpiar(f"Km de registro: {km_reg if km_reg is not None else '-'}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, limpiar(f"Km actual: {km_act if km_act is not None else '-'}"), new_x="LMARGIN", new_y="NEXT")
    if km_calc.get("km_recorridos") is not None:
        pdf.cell(0, 6, limpiar(f"Km recorridos: {km_calc['km_recorridos']:,}"), new_x="LMARGIN", new_y="NEXT")
    if km_calc.get("promedio_mensual") is not None:
        pdf.cell(0, 6, limpiar(f"Promedio aproximado: {km_calc['promedio_mensual']:,} km/mes"),
                  new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    if alertas:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, "Alertas de mantenimiento", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for a in alertas:
            linea = f"- {a['Pieza']} ({a.get('Marca') or 's/marca'}): {a['% consumido']}% de su vida util consumida"
            pdf.multi_cell(0, 6, limpiar(linea), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    if proyeccion:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, "Proyeccion de mantenimiento", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for p in proyeccion:
            linea = (f"- {p['Pieza']}: cambiada {p['Veces cambiada']} vez/veces, "
                      f"deberia {p['Veces que debería (según km)']} - atraso: {p['Atraso estimado']}")
            pdf.multi_cell(0, 6, limpiar(linea), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    if historial:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, "Historial de piezas cambiadas", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for h in historial:
            linea = f"- {h['Pieza']} ({h.get('Marca') or ''}) - {h.get('Fecha') or ''} - {h.get('Km instalación') or '-'} km"
            pdf.multi_cell(0, 6, limpiar(linea), new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


def _puntaje_como_encabezado(fila, siguientes):
    """Qué tan probable es que ESTA fila sea la de títulos de columna.

    No alcanza con mirar la fila sola: un título de tapa ("LISTA DE PRECIOS", "Vigente desde")
    también tiene texto. Lo que distingue al encabezado de verdad es que las filas de ABAJO
    tengan la misma cantidad de columnas llenas que él, y que abajo aparezcan números (precios,
    stock) donde arriba hay palabras."""
    celdas = [c for c in fila if c is not None and str(c).strip() != ""]
    if len(celdas) < 2:
        return -1

    textos = [str(c).strip() for c in celdas]
    puntaje = 0.0

    # Un encabezado son etiquetas cortas, no frases
    if all(len(t) <= 28 for t in textos):
        puntaje += 2
    largo_promedio = sum(len(t) for t in textos) / len(textos)
    if largo_promedio <= 18:
        puntaje += 1

    # Los títulos no se repiten entre sí
    if len(set(t.upper() for t in textos)) == len(textos):
        puntaje += 1.5

    # Casi ningún título es un número suelto
    numericos = sum(1 for t in textos if t.replace(".", "").replace(",", "").isdigit())
    puntaje -= numericos * 1.5

    # Palabras que titulan columnas en cualquier lista de repuestos
    texto_junto = " ".join(textos).upper()
    aciertos = sum(1 for pistas in PISTAS_COLUMNAS.values()
                   for pista in pistas if pista in texto_junto)
    puntaje += min(aciertos, 4) * 1.5

    # Y lo más decisivo: que abajo siga una tabla con la misma forma
    llenas = len(celdas)
    parecidas = 0
    con_numeros = 0
    for sig in siguientes:
        celdas_sig = [c for c in sig if c is not None and str(c).strip() != ""]
        if not celdas_sig:
            continue
        if abs(len(celdas_sig) - llenas) <= 1:
            parecidas += 1
        if any(isinstance(c, (int, float)) or
               str(c).strip().replace(".", "").replace(",", "").isdigit() for c in celdas_sig):
            con_numeros += 1
    puntaje += parecidas * 1.2
    puntaje += min(con_numeros, 3) * 0.8

    # Una fila de tapa suele tener texto largo en la primera celda y nada más
    if len(celdas) <= 2 and largo_promedio > 20:
        puntaje -= 3
    return puntaje


def detectar_fila_encabezado(filas, maximo=25):
    """En qué fila están los títulos de las columnas.

    Hace falta porque las listas casi nunca empiezan en la fila 1: traen el nombre del
    proveedor, el teléfono, la fecha de vigencia, filas vacías. La detección anterior agarraba
    la PRIMERA fila con dos textos, así que en una lista con título y fecha arriba se quedaba
    con 'Vigente desde... / IVA incluido' como si fueran los nombres de las columnas. A partir
    de ahí todo salía mal: el mapeo apuntaba a columnas equivocadas y las filas de arriba
    entraban como si fueran productos.
    """
    if not filas:
        return 0

    # Primero: ¿dónde arrancan los datos? Muchas listas empiezan con filas totalmente vacías.
    primera_con_datos = 0
    for idx, fila in enumerate(filas[:maximo]):
        if any(x is not None and str(x).strip() != "" for x in fila):
            primera_con_datos = idx
            break

    mejor_fila, mejor_puntaje = primera_con_datos, -999
    for idx in range(primera_con_datos, min(len(filas), maximo)):
        p = _puntaje_como_encabezado(filas[idx], filas[idx + 1:idx + 6])
        if p > mejor_puntaje:
            mejor_fila, mejor_puntaje = idx, p

    # Muchísimas listas de proveedor NO TIENEN encabezado: arrancan directo con el primer
    # producto. Sin este control, el detector igual elegía "la fila que más se parece a un
    # encabezado", que terminaba siendo una fila de datos — y ahí se pierde ese producto y se
    # toman sus valores como si fueran los nombres de las columnas.
    #
    # La señal de que NO hay encabezado: la primera fila con datos se parece a las que siguen.
    # Un encabezado real es distinto de sus datos (dice palabras donde abajo hay números).
    if _parece_fila_de_datos(filas[primera_con_datos], filas[primera_con_datos + 1:primera_con_datos + 6]):
        return max(primera_con_datos - 1, 0)

    return mejor_fila if mejor_puntaje > 0 else primera_con_datos


def _parece_fila_de_datos(fila, siguientes):
    """¿Esta fila es un producto más, y no los títulos de las columnas?

    Se compara con las que vienen abajo: si tiene el mismo tipo de contenido en las mismas
    posiciones —número donde abajo hay números, texto donde abajo hay texto— entonces es una
    fila de datos igual a las demás, y esta lista no tiene encabezado."""
    if not siguientes:
        return False

    def perfil(f):
        p = []
        for celda in f:
            if celda is None or str(celda).strip() == "":
                p.append("vacio")
            elif isinstance(celda, (int, float)) or str(celda).strip().replace(".", "").replace(",", "").isdigit():
                p.append("num")
            else:
                p.append("texto")
        return p

    mio = perfil(fila)
    llenas = sum(1 for x in mio if x != "vacio")
    # Tiene que ser una fila de TABLA, con al menos 3 columnas con contenido. Sin este piso, un
    # título de tapa suelto ("ACTUALIZACION DE PRECIOS") se parecía a las otras filas de tapa
    # que venían abajo y la lista se tomaba como si no tuviera encabezado.
    if llenas < 3:
        return False

    iguales = 0
    comparadas = 0
    for sig in siguientes:
        perfil_sig = perfil(sig)
        if sum(1 for x in perfil_sig if x != "vacio") < 3:
            continue
        comparadas += 1
        largo = min(len(mio), len(perfil_sig))
        if largo and sum(1 for i in range(largo) if mio[i] == perfil_sig[i]) / largo >= 0.8:
            iguales += 1
    return comparadas >= 2 and iguales / comparadas >= 0.6


def hojas_del_excel(archivo):
    """Las hojas del archivo con cuántas filas tiene cada una.

    Importa porque antes se leía siempre wb.active —la hoja que quedó abierta cuando el
    proveedor guardó el archivo— y podía ser la de instrucciones o una vacía. Si la lista
    estaba en otra hoja, no se importaba nada y no había forma de darse cuenta."""
    nombre = archivo if isinstance(archivo, str) else getattr(archivo, "name", "")
    if not nombre.lower().endswith((".xlsx", ".xlsm", ".xltx")):
        return []
    try:
        if not isinstance(archivo, str):
            archivo.seek(0)
        # SIN read_only a propósito. En modo read_only, openpyxl devuelve max_row = None para
        # muchos archivos (pasa con las listas reales de proveedor), así que todas las hojas
        # figuraban con 0 filas: el selector no servía para nada y la elección automática de
        # "la hoja con más filas" terminaba tomando la primera por descarte.
        wb = load_workbook(archivo, data_only=True)
        hojas = [(ws.title, ws.max_row or 0) for ws in wb.worksheets]
        wb.close()
        return hojas
    except Exception:
        return []


def _decodificar_texto(crudo):
    """Pasa los bytes de un archivo de texto a string, probando las codificaciones que se usan.

    Antes se decodificaba solo como UTF-8 y, si el archivo venía en otra, la lectura fallaba
    entera con 'No se pudo leer el archivo'. El problema es que Excel en español guarda los CSV
    en Windows-1252, no en UTF-8: cualquier lista con una 'ó' o una 'ñ' —o sea, casi todas—
    era imposible de importar y no había forma de saber por qué.

    El orden importa: primero las que pueden fallar (UTF-8 detecta bytes inválidos), y latin-1
    al final porque acepta cualquier byte y nunca falla, así que si va antes gana siempre y
    deja los acentos mal."""
    for codificacion in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return crudo.decode(codificacion)
        except (UnicodeDecodeError, AttributeError):
            continue
    return crudo.decode("latin-1", errors="replace")


def _detectar_separador(texto):
    """Con qué carácter están separadas las columnas.

    Se mira solo el ENCABEZADO y las primeras filas, no el archivo entero: en una lista larga
    las descripciones traen comas ('FILTRO, ACEITE, FORD') y al contar sobre todo el texto la
    coma le ganaba al separador verdadero. Y se incluye la tabulación, que antes no se
    contemplaba: un archivo separado por tabs quedaba todo en una sola columna."""
    lineas_muestra = [l for l in texto.splitlines()[:15] if l.strip()]
    if not lineas_muestra:
        return ","
    mejor, mejor_puntaje = ",", -1
    for cand in ("\t", ";", ",", "|"):
        cuentas = [l.count(cand) for l in lineas_muestra]
        if not cuentas or max(cuentas) == 0:
            continue
        # Un separador de verdad aparece la MISMA cantidad de veces en todas las filas
        parejas = sum(1 for x in cuentas if x == cuentas[0])
        puntaje = cuentas[0] * 2 + parejas
        if puntaje > mejor_puntaje:
            mejor, mejor_puntaje = cand, puntaje
    return mejor


def leer_excel(archivo, nrows=None, hoja=None):
    """Lee un archivo Excel, CSV o PDF (subido o por ruta) y devuelve una lista de listas (filas)."""
    nombre = archivo if isinstance(archivo, str) else getattr(archivo, "name", "")
    nombre_lower = nombre.lower()
    # Volver al principio del archivo: esta función se llama primero para la vista previa y
    # después para la carga completa. Si el puntero quedó al final de la primera lectura, la
    # segunda leía vacío — de ahí que algunos archivos "se subieran mal" o salieran incompletos.
    if not isinstance(archivo, str):
        try:
            archivo.seek(0)
        except Exception:
            pass

    if nombre_lower.endswith((".csv", ".txt", ".tsv")):
        import csv as csv_module
        if isinstance(archivo, str):
            crudo = open(archivo, "rb").read()
        else:
            archivo.seek(0)
            crudo = archivo.read()
            if isinstance(crudo, str):
                crudo = crudo.encode("utf-8")

        texto = _decodificar_texto(crudo)
        delimitador = _detectar_separador(texto)
        filas = []
        for i, row in enumerate(csv_module.reader(texto.splitlines(), delimiter=delimitador)):
            filas.append(row)
            if nrows and i + 1 >= nrows:
                break
        return filas

    if nombre_lower.endswith(".pdf"):
        import pdfplumber
        filas = []
        with pdfplumber.open(archivo) as pdf:
            for pagina in pdf.pages:
                # extract_tables() en plural: antes se usaba extract_table() (singular), que
                # devuelve SOLO la primera tabla de cada página. En las listas de precios que
                # traen la tabla partida en varios bloques por hoja, se perdía casi todo.
                tablas = pagina.extract_tables() or []
                encontro_algo = False
                for tabla in tablas:
                    for row in tabla:
                        if row and any(celda not in (None, "") for celda in row):
                            filas.append([celda if celda is not None else "" for celda in row])
                            encontro_algo = True
                            if nrows and len(filas) >= nrows:
                                return filas
                if not encontro_algo:
                    # Muchos PDF de proveedor no tienen líneas de tabla: son columnas alineadas
                    # con espacios. Acá NO sirve partir el texto por «dos o más espacios»:
                    # extract_text() colapsa los espacios múltiples en uno solo, así que toda la
                    # fila vuelve como una sola celda y no entra ni un producto. Comprobado
                    # generando un PDF de ese formato: devolvía 0 filas.
                    #
                    # Lo que sí funciona es mirar dónde está cada palabra en la hoja: si entre
                    # el final de una y el comienzo de la siguiente hay un hueco grande, ahí
                    # cambia la columna. Es el mismo criterio que usa el ojo al leerlo.
                    try:
                        palabras = pagina.extract_words() or []
                    except Exception:
                        palabras = []
                    renglones = {}
                    for w in palabras:
                        # Se agrupan por altura, redondeando: los caracteres de una misma línea
                        # nunca están exactamente a la misma altura.
                        clave = round(float(w["top"]) / 3)
                        renglones.setdefault(clave, []).append(w)

                    # Dónde EMPIEZA cada columna, mirando la página entera. Un umbral fijo de
                    # separación no alcanza: cuando una descripción larga llena su columna, el
                    # hueco con el precio es el de un espacio común y quedan pegados. Pero en
                    # estas listas las columnas están alineadas fila a fila, así que las
                    # posiciones donde arrancan las palabras se repiten — y esas repeticiones
                    # son los bordes de las columnas.
                    inicios = {}
                    for w in palabras:
                        x = round(float(w["x0"]) / 4) * 4
                        inicios[x] = inicios.get(x, 0) + 1
                    # Un borde de columna real aparece en CASI TODAS las filas, no en un tercio.
                    # Con el umbral bajo entraban como columna las posiciones donde arrancan
                    # palabras sueltas de la descripción, y la descripción terminaba partida en
                    # cinco pedazos. Se pide el 70% de los renglones.
                    minimo_filas = max(3, int(len(renglones) * 0.7))
                    bordes = sorted(x for x, veces in inicios.items() if veces >= minimo_filas)

                    for clave in sorted(renglones):
                        grupo = sorted(renglones[clave], key=lambda w: float(w["x0"]))
                        if bordes:
                            columnas_fila = {}
                            for w in grupo:
                                # A qué columna pertenece: el borde más cercano a su izquierda
                                x = float(w["x0"])
                                borde = max([b for b in bordes if b <= x + 3], default=bordes[0])
                                columnas_fila.setdefault(borde, []).append(w["text"])
                            celdas = [" ".join(columnas_fila[b]) for b in sorted(columnas_fila)]
                        else:
                            celdas, actual, fin_anterior = [], [], None
                            for w in grupo:
                                if fin_anterior is not None and float(w["x0"]) - fin_anterior > 8:
                                    celdas.append(" ".join(actual))
                                    actual = []
                                actual.append(w["text"])
                                fin_anterior = float(w["x1"])
                            if actual:
                                celdas.append(" ".join(actual))
                        celdas = [x.strip() for x in celdas if x.strip()]
                        if len(celdas) >= 2:
                            filas.append(celdas)
                            if nrows and len(filas) >= nrows:
                                return filas
        return filas

    wb = load_workbook(archivo, data_only=True)
    if hoja and hoja in wb.sheetnames:
        ws = wb[hoja]
    else:
        # Sin hoja elegida, se toma la que MÁS FILAS tiene, no la que quedó activa: la activa
        # es simplemente la que el proveedor tenía abierta al guardar, y muchas veces es la de
        # instrucciones o una en blanco.
        ws = max(wb.worksheets, key=lambda w: w.max_row or 0)
    filas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        filas.append(list(row))
        if nrows and i + 1 >= nrows:
            break
    return filas


# ============================================================
# IDEA 2: FICHA DIGITAL DEL VEHÍCULO (patente + historial de piezas)
# ============================================================
def get_or_create_vehiculo(patente, cliente_nombre="", cliente_telefono="", marca_auto="", modelo_auto="",
                            km_actual=None, anio="", motorizacion="", vin=""):
    patente = patente.strip().upper()
    vin = re.sub(r'\s', '', (vin or "").strip().upper())
    with db_lock:
        c.execute("SELECT id FROM vehiculos WHERE patente = ?", (patente,))
        row = c.fetchone()
        if row:
            vid = row["id"]
            c.execute(
                "UPDATE vehiculos SET "
                "cliente_nombre = COALESCE(NULLIF(?, ''), cliente_nombre), "
                "cliente_telefono = COALESCE(NULLIF(?, ''), cliente_telefono), "
                "marca_auto = COALESCE(NULLIF(?, ''), marca_auto), "
                "modelo_auto = COALESCE(NULLIF(?, ''), modelo_auto), "
                "vin = COALESCE(NULLIF(?, ''), vin), "
                "anio = COALESCE(NULLIF(?, ''), anio), "
                "motorizacion = COALESCE(NULLIF(?, ''), motorizacion), "
                "km_actual = COALESCE(?, km_actual), "
                "km_registro = COALESCE(km_registro, ?), "  # solo se fija si todavía no tenía uno
                "km_actualizado_fecha = CASE WHEN ? IS NOT NULL THEN datetime('now') ELSE km_actualizado_fecha END "
                "WHERE id = ?",
                (cliente_nombre.strip(), cliente_telefono.strip(), marca_auto.strip(), modelo_auto.strip(),
                 vin, anio.strip(), motorizacion.strip(), km_actual, km_actual, km_actual, vid)
            )
        else:
            c.execute(
                "INSERT INTO vehiculos (patente, cliente_nombre, cliente_telefono, marca_auto, modelo_auto, "
                "vin, anio, motorizacion, km_registro, km_actual, km_actualizado_fecha) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))",
                (patente, cliente_nombre.strip(), cliente_telefono.strip(), marca_auto.strip(),
                 modelo_auto.strip(), vin, anio.strip(), motorizacion.strip(), km_actual, km_actual)
            )
            c.execute("SELECT id FROM vehiculos WHERE patente = ?", (patente,))
            vid = c.fetchone()["id"]
        conn.commit()
    # Cada ficha cargada con VIN y modelo enseña el patrón sola. Así el mostrador va llenando
    # la tabla de modelos sin que nadie tenga que sentarse a cargarla: al décimo Gol que pasa,
    # el próximo VIN de Gol se autocompleta.
    if len(vin) == 17 and (modelo_auto.strip() or motorizacion.strip()):
        try:
            aprender_modelo_de_vin(vin, modelo_auto, marca_auto, motorizacion)
        except Exception:
            pass
    return vid


def enseniar_motor_vin(wmi, codigo, motor, notas=None):
    """Guarda que en esta marca, ese carácter en la 8ª posición del VIN es ese motor."""
    wmi, codigo = wmi.strip().upper(), (codigo or "").strip().upper()[:1]
    if not wmi or not codigo or not (motor or "").strip():
        return False
    with db_lock:
        c.execute("""INSERT INTO motores_vin (wmi, codigo, motor, notas) VALUES (?, ?, ?, ?)
                     ON CONFLICT(wmi, codigo) DO UPDATE SET motor = excluded.motor,
                        notas = excluded.notas, veces = veces + 1""",
                  (wmi, codigo, motor.strip(), (notas or "").strip() or None))
        conn.commit()
    return True


def olvidar_motor_vin(wmi, codigo):
    with db_lock:
        c.execute("DELETE FROM motores_vin WHERE wmi = ? AND codigo = ?",
                  (wmi.strip().upper(), (codigo or "").strip().upper()[:1]))
        conn.commit()
    return c.rowcount


def listar_motores_vin():
    c.execute("""SELECT m.wmi AS "WMI", m.codigo AS "8ª posición", m.motor AS "Motor",
                        COALESCE(f.fabricante, '—') AS "Fabricante", m.veces AS "Visto",
                        m.notas AS "Notas"
                 FROM motores_vin m LEFT JOIN fabricantes_vin f ON f.wmi = m.wmi
                 ORDER BY f.fabricante, m.codigo""")
    return filas_a_listas(c)


def aprender_modelo_de_vin(vin, modelo, marca="", motor=""):
    """Guarda los patrones VIN → modelo y VIN → motor a partir de un dato real.

    Nunca pisa lo que ya esté cargado: si el patrón existe, se respeta (puede haberlo corregido
    una persona a mano, y eso vale más que una deducción automática). Devuelve qué aprendió."""
    vin = re.sub(r'\s', '', (vin or "").strip().upper())
    modelo = (modelo or "").strip()
    motor = (motor or "").strip()
    if len(vin) != 17:
        return {"modelo": False, "motor": False}
    wmi, vds = vin[:3], vin[3:8]
    origen = f"aprendido de una ficha de vehículo{f' ({marca.strip()})' if marca.strip() else ''}"
    aprendido = {"modelo": False, "motor": False}

    if modelo:
        c.execute("SELECT 1 FROM modelos_vin WHERE wmi = ? AND vds IN (?, ?, ?)",
                  (wmi, vds, vds[:4], vds[:3]))
        if not c.fetchone():
            with db_lock:
                c.execute("INSERT OR IGNORE INTO modelos_vin (wmi, vds, modelo, motor, notas) "
                          "VALUES (?, ?, ?, ?, ?)",
                          (wmi, vds, modelo, motor or None, origen))
                conn.commit()
            aprendido["modelo"] = True

    if motor:
        codigo = vin[7]
        c.execute("SELECT 1 FROM motores_vin WHERE wmi = ? AND codigo = ?", (wmi, codigo))
        if not c.fetchone():
            with db_lock:
                c.execute("INSERT OR IGNORE INTO motores_vin (wmi, codigo, motor, notas) "
                          "VALUES (?, ?, ?, ?)", (wmi, codigo, motor, origen))
                conn.commit()
            aprendido["motor"] = True
    return aprendido


def buscar_vehiculo_por_vin(vin):
    """Busca una ficha por número de chasis. Si el auto ya pasó por el mostrador, esto da los
    datos REALES (modelo, año, motor, dueño e historial de piezas), no una estimación."""
    vin = re.sub(r'\s', '', (vin or "").strip().upper())
    if len(vin) != 17:
        return None
    c.execute("SELECT * FROM vehiculos WHERE vin = ?", (vin,))
    row = c.fetchone()
    return dict(row) if row else None


def aprender_modelos_de_fichas_existentes():
    """Recorre las fichas que ya tienen VIN y modelo cargados y arma la tabla de patrones de una.
    Sirve para no perder lo que ya está cargado cuando se estrena esta función."""
    c.execute("""SELECT vin, modelo_auto, marca_auto, motorizacion FROM vehiculos
                 WHERE vin IS NOT NULL AND LENGTH(vin) = 17
                   AND ((modelo_auto IS NOT NULL AND modelo_auto <> '')
                        OR (motorizacion IS NOT NULL AND motorizacion <> ''))""")
    modelos = motores = 0
    for fila in c.fetchall():
        r = aprender_modelo_de_vin(fila["vin"], fila["modelo_auto"] or "",
                                    fila["marca_auto"] or "", fila["motorizacion"] or "")
        modelos += 1 if r["modelo"] else 0
        motores += 1 if r["motor"] else 0
    return modelos, motores


def actualizar_km_registro(vehiculo_id, km_registro):
    """Corrige manualmente el km de registro (por si se cargó mal la primera vez)."""
    with db_lock:
        c.execute("UPDATE vehiculos SET km_registro = ? WHERE id = ?", (km_registro, vehiculo_id))
        conn.commit()


def buscar_vehiculo(patente):
    c.execute("SELECT * FROM vehiculos WHERE patente = ?", (patente.strip().upper(),))
    row = c.fetchone()
    return dict(row) if row else None


def calcular_km_recorridos(vehiculo):
    """A partir del km de registro y el km actual, calcula km recorridos y el promedio
    aproximado por mes (usando la fecha de creación de la ficha como punto de partida)."""
    km_registro = vehiculo.get("km_registro")
    km_actual = vehiculo.get("km_actual")
    resultado = {"km_recorridos": None, "promedio_mensual": None, "dias_transcurridos": None}
    if km_registro is None or km_actual is None:
        return resultado
    recorridos = km_actual - km_registro
    if recorridos < 0:
        return resultado
    resultado["km_recorridos"] = recorridos

    creado = vehiculo.get("created_at")
    if creado:
        try:
            fecha_creado = datetime.strptime(creado[:19], "%Y-%m-%d %H:%M:%S")
            dias = max((datetime.now() - fecha_creado).days, 1)
            resultado["dias_transcurridos"] = dias
            resultado["promedio_mensual"] = round(recorridos / dias * 30)
        except (ValueError, TypeError):
            pass
    return resultado


def agregar_pieza_historial(vehiculo_id, descripcion, marca_pieza, codigo_pieza, km_instalacion, vida_util_km, nota):
    with db_lock:
        c.execute(
            "INSERT INTO historial_piezas (vehiculo_id, descripcion_pieza, marca_pieza, codigo_pieza, "
            "km_instalacion, vida_util_km, nota) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (vehiculo_id, descripcion.strip(), marca_pieza.strip(), codigo_pieza.strip(),
             km_instalacion, vida_util_km, nota.strip())
        )
        conn.commit()


def listar_historial_vehiculo(vehiculo_id):
    c.execute("""SELECT id AS "ID", descripcion_pieza AS "Pieza", marca_pieza AS "Marca",
                 codigo_pieza AS "Código", km_instalacion AS "Km instalación",
                 vida_util_km AS "Vida útil (km)", fecha_instalacion AS "Fecha", nota AS "Nota"
                 FROM historial_piezas WHERE vehiculo_id = ? ORDER BY fecha_instalacion DESC""", (vehiculo_id,))
    return filas_a_listas(c)


def calcular_proyeccion_mantenimiento(vehiculo_id, km_recorridos):
    """Para cada tipo de pieza con vida útil cargada, compara cuántas veces se cambió
    realmente contra cuántas veces debería haberse cambiado según los km recorridos totales
    del vehículo desde que se registró."""
    if km_recorridos is None:
        return []
    c.execute("""SELECT descripcion_pieza, COUNT(*) AS veces_reales, AVG(vida_util_km) AS vida_util_prom
                 FROM historial_piezas
                 WHERE vehiculo_id = ? AND vida_util_km IS NOT NULL AND vida_util_km > 0
                 GROUP BY UPPER(descripcion_pieza)""", (vehiculo_id,))
    proyeccion = []
    for row in c.fetchall():
        vida_util_prom = row["vida_util_prom"]
        veces_esperadas = int(km_recorridos // vida_util_prom)
        atraso = veces_esperadas - row["veces_reales"]
        proyeccion.append({
            "Pieza": row["descripcion_pieza"],
            "Vida útil prom. (km)": round(vida_util_prom),
            "Veces cambiada": row["veces_reales"],
            "Veces que debería (según km)": veces_esperadas,
            "Atraso estimado": max(atraso, 0),
        })
    return sorted(proyeccion, key=lambda p: -p["Atraso estimado"])


def listar_vehiculos_atrasados():
    """Recorre todos los vehículos con km cargado y arma un ranking de los que tienen
    mantenimiento atrasado, ordenados por urgencia (el atraso más grande primero)."""
    c.execute("""SELECT id, patente, cliente_nombre, cliente_telefono, marca_auto, modelo_auto,
                 km_registro, km_actual, created_at FROM vehiculos""")
    vehiculos = [dict(r) for r in c.fetchall()]
    resultado = []
    for v in vehiculos:
        km_calc = calcular_km_recorridos(v)
        if km_calc["km_recorridos"] is None:
            continue
        proyeccion = calcular_proyeccion_mantenimiento(v["id"], km_calc["km_recorridos"])
        atrasadas = [p for p in proyeccion if p["Atraso estimado"] > 0]
        if atrasadas:
            resultado.append({
                "vehiculo": v,
                "piezas_atrasadas": atrasadas,
                "atraso_max": max(p["Atraso estimado"] for p in atrasadas),
            })
    resultado.sort(key=lambda r: -r["atraso_max"])
    return resultado


def calcular_alertas_vehiculo(vehiculo_id, km_actual):
    """Piezas que ya recorrieron el 85% o más de su vida útil estimada."""
    c.execute("""SELECT descripcion_pieza, marca_pieza, codigo_pieza, km_instalacion, vida_util_km
                 FROM historial_piezas
                 WHERE vehiculo_id = ? AND vida_util_km IS NOT NULL AND km_instalacion IS NOT NULL""",
              (vehiculo_id,))
    alertas = []
    for row in c.fetchall():
        recorridos = km_actual - row["km_instalacion"]
        if recorridos < 0 or not row["vida_util_km"]:
            continue
        porcentaje = recorridos / row["vida_util_km"]
        if porcentaje >= 0.85:
            alertas.append({
                "Pieza": row["descripcion_pieza"], "Marca": row["marca_pieza"], "Código": row["codigo_pieza"],
                "Km recorridos": recorridos, "Vida útil estimada": row["vida_util_km"],
                "% consumido": round(porcentaje * 100)
            })
    return sorted(alertas, key=lambda a: -a["% consumido"])


# ============================================================
# IDEA 3: SUSTITUCIÓN INTELIGENTE POR MEDIDAS MECÁNICAS
# ============================================================
def buscar_por_medidas(diam_int=None, diam_ext=None, ancho=None, paso_rosca=None, estrias=None, tolerancia_pct=5,
                        estrias_internas=None, estrias_externas=None, posicion_seguro=None, tiene_abs="Cualquiera",
                        diam_int_cara_b=None, diam_ext_cara_b=None,
                        diam_rosca_homocinetica=None, diam_copa=None,
                        diam_copa_superior=None, largo_total=None):
    condiciones = []
    params = []

    def rango(valor, campo):
        if valor:
            tol = valor * tolerancia_pct / 100.0
            condiciones.append(f"p.{campo} BETWEEN ? AND ?")
            params.extend([valor - tol, valor + tol])

    rango(diam_int, "diametro_interno")
    rango(diam_ext, "diametro_externo")
    rango(diam_int_cara_b, "diametro_interno_cara_b")
    rango(diam_ext_cara_b, "diametro_externo_cara_b")
    rango(diam_rosca_homocinetica, "diametro_rosca_homocinetica")
    rango(diam_copa, "diametro_copa")
    rango(diam_copa_superior, "diametro_copa_superior")
    rango(largo_total, "largo_total")
    rango(ancho, "ancho")
    if paso_rosca:
        condiciones.append("UPPER(p.paso_rosca) = ?")
        params.append(paso_rosca.strip().upper())
    if estrias:
        condiciones.append("p.cantidad_estrias = ?")
        params.append(estrias)
    if estrias_internas:
        condiciones.append("p.estrias_internas = ?")
        params.append(estrias_internas)
    if estrias_externas:
        condiciones.append("p.estrias_externas = ?")
        params.append(estrias_externas)
    if posicion_seguro:
        condiciones.append("UPPER(p.posicion_seguro) = ?")
        params.append(posicion_seguro.strip().upper())
    if tiene_abs != "Cualquiera":
        condiciones.append("p.tiene_abs = ?")
        params.append(1 if tiene_abs == "Sí" else 0)

    if not condiciones:
        return []

    query = f"""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", p.descripcion AS "Descripcion", m.nombre AS "Marca",
                p.diametro_interno AS "Diám. interno (cara A)", p.diametro_interno_cara_b AS "Diám. interno (cara B)",
                p.diametro_externo AS "Diám. externo (cara A)", p.diametro_externo_cara_b AS "Diám. externo (cara B)",
                p.diametro_rosca_homocinetica AS "Diám. rosca homocinética",
                p.diametro_copa AS "Diám. copa (base)",
                p.diametro_copa_superior AS "Diám. copa (boca)",
                p.largo_total AS "Largo total",
                p.ancho AS "Ancho",
                p.paso_rosca AS "Paso de rosca", p.cantidad_estrias AS "Estrías",
                p.estrias_internas AS "Estrías internas", p.estrias_externas AS "Estrías externas",
                p.posicion_seguro AS "Posición del seguro",
                CASE WHEN p.tiene_abs = 1 THEN 'Sí' WHEN p.tiene_abs = 0 THEN 'No' ELSE '' END AS "ABS",
                p.precio AS "Precio", p.stock AS "Stock"
                FROM productos p JOIN marcas m ON m.id = p.marca_id
                WHERE {" AND ".join(condiciones)} ORDER BY m.nombre LIMIT 100"""
    c.execute(query, params)
    return filas_a_listas(c)


def actualizar_medidas(producto_id, diam_int, diam_ext, ancho, paso_rosca, estrias, ubicacion,
                        estrias_internas=None, estrias_externas=None, posicion_seguro=None, tiene_abs="Cualquiera",
                        diam_int_cara_b=None, diam_ext_cara_b=None,
                        diam_rosca_homocinetica=None, diam_copa=None,
                        diam_copa_superior=None, largo_total=None):
    tiene_abs_valor = None if tiene_abs == "Cualquiera" else (1 if tiene_abs == "Sí" else 0)
    with db_lock:
        c.execute(
            "UPDATE productos SET diametro_interno=?, diametro_externo=?, ancho=?, paso_rosca=?, "
            "cantidad_estrias=?, ubicacion=?, estrias_internas=?, estrias_externas=?, posicion_seguro=?, "
            "tiene_abs=?, diametro_interno_cara_b=?, diametro_externo_cara_b=?, "
            "diametro_rosca_homocinetica=?, diametro_copa=?, diametro_copa_superior=?, "
            "largo_total=? WHERE id=?",
            (diam_int or None, diam_ext or None, ancho or None, (paso_rosca.strip() or None) if paso_rosca else None,
             estrias or None, (ubicacion.strip() or None) if ubicacion else None,
             estrias_internas or None, estrias_externas or None,
             (posicion_seguro.strip() or None) if posicion_seguro else None, tiene_abs_valor,
             diam_int_cara_b or None, diam_ext_cara_b or None,
             diam_rosca_homocinetica or None, diam_copa or None,
             diam_copa_superior or None, largo_total or None, producto_id)
        )
        conn.commit()


# ============================================================
# COMPARACIÓN VISUAL DE PIEZAS
# ============================================================
# La idea: que puedas sacarle una foto a la pieza en el mostrador y encontrarla en el catálogo
# aunque la foto de referencia sea del sitio del proveedor — otro ángulo, otro fondo, y otro
# color de pieza (la misma pieza de dos marcas viene pintada distinta).
#
# Cómo se aguanta cada una de esas diferencias:
#   fondo   → se recorta la pieza y se tira el resto antes de comparar
#   color   → se compara en blanco y negro, y también contra el negativo de la foto
#   luz     → se empareja el contraste por zonas (CLAHE) en las dos fotos
#   ángulo  → los puntos ORB aguantan giro y escala, y RANSAC verifica que las coincidencias
#             sean geométricamente coherentes (la misma pieza vista distinto) y no casualidad
#   piezas lisas → cuando no hay textura para agarrarse, queda la silueta (momentos de Hu)
#
# Lo que NO puede: un cambio de punto de vista grande (de frente contra de costado) es otra
# imagen para cualquier método de estos. Para eso se cargan varias fotos del mismo producto.

FIRMA_VERSION = 4   # 4 suma los rasgos gruesos de forma (objetos, aspecto, llenado)
MAX_LADO = 640
ORB_FEATURES_CATALOGO = 500
ORB_FEATURES_CONSULTA = 800


def _cv():
    import cv2
    import numpy as np
    return cv2, np


def _recortar_objeto(img, cv2, np):
    """Se queda con la pieza y tira el fondo.

    Por qué importa: la foto del catálogo del proveedor suele estar sobre fondo blanco de
    estudio, y la que sacás vos en el mostrador tiene el mostrador, la caja, la mano. Si no se
    recorta, la mitad de los puntos que se comparan son del fondo, que no tienen nada que ver
    entre una foto y la otra, y el parecido real de la pieza queda tapado.

    Es conservador: si el recorte da un resultado raro (agarra casi toda la foto, o una esquina
    minúscula), se deja la imagen entera. Recortar mal es peor que no recortar."""
    alto, ancho = img.shape[:2]
    area_total = alto * ancho

    suave = cv2.GaussianBlur(img, (5, 5), 0)
    bordes = cv2.Canny(suave, 30, 110)
    bordes = cv2.dilate(bordes, np.ones((7, 7), np.uint8), iterations=2)

    contornos, _ = cv2.findContours(bordes, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contornos:
        return img, None

    mayor = max(contornos, key=cv2.contourArea)
    area = cv2.contourArea(mayor)
    if area < area_total * 0.04 or area > area_total * 0.92:
        return img, mayor  # el recorte no aporta, pero la silueta sí sirve

    x, y, w, h = cv2.boundingRect(mayor)
    margen_x, margen_y = int(w * 0.08), int(h * 0.08)
    x0 = max(x - margen_x, 0)
    y0 = max(y - margen_y, 0)
    x1 = min(x + w + margen_x, ancho)
    y1 = min(y + h + margen_y, alto)
    if (x1 - x0) < 40 or (y1 - y0) < 40:
        return img, mayor
    return img[y0:y1, x0:x1], mayor


def _rasgos_de_forma(img, contorno, cv2, np):
    """Tres datos gruesos de la foto que sirven para descartar rápido, y que los puntos ORB y
    los momentos de Hu no miran.

    El caso que motivó esto: buscando una rótula aparecía un juego de descarbonización. Son
    fotos que no se parecen en nada para una persona — una es UNA pieza compacta y la otra son
    QUINCE juntas planas desparramadas — pero la comparación no tenía forma de notarlo, porque
    ORB no encuentra textura en una rótula lisa y los momentos de Hu solo miran el contorno más
    grande, así que comparaba la rótula contra UNA de las juntas del juego.

      objetos → cuántas piezas separadas hay en la foto. Distingue 'una pieza' de 'un juego'.
      aspecto → qué tan alargada es (siempre ≥1, sin importar si está parada o acostada).
      llenado → cuánto de su recuadro ocupa. Una rótula llena casi todo; una junta plana o un
                soporte fino, mucho menos."""
    rasgos = {"objetos": 1, "aspecto": 1.0, "llenado": 1.0}
    alto, ancho = img.shape[:2]
    area_total = float(alto * ancho)

    try:
        suave = cv2.GaussianBlur(img, (5, 5), 0)
        bordes = cv2.dilate(cv2.Canny(suave, 30, 110), np.ones((7, 7), np.uint8), iterations=2)
        contornos, _ = cv2.findContours(bordes, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # "Significativo" = al menos el 1% de la foto. Sin ese piso, cada mota de polvo del
        # mostrador contaría como una pieza más.
        grandes = [k for k in contornos if cv2.contourArea(k) > area_total * 0.01]
        rasgos["objetos"] = max(len(grandes), 1)
    except Exception:
        pass

    if contorno is not None:
        try:
            (_, _), (w, h), _ = cv2.minAreaRect(contorno)
            if w > 0 and h > 0:
                rasgos["aspecto"] = float(max(w, h) / min(w, h))
            x, y, bw, bh = cv2.boundingRect(contorno)
            if bw > 0 and bh > 0:
                rasgos["llenado"] = float(min(cv2.contourArea(contorno) / (bw * bh), 1.0))
        except Exception:
            pass
    return rasgos


def _penalizacion_por_rasgos(ra, rb):
    """Cuánto se castiga el parecido por diferencias gruesas de forma. Devuelve un multiplicador
    entre 0 y 1. Si alguna de las dos firmas es vieja y no tiene rasgos, no penaliza nada."""
    if not ra or not rb:
        return 1.0
    factor = 1.0

    # Una pieza suelta contra un juego de muchas: es el caso rótula vs descarbonización
    oa, ob = ra.get("objetos", 1), rb.get("objetos", 1)
    if min(oa, ob) <= 2 and max(oa, ob) >= 6:
        factor *= 0.12
    elif abs(oa - ob) >= 4:
        factor *= 0.5

    # Proporción: algo el doble de alargado que lo otro difícilmente sea la misma pieza
    aa, ab = ra.get("aspecto", 1.0), rb.get("aspecto", 1.0)
    razon = max(aa, ab) / max(min(aa, ab), 0.01)
    if razon > 2.5:
        factor *= 0.25
    elif razon > 1.7:
        factor *= 0.6

    # Cuánto llena su recuadro: separa lo macizo de lo plano o calado
    dif = abs(ra.get("llenado", 1.0) - rb.get("llenado", 1.0))
    if dif > 0.35:
        factor *= 0.4
    elif dif > 0.2:
        factor *= 0.75
    return factor


def _firma_de_forma(contorno, cv2, np):
    """Momentos de Hu de la silueta: describen la FORMA sin importar el tamaño, la rotación ni
    el color. Es lo único que queda cuando la pieza es lisa y no tiene textura para agarrarse
    (rótulas, rulemanes, bujes) — ahí los puntos característicos no dan nada."""
    if contorno is None:
        return None
    try:
        hu = cv2.HuMoments(cv2.moments(contorno)).flatten()
        # escala logarítmica: los valores crudos van de 1e-1 a 1e-60 y son incomparables
        return [float(-np.sign(v) * np.log10(abs(v) + 1e-30)) for v in hu]
    except Exception:
        return None


def _preparar(imagen_bytes):
    """Deja la imagen lista para comparar: gris, recortada al objeto, a un tamaño común y con
    el contraste emparejado. Devuelve (imagen, silueta, rasgos) o (None, None, None)."""
    cv2, np = _cv()
    arr = np.frombuffer(imagen_bytes, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_GRAYSCALE)
    if img is None:
        return None, None, None

    escala_previa = 900 / max(img.shape[:2])
    if escala_previa < 1:
        img = cv2.resize(img, None, fx=escala_previa, fy=escala_previa, interpolation=cv2.INTER_AREA)

    # CLAHE ANTES de buscar el objeto: si la foto tiene poco contraste (pieza clara sobre fondo
    # claro, foto velada) sin esto no se detecta ningún borde y el recorte no recorta nada.
    img = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8)).apply(img)

    recortada, contorno = _recortar_objeto(img, cv2, np)
    forma = _firma_de_forma(contorno, cv2, np)
    # Los rasgos se miden sobre la foto ENTERA, antes del recorte: el recorte se queda con la
    # pieza más grande, así que después del recorte un juego de 15 juntas parece una sola.
    rasgos = _rasgos_de_forma(img, contorno, cv2, np)

    escala = MAX_LADO / max(recortada.shape[:2])
    if escala < 1:
        recortada = cv2.resize(recortada, None, fx=escala, fy=escala, interpolation=cv2.INTER_AREA)
    elif escala > 2:
        recortada = cv2.resize(recortada, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)

    # Segunda pasada de contraste, ya sobre el recorte: ahora que el fondo no está, el ajuste
    # se reparte sobre la pieza en vez de gastarse en el mostrador.
    return cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8)).apply(recortada), forma, rasgos


def leer_codigo_de_barras(imagen_bytes):
    """Lee el código de barras o QR de una foto. Devuelve (lista de códigos, error).

    Es la forma más confiable de identificar un repuesto y no depende de nada que se pueda
    equivocar: no hay que reconocer la pieza, ni leer un grabado gastado, ni comparar siluetas.
    El código de la caja es exacto.

    Usa el lector que ya trae OpenCV, que la app instala igual para la comparación de fotos.
    Prueba con la imagen tal cual y también en escala de grises con el contraste emparejado:
    en el mostrador las fotos salen con poca luz y el negro del código se empasta."""
    # OJO: _cv() devuelve (cv2, np), no un tercer valor de error. Asumir lo contrario
    # reventaba con «not enough values to unpack» en la primera línea.
    try:
        cv2, np = _cv()
    except Exception as e:
        return [], f"No está disponible el lector de imágenes: {e}"
    try:
        img = cv2.imdecode(np.frombuffer(imagen_bytes, np.uint8), cv2.IMREAD_COLOR)
        if img is None:
            return [], "No pude leer la imagen."

        alto, ancho = img.shape[:2]
        if max(alto, ancho) > 1600:
            escala = 1600 / max(alto, ancho)
            img = cv2.resize(img, (int(ancho * escala), int(alto * escala)))

        gris = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        versiones = [img, gris,
                     cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8)).apply(gris)]

        encontrados = []
        try:
            lector = cv2.barcode.BarcodeDetector()
            for v in versiones:
                ok, textos, _tipos, _puntos = lector.detectAndDecodeWithType(v)
                if ok:
                    encontrados += [t for t in textos if t and t.strip()]
                if encontrados:
                    break
        except Exception:
            pass

        if not encontrados:
            # Muchos proveedores usan QR en vez de barras
            try:
                qr = cv2.QRCodeDetector()
                for v in versiones:
                    texto, _p, _s = qr.detectAndDecode(v)
                    if texto and texto.strip():
                        encontrados.append(texto.strip())
                        break
            except Exception:
                pass

        if not encontrados:
            return [], ("No encontré ningún código de barras en la foto. Probá más cerca, con "
                        "buena luz y el código derecho, ocupando buena parte de la pantalla.")
        # Sin repetidos, conservando el orden
        vistos, salida = set(), []
        for t in encontrados:
            limpio = t.strip()
            if limpio not in vistos:
                vistos.add(limpio)
                salida.append(limpio)
        return salida, None
    except Exception as e:
        return [], f"No se pudo leer: {type(e).__name__}: {e}"


def buscar_por_codigo_de_barras(codigo_leido):
    """Busca el código escaneado, probando también como código de producto.

    El EAN se carga como un código más del producto (así entró el `CODIGO_EAN` de tu lista de
    juntas), así que buscarlo normalmente ya lo encuentra y trae todas sus equivalencias."""
    clean = sanitizar(codigo_leido)
    if not clean:
        return [], None
    res = buscar_por_codigo(clean)
    if res:
        return res, None
    # Los EAN-13 a veces se cargan sin el dígito verificador, o con un 0 adelante
    for variante in (clean[:-1], clean.lstrip("0"), "0" + clean):
        if variante and variante != clean:
            res = buscar_por_codigo(variante)
            if res:
                return res, f"Se encontró como «{variante}» (variante del código escaneado)."
    return [], None


def calcular_firma_visual(imagen_bytes, es_consulta=False):
    """Saca la 'huella' visual de una foto. Devuelve (blob, estado).

    Guarda los puntos característicos (ORB) Y la silueta. Para la foto que se está buscando
    guarda además la versión en negativo: una misma pieza en negro y en aluminio da gradientes
    invertidos, y ORB compara claro-contra-oscuro, así que sin el negativo no se reconocerían
    entre sí. Se guarda solo del lado de la consulta para no duplicar el peso del catálogo."""
    try:
        cv2, np = _cv()
    except Exception:
        return None, "error"

    try:
        img, forma, rasgos = _preparar(imagen_bytes)
        if img is None:
            return None, "error"

        n = ORB_FEATURES_CONSULTA if es_consulta else ORB_FEATURES_CATALOGO
        orb = cv2.ORB_create(nfeatures=n, scaleFactor=1.2, nlevels=10,
                             edgeThreshold=15, fastThreshold=12)
        kp, desc = orb.detectAndCompute(img, None)

        payload = {
            "v": FIRMA_VERSION,
            "forma": forma,
            "rasgos": rasgos,
            "dim": [int(img.shape[1]), int(img.shape[0])],
            "kp": np.float32([k.pt for k in kp]) if kp else None,
            "desc": desc,
        }

        if es_consulta:
            kp_i, desc_i = orb.detectAndCompute(cv2.bitwise_not(img), None)
            payload["kp_inv"] = np.float32([k.pt for k in kp_i]) if kp_i else None
            payload["desc_inv"] = desc_i

        tiene_textura = desc is not None and len(desc) >= 12
        if not tiene_textura and forma is None:
            return None, "sin_detalle"
        estado = "ok" if tiene_textura else "solo_forma"
        return pickle.dumps(payload, protocol=4), estado
    except Exception:
        return None, "error"


def _cargar_firma(blob):
    """Lee una firma guardada. Acepta las viejas (que eran solo los descriptores sueltos) para
    no tener que rehacer todo el catálogo de golpe."""
    try:
        dato = pickle.loads(blob)
    except Exception:
        return None
    if isinstance(dato, dict) and dato.get("v"):
        return dato
    # Formato viejo: un array de descriptores pelado, sin puntos ni silueta
    return {"v": 1, "desc": dato, "kp": None, "forma": None}


def _parecido_de_forma(forma_a, forma_b):
    """0..100 según cuánto se parecen las siluetas."""
    if not forma_a or not forma_b:
        return 0.0
    try:
        # Solo los 3 primeros momentos, y con peso decreciente. Del cuarto en adelante son ruido
        # puro en fotos reales (el séptimo hasta cambia de signo si la pieza está espejada), y
        # metiéndolos la comparación daba cero siempre, incluso entre dos fotos de la misma pieza.
        pesos = (1.0, 0.7, 0.4)
        d = sum(w * abs(a - b) for w, a, b in zip(pesos, forma_a[:3], forma_b[:3]))
        return float(max(0.0, 100.0 * (1.0 - d / 2.5)))
    except Exception:
        return 0.0


def _puntaje_textura(desc_q, kp_q, desc_c, kp_c, cv2, np):
    """Cuenta coincidencias reales entre dos fotos y las verifica geométricamente.

    El filtro de Lowe saca las coincidencias ambiguas, y RANSAC exige además que todas caigan
    en una misma transformación coherente — o sea, que sean la misma pieza vista distinto, y no
    puntos sueltos que casualmente se parecen. Sin esa verificación, dos piezas metálicas
    cualesquiera dan decenas de 'coincidencias' y todo parece parecido a todo."""
    if desc_q is None or desc_c is None or len(desc_q) < 8 or len(desc_c) < 8:
        return 0.0, 0
    bf = cv2.BFMatcher(cv2.NORM_HAMMING)
    try:
        matches = bf.knnMatch(desc_q, desc_c, k=2)
    except Exception:
        return 0.0, 0

    buenos = [par[0] for par in matches if len(par) == 2 and par[0].distance < 0.78 * par[1].distance]
    if not buenos:
        return 0.0, 0

    inliers = 0
    if len(buenos) >= 8 and kp_q is not None and kp_c is not None:
        try:
            src = np.float32([kp_q[m.queryIdx] for m in buenos]).reshape(-1, 1, 2)
            dst = np.float32([kp_c[m.trainIdx] for m in buenos]).reshape(-1, 1, 2)
            _, mascara = cv2.findHomography(src, dst, cv2.RANSAC, 6.0, maxIters=2000)
            inliers = int(mascara.sum()) if mascara is not None else 0
        except Exception:
            inliers = 0

    # Las verificadas valen; las no verificadas cuentan poco (pueden ser casualidad)
    efectivas = inliers + len(buenos) * 0.15
    denominador = max(min(len(desc_q), len(desc_c)), 1)
    return float(min(100.0, 130.0 * efectivas / denominador)), inliers


def comparar_firmas(firma_consulta, blob_catalogo):
    """Devuelve (parecido 0-100, coincidencias verificadas, parecido de forma 0-100)."""
    try:
        cv2, np = _cv()
    except Exception:
        return 0.0, 0, 0.0

    fc = _cargar_firma(blob_catalogo)
    if not fc:
        return 0.0, 0, 0.0

    mejor_textura, mejor_inliers = 0.0, 0
    versiones = [(firma_consulta.get("desc"), firma_consulta.get("kp"))]
    if firma_consulta.get("desc_inv") is not None:
        versiones.append((firma_consulta.get("desc_inv"), firma_consulta.get("kp_inv")))

    for desc_q, kp_q in versiones:
        p, inl = _puntaje_textura(desc_q, kp_q, fc.get("desc"), fc.get("kp"), cv2, np)
        if p > mejor_textura:
            mejor_textura, mejor_inliers = p, inl

    forma = _parecido_de_forma(firma_consulta.get("forma"), fc.get("forma"))

    factor = _penalizacion_por_rasgos(firma_consulta.get("rasgos"), fc.get("rasgos"))

    if mejor_textura >= 5:
        # Hay textura para agarrarse: manda eso, la forma solo desempata
        total = 0.78 * mejor_textura + 0.22 * forma
    else:
        # Pieza lisa (rótulas, rulemanes, bujes): no hay ningún detalle que verificar, solo la
        # silueta. Sola vale poco, y encima los momentos de Hu comparan un contorno contra otro
        # sin enterarse de si la otra foto es un juego de quince piezas. Por eso acá los rasgos
        # gruesos no descuentan: mandan. Si no coinciden, esto no es un candidato.
        if factor < 0.5:
            return 0.0, 0, round(forma, 1)
        total = 0.35 * forma
    return round(min(total * factor, 100.0), 1), mejor_inliers, round(forma, 1)


def firma_de_consulta(imagen_bytes):
    blob, estado = calcular_firma_visual(imagen_bytes, es_consulta=True)
    if blob is None:
        return None, estado
    return pickle.loads(blob), estado


# ============================================================
# GUARDADO DE FOTOS (varias por producto)
# ============================================================

def agregar_foto_producto(producto_id, imagen_bytes, origen="subida", fuente=None,
                          hacer_principal=None, liviano=False):
    """Suma una foto más al producto. Devuelve (id_foto, estado).

    Un producto puede tener varias: la del catálogo del proveedor, la que sacaste vos, la de
    otra marca del mismo repuesto. Al buscar se compara contra TODAS y se queda con la mejor —
    que es lo único que realmente resuelve el cambio de ángulo, porque ninguna comparación
    reconoce una pieza de frente en una foto de costado."""
    from PIL import Image as PILImage
    import base64
    try:
        img = PILImage.open(io.BytesIO(imagen_bytes)).convert("RGB")
    except Exception:
        return None, "error"
    img.thumbnail((500, 500))
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=82)
    comprimida = buffer.getvalue()

    firma, estado = calcular_firma_visual(comprimida)
    thumb = generar_miniatura(comprimida)

    if liviano and fuente:
        # Modo liviano: se guarda el LINK, no la imagen. La foto de 500px pesa unos 45 KB dentro
        # de la base; la miniatura, 4 KB. Con 8.000 productos eso es la diferencia entre una base
        # de 500 MB (que ya no entra en el backup de GitHub ni sobrevive un reinicio) y una de
        # 200 MB. La imagen grande la muestra el navegador desde el sitio del proveedor.
        data_uri = fuente
    else:
        data_uri = "data:image/jpeg;base64," + base64.b64encode(comprimida).decode("ascii")

    with db_lock:
        c.execute("""INSERT INTO producto_fotos (producto_id, imagen_data, firma_blob, estado,
                                                 origen, fuente, firma_version)
                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                  (producto_id, data_uri, firma, estado, origen, fuente,
                   FIRMA_VERSION if firma else None))
        id_foto = c.lastrowid
        if hacer_principal is None:
            c.execute("SELECT imagen_url FROM productos WHERE id = ?", (producto_id,))
            fila = c.fetchone()
            hacer_principal = not (fila and fila["imagen_url"])
        if hacer_principal:
            c.execute("UPDATE productos SET imagen_url = ?, imagen_thumb = ?, imagen_orb_estado = ? "
                      "WHERE id = ?", (data_uri, thumb, estado, producto_id))
        conn.commit()
    return id_foto, estado


def listar_fotos_producto(producto_id):
    c.execute("""SELECT id, imagen_data, estado, origen, fuente, created_at
                 FROM producto_fotos WHERE producto_id = ? ORDER BY id""", (producto_id,))
    return [dict(r) for r in c.fetchall()]


def eliminar_foto_producto(id_foto):
    """Borra una foto. Si era la que se muestra en el buscador, la reemplaza por otra que quede."""
    c.execute("SELECT producto_id, imagen_data FROM producto_fotos WHERE id = ?", (id_foto,))
    fila = c.fetchone()
    if not fila:
        return False
    producto_id, data = fila["producto_id"], fila["imagen_data"]
    with db_lock:
        c.execute("DELETE FROM producto_fotos WHERE id = ?", (id_foto,))
        c.execute("SELECT imagen_url FROM productos WHERE id = ?", (producto_id,))
        actual = c.fetchone()
        if actual and actual["imagen_url"] == data:
            c.execute("SELECT imagen_data, estado FROM producto_fotos WHERE producto_id = ? "
                      "ORDER BY id LIMIT 1", (producto_id,))
            reemplazo = c.fetchone()
            if reemplazo:
                c.execute("UPDATE productos SET imagen_url = ?, imagen_thumb = ?, imagen_orb_estado = ? "
                          "WHERE id = ?",
                          (reemplazo["imagen_data"], generar_miniatura_de_data_uri(reemplazo["imagen_data"]),
                           reemplazo["estado"], producto_id))
            else:
                c.execute("UPDATE productos SET imagen_url = NULL, imagen_thumb = NULL, "
                          "imagen_orb_estado = NULL WHERE id = ?", (producto_id,))
        conn.commit()
    return True


def generar_miniatura_de_data_uri(data_uri):
    """Miniatura de una foto guardada. Si lo guardado es un link (modo liviano), el link mismo
    hace de miniatura: la baja el navegador."""
    import base64
    if not data_uri:
        return None
    if not data_uri.startswith("data:"):
        return data_uri
    try:
        return generar_miniatura(base64.b64decode(data_uri.split(",", 1)[1]))
    except Exception:
        return None


def contar_fotos_comparables():
    """(fotos listas, productos con al menos una foto lista, fotos sin procesar, fotos que no sirven)"""
    c.execute("SELECT COUNT(*), COUNT(DISTINCT producto_id) FROM producto_fotos WHERE firma_blob IS NOT NULL")
    listas, productos = c.fetchone()
    c.execute("SELECT COUNT(*) FROM producto_fotos WHERE firma_blob IS NULL AND (estado IS NULL OR estado = 'error')")
    pendientes = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM producto_fotos WHERE estado = 'sin_detalle'")
    no_sirven = c.fetchone()[0]
    return listas, productos, pendientes, no_sirven


def buscar_por_similitud_visual(imagen_bytes, top_n=8, minimo=8.0, progreso=None,
                                 familia=None):
    """Compara una foto contra todas las del catálogo y devuelve las más parecidas ordenadas.

    Son candidatos para revisar a mano, NUNCA una identificación confirmada: hay repuestos que
    de foto son idénticos y no son intercambiables (cambia el paso de rosca, la altura, el lado).
    Confirmá siempre por código o comparando la pieza física."""
    firma_consulta, estado = calcular_firma_visual(imagen_bytes, es_consulta=True)
    if firma_consulta is None:
        if estado == "error":
            return None, ("No se pudo procesar esa imagen. Puede estar dañada, o ser un formato que "
                          "el servidor no puede abrir.")
        return None, ("Esa foto no tiene nada de qué agarrarse para comparar: está muy borrosa, muy "
                      "oscura, o la pieza se confunde con el fondo. Probá de nuevo apoyándola sobre "
                      "un fondo liso de otro color, con buena luz y sin que salga movida.")
    import pickle as _pickle
    firma_consulta = _pickle.loads(firma_consulta)

    c.execute("SELECT COUNT(*) FROM producto_fotos WHERE firma_blob IS NOT NULL")
    total_fotos = c.fetchone()[0]
    if not total_fotos:
        return None, _mensaje_catalogo_visual_vacio()

    mejores = {}
    procesadas = 0
    # De a tandas: traer las firmas de todo el catálogo de una sola vez llenaría la memoria del
    # servidor con catálogos grandes y la app se reiniciaría en el medio de la búsqueda.
    c.execute("""SELECT f.id, f.producto_id, f.firma_blob, p.codigo_raw, p.descripcion,
                        p.precio, p.stock, m.nombre AS marca
                 FROM producto_fotos f
                 JOIN productos p ON p.id = f.producto_id
                 JOIN marcas m ON m.id = p.marca_id
                 WHERE f.firma_blob IS NOT NULL""")
    while True:
        tanda = c.fetchmany(200)
        if not tanda:
            break
        for fila in tanda:
            procesadas += 1
            if progreso and procesadas % 25 == 0:
                progreso(procesadas, total_fotos)
            # Filtro por tipo de pieza. Es lo que más ayuda y no sale de la foto: vos ya sabés
            # que estás buscando una rótula, así que no tiene sentido que la app te ofrezca un
            # juego de juntas. Descarta antes de comparar, así además va más rápido.
            if familia and clasificar_repuesto(fila["descripcion"]) != familia:
                continue
            try:
                puntaje, verificadas, forma = comparar_firmas(firma_consulta, fila["firma_blob"])
            except Exception:
                continue
            pid = fila["producto_id"]
            if pid not in mejores or puntaje > mejores[pid]["Parecido"]:
                mejores[pid] = {
                    "ID": pid,
                    "Codigo": fila["codigo_raw"],
                    "Descripcion": fila["descripcion"],
                    "Marca": fila["marca"],
                    "Precio": fila["precio"],
                    "Stock": fila["stock"],
                    "Parecido": puntaje,
                    "Coincidencias": verificadas,
                    "Forma": forma,
                }

    resultados = [r for r in mejores.values() if r["Parecido"] >= minimo]
    resultados.sort(key=lambda r: (-r["Parecido"], -r["Coincidencias"]))
    if not resultados:
        return [], ("Ninguna foto del catálogo se parece lo suficiente. Puede ser que el producto no "
                    "esté cargado con foto, que la foto de referencia sea de un ángulo demasiado "
                    "distinto (cargale a ese producto una segunda foto del ángulo que usás vos), o "
                    "que la pieza sea lisa y sin marcas: en rótulas, rulemanes y bujes no hay "
                    "ningún detalle para agarrarse y esta búsqueda casi nunca sirve. Para esos "
                    "casos anda mucho mejor **📐 Buscar por medidas mecánicas**.")
    return resultados[:top_n], None


def _mensaje_catalogo_visual_vacio():
    c.execute("SELECT COUNT(*) FROM producto_fotos")
    fotos = c.fetchone()[0]
    if fotos:
        return (f"Hay {fotos} foto(s) en el catálogo pero ninguna procesada todavía. Tocá "
                "«🔄 Procesar fotos pendientes» acá arriba y volvé a intentar.")
    return ("Todavía no hay ninguna foto en el catálogo para comparar. Se cargan desde "
            "Administrar → Medidas y fotos (subiendo la foto o pegando la dirección de la ficha "
            "del proveedor), o en tanda desde Estadísticas → Mantenimiento. "
            "Ojo: el *backup sin fotos* que se sube al repositorio no las lleva, así que después "
            "de un reinicio del hosting hay que volver a cargarlas.")


def nivel_de_parecido(fila):
    """Traduce el puntaje a algo que se pueda leer de un vistazo, sin dar falsa seguridad.

    Lo que manda son las coincidencias VERIFICADAS, no el porcentaje. Sin ninguna, el parecido
    salió solo de la silueta, y una silueta parecida no significa casi nada entre repuestos:
    antes eso podía llegar a mostrarse como «Media — mirala bien», que era prometer de más."""
    if fila["Coincidencias"] >= 25 and fila["Parecido"] >= 30:
        return "🟢 Fuerte — muchos detalles coinciden"
    if fila["Coincidencias"] >= 8:
        return "🟡 Media — mirala bien"
    if fila["Coincidencias"] >= 3:
        return "🟠 Floja — pocos detalles verificados"
    return "🔴 Solo la silueta — no confíes en esto"


def generar_miniatura(imagen_bytes, lado=110):
    """Versión chiquita de la foto, para las listas de resultados. La foto normal (400px) pesa
    unas 10 veces más y antes la búsqueda la traía entera por cada resultado, aunque en la tabla
    se vea del tamaño de una uña — con muchos resultados eso se nota, sobre todo en el celular."""
    from PIL import Image as PILImage
    import base64
    try:
        img = PILImage.open(io.BytesIO(imagen_bytes)).convert("RGB")
        img.thumbnail((lado, lado))
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=70)
        return "data:image/jpeg;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")
    except Exception:
        return None


def actualizar_imagen_producto(producto_id, imagen_bytes, origen="subida", fuente=None,
                                liviano=False):
    """Guarda una foto del producto. Devuelve True si además quedó lista para la búsqueda visual.
    Se conserva el nombre de antes porque lo usan las bajadas en tanda."""
    _, estado = agregar_foto_producto(producto_id, imagen_bytes, origen=origen, fuente=fuente,
                                       liviano=liviano)
    return estado == "ok"


def eliminar_imagen_producto(producto_id):
    """Saca TODAS las fotos del producto."""
    with db_lock:
        c.execute("DELETE FROM producto_fotos WHERE producto_id = ?", (producto_id,))
        c.execute("UPDATE productos SET imagen_url = NULL, imagen_thumb = NULL, imagen_orb_blob = NULL, "
                   "imagen_orb_estado = NULL WHERE id = ?", (producto_id,))
        conn.commit()


def imagenes_de_una_direccion(url, maximo=8):
    """Trae las fotos que haya en una dirección web. Devuelve (lista de (url, bytes), error).

    Si la dirección apunta directo a una imagen, baja esa. Si es la página de un producto, saca
    las fotos de adentro: mira la imagen que la página declara como principal (og:image, la que
    usan WhatsApp y Facebook para la vista previa) y después las <img> normales, salteando
    logos, íconos y banners."""
    import requests
    from urllib.parse import urljoin

    try:
        cabecera = requests.head(url, timeout=8, allow_redirects=True,
                                 headers={"User-Agent": "Mozilla/5.0 (compatible; EquivalenciasElChavo/1.0)"})
        tipo = cabecera.headers.get("Content-Type", "")
    except Exception:
        tipo = ""

    if "image" in tipo.lower() or url.lower().split("?")[0].endswith((".jpg", ".jpeg", ".png", ".webp")):
        datos, error = descargar_imagen(url)
        return ([(url, datos)], None) if datos else (None, error or "no se pudo bajar esa imagen")

    try:
        respuesta = requests.get(url, timeout=15,
                                 headers={"User-Agent": "Mozilla/5.0 (compatible; EquivalenciasElChavo/1.0)"})
        if respuesta.status_code != 200:
            return None, f"la página respondió {respuesta.status_code}"
        html = respuesta.text
    except Exception as e:
        return None, f"no se pudo abrir la página ({type(e).__name__})"

    candidatas, vistas = [], set()

    def sumar(src):
        if not src or src.startswith("data:"):
            return
        if any(x in src.lower() for x in ("logo", "icon", "sprite", "banner", "pixel", "avatar",
                                          "placeholder", "loading", ".svg", ".gif")):
            return
        absoluta = urljoin(url, src)
        if absoluta not in vistas:
            vistas.add(absoluta)
            candidatas.append(absoluta)

    for m in re.finditer(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)', html, re.I):
        sumar(m.group(1))
    for m in re.finditer(r'<img[^>]+?(?:data-src|data-original|src)=["\']([^"\']+)["\']', html, re.I):
        sumar(m.group(1))

    encontradas, errores = [], []
    for url_img in candidatas[:maximo * 3]:
        if len(encontradas) >= maximo:
            break
        datos, error = descargar_imagen(url_img)
        if datos and len(datos) > 8000:   # los íconos chiquitos no son fotos de producto
            encontradas.append((url_img, datos))
        elif error:
            errores.append(error)
    if encontradas:
        return encontradas, None
    return None, "no encontré ninguna foto de producto en esa página"


def migrar_imagenes_pendientes(limite=400):
    """Pone al día las firmas visuales: pasa las fotos viejas (que estaban sueltas en la ficha del
    producto) a la tabla de fotos, y calcula la firma de las que todavía no la tienen.
    Se saltea lo que ya se intentó y no dio, así el contador de pendientes no queda clavado."""
    import base64
    resumen = {"listas": 0, "sin_detalle": 0, "error": 0, "links": 0, "migradas": 0}

    # 1) Fotos que están en la ficha del producto pero todavía no en la tabla de fotos
    c.execute("""SELECT p.id, p.imagen_url FROM productos p
                 WHERE p.imagen_url IS NOT NULL AND p.imagen_url LIKE 'data:%'
                   AND NOT EXISTS (SELECT 1 FROM producto_fotos f WHERE f.producto_id = p.id)
                 LIMIT ?""", (limite,))
    for fila in c.fetchall():
        try:
            datos = base64.b64decode(fila["imagen_url"].split(",", 1)[1])
            firma, estado = calcular_firma_visual(datos)
            with db_lock:
                c.execute("""INSERT INTO producto_fotos (producto_id, imagen_data, firma_blob, estado,
                                                         origen, firma_version)
                             VALUES (?, ?, ?, ?, 'migrada', ?)""",
                          (fila["id"], fila["imagen_url"], firma, estado,
                           FIRMA_VERSION if firma else None))
                c.execute("UPDATE productos SET imagen_orb_estado = ? WHERE id = ?", (estado, fila["id"]))
                conn.commit()
            resumen["migradas"] += 1
            resumen["listas" if estado == "ok" else ("sin_detalle" if estado == "sin_detalle" else "error")] += 1
        except Exception:
            resumen["error"] += 1

    # 2) Fotos ya en la tabla pero sin firma calculada
    c.execute("""SELECT id, producto_id, imagen_data FROM producto_fotos
                 WHERE (
                        (firma_blob IS NULL AND (estado IS NULL OR estado = 'error'))
                        OR COALESCE(firma_version, 0) < ?
                       )
                   AND imagen_data LIKE 'data:%' LIMIT ?""", (FIRMA_VERSION, limite))
    for fila in c.fetchall():
        try:
            datos = base64.b64decode(fila["imagen_data"].split(",", 1)[1])
            firma, estado = calcular_firma_visual(datos)
            with db_lock:
                c.execute("UPDATE producto_fotos SET firma_blob = ?, estado = ?, firma_version = ? "
                          "WHERE id = ?",
                          (firma, estado, FIRMA_VERSION if firma else None, fila["id"]))
                conn.commit()
            resumen["listas" if estado == "ok" else ("sin_detalle" if estado == "sin_detalle" else "error")] += 1
        except Exception:
            resumen["error"] += 1

    # 3) Miniaturas faltantes y fotos que son link externo (esas hay que bajarlas desde Mantenimiento)
    c.execute("""SELECT id, imagen_url FROM productos
                 WHERE imagen_url IS NOT NULL AND imagen_thumb IS NULL LIMIT ?""", (limite,))
    for fila in c.fetchall():
        url = fila["imagen_url"]
        try:
            if url.startswith("data:"):
                thumb = generar_miniatura_de_data_uri(url)
            else:
                thumb = url          # link externo: lo carga el navegador
                resumen["links"] += 1
            if thumb:
                with db_lock:
                    c.execute("UPDATE productos SET imagen_thumb = ? WHERE id = ?", (thumb, fila["id"]))
                    conn.commit()
        except Exception:
            continue
    return resumen


def contar_fotos_pendientes_de_firma():
    c.execute("""SELECT (SELECT COUNT(*) FROM productos p
                         WHERE p.imagen_url LIKE 'data:%'
                           AND NOT EXISTS (SELECT 1 FROM producto_fotos f WHERE f.producto_id = p.id))
                      + (SELECT COUNT(*) FROM producto_fotos
                         WHERE (firma_blob IS NULL AND (estado IS NULL OR estado = 'error'))
                            OR COALESCE(firma_version, 0) < ?)""", (FIRMA_VERSION,))
    return c.fetchone()[0]


@st.cache_resource
def _ejecutar_migracion_orb_una_vez():
    """Corre migrar_imagenes_pendientes() una sola vez por proceso (no en cada rerun de
    Streamlit), con el mismo patrón que ya usa get_connection() para la conexión a la base."""
    return migrar_imagenes_pendientes()


_ejecutar_migracion_orb_una_vez()


# ============================================================
# IDEA 5: AUDITORÍA PREVENTIVA POR MUESTREO ALEATORIO
# ============================================================
def generar_auditoria_hoy(cantidad=8):
    """Genera (si no existe todavía) la muestra aleatoria de hoy, priorizando favoritos y productos con precio cargado."""
    hoy = datetime.now().strftime("%Y-%m-%d")
    with db_lock:
        c.execute("SELECT COUNT(*) FROM auditoria_diaria WHERE fecha = ?", (hoy,))
        if c.fetchone()[0] > 0:
            return False
        c.execute(
            "SELECT id, stock FROM productos WHERE favorito = 1 OR precio IS NOT NULL ORDER BY RANDOM() LIMIT ?",
            (cantidad,)
        )
        elegidos = c.fetchall()
        for row in elegidos:
            c.execute(
                "INSERT OR IGNORE INTO auditoria_diaria (fecha, producto_id, stock_sistema) VALUES (?, ?, ?)",
                (hoy, row["id"], row["stock"])
            )
        conn.commit()
    return True


def listar_auditoria_hoy():
    hoy = datetime.now().strftime("%Y-%m-%d")
    c.execute("""SELECT a.id AS "ID_auditoria", p.codigo_raw AS "Codigo", m.nombre AS "Marca",
                 a.stock_sistema AS "Stock sistema", a.stock_contado AS "Stock contado",
                 a.diferencia AS "Diferencia", a.resuelto AS "Resuelto"
                 FROM auditoria_diaria a JOIN productos p ON p.id = a.producto_id JOIN marcas m ON m.id = p.marca_id
                 WHERE a.fecha = ? ORDER BY a.resuelto ASC, p.codigo_raw""", (hoy,))
    return filas_a_listas(c)


def registrar_conteo_auditoria(auditoria_id, stock_contado):
    with db_lock:
        c.execute("SELECT stock_sistema FROM auditoria_diaria WHERE id = ?", (auditoria_id,))
        row = c.fetchone()
        diferencia = stock_contado - (row["stock_sistema"] or 0)
        c.execute(
            "UPDATE auditoria_diaria SET stock_contado=?, diferencia=?, resuelto=1 WHERE id=?",
            (stock_contado, diferencia, auditoria_id)
        )
        conn.commit()


# ============================================================
# IDEA 6: UBICACIÓN INTELIGENTE EN DEPÓSITO (matriz ABC)
# ============================================================
def calcular_matriz_abc(limite=300):
    """Clasifica productos en A/B/C usando la frecuencia de búsqueda como indicador de rotación
    (no hay módulo de ventas en la app, así que esto es una aproximación de demanda)."""
    c.execute("""SELECT p.id AS "ID", p.codigo_raw AS "Codigo", m.nombre AS "Marca",
                 p.ubicacion AS "Ubicación", COALESCE(p.veces_buscado, 0) AS "Veces buscado"
                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                 WHERE COALESCE(p.veces_buscado, 0) > 0 OR p.favorito = 1
                 ORDER BY COALESCE(p.veces_buscado, 0) DESC LIMIT ?""", (limite,))
    filas = filas_a_listas(c)
    total = len(filas)
    for i, f in enumerate(filas):
        if total <= 1 or i < max(1, round(total * 0.2)):
            f["Categoría"] = "A"
            f["Sugerencia"] = "Cerca de la entrada, a la altura de la cintura"
        elif i < round(total * 0.5):
            f["Categoría"] = "B"
            f["Sugerencia"] = "Zona intermedia"
        else:
            f["Categoría"] = "C"
            f["Sugerencia"] = "Estante superior o trasero"
    return filas


# ============================================================
# MODO MECÁNICO — DICCIONARIO DE CÓDIGOS OBD2 / DTC
# ============================================================
def buscar_dtc(codigo, fabricante_filtro="Todos"):
    codigo = codigo.strip().upper()
    query = """SELECT codigo AS "Código",
               CASE WHEN fabricante = '' THEN 'Genérico' ELSE fabricante END AS "Fabricante",
               descripcion AS "Descripción", sistema AS "Sistema",
               causas_posibles AS "Causas posibles" FROM codigos_dtc
               WHERE (codigo = ? OR codigo LIKE ?)"""
    params = [codigo, f"%{codigo}%"]
    if fabricante_filtro == "Genérico":
        query += " AND fabricante = ''"
    elif fabricante_filtro and fabricante_filtro != "Todos":
        query += " AND fabricante = ?"
        params.append(fabricante_filtro)
    query += " ORDER BY fabricante, codigo"
    c.execute(query, params)
    return filas_a_listas(c)


def agregar_dtc(codigo, descripcion, sistema, causas, fabricante=""):
    codigo = codigo.strip().upper()
    fabricante = fabricante.strip()
    with db_lock:
        c.execute(
            "INSERT INTO codigos_dtc (codigo, fabricante, descripcion, sistema, causas_posibles) "
            "VALUES (?, ?, ?, ?, ?) "
            "ON CONFLICT(codigo, fabricante) DO UPDATE SET descripcion=excluded.descripcion, "
            "sistema=excluded.sistema, causas_posibles=excluded.causas_posibles",
            (codigo, fabricante, descripcion.strip(), sistema.strip(), causas.strip())
        )
        conn.commit()


def importar_dtc_masivo(texto):
    """Importa códigos DTC pegados como texto, una línea por código:
    codigo;descripcion;sistema;causas;fabricante (fabricante es opcional, vacío = código genérico)."""
    cargados = 0
    with db_lock:
        for linea in texto.strip().splitlines():
            partes = [p.strip() for p in linea.split(";")]
            if len(partes) < 2 or not partes[0]:
                continue
            codigo = partes[0].upper()
            descripcion = partes[1]
            sistema = partes[2] if len(partes) > 2 else ""
            causas = partes[3] if len(partes) > 3 else ""
            fabricante = partes[4] if len(partes) > 4 else ""
            c.execute(
                "INSERT INTO codigos_dtc (codigo, fabricante, descripcion, sistema, causas_posibles) "
                "VALUES (?, ?, ?, ?, ?) "
                "ON CONFLICT(codigo, fabricante) DO UPDATE SET descripcion=excluded.descripcion, "
                "sistema=excluded.sistema, causas_posibles=excluded.causas_posibles",
                (codigo, fabricante, descripcion, sistema, causas)
            )
            cargados += 1
        conn.commit()
    return cargados


def contar_dtc():
    c.execute("SELECT COUNT(*) FROM codigos_dtc")
    return c.fetchone()[0]


def listar_fabricantes_dtc():
    c.execute("SELECT DISTINCT fabricante FROM codigos_dtc WHERE fabricante != '' ORDER BY fabricante")
    return [r["fabricante"] for r in c.fetchall()]


# ============================================================
# MODO MECÁNICO — LECTOR DE VIN
# ============================================================
# Primer carácter del VIN = región/país de fabricación (estándar ISO 3779, dato genérico).
PAISES_VIN = {
    # Primer carácter del VIN = región. Es lo más grueso; abajo se afina con dos caracteres.
    "1": "Estados Unidos", "4": "Estados Unidos", "5": "Estados Unidos",
    "2": "Canadá", "3": "México / Centroamérica",
    "6": "Australia / Oceanía", "7": "Nueva Zelanda / Oceanía",
    "8": "Sudamérica", "9": "Brasil / Sudamérica", "0": "Sudamérica",
    "A": "África", "B": "África", "C": "África", "D": "África", "E": "África",
    "F": "África", "G": "África", "H": "África",
    "J": "Japón", "K": "Corea del Sur", "L": "China", "M": "India / Asia del Sur",
    "N": "Turquía / Asia occidental", "P": "Filipinas / Asia", "R": "Taiwán / Asia",
    "S": "Reino Unido", "T": "Europa central", "U": "Europa del este",
    "V": "Francia / España", "W": "Alemania", "X": "Rusia / Europa del este",
    "Y": "Suecia / Finlandia", "Z": "Italia",
}
# Con los DOS primeros caracteres se distingue mucho mejor: "8" solo dice Sudamérica, pero
# "8A" es Argentina y "8B" Chile. Se usa esto primero y, si no está, se cae al de una letra.
PAISES_VIN_2 = {
    # --- Sudamérica ---
    "8A": "Argentina", "8B": "Argentina", "8C": "Argentina", "8D": "Argentina",
    "8E": "Argentina",
    "8F": "Chile", "8G": "Chile", "8H": "Chile", "8J": "Chile",
    "8K": "Ecuador", "8L": "Ecuador", "8M": "Ecuador",
    "8S": "Perú", "8T": "Perú",
    "8X": "Venezuela", "8Y": "Venezuela", "8Z": "Venezuela",
    "9A": "Brasil", "9B": "Brasil", "9C": "Brasil", "9D": "Brasil", "9E": "Brasil",
    "93": "Brasil", "94": "Brasil", "95": "Brasil", "96": "Brasil", "97": "Brasil",
    "98": "Brasil", "99": "Brasil",
    "9F": "Colombia", "9G": "Colombia", "9H": "Colombia",
    "9L": "Paraguay", "9M": "Paraguay",
    "9U": "Uruguay", "9V": "Uruguay",
    "9X": "Venezuela",
    # --- Norteamérica ---
    "1G": "Estados Unidos", "1F": "Estados Unidos", "2F": "Canadá", "2G": "Canadá",
    "3F": "México", "3G": "México", "3N": "México", "3V": "México",
    # --- Europa ---
    "SA": "Reino Unido", "SB": "Reino Unido", "SC": "Reino Unido", "SD": "Reino Unido",
    "SE": "Reino Unido", "SF": "Reino Unido", "SH": "Reino Unido", "SJ": "Reino Unido",
    "SK": "Reino Unido", "SL": "Reino Unido", "SM": "Reino Unido",
    "SU": "Polonia", "SN": "Alemania (este)",
    "TM": "República Checa", "TN": "República Checa", "TR": "Hungría", "TS": "Hungría",
    "TW": "Portugal", "TY": "Portugal", "TC": "Suiza", "TD": "Suiza",
    "UU": "Rumania", "U5": "Eslovaquia", "U6": "Eslovaquia",
    "VA": "Austria", "VN": "Francia", "VF": "Francia", "VG": "Francia", "VL": "Francia",
    "VS": "España", "VT": "España", "VV": "España", "VW": "España", "VX": "Serbia",
    "WA": "Alemania", "WB": "Alemania", "WD": "Alemania", "WE": "Alemania",
    "WF": "Alemania", "WJ": "Alemania", "WM": "Alemania", "WP": "Alemania",
    "WU": "Alemania", "WV": "Alemania", "W0": "Alemania",
    "XL": "Países Bajos", "XM": "Países Bajos", "XT": "Rusia", "XU": "Rusia",
    "XW": "Rusia / Uzbekistán", "X4": "Rusia", "X7": "Rusia",
    "YB": "Bélgica", "YC": "Bélgica", "YE": "Bélgica",
    "YK": "Finlandia", "YS": "Suecia", "YT": "Suecia", "YV": "Suecia", "Y6": "Ucrania",
    "ZA": "Italia", "ZB": "Italia", "ZC": "Italia", "ZD": "Italia", "ZF": "Italia",
    "ZG": "Italia", "ZH": "Italia", "ZJ": "Italia", "ZK": "Italia", "ZL": "Italia",
    "ZO": "Italia",
    # --- Asia ---
    "JA": "Japón", "JD": "Japón", "JF": "Japón", "JH": "Japón", "JK": "Japón",
    "JL": "Japón", "JM": "Japón", "JN": "Japón", "JS": "Japón", "JT": "Japón",
    "JY": "Japón",
    "KL": "Corea del Sur", "KM": "Corea del Sur", "KN": "Corea del Sur",
    "KP": "Corea del Sur",
    "LA": "China", "LB": "China", "LC": "China", "LD": "China", "LE": "China",
    "LF": "China", "LG": "China", "LH": "China", "LJ": "China", "LK": "China",
    "LL": "China", "LM": "China", "LP": "China", "LS": "China", "LT": "China",
    "LU": "China", "LV": "China", "LZ": "China", "L4": "China", "L5": "China",
    "MA": "India", "MB": "India", "MC": "India", "MD": "India", "ME": "India",
    "MH": "Indonesia", "MJ": "Indonesia", "ML": "Tailandia", "MM": "Tailandia",
    "MN": "Tailandia", "MP": "Tailandia", "MR": "Tailandia",
    "NL": "Turquía", "NM": "Turquía", "NA": "Irán", "NC": "Turquía",
    "PE": "Filipinas", "PL": "Malasia", "PN": "Malasia", "PP": "Singapur",
    "RF": "Taiwán", "RA": "Taiwán", "RL": "Taiwán",
    # --- África y Oceanía ---
    "AA": "Sudáfrica", "AC": "Sudáfrica", "AD": "Sudáfrica", "AF": "Sudáfrica",
    "AH": "Sudáfrica",
    "6A": "Australia", "6F": "Australia", "6G": "Australia", "6H": "Australia",
    "6M": "Australia", "6T": "Australia", "6U": "Australia",
    "7A": "Nueva Zelanda",
}
# Código de año en la 10ª posición del VIN (estándar, cíclico cada 30 años).
ANIOS_VIN = {
    "A": 1980, "B": 1981, "C": 1982, "D": 1983, "E": 1984, "F": 1985, "G": 1986, "H": 1987,
    "J": 1988, "K": 1989, "L": 1990, "M": 1991, "N": 1992, "P": 1993, "R": 1994, "S": 1995,
    "T": 1996, "V": 1997, "W": 1998, "X": 1999, "Y": 2000,
    "1": 2001, "2": 2002, "3": 2003, "4": 2004, "5": 2005, "6": 2006, "7": 2007, "8": 2008, "9": 2009,
}


def decodificar_vin(vin):
    vin = re.sub(r'\s', '', vin.strip().upper())
    resultado = {"vin": vin, "valido": False}
    if len(vin) != 17:
        resultado["error"] = "El VIN debe tener 17 caracteres."
        return resultado
    if any(ch in vin for ch in ("I", "O", "Q")):
        resultado["error"] = "El VIN no puede contener las letras I, O ni Q."
        return resultado

    resultado["valido"] = True
    wmi = vin[:3]
    resultado["wmi"] = wmi
    resultado["pais"] = PAISES_VIN_2.get(vin[:2]) or PAISES_VIN.get(vin[0], "Desconocido / no cargado")

    # Primero el WMI exacto de 3; si no está, el prefijo de 2. El orden importa y no es un
    # detalle: 'JA' es Isuzu pero 'JA3' es Mitsubishi, y '1H' es Honda mientras que '1HD' es
    # Harley-Davidson. Si se buscara el prefijo primero, esos casos darían la marca equivocada.
    resultado["fabricante"] = None
    resultado["fabricante_por_prefijo"] = False
    for clave, es_prefijo in ((wmi, False), (wmi[:2], True)):
        c.execute("SELECT fabricante, pais FROM fabricantes_vin WHERE wmi = ?", (clave,))
        fila = c.fetchone()
        if fila:
            resultado["fabricante"] = fila["fabricante"]
            resultado["fabricante_por_prefijo"] = es_prefijo
            if fila["pais"]:
                resultado["pais"] = fila["pais"]
            break

    # --- Modelo: NO sale del VIN, sale de lo que se haya enseñado ---
    # Las posiciones 4-8 (VDS) las define cada fabricante a su gusto, no hay norma que diga qué
    # significan. Así que se busca el patrón exacto y, si no está, uno más corto (4-7 y 4-6):
    # muchos fabricantes usan los primeros caracteres para la carrocería y los últimos para el
    # motor o el equipamiento, así que un patrón corto suele acertar la familia del modelo.
    vds = vin[3:8]
    resultado["vds"] = vds
    resultado["modelo"] = None
    resultado["modelo_exacto"] = False
    resultado["motor"] = None
    resultado["motor_origen"] = None
    for largo in (5, 4, 3):
        c.execute("SELECT vds, modelo, motor, notas FROM modelos_vin WHERE wmi = ? AND vds = ?",
                  (wmi, vds[:largo]))
        m = c.fetchone()
        if m:
            resultado["modelo"] = m["modelo"]
            resultado["modelo_notas"] = m["notas"]
            resultado["modelo_exacto"] = (largo == 5)
            if m["motor"]:
                resultado["motor"] = m["motor"]
                resultado["motor_origen"] = ("patrón exacto de este modelo" if largo == 5
                                              else "patrón de la familia del modelo")
            break

    # --- Motor por la 8ª posición ---
    # Es la posición que la norma de Norteamérica reserva para el código de motor, y que casi
    # todos los fabricantes usan para lo mismo aunque afuera no sea obligatorio. Generaliza
    # mejor que el modelo: el mismo código suele repetirse en toda la gama de la marca.
    # Solo se usa si el patrón del modelo no trajo ya un motor, que es más específico.
    codigo_motor = vin[7]
    resultado["codigo_motor"] = codigo_motor
    resultado["motor_por_norma"] = vin[0] in "12345"
    if not resultado["motor"]:
        c.execute("SELECT motor, notas FROM motores_vin WHERE wmi = ? AND codigo = ?",
                  (wmi, codigo_motor))
        mm = c.fetchone()
        if mm:
            resultado["motor"] = mm["motor"]
            resultado["motor_notas"] = mm["notas"]
            resultado["motor_origen"] = f"código «{codigo_motor}» en la 8ª posición"

    # --- Año: se puede ser preciso en los VIN de Norteamérica ---
    # El código de la 10ª posición se repite cada 30 años (la "D" sirve para 1983 y 2013).
    # Pero en los VIN emitidos para Norteamérica (país 1 a 5) hay una regla que lo desambigua:
    # si el 7° carácter es LETRA, el vehículo es del ciclo 2010 en adelante; si es NÚMERO, del
    # ciclo 1980-2009. Fuera de Norteamérica esa regla no se aplica (VW, por ejemplo, usa
    # 'ZZZ' en esas posiciones), así que ahí se muestran las dos posibilidades.
    letra_anio = vin[9]
    base_anio = ANIOS_VIN.get(letra_anio)
    resultado["anio_estimado"] = None
    resultado["anio_alternativo"] = None
    resultado["anio_preciso"] = False
    if base_anio:
        tope = datetime.now().year + 1
        es_norteamerica = vin[0] in "12345"
        if es_norteamerica:
            anio = base_anio + 30 if vin[6].isalpha() else base_anio
            if anio <= tope:
                resultado["anio_estimado"] = anio
                resultado["anio_preciso"] = True
        if not resultado["anio_preciso"]:
            candidatos = [a for a in sorted({base_anio, base_anio + 30}) if a <= tope]
            if candidatos:
                resultado["anio_estimado"] = candidatos[-1]
                if len(candidatos) > 1:
                    resultado["anio_alternativo"] = candidatos[0]

    # --- Dígito verificador (9ª posición) ---
    # Obligatorio en Norteamérica: se calcula con el resto del VIN, así que si no coincide es
    # porque hay un carácter mal tipeado. Sirve para no buscar repuestos de un auto equivocado.
    resultado["digito_verificador"] = None
    if vin[0] in "12345":
        valores = {**{str(d): d for d in range(10)},
                   "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
                   "J": 1, "K": 2, "L": 3, "M": 4, "N": 5, "P": 7, "R": 9,
                   "S": 2, "T": 3, "U": 4, "V": 5, "W": 6, "X": 7, "Y": 8, "Z": 9}
        pesos = [8, 7, 6, 5, 4, 3, 2, 10, 0, 9, 8, 7, 6, 5, 4, 3, 2]
        try:
            suma = sum(valores[ch] * peso for ch, peso in zip(vin, pesos))
            esperado = "X" if suma % 11 == 10 else str(suma % 11)
            resultado["digito_verificador"] = (vin[8] == esperado)
        except KeyError:
            resultado["digito_verificador"] = None

    # --- Lo más confiable de todo: que el auto ya esté en una ficha ---
    # Si este VIN exacto ya pasó por el mostrador, no hay nada que estimar: el modelo, el año y
    # el motor son los que cargó una persona mirando el auto. Eso pisa cualquier deducción, y
    # además trae la patente, el dueño y el historial de piezas.
    ficha = buscar_vehiculo_por_vin(vin)
    if ficha:
        resultado["vehiculo"] = ficha
        if ficha.get("modelo_auto"):
            resultado["modelo"] = ficha["modelo_auto"]
            resultado["modelo_exacto"] = True
            resultado["modelo_notas"] = f"de la ficha de la patente {ficha.get('patente') or '—'}"
        if ficha.get("motorizacion"):
            resultado["motor"] = ficha["motorizacion"]
            resultado["motor_origen"] = f"ficha de la patente {ficha.get('patente') or '—'}"
        if ficha.get("anio") and str(ficha["anio"]).strip().isdigit():
            resultado["anio_estimado"] = int(str(ficha["anio"]).strip())
            resultado["anio_alternativo"] = None
            resultado["anio_preciso"] = True

    return resultado


def agregar_fabricante_vin(wmi, fabricante, pais):
    wmi = wmi.strip().upper()
    with db_lock:
        c.execute(
            "INSERT INTO fabricantes_vin (wmi, fabricante, pais) VALUES (?, ?, ?) "
            "ON CONFLICT(wmi) DO UPDATE SET fabricante=excluded.fabricante, pais=excluded.pais",
            (wmi, fabricante.strip(), pais.strip())
        )
        conn.commit()


def enseniar_modelo_vin(wmi, vds, modelo, notas=None):
    """Guarda que este patrón de VIN corresponde a este modelo. La próxima vez se autocompleta."""
    wmi, vds = wmi.strip().upper(), vds.strip().upper()
    if not wmi or not vds or not modelo.strip():
        return False
    with db_lock:
        c.execute("""INSERT INTO modelos_vin (wmi, vds, modelo, notas) VALUES (?, ?, ?, ?)
                     ON CONFLICT(wmi, vds) DO UPDATE SET modelo = excluded.modelo,
                        notas = excluded.notas, veces = veces + 1""",
                  (wmi, vds, modelo.strip(), (notas or "").strip() or None))
        conn.commit()
    return True


def olvidar_modelo_vin(wmi, vds):
    with db_lock:
        c.execute("DELETE FROM modelos_vin WHERE wmi = ? AND vds = ?",
                  (wmi.strip().upper(), vds.strip().upper()))
        conn.commit()
    return c.rowcount


def listar_modelos_vin():
    c.execute("""SELECT m.wmi AS "WMI", m.vds AS "Patrón (VDS)", m.modelo AS "Modelo",
                        COALESCE(f.fabricante, '—') AS "Fabricante", m.notas AS "Notas"
                 FROM modelos_vin m LEFT JOIN fabricantes_vin f ON f.wmi = m.wmi
                 ORDER BY f.fabricante, m.modelo""")
    return filas_a_listas(c)


def listar_fabricantes_vin():
    c.execute("""SELECT wmi AS "WMI", fabricante AS "Fabricante", pais AS "País"
                 FROM fabricantes_vin ORDER BY fabricante""")
    return filas_a_listas(c)


# ============================================================
# MODO MECÁNICO — VISOR DE ESQUEMAS
# ============================================================
def guardar_esquema(titulo, marca_auto, modelo_auto, sistema, descripcion, imagen_bytes, imagen_nombre, generado_ia=False):
    with db_lock:
        c.execute(
            "INSERT INTO esquemas (titulo, marca_auto, modelo_auto, sistema, descripcion, imagen_blob, "
            "imagen_nombre, generado_ia) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (titulo.strip(), marca_auto.strip(), modelo_auto.strip(), sistema.strip(), descripcion.strip(),
             imagen_bytes, imagen_nombre, 1 if generado_ia else 0)
        )
        conn.commit()


# Pistas de piezas típicas en inglés por sistema, para ayudar al modelo gratuito (Flux) a acertar
# mejor el contenido — sin esto tiende a dibujar cualquier cosa genérica de "auto".
PARTES_TIPICAS_POR_SISTEMA = {
    "Motor": "engine block, pistons, cylinder head, timing chain, oil pan",
    "Refrigeración": "radiator, water pump, thermostat, coolant hoses, coolant expansion tank, cooling fan",
    "Retenes y juntas": "crankshaft seal, camshaft seal, gaskets, o-rings",
    "Frenos": "brake disc, brake caliper, brake pads, brake drum, brake hose",
    "Suspensión": "shock absorber, coil spring, control arm, ball joint, stabilizer bar",
    "Dirección": "steering rack, tie rod, steering column, power steering pump",
    "Transmisión": "gearbox, clutch disc, driveshaft, CV joint",
    "Embrague": "clutch disc, clutch pressure plate, release bearing, clutch cable",
    "Correas y distribución": "timing belt, timing belt tensioner, timing belt kit, pulleys",
    "Eléctrico": "alternator, starter motor, battery, wiring harness, fuse box",
    "Combustible": "fuel pump, fuel filter, fuel injectors, fuel tank, fuel lines",
    "Escape": "exhaust manifold, muffler, catalytic converter, exhaust pipe",
    "Aire acondicionado": "AC compressor, condenser, evaporator, AC hoses",
}

# Nombre del sistema en inglés, para no mezclar español dentro de un prompt en inglés.
SISTEMA_EN = {
    "Motor": "engine", "Refrigeración": "cooling", "Retenes y juntas": "seals and gaskets",
    "Frenos": "brake", "Suspensión": "suspension", "Dirección": "steering",
    "Transmisión": "transmission", "Embrague": "clutch", "Correas y distribución": "timing belt",
    "Eléctrico": "electrical", "Combustible": "fuel", "Escape": "exhaust",
    "Aire acondicionado": "air conditioning",
}


def generar_esquema_orientativo_ia(marca, modelo, motorizacion, sistema):
    """Genera una imagen orientativa/genérica (NO una foto real del vehículo) con Gemini.
    Requiere que la generación de imágenes esté habilitada en la API key correspondiente.
    Usa una key APARTE de la del resto de las funciones de IA (identificar_pieza_por_foto,
    extraer_datos_cedula, transcribir_audio, leer_remito_por_foto) — así, aunque esas otras se usen
    mucho y choquen contra el límite gratuito, nunca pueden generar un cobro por sí solas: la única
    key con facturación habilitada es esta, y solo la usa esta función."""
    from google import genai
    from PIL import Image as PILImage

    api_key = st.secrets.get("gemini_api_key_imagenes") if hasattr(st, "secrets") else None
    if not api_key:
        return None, (
            "No configuraste 'gemini_api_key_imagenes' en Streamlit Cloud (Settings → Secrets). "
            "A propósito es una key distinta de 'gemini_api_key', separada solo para esta función."
        )

    sistema_en = SISTEMA_EN.get(sistema, sistema)
    pistas = PARTES_TIPICAS_POR_SISTEMA.get(sistema, "")
    pistas_txt = f" Mostrá específicamente: {pistas}." if pistas else ""
    prompt = (
        f"Genera un diagrama técnico de despiece ('exploded view') en estilo línea/dibujo técnico "
        f"(como los planos de catálogos de repuestos), mostrando ÚNICAMENTE los componentes del sistema "
        f"de {sistema_en} de un automóvil {marca} {modelo} {motorizacion}.{pistas_txt} No dibujes la "
        f"carrocería completa del auto — solo estas piezas mecánicas, separadas entre sí (vista "
        f"explosionada), unidas por líneas finas, sobre fondo blanco liso, en blanco y negro o con líneas "
        f"oscuras simples. IMPORTANTE: no incluyas números, letras, flechas de referencia, texto ni logos "
        f"de ninguna marca dentro del dibujo — esos se agregan después por separado. Es una referencia "
        f"orientativa general de cómo se relacionan las piezas entre sí, no necesita ser exacto a ese "
        f"modelo puntual."
    )
    try:
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(model="gemini-2.5-flash-image", contents=[prompt])
        for part in response.candidates[0].content.parts:
            if getattr(part, "inline_data", None) is not None:
                img = PILImage.open(io.BytesIO(part.inline_data.data)).convert("RGB")
                salida = io.BytesIO()
                img.save(salida, format="JPEG", quality=90)
                registrar_uso_ia("Generar imagen orientativa (paga)", True)
                return salida.getvalue(), None
        registrar_uso_ia("Generar imagen orientativa (paga)", False)
        return None, "Gemini no devolvió ninguna imagen para ese pedido."
    except Exception as e:
        registrar_uso_ia("Generar imagen orientativa (paga)", False)
        texto_error = str(e)
        if "RESOURCE_EXHAUSTED" in texto_error or "429" in texto_error or "quota" in texto_error.lower():
            return None, (
                "Esta función no está habilitada en la API key configurada (el error dice 'limit: 0' "
                "para ese modelo). Hay que habilitarla en aistudio.google.com para la key "
                "'gemini_api_key_imagenes'. El resto de las funciones de IA usan una key distinta y "
                "separada, así que no se ven afectadas por esto."
            )
        return None, f"Error generando la imagen: {texto_error}"


def listar_marcas_esquemas():
    """Marcas con esquemas ya cargados, UNIDAS con las precargadas en el catálogo (sin imagen todavía)."""
    c.execute("""SELECT marca_auto AS marca FROM esquemas WHERE marca_auto IS NOT NULL AND TRIM(marca_auto) != ''
                 UNION
                 SELECT marca FROM esquemas_catalogo
                 ORDER BY marca""")
    return [r["marca"] for r in c.fetchall()]


def listar_modelos_esquemas(marca):
    c.execute("""SELECT modelo_auto AS modelo FROM esquemas
                 WHERE marca_auto = ? AND modelo_auto IS NOT NULL AND TRIM(modelo_auto) != ''
                 UNION
                 SELECT modelo FROM esquemas_catalogo WHERE marca = ?
                 ORDER BY modelo""", (marca, marca))
    return [r["modelo"] for r in c.fetchall()]


def listar_sistemas_esquemas(marca, modelo):
    c.execute("""SELECT DISTINCT sistema FROM esquemas
                 WHERE marca_auto = ? AND modelo_auto = ? AND sistema IS NOT NULL AND TRIM(sistema) != ''
                 ORDER BY sistema""", (marca, modelo))
    return [r["sistema"] for r in c.fetchall()]


def listar_esquemas_por_categoria(marca, modelo, sistema):
    c.execute("""SELECT id, titulo, descripcion, generado_ia FROM esquemas
                 WHERE marca_auto = ? AND modelo_auto = ? AND sistema = ? ORDER BY titulo""",
              (marca, modelo, sistema))
    return [dict(r) for r in c.fetchall()]


def agregar_vehiculo_catalogo(marca, modelo):
    with db_lock:
        c.execute("INSERT OR IGNORE INTO esquemas_catalogo (marca, modelo) VALUES (?, ?)",
                   (marca.strip(), modelo.strip()))
        conn.commit()


def eliminar_vehiculo_catalogo(marca, modelo):
    with db_lock:
        c.execute("DELETE FROM esquemas_catalogo WHERE marca = ? AND modelo = ?", (marca, modelo))
        conn.commit()


def listar_catalogo_precargado():
    """Marca/modelo precargados sin ningún esquema real cargado todavía (candidatos a borrar)."""
    c.execute("""SELECT marca, modelo FROM esquemas_catalogo ec
                 WHERE NOT EXISTS (
                     SELECT 1 FROM esquemas e WHERE e.marca_auto = ec.marca AND e.modelo_auto = ec.modelo
                 ) ORDER BY marca, modelo""")
    return [dict(r) for r in c.fetchall()]


def listar_esquemas(texto_filtro=""):
    if texto_filtro.strip():
        like = f"%{texto_filtro.strip().upper()}%"
        c.execute("""SELECT id, titulo, marca_auto, modelo_auto, sistema, descripcion, generado_ia FROM esquemas
                     WHERE UPPER(titulo) LIKE ? OR UPPER(marca_auto) LIKE ? OR UPPER(modelo_auto) LIKE ?
                        OR UPPER(sistema) LIKE ?
                     ORDER BY marca_auto, modelo_auto""", (like, like, like, like))
    else:
        c.execute("SELECT id, titulo, marca_auto, modelo_auto, sistema, descripcion, generado_ia FROM esquemas "
                   "ORDER BY marca_auto, modelo_auto")
    return [dict(row) for row in c.fetchall()]


def obtener_imagen_esquema(esquema_id):
    c.execute("SELECT imagen_blob FROM esquemas WHERE id = ?", (esquema_id,))
    row = c.fetchone()
    return row["imagen_blob"] if row else None


def eliminar_esquema(esquema_id):
    with db_lock:
        c.execute("DELETE FROM esquemas WHERE id = ?", (esquema_id,))
        conn.commit()


def agregar_punto_esquema(esquema_id, numero, nombre_pieza, codigo, pos_x=None, pos_y=None):
    """Agrega una pieza marcada dentro de un esquema. Si el código coincide con un producto
    ya cargado, lo vincula (producto_id); si no, igual guarda el código como texto de referencia.
    pos_x/pos_y son porcentajes (0-100) de dónde está la pieza en la imagen, para dibujar el marcador."""
    codigo = (codigo or "").strip()
    producto_id = None
    if codigo:
        clean = sanitizar(codigo)
        if clean:
            c.execute("SELECT id FROM productos WHERE codigo_clean = ? LIMIT 1", (clean,))
            fila = c.fetchone()
            if fila:
                producto_id = fila["id"]
    with db_lock:
        c.execute("SELECT COALESCE(MAX(orden), 0) + 1 FROM esquema_puntos WHERE esquema_id = ?", (esquema_id,))
        siguiente_orden = c.fetchone()[0]
        c.execute(
            "INSERT INTO esquema_puntos (esquema_id, numero, nombre_pieza, codigo, producto_id, pos_x, pos_y, orden) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (esquema_id, (numero or "").strip(), nombre_pieza.strip(), codigo, producto_id, pos_x, pos_y, siguiente_orden)
        )
        conn.commit()
    return producto_id is not None


def listar_puntos_esquema(esquema_id):
    c.execute("""SELECT id, numero, nombre_pieza, codigo, producto_id, pos_x, pos_y FROM esquema_puntos
                 WHERE esquema_id = ? ORDER BY orden""", (esquema_id,))
    return [dict(r) for r in c.fetchall()]


@st.cache_data(show_spinner=False, max_entries=40)
def imagen_esquema_lista_para_mostrar(imagen_bytes, firma_puntos, _puntos):
    """Dibuja los marcadores sobre la imagen y guarda el resultado en caché. Sin esto, cada
    refresco de pantalla vuelve a redibujar la imagen entera con Pillow — aunque el desplegable
    esté cerrado, porque Streamlit igual arma el contenido.
    El caché se invalida solo cuando cambia la imagen (se compara su contenido) o cuando cambian
    los puntos marcados ('firma_puntos'). '_puntos' va con guion bajo adelante para que Streamlit
    no intente usarlo como clave: es una lista de diccionarios y no le sirve para comparar."""
    return generar_imagen_con_marcadores(imagen_bytes, _puntos)


def firma_de_puntos(puntos):
    """Resumen corto de los puntos, para usar como clave de caché."""
    return tuple((p["id"], p.get("numero"), p.get("pos_x"), p.get("pos_y")) for p in puntos)


def generar_imagen_con_marcadores(imagen_bytes, puntos):
    """Dibuja círculos numerados sobre la imagen real, en las posiciones (%) que cargó el admin.
    Si la imagen está corrupta, devuelve la original sin marcadores en vez de romper la pantalla."""
    from PIL import Image, ImageDraw, UnidentifiedImageError

    puntos_con_pos = [p for p in puntos if p.get("pos_x") is not None and p.get("pos_y") is not None]
    if not puntos_con_pos:
        return imagen_bytes

    try:
        img = Image.open(io.BytesIO(imagen_bytes)).convert("RGB")
        ancho, alto = img.size
        draw = ImageDraw.Draw(img)
        radio = max(min(ancho, alto) // 40, 12)

        for i, p in enumerate(puntos_con_pos, start=1):
            x = int(p["pos_x"] / 100 * ancho)
            y = int(p["pos_y"] / 100 * alto)
            etiqueta = p.get("numero") or str(i)
            draw.ellipse([x - radio, y - radio, x + radio, y + radio], fill=(232, 163, 61), outline=(20, 20, 20), width=2)
            bbox = draw.textbbox((0, 0), etiqueta)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            draw.text((x - tw / 2, y - th / 2 - bbox[1]), etiqueta, fill=(20, 20, 20))

        salida = io.BytesIO()
        img.save(salida, format="JPEG", quality=90)
        return salida.getvalue()
    except (UnidentifiedImageError, OSError):
        return imagen_bytes


def eliminar_punto_esquema(punto_id):
    with db_lock:
        c.execute("DELETE FROM esquema_puntos WHERE id = ?", (punto_id,))
        conn.commit()


def guardar_presupuesto_mecanico(mecanico_id, cliente_nombre, items, mano_obra):
    total = sum(it["precio"] * it["cantidad"] for it in items) + mano_obra
    with db_lock:
        c.execute(
            "INSERT INTO presupuestos_mecanico (mecanico_id, cliente_nombre, items_json, mano_obra, total) "
            "VALUES (?, ?, ?, ?, ?)",
            (mecanico_id, cliente_nombre.strip(), json.dumps(items, ensure_ascii=False), mano_obra, total)
        )
        conn.commit()
    return total


def listar_presupuestos_mecanico(mecanico_id):
    c.execute("""SELECT id AS "ID", cliente_nombre AS "Cliente", items_json, mano_obra AS "Mano de obra",
                 total AS "Total", creado_en AS "Fecha" FROM presupuestos_mecanico
                 WHERE mecanico_id = ? ORDER BY id DESC""", (mecanico_id,))
    return filas_a_listas(c)


def generar_pdf_presupuesto_mecanico(nombre_mecanico, cliente_nombre, items, mano_obra, total):
    from fpdf import FPDF

    def limpiar(texto):
        return str(texto).encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Presupuesto", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, limpiar(f"Mecánico: {nombre_mecanico}"), new_x="LMARGIN", new_y="NEXT")
    if cliente_nombre:
        pdf.cell(0, 6, limpiar(f"Cliente: {cliente_nombre}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Fecha: {datetime.now():%d/%m/%Y %H:%M}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Repuestos", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    for it in items:
        subtotal = it["precio"] * it["cantidad"]
        linea = f"- {it['codigo']} ({it['marca']}) x{it['cantidad']} - ${subtotal:,.0f}"
        pdf.multi_cell(0, 6, limpiar(linea), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, limpiar(f"Mano de obra: ${mano_obra:,.0f}"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, limpiar(f"TOTAL: ${total:,.0f}"), new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


def mostrar_portal_mecanico():
    """Portal separado para mecánicos externos — no ven ninguna de las pestañas internas del
    negocio, solo esto: buscar repuestos, armar su presupuesto con su propia mano de obra, y
    ver sus presupuestos guardados anteriormente (nunca los de otros mecánicos)."""
    mostrar_avisos_pendientes()
    mecanico_id = st.session_state.get("mecanico_id")
    nombre_mecanico = st.session_state.get("admin_nombre", "")
    st.markdown(f"### 🔧 Portal de mecánico — {nombre_mecanico}")
    st.caption(
        "Buscá repuestos del catálogo, armá tu presupuesto con tu propia mano de obra, y mandaselo "
        "al cliente. Solo ves tus propios presupuestos guardados, no los de otros mecánicos."
    )

    if "presupuesto_mecanico_items" not in st.session_state:
        st.session_state["presupuesto_mecanico_items"] = []

    st.markdown("**🔍 Buscar repuesto**")
    texto_busqueda_mec = st.text_input("Código o descripción:", key="mec_busqueda")
    if texto_busqueda_mec.strip():
        clean_mec = sanitizar(texto_busqueda_mec)
        resultados_mec = buscar_por_codigo(clean_mec) if clean_mec else []
        if not resultados_mec:
            resultados_mec = buscar_por_texto(texto_busqueda_mec)
        if resultados_mec:
            for fila in resultados_mec[:20]:
                colm1, colm2, colm3 = st.columns([3, 1, 1])
                precio_txt = f"${fila['Precio']:,.0f}" if fila.get("Precio") else "s/precio"
                colm1.write(f"{fila['Marca']} - {fila['Codigo']} — {fila.get('Descripcion') or ''} ({precio_txt})")
                cantidad_mec = colm2.number_input("Cant.", min_value=1, value=1, step=1,
                                                    key=f"cant_mec_{fila['ID']}", label_visibility="collapsed")
                if colm3.button("➕", key=f"add_mec_{fila['ID']}"):
                    st.session_state["presupuesto_mecanico_items"].append({
                        "codigo": fila["Codigo"], "marca": fila["Marca"],
                        "descripcion": fila.get("Descripcion") or "",
                        "precio": float(fila.get("Precio") or 0), "cantidad": int(cantidad_mec)
                    })
                    st.rerun()
        else:
            st.caption("Sin resultados.")

    items_actuales = st.session_state["presupuesto_mecanico_items"]
    st.markdown("---")
    st.markdown(f"**📋 Presupuesto actual ({len(items_actuales)} ítem(s))**")
    if items_actuales:
        subtotal_repuestos = 0.0
        for i, it in enumerate(items_actuales):
            subtotal_item = it["precio"] * it["cantidad"]
            subtotal_repuestos += subtotal_item
            coli1, coli2 = st.columns([4, 1])
            coli1.write(f"{it['marca']} - {it['codigo']} x{it['cantidad']} — ${subtotal_item:,.0f}")
            coli2.button("🗑️", key=f"quitar_mec_{i}",
                          on_click=quitar_item_lista_sesion, args=("presupuesto_mecanico_items", i))

        cliente_nombre_mec = st.text_input("Nombre del cliente (opcional):", key="mec_cliente")
        mano_obra_mec = st.number_input("Mano de obra ($):", min_value=0.0, step=500.0, key="mec_mano_obra")
        total_mec = subtotal_repuestos + mano_obra_mec
        st.metric("Total", f"${total_mec:,.0f}")

        colb1, colb2, colb3 = st.columns(3)
        if colb1.button("💾 Guardar presupuesto", type="primary"):
            guardar_presupuesto_mecanico(mecanico_id, cliente_nombre_mec, items_actuales, mano_obra_mec)
            st.success("Presupuesto guardado.")
            st.session_state["presupuesto_mecanico_items"] = []
            st.rerun()

        pdf_bytes_mec = pdf_con_cache(
            "presupuesto_mecanico", generar_pdf_presupuesto_mecanico,
            nombre_mecanico, cliente_nombre_mec, items_actuales, mano_obra_mec, total_mec
        )
        colb2.download_button("📄 PDF", data=pdf_bytes_mec, file_name="presupuesto.pdf", mime="application/pdf")

        mensaje_mec = f"Presupuesto de {nombre_mecanico}:\n"
        for it in items_actuales:
            mensaje_mec += f"- {it['codigo']} ({it['marca']}) x{it['cantidad']} — ${it['precio']*it['cantidad']:,.0f}\n"
        mensaje_mec += f"Mano de obra: ${mano_obra_mec:,.0f}\nTOTAL: ${total_mec:,.0f}"
        url_wa_mec = "https://wa.me/?text=" + quote(mensaje_mec)
        colb3.link_button("📲 WhatsApp", url_wa_mec)
    else:
        st.caption("Todavía no agregaste ningún repuesto al presupuesto.")

    st.markdown("---")
    st.markdown("**📁 Mis presupuestos anteriores**")
    anteriores = listar_presupuestos_mecanico(mecanico_id)
    if anteriores:
        for p in anteriores[:20]:
            with st.expander(f"{p['Fecha']} — {p['Cliente'] or 'sin nombre'} — ${p['Total']:,.0f}"):
                items_p = json.loads(p["items_json"])
                for it in items_p:
                    st.write(f"- {it['codigo']} ({it['marca']}) x{it['cantidad']} — ${it['precio']*it['cantidad']:,.0f}")
                st.write(f"Mano de obra: ${p['Mano de obra']:,.0f}")
    else:
        st.caption("Todavía no guardaste ningún presupuesto.")


# ============================================================
# ENCABEZADO
# ============================================================
st.markdown(
    """
    <div class="app-header">
        <p class="app-header__eyebrow">Base de equivalencias de repuestos</p>
        <h1>🔧 Equivalencias El Chavo</h1>
        <p>Sistema de búsqueda de repuestos por equivalencia</p>
    </div>
    """,
    unsafe_allow_html=True
)

# Pantalla de login apenas se abre la app, con opción de seguir sin loguearse.
if not es_admin() and not st.session_state.get("saltar_login"):
    mostrar_login_inicial()
    st.stop()

if es_admin() or es_operador_o_admin() or st.session_state.get("nivel_usuario") == "mecanico":
    col_estado, col_modo, col_salir = st.columns([3, 1.4, 1])
    nombre_sesion = st.session_state.get("admin_nombre", "")
    etiquetas_nivel = {"admin": "administrador", "operador": "operador", "mecanico": "mecánico"}
    etiqueta_nivel = etiquetas_nivel.get(st.session_state.get("nivel_usuario"), "administrador")
    col_estado.caption(f"🔓 Sesión de {etiqueta_nivel} activa ({nombre_sesion}).")
    col_modo.selectbox("Vista:", ["📱 Celular", "💻 Computadora"], key="modo_vista",
                        label_visibility="collapsed",
                        help="Acomoda la pantalla según el dispositivo desde el que estés entrando.")
    if col_salir.button("Salir"):
        st.session_state.nivel_usuario = None
        st.session_state.admin_nombre = None
        st.session_state.mecanico_id = None
        st.rerun()
else:
    col_estado, col_modo = st.columns([3, 1.4])
    col_estado.caption(f"👤 Usando como: {obtener_usuario_actual()}")
    col_modo.selectbox("Vista:", ["📱 Celular", "💻 Computadora"], key="modo_vista",
                        label_visibility="collapsed",
                        help="Acomoda la pantalla según el dispositivo desde el que estés entrando.")

if st.session_state.get("nivel_usuario") == "mecanico":
    mostrar_portal_mecanico()
    st.stop()

if "lista_whatsapp" not in st.session_state:
    st.session_state.lista_whatsapp = []  # lista de códigos agregados para el mensaje

# ============================================================
# NAVEGACIÓN PRINCIPAL
# ============================================================
# Antes esto era st.tabs(). Se cambió por un selector guardado en la sesión por dos motivos:
#  1) st.tabs pierde en qué pestaña estabas cada vez que la página se refresca, y te devolvía
#     al Buscador (el problema de "me vuelve al inicio" al tocar un botón).
#  2) st.tabs dibuja TODAS las pestañas en cada refresco, aunque no las estés mirando —
#     con 8 pestañas haciendo consultas a la base, eso era trabajo al pedo. Ahora solo se
#     arma la sección que estás viendo, así que la app responde bastante más rápido.
# Aviso fuerte si la base quedó vacía: Streamlit Cloud borra el disco al redesplegar, y sin
# este cartel uno se entera recién cuando busca un código y no aparece nada.
c.execute("SELECT COUNT(*) FROM productos")
_total_productos = c.fetchone()[0]
if _total_productos == 0:
    st.error(
        "⚠️ **La base está vacía.** Esto pasa porque el servidor borra el disco de la app cuando "
        "se redespliega o se reinicia. Si tenés un backup (.db) descargado, restauralo desde "
        "**Estadísticas → 💾 Backup y config**. Para que no vuelva a pasar, mirá ahí abajo la "
        "sección de copia permanente."
    )
elif st.session_state.get("_restaurado_de_semilla"):
    st.info(
        f"♻️ La base se restauró sola desde la copia guardada en el repositorio "
        f"({_total_productos} productos). Lo cargado después de esa copia no está — "
        "acordate de actualizarla cada tanto."
    )

# Chequeo de salud, en cualquier sección. Va arriba a propósito: los controles de mantenimiento
# se fueron sumando de a uno y quedaron repartidos en siete pantallas distintas, ninguna avisa
# sola, y en la práctica nadie entra a mirarlas hasta que algo ya salió mal.
#
# El resultado se guarda en la sesión y se recalcula cada 3 minutos. Correrlo en CADA clic serían
# 75 ms de más en cada cosa que se toca, y estos números no cambian de un segundo al otro.
# Mantenimiento del día. Va acá porque es el único lugar por el que pasan todos, sin importar
# a qué sección entren. Se protege con try porque una tarea de fondo que falla NUNCA puede
# impedir que la app abra.
if not st.session_state.get("_tareas_dia_corridas"):
    st.session_state["_tareas_dia_corridas"] = True
    try:
        _hecho_hoy = tareas_automaticas_del_dia()
        if _hecho_hoy:
            st.session_state["_aviso_tareas"] = _hecho_hoy
    except Exception:
        pass

_hecho_hoy = st.session_state.pop("_aviso_tareas", None)
if _hecho_hoy:
    st.caption("🔧 Mantenimiento automático de hoy: " + " · ".join(_hecho_hoy))

_ahora = time.time()
_cache_salud = st.session_state.get("_salud_cache")
if not _cache_salud or _ahora - _cache_salud["momento"] > 180:
    try:
        _cache_salud = {"momento": _ahora, "problemas": diagnostico_de_salud()}
    except Exception:
        _cache_salud = {"momento": _ahora, "problemas": []}
    st.session_state["_salud_cache"] = _cache_salud

_problemas = _cache_salud["problemas"]
if _problemas:
    _graves = [p for p in _problemas if p["nivel"] == "alto"]
    _titulo = (f"⚠️ {len(_problemas)} cosa(s) para revisar"
               + (f" — {len(_graves)} importante(s)" if _graves else ""))
    with st.expander(_titulo, expanded=False):
        for _p in _problemas:
            _icono = "🔴" if _p["nivel"] == "alto" else "🟡"
            st.markdown(f"{_icono} **{_p['titulo']}**")
            st.caption(f"{_p['detalle']}  \n📍 {_p['donde']}")
        if st.button("🔄 Volver a revisar", key="refrescar_salud"):
            st.session_state.pop("_salud_cache", None)
            st.rerun()

GRUPOS_MANTENIMIENTO = ["🧹 Limpiar vínculos", "🧠 Calidad y aprendizaje", "📷 Fotos",
                        "🩺 Estado y papelera"]

PAGINAS = ["🔍 Buscador", "🔗 Vincular manual", "📁 Cargar Excel", "🗂️ Administrar",
           "📊 Estadísticas", "📋 Lista WhatsApp", "🚗 Vehículos", "🛠️ Modo Mecánico"]

if st.session_state.get("pagina_actual") not in PAGINAS:
    st.session_state["pagina_actual"] = PAGINAS[0]

# En los dos modos se usa st.radio en vez de un desplegable: el desplegable de Streamlit lleva
# un campo de texto adentro para filtrar, y en el celular eso abre el teclado cada vez que lo
# tocás, que es molesto para algo que se usa todo el tiempo. Con radio es un toque y listo.
# El CSS los muestra como pastillas: en el celular se acomodan solas en varias filas.
st.radio("Sección:", PAGINAS, key="pagina_actual", horizontal=True, label_visibility="collapsed")

pagina = st.session_state["pagina_actual"]

# Los avisos que quedaron guardados antes del último refresco. Van acá, arriba del contenido
# de la página, para que se vean sí o sí — sin esto, cada "Guardado" se perdía en el refresco.
mostrar_avisos_pendientes()

# ============================================================
# BUSCADOR
# ============================================================
if pagina == PAGINAS[0]:
    with st.expander("❓ Guía rápida — cómo usar esta app"):
        st.markdown("""
- **🔍 Buscador** — el corazón de la app. Buscá por código (acepta varios separados por coma) o por
  descripción. Los resultados muestran todas las marcas equivalentes, precio, stock y un link directo
  a la ficha del proveedor si lo cargaste.
- **🔗 Vincular manual** — cuando encontrás que dos o más códigos de distintos proveedores son la misma
  pieza y todavía no están relacionados, los agrupás acá de una sola vez.
- **📁 Cargar Excel** — subís la lista completa de un proveedor (Excel, CSV o PDF con tabla) y la app
  arma las equivalencias sola comparando código OEM. También lee remitos por foto.
- **🗂️ Administrar** — todo lo de mantenimiento: marcas, medidas de piezas, fotos de productos,
  combos relacionados, mensajería y cobros.
- **📊 Estadísticas** — números generales, backups, auditoría de stock y qué se buscó sin encontrar nada.
- **📋 Lista WhatsApp** — armá una cotización con varios productos y mandala por WhatsApp o como PDF.
- **🚗 Vehículos** — ficha por patente: historial de piezas, alertas de mantenimiento, y podés cargar
  los datos sacándole una foto a la cédula.
- **🛠️ Modo Mecánico** — diccionario de códigos de falla (DTC), lector de VIN, esquemas técnicos y
  un conversor de unidades.

Casi todo lo que edita o borra algo pide la contraseña de administrador la primera vez que lo usás.
        """)

    # Si se tocó un botón de sugerencia rápida (favorito o búsqueda reciente), precargamos el
    # campo de búsqueda ANTES de crear el widget — si se hace después de creado, Streamlit tira error.
    if "sugerencia_busqueda" in st.session_state:
        st.session_state["busqueda_input"] = st.session_state.pop("sugerencia_busqueda")

    # El carrito arriba de todo: si está armándose un presupuesto, tiene que estar a la vista.
    # Escondido en otra pantalla, la gente se olvida de lo que ya sumó y lo suma dos veces.
    if st.session_state.get("carrito"):
        _cart = st.session_state["carrito"]
        _total = sum((x["precio"] or 0) * x["cantidad"] for x in _cart.values())
        with st.expander(f"🛒 Presupuesto en armado — {len(_cart)} ítem(s) · ${_total:,.0f}",
                          expanded=False):
            for _pid, _item in list(_cart.items()):
                ci1, ci2, ci3 = st.columns([5, 2, 1])
                ci1.markdown(f"**{_item['marca']} {_item['codigo']}** — "
                              f"{(_item['descripcion'] or '')[:40]}")
                _nueva_cant = ci2.number_input("Cant.", min_value=1, value=_item["cantidad"],
                                                step=1, key=f"cant_cart_{_pid}",
                                                label_visibility="collapsed")
                if _nueva_cant != _item["cantidad"]:
                    st.session_state["carrito"][_pid]["cantidad"] = int(_nueva_cant)
                    st.rerun()
                if ci3.button("🗑️", key=f"quitar_cart_{_pid}"):
                    st.session_state["carrito"].pop(_pid, None)
                    # También la cantidad que quedó guardada del widget: sin esto, volver a
                    # sumar el mismo producto lo traía con la cantidad vieja en vez de 1.
                    st.session_state.pop(f"cant_cart_{_pid}", None)
                    st.rerun()

            _lineas = [f"{x['marca']} {x['codigo']} x{x['cantidad']} — "
                       f"${(x['precio'] or 0) * x['cantidad']:,.0f}" for x in _cart.values()]
            _texto = "\n".join(_lineas) + f"\n\nTOTAL: ${_total:,.0f}"
            st.caption("Para copiar y mandar por WhatsApp:")
            st.code(_texto, language=None)
            cb1, cb2 = st.columns(2)
            if cb1.button("🔒 Apartar todo el presupuesto"):
                _apartados, _fallaron = 0, []
                for _pid, _item in list(_cart.items()):
                    _ok, _msg = reservar_stock(_pid, _item["cantidad"], "presupuesto")
                    if _ok:
                        _apartados += 1
                        # Lo apartado sale del carrito. Si se quedara, tocar el botón otra vez
                        # —porque uno de los ítems falló, por ejemplo— volvía a apartar los que
                        # ya estaban apartados y el stock libre terminaba en cualquier cosa.
                        st.session_state["carrito"].pop(_pid, None)
                        st.session_state.pop(f"cant_cart_{_pid}", None)
                    else:
                        _fallaron.append(f"{_item['codigo']}: {_msg}")
                if _apartados:
                    avisar("success", f"Se apartaron {_apartados} ítem(s).")
                for _f in _fallaron:
                    avisar("warning", _f)
                st.rerun()
            if cb2.button("🗑️ Vaciar presupuesto"):
                st.session_state["carrito"] = {}
                st.rerun()

    # El escáner va PRIMERO, antes que la foto y las medidas. No es una preferencia de orden:
    # el código de barras es exacto, y todo lo demás —comparar siluetas, leer un grabado
    # gastado, medir— es aproximar. Poner primero lo confiable evita que alguien resuelva con
    # una adivinanza algo que la caja tenía escrito.
    with st.expander("📷 Escanear el código de barras de la caja"):
        explicar(
            "La forma más segura de identificar un repuesto: no hay nada que adivinar.",
            "El código de la caja es exacto. A diferencia de comparar fotos o leer un grabado, "
            "acá no hay interpretación posible: o lo lee o no lo lee.\n\nSacá la foto de cerca, "
            "con buena luz y el código derecho, ocupando buena parte de la pantalla. Sirve "
            "también para QR.", en_expander=True
        )
        foto_barras = st.camera_input("Apuntá al código de barras", key="cam_barras")
        if foto_barras is None:
            foto_barras = subir_archivo("...o subí una foto ya sacada:",
                                         ["jpg", "jpeg", "png"], "barras_archivo")
        if foto_barras is not None:
            datos_foto = foto_barras.getvalue() if hasattr(foto_barras, "getvalue") else foto_barras
            with st.spinner("Leyendo el código..."):
                codigos_leidos, err_barras = leer_codigo_de_barras(datos_foto)
            if err_barras:
                st.warning(err_barras)
            else:
                for cod_barra in codigos_leidos:
                    st.success(f"📷 Código leído: **`{cod_barra}`**")
                    res_barra, nota_barra = buscar_por_codigo_de_barras(cod_barra)
                    if nota_barra:
                        st.caption(nota_barra)
                    if res_barra:
                        st.success(f"✅ {len(res_barra)} coincidencia(s) — es una coincidencia "
                                    "exacta de código, no un parecido:")
                        st.dataframe(quitar_id(res_barra), use_container_width=True,
                                      hide_index=True)
                    else:
                        st.warning(
                            f"El código `{cod_barra}` no está cargado en tu catálogo."
                        )
                        parecidos_barra = codigos_por_tipeo(sanitizar(cod_barra))
                        if parecidos_barra:
                            st.caption("Códigos parecidos que sí tenés:")
                            st.dataframe(
                                [{"Código": x["Codigo"], "Marca": x["Marca"],
                                  "Descripción": x["Descripcion"], "Stock": x["Stock"]}
                                 for x in parecidos_barra],
                                use_container_width=True, hide_index=True)
                        ajeno_barra = identificar_codigo_ajeno(sanitizar(cod_barra))
                        if ajeno_barra:
                            st.info(f"🏭 Según el catálogo de {ajeno_barra['marca']}, es "
                                     f"{ajeno_barra['tipo'] or 'una pieza'} para "
                                     f"{ajeno_barra['cantidad_autos']} auto(s).")

    if es_operador_o_admin():
        with st.expander("🖼️ Buscar por parecido visual (experimental)"):
            explicar(
                "Compará una foto contra el catálogo. Sirve para acortar candidatos, nunca para "
                "confirmar una venta.",
                "No hace falta que sea la misma foto: aguanta otro ángulo, otro fondo y otro color "
                "de pieza — se compara en blanco y negro, recortando el fondo.\n\nAun así es "
                "**mucho** menos confiable que un código exacto. Con piezas lisas sin marcas ni "
                "grabado (rótulas, rulemanes, bulones) rinde mal por más buena que sea la foto. "
                "Para esas, **📐 Buscar por medidas mecánicas** anda mucho mejor.", en_expander=True
            )

            fotos_listas, prod_con_foto, fotos_pendientes, fotos_no_sirven = contar_fotos_comparables()
            fotos_pendientes = max(fotos_pendientes, contar_fotos_pendientes_de_firma())
            st.caption(
                f"📊 {fotos_listas} foto(s) listas para comparar, sobre {prod_con_foto} producto(s)" +
                (f" — {fotos_pendientes} pendiente(s) de procesar" if fotos_pendientes else "") +
                (f" — {fotos_no_sirven} sin detalle suficiente" if fotos_no_sirven else "")
            )

            if fotos_listas == 0 and fotos_pendientes == 0:
                st.info(_mensaje_catalogo_visual_vacio())

            # Se procesa una tanda chica sola, cada vez que se abre este panel. Antes había que
            # acordarse de tocar el botón, y como el comparador fue mejorando, las fotos viejas
            # quedaban con la firma vieja sin que nadie se enterara: la mejora no llegaba nunca
            # al catálogo que ya estaba cargado. La tanda es de 25 a propósito, para que abrir
            # el panel no se cuelgue.
            if fotos_pendientes:
                with st.spinner(f"Poniendo al día {min(fotos_pendientes, 25)} foto(s)..."):
                    _r = migrar_imagenes_pendientes(limite=25)
                if _r["listas"]:
                    st.caption(f"🔄 Se pusieron al día {_r['listas']} foto(s) automáticamente.")
                fotos_listas, prod_con_foto, fotos_pendientes, fotos_no_sirven = contar_fotos_comparables()
                fotos_pendientes = max(fotos_pendientes, contar_fotos_pendientes_de_firma())

            st.checkbox(
                "🤖 Traer fotos automáticamente, una tanda por día",
                value=obtener_config("fotos_automaticas", "0") == "1",
                key="fotos_auto_check",
                on_change=lambda: guardar_config(
                    "fotos_automaticas", "1" if st.session_state["fotos_auto_check"] else "0"),
                help="Cuando alguien abre la app, trae 15 fotos de las fichas del proveedor. "
                     "Con miles de productos, sentarse a esperar no es opción; así avanza solo. "
                     "Sale a internet, por eso lo elegís vos."
            )

            if fotos_pendientes and st.button(
                    f"🔄 Procesar las {fotos_pendientes} que faltan ahora"):
                with st.spinner("Procesando fotos..."):
                    r = migrar_imagenes_pendientes()
                partes = []
                if r["listas"]:
                    partes.append(f"{r['listas']} lista(s) para comparar")
                if r["sin_detalle"]:
                    partes.append(f"{r['sin_detalle']} sin detalle suficiente")
                if r["links"]:
                    partes.append(f"{r['links']} son links externos (bajalas desde Mantenimiento)")
                if r["error"]:
                    partes.append(f"{r['error']} con error")
                avisar("success", "Procesadas: " + (", ".join(partes) if partes else "no había nada pendiente") + ".")
                st.rerun()

            origen_visual = st.radio(
                "¿De dónde sale la foto a buscar?",
                ["📷 Subir una foto", "🔗 Pegar una dirección web"],
                horizontal=True, key="origen_foto_visual"
            )

            bytes_consulta = None

            if origen_visual.startswith("📷"):
                foto_visual = subir_archivo(
                    "Foto de la pieza:", ["png", "jpg", "jpeg"], "foto_visual",
                    label_visibility="collapsed"
                )
                if archivo_listo(foto_visual, "foto"):
                    bytes_consulta = foto_visual.getvalue()
                if foto_visual:
                    if st.button("🗑️ Usar otra foto", key="otra_foto_visual"):
                        olvidar_archivo("foto_visual")
                        st.session_state.pop("ocr_visual", None)
                        st.session_state.pop("resultado_visual", None)
                        st.rerun()
            else:
                st.caption(
                    "Pegá la dirección de la ficha del producto (la de tu proveedor, la de Mercado "
                    "Libre, la que sea) o la de la imagen sola. Se bajan las fotos de esa página y "
                    "elegís cuál usar — no hace falta guardar nada en el teléfono."
                )
                url_visual = st.text_input(
                    "Dirección web:", placeholder="https://...", key="url_foto_visual"
                ).strip()
                if st.button("⬇️ Traer fotos de esa dirección", disabled=not url_visual):
                    with st.spinner("Bajando..."):
                        encontradas, error_url = imagenes_de_una_direccion(url_visual)
                    if error_url:
                        st.error(error_url)
                        st.session_state.pop("fotos_de_url", None)
                    else:
                        st.session_state["fotos_de_url"] = encontradas
                        st.session_state.pop("foto_url_elegida", None)

                encontradas = st.session_state.get("fotos_de_url")
                if encontradas:
                    st.caption(f"Se encontraron {len(encontradas)} foto(s). Elegí la de la pieza:")
                    filas_img = st.columns(min(len(encontradas), 4))
                    for idx, (url_img, datos_img) in enumerate(encontradas[:8]):
                        with filas_img[idx % len(filas_img)]:
                            st.image(datos_img, use_container_width=True)
                            if st.button("Usar esta", key=f"usar_img_url_{idx}"):
                                st.session_state["foto_url_elegida"] = idx
                                st.rerun()
                    elegida = st.session_state.get("foto_url_elegida")
                    if elegida is not None and elegida < len(encontradas):
                        bytes_consulta = encontradas[elegida][1]
                        st.success(f"✅ Foto {elegida + 1} elegida.")

            # Si la IA ya reconoció qué tipo de pieza es, se propone solo. Es la señal que más
            # mejora la comparación y es la que menos cuesta: la foto ya se mandó igual.
            _datos_ocr = (st.session_state.get("ocr_visual") or (None, None))[0]
            _familias = ["No sé / buscar en todo"] + sorted(FAMILIAS_REPUESTO.keys())
            if _datos_ocr and _datos_ocr.get("tipo_pieza"):
                _sugerida = clasificar_repuesto(_datos_ocr["tipo_pieza"])
                if _sugerida in _familias and not st.session_state.get("familia_visual"):
                    st.session_state["familia_visual"] = _sugerida
                    st.caption(f"🤖 La IA vio que es «{_datos_ocr['tipo_pieza']}», así que preseleccioné "
                                f"**{_sugerida}**. Cambialo si no acertó.")

            familia_elegida = st.selectbox(
                "¿Qué tipo de pieza es? (recomendado)", _familias,
                key="familia_visual",
                help="Es lo que más mejora el resultado, y no sale de la foto: vos ya sabés qué "
                     "pieza tenés en la mano. Diciéndolo, la app deja de ofrecerte cosas de otro "
                     "rubro que apenas se parecen de silueta."
            )
            familia_filtro = None if familia_elegida.startswith("No sé") else familia_elegida

            # PRIMERO leer el código grabado, después comparar formas. El orden importa: casi
            # todo repuesto trae el código estampado, impreso o troquelado, y leerlo da una
            # respuesta EXACTA. Comparar siluetas es adivinar. Tener las dos cosas separadas en
            # pantallas distintas hacía que se usara la peor sin necesidad.
            if bytes_consulta is not None:
                st.markdown("**Paso 1 — buscar un código en la foto**")
                st.caption(
                    "Casi todas las piezas traen el código grabado, impreso en una etiqueta o "
                    "moldeado. Si se llega a leer, la respuesta es exacta y no hay nada que adivinar."
                )
                if st.button("🔤 Leer el código de la foto", key="leer_codigo_visual"):
                    with st.spinner("Leyendo la pieza..."):
                        datos_ocr, err_ocr = identificar_pieza_por_foto(bytes_consulta)
                    st.session_state["ocr_visual"] = (datos_ocr, err_ocr)

                datos_ocr, err_ocr = st.session_state.get("ocr_visual", (None, None))
                if err_ocr:
                    st.info(f"{err_ocr} — igual podés comparar por parecido más abajo.")
                elif datos_ocr:
                    cod_leido = (datos_ocr.get("codigo") or "").strip()
                    conf_ocr = (datos_ocr.get("confianza") or "s/d").strip().lower()
                    if not cod_leido:
                        st.warning(
                            "No se pudo leer ningún código en la foto. Probá con más luz y de más "
                            "cerca, buscando el lado donde esté grabado. Si la pieza no tiene "
                            "código visible, usá la comparación por parecido de acá abajo."
                        )
                    else:
                        st.success(f"Código leído: **`{cod_leido}`** (la IA lo da con confianza "
                                    f"{conf_ocr})")
                        res_ocr = buscar_por_codigo(sanitizar(cod_leido))
                        if res_ocr:
                            st.success(f"✅ Está en tu catálogo — {len(res_ocr)} coincidencia(s):")
                            st.dataframe(quitar_id(res_ocr), use_container_width=True, hide_index=True)
                            st.caption("Esto es una coincidencia exacta de código, no un parecido.")
                        else:
                            # El OCR se come una letra seguido: es exactamente el caso que
                            # resuelve la búsqueda por tipeo.
                            parecidos_ocr = codigos_por_tipeo(sanitizar(cod_leido))
                            if parecidos_ocr:
                                st.warning(
                                    f"«{cod_leido}» no está en tu catálogo, pero hay códigos que se "
                                    "escriben casi igual. Leer un carácter de más o de menos es lo "
                                    "más común al leer un grabado, así que fijate si es alguno:"
                                )
                                st.dataframe(
                                    [{"Código": x["Codigo"], "Marca": x["Marca"],
                                      "Descripción": x["Descripcion"], "Precio": x["Precio"],
                                      "Stock": x["Stock"]} for x in parecidos_ocr],
                                    use_container_width=True, hide_index=True
                                )
                            else:
                                st.info(f"«{cod_leido}» no figura en tu catálogo ni se parece a nada "
                                         "cargado. Podés probar la comparación por parecido.")
                st.markdown("**Paso 2 — comparar por parecido** (si no se pudo leer el código)")

            if st.button("🖼️ Comparar con el catálogo", disabled=bytes_consulta is None,
                         type="primary", key="btn_comparar_visual"):
                barra_v = st.progress(0.0, text="Comparando...")
                res_visual, error_visual = buscar_por_similitud_visual(
                    bytes_consulta, familia=familia_filtro,
                    progreso=lambda i, t: barra_v.progress(min(i / max(t, 1), 1.0),
                                                           text=f"Comparando {i} de {t} fotos...")
                )
                barra_v.empty()
                st.session_state["resultado_visual"] = (res_visual, error_visual)

            res_visual, error_visual = st.session_state.get("resultado_visual", (None, None))
            if error_visual:
                st.info(error_visual)
            elif res_visual:
                st.warning(
                    f"⚠️ {len(res_visual)} candidato(s) ordenados de más a menos parecido. "
                    "NINGUNO está confirmado: hay repuestos que de foto son idénticos y no son "
                    "intercambiables (cambia el paso de rosca, la altura, el lado). "
                    "Verificá por código o comparando la pieza en la mano antes de vender."
                )
                filas_visual = []
                for r in res_visual:
                    filas_visual.append({
                        "Código": r["Codigo"], "Descripción": r["Descripcion"], "Marca": r["Marca"],
                        "Confianza": nivel_de_parecido(r), "Parecido": f"{r['Parecido']:.0f}",
                        "Detalles que coinciden": r["Coincidencias"],
                        "Precio": r["Precio"], "Stock": r["Stock"],
                    })
                st.dataframe(filas_visual, use_container_width=True, hide_index=True)
                explicar(
                    "«Detalles que coinciden» son los puntos de la pieza que además dieron geométricamente "
                    "coherentes entre las dos fotos:",
                    "es el número que más conviene mirar. Muchos detalles y parecido alto = vale la pena "
                    "revisarla. Si el que buscabas no aparece, cargale a ese producto una segunda foto del "
                    "ángulo que usás vos y la próxima vez lo encuentra.", en_expander=True
                )

    modo = st.radio("Buscar por:", ["Código", "Descripción"], horizontal=True, key="modo_busqueda")

    c.execute("SELECT nombre FROM marcas ORDER BY nombre")
    lista_marcas = ["Todas"] + [r["nombre"] for r in c.fetchall()]

    if modo == "Código":
        with st.form("form_buscar_codigo"):
            col_busq, col_filt = st.columns([3, 1])
            with col_busq:
                busqueda = st.text_input(
                    "Ingresá uno o varios códigos (separados por coma):",
                    placeholder="Ej: W712/94, 036115561G...",
                    key="busqueda_input"
                )
            with col_filt:
                marca_filtro = st.selectbox("Filtrar por marca:", lista_marcas)
            # La búsqueda encadena: el código buscado trae sus equivalentes, y los equivalentes
            # de esos, y así. Cuanto más larga la cadena, más chances de que un eslabón esté mal
            # y aparezcan cosas que no entran. Este control corta esa cadena.
            opciones_saltos = {
                "Solo los directos (más confiable)": 1,
                "Hasta 3 saltos (recomendado)": 3,
                "Toda la cadena": None,
            }
            etiqueta_saltos = st.radio(
                "Qué tan lejos buscar:", list(opciones_saltos.keys()),
                index=1, horizontal=True, key="saltos_busqueda",
                help="«Directo» es lo que alguna lista puso en la misma fila que tu código. Cada "
                     "salto más se apoya en el vínculo anterior: si uno está mal cargado, todo lo "
                     "que cuelga de ahí también."
            )
            max_saltos = opciones_saltos[etiqueta_saltos]
            buscar_click = st.form_submit_button("🔍 Buscar Equivalencias", type="primary")

        # La búsqueda en sí (con sus efectos de una sola vez: guardar historial, contar
        # veces_buscado) se hace acá, solo cuando se tocó "Buscar". El resultado se guarda en
        # session_state y el DESPLIEGUE se hace más abajo, FUERA de este "if", para que los
        # botones de adentro (agregar a WhatsApp, favoritos, combos) sigan funcionando en los
        # reruns siguientes — si el despliegue dependiera de "buscar_click", cualquier otro botón
        # que se toque después haría que buscar_click vuelva a False y todo el bloque desaparezca
        # antes de que el click en el botón de adentro llegue a registrarse.
        if buscar_click:
            # Sin sacar los repetidos, pegar "ABC, abc" dibujaba dos veces los mismos widgets
            # con la misma key —todas se arman con el código limpio— y Streamlit corta la app
            # entera con "multiple widgets with the same key". Y pegar una lista con un código
            # duplicado es lo más común del mundo.
            codigos_buscados, _ya_vistos = [], set()
            for _crudo in busqueda.split(","):
                _crudo = _crudo.strip()
                if not _crudo:
                    continue
                _clave_dedup = sanitizar(_crudo) or _crudo.upper()
                if _clave_dedup in _ya_vistos:
                    continue
                _ya_vistos.add(_clave_dedup)
                codigos_buscados.append(_crudo)
            if not codigos_buscados:
                st.info("Ingresá al menos un código válido para buscar.")
                st.session_state.pop("ultima_busqueda_codigo", None)
            else:
                guardar_busqueda(busqueda.strip())
                resultados_guardados = []
                for codigo_individual in codigos_buscados:
                    clean = sanitizar(codigo_individual)
                    if not clean:
                        resultados_guardados.append(
                            {"codigo_individual": codigo_individual, "clean": None, "res": None}
                        )
                        continue
                    res = buscar_por_codigo(clean, marca_filtro, max_saltos)
                    if res:
                        incrementar_veces_buscado(clean)
                    else:
                        registrar_busqueda_sin_resultado(codigo_individual)
                    resultados_guardados.append(
                        {"codigo_individual": codigo_individual, "clean": clean, "res": res}
                    )
                st.session_state["ultima_busqueda_codigo"] = resultados_guardados

        if st.session_state.get("ultima_busqueda_codigo"):
            catalogos = listar_catalogos_externos()
            total_codigos_buscados = len(st.session_state["ultima_busqueda_codigo"])
            for item in st.session_state["ultima_busqueda_codigo"]:
                codigo_individual = item["codigo_individual"]
                clean = item["clean"]
                res = item["res"]
                if not clean:
                    st.warning(f"🔎 {codigo_individual} — código no válido, se omitió.")
                    continue

                if res:
                    etiqueta_resultado = f"🔎 {codigo_individual} — {len(res)} coincidencia" + ("s" if len(res) != 1 else "")
                else:
                    etiqueta_resultado = f"🔎 {codigo_individual} — sin resultados"

                with st.expander(etiqueta_resultado, expanded=(total_codigos_buscados == 1)):
                    if res:
                        st.success(f"Se encontraron {len(res)} coincidencias:")


                        # Filtro de stock: un botón que ahorra scroll en cada consulta. Va
                        # antes de la tabla porque decide QUÉ se muestra, no cómo.
                        fs1, fs2 = st.columns([3, 2])
                        solo_stock = fs1.checkbox(
                            f"📦 Solo con stock ({sum(1 for f in res if (f.get('Stock') or 0) > 0)}"
                            f" de {len(res)})", key=f"solo_stock_{clean}")
                        if solo_stock:
                            con_stock = [f for f in res if (f.get("Stock") or 0) > 0]
                            if con_stock:
                                res = con_stock
                            else:
                                # Antes se volvía a la lista completa sin decir nada: se tildaba
                                # "solo con stock" y aparecían igual los que no tienen, así que
                                # parecía que el filtro estaba roto.
                                st.info("Ninguno de estos tiene stock. Se muestran todos.")

                        # Copiar el código para pegarlo en facturación o WhatsApp. st.code trae
                        # el botón de copiar incorporado, así que no hace falta JavaScript.
                        if fs2.checkbox("📋 Códigos para copiar", key=f"copiar_{clean}"):
                            st.code("\n".join(f["Codigo"] for f in res), language=None)

                        # Stock libre = lo que hay menos lo apartado en presupuestos. Es el
                        # número que importa al prometerle algo a un cliente; el stock a secas
                        # puede estar comprometido con otro que ya lo cotizó.
                        #
                        # Va en una columna aparte y NUMÉRICA a propósito. Antes se escribía
                        # "3 (de 4, el resto apartado)" encima de Stock, y estos diccionarios son
                        # los mismos que quedan guardados en session_state: al refresco siguiente
                        # Stock ya era un texto, el filtro de arriba hacía texto > 0 y la pantalla
                        # se cerraba con TypeError. Con una sola reserva activa alcanzaba con
                        # tocar cualquier botón de la búsqueda, y el texto además se anidaba en
                        # cada vuelta: "3 (de 3 (de 4, el resto apartado)...".
                        try:
                            libres = stock_libre_de_varios([f["ID"] for f in res])
                        except Exception:
                            libres = {}
                        hay_reservas = any(f["ID"] in libres and libres[f["ID"]] != (f.get("Stock") or 0)
                                            for f in res)
                        if hay_reservas:
                            for f in res:
                                f["Libre"] = libres.get(f["ID"], f.get("Stock") or 0)
                            st.caption("**Libre** es lo que queda sin apartar: es el número que "
                                       "se le puede prometer a un cliente.")

                        # El margen es información sensible: la ve el dueño y el administrador,
                        # no cualquiera que atienda el mostrador.
                        if es_admin():
                            mejor_marg = mejor_margen_entre_equivalentes(res)
                            agregar_margen(res)
                            if mejor_marg:
                                st.info(
                                    f"💰 **Mejor margen entre los que tenés en stock:** "
                                    f"{mejor_marg['marca']} {mejor_marg['codigo']} — te deja "
                                    f"${mejor_marg['ganancia']:,.0f} "
                                    f"(${mejor_marg['diferencia']:,.0f} más que el peor de la lista)."
                                )
                        puentes_res = puentes_en_el_resultado([f["ID"] for f in res])
                        # comparar precios a ojo cuando hay varias marcas equivalentes.
                        candidatos_precio = [f for f in res if f.get("Precio") and (f.get("Stock") or 0) > 0]
                        id_mas_barato = min(candidatos_precio, key=lambda f: f["Precio"])["ID"] if candidatos_precio else None
                        for f in res:
                            f["💰"] = "🏆 Más barato en stock" if f["ID"] == id_mas_barato else ""
                        # Las claves con guión bajo son internas —el costo entre ellas— y no
                        # salen a pantalla ni al Excel. Se sacan ACÁ, al dibujar, y no borrándolas
                        # de la fila: la fila se reusa en el refresco siguiente y borrarle datos
                        # es lo que hacía desaparecer el margen después del primer toque.
                        mostrar = quitar_id([{k: v for k, v in f.items() if not k.startswith("_")}
                                              for f in res])
                        st.dataframe(
                            mostrar, use_container_width=True, hide_index=True,
                            column_config={
                                "Imagen": st.column_config.ImageColumn("Imagen", width="small"),
                                "Ficha": st.column_config.LinkColumn("Ficha", display_text="Ver en proveedor ↗")
                            }
                        )

                        # La pregunta que sigue siempre a «¿lo tenés?»: ¿le sirve al auto del
                        # cliente? La app tiene el dato en los catálogos de fabricante y no lo
                        # usaba acá. Contestar mal esto es lo más caro del mostrador: la pieza
                        # vuelve, y con ella se va la confianza.
                        autos_cod = autos_de_un_codigo(clean)
                        if autos_cod:
                            if seccion_plegable(
                                    f"🚗 ¿Le sirve a qué autos? ({len(autos_cod)} aplicaciones)",
                                    key=f"aplic_{clean}"):
                                cver1, cver2, cver3 = st.columns([2, 2, 1])
                                marca_ver = cver1.text_input("Marca del auto:",
                                                              key=f"ver_marca_{clean}",
                                                              placeholder="Ej: FORD").strip()
                                modelo_ver = cver2.text_input("Modelo:", key=f"ver_modelo_{clean}",
                                                               placeholder="Ej: FIESTA").strip()
                                anio_ver = cver3.number_input("Año:", min_value=0, max_value=2100,
                                                               value=0, step=1,
                                                               key=f"ver_anio_{clean}")
                                if marca_ver:
                                    sirve, motivo = le_sirve_a_este_auto(
                                        clean, marca_ver, modelo_ver, anio_ver or None)
                                    if sirve is True:
                                        st.success(f"✅ **Sí le sirve.** {motivo}")
                                    elif sirve is False:
                                        st.error(f"❌ **No según el catálogo.** {motivo}")
                                    else:
                                        st.info(motivo)
                                st.caption("Todos los autos para los que el fabricante da este código:")
                                st.dataframe(autos_cod, use_container_width=True, hide_index=True)

                        # Los avisos de calidad van DESPUÉS de la tabla, y agrupados en un solo
                        # desplegable. Antes iban arriba: dos alertas rojas y el explorador de
                        # caminos empujaban el resultado —que es a lo que uno vino— abajo de todo.
                        # La advertencia sigue estando, pero deja de tapar la respuesta.
                        _hay_indirectos = any(
                            f.get("Cadena", "").startswith(("🟡", "🔴")) for f in res)
                        if puentes_res or _hay_indirectos:
                            if seccion_plegable(
                                "⚠️ Revisar la calidad de estos resultados" if puentes_res
                                else "🧭 ¿Por qué apareció alguno de estos?",
                                key=f"calidad_{clean}", abierto=bool(puentes_res)
                            ):
                                # Aviso de contaminación. Va acá y no solo en Mantenimiento porque es
                                # justo en el momento de vender cuando importa saber que estos resultados
                                # pueden no ser reales.
                                # De qué lista salió cada vínculo directo. Se agrega solo si alguna
                                # importación dejó rastro; las equivalencias viejas no lo tienen.
                                id_buscado = next((f["ID"] for f in res if sanitizar(f["Codigo"]) == clean), None)
                                if id_buscado:
                                    origenes = origenes_de_los_vinculos_directos(
                                        id_buscado, [f["ID"] for f in res]
                                    )
                                    if origenes:
                                        for f in res:
                                            lote_f = origenes.get(f["ID"])
                                            f["Vino de"] = lote_f.split(" · ")[0] if lote_f else ""

                                # Además del código puente clásico: el vínculo suelto que está uniendo
                                # dos familias enteras. No lo detecta ningún otro control, porque no hay
                                # ningún código con muchos enlaces — es un solo eslabón mal puesto.
                                if id_buscado and len(res) >= 8:
                                    try:
                                        uniones = vinculos_que_unen_familias(id_buscado)
                                    except Exception:
                                        uniones = []
                                    if uniones and uniones[0]["Confianza del vínculo"] < 50:
                                        u = uniones[0]
                                        st.error(
                                            f"🔗 **Un solo vínculo está uniendo dos grupos de repuestos.** "
                                            f"«{u['Código A']}» ({u['Marca A']}) ↔ «{u['Código B']}» "
                                            f"({u['Marca B']}) separa {u['Separa']}, y su confianza es "
                                            f"{u['Confianza del vínculo']}/100.\n\n"
                                            "Cortando **ese solo vínculo** se separan las dos familias. "
                                            "Está en **Estadísticas → Mantenimiento → Vínculos que unen "
                                            "familias**."
                                        )
                                if puentes_res:
                                    nombres = ", ".join(f"«{p['Código']}» ({p['Vínculos']} vínculos)"
                                                         for p in puentes_res[:3])
                                    st.error(
                                        f"🌉 **Ojo: estos resultados pasan por un código puente.** {nombres} "
                                        "está vinculado a demasiadas cosas, así que arrastra acá repuestos "
                                        "de otros rubros que no tienen nada que ver. Fijate la columna "
                                        "**Cadena**: lo marcado como 🟢 directo es lo confiable. "
                                        "Para arreglarlo de raíz: **Estadísticas → Mantenimiento → "
                                        "Códigos puente**."
                                    )

                                # ¿Por qué apareció este resultado? Muestra la cadena de vínculos que lo
                                # trajo, con el puntaje y la lista de origen de cada paso. Es lo que
                                # convierte "el camino es débil" en algo accionable: se ve cuál cortar.
                                # Sin expander anidado: Streamlit no los admite y tira excepción.
                                # Como este bloque YA está dentro de uno, va como subtítulo.
                                indirectos = [f for f in res if f.get("Cadena", "").startswith(("🟡", "🔴"))]
                                if id_buscado and indirectos:
                                    # Las keys llevan el código de ESTA vuelta. Eran fijas, y este
                                    # bloque está adentro del bucle que recorre los códigos
                                    # buscados: pidiendo dos a la vez ("P-1, Q-1") y abriendo esta
                                    # sección en los dos, Streamlit encontraba la misma key dos
                                    # veces y cerraba la app entera. Y buscar varios códigos
                                    # separados por coma es justo lo que el campo invita a hacer.
                                    if True:
                                        st.markdown("**🧭 ¿Por qué apareció alguno de estos?**")
                                        etiquetas_por_que = {
                                            f"{f['Codigo']} ({f['Marca']}) — {f['Cadena']}, {f['Confianza']}": f["ID"]
                                            for f in indirectos
                                        }
                                        elegido_pq = st.selectbox("Elegí un resultado:",
                                                                   list(etiquetas_por_que.keys()),
                                                                   key=f"por_que_resultado_{clean}")
                                        camino = camino_entre(id_buscado, etiquetas_por_que[elegido_pq])
                                        if not camino:
                                            st.caption("No pude reconstruir el camino.")
                                        else:
                                            st.caption(
                                                "Se muestra el camino MÁS confiable de los que existen. "
                                                "El paso con menor puntaje es el que decide si este "
                                                "resultado sirve o no."
                                            )
                                            peor_paso = min(camino, key=lambda x: x["Confianza"])
                                            for paso in camino:
                                                marca_paso = ("🔴" if paso is peor_paso and paso["Confianza"] < 50
                                                               else "🟢" if paso["Confianza"] >= 70
                                                               else "🟡")
                                                st.markdown(f"{marca_paso} **{paso['Paso']}** — "
                                                             f"{paso['Confianza']}/100 · {paso['Vino de']}")
                                            if peor_paso["Confianza"] < 50:
                                                st.warning(
                                                    f"El eslabón flojo es **{peor_paso['Paso']}** "
                                                    f"({peor_paso['Confianza']}/100). Cortando ese, este "
                                                    "resultado deja de aparecer."
                                                )
                                                if st.button("✂️ Cortar ese vínculo",
                                                              key=f"cortar_paso_debil_{clean}"):
                                                    borrar_equivalencias_dudosas(
                                                        [(peor_paso["_a"], peor_paso["_b"])])
                                                    invalidar_salud()
                                                    avisar("success", "Vínculo cortado.")
                                                    st.rerun()

                                # Marca la opción más barata ENTRE LAS QUE TIENEN STOCK, para no tener que

                        # Botones de link aparte, para no depender de scrollear la tabla al costado en el celular.
                        # La key incluye el código buscado (clean) además del ID: si se buscan varios códigos
                        # a la vez y dos están vinculados entre sí, el mismo producto puede aparecer en más de
                        # un resultado — sin el prefijo de clean, la key se repetiría y Streamlit tira error.
                        con_ficha = [f for f in res if f.get("Ficha")]
                        if con_ficha:
                            for f in con_ficha:
                                st.link_button(
                                    f"🔗 Ver {f['Codigo']} ({f['Marca']}) en el sitio del proveedor",
                                    f["Ficha"], key=f"link_ficha_{clean}_{f['ID']}"
                                )

                        # Combos: piezas que suelen cambiarse junto con lo que se encontró
                        combos_encontrados = {}
                        for f in res:
                            for disp, items in buscar_combos_para_descripcion(f.get("Descripcion", "")).items():
                                combos_encontrados.setdefault(disp, set()).update(items)
                        if combos_encontrados:
                            st.markdown("**💡 Suelen cambiarse junto con esto:**")
                            for disp, items_set in combos_encontrados.items():
                                items = sorted(items_set)
                                st.caption(f"Relacionado con: {disp}")
                                item_cols = st.columns(len(items))
                                for col_item, item in zip(item_cols, items):
                                    if col_item.button(f"🔍 {item}", key=f"combo_{clean}_{disp}_{item}"):
                                        res_item = buscar_por_texto(item)
                                        if res_item:
                                            con_stock = any((r.get("Stock") or 0) > 0 for r in res_item)
                                            if not con_stock:
                                                st.error(f"⚠️ Tenés '{item}' cargado pero SIN STOCK en ningún proveedor.")
                                            st.dataframe(quitar_id(res_item), use_container_width=True, hide_index=True)
                                        else:
                                            st.error(f"⚠️ No tenés '{item}' cargado en la base — vas a necesitar pedirlo.")

                        col_dl, col_add = st.columns(2)
                        with col_dl:
                            st.download_button(
                                "⬇️ Descargar (Excel)",
                                data=to_excel_bytes(mostrar),
                                file_name=f"equivalencias_{clean}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_{clean}"
                            )
                        with col_add:
                            if st.button("📋 Agregar a lista de WhatsApp", key=f"add_wa_{clean}"):
                                st.session_state.lista_whatsapp.append({
                                    "codigo_buscado": codigo_individual,
                                    "resultados": res
                                })
                                st.success("Agregado a la lista. Andá a la pestaña 'Lista WhatsApp' para armarla.")

                        st.markdown("**🛒 ¿Se lo llevó? / 📌 ¿Falta stock?**")
                        st.caption(
                            "Marcá cuál se llevó el cliente: con eso el sistema va aprendiendo qué "
                            "sirve para qué, y después te propone equivalencias nuevas en "
                            "Estadísticas → Equivalencias sugeridas. Si falta stock, 'Pedir' lo manda "
                            "a la lista de reposición."
                        )
                        for fila_stock in res:
                            colr1, colr2, colr3 = st.columns([3, 1, 1])
                            colr1.write(f"{fila_stock['Marca']} - {fila_stock['Codigo']} (stock actual: {fila_stock.get('Stock') if fila_stock.get('Stock') is not None else 's/d'})")
                            colr2.button("🛒 Se llevó", key=f"vendido_{fila_stock['ID']}_{clean}",
                                          on_click=registrar_venta, args=(fila_stock["ID"], codigo_individual),
                                          help="Anota la venta para ir descubriendo equivalencias solas")
                            colr3.button("📌 Pedir", key=f"pedir_repo_{fila_stock['ID']}_{clean}",
                                          on_click=solicitar_reposicion, args=(fila_stock["ID"],),
                                          help="Marcar para reposición")

                        # Marcar favoritos / editar precio y stock
                        if seccion_plegable("✏️ Marcar favorito / editar precio, costo y stock",
                                             key=f"editar_{clean}"):
                            for fila in res:
                                colF, colC, colP, colS, colG, colH = cols([0.5, 1.7, 1.1, 0.9, 0.7, 0.7])
                                es_fav = bool(fila.get("Favorito"))
                                nuevo_fav = colF.checkbox("⭐", value=es_fav, key=f"fav_{fila['ID']}_{clean}")
                                if nuevo_fav != es_fav:
                                    alternar_favorito(fila["ID"], nuevo_fav)
                                colC.write(f"{fila['Marca']} - {fila['Codigo']}")
                                nuevo_precio = colP.number_input(
                                    "Precio", value=float(fila.get("Precio") or 0),
                                    key=f"precio_{fila['ID']}_{clean}", min_value=0.0, step=100.0,
                                    label_visibility="collapsed"
                                )
                                nuevo_stock = colS.number_input(
                                    "Stock", value=int(fila.get("Stock") or 0),
                                    key=f"stock_{fila['ID']}_{clean}", min_value=0, step=1,
                                    label_visibility="collapsed"
                                )
                                if colG.button("💾", key=f"save_{fila['ID']}_{clean}"):
                                    actualizar_precio_stock(
                                        fila["ID"], nuevo_precio, nuevo_stock,
                                        st.session_state.get(f"costo_{fila['ID']}_{clean}"))
                                    st.success("Guardado.")
                                if es_admin():
                                    c.execute("SELECT precio_costo FROM productos WHERE id = ?",
                                              (fila["ID"],))
                                    _fc = c.fetchone()
                                    _costo_actual = float((_fc["precio_costo"] if _fc else 0) or 0)
                                    cc1, cc2 = st.columns([2, 3])
                                    cc1.number_input(
                                        "Costo", value=_costo_actual, min_value=0.0, step=100.0,
                                        key=f"costo_{fila['ID']}_{clean}",
                                        help="Lo que te cuesta a vos. Solo lo ve el administrador."
                                    )
                                    _pct, _pesos = margen_de(nuevo_precio,
                                                              st.session_state.get(
                                                                  f"costo_{fila['ID']}_{clean}"))
                                    if _pct is not None:
                                        cc2.caption(f"Margen: **{_pct:.0f}%** "
                                                     f"(${_pesos:,.0f} por unidad)")
                                    else:
                                        cc2.caption("Cargá el costo para ver el margen.")

                                if colH.button("📈", key=f"hist_precio_{fila['ID']}_{clean}", help="Ver historial de precio"):
                                    st.session_state[f"mostrar_hist_{fila['ID']}"] = True
                                if st.session_state.get(f"mostrar_hist_{fila['ID']}"):
                                    historial_p = historial_precio_producto(fila["ID"])
                                    if historial_p:
                                        st.dataframe(historial_p, use_container_width=True, hide_index=True)
                                    else:
                                        st.caption("Todavía no hay cambios de precio registrados para este producto.")

                        if catalogos:
                            st.caption("Buscar este código también en:")
                            # OJO con el nombre: 'cols' es una función de esta app (arma columnas
                            # que se apilan bien en el celular). Llamar a la variable igual la
                            # pisaba a nivel global, y a partir de ahí cualquier cols(3) posterior
                            # reventaba con "'list' object is not callable". Solo saltaba cuando la
                            # marca tenía catálogo cargado, por eso era intermitente.
                            columnas_catalogos = st.columns(len(catalogos))
                            for col, cat in zip(columnas_catalogos, catalogos):
                                with col:
                                    st.link_button(f"🌐 {cat['nombre']}", cat["url"],
                                                    use_container_width=True, key=f"link_{cat['id']}_{clean}")
                    else:
                        st.warning("No hay ningún producto con ese código exacto.")

                        # ¿Este código fue reemplazado por otro? Es la causa más tonta de
                        # perder una venta: la pieza existe, cambió de número, y se le dice al
                        # cliente que no se fabrica más.
                        _cadena_r = cadena_de_reemplazos(clean)
                        if _cadena_r:
                            _ultimo = _cadena_r[-1]
                            # La variable acá se llama codigo_individual: es el código que se
                            # está resolviendo en esta vuelta del bucle, no el texto del campo.
                            _ruta = " → ".join([str(codigo_individual).strip()] +
                                                [x["codigo"] for x in _cadena_r])
                            st.success(
                                f"🔄 **Ese código fue reemplazado.** {_ruta}\n\n"
                                f"El vigente es **{_ultimo['codigo']}**"
                                + (f" ({_ultimo['marca']})" if _ultimo["marca"] else "")
                                + (f". {_ultimo['nota']}" if _ultimo["nota"] else ".")
                            )
                            _res_nuevo = buscar_por_codigo(_ultimo["clean"])
                            if _res_nuevo:
                                st.dataframe(quitar_id(_res_nuevo), use_container_width=True,
                                              hide_index=True)
                            else:
                                st.caption("Tampoco tenés cargado el código nuevo.")

                        # Antes de darse por vencido: ¿lo conoce algún catálogo de fabricante?
                        # Ahí figura qué pieza es y a qué autos le va, y con eso se puede
                        # ofrecer un equivalente en vez de perder la venta.
                        ajeno = identificar_codigo_ajeno(clean)
                        if ajeno:
                            st.info(
                                f"🏭 **Ese código existe: es {ajeno['tipo'] or 'una pieza'} de "
                                f"{ajeno['marca']}**, y según su catálogo le va a "
                                f"{ajeno['cantidad_autos']} auto(s). No lo tenés cargado."
                            )
                            # Sin expander: este bloque ya está dentro del expander de
                            # resultados por código, y Streamlit no admite anidarlos.
                            if ajeno["autos"]:
                                if seccion_plegable(
                                        f"Ver a qué autos le va ({len(ajeno['autos'])})",
                                        key=f"autos_ajeno_{clean}"):
                                    st.dataframe(
                                        [{"Marca": a["marca_auto"], "Modelo": a["modelo_auto"],
                                          "Motor": a["motor"],
                                          "Años": f"{a['anio_desde'] or ''}-{a['anio_hasta'] or ''}"}
                                         for a in ajeno["autos"]],
                                        use_container_width=True, hide_index=True
                                    )
                            reemplazos = equivalentes_para_los_mismos_autos(clean)
                            if reemplazos:
                                st.success(
                                    f"✅ **Tenés {len(reemplazos)} repuesto(s) que sirven para los "
                                    "mismos autos**, según el catálogo de sus propios fabricantes:"
                                )
                                st.dataframe(reemplazos, use_container_width=True, hide_index=True)
                                st.caption(
                                    "Los que tienen stock van primero. Confirmá con el cliente "
                                    "el modelo y el año antes de cerrar: el catálogo dice a qué "
                                    "autos le va la pieza, no si es exactamente la que traía."
                                )

                        # Error de tipeo. Se muestra ANTES que las familias porque es lo más
                        # probable: en el mostrador se tipea rápido y se dan vuelta dos dígitos.
                        # Lo que ya había solo encontraba códigos que empiezan igual, así que
                        # W71249 por W71294 no daba nada aunque fuera el mismo filtro.
                        tipeos = codigos_por_tipeo(clean)
                        if tipeos:
                            st.info(f"⌨️ ¿No será alguno de estos? Se escriben casi igual a lo que "
                                     f"pusiste ({len(tipeos)} encontrado(s)):")
                            st.dataframe(
                                [{"Código": t["Codigo"], "Marca": t["Marca"],
                                  "Descripción": t["Descripcion"], "Precio": t["Precio"],
                                  "Stock": t["Stock"],
                                  "Diferencia": "1 carácter" if t["_dist"] == 1 else "2 caracteres"}
                                 for t in tipeos],
                                use_container_width=True, hide_index=True
                            )

                        # Familias de códigos: pedís "TC-421" y en la base están "TC-421-15",
                        # "TC-421-20", etc. Antes esto no aparecía por ningún lado.
                        parecidos = buscar_codigos_parecidos(clean)
                        if parecidos:
                            st.info(
                                f"🔎 Pero hay {len(parecidos)} código(s) que empiezan igual o lo "
                                "contienen — puede ser una familia con variantes (espesor, lado, medida):"
                            )
                            for p in parecidos[:12]:
                                equivalentes_p = buscar_por_codigo(p["_clean"])
                                otros_p = [e for e in equivalentes_p if e["ID"] != p["ID"]]
                                resumen_p = (f"{len(otros_p)} equivalencia(s)" if otros_p
                                              else "sin equivalencias cargadas")
                                if seccion_plegable(f"🔎 {p['Marca']} · {p['Codigo']} — {resumen_p}",
                                                     key=f"detalle_{clean}_{p['ID']}"):
                                    if p.get("Descripcion"):
                                        st.caption(p["Descripcion"])
                                    if equivalentes_p:
                                        st.dataframe(quitar_id(equivalentes_p),
                                                      use_container_width=True, hide_index=True)
                                    else:
                                        st.caption("Todavía no tiene equivalencias cargadas.")
                                    st.button("🔎 Abrir este código como búsqueda",
                                               key=f"abrir_parecido_{p['ID']}",
                                               on_click=cb_ver_equivalencias, args=(p["Codigo"],))
                            if len(parecidos) > 12:
                                st.caption(f"(mostrando 12 de {len(parecidos)})")
                        else:
                            parcial = buscar_por_texto(clean)
                            if parcial:
                                st.info("¿Quisiste decir alguno de estos? Tocá el código para ver sus equivalencias:")
                                mostrar_lista_clickeable(parcial, f"sug_{clean}", limite=12)
    else:
        with st.expander("🎙️ Buscar por voz"):
            st.caption(
                "Grabá diciendo lo que buscás — la IA lo transcribe y lo busca con el buscador de "
                "siempre. No es un asistente que entienda pedidos complejos, es simplemente hablar "
                "en vez de tipear."
            )
            audio_busqueda = st.audio_input("Grabar:", key="audio_busqueda_voz")
            if audio_busqueda and st.button("🔍 Transcribir y buscar"):
                with st.spinner("Transcribiendo..."):
                    mime_audio = audio_busqueda.type or "audio/wav"
                    texto_voz, error_voz = transcribir_audio(audio_busqueda.getvalue(), mime_audio)
                if error_voz:
                    st.error(error_voz)
                else:
                    st.session_state["texto_desde_voz"] = texto_voz
                    st.rerun()

        # Precargar el texto transcripto ANTES de crear el widget del form — si se hace después
        # de que ya se dibujó en pantalla, Streamlit tira un error.
        if "texto_desde_voz" in st.session_state:
            st.session_state["texto_input"] = st.session_state.pop("texto_desde_voz")

        with st.form("form_buscar_texto"):
            texto = st.text_input(
                "Ingresá parte de una descripción:",
                placeholder="Ej: ruleman delantero gol (no hace falta el orden exacto)",
                key="texto_input"
            )
            buscar_texto_click = st.form_submit_button("🔍 Buscar por Descripción", type="primary")

        # Igual que en la búsqueda por código: los resultados se guardan en la sesión en vez de
        # depender de que el botón "Buscar" haya sido lo último que se tocó. Si no, al apretar
        # cualquier botón de adentro el bloque entero desaparece y el click se pierde.
        if buscar_texto_click:
            if not texto.strip():
                st.info("Ingresá un texto para buscar.")
                st.session_state.pop("ultima_busqueda_texto", None)
            else:
                guardar_busqueda(texto.strip())
                st.session_state["ultima_busqueda_texto"] = {
                    "texto": texto.strip(), "res": buscar_por_texto(texto)
                }

        busqueda_texto_guardada = st.session_state.get("ultima_busqueda_texto")
        if busqueda_texto_guardada:
            res_texto = busqueda_texto_guardada["res"]
            texto_pedido = busqueda_texto_guardada["texto"]
            if res_texto:
                st.success(f"Se encontraron {len(res_texto)} coincidencia(s):")
                st.dataframe(quitar_id(res_texto), use_container_width=True, hide_index=True)
                mostrar_lista_clickeable(
                    res_texto, "txt_click", limite=15,
                    nota="👆 Tocá cualquier código para abrirlo con todas sus equivalencias:"
                )

                # La búsqueda por descripción solo hace coincidir texto: encuentra el producto,
                # pero no sus equivalentes. Acá se abre la red de equivalencias de cada resultado,
                # igual que hace la búsqueda por código, para no perder de vista las otras marcas.
                st.markdown("**🔗 Equivalencias de cada resultado**")
                st.caption("Cada uno abre las otras marcas que sirven, con precio y stock.")
                for fila_txt in res_texto[:15]:
                    clean_txt = sanitizar(fila_txt["Codigo"])
                    equivalentes = buscar_por_codigo(clean_txt) if clean_txt else []
                    otros = [e for e in equivalentes if e["ID"] != fila_txt["ID"]]
                    resumen = (f"{len(otros)} equivalencia(s)" if otros else "sin equivalencias cargadas")
                    with st.expander(f"🔎 {fila_txt['Marca']} · {fila_txt['Codigo']} — {resumen}"):
                        if fila_txt.get("Descripcion"):
                            st.caption(fila_txt["Descripcion"])
                        if equivalentes:
                            candidatos_precio = [f for f in equivalentes
                                                  if f.get("Precio") and (f.get("Stock") or 0) > 0]
                            id_barato = (min(candidatos_precio, key=lambda f: f["Precio"])["ID"]
                                          if candidatos_precio else None)
                            for f in equivalentes:
                                f["💰"] = "🏆 Más barato en stock" if f["ID"] == id_barato else ""
                            st.dataframe(quitar_id(equivalentes), use_container_width=True, hide_index=True)
                        else:
                            st.caption("Este producto todavía no tiene equivalencias cargadas.")

                        # Carrito: ir juntando de varias búsquedas y armar el presupuesto al
                        # final. Sin esto había que anotar los códigos aparte y volver a
                        # buscarlos uno por uno.
                        #
                        # OJO con los nombres: este bloque vive adentro de la búsqueda por
                        # DESCRIPCIÓN, donde el resultado de esta vuelta es fila_txt y sus
                        # equivalentes están en 'otros'. Estaba escrito con los nombres de la
                        # búsqueda por CÓDIGO (res, clean), que acá no existen, así que entrar a
                        # "Descripción" tiraba NameError y la sección entera quedaba en blanco.
                        # La clave de los widgets va por ID y no por código limpio: dos filas
                        # distintas pueden limpiar al mismo texto y ahí Streamlit corta la app
                        # por claves repetidas.
                        st.session_state.setdefault("carrito", {})
                        opciones_carrito = [fila_txt] + otros
                        cc1, cc2 = st.columns([3, 2])
                        agregar_cod = cc1.selectbox(
                            "🛒 Sumar al presupuesto:",
                            ["(elegir)"] + [f"{f['Marca']} {f['Codigo']}" for f in opciones_carrito],
                            key=f"al_carrito_{fila_txt['ID']}")
                        if agregar_cod != "(elegir)" and cc2.button(
                                "➕ Sumar", key=f"btn_carrito_{fila_txt['ID']}"):
                            elegido_c = next((f for f in opciones_carrito
                                               if f"{f['Marca']} {f['Codigo']}" == agregar_cod), None)
                            if elegido_c:
                                st.session_state["carrito"][elegido_c["ID"]] = {
                                    "codigo": elegido_c["Codigo"], "marca": elegido_c["Marca"],
                                    "descripcion": elegido_c.get("Descripcion") or "",
                                    "precio": elegido_c.get("Precio") or 0, "cantidad": 1,
                                }
                                avisar("success", f"{elegido_c['Codigo']} sumado al presupuesto.")
                                st.rerun()

                        # Apartar mientras el cliente lo piensa. Va acá y no en otra pantalla
                        # porque el momento de reservar es este: con el presupuesto recién hecho.
                        # OJO con los nombres: acá el código limpio es clean_txt y la lista es
                        # 'equivalentes'. Usar los de la búsqueda por código ('clean', 'res')
                        # tira NameError y corta el renderizado de toda la pantalla.
                        _para_apartar = equivalentes or [fila_txt]
                        if seccion_plegable("🔒 Apartar para un presupuesto",
                                             key=f"apartar_txt_{fila_txt['ID']}"):
                            # Una sola consulta para todos: preguntarlo de a uno acá eran dos
                            # por fila para filtrar y dos más por fila para armar la etiqueta.
                            libres_ap = stock_libre_de_varios([f["ID"] for f in _para_apartar])
                            con_stock_ap = [f for f in _para_apartar
                                             if (libres_ap.get(f["ID"]) or 0) > 0]
                            if not con_stock_ap:
                                st.caption("Ninguno de estos tiene stock libre para apartar.")
                            else:
                                etq_ap = {
                                    f"{f['Marca']} {f['Codigo']} — {libres_ap[f['ID']]} libre(s)":
                                        f["ID"] for f in con_stock_ap
                                }
                                elegido_ap = st.selectbox("¿Cuál?", list(etq_ap.keys()),
                                                           key=f"cual_apartar_txt_{fila_txt['ID']}")
                                ap1, ap2 = st.columns([1, 3])
                                cant_ap = ap1.number_input(
                                    "Cantidad", min_value=1, value=1, step=1,
                                    key=f"cant_apartar_txt_{fila_txt['ID']}")
                                cliente_ap = ap2.text_input(
                                    "¿Para quién?", key=f"cliente_apartar_txt_{fila_txt['ID']}",
                                    placeholder="Nombre o patente, para saber a quién reclamarle")
                                if st.button("🔒 Apartar", key=f"btn_apartar_txt_{fila_txt['ID']}"):
                                    ok_ap, msg_ap = reservar_stock(
                                        etq_ap[elegido_ap], cant_ap, cliente_ap)
                                    if ok_ap:
                                        avisar("success", msg_ap)
                                        st.rerun()
                                    else:
                                        st.error(msg_ap)

                        st.markdown("**🛒 ¿Cuál se llevó el cliente?**")
                        st.caption(
                            "Marcá el que se lleva — vale también si se lleva un equivalente y no "
                            "el que apareció en la búsqueda. Así el sistema aprende esa relación."
                        )
                        for f in (equivalentes or [fila_txt]):
                            cv1, cv2 = st.columns([4, 1])
                            precio_txt = f"${f['Precio']:,.0f}" if f.get("Precio") else "s/precio"
                            cv1.write(f"{f['Marca']} - {f['Codigo']} ({precio_txt}, "
                                       f"stock: {f.get('Stock') if f.get('Stock') is not None else 's/d'})")
                            cv2.button("🛒 Se llevó", key=f"vendido_txt_{fila_txt['ID']}_{f['ID']}",
                                        on_click=registrar_venta, args=(f["ID"], texto_pedido))
                if len(res_texto) > 15:
                    st.caption(f"(mostrando las primeras 15 de {len(res_texto)} — afiná la búsqueda "
                                "para ver menos resultados)")
            else:
                st.warning("No se encontraron productos con esa descripción.")

                # Antes de rendirse: entender qué pidió el cliente. En el mostrador nadie dice
                # un código, dice «pastillas para un Gol 1.6». De esa frase se puede sacar el
                # rubro y el auto, y con eso ofrecer algo en vez de nada.
                pedido = interpretar_pedido_hablado(texto_pedido)
                if pedido and (pedido["familia"] or pedido["marca_auto"]):
                    partes = []
                    if pedido["familia"]:
                        partes.append(f"**{pedido['familia']}**")
                    if pedido["marca_auto"]:
                        auto = pedido["marca_auto"]
                        if pedido["modelo"]:
                            auto += f" {pedido['modelo']}"
                        if pedido["cilindrada"]:
                            auto += f" {pedido['cilindrada']}"
                        partes.append(f"para un **{auto}**")
                    st.info("🤔 Entendí que buscás " + " ".join(partes) + ".")

                    alternativas = buscar_por_pieza_y_auto(
                        pedido["familia"], pedido["marca_auto"], pedido["modelo"],
                        pedido["cilindrada"])
                    if alternativas:
                        st.success(f"Esto es lo que tenés que puede servir ({len(alternativas)}):")
                        st.dataframe(quitar_id(alternativas), use_container_width=True,
                                      hide_index=True)
                        st.caption("Con stock primero. Confirmá el modelo y el año antes de cerrar.")
                    elif pedido["marca_auto"]:
                        # Aflojar el filtro: quizá no tenés ese rubro para ese auto, pero sí algo
                        sin_familia = buscar_por_pieza_y_auto(
                            None, pedido["marca_auto"], pedido["modelo"], pedido["cilindrada"])
                        if sin_familia:
                            st.info(
                                f"No tenés {pedido['familia'] or 'esa pieza'} para ese auto, pero "
                                f"sí {len(sin_familia)} repuesto(s) de otros rubros:"
                            )
                            st.dataframe(quitar_id(sin_familia[:20]),
                                          use_container_width=True, hide_index=True)

    with st.expander("📦 Armar pedido (ordenado por ubicación en depósito)"):
        st.caption(
            "Pegá varios códigos separados por coma — te devuelve la lista ordenada por ubicación "
            "en el depósito, para juntar todo en un solo recorrido en vez de ir y volver."
        )
        codigos_picking = st.text_input(
            "Códigos del pedido (separados por coma):",
            placeholder="Ej: W712/94, 036115561G, 24427...", key="picking_codigos"
        )
        if st.button("📦 Ordenar para picking"):
            if not codigos_picking.strip():
                st.info("Pegá al menos un código.")
            else:
                res_picking = armar_lista_picking(codigos_picking)
                if res_picking:
                    sin_ubicacion = [r for r in res_picking if not r["Ubicación"]]
                    st.success(f"Se encontraron {len(res_picking)} de los códigos pedidos:")
                    st.dataframe(res_picking, use_container_width=True, hide_index=True)
                    if sin_ubicacion:
                        st.caption(
                            f"⚠️ {len(sin_ubicacion)} producto(s) todavía no tienen ubicación cargada "
                            "(aparecen al final) — cargala desde 'Administrar' para que la próxima vez "
                            "el orden sea completo."
                        )
                else:
                    st.warning("No encontré ninguno de esos códigos en el catálogo.")

    with st.expander("📐 Buscar por medidas mecánicas (cuando no hay código ni equivalencia cargada)"):
        st.caption(
            "Para piezas de autos antiguos, importados o fuera de catálogo: medí la pieza rota con un "
            "calibre y buscá alternativas que compartan esas cotas, aunque no tengan equivalencia registrada."
        )
        cm1, cm2, cm3 = cols(3)
        m_diam_int = cm1.number_input("Diámetro interno (mm)", min_value=0.0, step=0.1, value=0.0, key="med_di")
        m_diam_ext = cm2.number_input("Diámetro externo (mm)", min_value=0.0, step=0.1, value=0.0, key="med_de")
        m_ancho = cm3.number_input("Ancho (mm)", min_value=0.0, step=0.1, value=0.0, key="med_an")
        cm4, cm5, cm6 = cols(3)
        m_paso = cm4.text_input("Paso de rosca (opcional)", key="med_paso", placeholder="Ej: M12x1.5")
        m_estrias = cm5.number_input("Cantidad de estrías (opcional)", min_value=0, step=1, value=0, key="med_estrias")
        m_tolerancia = cm6.slider("Tolerancia (%)", min_value=1, max_value=15, value=5, key="med_tol")

        st.markdown("**↔️ Segunda cara (opcional, para piezas con distinta medida de cada lado)**")
        st.caption(
            "Ej: un retén con labio interior de un diámetro de un lado y otro del otro, o un tensor "
            "con el interior escalonado (17mm de una cara, 8mm de la otra)."
        )
        cb1, cb2 = st.columns(2)
        m_diam_int_b = cb1.number_input("Diámetro interno cara B (mm)", min_value=0.0, step=0.1, value=0.0, key="med_di_b")
        m_diam_ext_b = cb2.number_input("Diámetro externo / labio exterior cara B (mm)", min_value=0.0, step=0.1,
                                         value=0.0, key="med_de_b")

        st.markdown("**🔩 Homocinéticas (opcional)**")
        ch1, ch2 = st.columns(2)
        m_estrias_int = ch1.number_input("Estrías internas", min_value=0, step=1, value=0, key="med_estrias_int")
        m_estrias_ext = ch2.number_input("Estrías externas", min_value=0, step=1, value=0, key="med_estrias_ext")
        ch3, ch4 = st.columns(2)
        m_seguro = ch3.text_input("Posición del seguro", key="med_seguro", placeholder="Ej: 1er ranura, a 12mm")
        m_abs = ch4.selectbox("¿Tiene ABS?", ["Cualquiera", "Sí", "No"], key="med_abs")
        ch5, ch6 = st.columns(2)
        m_rosca_homo = ch5.number_input("Diámetro de rosca (mm)", min_value=0.0, step=0.1, value=0.0, key="med_rosca_homo")
        m_largo_total = ch6.number_input(
            "Largo total (mm)", min_value=0.0, step=0.5, value=0.0, key="med_largo_total",
            help="De punta a punta — es la medida que más rápido descarta."
        )
        st.caption("La copa es cónica: cargá los dos diámetros si los tenés (base y boca).")
        ch7, ch8 = st.columns(2)
        m_copa = ch7.number_input("Diám. copa — base (mm)", min_value=0.0, step=0.1, value=0.0, key="med_copa")
        m_copa_sup = ch8.number_input("Diám. copa — boca / superior (mm)", min_value=0.0, step=0.1,
                                       value=0.0, key="med_copa_sup")

        if st.button("📐 Buscar por medidas"):
            res_medidas = buscar_por_medidas(
                m_diam_int or None, m_diam_ext or None, m_ancho or None,
                m_paso or None, m_estrias or None, m_tolerancia,
                m_estrias_int or None, m_estrias_ext or None, m_seguro or None, m_abs,
                m_diam_int_b or None, m_diam_ext_b or None,
                m_rosca_homo or None, m_copa or None,
                m_copa_sup or None, m_largo_total or None
            )
            if res_medidas:
                st.success(f"Se encontraron {len(res_medidas)} pieza(s) con medidas compatibles:")
                st.dataframe(quitar_id(res_medidas), use_container_width=True, hide_index=True)
            else:
                st.warning(
                    "Sin resultados. Puede ser que no haya piezas con esas medidas cargadas todavía — "
                    "cargalas desde la pestaña 'Administrar' a medida que las vayas midiendo."
                )

    if es_operador_o_admin():
        with st.expander("📷 Identificar pieza por foto (con IA)"):
            st.caption(
                "Sacale una foto a la pieza o subí una que ya tengas. La IA busca un código visible "
                "y, si lo encuentra, lo busca directo en tu catálogo."
            )
            foto = recordar_archivo(st.file_uploader(
                "Foto de la pieza:", type=["png", "jpg", "jpeg"], key="foto_identificar_pieza",
                label_visibility="collapsed"
            ), "foto_pieza")
            foto_ok = archivo_listo(foto, "foto")
            if foto and st.button("🗑️ Usar otra foto", key="otra_foto_pieza"):
                olvidar_archivo("foto_pieza")
                st.rerun()
            if st.button("🔍 Identificar", disabled=not foto_ok):
                with st.spinner("Consultando..."):
                    datos_pieza, error = identificar_pieza_por_foto(foto.getvalue())
                if error:
                    st.error(error)
                elif datos_pieza:
                    st.session_state["datos_pieza_foto"] = datos_pieza
                    st.session_state["buscar_tipo_pieza_click"] = False

            if st.session_state.get("datos_pieza_foto"):
                datos_pieza = st.session_state["datos_pieza_foto"]
                codigo_detectado = (datos_pieza.get("codigo") or "").strip()
                confianza = (datos_pieza.get("confianza") or "").strip().lower()

                if codigo_detectado:
                    st.success(f"**Código detectado: `{codigo_detectado}`** (confianza de la IA: {confianza or 's/d'})")
                    if confianza in ("media", "baja"):
                        st.caption(
                            "⚠️ La propia IA no está muy segura de haber leído bien el código — "
                            "confirmalo mirando la pieza antes de vender."
                        )
                else:
                    st.warning("No se distinguió ningún código legible en la foto.")
                if datos_pieza.get("marca_visible"):
                    st.caption(f"Marca visible en la pieza: {datos_pieza['marca_visible']}")
                if datos_pieza.get("tipo_pieza"):
                    st.caption(f"Tipo de pieza (según la IA): {datos_pieza['tipo_pieza']}")

                if codigo_detectado:
                    clean_foto = sanitizar(codigo_detectado)
                    res_foto = buscar_por_codigo(clean_foto) if clean_foto else []
                    if res_foto:
                        incrementar_veces_buscado(clean_foto)
                        st.success(f"✅ Coincidencia CONFIRMADA en tu catálogo — {len(res_foto)} resultado(s):")
                        st.caption(
                            "Esto es un match exacto por código, con las equivalencias que ya tenés "
                            "cargadas — no es una suposición de la IA."
                        )
                        st.dataframe(quitar_id(res_foto), use_container_width=True, hide_index=True)
                    else:
                        st.info(
                            f"El código `{codigo_detectado}` no coincide con nada cargado — puede que la "
                            "IA haya leído mal algún carácter, o que sea un código que todavía no tenés."
                        )
                        if datos_pieza.get("tipo_pieza") and st.button("🔍 Buscar por el tipo de pieza en vez del código"):
                            st.session_state["buscar_tipo_pieza_click"] = True

                if (not codigo_detectado or (codigo_detectado and st.session_state.get("buscar_tipo_pieza_click"))) \
                        and datos_pieza.get("tipo_pieza"):
                    if not codigo_detectado:
                        mostrar_tipo = st.button("🔍 Buscar por el tipo de pieza")
                    else:
                        mostrar_tipo = True
                    if mostrar_tipo:
                        tipo_pieza_texto = datos_pieza["tipo_pieza"]
                        res_tipo = buscar_por_texto(tipo_pieza_texto)
                        busqueda_usada = tipo_pieza_texto
                        if not res_tipo:
                            # La frase completa no encontró nada — reintenta con menos palabras
                            # (más amplio), por si tus descripciones no usan las mismas palabras
                            # exactas que eligió la IA (ej: "rótula de suspensión" vs "ROTULA DERECHA").
                            palabras_tipo = tipo_pieza_texto.split()
                            for n in range(len(palabras_tipo) - 1, 0, -1):
                                intento = " ".join(palabras_tipo[:n])
                                res_tipo = buscar_por_texto(intento)
                                if res_tipo:
                                    busqueda_usada = intento
                                    break
                        if res_tipo:
                            if busqueda_usada != tipo_pieza_texto:
                                st.caption(
                                    f"No encontré nada con \"{tipo_pieza_texto}\" completo — probé de nuevo "
                                    f"solo con \"{busqueda_usada}\" y esto apareció (todavía menos preciso, "
                                    "revisá con más cuidado):"
                                )
                            if len(res_tipo) > 1:
                                st.warning(
                                    f"⚠️ Encontré {len(res_tipo)} pieza(s) parecida(s) por palabras clave — "
                                    "NINGUNA está confirmada como la exacta, es solo una búsqueda por texto. "
                                    "Si son piezas como rótulas, retenes, etc. que varían por modelo de auto, "
                                    "comparalas físicamente (o por medidas, en '📐 Buscar por medidas mecánicas') "
                                    "antes de vender la que sea."
                                )
                            else:
                                st.info(
                                    "Encontré 1 coincidencia por palabras clave — tampoco está confirmada, "
                                    "revisala antes de vender."
                                )
                            st.dataframe(quitar_id(res_tipo)[:15], use_container_width=True, hide_index=True)
                            if len(res_tipo) > 15:
                                st.caption(f"Mostrando las primeras 15 de {len(res_tipo)} coincidencias.")
                        else:
                            st.caption("No encontré nada parecido en la base por ese tipo de pieza.")

    historial = historial_reciente()
    if historial:
        st.caption("🕘 Búsquedas recientes:")
        cols_hist = st.columns(min(len(historial), 5))
        for i, termino in enumerate(historial[:5]):
            if cols_hist[i % 5].button(termino, key=f"sugerencia_hist_{i}_{termino}", use_container_width=True):
                st.session_state["sugerencia_busqueda"] = termino
                st.rerun()

    favoritos = listar_favoritos()
    if favoritos:
        with st.expander(f"⭐ Favoritos ({len(favoritos)})"):
            for fila_fav in favoritos[:8]:
                colf1, colf2 = st.columns([4, 1])
                colf1.write(f"{fila_fav.get('Codigo') or ''} — {fila_fav.get('Marca') or ''}")
                if colf2.button("🔍", key=f"sugerencia_fav_{fila_fav['ID']}"):
                    st.session_state["sugerencia_busqueda"] = fila_fav.get("Codigo") or ""
                    st.rerun()
            st.dataframe(quitar_id(favoritos), use_container_width=True, hide_index=True)

# ============================================================
# VINCULAR MANUAL
# ============================================================
def vincular_grupo_equivalencias(productos_info, nivel, nota, verificar):
    """Crea (o reutiliza) cada producto de la lista y los vincula a TODOS entre sí — así se puede
    armar de una un grupo de equivalencias con productos de varios proveedores distintos,
    en vez de tener que ir vinculando de a pares."""
    ids = []
    with db_lock:
        for p in productos_info:
            clean = sanitizar(p["codigo"])
            marca_id = get_or_create_marca(p["marca"])
            pid = get_or_create_producto(
                p["codigo"].strip(), clean, p.get("descripcion", "").strip(),
                marca_id, p.get("imagen_url", "").strip() or None
            )
            ids.append(pid)
        v = 1 if verificar else 0
        pares = 0
        for i in range(len(ids)):
            for j in range(len(ids)):
                if i == j or ids[i] == ids[j]:
                    continue
                c.execute(
                    "INSERT OR REPLACE INTO equivalencias "
                    "(producto_a_id, producto_b_id, created_at, verificada, nivel, nota) "
                    "VALUES (?, ?, datetime('now'), ?, ?, ?)",
                    (ids[i], ids[j], v, nivel, nota.strip())
                )
                pares += 1
        conn.commit()
    return len(set(ids)), pares


if pagina == PAGINAS[1]:
    st.subheader("Vincular varios códigos como equivalentes")
    explicar(
        "Armá un grupo de códigos — de la marca/proveedor que sea, se pueden mezclar — y "
        "vinculalos todos entre sí de una sola vez.",
        "Antes había que hacerlo de a pares; ahora si tenés 5 productos de 5 proveedores "
        "distintos que son lo mismo, los sumás todos a la tanda y los vinculás juntos."
    )

    c.execute("SELECT id, nombre FROM marcas ORDER BY nombre")
    nombres_marcas = [m["nombre"] for m in c.fetchall()]

    if "grupo_equivalencia" not in st.session_state:
        st.session_state["grupo_equivalencia"] = []

    # Si se vino desde "Productos sin equivalencias" (pestaña Administrar) con un código para
    # precargar, hay que fijar estos valores ANTES de crear los widgets de abajo — si se hace
    # después de que ya se dibujaron en pantalla, Streamlit tira un error.
    if "vincular_pendiente" in st.session_state:
        pendiente = st.session_state.pop("vincular_pendiente")
        st.session_state["nuevo_codigo_grupo"] = pendiente.get("cod_a", "")
        st.session_state["nueva_desc_grupo"] = pendiente.get("desc_a", "")
        if pendiente.get("marca_a") in nombres_marcas:
            st.session_state["nueva_marca_opcion_grupo"] = pendiente["marca_a"]

    st.markdown("**➕ Agregar un código a la tanda**")
    cg1, cg2 = st.columns(2)
    nuevo_codigo = cg1.text_input("Código", key="nuevo_codigo_grupo")
    marca_opcion = cg2.selectbox("Marca", nombres_marcas + ["➕ Nueva marca..."], key="nueva_marca_opcion_grupo")
    nueva_marca = st.text_input("Nombre de la nueva marca", key="nueva_marca_texto_grupo") \
        if marca_opcion == "➕ Nueva marca..." else marca_opcion
    cg3, cg4 = st.columns(2)
    nueva_desc = cg3.text_input("Descripción (opcional)", key="nueva_desc_grupo")
    nueva_img = cg4.text_input("URL de foto (opcional)", key="nueva_img_grupo", placeholder="https://...")

    if st.button("➕ Agregar a la tanda"):
        clean = sanitizar(nuevo_codigo)
        if not clean:
            st.warning("Completá el código.")
        elif not nueva_marca or not nueva_marca.strip():
            st.warning("Completá la marca.")
        else:
            ya_esta = any(
                sanitizar(it["codigo"]) == clean and it["marca"].strip().upper() == nueva_marca.strip().upper()
                for it in st.session_state["grupo_equivalencia"]
            )
            if ya_esta:
                st.warning("Ese código con esa marca ya está en la tanda.")
            else:
                st.session_state["grupo_equivalencia"].append({
                    "codigo": nuevo_codigo.strip(), "marca": nueva_marca.strip(),
                    "descripcion": nueva_desc.strip(), "imagen_url": nueva_img.strip()
                })
                for k in ["nuevo_codigo_grupo", "nueva_desc_grupo", "nueva_img_grupo"]:
                    st.session_state.pop(k, None)
                st.rerun()

    grupo = st.session_state["grupo_equivalencia"]
    if grupo:
        st.markdown(f"**📋 Tanda actual ({len(grupo)} código{'s' if len(grupo) != 1 else ''}):**")
        for i, it in enumerate(grupo):
            colg1, colg2 = st.columns([5, 1])
            colg1.write(f"{it['marca']}: {it['codigo']}" + (f" — {it['descripcion']}" if it['descripcion'] else ""))
            colg2.button("🗑️", key=f"quitar_grupo_{i}",
                          on_click=quitar_item_lista_sesion, args=("grupo_equivalencia", i))

        st.markdown("---")
        nivel_equiv = st.selectbox(
            "Nivel de equivalencia (aplica a todo el grupo):",
            ["Exacta", "Reemplazo con modificación", "Solo alternativa de menor calidad"],
            help="Qué tan intercambiables son en la práctica."
        )
        nota_tecnica = st.text_input(
            "Nota técnica (opcional, aplica a todo el grupo):",
            placeholder="Ej: Equivale pero requiere cambiar la ficha eléctrica"
        )
        verificar = st.checkbox("✅ Marcar como verificada", value=True,
                                 help="Verificada = confirmaste vos mismo que son intercambiables.")

        if len(grupo) < 2:
            st.info("Agregá al menos 2 códigos a la tanda para poder vincularlos.")
        elif st.button(f"🔗 Vincular los {len(grupo)} códigos entre sí", type="primary"):
            cantidad_prod, cantidad_pares = vincular_grupo_equivalencias(grupo, nivel_equiv, nota_tecnica, verificar)
            st.success(f"Listo: {cantidad_prod} productos quedaron vinculados entre sí ({cantidad_pares} relaciones creadas).")
            st.session_state["grupo_equivalencia"] = []
            st.rerun()
    else:
        st.caption("Todavía no agregaste ningún código a la tanda.")

# ============================================================
# CARGAR EXCEL
# ============================================================
if pagina == PAGINAS[2]:
    if not pedir_password_admin("cargar listas de proveedores"):
        pass
    else:
        st.subheader("Cargar nueva planilla (.xlsx / .csv / .pdf)")

        # El proveedor casi siempre está en el nombre del archivo ("ILLINOIS 17 07 2026.xlsx").
        # Se propone solo, pero queda editable: nunca se pisa lo que la persona haya escrito.
        _archivo_previo = st.session_state.get("_archivo_lista")
        if _archivo_previo and not st.session_state.get("nombre_prov_carga"):
            c.execute("SELECT nombre FROM marcas")
            _marcas_conocidas = [r["nombre"] for r in c.fetchall()]
            _sugerido_prov = adivinar_proveedor(_archivo_previo["nombre"], _marcas_conocidas)
            if _sugerido_prov:
                st.session_state["nombre_prov_carga"] = _sugerido_prov

        nombre_prov = st.text_input(
            "Nombre de la Marca / Proveedor:", placeholder="Ej: Mahle, Bosch, Mann...",
            key="nombre_prov_carga",
            help="Se propone solo a partir del nombre del archivo. Corregilo si no acertó."
        )

        metodo = st.radio(
            "¿Cómo querés indicar el archivo?",
            ["Subir archivo", "Escribir la ruta en el teléfono"],
            horizontal=True,
            help="Si el botón de subir no responde en el navegador del celular, usá la opción de ruta."
        )

        archivo = None

        if metodo == "Subir archivo":
            # La primera subida que llegue bien queda guardada, así no se pierde si el widget
            # se vacía por un refresco o un corte de conexión (el clásico "lo subo y no lo toma").
            # .txt y .tsv también: muchos proveedores exportan así, y hasta ahora el selector
            # ni siquiera los dejaba elegir aunque la app sabe leerlos.
            archivo = subir_archivo("Seleccioná el archivo",
                                     ["xlsx", "xlsm", "csv", "txt", "tsv", "pdf"], "lista")
            if archivo:
                ca1, ca2 = st.columns([3, 1])
                with ca1:
                    archivo_listo(archivo, "archivo")
                with ca2:
                    boton_otro_archivo("lista")

                # ¿Es un archivo que generó la propia app? Importarlo mete de vuelta lo que
                # se había separado a propósito.
                _generado, _que_es = es_archivo_generado_por_la_app(getattr(archivo, "name", ""))
                if _generado:
                    st.error(
                        f"🔄 **Este archivo lo generó la propia app**: es {_que_es}. "
                        "No es una lista de proveedor.\n\n"
                        "Importarlo vuelve a meter en la base justo lo que se había separado, y "
                        "además crea una marca con el nombre del archivo. Si en tus marcas ves "
                        "algo como «FILAS OMITIDAS», salió de acá."
                    )

                # ¿Esta misma planilla ya se importó? Se compara el CONTENIDO, no el nombre:
                # el archivo suele llegar renombrado y así se reconoce igual.
                previa = importacion_previa(huella_de_archivo(archivo.getvalue()))
                if previa:
                    st.warning(
                        f"🔁 Esta misma planilla ya se importó el "
                        f"**{str(previa['fecha'])[:16]}** como **{previa['marca']}** "
                        f"({previa['filas_cargadas']} filas). Si la volvés a importar vas a "
                        "revisar de nuevo los mismos vínculos, y los precios que hayas corregido "
                        "a mano desde entonces se pisan con los de esta lista."
                    )
            else:
                archivo_listo(None, "archivo")
            if archivo and archivo.name.lower().endswith(".pdf"):
                st.caption(
                    "📄 PDF: funciona mejor con catálogos que tienen tablas reales (no una imagen escaneada). "
                    "Revisá bien la vista previa antes de importar, el resultado puede variar según el PDF."
                )
        else:
            st.caption(
                "Ejemplo: /storage/emulated/0/Download/lista.xlsx "
                "(si el archivo está en Descargas, esa es la ruta de siempre)."
            )
            ruta_archivo = st.text_input("Ruta completa del archivo en el teléfono:",
                                          placeholder="/storage/emulated/0/Download/lista.xlsx")
            EXTENSIONES_OK = (".xlsx", ".xlsm", ".csv", ".txt", ".tsv", ".pdf")
            if ruta_archivo:
                import os
                if not os.path.isfile(ruta_archivo):
                    st.error("No se encontró un archivo en esa ruta. Revisá que esté bien escrita.")
                elif not ruta_archivo.lower().endswith(EXTENSIONES_OK):
                    st.error("El archivo debe terminar en " + ", ".join(EXTENSIONES_OK))
                else:
                    archivo = ruta_archivo

        # --- Mapeo dinámico de columnas ---
        # OJO: estos valores por defecto NO son decorativos. Los selectores se crean adentro de
        # un "else" (solo si el archivo tiene columnas), pero el bloque de importar está afuera:
        # sin estas líneas, un archivo sin columnas hace que el importador rompa con NameError.
        todas_filas = None
        idx_prov = idx_oem = 0
        idx_desc = None
        idx_precio = idx_stock = None
        tope_salto = 200

        hoja_elegida = None
        if archivo:
            # Selector de hoja. Antes se leía siempre la que quedó activa al guardar el archivo,
            # que muchas veces es la de instrucciones o una en blanco: la lista no se importaba
            # y no había ninguna señal de por qué.
            hojas = hojas_del_excel(archivo)
            if len(hojas) > 1:
                etiquetas_hojas = {f"{n} ({f} filas)": n for n, f in hojas}
                sugerida = max(hojas, key=lambda h: h[1])[0]
                indice_sug = [n for n, _ in hojas].index(sugerida)
                elegida_lbl = st.selectbox(
                    f"El archivo tiene {len(hojas)} hojas — ¿cuál es la lista?",
                    list(etiquetas_hojas.keys()), index=indice_sug, key="hoja_excel",
                    help="Se propone la que más filas tiene."
                )
                hoja_elegida = etiquetas_hojas[elegida_lbl]
            try:
                todas_filas = leer_excel(archivo, nrows=200, hoja=hoja_elegida)
                if isinstance(archivo, object) and not isinstance(archivo, str):
                    archivo.seek(0)
            except Exception as e:
                st.error(f"No se pudo leer el archivo: {e}")
                todas_filas = None

        if todas_filas:
            # Detectar automáticamente la fila de encabezado como punto de partida
            # Fila de títulos: la detecta mirando la forma de la tabla, no solo esta fila.
            # Queda EDITABLE porque ninguna detección acierta siempre, y equivocarse acá arruina
            # toda la importación: desalinea el mapeo y mete las filas de arriba como productos.
            header_auto = detectar_fila_encabezado(todas_filas)
            header_row = st.number_input(
                "Fila donde están los títulos de las columnas:",
                min_value=1, max_value=max(len(todas_filas), 1), value=header_auto + 1,
                step=1, key="fila_encabezado",
                help="Se detecta sola. Corregila si la vista previa no muestra los títulos "
                     "correctos en la primera fila."
            ) - 1
            encabezado = todas_filas[header_row]
            preview_filas = todas_filas[header_row:header_row + 6]

            st.write("Vista previa (primeras filas detectadas):")
            st.dataframe(preview_filas, use_container_width=True)

            if len(encabezado) < 1:
                st.error("El archivo no tiene ninguna columna con datos.")
            else:
                st.markdown("**Mapeo de columnas** — revisá que coincida con tu archivo (se sugiere automáticamente):")
                sugerido = adivinar_columnas(encabezado)
                # Si el encabezado no tiene títulos de verdad (la lista arranca directo en los
                # datos), se deduce mirando los valores. Sin esto la detección caía en la
                # columna 0, que en muchas listas está vacía, y no entraba ni un producto.
                titulos_utiles = sum(1 for x in encabezado
                                     if x is not None and str(x).strip() and not str(x).strip()[0].isdigit())
                if titulos_utiles < 2:
                    por_datos = adivinar_columnas_por_datos(
                        todas_filas[header_row + 1:header_row + 60], len(encabezado)
                    )
                    if por_datos["prov"] is not None:
                        sugerido = por_datos
                        st.info(
                            "ℹ️ Esta lista no trae títulos de columna, así que deduje qué es cada "
                            "una mirando los datos. **Revisá la tabla de más abajo** antes de importar."
                        )
                idx_prov_auto = sugerido["prov"]
                idx_oem_auto = sugerido["oem"] if sugerido["oem"] is not None else min(1, len(encabezado) - 1)
                idx_desc_auto = sugerido["desc"]
                idx_precio_sug, idx_stock_sug = sugerido["precio"], sugerido["stock"]

                opciones_cols = [f"Columna {i}: {str(v)[:20] if v else '(sin título)'}"
                                  for i, v in enumerate(encabezado)]

                # Si ya se importó una lista de este proveedor, se arranca con el mismo mapeo
                # que funcionó la vez anterior en vez de tener que acertarle de nuevo.
                mapeo_previo = leer_mapeo_columnas(nombre_prov)
                if mapeo_previo:
                    st.success(
                        f"💾 Se recordó cómo mapeaste las columnas la última vez que importaste "
                        f"una lista de **{nombre_prov.strip().upper()}** ({mapeo_previo['fecha'][:10]}). "
                        "Ya viene preseleccionado — revisá que coincida con este archivo."
                    )
                    def _valido(indice):
                        return indice if (indice is not None and 0 <= indice < len(opciones_cols)) else None
                    if _valido(mapeo_previo["idx_prov"]) is not None:
                        idx_prov_auto = mapeo_previo["idx_prov"]
                    idx_oem_auto = _valido(mapeo_previo["idx_oem"])
                    idx_desc_auto = _valido(mapeo_previo["idx_desc"])

                c_p, c_o, c_d = cols(3)
                with c_p:
                    idx_prov = st.selectbox("Código Proveedor:", range(len(opciones_cols)),
                                             format_func=lambda x: opciones_cols[x], index=idx_prov_auto)
                with c_o:
                    opciones_oem = [None] + list(range(len(opciones_cols)))
                    idx_oem = st.selectbox(
                        "Código OEM / Equivalente:", opciones_oem,
                        format_func=lambda x: "Ninguna (la lista no lo trae)" if x is None else opciones_cols[x],
                        index=opciones_oem.index(idx_oem_auto),
                        help="Si la lista no tiene columna de OEM, elegí 'Ninguna': los productos se "
                             "cargan igual y quedan buscables, solo que sin equivalencia."
                    )
                with c_d:
                    opciones_desc = [None] + list(range(len(opciones_cols)))
                    idx_default_desc = opciones_desc.index(idx_desc_auto) if idx_desc_auto is not None else 0
                    idx_desc = st.selectbox("Descripción (opcional):", opciones_desc,
                                             format_func=lambda x: "Ninguna" if x is None else opciones_cols[x],
                                             index=idx_default_desc)

                c_pr, c_st = cols(2)
                opciones_num = [None] + list(range(len(opciones_cols)))
                # El mapeo que ya funcionó con este proveedor manda; si no hay, lo detectado
                # por el título de la columna.
                idx_precio_auto = (_valido(mapeo_previo["idx_precio"]) if mapeo_previo else None)
                if idx_precio_auto is None:
                    idx_precio_auto = idx_precio_sug
                idx_stock_auto = (_valido(mapeo_previo["idx_stock"]) if mapeo_previo else None)
                if idx_stock_auto is None:
                    idx_stock_auto = idx_stock_sug
                with c_pr:
                    idx_precio = st.selectbox(
                        "Precio (opcional):", opciones_num,
                        format_func=lambda x: "Ninguna" if x is None else opciones_cols[x],
                        index=opciones_num.index(idx_precio_auto) if idx_precio_auto is not None else 0,
                        help="Si la elegís, la lista actualiza los precios al importar."
                    )
                with c_st:
                    idx_stock = st.selectbox(
                        "Stock (opcional):", opciones_num,
                        format_func=lambda x: "Ninguna" if x is None else opciones_cols[x],
                        index=opciones_num.index(idx_stock_auto) if idx_stock_auto is not None else 0
                    )

                if idx_precio is not None:
                    st.session_state.setdefault("tope_salto_precio", 200)
                    tope_salto = st.select_slider(
                        "Frenar precios que cambien más de:",
                        options=[50, 100, 200, 400, 800, 0],
                        format_func=lambda x: "No frenar ninguno" if x == 0 else f"{x}%",
                        key="tope_salto_precio",
                        help="Un precio que se multiplica o se divide de golpe casi nunca es un "
                             "aumento: es la columna equivocada, el separador de decimales al "
                             "revés, o un precio por bulto donde iba el unitario. Los que superen "
                             "este límite se cargan igual como producto, pero el precio queda "
                             "frenado para que lo mires vos."
                    )
                else:
                    tope_salto = 200

                # --- Diagnóstico ANTES de importar ---
                # Es la respuesta a "se carga mal y no sé por qué": muestra qué entendió la app
                # de cada columna con valores reales del archivo, y cuántas filas se van a
                # perder y por qué motivo. Todo esto antes se descubría después de importar.
                st.markdown("**🔎 Qué está entendiendo la app de tu lista**")
                diag = diagnosticar_lista(todas_filas, header_row, idx_prov, idx_oem, idx_desc)

                if diag["total"] == 0:
                    st.error(
                        "No hay ninguna fila de datos debajo de los títulos. Revisá la fila de "
                        "encabezado de más arriba: si apunta a la última fila con texto, no queda "
                        "nada para importar."
                    )
                else:
                    resumen = []
                    for etiqueta, idx, tipo in (("Código de proveedor", idx_prov, "codigo"),
                                                 ("Código de fábrica", idx_oem, "codigo"),
                                                 ("Descripción", idx_desc, "descripcion")):
                        if idx is None:
                            resumen.append({"Campo": etiqueta, "Columna": "— sin asignar",
                                            "Diagnóstico": "", "Ejemplos del archivo": ""})
                            continue
                        clave = {"Código de proveedor": "ejemplos_prov",
                                 "Código de fábrica": "ejemplos_oem",
                                 "Descripción": "ejemplos_desc"}[etiqueta]
                        estado, motivo = _pinta_columna(diag[clave], tipo)
                        resumen.append({
                            "Campo": etiqueta,
                            "Columna": opciones_cols[idx] if idx < len(opciones_cols) else str(idx),
                            "Diagnóstico": estado + (f" — {motivo}" if motivo else ""),
                            "Ejemplos del archivo": " · ".join(diag[clave][:3]) or "(vacío)",
                        })
                    st.dataframe(resumen, use_container_width=True, hide_index=True)

                    d1, d2, d3 = cols(3)
                    d1.metric("Filas que entran", diag["ok"])
                    d2.metric("Sin código", diag["sin_codigo"] + diag["codigo_basura"])
                    d3.metric("Con código de fábrica", diag["con_oem"])

                    porcentaje_ok = diag["ok"] / max(diag["total"], 1)
                    if porcentaje_ok < 0.5:
                        st.error(
                            f"🔴 Solo entrarían {diag['ok']} de {diag['total']} filas de la muestra. "
                            "Eso casi siempre significa que el mapeo apunta a la columna equivocada "
                            "o que la fila de encabezado está mal. **Mirá los ejemplos de arriba**: "
                            "si en «Código de proveedor» ves descripciones o cantidades en vez de "
                            "códigos, ahí está el problema. No importes así."
                        )
                    elif diag["codigo_basura"]:
                        st.warning(
                            f"⚠️ {diag['codigo_basura']} fila(s) tienen en la columna de código algo "
                            "que no es un código (números sueltos de 1 o 2 dígitos). Se van a saltear."
                        )
                    else:
                        st.success(f"✅ Entrarían {diag['ok']} de {diag['total']} filas de la muestra.")

                    # Antes que cualquier otra cosa: avisar si este archivo no es una lista de
                    # precios sino un catálogo de aplicaciones. Importarlo acá carga los modelos
                    # de auto como códigos de repuesto, y eso ensucia la base entera.
                    es_aplic, motivo_aplic = parece_catalogo_de_aplicaciones(todas_filas)
                    if es_aplic:
                        st.error(
                            "🏭 **Esto no parece una lista de precios, sino un catálogo de "
                            f"aplicaciones** ({motivo_aplic}).\n\n"
                            "Si lo importás acá, se van a cargar los **modelos de auto como si "
                            "fueran códigos de repuesto** (A4, Q3, Golf...) y los años como "
                            "precios.\n\n"
                            "Para este archivo andá a **Estadísticas → Mantenimiento → "
                            "🏭 Catálogo de aplicaciones**: ahí se lee bien y sirve para que la "
                            "búsqueda por vehículo sepa qué repuesto le va a cada auto."
                        )

                    # Aviso clave: una lista sin columna de código de fábrica no puede generar
                    # equivalencias reales. Si igual se activa "buscarlas en la descripción",
                    # lo que sale son modelos de auto, medidas y cilindradas tomados por códigos,
                    # y cada uno de esos vincula entre sí todas las filas donde aparece.
                    if idx_oem is None:
                        st.warning(
                            "🔗 **Esta lista no trae columna de código de fábrica**, así que no "
                            "puede generar equivalencias: cada fila es un producto suelto con su "
                            "precio. Eso está perfecto para **cargar y actualizar precios**.\n\n"
                            "Lo que NO conviene es activar «buscar el código de fábrica dentro de "
                            "la descripción» en una lista así. En descripciones como "
                            "«...CRUZE AVEO ASTRA - 1.4/1.8 - F14D4 Z18XER» lo que se extrae son "
                            "modelos de motor, medidas y cilindradas, no códigos — y cada valor "
                            "repetido vincula entre sí todas las filas donde aparece. **Es la "
                            "forma más rápida de llenar la base de equivalencias falsas.**"
                        )

                    if diag["fechas"]:
                        st.error(
                            f"📅 **{diag['fechas']} código(s) llegaron convertidos en fecha** "
                            f"(por ejemplo: {', '.join(diag['ejemplos_fechas'][:3])}). "
                            "Esto lo hace Excel solo: códigos como «12-15», «3/8» o «8-10» los "
                            "toma por fechas al guardar el archivo, y ya no hay forma de saber "
                            "cuál era el original.\n\n"
                            "**Cómo arreglarlo:** pedile la lista al proveedor en **.csv**, o abrí "
                            "el Excel, seleccioná la columna de códigos, ponela en formato "
                            "**Texto** y volvé a pegar los datos. Estas filas se van a saltear: "
                            "es preferible eso a cargar un código inventado que nunca va a coincidir."
                        )

                    if diag["columnas_desiguales"] > diag["total"] * 0.3:
                        st.warning(
                            f"⚠️ {diag['columnas_desiguales']} fila(s) tienen bastantes menos "
                            "columnas que el encabezado. Suele pasar cuando la planilla trae "
                            "subtítulos por rubro en el medio, o cuando la fila de encabezado "
                            "elegida no es la correcta."
                        )
                    if diag["vacias"]:
                        st.caption(f"({diag['vacias']} fila(s) vacías, se ignoran solas)")

                st.markdown("---")
                st.markdown("**Cuando la lista no trae el código de fábrica**")
                buscar_oem_en_desc = st.checkbox(
                    "🔎 Buscar códigos de fábrica dentro de la descripción",
                    value=bool(mapeo_previo["buscar_oem_en_desc"]) if mapeo_previo else False,
                    help="Muchas listas meten el OEM en el texto ('... ORIG 6Q0407365'). Esto lo saca de ahí. "
                         "Es conservador a propósito: ante la duda no lo toma, para no ensuciar la base."
                )
                prov_es_oem = st.checkbox(
                    "🏭 El código del proveedor YA es el código de fábrica",
                    value=bool(mapeo_previo["prov_es_oem"]) if mapeo_previo else False,
                    help="Para listas donde el proveedor usa directamente el código original. Se carga "
                         "también como OEM y quedan vinculados, así engancha con las listas de otros proveedores."
                )

                if buscar_oem_en_desc and idx_desc is not None:
                    muestras = []
                    for fila_prev in todas_filas[header_row + 1:header_row + 60]:
                        texto_desc = valor_o_vacio(fila_prev[idx_desc]) if idx_desc < len(fila_prev) else ""
                        hallados = extraer_codigos_de_texto(texto_desc)
                        if hallados:
                            muestras.append({"Descripción": texto_desc[:60], "Detecta": ", ".join(hallados)})
                        if len(muestras) >= 8:
                            break
                    if muestras:
                        st.caption("Así quedaría (muestra de las primeras filas) — revisá antes de importar:")
                        st.dataframe(muestras, use_container_width=True, hide_index=True)
                    else:
                        st.caption("En las primeras filas no encontré códigos dentro de la descripción.")
                elif buscar_oem_en_desc and idx_desc is None:
                    st.warning("Para buscar códigos en la descripción, elegí primero la columna de descripción.")

        st.markdown("**🔒 Antes de importar: ¿qué hacemos con las equivalencias?**")
        cargar_directo = st.radio(
            "Equivalencias de esta lista:",
            ["Mandarlas a revisar (recomendado)", "Cargarlas directo"],
            key="modo_carga_equivalencias", label_visibility="collapsed",
            help="Una lista puede generar miles de vínculos de una sola vez."
        ) == "Cargarlas directo"
        if cargar_directo:
            st.warning(
                "⚠️ Se van a cargar sin revisar. Si la columna de código de fábrica tiene algún "
                "error, esos vínculos equivocados quedan en la base y después es difícil encontrarlos. "
                "Usalo solo con listas de proveedores en las que confiés plenamente."
            )
        else:
            explicar(
                "Los productos y precios se cargan igual y quedan buscables enseguida.",
                "Lo único que espera son las **equivalencias**: te esperan en Estadísticas → 🔗 "
                "Equivalencias sugeridas, ya separadas entre las limpias y las que tienen algo raro."
            )

        procesar = st.button("📥 Procesar e Importar Lista", type="primary")

        if procesar:
            if not archivo:
                st.warning("Indicá un archivo primero (subilo o escribí su ruta).")
            elif not nombre_prov.strip():
                st.warning("Ingresá el nombre de la marca / proveedor.")
            elif not todas_filas:
                st.warning("No se pudo leer el archivo, revisá el formato.")
            elif len(encabezado) < 1:
                st.warning("El archivo no tiene ninguna columna con datos.")
            else:
                try:
                    # La MISMA fila que se eligió arriba. Si acá se volviera a detectar por
                    # separado, la vista previa y la carga real podrían leer filas distintas y
                    # uno importaría algo diferente de lo que vio.
                    header_row = max(0, st.session_state.get("fila_encabezado", 1) - 1)

                    # Releer completo (leer_excel con nrows=200 antes era solo para la vista previa)
                    # La MISMA hoja que se eligió arriba: si acá se releyera sin ese dato, se
                    # importaría una hoja distinta de la que se vio en la vista previa.
                    todas_filas_completas = leer_excel(archivo, hoja=hoja_elegida)
                    filas_datos = todas_filas_completas[header_row + 1:]

                    cargados = 0
                    cargados_sin_equiv = 0
                    omitidos = 0
                    descartados_cortos = 0
                    precios_actualizados = 0
                    precios_frenados = []
                    filas_omitidas = []
                    eq_batch = set()  # inserción en lote: se acumulan los pares y se insertan todos juntos al final
                    progreso = st.progress(0, text="Procesando filas...")
                    total = len(filas_datos)

                    # Todo el trabajo de escritura va con el candado tomado, para que ninguna otra
                    # persona pueda buscar/escribir a mitad de una importación larga y quede todo trabado.
                    with db_lock:
                        prov_id = get_or_create_marca(nombre_prov, "PROVEEDOR")
                        oem_id = get_or_create_marca("OEM / FABRICA", "OEM")

                        for n, fila in enumerate(filas_datos):
                            def celda(idx):
                                return fila[idx] if idx is not None and idx < len(fila) else None

                            raw_p_cell = valor_codigo(celda(idx_prov))
                            raw_o_cell = valor_codigo(celda(idx_oem)) if idx_oem is not None else ""
                            desc = separar_texto_pegado(valor_o_vacio(celda(idx_desc)))

                            # OJO con el fallback: antes, si dividir_codigos no devolvía nada se
                            # usaba la celda cruda igual — así que un '1' suelto entraba lo mismo.
                            # Ahora la celda cruda también tiene que pasar el filtro.
                            codigos_prov = dividir_codigos(raw_p_cell)
                            if not codigos_prov and raw_p_cell and es_codigo_util(raw_p_cell):
                                codigos_prov = [raw_p_cell]
                            codigos_oem = dividir_codigos(raw_o_cell)
                            if not codigos_oem and raw_o_cell and es_codigo_util(raw_o_cell):
                                codigos_oem = [raw_o_cell]

                            # Para poder avisar al final cuántos se filtraron por ser números sueltos
                            if raw_p_cell and not codigos_prov:
                                descartados_cortos += 1
                            if raw_o_cell and not codigos_oem:
                                descartados_cortos += 1

                            # Si la lista no trae OEM, se intenta sacarlo de la descripción
                            if not codigos_oem and buscar_oem_en_desc and desc:
                                codigos_oem = extraer_codigos_de_texto(desc)

                            # Sin código de proveedor no hay nada que cargar: esa fila sí se omite
                            if not codigos_prov:
                                omitidos += 1
                                filas_omitidas.append({"Proveedor": raw_p_cell, "OEM": raw_o_cell,
                                                        "Descripcion": desc, "Motivo": "sin código de proveedor"})
                                if total and n % 25 == 0:
                                    progreso.progress(min((n + 1) / total, 1.0))
                                continue

                            precio_fila = leer_numero(celda(idx_precio)) if idx_precio is not None else None
                            stock_fila = leer_numero(celda(idx_stock)) if idx_stock is not None else None

                            ids_prov = []
                            for raw_p in codigos_prov:
                                clean_p = sanitizar(raw_p)
                                if clean_p:
                                    pid_nuevo = get_or_create_producto(raw_p, clean_p, desc, prov_id)
                                    ids_prov.append(pid_nuevo)
                                    if precio_fila is not None or stock_fila is not None:
                                        c.execute("SELECT precio FROM productos WHERE id = ?", (pid_nuevo,))
                                        _f = c.fetchone()
                                        precio_viejo = _f["precio"] if _f else None
                                        raro, motivo = salto_de_precio_sospechoso(
                                            precio_viejo, precio_fila, tope_salto
                                        )
                                        if raro:
                                            # El producto se carga igual; lo único que no se pisa
                                            # es el precio, para no cotizar con un número roto.
                                            precios_frenados.append({
                                                "Código": raw_p, "Descripción": (desc or "")[:60],
                                                "Precio actual": precio_viejo,
                                                "Precio de la lista": precio_fila,
                                                "Motivo": motivo,
                                                "_id": pid_nuevo,
                                            })
                                            if stock_fila is not None:
                                                c.execute("UPDATE productos SET stock = ? WHERE id = ?",
                                                          (int(stock_fila), pid_nuevo))
                                        else:
                                            if precio_fila is not None and precio_viejo != precio_fila:
                                                c.execute("INSERT INTO historial_precios (producto_id, precio) "
                                                          "VALUES (?, ?)", (pid_nuevo, precio_fila))
                                                precios_actualizados += 1
                                            c.execute(
                                                "UPDATE productos SET "
                                                "precio = COALESCE(?, precio), stock = COALESCE(?, stock) "
                                                "WHERE id = ?",
                                                (precio_fila,
                                                 int(stock_fila) if stock_fila is not None else None,
                                                 pid_nuevo)
                                            )

                            if not ids_prov:
                                omitidos += 1
                                filas_omitidas.append({"Proveedor": raw_p_cell, "OEM": raw_o_cell,
                                                        "Descripcion": desc, "Motivo": "código de proveedor no válido"})
                                if total and n % 25 == 0:
                                    progreso.progress(min((n + 1) / total, 1.0))
                                continue

                            # El código del proveedor ES el de fábrica: se carga también como OEM
                            # para que enganche con las listas de otros proveedores.
                            if prov_es_oem:
                                for raw_p in codigos_prov:
                                    clean_p = sanitizar(raw_p)
                                    if clean_p and clean_p not in {sanitizar(x) for x in codigos_oem}:
                                        codigos_oem.append(raw_p)

                            ids_oem = []
                            for raw_o in codigos_oem:
                                clean_o = sanitizar(raw_o)
                                if clean_o:
                                    ids_oem.append(get_or_create_producto(raw_o, clean_o, desc, oem_id))

                            # ANTES: sin OEM se descartaba la fila entera y el producto ni se cargaba.
                            # Ahora el producto queda cargado igual (buscable por código y por
                            # descripción), solo que sin equivalencia hasta que aparezca de otra lista
                            # o se vincule a mano.
                            if not ids_oem:
                                cargados_sin_equiv += len(ids_prov)
                                # aunque no haya OEM, los códigos de la misma celda se vinculan entre sí
                                for pid in ids_prov:
                                    for pid2 in ids_prov:
                                        if pid2 != pid:
                                            eq_batch.add((min(pid, pid2), max(pid, pid2)))
                                if total and n % 25 == 0:
                                    progreso.progress(min((n + 1) / total, 1.0))
                                continue

                            # Cada par se guarda UNA sola vez, siempre con el id menor primero.
                            # Antes se agregaban las dos direcciones —(a,b) y (b,a)—, y como la
                            # clave primaria es el par ordenado, quedaban dos filas por cada
                            # equivalencia. Eso duplicaba la tabla, contaba doble los vínculos de
                            # cada código (los "códigos puente" mostraban 200 donde había 100) y
                            # hacía que el control de precios listara cada par dos veces.
                            # La búsqueda nunca lo notó porque consulta las dos columnas con OR.
                            for pid in ids_prov:
                                for oid in ids_oem:
                                    if pid != oid:
                                        eq_batch.add((min(pid, oid), max(pid, oid)))
                                for pid2 in ids_prov:
                                    if pid2 != pid:
                                        eq_batch.add((min(pid, pid2), max(pid, pid2)))

                            cargados += 1
                            if total and n % 25 == 0:
                                progreso.progress(min((n + 1) / total, 1.0))

                            # En autocommit cada sentencia se confirma sola, así que este commit
                            # ya no hace nada (queda porque es inofensivo y evita tocar 89 lugares
                            # iguales). Se midió: la importación de 10.000 filas tarda lo mismo,
                            # 0,33 s contra 0,36 s, porque en modo WAL confirmar es barato.
                            if n % 300 == 0 and n > 0:
                                conn.commit()

                        # Inserción en lote: mucho más rápido que insertar de a un vínculo por vez
                        if eq_batch:
                            # No revivir vínculos que ya fueron rechazados en una revisión anterior
                            rechazados_antes = pares_rechazados()
                            eq_batch = {p for p in eq_batch if p not in rechazados_antes}
                        # El nombre del lote se arma SIEMPRE, vayan los vínculos a revisión o
                        # directo: es la etiqueta que después permite deshacer toda la lista.
                        lote_importacion = (f"{nombre_prov.upper()} · "
                                             f"{getattr(archivo, 'name', 'lista')} · "
                                             f"{datetime.now():%d/%m %H:%M}")
                        if eq_batch:
                            if cargar_directo:
                                c.executemany(
                                    "INSERT OR IGNORE INTO equivalencias (producto_a_id, producto_b_id, created_at, lote) "
                                    "VALUES (?, ?, datetime('now'), ?)",
                                    [(a, b, lote_importacion) for a, b in eq_batch]
                                )
                            else:
                                c.executemany(
                                    "INSERT OR IGNORE INTO equivalencias_pendientes "
                                    "(producto_a_id, producto_b_id, origen, lote) VALUES (?, ?, 'lista_proveedor', ?)",
                                    [(a, b, lote_importacion) for a, b in eq_batch]
                                )

                        try:
                            _huella = huella_de_archivo(archivo.getvalue())
                        except Exception:
                            _huella = None
                        c.execute(
                            "INSERT INTO importaciones (marca, archivo, filas_cargadas, filas_omitidas, huella, lote) "
                            "VALUES (?, ?, ?, ?, ?, ?)",
                            (nombre_prov.upper(), getattr(archivo, "name", str(archivo)),
                             cargados, omitidos, _huella, lote_importacion)
                        )
                        conn.commit()

                    progreso.empty()

                    st.success(f"Se importaron {cargados} filas con equivalencia.")
                    # Los números del chequeo de salud cambiaron: que se recalculen
                    invalidar_salud()

                    # Puntuar los vínculos que acaban de entrar, para que el buscador ya los
                    # use. Antes había que acordarse de tocar «Calcular la confianza».
                    try:
                        recalcular_confianzas(limite=4000)
                    except Exception:
                        pass

                    # Informe de lo que pasó con ESTA lista, mientras se la tiene fresca
                    try:
                        _inf = informe_post_importacion(lote_importacion, nombre_prov, cargados)
                    except Exception:
                        _inf = {"puntos": [], "vinculos_nuevos": 0}
                    if _inf["puntos"]:
                        st.markdown("#### 🔎 Qué conviene revisar de esta lista")
                        for nivel, titulo, detalle, donde in _inf["puntos"]:
                            (st.error if nivel == "alto" else st.warning)(
                                f"**{titulo}**\n\n{detalle}\n\n📍 {donde}")
                    elif _inf["vinculos_nuevos"]:
                        st.info(f"✅ Entraron {_inf['vinculos_nuevos']} vínculo(s) y ninguno "
                                 "disparó alarmas.")
                    # Recordar el mapeo que funcionó, para la próxima lista de este proveedor
                    guardar_mapeo_columnas(nombre_prov, idx_prov, idx_oem, idx_desc,
                                            idx_precio, idx_stock, buscar_oem_en_desc, prov_es_oem)
                    if not cargar_directo and eq_batch:
                        st.info(
                            f"🔒 {len(eq_batch)} vínculo(s) quedaron **esperando tu revisión** — todavía "
                            "no están cargados como equivalencias. Andá a Estadísticas → "
                            "🔗 Equivalencias sugeridas para aprobarlos (podés hacerlo en bloque)."
                        )
                    if cargados_sin_equiv:
                        st.info(
                            f"📦 Además se cargaron {cargados_sin_equiv} producto(s) que no traían código "
                            "de fábrica. Quedan buscables por código y por descripción; les va a aparecer "
                            "la equivalencia sola cuando el mismo código llegue desde la lista de otro "
                            "proveedor, o podés vincularlos a mano desde 'Vincular manual'."
                        )
                    if precios_actualizados:
                        st.success(f"💲 Se actualizaron {precios_actualizados} precio(s).")

                    if precios_frenados:
                        st.warning(
                            f"🛑 {len(precios_frenados)} precio(s) quedaron **sin actualizar** porque "
                            "el cambio no parece un aumento sino un error de la lista. El producto se "
                            "cargó igual; lo único que no se tocó es el precio."
                        )
                        st.dataframe(precios_frenados[:100], use_container_width=True, hide_index=True,
                                      column_config={"_id": None})
                        st.download_button(
                            "⬇️ Bajar la lista completa de precios frenados",
                            data=to_excel_bytes([{k: v for k, v in f.items() if k != "_id"}
                                                  for f in precios_frenados]),
                            file_name="precios_frenados.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.caption(
                            "Si mirás la lista y los precios están bien (por ejemplo, hubo un "
                            "aumento fuerte de verdad), volvé a importar subiendo el límite."
                        )
                        if st.checkbox("Revisé la lista y los precios de la planilla son correctos",
                                        key="confirmar_precios_frenados"):
                            if st.button(f"💲 Aplicar igual esos {len(precios_frenados)} precios"):
                                with db_lock:
                                    for f in precios_frenados:
                                        if f["Precio de la lista"] is not None:
                                            c.execute("INSERT INTO historial_precios (producto_id, precio) "
                                                       "VALUES (?, ?)", (f["_id"], f["Precio de la lista"]))
                                            c.execute("UPDATE productos SET precio = ? WHERE id = ?",
                                                       (f["Precio de la lista"], f["_id"]))
                                    conn.commit()
                                st.success(f"Se aplicaron {len(precios_frenados)} precio(s).")

                    if descartados_cortos:
                        st.caption(
                            f"🧹 Se ignoraron {descartados_cortos} valor(es) de las columnas de código "
                            "por ser un número suelto de 1 o 2 dígitos (cantidad, número de orden, "
                            "bulto). No son códigos y ensuciaban las equivalencias."
                        )
                    if omitidos:
                        st.warning(f"Se omitieron {omitidos} filas porque no tenían código de proveedor.")
                        st.dataframe(filas_omitidas, use_container_width=True)
                        st.download_button(
                            "⬇️ Descargar filas omitidas",
                            data=to_excel_bytes(filas_omitidas),
                            file_name="filas_omitidas.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    # Detección de posibles duplicados / errores de tipeo dentro de la marca recién cargada
                    sospechosos = detectar_posibles_duplicados(prov_id)
                    if sospechosos is None:
                        st.caption(
                            "ℹ️ La marca tiene demasiados códigos cargados como para revisar duplicados "
                            "automáticamente sin demorar la página."
                        )
                    elif sospechosos:
                        st.warning(
                            f"⚠️ Encontré {len(sospechosos)} par(es) de códigos muy parecidos dentro de "
                            f"'{nombre_prov}' — podrían ser errores de tipeo. Revisalos:"
                        )
                        st.dataframe(sospechosos, use_container_width=True, hide_index=True)
                except Exception as e:
                    st.error(f"Error procesando la lista: {e}")

        st.markdown("---")
        st.markdown("**📄 Cargar remito por foto (con IA)**")
        explicar(
            "Sacale una foto o subí una imagen del remito/factura de un proveedor.",
            "La IA lee los ítems y te arma una lista para revisar — no toca el stock hasta que vos "
            "confirmes. Solo actualiza cantidad de los códigos que ya existen en tu catálogo; los "
            "que no coincidan con nada, los tenés que cargar por 'Vincular manual' o con un Excel "
            "nuevo."
        )
        foto_remito = subir_archivo("Foto del remito:", ["png", "jpg", "jpeg"], "remito")
        remito_ok = archivo_listo(foto_remito, "foto del remito")
        if foto_remito:
            boton_otro_archivo("remito", "🗑️ Usar otra foto", key="otra_foto_remito")
        if st.button("🔍 Leer remito", disabled=not remito_ok):
            with st.spinner("Leyendo remito..."):
                items_leidos, error_remito = leer_remito_por_foto(foto_remito.getvalue())
            if error_remito:
                st.error(error_remito)
            elif not items_leidos:
                st.warning("No se pudo leer ningún ítem en esa imagen.")
            else:
                st.session_state["items_remito"] = cotejar_items_remito(items_leidos)

        if st.session_state.get("items_remito"):
            items_actuales = st.session_state["items_remito"]
            coinciden = [i for i in items_actuales if i["_producto_id"]]
            no_coinciden = [i for i in items_actuales if not i["_producto_id"]]
            st.success(f"Se leyeron {len(items_actuales)} ítem(s) — {len(coinciden)} coinciden con tu catálogo.")
            st.dataframe(
                [{k: v for k, v in i.items() if not k.startswith("_")} for i in items_actuales],
                use_container_width=True, hide_index=True
            )
            if no_coinciden:
                st.caption(
                    f"⚠️ {len(no_coinciden)} ítem(s) no coinciden con ningún código cargado — "
                    "revisá si están mal leídos o si son productos nuevos para vos."
                )
            colr1, colr2 = st.columns(2)
            if coinciden and colr1.button(f"💾 Sumar stock de los {len(coinciden)} que coinciden"):
                actualizados = aplicar_carga_remito(items_actuales)
                st.success(f"Stock actualizado en {actualizados} producto(s).")
                st.session_state.pop("items_remito", None)
                st.rerun()
            if colr2.button("🗑️ Descartar esta lectura"):
                st.session_state.pop("items_remito", None)
                st.rerun()

# ============================================================
# ADMINISTRAR
# ============================================================
def exportar_configuracion_txt():
    """Junta combos de repuestos, códigos DTC cargados y fabricantes VIN en un solo archivo de texto,
    para respaldo aparte de la base completa o para copiarle la configuración a otra sucursal."""
    lineas = []
    lineas.append(f"# Exportación de configuración — Equivalencias El Chavo — {datetime.now():%d/%m/%Y %H:%M}")
    lineas.append("")

    lineas.append("## COMBOS DE REPUESTOS RELACIONADOS (disparador;item)")
    for combo in listar_combos():
        for item in combo["items"]:
            lineas.append(f"{combo['disparador']};{item}")
    lineas.append("")

    lineas.append("## CÓDIGOS DTC (codigo;descripcion;sistema;causas;fabricante — mismo formato que la carga masiva)")
    c.execute("SELECT codigo, descripcion, sistema, causas_posibles, fabricante FROM codigos_dtc ORDER BY fabricante, codigo")
    for row in c.fetchall():
        lineas.append(f"{row['codigo']};{row['descripcion']};{row['sistema'] or ''};{row['causas_posibles'] or ''};{row['fabricante'] or ''}")
    lineas.append("")

    lineas.append("## FABRICANTES POR WMI - lector de VIN (wmi;fabricante;pais)")
    for f in listar_fabricantes_vin():
        lineas.append(f"{f['WMI']};{f['Fabricante']};{f.get('País') or ''}")

    return "\n".join(lineas)


if pagina == PAGINAS[3]:
    st.subheader("🗂️ Administrar")

    SUB_ADMIN = ["🏷️ Marcas", "📦 Productos", "💬 Mensajería y cobros", "🧩 Combos", "🧹 Mantenimiento", "👥 Usuarios"]
    if st.session_state.get("sub_admin") not in SUB_ADMIN:
        st.session_state["sub_admin"] = SUB_ADMIN[0]
    st.radio("Sub-sección:", SUB_ADMIN, key="sub_admin", horizontal=True,
             label_visibility="collapsed")
    sub_admin = st.session_state["sub_admin"]

    c.execute("""SELECT m.id, m.nombre, m.tipo, COUNT(p.id) AS productos
                 FROM marcas m LEFT JOIN productos p ON p.marca_id = m.id
                 GROUP BY m.id ORDER BY m.nombre""")
    marcas_info = c.fetchall()

    if sub_admin == SUB_ADMIN[0]:
        if not marcas_info:
            st.info("Todavía no hay marcas cargadas.")
        else:
            tabla_marcas = [{"Marca": m["nombre"], "Tipo": m["tipo"], "Productos cargados": m["productos"]}
                             for m in marcas_info]
            st.dataframe(tabla_marcas, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("**🔗 Link a la ficha del proveedor (en vez de guardar la foto)**")
            explicar(
                "Por cada marca/proveedor podés cargar un patrón de URL con `{codigo}` donde va el "
                "código del producto.",
                "La app arma el link automáticamente para cada resultado de búsqueda, sin copiar "
                "ninguna imagen — así podés sumar Taranto y cualquier otro proveedor que uses, cada uno "
                "con su propio patrón. Ejemplo: `https://www.taranto.com.ar/busqueda?q={codigo}`"
            )
            nombres_para_link = [m["nombre"] for m in marcas_info]
            marca_link = st.selectbox("Marca:", nombres_para_link, key="marca_link_ficha")
            id_marca_link = next(m["id"] for m in marcas_info if m["nombre"] == marca_link)
            c.execute("SELECT url_ficha_template FROM marcas WHERE id = ?", (id_marca_link,))
            template_actual = c.fetchone()["url_ficha_template"] or ""
            nuevo_template = st.text_input(
                "Patrón de URL (usá {codigo} donde va el código):", value=template_actual,
                placeholder="https://www.taranto.com.ar/busqueda?q={codigo}", key="input_template_link"
            )
            if st.button("💾 Guardar patrón de link"):
                if nuevo_template.strip() and "{codigo}" not in nuevo_template:
                    st.warning("El patrón tiene que incluir '{codigo}' en algún lado, si no todos los links quedan iguales.")
                else:
                    with db_lock:
                        c.execute("UPDATE marcas SET url_ficha_template = ? WHERE id = ?",
                                  (nuevo_template.strip() or None, id_marca_link))
                        conn.commit()
                    avisar("success", f"Patrón de link guardado para '{marca_link}'.")
                    st.rerun()

            st.markdown("---")
            duplicadas = marcas_probablemente_duplicadas()
            if duplicadas:
                st.markdown("**🔎 Marcas que parecen ser la misma**")
                explicar(
                    "Detectadas por el nombre o porque comparten códigos.",
                    "Una lista viene como «MAHLE» y la siguiente como «MAHLE FILTER», o alguien "
                    "la tipeó distinto. Quedan como dos proveedores separados y el buscador deja "
                    "de encontrar equivalencias que en realidad existen.\n\nLa señal más fuerte "
                    "no es el nombre sino los **códigos compartidos**: si dos marcas tienen los "
                    "mismos códigos son el mismo proveedor, aunque se llamen distinto. Ese caso "
                    "no lo agarra ningún chequeo por nombre."
                )
                st.dataframe(quitar_id(duplicadas), use_container_width=True, hide_index=True)
                st.caption(
                    "Fusionalas abajo, poniendo como origen la que quieras eliminar. Los "
                    "productos que existan en las dos se juntan conservando precio, stock y "
                    "equivalencias."
                )
                st.markdown("---")

            st.markdown("**🔀 Fusionar marcas duplicadas**")
            st.caption(
                "Útil cuando una marca quedó cargada con nombres distintos por error de tipeo "
                "(ej: 'MANN' y 'MANN FILTER'). Mueve todos los productos de una a la otra."
            )
            nombres_para_fusion = [m["nombre"] for m in marcas_info]
            colOrig, colDest = st.columns(2)
            marca_origen = colOrig.selectbox("Marca a eliminar (origen):", nombres_para_fusion, key="fusion_origen")
            marca_destino = colDest.selectbox("Marca a conservar (destino):", nombres_para_fusion, key="fusion_destino")
            if st.button("🔀 Fusionar", disabled=(marca_origen == marca_destino)):
                if pedir_password_admin("fusionar marcas"):
                    id_origen = next(m["id"] for m in marcas_info if m["nombre"] == marca_origen)
                    id_destino = next(m["id"] for m in marcas_info if m["nombre"] == marca_destino)
                    movidos, fusionados = fusionar_marcas(id_origen, id_destino)
                    detalle = f"{movidos} producto(s) movidos"
                    if fusionados:
                        detalle += (f" y {fusionados} fusionados con los que ya existían "
                                     "en la marca destino con el mismo código")
                    avisar("success", f"'{marca_origen}' se fusionó dentro de '{marca_destino}': "
                                       f"{detalle}.")
                    invalidar_salud()
                    st.rerun()

            st.markdown("---")
            st.markdown("**💲 Aumentar/bajar precios por porcentaje**")
            st.caption("Aplica el ajuste a todos los productos con precio cargado de la marca elegida.")
            marca_precio = st.selectbox("Marca:", nombres_para_fusion, key="marca_ajuste_precio")
            porcentaje = st.number_input("Porcentaje (usá negativo para bajar, ej: -5):", value=0.0, step=1.0)
            if st.button("💲 Aplicar ajuste de precios", disabled=(porcentaje == 0)):
                if pedir_password_admin("ajustar precios masivamente"):
                    id_marca_precio = next(m["id"] for m in marcas_info if m["nombre"] == marca_precio)
                    afectados = aumentar_precios_por_marca(id_marca_precio, porcentaje)
                    st.success(f"Se ajustaron {afectados} precio(s) de '{marca_precio}' en {porcentaje:+.1f}%.")

            st.markdown("---")
            st.markdown("**Eliminar una marca** (borra también sus productos y equivalencias asociadas)")
            marca_a_borrar = st.selectbox("Elegí una marca", [m["nombre"] for m in marcas_info])
            confirmar = st.checkbox(f"Confirmo que quiero borrar '{marca_a_borrar}' y todo lo asociado")
            if st.button("🗑️ Eliminar marca", disabled=not confirmar):
                if pedir_password_admin("eliminar una marca"):
                    eliminar_marca_con_papelera(marca_a_borrar)
                    avisar("success", f"Marca '{marca_a_borrar}' eliminada (podés restaurarla desde la papelera).")
                    st.rerun()

        st.markdown("**Catálogos externos**")
        st.caption("Agregá los sitios de proveedores que querés que aparezcan como botones al buscar un código.")

        catalogos = listar_catalogos_externos()
        if catalogos:
            for cat in catalogos:
                colA, colB, colC = st.columns([2, 5, 1])
                colA.write(cat["nombre"])
                colB.write(cat["url"])
                if colC.button("🗑️", key=f"del_cat_{cat['id']}"):
                    eliminar_catalogo_externo(cat["id"])
                    st.rerun()
        else:
            st.caption("Todavía no agregaste ningún catálogo externo.")

        with st.form("nuevo_catalogo", clear_on_submit=True):
            colN, colU = st.columns(2)
            nombre_cat = colN.text_input("Nombre del proveedor", placeholder="Ej: Wega")
            url_cat = colU.text_input("URL del catálogo", placeholder="Ej: wegamotors.com")
            agregar = st.form_submit_button("➕ Agregar catálogo")
            if agregar:
                if not nombre_cat.strip() or not url_cat.strip():
                    st.warning("Completá nombre y URL.")
                else:
                    agregar_catalogo_externo(nombre_cat, url_cat)
                    avisar("success", f"'{nombre_cat}' agregado.")
                    st.rerun()

    if sub_admin == SUB_ADMIN[1]:
        st.markdown("**Buscar y editar un producto puntual**")
        texto_prod = st.text_input("Buscar producto por código o descripción", key="admin_buscar")
        if texto_prod.strip():
            res_admin = buscar_por_texto(texto_prod)
            if not res_admin:
                clean_admin = sanitizar(texto_prod)
                if clean_admin:
                    res_admin = buscar_por_codigo(clean_admin)
            if res_admin:
                st.dataframe(res_admin, use_container_width=True, hide_index=True)
                st.caption(
                    "¿Necesitás borrar un producto? Está en '🧹 Mantenimiento' → separado a propósito "
                    "de la edición, para que un descuido acá no borre nada."
                )

                st.markdown("**📐 Cargar medidas mecánicas / ubicación en depósito**")
                opciones_prod = {f"{f['Codigo']} ({f['Marca']}) — ID {f['ID']}": f['ID'] for f in res_admin}
                elegido_label = st.selectbox("Elegí el producto a editar:", list(opciones_prod.keys()), key="sel_medidas")
                id_medidas = opciones_prod[elegido_label]
                c.execute(
                    "SELECT diametro_interno, diametro_externo, ancho, paso_rosca, cantidad_estrias, ubicacion, "
                    "estrias_internas, estrias_externas, posicion_seguro, tiene_abs, "
                    "diametro_interno_cara_b, diametro_externo_cara_b, "
                    "diametro_rosca_homocinetica, diametro_copa, diametro_copa_superior, largo_total "
                    "FROM productos WHERE id = ?", (id_medidas,)
                )
                actual = c.fetchone()
                em1, em2, em3 = cols(3)
                e_diam_int = em1.number_input("Diám. interno cara A (mm)", min_value=0.0, step=0.1,
                                               value=float(actual["diametro_interno"] or 0), key="e_di")
                e_diam_ext = em2.number_input("Diám. externo cara A (mm)", min_value=0.0, step=0.1,
                                               value=float(actual["diametro_externo"] or 0), key="e_de")
                e_ancho = em3.number_input("Ancho (mm)", min_value=0.0, step=0.1,
                                            value=float(actual["ancho"] or 0), key="e_an")
                em4, em5, em6 = cols(3)
                e_paso = em4.text_input("Paso de rosca", value=actual["paso_rosca"] or "", key="e_paso")
                e_estrias = em5.number_input("Cantidad de estrías", min_value=0, step=1,
                                              value=int(actual["cantidad_estrias"] or 0), key="e_estrias")
                e_ubicacion = em6.text_input("Ubicación en depósito", value=actual["ubicacion"] or "",
                                              placeholder="Ej: Pasillo 3, estante B", key="e_ubic")

                st.markdown("**↔️ Segunda cara (opcional)**")
                st.caption(
                    "Para piezas con distinta medida de cada lado — retenes con labio interior/exterior "
                    "escalonado, tensores con el interior de un diámetro de un lado y otro del otro, etc."
                )
                eb1, eb2 = st.columns(2)
                e_diam_int_b = eb1.number_input("Diám. interno cara B (mm)", min_value=0.0, step=0.1,
                                                 value=float(actual["diametro_interno_cara_b"] or 0), key="e_di_b")
                e_diam_ext_b = eb2.number_input("Diám. externo / labio exterior cara B (mm)", min_value=0.0, step=0.1,
                                                 value=float(actual["diametro_externo_cara_b"] or 0), key="e_de_b")

                st.markdown("**🔩 Homocinéticas**")
                eh1, eh2 = st.columns(2)
                e_estrias_int = eh1.number_input("Estrías internas", min_value=0, step=1,
                                                  value=int(actual["estrias_internas"] or 0), key="e_estrias_int")
                e_estrias_ext = eh2.number_input("Estrías externas", min_value=0, step=1,
                                                  value=int(actual["estrias_externas"] or 0), key="e_estrias_ext")
                eh3, eh4 = st.columns(2)
                e_seguro = eh3.text_input("Posición del seguro", value=actual["posicion_seguro"] or "",
                                           placeholder="Ej: 1er ranura, a 12mm", key="e_seguro")
                abs_actual = "Cualquiera" if actual["tiene_abs"] is None else ("Sí" if actual["tiene_abs"] else "No")
                e_abs = eh4.selectbox("¿Tiene ABS?", ["Cualquiera", "Sí", "No"],
                                       index=["Cualquiera", "Sí", "No"].index(abs_actual), key="e_abs")
                eh5, eh6 = st.columns(2)
                e_rosca_homo = eh5.number_input("Diámetro de rosca (mm)", min_value=0.0, step=0.1,
                                                 value=float(actual["diametro_rosca_homocinetica"] or 0), key="e_rosca_homo")
                e_largo_total = eh6.number_input(
                    "Largo total (mm)", min_value=0.0, step=0.5,
                    value=float(actual["largo_total"] or 0), key="e_largo_total",
                    help="De punta a punta. Es la medida que más rápido descarta: dos homocinéticas "
                         "con las mismas estrías y la misma copa pero distinto largo no entran en "
                         "el mismo auto."
                )
                explicar(
                    "La copa es cónica, así que se cargan sus dos diámetros:",
                    "el de la **base** (donde se une al eje) y el de la **boca** (el borde abierto). Con "
                    "uno solo no se distinguen dos copas que arrancan igual y terminan distinto — y son "
                    "justo esas las que no se pueden intercambiar."
                )
                eh7, eh8 = st.columns(2)
                e_copa = eh7.number_input("Diám. copa — base (mm)", min_value=0.0, step=0.1,
                                           value=float(actual["diametro_copa"] or 0), key="e_copa")
                e_copa_sup = eh8.number_input("Diám. copa — boca / superior (mm)", min_value=0.0, step=0.1,
                                               value=float(actual["diametro_copa_superior"] or 0),
                                               key="e_copa_sup")

                if st.button("💾 Guardar medidas y ubicación"):
                    actualizar_medidas(id_medidas, e_diam_int, e_diam_ext, e_ancho, e_paso, e_estrias, e_ubicacion,
                                        e_estrias_int, e_estrias_ext, e_seguro, e_abs,
                                        e_diam_int_b, e_diam_ext_b, e_rosca_homo, e_copa,
                                        e_copa_sup, e_largo_total)
                    st.success("Guardado.")

                st.markdown("**📷 Fotos del producto**")
                explicar(
                    "Cargá varias del mismo producto, y cuanto más distintas entre sí, mejor.",
                    "De frente, de costado, la de la ficha del proveedor. Ninguna comparación "
                    "reconoce una pieza de frente en una foto sacada de costado, así que la única "
                    "forma de cubrir los dos ángulos es tener los dos."
                )

                fotos_actuales = listar_fotos_producto(id_medidas)
                if fotos_actuales:
                    etiquetas_estado = {
                        "ok": "🟢 lista para comparar",
                        "sin_detalle": "🟡 se ve, pero no sirve para comparar",
                        "error": "🔴 no se pudo procesar",
                    }
                    columnas_fotos = st.columns(min(len(fotos_actuales), 4))
                    for idx_f, foto in enumerate(fotos_actuales):
                        with columnas_fotos[idx_f % len(columnas_fotos)]:
                            st.image(foto["imagen_data"], use_container_width=True)
                            st.caption(etiquetas_estado.get(foto["estado"], "⚪ sin procesar"))
                            if st.button("🗑️", key=f"del_foto_{foto['id']}", help="Borrar esta foto"):
                                eliminar_foto_producto(foto["id"])
                                st.rerun()
                else:
                    st.caption("Todavía sin fotos.")

                origen_foto_nueva = st.radio(
                    "Agregar una foto:", ["📷 Subir", "🔗 Desde una dirección web"],
                    horizontal=True, key="origen_foto_producto"
                )

                if origen_foto_nueva.startswith("📷"):
                    foto_producto = subir_archivo(
                        "Foto (se guarda comprimida y aparece en la columna 'Imagen' del buscador):",
                        ["png", "jpg", "jpeg"], "foto_producto"
                    )
                    producto_foto_ok = archivo_listo(foto_producto, "foto")
                    if foto_producto:
                        boton_otro_archivo("foto_producto", "🗑️ Usar otra foto", key="otra_foto_producto")
                    if st.button("💾 Agregar esta foto", disabled=not producto_foto_ok):
                        _, estado_foto = agregar_foto_producto(id_medidas, foto_producto.getvalue())
                        if estado_foto == "ok":
                            st.success("Foto agregada y lista para la búsqueda por parecido.")
                        elif estado_foto == "sin_detalle":
                            st.warning(
                                "Foto agregada — se va a ver en el buscador, pero **no sirve para "
                                "comparar por parecido**: no tiene detalles distintivos suficientes "
                                "(pieza lisa, fondo del mismo tono, poca luz o movida). Para que "
                                "sirva, sacala apoyada sobre un fondo liso de OTRO color, con buena "
                                "luz, y que se lea el grabado o la marca de la pieza."
                            )
                        else:
                            st.error("No se pudo procesar esa imagen. Probá con otra.")
                        olvidar_archivo("foto_producto")
                        st.rerun()
                else:
                    st.caption(
                        "Pegá la dirección de la ficha del proveedor o la de la imagen directa. Se "
                        "bajan las fotos de esa página y elegís cuáles guardar."
                    )
                    url_foto_prod = st.text_input(
                        "Dirección web:", placeholder="https://...", key="url_foto_producto"
                    ).strip()
                    if st.button("⬇️ Traer fotos de esa dirección", disabled=not url_foto_prod,
                                 key="btn_traer_fotos_prod"):
                        with st.spinner("Bajando..."):
                            halladas, error_ph = imagenes_de_una_direccion(url_foto_prod)
                        if error_ph:
                            st.error(error_ph)
                            st.session_state.pop("fotos_url_producto", None)
                        else:
                            st.session_state["fotos_url_producto"] = halladas
                            st.rerun()

                    halladas = st.session_state.get("fotos_url_producto")
                    if halladas:
                        st.caption(f"{len(halladas)} foto(s) encontradas — guardá las que sean de la pieza:")
                        cols_ph = st.columns(min(len(halladas), 4))
                        for idx_h, (url_h, datos_h) in enumerate(halladas[:8]):
                            with cols_ph[idx_h % len(cols_ph)]:
                                st.image(datos_h, use_container_width=True)
                                if st.button("💾 Guardar", key=f"guardar_foto_url_{idx_h}"):
                                    _, est_h = agregar_foto_producto(
                                        id_medidas, datos_h, origen="url", fuente=url_h
                                    )
                                    if est_h == "ok":
                                        st.success("Guardada y lista para comparar.")
                                    elif est_h == "sin_detalle":
                                        st.warning("Guardada, pero sin detalle suficiente para comparar.")
                                    else:
                                        st.error("No se pudo procesar esa imagen.")
                                    st.rerun()
                        if st.button("✖️ Cerrar estas fotos", key="cerrar_fotos_url_prod"):
                            st.session_state.pop("fotos_url_producto", None)
                            st.rerun()

                if fotos_actuales and st.button("🗑️ Sacar TODAS las fotos de este producto"):
                    eliminar_imagen_producto(id_medidas)
                    avisar("success", "Fotos eliminadas.")
                    st.rerun()
            else:
                st.info("Sin resultados.")

        st.markdown("**🧩 Productos sin equivalencias**")
        st.caption(
            "Esta sección puede ser pesada, así que se calcula solo cuando la pedís (no en cada búsqueda)."
        )

        if "mostrar_huerfanos" not in st.session_state:
            st.session_state.mostrar_huerfanos = False

        col_ver, col_ocultar = st.columns(2)
        if col_ver.button("📋 Mostrar productos sin equivalencias"):
            st.session_state.mostrar_huerfanos = True
            st.rerun()
        if st.session_state.mostrar_huerfanos and col_ocultar.button("🙈 Ocultar"):
            st.session_state.mostrar_huerfanos = False
            st.rerun()

        if st.session_state.mostrar_huerfanos:
            total_sin_eq = contar_productos_sin_equivalencias()
            st.write(f"Total: **{total_sin_eq}** producto(s) sin ninguna equivalencia.")

            if total_sin_eq == 0:
                st.info("¡Todos los productos tienen al menos una equivalencia! 🎉")
            else:
                c.execute("SELECT nombre FROM marcas ORDER BY nombre")
                marcas_para_filtro = ["Todas"] + [r["nombre"] for r in c.fetchall()]
                marca_filtro_huerfanos = st.selectbox("Filtrar por marca:", marcas_para_filtro, key="filtro_huerfanos")
                cantidad_mostrar = st.selectbox("Mostrar en pantalla:", [25, 50, 100], index=0,
                                                 help="La descarga en Excel siempre incluye todo, esto es solo lo que se dibuja en pantalla.")

                pendientes_completo = listar_productos_sin_equivalencias(marca_filtro_huerfanos, limite=2000)
                if pendientes_completo:
                    st.download_button(
                        "⬇️ Descargar lista completa (Excel)",
                        data=to_excel_bytes(quitar_id(pendientes_completo)),
                        file_name="productos_sin_equivalencias.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    pendientes = pendientes_completo[:cantidad_mostrar]
                    st.caption(f"Mostrando {len(pendientes)} de {len(pendientes_completo)}.")
                    for fila in pendientes:
                        colC, colM, colB = st.columns([3, 2, 1.2])
                        colC.write(f"{fila['Codigo']}" + (f" — {fila['Descripcion']}" if fila.get('Descripcion') else ""))
                        colM.write(fila['Marca'])
                        if colB.button("🔗 Usar", key=f"usar_huerfano_{fila['ID']}"):
                            st.session_state["vincular_pendiente"] = {
                                "cod_a": fila["Codigo"],
                                "marca_a": fila["Marca"],
                                "desc_a": fila.get("Descripcion") or ""
                            }
                            avisar("success", "Cargado. Andá a la pestaña '🔗 Vincular manual' para completar el Código B.")
                            st.rerun()
                else:
                    st.info("Sin resultados para esa marca.")

    if sub_admin == SUB_ADMIN[2]:
        st.markdown("**💬 Texto del mensaje de WhatsApp**")
        st.caption(
            "Personalizá el encabezado y el pie del mensaje que se arma en 'Lista WhatsApp' — por "
            "ejemplo para poner el nombre real de tu local, un teléfono de contacto, horarios, etc."
        )
        encabezado_actual = obtener_config("whatsapp_encabezado", "🔧 *Equivalencias El Chavo*")
        pie_actual = obtener_config("whatsapp_pie", "")
        nuevo_encabezado = st.text_input("Encabezado del mensaje:", value=encabezado_actual, key="wa_encabezado_in")
        nuevo_pie = st.text_area("Pie del mensaje (opcional):", value=pie_actual, key="wa_pie_in",
                                  placeholder="Ej: 📍 Av. Siempreviva 742 - Horario: L a V 9 a 18hs")
        if st.button("💾 Guardar textos del mensaje"):
            guardar_config("whatsapp_encabezado", nuevo_encabezado.strip() or "🔧 *Equivalencias El Chavo*")
            guardar_config("whatsapp_pie", nuevo_pie.strip())
            avisar("success", "Guardado.")
            st.rerun()

        st.markdown("---")
        st.markdown("**💳 Alias para QR de transferencia**")
        explicar(
            "Cargá los alias/CBU que usás (Mercado Pago, distintos bancos, etc.).",
            "Al armar una cotización en 'Lista WhatsApp' vas a poder elegir cuál de estos usar para "
            "el QR del PDF. Si subís el QR real que te da tu banco/Mercado Pago/MODO, se usa ese "
            "(funciona de verdad para transferir). Si no subís nada, se genera uno con el alias/CBU "
            "como texto plano — sirve para no tipear a mano, pero no lo van a reconocer como QR de "
            "pago."
        )
        alias_cargados = listar_alias_transferencia()
        if alias_cargados:
            st.dataframe(
                [{k: v for k, v in a.items() if k not in ("ID", "TieneQrReal")} | {"QR real": "✅" if a["TieneQrReal"] else "—"}
                 for a in alias_cargados],
                use_container_width=True, hide_index=True
            )

        opciones_alias_edit = ["➕ Nuevo alias..."] + [f"{a['Nombre']} (editar)" for a in alias_cargados]
        alias_opcion_edit = st.selectbox("Elegí qué hacer:", opciones_alias_edit, key="alias_opcion_edit")
        alias_actual = None
        if alias_opcion_edit != "➕ Nuevo alias...":
            nombre_buscar = alias_opcion_edit.replace(" (editar)", "")
            alias_actual = next(a for a in alias_cargados if a["Nombre"] == nombre_buscar)

        cae1, cae2 = st.columns(2)
        nombre_alias_in = cae1.text_input("Nombre (ej: Mercado Pago, Banco Galicia)",
                                           value=(alias_actual or {}).get("Nombre", ""), key="alias_nombre_in")
        alias_in = cae2.text_input("Alias", value=(alias_actual or {}).get("Alias", ""), key="alias_alias_in")
        cae3, cae4 = st.columns(2)
        cbu_in = cae3.text_input("CBU/CVU (opcional)", value=(alias_actual or {}).get("CBU", ""), key="alias_cbu_in")
        titular_in = cae4.text_input("Titular (opcional)", value=(alias_actual or {}).get("Titular", ""), key="alias_titular_in")

        archivo_qr_real = subir_archivo(
            "QR real (opcional — el que te dio tu banco/Mercado Pago/MODO):",
            ["png", "jpg", "jpeg"], "qr_real"
        )
        if archivo_qr_real:
            archivo_listo(archivo_qr_real, "QR")
            boton_otro_archivo("qr_real", "🗑️ Usar otro QR", key="otro_qr_real")
        if alias_actual and alias_actual["TieneQrReal"]:
            st.caption("✅ Este alias ya tiene un QR real cargado. Subí uno nuevo para reemplazarlo.")

        cbtn1, cbtn2, cbtn3 = st.columns(3)
        if cbtn1.button("💾 Guardar alias"):
            if not nombre_alias_in.strip() or not alias_in.strip():
                st.warning("Completá al menos el nombre y el alias.")
            else:
                guardar_alias_transferencia(
                    nombre_alias_in, alias_in, cbu_in, titular_in,
                    alias_id=(alias_actual["ID"] if alias_actual else None),
                    qr_real_bytes=(archivo_qr_real.getvalue() if archivo_qr_real else None)
                )
                avisar("success", "Alias guardado.")
                st.rerun()
        if alias_actual and alias_actual["TieneQrReal"] and cbtn2.button("🗑️ Sacar el QR real"):
            eliminar_qr_real(alias_actual["ID"])
            avisar("success", "QR real eliminado — vuelve a usar el de texto plano.")
            st.rerun()
        if alias_actual and cbtn3.button("🗑️ Eliminar este alias"):
            eliminar_alias_transferencia(alias_actual["ID"])
            avisar("success", "Alias eliminado.")
            st.rerun()

    if sub_admin == SUB_ADMIN[3]:
        st.markdown("**🧩 Combos de repuestos relacionados**")
        st.caption(
            "Cuando alguien busca un producto cuya descripción contenga el 'disparador', la app va a "
            "sugerir estos ítems relacionados con un botón para buscarlos también. Ej: disparador "
            "'correa de distribucion' → ítems 'Kit de distribución', 'Tensor', 'Bomba de agua'."
        )
        combos_actuales = listar_combos()
        if combos_actuales:
            st.dataframe(
                [{"Disparador": c_["disparador"], "Ítems sugeridos": ", ".join(c_["items"])} for c_ in combos_actuales],
                use_container_width=True, hide_index=True
            )
        disparador_edit = st.text_input(
            "Disparador (palabra/frase que aparece en la descripción del producto):",
            placeholder="Ej: correa de distribucion", key="combo_disparador"
        )
        items_edit = st.text_area(
            "Ítems sugeridos (uno por línea):",
            placeholder="Kit de distribución\nTensor de distribución\nBomba de agua",
            key="combo_items", height=100
        )
        cc1, cc2 = st.columns(2)
        if cc1.button("💾 Guardar combo"):
            if not disparador_edit.strip() or not items_edit.strip():
                st.warning("Completá el disparador y al menos un ítem.")
            else:
                guardar_combo(disparador_edit, items_edit.strip().splitlines())
                avisar("success", f"Combo para '{disparador_edit.strip()}' guardado.")
                st.rerun()
        if cc2.button("🗑️ Eliminar combo (según el disparador de arriba)"):
            if disparador_edit.strip():
                eliminar_combo(disparador_edit)
                avisar("success", f"Combo para '{disparador_edit.strip()}' eliminado.")
                st.rerun()

    if sub_admin == SUB_ADMIN[4]:
        st.markdown("**🗑️ Eliminar un producto puntual**")
        st.caption(
            "Separado a propósito de la edición de medidas/fotos, para que buscar y editar un producto "
            "no te deje el botón de borrar a mano por accidente."
        )
        texto_prod_borrar = st.text_input("Buscar el producto a borrar (por código o descripción):", key="mant_buscar_borrar")
        if texto_prod_borrar.strip():
            res_borrar = buscar_por_texto(texto_prod_borrar)
            if not res_borrar:
                clean_borrar = sanitizar(texto_prod_borrar)
                if clean_borrar:
                    res_borrar = buscar_por_codigo(clean_borrar)
            if res_borrar:
                opciones_borrar = {f"{f['Codigo']} ({f['Marca']}) — ID {f['ID']}": f['ID'] for f in res_borrar}
                elegido_borrar_label = st.selectbox("Elegí el producto a eliminar:", list(opciones_borrar.keys()),
                                                     key="mant_sel_borrar")
                id_a_borrar = opciones_borrar[elegido_borrar_label]
                st.caption(
                    "⚠️ Se borran también sus equivalencias con otros productos. Si lo restaurás desde "
                    "la papelera, el producto vuelve pero **sin** esos vínculos — hay que volver a "
                    "vincularlo manualmente."
                )
                confirmar_borrado = st.checkbox(f"Confirmo que quiero borrar '{elegido_borrar_label}'",
                                                 key="mant_confirmar_borrar")
                if st.button("🗑️ Eliminar producto", disabled=not confirmar_borrado):
                    if pedir_password_admin("eliminar un producto"):
                        c.execute("SELECT * FROM productos WHERE id = ?", (id_a_borrar,))
                        fila_producto = c.fetchone()
                        if fila_producto:
                            mover_a_papelera("producto", dict(fila_producto))
                        with db_lock:
                            c.execute("DELETE FROM productos WHERE id = ?", (id_a_borrar,))
                            conn.commit()
                        avisar("success", "Producto eliminado (podés restaurarlo desde la papelera, más abajo).")
                        st.rerun()
            else:
                st.caption("Sin resultados.")

        st.markdown("---")
        # Mantenimiento quedó con 15 herramientas apiladas en un solo scroll interminable.
        # Se agrupan por lo que uno viene a hacer, no por el orden en que se fueron sumando:
        # así se entra directo a lo que se necesita en vez de bajar buscándolo.
        _tabs_mant = st.tabs([g[0] for g in GRUPOS_MANTENIMIENTO])

        with _tabs_mant[0]:
            st.markdown("**🌉 Códigos puente — los que rompen la búsqueda**")
            explicar(
                "Códigos vinculados a demasiadas cosas. Cortar uno limpia miles de resultados falsos.",
                "La búsqueda encadena: trae los equivalentes de tu código, y los de esos, y así. "
                "Un código mal vinculado no ensucia solo su fila: **fusiona todas las familias que "
                "toca**.\n\nUn filtro legítimo puede tener 10 o 15 equivalencias entre marcas. Si "
                "aparece con 200, casi seguro se cargó mal y quedó de puente entre repuestos que no "
                "tienen nada que ver."
            )
            st.session_state.setdefault("minimo_puente", 15)
            minimo_puente = st.select_slider(
                "Mostrar los que tengan más de:", options=[15, 30, 50, 100, 200],
                key="minimo_puente"
            )
            puentes = codigos_puente(int(minimo_puente))
            if puentes:
                st.warning(f"⚠️ {len(puentes)} código(s) con más de {minimo_puente} vínculos.")
                st.caption(
                    "Mirá la columna «Marcas distintas»: un repuesto real se vincula con unas pocas "
                    "marcas. Uno que toca 20 marcas distintas casi nunca es legítimo."
                )
                st.dataframe(puentes, use_container_width=True, hide_index=True)
                st.caption(
                    "Si alguno de estos está bien —hay repuestos que legítimamente equivalen a "
                    "decenas—, aprobalo y deja de aparecer acá y en el aviso del buscador."
                )
                cod_aprobar = st.text_input("Código a APROBAR (está bien así):",
                                             key="cod_aprobar_puente").strip()
                if cod_aprobar:
                    c.execute("SELECT id, codigo_raw, descripcion FROM productos WHERE codigo_clean = ?",
                              (sanitizar(cod_aprobar),))
                    obj_ap = c.fetchone()
                    if not obj_ap:
                        st.error("No encontré ese código.")
                    else:
                        nota_ap = st.text_input("Nota (opcional):", key="nota_aprobar_puente",
                                                 placeholder="Ej: filtro común, va en toda la gama")
                        if st.button(f"✅ Aprobar {obj_ap['codigo_raw']} — sus vínculos están bien"):
                            aprobar_puente(obj_ap["id"], nota_ap)
                            invalidar_salud()
                            avisar("success", f"{obj_ap['codigo_raw']} quedó aprobado: no vuelve a "
                                               "aparecer como código puente.")
                            st.rerun()

                aprobados = puentes_aprobados_ids()
                if aprobados:
                    c.execute(f"""SELECT p.codigo_raw AS "Código", p.descripcion AS "Descripción",
                                         pa.nota AS "Nota", substr(pa.fecha,1,10) AS "Aprobado el",
                                         p.id AS "_id"
                                  FROM puentes_aprobados pa JOIN productos p ON p.id = pa.producto_id
                                  ORDER BY pa.fecha DESC""")
                    lista_ap = filas_a_listas(c)
                    with st.expander(f"✅ Códigos puente aprobados ({len(lista_ap)})"):
                        st.dataframe(quitar_id(lista_ap), use_container_width=True, hide_index=True)
                        quitar_ap = st.text_input("Código a desaprobar (volver a vigilarlo):",
                                                   key="cod_desaprobar").strip()
                        if quitar_ap and st.button("↩️ Volver a vigilarlo"):
                            c.execute("SELECT id FROM productos WHERE codigo_clean = ?",
                                      (sanitizar(quitar_ap),))
                            f_des = c.fetchone()
                            if f_des and desaprobar_puente(f_des["id"]):
                                invalidar_salud()
                                avisar("success", f"{quitar_ap} vuelve a vigilarse.")
                                st.rerun()
                            else:
                                st.warning("No estaba aprobado.")

                cod_cortar = st.text_input(
                    "Código al que cortarle TODOS los vínculos:", key="cod_cortar_puente",
                    help="El producto no se borra: queda en la base con su precio y su stock. Lo único "
                         "que se corta son las equivalencias, que es lo que está mal."
                ).strip()
                if cod_cortar:
                    clean_cortar = sanitizar(cod_cortar)
                    c.execute("SELECT id, codigo_raw, descripcion FROM productos WHERE codigo_clean = ?",
                              (clean_cortar,))
                    objetivo = c.fetchone()
                    if not objetivo:
                        st.error("No encontré ese código.")
                    else:
                        red = tamano_de_la_red(objetivo["id"])
                        st.info(f"**{objetivo['codigo_raw']}** — {objetivo['descripcion'] or 'sin descripción'}. "
                                 f"Hoy está encadenado con {red}{'+' if red >= 500 else ''} producto(s).")
                        if st.button(f"✂️ Cortar todos los vínculos de {objetivo['codigo_raw']}"):
                            if pedir_password_admin("cortar vínculos"):
                                n = cortar_vinculos_de(objetivo["id"])
                                invalidar_salud()
                                avisar("success", f"Se cortaron {n} vínculo(s). El producto quedó en la base.")
                                st.rerun()
            else:
                st.caption(f"✅ Ningún código con más de {minimo_puente} vínculos.")
            st.markdown("**🔗 Vínculos que unen dos familias de repuestos**")
            explicar(
                "Dos grupos sanos pegados por un solo vínculo malo. Cortándolo se separan.",
                "Es el caso que ningún otro control agarra: los filtros por un lado y los "
                "amortiguadores por el otro, perfectamente legítimos entre sí, unidos por UN "
                "vínculo mal cargado.\n\nNingún código tiene muchos enlaces, así que no aparece "
                "como código puente. Pero ese eslabón solo sostiene toda la unión."
            )
            cod_red = st.text_input("Código desde el que analizar la red:", key="cod_red_familias",
                                     placeholder="Poné uno de los que te devuelve resultados raros"
                                     ).strip()
            if cod_red:
                c.execute("SELECT id, codigo_raw FROM productos WHERE codigo_clean = ?",
                          (sanitizar(cod_red),))
                obj_red = c.fetchone()
                if not obj_red:
                    st.error("No encontré ese código.")
                else:
                    with st.spinner("Analizando la red..."):
                        uniones = vinculos_que_unen_familias(obj_red["id"])
                    if not uniones:
                        st.success("✅ Ningún vínculo suelto está uniendo grupos grandes en esta red.")
                    else:
                        st.warning(f"⚠️ {len(uniones)} vínculo(s) sostienen la unión de dos grupos. "
                                    "Los más equilibrados y de menor confianza van primero.")
                        st.dataframe(quitar_id(uniones), use_container_width=True, hide_index=True)
                        etiquetas_u = {
                            f"{u['Código A']} ↔ {u['Código B']} (separa {u['Separa']}, "
                            f"confianza {u['Confianza del vínculo']})": (u["_a"], u["_b"])
                            for u in uniones
                        }
                        elegido_u = st.selectbox("¿Cuál cortar?", list(etiquetas_u.keys()),
                                                  key="union_a_cortar")
                        if st.button("✂️ Cortar ese vínculo"):
                            n = borrar_equivalencias_dudosas([etiquetas_u[elegido_u]])
                            invalidar_salud()
                            avisar("success", f"Se cortó el vínculo. Las dos familias quedaron separadas.")
                            st.rerun()
            st.markdown("**🔍 Revisar los vínculos que YA están cargados**")
            explicar(
                "El análisis de confianza mira los vínculos pendientes de revisión, pero el problema "
                "grande está en los que ya entraron:",
                "los que cargaron importaciones viejas que nadie revisó. Esto les pasa el mismo "
                "análisis y te muestra los peores. Hasta ahora la única forma de encontrarlos era "
                "tropezarse con uno buscando un código."
            )
            if st.button("🔎 Analizar los vínculos cargados"):
                with st.spinner("Analizando..."):
                    st.session_state["dudosas_cargadas"] = auditar_equivalencias_cargadas()

            if st.session_state.get("dudosas_cargadas"):
                dudosas, revisadas = st.session_state["dudosas_cargadas"]
                if not dudosas:
                    st.success(f"✅ Se revisaron {revisadas:,} vínculos y ninguno quedó por debajo del "
                                "umbral de confianza.")
                else:
                    st.warning(
                        f"⚠️ De {revisadas:,} vínculos revisados, **{len(dudosas)} tienen evidencia en "
                        "contra**. Están ordenados de peor a mejor, con el motivo al lado."
                    )
                    st.dataframe(quitar_id(dudosas), use_container_width=True, hide_index=True)
                    st.download_button(
                        "⬇️ Bajarlos en Excel antes de decidir",
                        data=to_excel_bytes(quitar_id(dudosas)),
                        file_name="vinculos_dudosos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    # Las opciones se arman a partir de cuántos hay, y el valor por defecto se
                    # elige de esa misma lista. Con value=min(25, N) y N=7, el valor 7 podía no
                    # estar entre las opciones y Streamlit tira excepción: es el mismo error que
                    # ya rompió la carga de fotos una vez.
                    opciones_cortar = sorted({n for n in (10, 25, 50, 100, len(dudosas))
                                               if n <= len(dudosas)}) or [len(dudosas)]
                    st.session_state.setdefault("cuantos_dudosos", opciones_cortar[0])
                    if st.session_state["cuantos_dudosos"] not in opciones_cortar:
                        st.session_state["cuantos_dudosos"] = opciones_cortar[0]
                    cuantos_cortar = st.select_slider(
                        "Cortar los peores:", options=opciones_cortar, key="cuantos_dudosos"
                    )
                    st.caption(
                        "Se cortan los vínculos, NO los productos: precios, stock e historial quedan "
                        "como están. Y quedan anotados como rechazados, así que una reimportación de "
                        "la misma lista no los vuelve a crear."
                    )
                    if st.checkbox("Miré la lista y entiendo qué se corta", key="confirmar_dudosas"):
                        if st.button(f"✂️ Cortar los {cuantos_cortar} peores", type="primary"):
                            pares = [(x["_a"], x["_b"]) for x in dudosas[:int(cuantos_cortar)]]
                            n = borrar_equivalencias_dudosas(pares)
                            st.session_state.pop("dudosas_cargadas", None)
                            invalidar_salud()
                            avisar("success", f"Se cortaron {n} vínculo(s). Los productos quedaron intactos.")
                            st.rerun()
            st.markdown("**💲 Precios que no cierran entre equivalentes**")
            explicar(
                "Dos repuestos que hacen lo mismo pueden costar distinto según la marca, pero no ocho "
                "veces distinto.",
                "Cuando pasa, una de dos cosas está mal — y las dos cuestan plata: o el **precio** "
                "(columna equivocada, separador de decimales al revés), o la **equivalencia** (los "
                "vinculó una lista mal cargada y no son la misma pieza)."
            )
            st.session_state.setdefault("factor_precio_raro", 8)
            factor_precio = st.select_slider(
                "Mostrar los que se diferencien más de:", options=[4, 6, 8, 15, 30],
                format_func=lambda x: f"{x} veces", key="factor_precio_raro"
            )
            incoherentes = precios_incoherentes_entre_equivalentes(int(factor_precio))
            if incoherentes:
                st.warning(f"⚠️ {len(incoherentes)} par(es) de equivalentes con precios muy distintos.")
                st.dataframe(quitar_id(incoherentes), use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇️ Bajar la lista en Excel",
                    data=to_excel_bytes(quitar_id(incoherentes)),
                    file_name="precios_incoherentes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.caption(
                    "Revisá primero los de arriba. Si el precio está bien, entonces lo que está mal "
                    "es el vínculo: cortalo desde «Vincular manual» o con los códigos puente de acá abajo."
                )
            else:
                st.caption(f"✅ Ningún par de equivalentes se diferencia más de {factor_precio} veces.")
            st.markdown("**🗑️ Códigos que son solo un número suelto**")
            explicar(
                "Códigos de 1 o 2 caracteres ('1', '12', '1S'). Casi siempre son cantidades coladas.",
                "Como la app vincula todo lo que aparece en la misma fila, el '1' de una lista queda "
                "como equivalente del '1' de otra, y por ahí se cuelan equivalencias entre repuestos "
                "que no tienen nada que ver.\n\nLas listas nuevas ya los filtran solas; esto limpia "
                "lo que quedó de antes."
            )
            cantidad_basura = contar_codigos_basura()
            if cantidad_basura:
                st.warning(f"⚠️ Hay {cantidad_basura} producto(s) con un código así.")
                muestra_basura = listar_codigos_basura(limite=200)
                with st.expander(f"👀 Ver los primeros {len(muestra_basura)} antes de borrar"):
                    st.dataframe(muestra_basura, use_container_width=True, hide_index=True)
                    st.download_button(
                        "⬇️ Descargar la lista completa",
                        data=to_excel_bytes(listar_codigos_basura(limite=100000)),
                        file_name="codigos_basura.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                st.caption(
                    "Al borrarlos se van también las equivalencias que colgaban de ellos. Esto NO va a "
                    "la papelera: bajate la lista de arriba si querés dejar constancia."
                )
                if st.checkbox("Ya revisé la lista y entiendo que se borran definitivamente",
                                key="confirmar_borrar_basura"):
                    if st.button(f"🗑️ Borrar los {cantidad_basura} códigos y sus equivalencias"):
                        if pedir_password_admin("borrar códigos basura"):
                            borrados = borrar_codigos_basura()
                            invalidar_salud()
                            avisar("success", f"Se borraron {borrados} producto(s) con código basura.")
                            st.rerun()
            else:
                st.caption("✅ Ningún código de 1 o 2 dígitos en la base.")
            st.markdown("**🔢 Códigos que quedaron con '.0'**")
            st.caption(
                "Cuando una lista de Excel trae el código como número, llega con un decimal pegado "
                "(2776400.0). Además de verse mal, eso los volvía imposibles de encontrar: al buscarlos "
                "quedaban con un cero de más. Esto los deja como corresponde."
            )
            cantidad_decimal = contar_codigos_con_decimal()
            if cantidad_decimal:
                st.warning(f"⚠️ Hay {cantidad_decimal} producto(s) con el código terminado en '.0'.")
                if st.button(f"🔧 Arreglar los {cantidad_decimal} códigos"):
                    arreglados = reparar_codigos_con_decimal()
                    # Sin el refresco, el cartel de arriba seguía mostrando el número viejo y
                    # parecía que el botón no hacía nada. El aviso se guarda para que sobreviva.
                    avisar("success", f"Se arreglaron {arreglados} código(s) terminados en '.0'.")
                    invalidar_salud()
                    st.rerun()
            else:
                st.caption("✅ Ningún código con ese problema.")
            st.markdown("**📝 Descripciones con las columnas pegadas**")
            st.caption(
                "Algunas listas exportan varias columnas sin espacio entre medio "
                "(«Junta Tapa de CilindrosFORDTAUNUS»). Esto las separa para que se lean."
            )
            pegadas = contar_descripciones_pegadas()
            if pegadas:
                st.warning(f"⚠️ Hay al menos {pegadas} descripción(es) con ese problema.")
                if st.button("🔧 Separar las descripciones pegadas"):
                    arregladas = reparar_descripciones_pegadas()
                    st.success(f"Se separaron {arregladas} descripción(es).")
            else:
                st.caption("✅ Ninguna descripción con ese problema.")

        with _tabs_mant[1]:
            st.markdown("**🎯 Puntuar los vínculos para el buscador**")
            try:
                c.execute("SELECT COUNT(*) FROM equivalencias WHERE confianza IS NULL")
                sin_puntuar = c.fetchone()[0]
                c.execute("SELECT COUNT(*) FROM equivalencias")
                total_eq = c.fetchone()[0]
            except sqlite3.OperationalError:
                sin_puntuar = total_eq = 0
            explicar(
                "Le pone puntaje a cada vínculo y lo guarda.",
                "El buscador lo usa para decirte, en cada resultado, qué tan sólido es el camino por el "
                "que llegó: **una cadena vale lo que su eslabón más flojo**. Un resultado a dos saltos "
                "por vínculos buenos es más confiable que uno directo colgado de un vínculo malo."
            )
            if sin_puntuar:
                st.info(f"Hay {sin_puntuar:,} vínculo(s) sin puntuar de {total_eq:,}. "
                         "Mientras tanto cuentan como neutros.")
            if total_eq and st.button("🎯 Calcular la confianza de todos los vínculos"):
                barra_conf = st.progress(0.0, text="Puntuando...")
                n = recalcular_confianzas(
                    progreso=lambda i, t: barra_conf.progress(min(i / max(t, 1), 1.0),
                                                              text=f"Puntuando {i:,} de {t:,}...")
                )
                barra_conf.empty()
                invalidar_salud()
                avisar("success", f"Se puntuaron {n:,} vínculo(s). El buscador ya lo está usando.")
                st.rerun()
            st.markdown("**🧾 Equivalencias que confirmó el mostrador**")
            explicar(
                "Cada venta guarda qué código te pidieron y cuál le vendiste.",
                "Cuando son distintos y se repite, eso es una equivalencia confirmada en la práctica: "
                "alguien decidió que servía, el cliente se lo llevó y no volvió a reclamar. Pesa más "
                "que cualquier lista de proveedor — una lista dice lo que el proveedor cree, esto es lo "
                "que pasó."
            )
            try:
                sustituciones = sustituciones_reales()
            except sqlite3.OperationalError:
                sustituciones = []
            if not sustituciones:
                st.caption(
                    "Todavía no hay ninguna repetida. Aparecen solas a medida que vendés reemplazos: "
                    "hace falta que la misma sustitución se dé al menos dos veces, porque una sola "
                    "puede ser un error de tipeo."
                )
            else:
                sin_vincular = [x for x in sustituciones if not x["_ya"]]
                st.dataframe(quitar_id(sustituciones), use_container_width=True, hide_index=True)
                if sin_vincular:
                    st.warning(
                        f"⚠️ {len(sin_vincular)} de estas sustituciones **todavía no están cargadas "
                        "como equivalencia**. Son ventas que ya hiciste: el buscador debería "
                        "encontrarlas solo la próxima vez."
                    )
                    if st.button(f"📥 Mandar esas {len(sin_vincular)} a revisión"):
                        lote_v = f"CONFIRMADAS EN EL MOSTRADOR · {datetime.now():%d/%m %H:%M}"
                        n = guardar_equivalencias_derivadas(
                            [(x["_a"], x["_b"]) for x in sin_vincular], lote_v)
                        invalidar_salud()
                        avisar("success", f"{n} equivalencia(s) quedaron para revisar.")
                        st.rerun()
                else:
                    st.success("✅ Todas las sustituciones repetidas ya están cargadas.")
            st.markdown("**📚 Lo que la app aprendió de tus decisiones**")
            explicar(
                "Cada vez que aprobás o descartás un vínculo, esa decisión queda guardada.",
                "Acá se ve qué patrones sacó de eso y los está usando para puntuar los vínculos nuevos. "
                "Se muestra a propósito: un sistema que aprende a escondidas no se puede corregir."
            )
            patrones_vistos = aprender_de_las_decisiones()
            if not patrones_vistos["marcas"]:
                st.caption(
                    f"Todavía no hay patrones. Llevás {patrones_vistos['total']} decisión(es) "
                    f"revisadas; hacen falta al menos {MINIMO_PARA_APRENDER} sobre una misma "
                    "combinación de marcas para sacar una conclusión. Con menos, la regla sería peor "
                    "que no tener regla."
                )
            else:
                filas_patron = []
                for (ma, mb), d in sorted(patrones_vistos["marcas"].items(),
                                           key=lambda x: -x[1]["decisiones"]):
                    if d["tasa_ok"] >= 0.85:
                        efecto = "✅ suma confianza"
                    elif d["tasa_ok"] <= 0.20:
                        efecto = "❌ resta confianza"
                    else:
                        efecto = "— sin efecto (ni claro que sí ni que no)"
                    filas_patron.append({
                        "Marcas": f"{ma} ↔ {mb}",
                        "Revisados": d["decisiones"],
                        "Aprobaste": f"{d['tasa_ok']*100:.0f}%",
                        "Efecto en los vínculos nuevos": efecto,
                    })
                st.dataframe(filas_patron, use_container_width=True, hide_index=True)
                st.caption(
                    "Si algún patrón no te cierra, corregilo revisando algunos vínculos de esa "
                    "combinación al revés: la app se reajusta sola con las decisiones nuevas."
                )
            st.markdown("**🏭 Catálogo de aplicaciones (qué repuesto le va a cada auto)**")
            explicar(
                "El catálogo que dice a qué auto le va cada repuesto. NGK, Bosch, Mann y SKF "
                "publican el suyo gratis.",
                "Es distinto de una lista de precios: una lista dice cuánto cuesta un código, un "
                "catálogo de aplicaciones dice **a qué auto le va**.\n\nBuscalo en el sitio del "
                "fabricante como «catálogo de aplicaciones» y subilo acá. Con esto, buscar por VIN "
                "o por vehículo deja de adivinar desde las descripciones del proveedor."
            )
            try:
                c.execute("""SELECT marca_repuesto AS "Marca", COUNT(*) AS "Aplicaciones",
                                    COUNT(DISTINCT marca_auto) AS "Marcas de auto",
                                    COUNT(DISTINCT codigo) AS "Códigos"
                             FROM aplicaciones GROUP BY marca_repuesto ORDER BY 2 DESC""")
                ya_cargadas = filas_a_listas(c)
            except sqlite3.OperationalError:
                ya_cargadas = []
            if ya_cargadas:
                st.dataframe(ya_cargadas, use_container_width=True, hide_index=True)

            arch_aplic = subir_archivo("Catálogo de aplicaciones (.pdf o .xlsx):",
                                        ["pdf", "xlsx", "csv"], "aplicaciones")
            ca_m, ca_t = cols(2)
            marca_aplic = ca_m.text_input("¿De qué marca de repuesto es este catálogo?",
                                           placeholder="Ej: NGK, Bosch, Mann", key="marca_aplic").strip()
            # El tipo de pieza no es un dato decorativo: es lo que después permite cruzar catálogos
            # sin mezclar una bujía con un filtro por el solo hecho de ir al mismo auto.
            tipo_aplic = ca_t.selectbox(
                "¿Qué tipo de pieza trae?", ["(elegir)"] + sorted(FAMILIAS_REPUESTO.keys()),
                key="tipo_aplic",
                help="Importante: con esto la app puede después cruzar los catálogos de dos "
                     "fabricantes y deducir equivalencias, sin confundir rubros distintos."
            )
            tipo_aplic = "" if tipo_aplic == "(elegir)" else tipo_aplic
            if arch_aplic:
                archivo_listo(arch_aplic, "catálogo")
                boton_otro_archivo("aplicaciones", "🗑️ Usar otro", key="otro_aplic")
                if st.button("🔎 Leer el catálogo", disabled=not (marca_aplic and tipo_aplic)):
                    with st.spinner("Leyendo... en un PDF grande puede tardar un rato"):
                        try:
                            tablas = tablas_de_archivo(arch_aplic)
                            apps = []
                            for tabla in tablas:
                                # OJO con el nombre: 'cols' es una función de esta app. Llamar así a
                                # la variable la pisa a nivel global y cualquier cols(3) posterior
                                # revienta. Ya pasó antes; por eso la auditoría lo chequea.
                                posiciones = detectar_columnas_aplicaciones(tabla)
                                apps += parsear_catalogo_aplicaciones(
                                    tabla, posiciones["modelo"], posiciones["motor"],
                                    posiciones["comb"], posiciones["anios"], posiciones["codigo"]
                                )
                        except Exception as e:
                            apps = []
                            st.error(f"No se pudo leer: {type(e).__name__}: {e}")
                    if apps:
                        st.session_state["aplic_leidas"] = apps
                        if len({a["marca_auto"] for a in apps}) < 2:
                            st.warning(
                                "⚠️ Se reconoció una sola marca de auto. Si este archivo en realidad "
                                "es una **lista de precios**, no va acá: cargala en "
                                "**📁 Cargar Excel**."
                            )
                        marcas_detectadas = sorted({a["marca_auto"] for a in apps})
                        st.success(f"Se reconocieron {len(apps):,} aplicaciones de "
                                   f"{len(marcas_detectadas)} marca(s) de auto.")
                        st.caption("Revisá esta muestra antes de guardar:")
                        st.dataframe(apps[:40], use_container_width=True, hide_index=True)
                    elif arch_aplic:
                        st.warning(
                            "No reconocí aplicaciones en ese archivo. Esta lectura espera la forma "
                            "habitual de estos catálogos: la marca del auto sola en una fila, y "
                            "debajo modelo, motorización, años y código."
                        )

            if st.session_state.get("aplic_leidas"):
                apps_pend = st.session_state["aplic_leidas"]
                if st.button(f"💾 Guardar las {len(apps_pend):,} aplicaciones", type="primary"):
                    n = guardar_aplicaciones(apps_pend, marca_aplic,
                                              getattr(arch_aplic, "name", "catálogo"), tipo_aplic)
                    st.session_state.pop("aplic_leidas", None)
                    invalidar_salud()
                    avisar("success", f"Se guardaron {n:,} aplicaciones de {marca_aplic.upper()}.")
                    st.rerun()

            try:
                c.execute("""SELECT COUNT(DISTINCT marca_repuesto) FROM aplicaciones
                             WHERE COALESCE(tipo_pieza,'') <> ''""")
                marcas_con_tipo = c.fetchone()[0]
            except sqlite3.OperationalError:
                marcas_con_tipo = 0

            if marcas_con_tipo >= 2:
                st.markdown("**🔗 Equivalencias deducidas cruzando catálogos**")
                explicar(
                    "Si dos fabricantes dicen que sus piezas van al mismo auto, son intercambiables.",
                    "Ninguna lista de proveedor te lo dice: sale de cruzar catálogos que ya "
                    "tenés.\n\nSolo se cruzan códigos del **mismo tipo de pieza**, de fabricantes "
                    "**distintos**, y que coincidan en **varios autos** — con uno solo podría ser "
                    "casualidad."
                )
                if st.button("🔗 Buscar equivalencias entre catálogos"):
                    with st.spinner("Cruzando..."):
                        st.session_state["derivadas"] = derivar_equivalencias_de_aplicaciones()
                derivadas = st.session_state.get("derivadas")
                if derivadas is not None:
                    if not derivadas:
                        st.info(
                            "No salió ninguna. Puede ser que los catálogos cargados sean de tipos de "
                            "pieza distintos, o que sus códigos todavía no estén en tus listas."
                        )
                    else:
                        st.success(f"Se dedujeron {len(derivadas)} equivalencia(s) posibles.")
                        st.dataframe(quitar_id(derivadas), use_container_width=True, hide_index=True)
                        st.caption(
                            "No se cargan directo: van a la cola de revisión, donde el análisis de "
                            "confianza las evalúa como a cualquier otra. Por buena que sea la "
                            "deducción, sigue siendo una deducción."
                        )
                        if st.button(f"📥 Mandar las {len(derivadas)} a revisión", type="primary"):
                            lote_der = f"CATÁLOGOS DE FABRICANTE · {datetime.now():%d/%m %H:%M}"
                            n = guardar_equivalencias_derivadas(
                                [(x["_a"], x["_b"]) for x in derivadas], lote_der)
                            st.session_state.pop("derivadas", None)
                            invalidar_salud()
                            avisar("success", f"{n} equivalencia(s) quedaron para revisar.")
                            st.rerun()
                st.markdown("---")

            st.markdown("**↩️ Deshacer una importación**")
            explicar(
                "Saca de una todos los vínculos que dejó una lista.",
                "Es la red de seguridad que faltaba: hasta ahora, si una lista venía mal mapeada, la "
                "única salida era borrar de a uno entre miles. **Los productos no se tocan**: precios, "
                "stock, ubicación e historial quedan como están. Se deshace solo la parte peligrosa, "
                "que son los vínculos."
            )
            importaciones = listar_importaciones_deshacibles()
            if importaciones:
                st.dataframe(quitar_id(importaciones), use_container_width=True, hide_index=True)
                etiquetas_imp = {
                    f"{i['Marca']} · {i['Archivo']} · {i['Fecha']} "
                    f"({i['Vínculos vivos']} vivos, {i['Sin revisar']} sin revisar)": i["_lote"]
                    for i in importaciones if (i["Vínculos vivos"] or i["Sin revisar"])
                }
                if etiquetas_imp:
                    elegida_imp = st.selectbox("¿Cuál deshacer?", list(etiquetas_imp.keys()),
                                                key="importacion_deshacer")
                    lote_elegido = etiquetas_imp[elegida_imp]
                    previo = previsualizar_deshacer(lote_elegido)
                    st.warning(
                        f"Se van a borrar **{previo['vinculos']} vínculo(s) ya cargados** y "
                        f"**{previo['pendientes']} sin revisar**. Quedan registrados como rechazados, "
                        "así que si volvés a importar la misma lista no se vuelven a crear solos."
                    )
                    if st.checkbox("Entiendo que esto borra esos vínculos", key="confirmar_deshacer_imp"):
                        if st.button("↩️ Deshacer esta importación", type="primary"):
                            if pedir_password_admin("deshacer una importación"):
                                nv, npd = deshacer_importacion(lote_elegido)
                                invalidar_salud()
                                avisar("success", f"Se deshicieron {nv} vínculo(s) cargados y {npd} pendientes.")
                                st.rerun()
            else:
                st.caption(
                    "Todavía no hay importaciones con origen registrado. Las listas que importes de "
                    "acá en adelante van a poder deshacerse; las anteriores no guardaron de dónde "
                    "venía cada vínculo."
                )

            espejadas = contar_equivalencias_espejadas()
            if espejadas:
                st.markdown("---")
                st.markdown("**🔁 Equivalencias anotadas dos veces**")
                st.warning(
                    f"⚠️ Hay {espejadas:,} equivalencia(s) guardadas por duplicado: la misma relación "
                    "anotada en las dos direcciones (A↔B y B↔A). No son vínculos distintos, es la "
                    "misma información dos veces."
                )
                explicar(
                    "No cambia lo que encuentra el buscador —consulta las dos columnas igual—, pero sí "
                    "infla todos los conteos:",
                    "un código con 100 equivalencias reales figura con 200, y el control de precios lista "
                    "cada par dos veces. Unificarlas no borra ninguna equivalencia, solo deja una sola fila "
                    "por cada una."
                )
                if st.button("🔁 Unificar duplicadas"):
                    borradas, vueltas = unificar_equivalencias_espejadas()
                    invalidar_salud()
                    avisar("success", f"Se unificaron {borradas:,} duplicadas "
                                       f"y se ordenaron {vueltas:,}.")
                    st.rerun()

        with _tabs_mant[2]:
            st.markdown("**📷 Traer fotos de productos en tanda**")
            st.caption(
                "En vez de cargarlas de a una. No existe ninguna base pública y gratuita de fotos por "
                "número de parte (la del rubro, TecDoc, es paga), así que las dos fuentes confiables son: "
                "el link que ya venga en tu lista, o la ficha del propio proveedor."
            )

            modo_liviano = st.checkbox(
                "Modo liviano (recomendado): guardar el link de la foto, no la foto entera",
                value=True, key="fotos_modo_liviano",
                help="La foto se sigue viendo y se sigue pudiendo buscar por parecido, porque la firma "
                     "visual y la miniatura sí se guardan. Lo que no se guarda es la imagen grande: "
                     "esa la trae el navegador desde el sitio del proveedor."
            )
            kb_por_foto = peso_estimado_por_foto(modo_liviano)

            pendientes_url = contar_fotos_por_bajar()
            if pendientes_url:
                st.info(f"Hay {pendientes_url} producto(s) con la foto como link externo, sin bajar.")
                # OJO: en select_slider el value TIENE que ser uno de los options. Antes acá iba
                # min(500, pendientes) y con 347 pendientes tiraba excepción y se caía toda la
                # pestaña de Mantenimiento — de ahí que la carga de fotos apareciera "rota".
                opciones_url = [100, 250, 500, 1000, 2000]
                st.session_state.setdefault("cuantas_fotos_url", 500)
                cuantas_url = st.select_slider("¿Cuántas bajar?", options=opciones_url,
                                                key="cuantas_fotos_url")
                if st.button(f"⬇️ Bajar {cuantas_url} fotos de esos links"):
                    barra = st.progress(0.0, text="Bajando fotos...")
                    bajadas, fallidas = bajar_fotos_pendientes(
                        int(cuantas_url), liviano=modo_liviano,
                        progreso=lambda i, t: barra.progress(i / max(t, 1), text=f"Foto {i} de {t}...")
                    )
                    barra.empty()
                    st.success(f"Se bajaron {bajadas} foto(s).")
                    if fallidas:
                        with st.expander(f"⚠️ {len(fallidas)} no se pudieron bajar"):
                            for url_f, motivo in fallidas[:30]:
                                st.caption(f"- {str(url_f)[:60]}: {motivo}")
            else:
                st.caption("✅ No hay fotos pendientes de bajar desde links.")

            c.execute("""SELECT m.id, m.nombre, COUNT(p.id) AS sin_foto
                         FROM marcas m JOIN productos p ON p.marca_id = m.id
                         WHERE m.url_ficha_template IS NOT NULL AND m.url_ficha_template <> ''
                           AND p.imagen_url IS NULL
                           AND (p.foto_busqueda_estado IS NULL OR p.foto_busqueda_estado = 'error')
                         GROUP BY m.id HAVING sin_foto > 0 ORDER BY sin_foto DESC""")
            marcas_con_catalogo = [dict(r) for r in c.fetchall()]

            if marcas_con_catalogo:
                st.markdown("**Traerlas desde la ficha del proveedor**")
                opciones_mc = {f"{m['nombre']} ({m['sin_foto']} por probar)": m["id"] for m in marcas_con_catalogo}
                elegida_mc = st.selectbox("Marca:", list(opciones_mc.keys()), key="marca_fotos_catalogo")
                id_marca_fotos = opciones_mc[elegida_mc]

                etiquetas_filtro = {
                    "stock": "Solo los que tienen stock",
                    "precio": "Solo los que tienen precio cargado",
                    "todos": "Todos los códigos de la marca",
                }
                filtro_fotos = st.radio(
                    "¿Para cuáles traer foto?", list(etiquetas_filtro.keys()),
                    format_func=lambda k: etiquetas_filtro[k], key="filtro_fotos_catalogo",
                    help="Traer foto de TODO un catálogo de miles de códigos llena la base y hace que "
                         "el backup ya no entre en GitHub. Empezar por los que tienen stock cubre lo "
                         "que realmente movés."
                )
                faltan_marca = contar_fotos_por_traer_de_catalogo(id_marca_fotos, filtro_fotos)
                faltan_todos = contar_fotos_por_traer_de_catalogo(id_marca_fotos, "todos")
                if filtro_fotos != "todos":
                    st.caption(f"Con ese filtro quedan {faltan_marca:,} de {faltan_todos:,} códigos.")

                explicar(
                    "Entra a la ficha de cada código y busca la foto ahí.",
                    "Ahora va de a 6 fichas a la vez en vez de una por una, así que rinde bastante más. Los "
                    "códigos cuya ficha no tiene foto quedan marcados y no se vuelven a consultar, para que "
                    "cada tanda avance de verdad."
                )

                cf_a, cf_b = st.columns(2)
                with cf_a:
                    st.session_state.setdefault("tanda_fotos_catalogo", 500)
                    tanda_fotos = st.select_slider(
                        "Fotos por tanda:", options=[100, 250, 500, 1000, 2000],
                        key="tanda_fotos_catalogo"
                    )
                with cf_b:
                    st.session_state.setdefault("hilos_fotos", 6)
                    hilos_fotos = st.select_slider(
                        "Fichas a la vez:", options=[3, 6, 10, 15], key="hilos_fotos",
                        help="Más rápido, pero es el sitio del proveedor el que atiende: si te pasás "
                             "puede empezar a rechazar los pedidos y no baja ninguna."
                    )

                mb_estimados = faltan_marca * kb_por_foto / 1024
                st.caption(
                    f"Cada foto ocupa unos {kb_por_foto} KB en la base "
                    f"({'modo liviano' if modo_liviano else 'guardando la imagen entera'}). "
                    f"Las {faltan_marca:,} de este filtro serían unos {mb_estimados:,.0f} MB."
                )
                if mb_estimados > 80:
                    st.warning(
                        f"⚠️ {mb_estimados:,.0f} MB no entran en el backup de GitHub (tope 100 MB), y "
                        "como el hosting borra el disco al reiniciar y restaura desde ahí, esas fotos "
                        "se te van a perder en cada reinicio. "
                        + ("Probá con «solo los que tienen stock»: es lo que realmente movés."
                            if filtro_fotos == "todos" else
                            "Aun con el filtro es mucho: convendría traerlas de a poco, empezando por "
                            "las marcas que más vendés.")
                    )

                en_curso = st.session_state.get("bajada_fotos_en_curso")

                bc1, bc2 = st.columns(2)
                if not en_curso:
                    if bc1.button(f"🌐 Traer una tanda de {tanda_fotos}", type="primary"):
                        st.session_state["bajada_fotos_en_curso"] = {
                            "marca_id": id_marca_fotos, "restantes": int(tanda_fotos),
                            "bajadas": 0, "sin_foto": 0, "fallos": 0, "hasta_terminar": False,
                        }
                        st.rerun()
                    if bc2.button(f"♾️ Seguir hasta terminar las {faltan_marca:,}",
                                   disabled=not faltan_marca):
                        st.session_state["bajada_fotos_en_curso"] = {
                            "marca_id": id_marca_fotos, "restantes": faltan_marca,
                            "bajadas": 0, "sin_foto": 0, "fallos": 0, "hasta_terminar": True,
                        }
                        st.rerun()
                else:
                    if st.button("⏹️ Detener", type="primary"):
                        st.session_state.pop("bajada_fotos_en_curso", None)
                        st.rerun()
                    if en_curso["marca_id"] != id_marca_fotos:
                        # Si no se avisa, cambiar de marca en el selector deja la bajada colgada:
                        # el bloque de abajo no corre, no se refresca sola, y parece que se trabó.
                        c.execute("SELECT nombre FROM marcas WHERE id = ?", (en_curso["marca_id"],))
                        otra = c.fetchone()
                        st.warning(
                            f"Hay una bajada en curso de **{otra['nombre'] if otra else 'otra marca'}** "
                            f"({en_curso['bajadas']} foto(s) hasta ahora). Volvé a elegir esa marca para "
                            "que siga, o tocá Detener."
                        )

                # La bajada se hace en pedazos, y entre pedazo y pedazo la pantalla se refresca. Si se
                # hiciera todo de un saque, una tanda de 2.000 fotos serían 20 minutos con la pantalla
                # colgada — y el navegador o el hosting cortan la conexión mucho antes de eso.
                if en_curso and en_curso["marca_id"] == id_marca_fotos:
                    # Pedazos chicos a propósito: entre pedazo y pedazo es cuando se puede tocar
                    # "Detener", así que con tandas grandes el botón tardaría un minuto en responder.
                    pedazo = min(en_curso["restantes"], 90)
                    barra2 = st.progress(
                        0.0, text=f"Consultando fichas... (llevamos {en_curso['bajadas']} foto(s))"
                    )
                    bajadas2, fallidas2, sin_foto2 = bajar_fotos_desde_catalogo(
                        id_marca_fotos, pedazo, liviano=modo_liviano, hilos=int(hilos_fotos),
                        filtro=filtro_fotos,
                        progreso=lambda i, t: barra2.progress(min(i / max(t, 1), 1.0),
                                                              text=f"Ficha {i} de {t}...")
                    )
                    barra2.empty()
                    en_curso["bajadas"] += bajadas2
                    en_curso["sin_foto"] += sin_foto2
                    en_curso["fallos"] += max(len(fallidas2) - sin_foto2, 0)
                    procesadas = bajadas2 + len(fallidas2)
                    en_curso["restantes"] -= max(procesadas, 1)

                    if procesadas == 0 or en_curso["restantes"] <= 0:
                        st.session_state.pop("bajada_fotos_en_curso", None)
                        st.success(
                            f"Listo: {en_curso['bajadas']} foto(s) traídas, "
                            f"{en_curso['sin_foto']} código(s) sin foto en la ficha, "
                            f"{en_curso['fallos']} con error de red."
                        )
                        if fallidas2:
                            with st.expander(f"Ver los últimos {min(len(fallidas2), 30)} que no salieron"):
                                for cod_f, motivo in fallidas2[:30]:
                                    st.caption(f"- {cod_f}: {motivo}")
                    else:
                        st.session_state["bajada_fotos_en_curso"] = en_curso
                        st.caption(
                            f"Van {en_curso['bajadas']} foto(s) — quedan unas {en_curso['restantes']:,}. "
                            "Dejá esta pantalla abierta; sigue sola."
                        )
                        st.rerun()
            else:
                st.caption(
                    "Para traer fotos del catálogo hace falta cargar la dirección de la ficha de la marca "
                    "en Administrar → Marcas."
                )

            c.execute("SELECT COUNT(*) FROM productos WHERE foto_busqueda_estado = 'sin_foto'")
            marcados_sin_foto = c.fetchone()[0]
            if marcados_sin_foto:
                st.caption(
                    f"🔕 {marcados_sin_foto:,} código(s) quedaron marcados como «la ficha no tiene foto» "
                    "y ya no se vuelven a consultar."
                )
                if st.button("🔄 Volver a probar esos códigos"):
                    avisar("success", f"Se rehabilitaron {reintentar_codigos_sin_foto()} código(s).")
                    st.rerun()

        with _tabs_mant[3]:
            st.markdown("**🔍 Salud de los datos**")
            st.caption(
                "Revisa la base en busca de cosas rotas o inconsistentes — útil para detectar corrupción "
                "de datos antes de encontrártela buscando un producto."
            )
            if st.button("🔍 Revisar salud de los datos"):
                reporte_salud = chequear_integridad_bd()
                total_problemas = sum(r["Problemas"] for r in reporte_salud)
                if total_problemas == 0:
                    st.success("✅ Todo en orden, no se encontró ningún problema.")
                else:
                    st.warning(f"⚠️ Se encontraron {total_problemas} problema(s) en total.")
                st.dataframe(
                    [r for r in reporte_salud],
                    use_container_width=True, hide_index=True,
                    column_config={"Problemas": st.column_config.NumberColumn(
                        "Problemas", help="0 está bien; más de 0 conviene revisarlo"
                    )}
                )
            st.markdown("**Limpieza de la base**")
            st.caption(
                "Con el tiempo pueden quedar códigos cargados por error que no están vinculados a "
                "ninguna equivalencia. Este botón los borra."
            )
            if st.button("🧹 Borrar productos sin ninguna equivalencia"):
                if pedir_password_admin("borrar productos sin equivalencias"):
                    borrados = depurar_huerfanos()
                    if borrados:
                        st.success(f"Se borraron {borrados} producto(s) sin equivalencias.")
                    else:
                        st.info("No había productos sueltos para borrar.")
            st.markdown("**🗑️ Papelera**")
            explicar(
                "Cuando borrás una marca entera, un combo, un alias de transferencia o un producto "
                "puntual (por ID), queda acá guardado por si te equivocaste.",
                "Se borra en forma permanente solo cuando vos lo pedís o pasan más de 30 días. "
                "(Fusionar marcas y restaurar un backup completo siguen siendo irreversibles — esos no "
                "pasan por acá.)"
            )
            items_papelera = listar_papelera()
            if not items_papelera:
                st.caption("La papelera está vacía.")
            else:
                iconos_tipo = {"combo": "🧩", "alias": "💳", "producto": "📦", "marca": "🏷️"}
                for item in items_papelera:
                    colp1, colp2, colp3 = st.columns([3, 1, 1])
                    icono = iconos_tipo.get(item["Tipo"], "🗑️")
                    colp1.write(
                        f"{icono} {item['Tipo'].capitalize()}: **{item['Detalle']}** — "
                        f"eliminado por {item['Eliminado por'] or 'alguien'} el {item['Fecha']}"
                    )
                    colp2.button("↩️ Restaurar", key=f"restaurar_papelera_{item['ID']}",
                                  on_click=cb_restaurar_papelera, args=(item["ID"],))
                    colp3.button("🗑️", key=f"borrar_papelera_{item['ID']}", help="Borrar en forma permanente, sin restaurar",
                                  on_click=borrar_papelera_definitivo, args=(item["ID"],))

                resultado_papelera = st.session_state.pop("resultado_papelera", None)
                if resultado_papelera:
                    tipo_res, msg_res = resultado_papelera
                    (st.success if tipo_res == "ok" else st.error)(msg_res)

                st.caption("También se limpia sola: lo que lleva más de 30 días acá se borra en forma permanente.")
                st.button("🧹 Vaciar ahora lo de más de 30 días", on_click=vaciar_papelera_antigua, args=(30,))

        if sub_admin == SUB_ADMIN[5]:
            if not pedir_password_admin("gestionar usuarios"):
                pass
            else:
                st.markdown("**👤 Empleados (admin / operador)**")
                explicar(
                    "Cuentas creadas desde acá, sin necesidad de tocar la configuración de Streamlit Cloud.",
                    "'Admin' puede todo, incluso borrar y configurar. 'Operador' puede usar las funciones "
                    "de IA y cargar cosas, pero no borrar ni configurar nada sensible."
                )
                usuarios_actuales = listar_usuarios()
                if usuarios_actuales:
                    st.dataframe(usuarios_actuales, use_container_width=True, hide_index=True)

                cu1, cu2 = st.columns(2)
                nombre_nuevo_usuario = cu1.text_input("Nombre:", key="nuevo_usuario_nombre")
                password_nuevo_usuario = cu2.text_input("Contraseña:", type="password", key="nuevo_usuario_pass")
                rol_nuevo_usuario = st.selectbox("Rol:", ["operador", "admin"], key="nuevo_usuario_rol")
                if st.button("➕ Crear empleado"):
                    if not nombre_nuevo_usuario.strip() or not password_nuevo_usuario:
                        st.warning("Completá nombre y contraseña.")
                    else:
                        try:
                            crear_usuario(nombre_nuevo_usuario, password_nuevo_usuario, rol_nuevo_usuario)
                            avisar("success", f"Empleado '{nombre_nuevo_usuario}' creado.")
                            st.rerun()
                        except sqlite3.IntegrityError:
                            st.error("Ya existe un empleado con ese nombre.")

                if usuarios_actuales:
                    st.markdown("**Gestionar un empleado existente**")
                    opciones_usuario = {u["Nombre"]: u["ID"] for u in usuarios_actuales}
                    usuario_elegido = st.selectbox("Elegí un empleado:", list(opciones_usuario.keys()), key="sel_usuario_gestionar")
                    usuario_id_sel = opciones_usuario[usuario_elegido]
                    cug1, cug2, cug3 = cols(3)
                    nueva_pass_usuario = cug1.text_input("Nueva contraseña (opcional):", type="password", key="usuario_nueva_pass")
                    if cug1.button("💾 Cambiar contraseña"):
                        if nueva_pass_usuario:
                            cambiar_password_usuario(usuario_id_sel, nueva_pass_usuario)
                            st.success("Contraseña actualizada.")
                        else:
                            st.warning("Escribí la nueva contraseña primero.")
                    usuario_activo_actual = next(u["Activo"] == "Sí" for u in usuarios_actuales if u["ID"] == usuario_id_sel)
                    if cug2.button("🚫 Desactivar" if usuario_activo_actual else "✅ Reactivar"):
                        activar_desactivar_usuario(usuario_id_sel, not usuario_activo_actual)
                        st.rerun()
                    if cug3.button("🗑️ Eliminar empleado"):
                        eliminar_usuario(usuario_id_sel)
                        avisar("success", "Empleado eliminado.")
                        st.rerun()

                st.markdown("---")
                st.markdown("**🔧 Mecánicos externos**")
                st.caption(
                    "Cuentas separadas para mecánicos que no son empleados tuyos — solo ven su propio "
                    "portal para armar presupuestos con su mano de obra, nunca las secciones internas."
                )
                mecanicos_actuales = listar_mecanicos()
                if mecanicos_actuales:
                    st.dataframe(mecanicos_actuales, use_container_width=True, hide_index=True)

                cm1, cm2 = st.columns(2)
                nombre_nuevo_mecanico = cm1.text_input("Nombre:", key="nuevo_mecanico_nombre")
                password_nuevo_mecanico = cm2.text_input("Contraseña:", type="password", key="nuevo_mecanico_pass")
                if st.button("➕ Crear mecánico"):
                    if not nombre_nuevo_mecanico.strip() or not password_nuevo_mecanico:
                        st.warning("Completá nombre y contraseña.")
                    else:
                        try:
                            crear_mecanico(nombre_nuevo_mecanico, password_nuevo_mecanico)
                            avisar("success", f"Mecánico '{nombre_nuevo_mecanico}' creado.")
                            st.rerun()
                        except sqlite3.IntegrityError:
                            st.error("Ya existe un mecánico con ese nombre.")

                if mecanicos_actuales:
                    st.markdown("**Gestionar un mecánico existente**")
                    opciones_mecanico = {m["Nombre"]: m["ID"] for m in mecanicos_actuales}
                    mecanico_elegido = st.selectbox("Elegí un mecánico:", list(opciones_mecanico.keys()), key="sel_mecanico_gestionar")
                    mecanico_id_sel = opciones_mecanico[mecanico_elegido]
                    cmg1, cmg2 = st.columns(2)
                    mecanico_activo_actual = next(m["Activo"] == "Sí" for m in mecanicos_actuales if m["ID"] == mecanico_id_sel)
                    if cmg1.button("🚫 Desactivar" if mecanico_activo_actual else "✅ Reactivar", key="toggle_mecanico"):
                        activar_desactivar_mecanico(mecanico_id_sel, not mecanico_activo_actual)
                        st.rerun()
                    if cmg2.button("🗑️ Eliminar mecánico"):
                        eliminar_mecanico(mecanico_id_sel)
                        avisar("success", "Mecánico eliminado.")
                        st.rerun()
# ============================================================
# ESTADÍSTICAS
# ============================================================
if pagina == PAGINAS[4]:
    st.subheader("📊 Estadísticas")

    SUB_STATS = ["📈 Resumen", "📥 Importaciones", "💾 Backup y config", "🧮 Auditoría y depósito",
               "🔎 Búsquedas sin resultado", "📌 Para pedir", "🔗 Equivalencias sugeridas"]
    if st.session_state.get("sub_stats") not in SUB_STATS:
        st.session_state["sub_stats"] = SUB_STATS[0]
    st.radio("Sub-sección:", SUB_STATS, key="sub_stats", horizontal=True,
             label_visibility="collapsed")
    sub_stats = st.session_state["sub_stats"]

    if sub_stats == SUB_STATS[0]:
        st.subheader("Estadísticas generales")

        c.execute("SELECT COUNT(*) FROM marcas")
        total_marcas = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM productos")
        total_productos = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM equivalencias")
        total_equiv = c.fetchone()[0]

        m1, m2, m3 = st.columns(3)
        m1.metric("Marcas registradas", total_marcas)
        m2.metric("Códigos cargados", total_productos)
        m3.metric("Vínculos de equivalencia", total_equiv // 2 if total_equiv else 0)

        st.markdown("---")
        c.execute("""SELECT m.nombre, COUNT(p.id) AS productos
                     FROM marcas m LEFT JOIN productos p ON p.marca_id = m.id
                     GROUP BY m.id ORDER BY productos DESC LIMIT 15""")
        top_marcas = c.fetchall()
        if top_marcas:
            st.markdown("---")
            st.markdown("**📊 Top marcas por cantidad de códigos cargados**")
            chart_data = {"Marca": [t["nombre"] for t in top_marcas],
                           "Productos": [t["productos"] for t in top_marcas]}
            st.bar_chart(chart_data, x="Marca", y="Productos")

        st.markdown("---")
        st.markdown("**🤖 Uso de las funciones de IA (últimos 30 días)**")
        st.caption(
            "Las primeras 4 funciones usan una API key; en el peor caso fallan por límite de uso y "
            "hay que reintentar. 'Generar imagen orientativa' usa una key aparte, configurada por separado."
        )
        uso_ia_actual = resumen_uso_ia()
        if uso_ia_actual:
            st.dataframe(uso_ia_actual, use_container_width=True, hide_index=True)
        else:
            st.caption("Todavía no se usó ninguna función de IA.")

    if sub_stats == SUB_STATS[1]:
        st.markdown("**Historial de importaciones**")
        c.execute("""SELECT marca AS Marca, archivo AS Archivo, filas_cargadas AS Cargadas,
                     filas_omitidas AS Omitidas, fecha AS Fecha FROM importaciones
                     ORDER BY fecha DESC LIMIT 20""")
        imports = filas_a_listas(c)
        if imports:
            st.dataframe(imports, use_container_width=True, hide_index=True)
        else:
            st.caption("Todavía no se registraron importaciones.")

    if sub_stats == SUB_STATS[2]:
        cantidad_fotos, mb_fotos = peso_de_las_fotos()
        c.execute("SELECT COUNT(*) FROM productos")
        total_prod_backup = c.fetchone()[0]

        st.markdown("**🗄️ Backup de la base**")

        # Lo primero: cuánto se perdería HOY. Es la diferencia entre lo que hay en el disco (que
        # se borra) y lo que hay en el repositorio (que sobrevive).
        _riesgo = None
        try:
            _riesgo = cuanto_perderias_si_reinicia()
        except Exception:
            pass
        if _riesgo:
            if not _riesgo["hay_semilla"]:
                st.error(
                    f"🔴 **No hay copia en el repositorio.** Tenés "
                    f"{_riesgo['productos_ahora']:,} producto(s) viviendo solo en el disco del "
                    "servidor, que se borra en cada reinicio. **Hoy un reinicio borra todo.**"
                )
            elif _riesgo["en_riesgo"]:
                st.warning(
                    f"⚠️ **{_riesgo['en_riesgo']:,} producto(s) y "
                    f"{_riesgo['equivalencias_en_riesgo']:,} vínculo(s) viven solo en el disco.**\n\n"
                    f"La copia del repositorio es del {_riesgo['fecha_semilla']} y tiene "
                    f"{_riesgo['productos_semilla']:,} productos; hoy tenés "
                    f"{_riesgo['productos_ahora']:,}. Si el servidor reinicia, esa diferencia "
                    "se pierde."
                )
            else:
                st.success(
                    f"✅ La copia del repositorio está al día ({_riesgo['productos_semilla']:,} "
                    f"productos, del {_riesgo['fecha_semilla']}). Un reinicio no te haría perder nada."
                )
            # Si están los secretos de GitHub, se puede resolver de un botón en vez de a mano.
            _cfg_gh = config_github()
            if _cfg_gh:
                st.info(
                    f"🔗 **Subida automática configurada** hacia `{_cfg_gh['repo']}`. "
                    "No hace falta bajar nada ni entrar a GitHub."
                )
                _ult_gh = obtener_config("ultimo_backup_github", "")
                if _ult_gh:
                    st.caption(f"Última subida automática: {_ult_gh}")
                if st.button("☁️ Subir el backup al repositorio ahora", type="primary"):
                    with st.spinner("Armando la copia y subiéndola..."):
                        _datos = generar_backup_sin_fotos()
                        _ok, _msg = subir_backup_a_github(_datos)
                    if _ok:
                        invalidar_salud()
                        avisar("success", f"☁️ {_msg}")
                        st.rerun()
                    else:
                        st.error(_msg)
            else:
                explicar(
                    "Se puede automatizar: que la app suba el backup sola al repositorio.",
                    "En **Settings → Secrets** de Streamlit Cloud, agregá estas dos líneas:\n\n"
                    "```\ngithub_token = \"ghp_tu_token\"\ngithub_repo = \"usuario/repositorio\"\n"
                    "```\n\nEl token se saca en GitHub → Settings → Developer settings → "
                    "Personal access tokens, con permiso de escritura (`Contents: read and "
                    "write`) sobre ese repositorio.\n\nCon eso configurado aparece un botón "
                    "para subir la copia de una, y además se sube sola una vez por día. "
                    "Mientras no lo configures, hay que hacerlo a mano como hasta ahora."
                )

            explicar(
                "Bajar el backup no alcanza: hay que subirlo al repositorio.",
                "El servidor borra su disco cada vez que la app se reinicia o se "
                "redespliega. Lo único que sobrevive son los archivos del **repositorio de "
                "GitHub**, porque son parte del despliegue.\n\nPor eso el backup se restaura "
                "solo desde un archivo llamado `datos_iniciales.db` que tiene que estar ahí. "
                "Bajarlo a tu teléfono te sirve a vos, pero **no protege a la app**: hasta que "
                "ese archivo no esté en GitHub, un reinicio se lleva todo lo cargado desde la "
                "última vez.\n\n**Los pasos:** bajá el backup de acá abajo → entrá al "
                "repositorio en GitHub → subí el archivo con el nombre exacto "
                "`datos_iniciales.db`, reemplazando el que está."
            )
            st.markdown("---")

        if cantidad_fotos:
            st.error(
                "🔴 **Importante sobre las fotos.** El hosting borra el disco cada vez que la app se "
                "reinicia o se redespliega, y al arrancar se restaura sola desde el `datos_iniciales.db` "
                "del repositorio. Ese archivo lo genera el **backup sin fotos**, que las saca a "
                "propósito para no pasarse del límite de GitHub. Resultado: **cada reinicio te borra "
                "todas las fotos**, y por eso la búsqueda por parecido aparece sin nada para comparar."
            )
            with st.expander("¿Y entonces qué hago con las fotos?"):
                st.markdown(
                    "- **Si son pocas y el archivo entra en GitHub (menos de 100 MB):** usá el "
                    "**backup completo** de acá abajo y subilo al repositorio renombrado a "
                    "`datos_iniciales.db`. Ahí sí sobreviven los reinicios.\n"
                    "- **Si ya no entra:** subí el backup sin fotos (para no perder el catálogo) y "
                    "recuperá las fotos después desde **Mantenimiento → Traer fotos en tanda**, que "
                    "las vuelve a bajar de las fichas de los proveedores sin cargarlas a mano.\n"
                    "- **Lo más prolijo a futuro:** guardar las fotos por dirección web en vez de "
                    "adentro de la base, y dejar cargada la dirección del catálogo de cada marca en "
                    "**Administrar → Marcas**. Así el backup queda liviano y las fotos se vuelven a "
                    "traer solas."
                )

        if mb_fotos > 60:
            st.warning(
                f"⚠️ Las fotos ({cantidad_fotos} productos) ocupan unos {mb_fotos:,.0f} MB. "
                "GitHub no acepta archivos de más de 100 MB, así que para la copia del repositorio "
                "conviene usar el **backup sin fotos** de acá abajo."
            )
        elif cantidad_fotos:
            st.caption(f"Las fotos de {cantidad_fotos} producto(s) ocupan {mb_fotos:,.1f} MB de la base.")

        cbk1, cbk2 = st.columns(2)
        with cbk1:
            st.markdown("*Completo (con fotos)*")
            if st.button("🗄️ Preparar backup completo"):
                with open(DB_PATH, "rb") as f:
                    st.session_state["backup_bytes"] = f.read()
            if "backup_bytes" in st.session_state:
                st.download_button(
                    f"⬇️ Descargar ({len(st.session_state['backup_bytes'])/(1024*1024):,.0f} MB)",
                    data=st.session_state["backup_bytes"],
                    file_name=f"equivalencias_backup_{datetime.now():%Y%m%d}.db",
                    on_click=marcar_backup_hecho
                )
        with cbk2:
            st.markdown("*Liviano (sin fotos) — para el repositorio*")
            if st.button("🪶 Preparar backup sin fotos"):
                with st.spinner("Armando..."):
                    st.session_state["backup_liviano"] = generar_backup_sin_fotos()
            if "backup_liviano" in st.session_state:
                st.download_button(
                    f"⬇️ Descargar ({len(st.session_state['backup_liviano'])/(1024*1024):,.1f} MB)",
                    data=st.session_state["backup_liviano"],
                    file_name="datos_iniciales.db",
                    help="Ya viene con el nombre listo para subir al repositorio",
                    on_click=marcar_backup_hecho
                )
                st.caption(
                    f"Lleva los {total_prod_backup} productos con precios, equivalencias, vehículos e "
                    "historial. ⚠️ **No lleva las fotos**: si restaurás desde este archivo hay que "
                    "volver a traerlas desde Mantenimiento."
                )

        st.markdown("---")
        st.markdown("**📦 Exportar configuración (sin el catálogo de productos)**")
        st.caption(
            "Combos de repuestos, códigos DTC y fabricantes por WMI en un solo archivo de texto — útil "
            "como respaldo liviano aparte del backup completo, o para copiarle la configuración a otra "
            "sucursal sin duplicar todo el catálogo de productos."
        )
        st.download_button(
            "⬇️ Descargar configuración (.txt)",
            data=exportar_configuracion_txt(),
            file_name=f"configuracion_{datetime.now():%Y%m%d}.txt",
            mime="text/plain"
        )

        st.markdown("---")
        st.markdown("**🛡️ Copia permanente (para que no se borre nunca más)**")
        explicar(
            "El servidor borra el disco de la app cada vez que se redespliega o se reinicia — por "
            "eso se pierde la base.",
            "Pero los archivos que están en el repositorio de GitHub **sí** sobreviven, porque "
            "forman parte del despliegue. Con esto se aprovecha eso:"
        )
        st.markdown("""
1. Tocá **🪶 Preparar backup sin fotos** acá arriba y descargalo (ya viene con el nombre correcto).
2. Subilo al repositorio de GitHub, al lado de `app.py`.

Listo: cada vez que el servidor borre el disco, la app se levanta sola con esos datos.
Repetí los 2 pasos cada tanto (una vez por semana, o después de cargar una lista grande).

**¿Por qué sin fotos?** Porque son lo que más pesa y GitHub rechaza archivos de más de 100 MB.
Sin ellas el archivo queda chico, y las fotos se vuelven a traer solas desde
Administrar → Mantenimiento.
        """)
        if os.path.exists(ARCHIVO_SEMILLA):
            try:
                marca_tiempo = datetime.fromtimestamp(os.path.getmtime(ARCHIVO_SEMILLA))
                peso = os.path.getsize(ARCHIVO_SEMILLA) / (1024 * 1024)
                st.success(f"✅ Hay una copia en el repositorio ({peso:,.1f} MB, del {marca_tiempo:%d/%m/%Y}).")
            except Exception:
                st.success("✅ Hay una copia en el repositorio.")
        else:
            st.warning(
                "⚠️ Todavía no hay ninguna copia en el repositorio. Mientras no la subas, cada "
                "redespliegue borra todo lo cargado."
            )

        st.markdown("---")
        st.markdown("**♻️ Restaurar desde un backup**")
        st.caption(
            "⚠️ Esto reemplaza TODA la base actual por la del archivo que subas. "
            "Usalo si el hosting se reinició y perdiste datos, o para volver a un backup anterior."
        )
        archivo_restaurar = subir_archivo("Subí un archivo .db de backup:", ["db"], "restaurar")
        if archivo_restaurar:
            archivo_listo(archivo_restaurar, "backup")
            boton_otro_archivo("restaurar", "🗑️ Usar otro backup", key="otro_backup")
        confirmar_restore = st.checkbox("Entiendo que esto borra los datos actuales y los reemplaza")
        if st.button("♻️ Restaurar backup", disabled=not (archivo_restaurar and confirmar_restore)):
            if pedir_password_admin("restaurar un backup"):
                restaurar_backup(archivo_restaurar)
                avisar("success", "Backup restaurado. Recargando...")
                st.rerun()

    if sub_stats == SUB_STATS[3]:
        st.markdown("**🧮 Auditoría diaria de stock (muestreo aleatorio)**")
        st.caption(
            "Todas las mañanas se puede generar una lista corta de productos al azar (priorizando favoritos "
            "y los que tienen precio cargado) para contarlos a mano en 5 minutos y detectar descalces antes de que se acumulen."
        )
        cant_auditoria = st.number_input("Cantidad de productos a auditar hoy:", min_value=3, max_value=20, value=8, step=1)
        if st.button("🎲 Generar auditoría de hoy"):
            generada = generar_auditoria_hoy(cant_auditoria)
            if generada:
                avisar("success", "Auditoría de hoy generada.")
                st.rerun()
            else:
                st.info("Ya había una auditoría generada para hoy (ver abajo).")

        auditoria_hoy = listar_auditoria_hoy()
        if auditoria_hoy:
            for item in auditoria_hoy:
                colA1, colA2, colA3 = st.columns([3, 1.5, 1])
                colA1.write(f"**{item['Codigo']}** ({item['Marca']}) — sistema: {item['Stock sistema']}")
                if item["Resuelto"]:
                    signo = "✅ OK" if item["Diferencia"] == 0 else f"⚠️ Diferencia: {item['Diferencia']:+d}"
                    colA2.write(f"Contado: {item['Stock contado']} — {signo}")
                else:
                    contado = colA2.number_input("Contado", min_value=0, step=1, key=f"conteo_{item['ID_auditoria']}",
                                                  label_visibility="collapsed")
                    if colA3.button("💾", key=f"guardar_conteo_{item['ID_auditoria']}"):
                        registrar_conteo_auditoria(item["ID_auditoria"], contado)
                        st.rerun()
        else:
            st.caption("Todavía no generaste la auditoría de hoy.")

        st.markdown("---")
        st.markdown("**📦 Matriz ABC — ubicación sugerida en depósito**")
        st.caption(
            "Como la app no tiene un módulo de ventas, la rotación se aproxima con la cantidad de veces que "
            "se buscó cada código. Los más buscados (A) conviene tenerlos más a mano."
        )
        matriz = calcular_matriz_abc()
        if matriz:
            st.dataframe(quitar_id(matriz), use_container_width=True, hide_index=True)
            st.caption("Para cargar o corregir la ubicación de un producto, andá a la pestaña 'Administrar'.")
        else:
            st.caption("Todavía no hay suficientes búsquedas registradas para armar la matriz.")

    if sub_stats == SUB_STATS[4]:
        st.markdown("**🔎 Códigos buscados sin resultado**")
        st.caption("Qué te están pidiendo los clientes que todavía no tenés cargado.")
        fallidas = listar_busquedas_sin_resultado()
        if fallidas:
            st.dataframe(fallidas, use_container_width=True, hide_index=True)
        else:
            st.caption("Sin registros todavía.")

    if sub_stats == SUB_STATS[5]:
        st.markdown("**⏳ Lo que se va a acabar**")
        explicar(
            "Calculado con el ritmo real de venta de cada producto y el stock que queda.",
            "La reposición hasta ahora era 100% manual: alguien tenía que acordarse de tocar "
            "«Pedir». Y de lo que uno no se acuerda es justamente de lo que se vende parejo "
            "todos los días — el filtro común que nadie mira hasta que un cliente lo pide y no "
            "está.\n\nSe ignora lo que se vendió una o dos veces: con eso no se puede calcular "
            "un ritmo, es ruido."
        )
        cq1, cq2 = cols(2)
        st.session_state.setdefault("dias_aviso_quiebre", 21)
        dias_aviso = cq1.select_slider("Avisar cuando queden menos de:",
                                        options=[7, 14, 21, 30, 60],
                                        format_func=lambda x: f"{x} días",
                                        key="dias_aviso_quiebre")
        st.session_state.setdefault("min_ventas_quiebre", 3)
        min_ventas = cq2.select_slider("Contar solo lo vendido al menos:",
                                        options=[3, 5, 10, 20],
                                        format_func=lambda x: f"{x} veces",
                                        key="min_ventas_quiebre")
        try:
            por_quebrar = productos_por_quebrar(int(dias_aviso), int(min_ventas))
        except sqlite3.OperationalError:
            por_quebrar = []
        if not por_quebrar:
            st.success("✅ Nada a punto de quebrarse, según lo que se vendió estos meses.")
        else:
            sin_stock = [x for x in por_quebrar if x["Stock"] <= 0]
            if sin_stock:
                st.error(f"🔴 {len(sin_stock)} producto(s) que se venden seguido están **en cero**.")
            st.dataframe(quitar_id(por_quebrar), use_container_width=True, hide_index=True)
            etiquetas_q = {f"{x['Código']} ({x['Marca']}) — {x['Se acaba en']}": x["_id"]
                           for x in por_quebrar}
            elegidos_q = st.multiselect("Marcar para pedir:", list(etiquetas_q.keys()),
                                         key="quiebre_a_pedir")
            if elegidos_q and st.button(f"📌 Marcar {len(elegidos_q)} para reposición"):
                for e in elegidos_q:
                    solicitar_reposicion(etiquetas_q[e])
                avisar("success", f"{len(elegidos_q)} producto(s) marcados para pedir.")
                st.rerun()

        st.markdown("---")
        st.markdown("**📈 Cuánto te aumentó cada proveedor**")
        explicar(
            "Medido sobre tu propio historial de precios, no sobre lo que dicen las listas.",
            "La app viene guardando cada cambio de precio y solo lo mostraba producto por "
            "producto. Pero la pregunta que importa no es cuánto aumentó un filtro, sino cuánto "
            "aumentó el proveedor: con eso se decide a quién comprarle y con quién hay que "
            "hablar.\n\nSe usa la **mediana**, no el promedio: un solo precio mal cargado —de "
            "esos que llegan con el separador de decimales al revés— alcanza para inflar un "
            "promedio y hacer parecer que un proveedor aumentó 400%."
        )
        st.session_state.setdefault("meses_variacion", 6)
        meses_var = st.select_slider("Comparar contra los precios de hace:",
                                      options=[3, 6, 12, 24],
                                      format_func=lambda x: f"{x} meses",
                                      key="meses_variacion")
        try:
            variacion = variacion_de_precios_por_marca(int(meses_var))
        except sqlite3.OperationalError:
            variacion = []
        if not variacion:
            st.caption(
                "Hace falta haber importado precios de una misma marca en dos momentos "
                "distintos para poder comparar. Aparece solo a medida que vas cargando listas."
            )
        else:
            st.dataframe(quitar_id(variacion), use_container_width=True, hide_index=True)
            st.caption("Ordenado de mayor a menor aumento.")

        st.markdown("---")
        st.markdown("**💰 A quién conviene comprarle**")
        explicar(
            "Comparando solo repuestos que son equivalentes entre sí.",
            "Comparar el precio promedio de dos marcas no dice nada: una puede vender frenos "
            "caros y la otra filtros baratos. Acá cada comparación es entre dos códigos que "
            "hacen el mismo trabajo, y solo se usan los vínculos con confianza razonable — si "
            "la equivalencia es dudosa, la comparación de precios también lo es."
        )
        try:
            conviene = quien_conviene_por_rubro()
        except sqlite3.OperationalError:
            conviene = []
        if not conviene:
            st.caption(
                "Hace falta tener el mismo repuesto de dos marcas distintas, vinculados entre sí "
                "y con precio cargado en los dos."
            )
        else:
            st.dataframe(quitar_id(conviene), use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**🧊 Clavos: lo que no se mueve**")
        explicar(
            "Stock que hace meses no se vende. Es plata dormida en el estante.",
            "Se cruza el stock con la última venta de cada producto. Ordenado por **plata "
            "parada** —precio por cantidad—, no por cuántos días lleva: 40 unidades de algo "
            "barato molestan menos que 2 de algo caro.\n\nLo que nunca se vendió cuenta solo "
            "si hace rato que está cargado: un producto que entró la semana pasada todavía no "
            "tuvo su chance."
        )
        st.session_state.setdefault("dias_clavo", 180)
        _dias_clavo = st.select_slider("Sin vender desde hace más de:",
                                        options=[90, 180, 365, 730],
                                        format_func=lambda x: f"{x} días" if x < 365
                                                              else f"{x // 365} año(s)",
                                        key="dias_clavo")
        try:
            _clavos = productos_estancados(int(_dias_clavo))
        except Exception:
            _clavos = []
        if not _clavos:
            st.success("✅ Nada estancado con ese criterio.")
        else:
            _plata = sum(x["Plata parada"] or 0 for x in _clavos)
            st.warning(f"⚠️ {len(_clavos)} producto(s) con **${_plata:,.0f}** inmovilizados.")
            st.dataframe(quitar_id([{k: v for k, v in x.items() if not k.startswith("_")}
                                     for x in _clavos]),
                          use_container_width=True, hide_index=True)
            st.caption(
                "Sirve para decidir una liquidación, o para priorizarlos cuando alguien pide "
                "un equivalente y tenés varios que sirven igual."
            )

        st.markdown("---")
        st.markdown("**🔄 Códigos reemplazados por el fabricante**")
        explicar(
            "Cuando un código deja de fabricarse y lo reemplaza otro.",
            "No es una equivalencia común: tiene dirección. El viejo se discontinúa y el nuevo "
            "lo reemplaza, pero no al revés.\n\nCargarlo evita la forma más tonta de perder "
            "una venta: alguien busca el código viejo, no aparece, y se le dice al cliente que "
            "no se fabrica más — cuando en realidad existe con otro número.\n\nSigue la cadena "
            "completa: si A fue reemplazado por B y B por C, buscando A te lleva hasta C."
        )
        cr1, cr2 = cols(2)
        _viejo = cr1.text_input("Código viejo:", key="reemp_viejo",
                                 placeholder="El que ya no se fabrica").strip()
        _nuevo = cr2.text_input("Lo reemplaza:", key="reemp_nuevo",
                                 placeholder="El código vigente").strip()
        _nota_r = st.text_input("Nota (opcional):", key="reemp_nota",
                                 placeholder="Ej: cambia el largo de rosca, revisar antes")
        if st.button("💾 Guardar el reemplazo", disabled=not (_viejo and _nuevo)):
            _ok_r, _msg_r = guardar_reemplazo(_viejo, _nuevo, "", _nota_r)
            if _ok_r:
                avisar("success", _msg_r)
                st.rerun()
            else:
                st.error(_msg_r)
        try:
            c.execute("""SELECT codigo_viejo AS "Ya no se fabrica",
                                codigo_nuevo AS "Lo reemplaza", nota AS "Nota",
                                substr(fecha, 1, 10) AS "Cargado"
                         FROM reemplazos_codigo ORDER BY fecha DESC LIMIT 200""")
            _lista_r = filas_a_listas(c)
        except sqlite3.OperationalError:
            _lista_r = []
        if _lista_r:
            st.dataframe(_lista_r, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**🚫 Puede que ya no se fabriquen**")
        explicar(
            "Códigos que faltaron en las últimas listas del proveedor.",
            "Cuando llega una lista nueva, los productos que vienen se actualizan y los que no "
            "vienen quedan intactos. Si un código faltó en las últimas listas, casi seguro el "
            "proveedor lo discontinuó o le cambió el número — pero en la base sigue figurando "
            "con su precio viejo.\n\nEso cuesta de dos formas: se cotiza algo que ya no existe "
            "y el cliente se va cuando no llega, y el stock muerto ocupa lugar sin que nadie lo "
            "note.\n\n**No son certezas.** Un producto puede faltar en una lista simplemente "
            "porque el proveedor lo mandó aparte. Son candidatos para que los mires."
        )
        st.session_state.setdefault("listas_discont", 2)
        listas_disc = st.select_slider(
            "Considerar discontinuado si faltó en las últimas:",
            options=[2, 3, 4, 5], format_func=lambda x: f"{x} listas",
            key="listas_discont",
            help="Con 2 aparecen más candidatos y más falsos; con 4 o 5, solo los que "
                 "vienen faltando hace rato."
        )
        try:
            discont, revisados_disc = productos_probablemente_discontinuados(
                listas_seguidas=int(listas_disc))
        except sqlite3.OperationalError:
            discont, revisados_disc = [], 0

        if not revisados_disc:
            st.caption(
                "Hace falta haber importado al menos 2 listas de una misma marca para poder "
                "comparar. Con una sola no hay con qué."
            )
        elif not discont:
            st.success(f"✅ De {revisados_disc:,} producto(s) revisados, ninguno faltó en las "
                        f"últimas {listas_disc} listas.")
        else:
            con_stock = [x for x in discont if (x["_stock"] or 0) > 0]
            if con_stock:
                st.warning(
                    f"⚠️ {len(con_stock)} de estos **tienen stock**. Si el proveedor ya no los "
                    "manda, eso es mercadería que conviene liquidar antes de que quede muerta."
                )
            st.dataframe(quitar_id([{k: v for k, v in x.items() if not k.startswith("_")}
                                     for x in discont]),
                          use_container_width=True, hide_index=True)
            st.download_button(
                "⬇️ Bajar la lista en Excel",
                data=to_excel_bytes([{k: v for k, v in x.items() if not k.startswith("_")}
                                      for x in discont]),
                file_name="posibles_discontinuados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.caption(
                "Antes de dar de baja alguno, confirmalo con el proveedor: puede haberle "
                "cambiado el número y seguir existiendo con otro código."
            )

        st.markdown("---")
        st.markdown("**🔒 Stock apartado en presupuestos**")
        explicar(
            "Lo que está reservado y todavía no se vendió.",
            "Con varios atendiendo a la vez, vender dos veces la misma pieza es cuestión de "
            "tiempo: uno cotiza 4 pastillas, el otro ve stock 4 y las vende.\n\nEl buscador "
            "muestra el stock **libre** cuando hay algo apartado, así nadie promete lo que ya "
            f"está comprometido. Las reservas de más de {DIAS_VENCIMIENTO_RESERVA} días se "
            "liberan solas: un presupuesto que nadie contestó no puede seguir bloqueando "
            "mercadería para siempre."
        )
        _vencidas = vencer_reservas_viejas()
        if _vencidas:
            st.caption(f"🔓 Se liberaron {_vencidas} reserva(s) que pasaron los "
                        f"{DIAS_VENCIMIENTO_RESERVA} días.")
        _reservas = reservas_activas()
        if not _reservas:
            st.caption("No hay nada apartado ahora mismo.")
        else:
            st.dataframe(quitar_id(_reservas), use_container_width=True, hide_index=True)
            _etq = {f"{r['Código']} ({r['Marca']}) — {r['Apartadas']} u. "
                    f"para {r['Cliente'] or 'sin nombre'}": r["_id"] for r in _reservas}
            _elegida = st.selectbox("Cerrar una reserva:", list(_etq.keys()), key="reserva_cerrar")
            rc1, rc2 = cols(2)
            if rc1.button("✅ Se vendió (descontar del stock)"):
                cerrar_reserva(_etq[_elegida], vendida=True)
                avisar("success", "Reserva cerrada y stock descontado.")
                st.rerun()
            if rc2.button("🔓 Se cayó (liberar sin descontar)"):
                cerrar_reserva(_etq[_elegida], vendida=False)
                avisar("success", "Reserva liberada.")
                st.rerun()

        st.markdown("---")
        st.markdown("**🙋 Pedidos marcados por empleados**")
        st.caption(
            "Cuando alguien busca algo y toca '📌 Pedir' en el buscador, aparece acá para que decidas "
            "qué comprarle a cada proveedor."
        )
        pedidos = listar_pedidos_reposicion("pendiente")
        seleccionados = []
        if pedidos:
            for p in pedidos:
                colp1, colp2, colp3 = st.columns([4, 1, 1])
                stock_txt = p["Stock actual"] if p["Stock actual"] is not None else "s/d"
                marcado = colp1.checkbox(
                    f"{p['Marca']} - {p['Codigo']} — {p['Descripcion'] or ''} "
                    f"(stock: {stock_txt}, pedido {p['Veces pedido']}x, último: {p['Último en pedirlo']})",
                    key=f"chk_pedido_{p['ID']}"
                )
                if marcado:
                    seleccionados.append(p)
                # on_click en vez de "if boton: accion + st.rerun()": el callback corre ANTES de
                # que Streamlit refresque la página, así la lista ya sale actualizada sin tener que
                # forzar un st.rerun() — que es lo que hacía perder la pestaña y volver al inicio.
                colp2.button("✅", key=f"resuelto_{p['ID']}", help="Marcar como resuelto",
                              on_click=marcar_pedido_resuelto, args=(p["ID"],))
                colp3.button("🗑️", key=f"descartar_{p['ID']}", help="Descartar (no hace falta pedirlo)",
                              on_click=descartar_pedido_reposicion, args=(p["ID"],))
        else:
            st.caption("Ningún empleado marcó nada para pedir todavía.")

        st.markdown("---")
        st.markdown("**📦 Favoritos con poco stock**")
        umbral_stock = st.number_input("Alertar cuando el stock sea menor o igual a:", min_value=0, value=2, step=1,
                                        key="umbral_para_pedir")
        stock_bajo = listar_favoritos_stock_bajo(umbral_stock)
        if stock_bajo:
            for f in stock_bajo:
                stock_txt_f = f["Stock"] if f["Stock"] is not None else "s/d"
                marcado_f = st.checkbox(
                    f"{f['Marca']} - {f['Codigo']} — {f['Descripcion'] or ''} (stock: {stock_txt_f})",
                    key=f"chk_stockbajo_{f['ID']}"
                )
                if marcado_f:
                    seleccionados.append(f)
        else:
            st.caption("Ningún favorito con stock bajo por ahora.")

        if seleccionados:
            st.markdown("---")
            st.markdown(f"**📲 Armar mensaje para el proveedor ({len(seleccionados)} ítem(s) elegidos)**")
            por_marca = {}
            for item in seleccionados:
                por_marca.setdefault(item["Marca"], []).append(item)
            for marca, items in por_marca.items():
                lineas_msg = [f"Hola! Necesito reponer estos productos de {marca}:"]
                for it in items:
                    stock_it = it.get("Stock actual", it.get("Stock"))
                    lineas_msg.append(
                        f"- {it['Codigo']} ({it.get('Descripcion') or ''}) — "
                        f"quedan {stock_it if stock_it is not None else 's/d'}"
                    )
                mensaje_reposicion = "\n".join(lineas_msg)
                with st.expander(f"📨 {marca} ({len(items)} ítem(s))"):
                    st.text_area("Mensaje:", value=mensaje_reposicion, height=120, key=f"msg_repo_{marca}")
                    url_wa_repo = "https://wa.me/?text=" + quote(mensaje_reposicion)
                    st.link_button(f"📲 Abrir WhatsApp para {marca}", url_wa_repo, key=f"wa_repo_{marca}")

    if sub_stats == SUB_STATS[6]:
        st.markdown("**🔍 Revisar las equivalencias que ya están cargadas**")
        explicar(
            "Pasa las mismas alarmas por todo lo que se cargó antes (listas viejas, vínculos hechos "
            "a mano).",
            "Lo que marques como correcto no vuelve a aparecer; lo que borres queda descartado para "
            "siempre, aunque vuelvas a importar la misma lista."
        )
        if st.button("🔍 Auditar lo ya cargado"):
            with st.spinner("Revisando..."):
                st.session_state["resultado_auditoria"] = auditar_equivalencias_existentes()

        resultado_aud = st.session_state.get("resultado_auditoria")
        if resultado_aud:
            ma1, ma2, ma3, ma4 = st.columns(4)
            ma1.metric("Vínculos revisados", resultado_aud["total_revisados"])
            ma2.metric("Códigos raros", len(resultado_aud.get("codigos_malos", [])))
            ma3.metric("Conflictos", len(resultado_aud["conflictos"]))
            ma4.metric("Medidas que no dan", len(resultado_aud["por_medidas"]))

            if resultado_aud.get("quedo_corta"):
                st.warning(
                    f"⚠️ **La revisión quedó corta.** Se miraron {resultado_aud['total_revisados']:,} "
                    f"de {resultado_aud['total_en_base']:,} vínculos que hay cargados. Resolvé estos "
                    "y volvé a auditar para seguir con el resto — todavía puede haber problemas sin ver."
                )

            # 1) Códigos que no parecen códigos: lo que deja una importación mal mapeada.
            #    Va primero porque un solo producto basura ensucia decenas de vínculos.
            if resultado_aud.get("codigos_malos"):
                st.markdown("**🚫 Códigos que no parecen códigos de repuesto**")
                st.caption(
                    "Suelen venir de una importación donde la columna del código en realidad tenía "
                    "medidas, cantidades o pedazos de la descripción. Cortarles los vínculos limpia "
                    "el problema; el producto queda por si lo querés corregir a mano."
                )
                for cm in resultado_aud["codigos_malos"][:25]:
                    cc1, cc2 = st.columns([3, 1])
                    cc1.markdown(f"**{cm['marca']}** · `{cm['codigo']}` — {cm['motivo']}  \n"
                                  f"<small>{cm['vinculos']} vínculo(s)</small>", unsafe_allow_html=True)
                    cc2.button("✂️ Cortar sus vínculos", key=f"cortar_malo_{cm['id']}",
                                on_click=cb_auditoria_cortar_todos, args=(cm["id"],))
                if len(resultado_aud["codigos_malos"]) > 25:
                    st.caption(f"(mostrando 25 de {len(resultado_aud['codigos_malos'])})")
                st.markdown("---")

            # 2) Productos basura: lo primero a resolver, porque un solo producto mal cargado
            #    puede estar ensuciando cientos de códigos a la vez.
            if resultado_aud["productos_sospechosos"]:
                st.markdown("**🚩 Productos con muchísimos vínculos (revisalos primero)**")
                explicar(
                    "Una pieza real rara vez equivale a más de 10 códigos de fábrica.",
                    "Si un producto tiene decenas, casi siempre es basura de una importación mal mapeada — "
                    "por ejemplo un código '1' que quedó de una columna equivocada. Cortarle los vínculos "
                    "de una limpia el problema entero."
                )
                # Sin desplegables: Streamlit los cierra en cada refresco, y como cada botón
                # provoca uno, se cerraba la ventana justo cuando estabas revisando.
                for p in resultado_aud["productos_sospechosos"][:15]:
                    desc = p["descripcion"] or "_(sin descripción)_"
                    ps1, ps2 = st.columns([3, 1])
                    ps1.markdown(f"**{p['marca']}** · `{p['codigo']}` — {desc}  \n"
                                  f"<small>vinculado a {p['cantidad']} códigos distintos</small>",
                                  unsafe_allow_html=True)
                    ps2.button(f"✂️ Cortar {p['cantidad']}",
                                key=f"cortar_todo_{p['id']}", type="primary",
                                on_click=cb_auditoria_cortar_todos, args=(p["id"],),
                                help="El producto queda; solo se cortan todas sus equivalencias")
                if len(resultado_aud["productos_sospechosos"]) > 15:
                    st.caption(f"(mostrando 15 de {len(resultado_aud['productos_sospechosos'])})")
                st.markdown("---")

            # 2) Conflictos agrupados: un código de fábrica apuntando a varios productos
            if resultado_aud["conflictos"]:
                st.markdown("**⚠️ Un código de fábrica apuntando a varios productos del mismo proveedor**")
                st.caption(
                    "Acá se ven juntos todos los productos a los que apunta cada código, para poder "
                    "comparar y cortar el que sobra. Normalmente uno tiene descripción real y el otro "
                    "es el que quedó mal."
                )
                total_conf = len(resultado_aud["conflictos"])
                por_pag_conf = 8
                pags_conf = (total_conf - 1) // por_pag_conf + 1
                if pags_conf > 1:
                    pag_conf = st.number_input(
                        f"Página (de {pags_conf}) — {por_pag_conf} conflictos por página:",
                        min_value=1, max_value=pags_conf, value=1, step=1, key="pagina_conflictos"
                    )
                else:
                    pag_conf = 1
                desde_conf = (int(pag_conf) - 1) * por_pag_conf
                for g in resultado_aud["conflictos"][desde_conf:desde_conf + por_pag_conf]:
                    st.markdown(f"**⚠️ {g['codigo_oem']} → {len(g['productos'])} productos "
                                 f"de {g['marca_proveedor']}**")
                    if g["descripcion_oem"]:
                        st.caption(g["descripcion_oem"])
                    for p in g["productos"]:
                        desc = p["descripcion"] or "⚠️ _(sin descripción — sospechoso)_"
                        marca_ok = " · ya revisado" if p["revisado_ok"] else ""
                        cg1, cg2, cg3 = st.columns([3, 1, 1])
                        cg1.markdown(f"**`{p['codigo']}`** — {desc}  \n"
                                      f"<small>{p['vinculos_totales']} vínculos en total{marca_ok}</small>",
                                      unsafe_allow_html=True)
                        cg2.button("🗑️ Cortar", key=f"cortar_par_{g['codigo_oem']}_{p['id']}",
                                    on_click=cb_auditoria_eliminar, args=(p["par"][0], p["par"][1]),
                                    help="Corta solo este vínculo")
                        cg3.button("✅ Dejar", key=f"dejar_par_{g['codigo_oem']}_{p['id']}",
                                    on_click=cb_auditoria_dejar, args=([p["par"]],),
                                    help="Es correcto; no volver a marcarlo")
                    st.markdown("")
                st.markdown("---")

            # 3) Medidas contradictorias
            if resultado_aud["por_medidas"]:
                st.markdown("**📐 Vínculos donde las medidas cargadas no coinciden**")
                for m in resultado_aud["por_medidas"][:30]:
                    st.markdown(f"**{m['marca_a']} `{m['cod_a']}`** — {m['desc_a'] or '_(sin descripción)_'}  \n"
                                 f"**{m['marca_b']} `{m['cod_b']}`** — {m['desc_b'] or '_(sin descripción)_'}")
                    st.caption(f"📐 {m['detalle']}")
                    mb1, mb2 = st.columns(2)
                    mb1.button("✅ Está bien, dejalo", key=f"med_ok_{m['a']}_{m['b']}",
                                on_click=cb_auditoria_dejar, args=([(m["a"], m["b"])],))
                    mb2.button("🗑️ Borrar el vínculo", key=f"med_del_{m['a']}_{m['b']}",
                                on_click=cb_auditoria_eliminar, args=(m["a"], m["b"]))
                if len(resultado_aud["por_medidas"]) > 30:
                    st.caption(f"(mostrando 30 de {len(resultado_aud['por_medidas'])})")

            if (not resultado_aud["conflictos"] and not resultado_aud["por_medidas"]
                    and not resultado_aud["productos_sospechosos"]
                    and not resultado_aud.get("codigos_malos")):
                st.success("✅ No se encontró nada sospechoso entre los vínculos ya cargados.")

        st.markdown("---")

        lotes_pendientes = resumen_lotes_pendientes()
        if lotes_pendientes:
            total_pendientes = sum(l["cantidad"] for l in lotes_pendientes)
            st.markdown("**📄 Vínculos de listas de proveedor esperando revisión**")
            explicar(
                "Estos llegaron al importar una lista y todavía NO están cargados.",
                "Se separan solos entre los que no tienen nada raro y los que dispararon alguna alarma, "
                "para que apruebes en bloque los limpios y mires con lupa solo los pocos sospechosos."
            )

            with st.expander(f"🧹 Descartar TODO lo pendiente ({total_pendientes:,} vínculos de "
                              f"{len(lotes_pendientes)} lista(s))"):
                st.warning(
                    "Borra de una todos los vínculos que están esperando revisión, de todas las listas. "
                    "**Los productos y precios no se tocan** — solo se descartan las equivalencias "
                    "pendientes. Es lo que conviene cuando una importación quedó mal mapeada: descartás "
                    "todo y volvés a importar con las columnas correctas."
                )
                confirmar_todo = st.checkbox("Confirmo que quiero descartar todo lo pendiente",
                                              key="confirmar_descartar_todo")
                if st.button("🧹 Descartar todo", disabled=not confirmar_todo, type="primary"):
                    borrados = 0
                    for l in lotes_pendientes:
                        borrados += rechazar_pendientes(l["lote"], None)
                    avisar("success", f"Se descartaron {borrados:,} vínculo(s) pendientes.")
                    st.rerun()

            # Una lista por vez, elegida con un selector. Antes cada lista estaba dentro de un
            # desplegable, y como Streamlit los cierra al refrescar la pantalla, cada botón que
            # tocabas te cerraba la ventana entera — imposible revisar 300 vínculos así.
            opciones_lotes = {f"{l['lote']} — {l['cantidad']} vínculo(s)": l for l in lotes_pendientes}
            etiqueta_lote = st.selectbox("Lista a revisar:", list(opciones_lotes.keys()),
                                          key="lote_en_revision")
            lote_info = opciones_lotes[etiqueta_lote]

            total_lote = contar_pendientes_del_lote(lote_info["lote"])
            ca1, ca2 = st.columns([2, 1])
            with ca1:
                st.session_state.setdefault("cuantos_pendientes", 1000)
                cuantos = st.select_slider(
                    "¿Cuántos analizar por vez?",
                    options=[400, 1000, 2500, 5000, 10000],
                    key="cuantos_pendientes",
                    help="Analizar más tarda un poco más, pero te evita repetir la vuelta muchas veces."
                )
            paginas_lote = max((total_lote - 1) // int(cuantos) + 1, 1)
            with ca2:
                tanda_lote = st.number_input(
                    f"Tanda (de {paginas_lote}):", min_value=1, max_value=paginas_lote,
                    value=1, step=1, key=f"tanda_lote_{lote_info['lote']}"
                ) if paginas_lote > 1 else 1

            with st.spinner("Analizando..."):
                limpias, sospechosas = analizar_lote_pendiente(
                    lote_info["lote"], limite=int(cuantos), desde=(int(tanda_lote) - 1) * int(cuantos)
                )
            analizados = len(limpias) + len(sospechosas)
            st.caption(f"Analizados {analizados:,} de {total_lote:,} vínculo(s) de esta lista." +
                       (f" Quedan {total_lote - analizados:,} — cambiá de tanda para verlos."
                        if total_lote > analizados else ""))

            # Se agrupa por confianza, no por "tiene alarma / no tiene". Con 397 alarmas planas
            # había que mirarlas de a una; así se ve de una que la mayoría es descartable y solo
            # un puñado merece atención.
            patrones_ui = aprender_de_las_decisiones()
            todos_evaluados = limpias + sospechosas
            por_nivel = {"🟢": [], "🟡": [], "🟠": [], "🔴": []}
            for x in todos_evaluados:
                por_nivel[nivel_de_confianza(x.get("confianza", 50))[0][:1]].append(x)

            mn1, mn2, mn3, mn4 = st.columns(4)
            mn1.metric("🟢 Muy probables", len(por_nivel["🟢"]))
            mn2.metric("🟡 Probables", len(por_nivel["🟡"]))
            mn3.metric("🟠 Dudosas", len(por_nivel["🟠"]))
            mn4.metric("🔴 Casi seguro mal", len(por_nivel["🔴"]))

            # Cuando el problema es UN producto que aparece en decenas de pendientes, se resuelve
            # de una. Antes había que aprobar o descartar cada vínculo por separado, aunque los
            # cincuenta dijeran exactamente lo mismo.
            culpables = productos_que_mas_ensucian(lote_info["lote"])
            if culpables:
                total_culpa = sum(x["Pendientes que genera"] for x in culpables)
                st.warning(
                    f"🎯 **{len(culpables)} producto(s) con código dudoso generan {total_culpa} "
                    "de estos pendientes.** Resolvelos de una en vez de vínculo por vínculo."
                )
                st.dataframe(quitar_id(culpables), use_container_width=True, hide_index=True)
                etiquetas_culpa = {
                    f"{x['Código']} ({x['Marca']}) — {x['Pendientes que genera']} pendientes": x["_id"]
                    for x in culpables
                }
                elegido_culpa = st.selectbox("¿Cuál resolver?", list(etiquetas_culpa.keys()),
                                              key=f"culpable_{lote_info['lote']}")
                cc1, cc2 = st.columns(2)
                if cc1.button("🚫 Descartar TODOS sus pendientes", key=f"desc_culpa_{lote_info['lote']}"):
                    n = rechazar_pendientes_de_producto(etiquetas_culpa[elegido_culpa],
                                                         lote_info["lote"])
                    invalidar_salud()
                    avisar("success", f"Se descartaron {n} vínculo(s) de una.")
                    st.rerun()
                if cc2.button("✅ El código está bien, no volver a marcarlo",
                               key=f"ok_culpa_{lote_info['lote']}"):
                    aprobar_puente(etiquetas_culpa[elegido_culpa], "código validado a mano")
                    invalidar_salud()
                    avisar("success", "Anotado: ese código deja de marcarse como problema.")
                    st.rerun()

            if por_nivel["🔴"]:
                muestra_mal = por_nivel["🔴"][:8]
                with st.expander(f"🔴 Ver por qué las {len(por_nivel['🔴'])} peores están mal"):
                    for x in muestra_mal:
                        st.markdown(f"**{x['cod_a']}** ({x['marca_a']}) ↔ "
                                     f"**{x['cod_b']}** ({x['marca_b']}) — {x['confianza']:.0f}/100")
                        for tipo, texto in x.get("senales", []):
                            st.caption(("❌ " if tipo == "mal" else "✅ ") + texto)

            ml1, ml2 = st.columns(2)
            ml1.metric("Sin nada raro", len(limpias))
            ml2.metric("Con alguna alarma", len(sospechosas))

            if analizados and len(sospechosas) / analizados > 0.7:
                st.error(
                    "🔴 Más del 70% de esta lista dispara alarmas. Eso no es que tengas mala suerte: "
                    "casi siempre significa que la importación quedó **mal mapeada** — la columna que "
                    "se tomó como código en realidad traía cantidades, medidas o pedazos de la "
                    "descripción. Antes de revisar de a uno, conviene descartar toda la lista y "
                    "volver a importarla revisando bien el mapeo de columnas."
                )

            bl1, bl2 = st.columns(2)
            pares_limpios = []
            for x in limpias:
                pares_limpios.extend([(x["a"], x["b"]), (x["b"], x["a"])])
            bl1.button(f"✅ Aprobar los {len(limpias)} sin alarmas",
                        key=f"apr_limpias_{lote_info['lote']}", type="primary",
                        disabled=not limpias,
                        on_click=aprobar_pendientes, args=(lote_info["lote"], pares_limpios))
            bl2.button("🚫 Descartar toda esta lista",
                        key=f"rec_lote_{lote_info['lote']}",
                        on_click=rechazar_pendientes, args=(lote_info["lote"], None),
                        help="Los productos y precios quedan; solo se descartan los vínculos")

            # Descartar en bloque los que el análisis ya dio por perdidos. Sin esto, la app
            # marcaba cientos como «casi seguro mal» y después te los hacía resolver de a uno,
            # que es exactamente lo que el análisis venía a evitar.
            if por_nivel["🔴"]:
                pares_rojos = []
                for x in por_nivel["🔴"]:
                    pares_rojos.extend([(x["a"], x["b"]), (x["b"], x["a"])])
                st.error(
                    f"🔴 Hay **{len(por_nivel['🔴'])} vínculos «casi seguro mal»** en esta tanda. "
                    "No hace falta mirarlos de a uno: el análisis ya tiene evidencia en contra de "
                    "todos (rubros distintos, medidas que no dan, o tu propio historial)."
                )
                if st.button(f"🚫 Descartar los {len(por_nivel['🔴'])} marcados 🔴",
                              key=f"rec_rojos_{lote_info['lote']}", type="primary"):
                    rechazar_pendientes(lote_info["lote"], pares_rojos)
                    invalidar_salud()
                    avisar("success", f"Se descartaron {len(por_nivel['🔴'])} vínculo(s) de una.")
                    st.rerun()

            # Y cuando tu propio historial es contundente sobre una combinación de marcas,
            # ofrecer resolver TODA esa combinación de una vez.
            combinaciones = {}
            for x in limpias + sospechosas:
                clave = tuple(sorted((x.get("marca_a", ""), x.get("marca_b", ""))))
                combinaciones.setdefault(clave, []).append(x)
            for (ma, mb), items in sorted(combinaciones.items(), key=lambda t: -len(t[1])):
                dato = patrones_ui["marcas"].get((ma, mb)) if patrones_ui else None
                if not dato or dato["tasa_ok"] > 0.05 or dato["decisiones"] < 50:
                    continue
                pares_comb = []
                for x in items:
                    pares_comb.extend([(x["a"], x["b"]), (x["b"], x["a"])])
                st.warning(
                    f"📚 **De {dato['decisiones']:,} vínculos {ma}↔{mb} que revisaste, descartaste "
                    f"el {(1-dato['tasa_ok'])*100:.0f}%.** En esta tanda hay {len(items)} más de "
                    "esa misma combinación. Si van a terminar igual, resolvelos de una."
                )
                if st.button(f"🚫 Descartar los {len(items)} de {ma}↔{mb}",
                              key=f"rec_comb_{lote_info['lote']}_{ma}_{mb}"):
                    rechazar_pendientes(lote_info["lote"], pares_comb)
                    invalidar_salud()
                    avisar("success", f"Se descartaron {len(items)} vínculo(s) de {ma}↔{mb}.")
                    st.rerun()
                break   # de a una combinación por vez, para no llenar la pantalla

            if limpias:
                with st.expander(f"Ver los {len(limpias)} sin alarmas"):
                    st.dataframe(
                        [{"Código A": x["cod_a"], "Marca A": x["marca_a"],
                           "Código B": x["cod_b"], "Marca B": x["marca_b"]} for x in limpias[:500]],
                        use_container_width=True, hide_index=True
                    )
                    if len(limpias) > 500:
                        st.caption(f"Se muestran 500 de {len(limpias)}; el botón de aprobar los toma a todos.")

            if sospechosas:
                st.markdown("---")
                st.warning(f"⚠️ {len(sospechosas)} vínculo(s) con algo raro — revisalos:")
                # De a tandas por pantalla: con cientos, la página se vuelve imposible de usar
                por_pagina = st.radio("Mostrar de a:", [10, 25, 50], horizontal=True,
                                       key="sosp_por_pagina")
                paginas = (len(sospechosas) - 1) // por_pagina + 1
                if paginas > 1:
                    pagina_sosp = st.number_input(
                        f"Página (de {paginas}) — cada una trae {por_pagina}:",
                        min_value=1, max_value=paginas, value=1, step=1, key="pagina_sospechosas"
                    )
                else:
                    pagina_sosp = 1
                desde = (int(pagina_sosp) - 1) * por_pagina
                pagina_actual = sospechosas[desde:desde + por_pagina]

                # Agrupados por MOTIVO, no uno debajo del otro. Cuando 40 vínculos fallan por lo
                # mismo —"«1S» es demasiado corto"— repetir la explicación 40 veces obliga a
                # scrollear y a decidir 40 veces algo que es una sola decisión. Agrupados, se lee
                # el motivo una vez y se resuelve el grupo entero.
                por_motivo = {}
                for s in pagina_actual:
                    clave = s["alarmas"][0] if s["alarmas"] else "Sin alarma puntual"
                    por_motivo.setdefault(clave, []).append(s)

                for motivo, items in sorted(por_motivo.items(), key=lambda x: -len(x[1])):
                    peor_grupo = min(x["confianza"] for x in items)
                    icono = "🔴" if peor_grupo < 30 else "🟠" if peor_grupo < 55 else "🟡"
                    st.markdown(f"{icono} **{motivo}** — {len(items)} vínculo(s)")

                    pares_grupo = []
                    for x in items:
                        pares_grupo.extend([(x["a"], x["b"]), (x["b"], x["a"])])
                    if len(items) > 1:
                        gb1, gb2 = st.columns(2)
                        gb1.button(f"✅ Los {len(items)} están bien",
                                    key=f"apr_grupo_{lote_info['lote']}_{abs(hash(motivo))}",
                                    on_click=aprobar_pendientes,
                                    args=(lote_info["lote"], pares_grupo))
                        gb2.button(f"🚫 Descartar los {len(items)}",
                                    key=f"rec_grupo_{lote_info['lote']}_{abs(hash(motivo))}",
                                    on_click=rechazar_pendientes,
                                    args=(lote_info["lote"], pares_grupo))

                    with st.expander(f"Ver los {len(items)} uno por uno"):
                        for s in items:
                            st.markdown(f"**{s['marca_a']} {s['cod_a']} ↔ "
                                         f"{s['marca_b']} {s['cod_b']}** · {s['confianza']:.0f}/100")
                            # Solo las alarmas ADICIONALES: la del título ya se leyó arriba
                            for alarma in s["alarmas"][1:]:
                                st.caption(f"   {alarma}")
                            sb1, sb2 = st.columns(2)
                            sb1.button("✅ Igual es correcto", key=f"apr_sosp_{s['a']}_{s['b']}",
                                        on_click=aprobar_pendientes,
                                        args=(lote_info["lote"], [(s["a"], s["b"]), (s["b"], s["a"])]))
                            sb2.button("🚫 Descartar", key=f"rec_sosp_{s['a']}_{s['b']}",
                                        on_click=rechazar_pendientes,
                                        args=(lote_info["lote"], [(s["a"], s["b"]), (s["b"], s["a"])]))
                            st.markdown("")
                    st.markdown("")
            st.markdown("---")

        st.markdown("**🛒 Equivalencias que aparecieron solas en el mostrador**")
        explicar(
            "Cuando un cliente pide un código y termina llevándose otro, eso es una equivalencia "
            "que pasó de verdad.",
            "**Nada se carga solo, nunca**: acá se juntan las sugerencias con el detalle de en qué "
            "se basa cada una, y una persona decide. Están ordenadas por cuánto respaldo tienen — "
            "mirá primero las verdes, y desconfiá de las que tengan evidencia en contra."
        )

        cfg1, cfg2 = cols(2)
        min_veces_cfg = cfg1.number_input("Mostrar cuando se repitió al menos:", min_value=1, value=2, step=1,
                                           key="cfg_min_veces",
                                           help="Con 1 vas a ver más sugerencias, pero también más ruido.")
        dias_cfg = cfg2.number_input("Mirar los últimos (días):", min_value=7, value=180, step=30,
                                      key="cfg_dias_equiv")

        candidatas = descubrir_equivalencias_candidatas(int(min_veces_cfg), int(dias_cfg))

        if not candidatas:
            c.execute("SELECT COUNT(*) FROM ventas_registradas")
            ventas_totales = c.fetchone()[0]
            if ventas_totales == 0:
                st.info(
                    "Todavía no hay ventas marcadas. Cuando busques algo en el Buscador y el cliente "
                    "se lleve una pieza, tocá '🛒 Se llevó' — con eso se empieza a alimentar esto."
                )
            else:
                st.caption(
                    f"Hay {ventas_totales} venta(s) registrada(s), pero todavía no se repitió ningún "
                    "caso lo suficiente como para sugerirlo (o ya están todos cargados como equivalentes)."
                )
        else:
            st.success(f"{len(candidatas)} sugerencia(s) para revisar:")
            marcas_disponibles = [m["nombre"] for m in
                                   c.execute("SELECT nombre FROM marcas ORDER BY nombre").fetchall()]

            for cand in candidatas:
                etiqueta_origen = ("marcado en el mostrador" if cand["origen"] == "mostrador"
                                    else "deducida: se vendió justo después de buscar eso sin resultado")
                with st.expander(
                    f"{cand['confianza']}  ·  {cand['codigo_pedido']} → {cand['codigo_vendido']} "
                    f"({cand['marca_vendida']}) — pasó {cand['veces']} vez/veces"
                ):
                    st.write(f"**Pidieron:** {cand['codigo_pedido']}")
                    st.write(f"**Se llevaron:** {cand['codigo_vendido']} ({cand['marca_vendida']}) "
                              f"{cand['descripcion']}")

                    st.markdown("**En qué se basa esto:**")
                    nombres_evidencia = {
                        "catalogo_oficial": "🌐 Aparece en la ficha oficial del proveedor",
                        "catalogo_no_lo_lista": "🌐 NO aparece en la ficha oficial",
                        "lista_proveedor": "📄 Venían relacionados en una lista del proveedor",
                        "medidas": "📐 Las medidas mecánicas coinciden",
                        "medidas_no_coinciden": "📐 Las medidas NO coinciden",
                        "mostrador": "🛒 Se repitió en el mostrador",
                    }
                    for ev in cand["evidencias"]:
                        st.caption(f"- {nombres_evidencia.get(ev['tipo'], ev['tipo'])}: {ev['detalle']}")
                    st.caption(f"({etiqueta_origen})")

                    # Verificación contra el catálogo del proveedor: la evidencia más fuerte,
                    # porque no depende de que nadie opine — el código está escrito en la
                    # página del proveedor o no está.
                    c.execute("""SELECT m.url_ficha_template, p.codigo_raw
                                 FROM productos p JOIN marcas m ON m.id = p.marca_id
                                 WHERE p.id = ?""", (cand["producto_id"],))
                    info_ficha = c.fetchone()
                    tiene_catalogo = bool(info_ficha and info_ficha["url_ficha_template"])
                    if st.button("🌐 Verificar en el catálogo del proveedor",
                                  key=f"verif_{cand['codigo_clean']}_{cand['producto_id']}",
                                  disabled=not tiene_catalogo,
                                  help=("Abre la ficha oficial y fija si el código pedido aparece ahí"
                                        if tiene_catalogo else
                                        f"La marca {cand['marca_vendida']} no tiene cargada la dirección "
                                        "de su catálogo (se carga en Administrar → Marcas)")):
                        url_ficha = info_ficha["url_ficha_template"].replace(
                            "{codigo}", quote(info_ficha["codigo_raw"], safe="")
                        )
                        with st.spinner("Consultando la ficha del proveedor..."):
                            encontrado, detalle_verif = verificar_en_catalogo_oficial(
                                cand["codigo_pedido"], url_ficha
                            )
                        if encontrado is True:
                            guardar_evidencia(cand["codigo_clean"], cand["producto_id"],
                                               "catalogo_oficial", detalle_verif)
                            st.success("✅ " + detalle_verif)
                        elif encontrado is False:
                            guardar_evidencia(cand["codigo_clean"], cand["producto_id"],
                                               "catalogo_no_lo_lista", detalle_verif)
                            st.warning("⚠️ " + detalle_verif)
                        else:
                            st.info(detalle_verif)

                    marca_nueva = None
                    if not cand["pedido_ya_cargado"]:
                        st.warning(
                            f"El código '{cand['codigo_pedido']}' no está cargado como producto. "
                            "Si confirmás, se crea con la marca que elijas y recién ahí se vinculan."
                        )
                        marca_nueva = st.selectbox(
                            "Marca para el código nuevo:", marcas_disponibles,
                            key=f"marca_cand_{cand['codigo_clean']}_{cand['producto_id']}"
                        ) if marcas_disponibles else None

                    cb1, cb2 = st.columns(2)
                    if cb1.button("✅ Confirmar equivalencia",
                                   key=f"conf_cand_{cand['codigo_clean']}_{cand['producto_id']}",
                                   type="primary"):
                        ok, error_conf = confirmar_candidata(
                            cand["codigo_clean"], cand["producto_id"], cand["codigo_pedido"], marca_nueva
                        )
                        if ok:
                            st.success("Equivalencia cargada. Ya aparece al buscar cualquiera de los dos códigos.")
                        else:
                            st.error(error_conf)
                    cb2.button("🚫 No es equivalente",
                                key=f"desc_cand_{cand['codigo_clean']}_{cand['producto_id']}",
                                on_click=descartar_candidata,
                                args=(cand["codigo_clean"], cand["producto_id"]),
                                help="No se vuelve a sugerir")

        st.markdown("---")
        st.markdown("**🛒 Últimas ventas marcadas**")
        c.execute("""SELECT v.fecha AS "Fecha", v.termino_pedido AS "Pidieron",
                            p.codigo_raw AS "Se llevaron", m.nombre AS "Marca", v.usuario AS "Empleado"
                     FROM ventas_registradas v
                     JOIN productos p ON p.id = v.producto_id
                     JOIN marcas m ON m.id = p.marca_id
                     ORDER BY v.id DESC LIMIT 25""")
        ultimas_ventas = filas_a_listas(c)
        if ultimas_ventas:
            st.dataframe(ultimas_ventas, use_container_width=True, hide_index=True)
        else:
            st.caption("Todavía no se marcó ninguna venta.")

# ============================================================
# LISTA PARA WHATSAPP
# ============================================================
if pagina == PAGINAS[5]:
    st.subheader("Armar lista de productos para enviar por WhatsApp")
    st.caption(
        "Buscá códigos en la pestaña Buscador y tocá '📋 Agregar a lista de WhatsApp'. "
        "Acá se arma un mensaje agrupado por producto, con las equivalencias y precios de cada marca."
    )

    lista = st.session_state.lista_whatsapp

    if not lista:
        st.info("Todavía no agregaste ningún producto a la lista. Andá a Buscador y agregá alguno.")
    else:
        st.markdown(f"**{len(lista)} producto(s) en la lista:**")
        for i, item in enumerate(lista):
            colT, colX = st.columns([5, 1])
            colT.write(f"{i + 1}. {item['codigo_buscado']} ({len(item['resultados'])} equivalencias)")
            if colX.button("🗑️", key=f"quitar_wa_{i}"):
                lista.pop(i)
                st.rerun()

        st.markdown("---")
        incluir_precio = st.checkbox("Incluir precios en el mensaje", value=True)
        incluir_stock = st.checkbox("Incluir stock en el mensaje", value=False)

        # Armado del texto del mensaje, agrupado por producto buscado
        encabezado_wa = obtener_config("whatsapp_encabezado", "🔧 *Equivalencias El Chavo*")
        pie_wa = obtener_config("whatsapp_pie", "")
        partes = [f"{encabezado_wa}\n"]
        for item in lista:
            partes.append(f"\n📦 *{item['codigo_buscado']}*")
            for fila in item["resultados"]:
                linea = f"  • {fila['Marca']}: {fila['Codigo']}"
                if fila.get("Descripcion"):
                    linea += f" - {fila['Descripcion']}"
                extras = []
                if incluir_precio and fila.get("Precio"):
                    extras.append(f"${fila['Precio']:,.0f}")
                if incluir_stock and fila.get("Stock") is not None:
                    extras.append(f"Stock: {fila['Stock']}")
                if extras:
                    linea += " (" + " · ".join(extras) + ")"
                partes.append(linea)
        if pie_wa.strip():
            partes.append(f"\n{pie_wa}")
        mensaje = "\n".join(partes)

        st.text_area("Vista previa del mensaje:", value=mensaje, height=300)

        alias_disponibles = listar_alias_transferencia()
        alias_elegido = None
        qr_real_para_pdf = None
        if alias_disponibles:
            opciones_alias = ["Sin QR de transferencia"] + [a["Nombre"] for a in alias_disponibles]
            alias_sel = st.selectbox("Alias para el QR del PDF (opcional):", opciones_alias, key="alias_para_pdf")
            if alias_sel != "Sin QR de transferencia":
                alias_elegido = next(a for a in alias_disponibles if a["Nombre"] == alias_sel)
                if alias_elegido["TieneQrReal"]:
                    qr_real_para_pdf = obtener_qr_real(alias_elegido["ID"])
                    st.caption("✅ Se va a usar el QR real que subiste para este alias.")
                else:
                    st.caption("ℹ️ Este alias no tiene QR real cargado — se va a generar uno con el alias/CBU como texto.")
        else:
            st.caption(
                "Todavía no cargaste ningún alias/CBU — podés hacerlo en 'Administrar' → "
                "'💳 Alias para QR de transferencia' si querés que la cotización incluya uno."
            )

        import urllib.parse
        url_whatsapp = "https://wa.me/?text=" + urllib.parse.quote(mensaje)
        col_wa, col_pdf = st.columns(2)
        col_wa.link_button("📲 Abrir en WhatsApp", url_whatsapp, type="primary", use_container_width=True)
        pdf_bytes = pdf_con_cache("cotizacion", generar_pdf_cotizacion, lista, incluir_precio,
                                   incluir_stock, alias_elegido, qr_real_para_pdf)
        col_pdf.download_button(
            "📄 Descargar cotización (PDF)", data=pdf_bytes,
            file_name=f"cotizacion_{datetime.now():%Y%m%d_%H%M}.pdf",
            mime="application/pdf", use_container_width=True
        )

        if st.button("🗑️ Vaciar toda la lista"):
            st.session_state.lista_whatsapp = []
            st.rerun()

# ============================================================
# VEHÍCULOS (ficha digital / historial de piezas)
# ============================================================
if pagina == PAGINAS[6]:
    st.subheader("🚗 Ficha digital del vehículo")
    st.caption(
        "Registrá la patente de un cliente frecuente junto con las piezas que le fuiste cambiando. "
        "La app avisa cuándo una pieza ya recorrió casi toda su vida útil estimada."
    )

    vehiculos_atrasados = listar_vehiculos_atrasados()
    with st.expander(f"⚠️ Vehículos con mantenimiento atrasado ({len(vehiculos_atrasados)})", expanded=bool(vehiculos_atrasados)):
        if not vehiculos_atrasados:
            st.caption(
                "Ninguno detectado por ahora (o todavía no cargaste km de registro/actual en los vehículos)."
            )
        else:
            st.caption("Ordenados por urgencia — el que tiene la pieza más atrasada aparece primero.")
            for item in vehiculos_atrasados:
                v = item["vehiculo"]
                nombre_auto = f"{v.get('marca_auto') or ''} {v.get('modelo_auto') or ''}".strip()
                piezas_txt = ", ".join(f"{p['Pieza']} (x{p['Atraso estimado']})" for p in item["piezas_atrasadas"])
                colv1, colv2 = st.columns([4, 1])
                colv1.write(f"**{v['patente']}** {nombre_auto} — {v.get('cliente_nombre') or 'sin nombre'}")
                colv1.caption(f"Atrasado: {piezas_txt}")
                if colv2.button("👁️ Ver", key=f"ver_atrasado_{v['id']}"):
                    st.session_state["patente_buscar"] = v["patente"]
                    st.rerun()

    st.markdown("---")
    st.markdown("**Buscar / registrar un vehículo**")

    with st.expander("📷 Cargar por foto de cédula/título (con IA)"):
        st.caption(
            "Sacale una foto a la cédula verde/azul o al título. La IA lee patente, marca, modelo, "
            "año y motorización — **siempre revisá los datos antes de guardar**, un OCR puede "
            "confundir letras o números parecidos."
        )
        foto_cedula = subir_archivo("Foto de la cédula/título:", ["png", "jpg", "jpeg"], "cedula")
        cedula_ok = archivo_listo(foto_cedula, "foto de la cédula")
        if foto_cedula:
            boton_otro_archivo("cedula", "🗑️ Usar otra foto", key="otra_foto_cedula")
        if st.button("🔍 Leer datos", disabled=not cedula_ok):
            with st.spinner("Leyendo..."):
                datos_cedula, error_cedula = extraer_datos_cedula(foto_cedula.getvalue())
            if error_cedula:
                st.error(error_cedula)
            else:
                st.session_state["datos_cedula_leidos"] = datos_cedula
        if st.session_state.get("datos_cedula_leidos"):
            datos = st.session_state["datos_cedula_leidos"]
            st.json(datos)
            if st.button("✅ Usar estos datos"):
                st.session_state["cedula_pendiente"] = datos
                st.session_state.pop("datos_cedula_leidos", None)
                st.rerun()

    # Si se leyó una cédula, precargar los campos ANTES de crear los widgets — si se hace
    # después de que ya se dibujaron en pantalla, Streamlit tira un error.
    if "cedula_pendiente" in st.session_state:
        datos = st.session_state.pop("cedula_pendiente")
        if datos.get("patente"):
            st.session_state["patente_buscar"] = str(datos["patente"]).strip().upper()
        st.session_state["form_marca_auto"] = datos.get("marca") or ""
        st.session_state["form_modelo_auto"] = datos.get("modelo") or ""
        st.session_state["form_anio_auto"] = str(datos.get("anio") or "")
        st.session_state["form_motorizacion_auto"] = datos.get("motorizacion") or ""

    # El VIN que se resolvió a patente en la vuelta anterior se vuelca ACÁ, antes de dibujar el
    # campo. Escribirlo después de dibujado —que es lo que se hacía— Streamlit no lo permite:
    # tira StreamlitAPIException y corta la pantalla. Pasaba con solo pegar un chasis de 17.
    if "patente_pendiente" in st.session_state:
        st.session_state["patente_buscar"] = st.session_state.pop("patente_pendiente")

    patente_input = st.text_input(
        "Patente o VIN:", placeholder="Ej: AB123CD — o el chasis completo", key="patente_buscar"
    ).strip().upper()

    # Si lo que pegaron es un VIN de 17, se resuelve a la patente y se sigue como siempre.
    if len(re.sub(r"\s", "", patente_input)) == 17:
        ficha_por_vin = buscar_vehiculo_por_vin(patente_input)
        if ficha_por_vin:
            patente_input = ficha_por_vin["patente"]
            avisar("success", f"🔎 Ese chasis es la patente **{patente_input}**.")
            st.session_state["patente_pendiente"] = patente_input
            st.rerun()
        else:
            st.info(
                "Ese VIN no está en ninguna ficha todavía. Cargá la patente del auto y poné el "
                "VIN en los datos del vehículo: la próxima vez lo encontrás pegando el chasis."
            )

    if patente_input:
        vehiculo = buscar_vehiculo(patente_input)

        with st.expander("✏️ Datos del cliente / vehículo", expanded=(vehiculo is None)):
            with st.form("form_vehiculo"):
                cv1, cv2 = st.columns(2)
                cliente_nombre = cv1.text_input("Nombre del cliente", value=(vehiculo or {}).get("cliente_nombre") or "")
                cliente_tel = cv2.text_input("Teléfono", value=(vehiculo or {}).get("cliente_telefono") or "")
                cv3, cv4 = st.columns(2)
                marca_auto = cv3.text_input("Marca del auto", value=(vehiculo or {}).get("marca_auto") or "",
                                             key="form_marca_auto")
                modelo_auto = cv4.text_input("Modelo", value=(vehiculo or {}).get("modelo_auto") or "",
                                              key="form_modelo_auto")
                cv5, cv6, cv7 = cols(3)
                anio_auto = cv5.text_input("Año", value=(vehiculo or {}).get("anio") or "", key="form_anio_auto")
                motorizacion_auto = cv6.text_input("Motorización", value=(vehiculo or {}).get("motorizacion") or "",
                                                    key="form_motorizacion_auto")
                km_actual_input = cv7.number_input(
                    "Km actual", min_value=0, step=1000,
                    value=int((vehiculo or {}).get("km_actual") or 0)
                )
                vin_auto = st.text_input(
                    "VIN / número de chasis (opcional)", value=(vehiculo or {}).get("vin") or "",
                    key="form_vin_auto", max_chars=17,
                    help="Cargarlo sirve para dos cosas: podés encontrar este auto por el chasis "
                         "además de por la patente, y la app aprende sola qué modelo corresponde "
                         "a ese patrón de VIN, así el próximo auto igual se completa solo."
                ).strip().upper()
                guardar_vehiculo = st.form_submit_button("💾 Guardar vehículo", type="primary")
            if guardar_vehiculo:
                vin_limpio = re.sub(r"\s", "", vin_auto)
                if vin_limpio and len(vin_limpio) != 17:
                    st.warning("El VIN tiene que tener 17 caracteres — se guardó el resto sin él.")
                    vin_limpio = ""
                get_or_create_vehiculo(patente_input, cliente_nombre, cliente_tel, marca_auto, modelo_auto,
                                        km_actual_input or None, anio_auto, motorizacion_auto,
                                        vin=vin_limpio)
                st.success(f"Vehículo {patente_input} guardado.")
                if vin_limpio and modelo_auto.strip():
                    st.caption(f"📚 De paso quedó aprendido que el patrón {vin_limpio[:3]}-"
                               f"{vin_limpio[3:8]} es un {modelo_auto.strip()}.")
                st.rerun()

        vehiculo = buscar_vehiculo(patente_input)
        if vehiculo:
            km_actual = vehiculo.get("km_actual")
            km_registro = vehiculo.get("km_registro")
            st.write(
                f"**{vehiculo.get('marca_auto') or ''} {vehiculo.get('modelo_auto') or ''}** — "
                f"Cliente: {vehiculo.get('cliente_nombre') or 'sin nombre'}"
            )

            km_calc = calcular_km_recorridos(vehiculo)
            mk1, mk2, mk3 = st.columns(3)
            mk1.metric("Km de registro", km_registro if km_registro is not None else "—")
            mk2.metric("Km actual", km_actual if km_actual is not None else "—")
            mk3.metric("Km recorridos", km_calc["km_recorridos"] if km_calc["km_recorridos"] is not None else "—")
            if km_calc["promedio_mensual"] is not None:
                st.caption(
                    f"📈 Promedio aproximado: **{km_calc['promedio_mensual']:,} km/mes** "
                    f"(en base a {km_calc['dias_transcurridos']} día(s) desde que se registró el vehículo)."
                )

            with st.expander("✏️ Corregir km de registro (solo si se cargó mal la primera vez)"):
                st.caption(
                    "El km de registro queda fijo automáticamente la primera vez que cargás el vehículo. "
                    "Usá esto solo para corregir un error de tipeo — cambiarlo afecta los cálculos de abajo."
                )
                nuevo_km_registro = st.number_input(
                    "Km de registro correcto", min_value=0, step=1000,
                    value=int(km_registro or 0), key="corregir_km_registro"
                )
                if st.button("💾 Corregir km de registro"):
                    actualizar_km_registro(vehiculo["id"], nuevo_km_registro or None)
                    avisar("success", "Km de registro actualizado.")
                    st.rerun()

            alertas = []
            if km_actual is not None:
                alertas = calcular_alertas_vehiculo(vehiculo["id"], km_actual)
                if alertas:
                    st.warning(f"⚠️ {len(alertas)} pieza(s) cerca de cumplir su vida útil estimada:")
                    st.dataframe(alertas, use_container_width=True, hide_index=True)
                else:
                    st.info("Sin alertas de mantenimiento por ahora.")

            st.markdown("---")
            st.markdown("**➕ Agregar pieza al historial**")
            with st.form("form_pieza", clear_on_submit=True):
                cp1, cp2 = st.columns(2)
                desc_pieza = cp1.text_input("Descripción de la pieza", placeholder="Ej: Kit de distribución")
                marca_pieza = cp2.text_input("Marca de la pieza", placeholder="Ej: SKF")
                cp3, cp4, cp5 = st.columns(3)
                codigo_pieza = cp3.text_input("Código (opcional)")
                km_instalacion = cp4.number_input("Km al instalarla", min_value=0, step=1000,
                                                    value=int(km_actual or 0))
                vida_util = cp5.number_input("Vida útil estimada (km, opcional)", min_value=0, step=5000, value=0)
                nota_pieza = st.text_input("Nota (opcional)")
                agregar_pieza = st.form_submit_button("➕ Agregar al historial", type="primary")
            if agregar_pieza:
                if not desc_pieza.strip():
                    st.warning("Completá la descripción de la pieza.")
                else:
                    agregar_pieza_historial(vehiculo["id"], desc_pieza, marca_pieza, codigo_pieza,
                                             km_instalacion or None, vida_util or None, nota_pieza)
                    avisar("success", "Pieza agregada al historial.")
                    st.rerun()

            st.markdown("---")
            st.markdown("**📋 Historial completo**")
            historial_vehiculo = listar_historial_vehiculo(vehiculo["id"])
            if historial_vehiculo:
                st.dataframe(quitar_id(historial_vehiculo), use_container_width=True, hide_index=True)
            else:
                st.caption("Todavía no hay piezas registradas para este vehículo.")

            st.markdown("---")
            st.markdown("**🔧 Proyección de mantenimiento**")
            st.caption(
                "Compara, para cada tipo de pieza con vida útil cargada, cuántas veces se cambió "
                "realmente contra cuántas veces debería haberse cambiado según los km recorridos "
                "totales desde que se registró el vehículo."
            )
            proyeccion = []
            # 'atrasadas' también se inicializa acá aunque más abajo siempre se asigne en las dos
            # ramas del else: si km_recorridos es None, nunca se asignaba, y el único motivo por
            # el que no explotaba es que "if proyeccion and atrasadas" corta antes de leerla.
            # Alcanzaba con dar vuelta esa condición para romper la ficha del vehículo.
            atrasadas = []
            if km_calc["km_recorridos"] is None:
                st.info(
                    "Para calcular esto hace falta el km de registro y el km actual del vehículo "
                    "(completá 'Kilometraje actual' arriba si todavía no lo cargaste)."
                )
            else:
                proyeccion = calcular_proyeccion_mantenimiento(vehiculo["id"], km_calc["km_recorridos"])
                if proyeccion:
                    atrasadas = [p for p in proyeccion if p["Atraso estimado"] > 0]
                    if atrasadas:
                        st.warning(f"⚠️ {len(atrasadas)} pieza(s) con cambios atrasados según el kilometraje:")
                    st.dataframe(proyeccion, use_container_width=True, hide_index=True)
                else:
                    st.caption("Todavía no hay piezas con vida útil cargada para proyectar.")
                    atrasadas = []

            st.markdown("---")
            st.markdown("**📤 Compartir con el cliente**")
            col_pdf, col_wa = st.columns(2)
            with col_pdf:
                st.download_button(
                    "📄 Descargar ficha en PDF",
                    data=pdf_con_cache("ficha_vehiculo", generar_pdf_ficha_vehiculo,
                                        vehiculo, km_calc, alertas, proyeccion, historial_vehiculo),
                    file_name=f"ficha_{vehiculo['patente']}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            with col_wa:
                if proyeccion and atrasadas:
                    nombre_cliente = vehiculo.get("cliente_nombre") or ""
                    nombre_auto_msg = f"{vehiculo.get('marca_auto') or ''} {vehiculo.get('modelo_auto') or ''}".strip()
                    piezas_atrasadas_txt = ", ".join(p["Pieza"] for p in atrasadas)
                    mensaje_wa = (
                        f"Hola {nombre_cliente}! Te escribimos de El Chavo. Revisando el kilometraje de tu "
                        f"{nombre_auto_msg} ({vehiculo['patente']}), notamos que tenés atrasado el cambio de: "
                        f"{piezas_atrasadas_txt}. ¿Coordinamos un turno?"
                    ).strip()
                    tel_limpio = re.sub(r"\D", "", vehiculo.get("cliente_telefono") or "")
                    url_wa_vehiculo = (
                        f"https://wa.me/{tel_limpio}" if tel_limpio else "https://wa.me/"
                    ) + "?text=" + quote(mensaje_wa)
                    st.link_button("📲 Avisar atraso por WhatsApp", url_wa_vehiculo,
                                    type="primary", use_container_width=True)
                else:
                    st.caption("Sin atrasos detectados todavía para avisar por WhatsApp.")

# ============================================================
# MODO MECÁNICO
# ============================================================
if pagina == PAGINAS[7]:
    st.subheader("🛠️ Modo Mecánico")

    SUB_MEC = ["📖 Códigos DTC", "🔢 Chasis / VIN", "🚙 Repuestos por vehículo",
               "🗺️ Esquemas", "🧮 Conversor de unidades"]
    if st.session_state.get("sub_mec") not in SUB_MEC:
        st.session_state["sub_mec"] = SUB_MEC[0]
    st.radio("Sub-sección:", SUB_MEC, key="sub_mec", horizontal=True,
             label_visibility="collapsed")
    sub_mec = st.session_state["sub_mec"]

    # -------- Diccionario de códigos OBD2 / DTC --------
    if sub_mec == SUB_MEC[0]:
        st.caption(
            f"Diccionario de códigos de falla OBD2/DTC. Arranca con {contar_dtc()} códigos genéricos "
            "estándar (no específicos de marca) — sumá los que te falten con el formulario de abajo. "
            "Los códigos P1xxx u otros específicos de fabricante se cargan indicando la marca, porque "
            "el mismo número puede significar algo distinto según el auto."
        )
        fabricantes_dtc = ["Todos", "Genérico"] + listar_fabricantes_dtc()
        cdb1, cdb2 = st.columns([2, 1])
        codigo_buscar = cdb1.text_input("Buscar código:", placeholder="Ej: P0301 o P1105", key="dtc_buscar")
        filtro_fab_dtc = cdb2.selectbox("Fabricante:", fabricantes_dtc, key="dtc_filtro_fab")
        if codigo_buscar.strip():
            res_dtc = buscar_dtc(codigo_buscar, filtro_fab_dtc)
            if res_dtc:
                st.dataframe(res_dtc, use_container_width=True, hide_index=True)
            else:
                st.warning("No tengo ese código cargado todavía (con ese filtro de fabricante). Podés agregarlo abajo.")

        with st.expander("➕ Agregar / corregir un código"):
            st.caption("Dejá 'Fabricante' vacío si es un código genérico (P0xxx). Completalo si es específico de una marca (ej: Ford, Toyota).")
            with st.form("form_dtc", clear_on_submit=True):
                cd1, cd2, cd3 = st.columns(3)
                nuevo_codigo = cd1.text_input("Código (ej: P0301)")
                nuevo_fabricante = cd2.text_input("Fabricante (opcional)", placeholder="Ej: Ford")
                nuevo_sistema = cd3.text_input("Sistema (ej: Motor, Transmisión)")
                nueva_desc = st.text_input("Descripción")
                nuevas_causas = st.text_input("Causas posibles (opcional)")
                guardar_dtc_btn = st.form_submit_button("💾 Guardar código", type="primary")
            if guardar_dtc_btn:
                if not nuevo_codigo.strip() or not nueva_desc.strip():
                    st.warning("Completá al menos el código y la descripción.")
                else:
                    agregar_dtc(nuevo_codigo, nueva_desc, nuevo_sistema, nuevas_causas, nuevo_fabricante)
                    etiqueta_fab = f" ({nuevo_fabricante.strip()})" if nuevo_fabricante.strip() else " (genérico)"
                    avisar("success", f"Código {nuevo_codigo.upper()}{etiqueta_fab} guardado.")
                    st.rerun()

        with st.expander("📋 Carga masiva de códigos (pegar texto)"):
            st.caption(
                "Un código por línea, formato: `codigo;descripción;sistema;causas;fabricante` "
                "(sistema, causas y fabricante son opcionales — dejá fabricante vacío para códigos genéricos)."
            )
            texto_dtc = st.text_area("Pegá los códigos acá:", height=150, key="dtc_masivo",
                                      placeholder="P0455;Fuga grande en sistema EVAP;Emisiones;Tapa de nafta, manguera\n"
                                                   "P1105;Solenoide de presión de combustible;Motor;;Chrysler")
            if st.button("📥 Importar códigos"):
                if texto_dtc.strip():
                    cargados_dtc = importar_dtc_masivo(texto_dtc)
                    avisar("success", f"Se cargaron/actualizaron {cargados_dtc} código(s).")
                    st.rerun()
                else:
                    st.warning("Pegá al menos un código.")

    # Definidas acá porque las usan las dos vistas de esquemas de más abajo.
    CATEGORIAS_ESQUEMA = [
        "Motor", "Refrigeración", "Retenes y juntas", "Frenos", "Suspensión", "Dirección",
        "Transmisión", "Embrague", "Correas y distribución", "Eléctrico", "Combustible",
        "Escape", "Aire acondicionado", "Otro"
    ]

    def mostrar_lista_esquemas(lista_esq):
        if not lista_esq:
            st.caption("No hay esquemas cargados acá todavía.")
            return
        for esq in lista_esq:
            titulo_expander = f"🗺️ {esq['titulo']}" + (" 🤖 (orientativo, generado por IA)" if esq.get("generado_ia") else "")
            with st.expander(titulo_expander):
                if esq.get("generado_ia"):
                    st.warning(
                        "🤖 Esta imagen fue generada por IA como referencia orientativa — "
                        "NO es una foto real de este vehículo. No la uses para identificar piezas con precisión."
                    )
                if esq.get("descripcion"):
                    st.write(esq["descripcion"])
                img_bytes = obtener_imagen_esquema(esq["id"])
                puntos = listar_puntos_esquema(esq["id"])
                if img_bytes:
                    imagen_a_mostrar = imagen_esquema_lista_para_mostrar(
                        img_bytes, firma_de_puntos(puntos), puntos
                    )
                    st.image(imagen_a_mostrar, use_container_width=True)
                    if any(p.get("pos_x") is not None for p in puntos):
                        st.caption("Los números marcados en la foto corresponden a la lista de piezas de abajo.")

                # Piezas marcadas en el esquema, con búsqueda directa por código
                if puntos:
                    st.markdown("**🔩 Piezas de este esquema**")
                    for punto in puntos:
                        etiqueta = f"{punto['numero']}. " if punto.get("numero") else ""
                        cp1, cp2 = st.columns([3, 1])
                        cp1.write(f"{etiqueta}{punto['nombre_pieza']}" + (f" — `{punto['codigo']}`" if punto.get("codigo") else ""))
                        if punto.get("codigo"):
                            if cp2.button("🔍 Buscar", key=f"buscar_punto_{punto['id']}"):
                                clean_punto = sanitizar(punto["codigo"])
                                res_punto = buscar_por_codigo(clean_punto) if clean_punto else []
                                if not res_punto:
                                    res_punto = buscar_por_texto(punto["nombre_pieza"])
                                if res_punto:
                                    st.dataframe(quitar_id(res_punto), use_container_width=True, hide_index=True)
                                else:
                                    st.error(f"No encontré '{punto['codigo']}' ni '{punto['nombre_pieza']}' en la base.")
                        if es_admin():
                            cp2.button("🗑️", key=f"del_punto_{punto['id']}",
                                        on_click=eliminar_punto_esquema, args=(punto["id"],))

                if es_operador_o_admin():
                    if seccion_plegable("➕ Agregar pieza a este esquema",
                                         key=f"agregar_pieza_{esq['id']}"):
                        if img_bytes:
                            st.caption(
                                "Mirá la foto de arriba y estimá en qué parte está la pieza: "
                                "0% = borde izquierdo/superior, 100% = borde derecho/inferior."
                            )
                        cpp1, cpp2, cpp3 = st.columns([1, 2, 2])
                        num_punto = cpp1.text_input("N°", key=f"num_punto_{esq['id']}", placeholder="1")
                        nombre_punto = cpp2.text_input("Nombre de la pieza", key=f"nombre_punto_{esq['id']}")
                        codigo_punto = cpp3.text_input("Código (opcional)", key=f"codigo_punto_{esq['id']}")
                        marcar_posicion = st.checkbox(
                            "Marcar posición en la foto", value=bool(img_bytes), key=f"marcar_pos_{esq['id']}",
                            disabled=not img_bytes
                        )
                        pos_x_punto, pos_y_punto = None, None
                        if marcar_posicion and img_bytes:
                            cpx, cpy = st.columns(2)
                            pos_x_punto = cpx.slider("Posición horizontal (%)", 0, 100, 50, key=f"posx_{esq['id']}")
                            pos_y_punto = cpy.slider("Posición vertical (%)", 0, 100, 50, key=f"posy_{esq['id']}")
                            vista_previa = generar_imagen_con_marcadores(
                                img_bytes,
                                puntos + [{"numero": num_punto or "?", "pos_x": pos_x_punto, "pos_y": pos_y_punto}]
                            )
                            st.image(vista_previa, use_container_width=True, caption="Vista previa de dónde quedaría el marcador")
                        if st.button("💾 Agregar pieza", key=f"agregar_punto_{esq['id']}"):
                            if not nombre_punto.strip():
                                st.warning("Completá el nombre de la pieza.")
                            else:
                                vinculado = agregar_punto_esquema(
                                    esq["id"], num_punto, nombre_punto, codigo_punto, pos_x_punto, pos_y_punto
                                )
                                if codigo_punto.strip() and not vinculado:
                                    st.info(
                                        "Pieza agregada. El código no coincide con ningún producto cargado "
                                        "todavía, pero igual queda guardado como referencia."
                                    )
                                else:
                                    st.success("Pieza agregada.")
                                st.rerun()

                    if st.button("🗑️ Eliminar este esquema", key=f"del_esq_{esq['id']}"):
                        eliminar_esquema(esq["id"])
                        st.rerun()

    # -------- Chasis / VIN (pantalla única) --------
    if sub_mec == SUB_MEC[1]:
        panel_vin(clave="vin_mec")

        st.markdown("---")
        st.markdown("**⚙️ Lo que la app fue aprendiendo**")
        st.caption(
            "Estas tablas son tuyas: se llenan solas con cada ficha de vehículo que cargues con "
            "VIN, y con lo que le enseñes arriba. Acá se revisan y se corrigen."
        )

        c.execute("""SELECT COUNT(*) FROM vehiculos WHERE vin IS NOT NULL AND LENGTH(vin) = 17
                     AND ((modelo_auto IS NOT NULL AND modelo_auto <> '')
                          OR (motorizacion IS NOT NULL AND motorizacion <> ''))""")
        fichas_con_vin = c.fetchone()[0]
        if fichas_con_vin:
            st.caption(f"Tenés {fichas_con_vin} ficha(s) de vehículo con VIN cargado.")
            if st.button("📚 Aprender modelos y motores de esas fichas"):
                n_mod, n_mot = aprender_modelos_de_fichas_existentes()
                avisar("success", f"Se aprendieron {n_mod} modelo(s) y {n_mot} motor(es) de tus fichas.")
                st.rerun()

        fabricantes_cargados = listar_fabricantes_vin()
        with st.expander(f"🏭 Fabricantes por WMI ({len(fabricantes_cargados)})"):
            st.caption(
                "Los 3 primeros caracteres del VIN. Vienen 328 cargados de la lista pública "
                "internacional; podés corregir o agregar los que falten."
            )
            if fabricantes_cargados:
                st.dataframe(fabricantes_cargados, use_container_width=True, hide_index=True)
            with st.form("form_wmi_admin", clear_on_submit=True):
                cw1, cw2, cw3 = st.columns(3)
                nuevo_wmi = cw1.text_input("WMI (3 caracteres)", max_chars=3)
                nuevo_fabricante = cw2.text_input("Fabricante")
                nuevo_pais_vin = cw3.text_input("País")
                if st.form_submit_button("💾 Guardar WMI"):
                    if len(nuevo_wmi.strip()) != 3 or not nuevo_fabricante.strip():
                        st.warning("El WMI debe tener 3 caracteres y el fabricante es obligatorio.")
                    else:
                        agregar_fabricante_vin(nuevo_wmi, nuevo_fabricante, nuevo_pais_vin)
                        avisar("success", f"WMI {nuevo_wmi.upper()} guardado.")
                        st.rerun()

        modelos_cargados = listar_modelos_vin()
        with st.expander(f"🚗 Modelos aprendidos ({len(modelos_cargados)})"):
            if modelos_cargados:
                st.dataframe(modelos_cargados, use_container_width=True, hide_index=True)
                cbm1, cbm2 = st.columns(2)
                wmi_borrar = cbm1.text_input("WMI a borrar", max_chars=3, key="wmi_borrar_modelo")
                vds_borrar = cbm2.text_input("Patrón (VDS) a borrar", max_chars=5, key="vds_borrar_modelo")
                if st.button("🗑️ Borrar ese patrón", disabled=not (wmi_borrar and vds_borrar)):
                    if olvidar_modelo_vin(wmi_borrar, vds_borrar):
                        st.success("Patrón borrado.")
                    else:
                        st.warning("No encontré ese patrón.")
                    st.rerun()
            else:
                st.caption(
                    "Todavía ninguno. Cada ficha de vehículo que cargues con VIN y modelo suma uno."
                )

        motores_cargados = listar_motores_vin()
        with st.expander(f"⚙️ Motores aprendidos ({len(motores_cargados)})"):
            st.caption(
                "La 8ª posición del VIN es el código de motor. Es el patrón que mejor rinde: el "
                "mismo código se repite en toda la gama de la marca, así que enseñarlo una vez "
                "sirve para los otros modelos."
            )
            if motores_cargados:
                st.dataframe(motores_cargados, use_container_width=True, hide_index=True)
                cbt1, cbt2 = st.columns(2)
                wmi_bm = cbt1.text_input("WMI a borrar", max_chars=3, key="wmi_borrar_motor")
                cod_bm = cbt2.text_input("Código (8ª posición)", max_chars=1, key="cod_borrar_motor")
                if st.button("🗑️ Borrar ese motor", disabled=not (wmi_bm and cod_bm)):
                    if olvidar_motor_vin(wmi_bm, cod_bm):
                        st.success("Borrado.")
                    else:
                        st.warning("No encontré ese código.")
                    st.rerun()
            else:
                st.caption("Todavía ninguno.")

    if sub_mec == SUB_MEC[2]:
        st.markdown("**🚙 Repuestos por vehículo**")
        st.caption(
            "Buscá lo que le entra a un auto entrando por marca y modelo, en vez de por código. "
            "Sale de las descripciones de tus propias listas — o sea que crece solo cada vez que "
            "importás un proveedor nuevo."
        )

        c.execute("SELECT COUNT(*) FROM productos")
        version_catalogo = c.fetchone()[0]   # cambia al cargar listas: refresca el caché
        disponibles = marcas_vehiculo_disponibles(version_catalogo)

        if not disponibles:
            st.info(
                "Todavía no se detectó ninguna marca de vehículo en las descripciones del catálogo. "
                "Aparecen solas a medida que vas importando listas de proveedores."
            )
        else:
            etiquetas_marcas = [f"{m} ({n} productos)" for m, n in disponibles]
            mapa_marcas = {f"{m} ({n} productos)": m for m, n in disponibles}

            st.info(
                "🔢 ¿Tenés el número de chasis? Andá a **🛠️ Modo Mecánico → 🔢 Chasis / VIN**: "
                "pegás el VIN y te da directamente los repuestos que lleva ese auto, sin tener "
                "que elegir marca y modelo acá a mano."
            )

            etiqueta_elegida = st.selectbox("Marca del vehículo:", etiquetas_marcas,
                                             key="sel_marca_vehiculo")
            marca_elegida = mapa_marcas[etiqueta_elegida]
            por_categoria = catalogo_por_vehiculo(marca_elegida, version_catalogo)

            if not por_categoria:
                st.caption("No se pudo separar la categoría de esas descripciones.")
            else:
                # El modelo sale de una lista detectada del propio catálogo, y el año filtra
                # usando los rangos que traen las descripciones ("1974/81", "1998/...").
                # Es lo más parecido a un catálogo de aplicaciones que se puede armar sin
                # comprar una base licenciada: cubre lo que vos vendés, no todo el mercado.
                modelos_detectados = modelos_de_marca(marca_elegida, version_catalogo)
                cmv1, cmv2 = cols(2)
                with cmv1:
                    if modelos_detectados:
                        opciones_mod = ["Todos los modelos"] + [f"{m} ({n})" for m, n in modelos_detectados[:120]]
                        mapa_mod = {f"{m} ({n})": m for m, n in modelos_detectados[:120]}

                        if st.session_state.get("sel_modelo_vehiculo") not in opciones_mod:
                            # Cambió la marca y el modelo guardado ya no existe: sin esto el
                            # selector tira excepción y se cae la pantalla.
                            st.session_state.pop("sel_modelo_vehiculo", None)

                        mod_etiqueta = st.selectbox("Modelo / motor:", opciones_mod, key="sel_modelo_vehiculo")
                        modelo_elegido = mapa_mod.get(mod_etiqueta)
                    else:
                        modelo_elegido = None
                        st.caption("No se detectaron modelos para esta marca.")
                with cmv2:
                    anio_filtro = st.number_input(
                        "Año del vehículo (0 = cualquiera):", min_value=0, max_value=2030,
                        value=0, step=1, key="anio_vehiculo_filtro",
                        help="Usa los rangos que traen las descripciones. Lo que no aclara años "
                             "se muestra igual, marcado aparte."
                    )

                filtro_modelo = st.text_input(
                    "Además, filtrar por texto (opcional):",
                    key="filtro_modelo_vehiculo", placeholder="Ej: 1.6, inyección, turbo..."
                ).strip().upper()

                categorias = sorted(por_categoria.keys(), key=lambda k: -len(por_categoria[k]))
                opciones_cat = ["Todas las categorías"] + [f"{c_} ({len(por_categoria[c_])})"
                                                            for c_ in categorias]
                mapa_cat = {f"{c_} ({len(por_categoria[c_])})": c_ for c_ in categorias}
                cat_etiqueta = st.selectbox("Categoría (opcional):", opciones_cat,
                                             key="sel_categoria_vehiculo")

                if cat_etiqueta == "Todas las categorías":
                    items = [x for lista in por_categoria.values() for x in lista]
                else:
                    items = list(por_categoria[mapa_cat[cat_etiqueta]])

                if modelo_elegido:
                    items = [x for x in items if modelo_elegido in (x["Descripcion"] or "").upper()]
                if filtro_modelo:
                    palabras = [p for p in filtro_modelo.split() if p]
                    items = [x for x in items
                             if all(p in (x["Descripcion"] or "").upper() for p in palabras)]

                # El año separa en tres grupos: sirve, no sirve, y "la descripción no lo aclara"
                sin_dato_anio = []
                if anio_filtro:
                    coinciden, sin_dato = [], []
                    for x in items:
                        r = sirve_para_anio(x["Descripcion"], int(anio_filtro))
                        if r is True:
                            coinciden.append(x)
                        elif r is None:
                            sin_dato.append(x)
                    items, sin_dato_anio = coinciden, sin_dato

                hay_filtro = modelo_elegido or filtro_modelo or anio_filtro or \
                    cat_etiqueta != "Todas las categorías"
                if not hay_filtro:
                    st.info(
                        f"Hay {len(items)} repuesto(s) de **{marca_elegida}**. Elegí el modelo "
                        "(o escribí el año) para achicar la lista."
                    )
                    items = items[:25]
                else:
                    resumen_filtro = " · ".join(filter(None, [
                        marca_elegida, modelo_elegido,
                        f"año {int(anio_filtro)}" if anio_filtro else None,
                    ]))
                    st.success(f"**{len(items)}** repuesto(s) para {resumen_filtro}")

                if items:
                    st.caption("👆 Tocá un código para ver todas sus equivalencias:")
                    for it in items[:30]:
                        cvv1, cvv2 = st.columns([1.2, 3])
                        cvv1.button(f"🔎 {it['Codigo']}", key=f"veh_{it['ID']}",
                                     on_click=cb_ver_equivalencias, args=(it["Codigo"],))
                        precio_v = f"${it['Precio']:,.0f}" if it.get("Precio") else "s/precio"
                        stock_v = it.get("Stock")
                        detalle = it.get("_aplicacion") or it.get("Descripcion") or ""
                        desde_a, hasta_a = extraer_anios(it["Descripcion"])
                        anios_txt = ""
                        if desde_a:
                            anios_txt = f" · {desde_a}–{hasta_a if hasta_a else 'en adelante'}"
                        cvv2.caption(f"**{it['_categoria']}**{anios_txt} · {detalle[:64]}  \n"
                                      f"{it['Marca']} · {precio_v} · stock "
                                      f"{stock_v if stock_v is not None else 's/d'}")
                    if len(items) > 30:
                        st.caption(f"(mostrando 30 de {len(items)})")
                elif hay_filtro:
                    st.warning("Ningún repuesto coincide con esa combinación.")

                if sin_dato_anio:
                    with st.expander(f"❔ {len(sin_dato_anio)} repuesto(s) que no aclaran el año "
                                      "(pueden servir igual)"):
                        st.caption(
                            "La descripción no dice para qué años es, así que no se puede afirmar "
                            "ni descartar. Se muestran aparte para que decidas vos."
                        )
                        for it in sin_dato_anio[:20]:
                            csd1, csd2 = st.columns([1.2, 3])
                            csd1.button(f"🔎 {it['Codigo']}", key=f"vehsd_{it['ID']}",
                                         on_click=cb_ver_equivalencias, args=(it["Codigo"],))
                            csd2.caption(f"**{it['_categoria']}** · "
                                          f"{(it.get('_aplicacion') or '')[:64]}")
                        if len(sin_dato_anio) > 20:
                            st.caption(f"(mostrando 20 de {len(sin_dato_anio)})")

            esquemas_veh = esquemas_de_vehiculo(marca_elegida)
            if esquemas_veh:
                st.markdown("---")
                st.markdown(f"**🗺️ Esquemas cargados de {marca_elegida}**")
                for e in esquemas_veh[:10]:
                    st.caption(f"• {e['titulo']} — {e['modelo_auto'] or ''} {e['sistema'] or ''}")
                st.caption("Se ven completos, con las piezas marcadas, en la sección Esquemas.")


    if sub_mec == SUB_MEC[3]:
        st.caption(
            "Diagramas organizados por Marca › Vehículo › Sistema, donde cada pieza marcada tiene "
            "su código vinculado al catálogo — así se busca directo desde el dibujo, ya sea en el "
            "taller o en el mostrador de una casa de repuestos. Las imágenes las tenés que subir vos."
        )
        modo_esq = st.radio(
            "¿Cómo lo buscás?", ["📂 Explorar por categoría", "🔎 Buscar por texto"],
            horizontal=True, key="modo_esquemas"
        )

        if modo_esq.startswith("📂"):
            marcas_esq = listar_marcas_esquemas()
            if not marcas_esq:
                st.info("Todavía no hay esquemas cargados con marca definida. Subí el primero más abajo.")
            else:
                marca_sel = st.selectbox("Marca:", marcas_esq, key="esq_marca_sel")
                modelos_esq = listar_modelos_esquemas(marca_sel)
                if not modelos_esq:
                    st.caption(f"No hay vehículos cargados todavía para {marca_sel}.")
                else:
                    modelo_sel = st.selectbox("Vehículo / modelo:", modelos_esq, key="esq_modelo_sel")
                    sistemas_esq = listar_sistemas_esquemas(marca_sel, modelo_sel)
                    if not sistemas_esq:
                        st.caption("No hay esquemas con sistema/parte definida para este vehículo.")
                    else:
                        sistema_sel = st.selectbox("Sistema / parte:", sistemas_esq, key="esq_sistema_sel")
                        mostrar_lista_esquemas(listar_esquemas_por_categoria(marca_sel, modelo_sel, sistema_sel))
        else:
            filtro_esq = st.text_input("Buscar esquema (título, marca, modelo o sistema):", key="esq_filtro")
            mostrar_lista_esquemas(listar_esquemas(filtro_esq))

        st.markdown("---")
        if not pedir_password_admin("subir esquemas nuevos"):
            pass
        else:
            marcas_existentes = listar_marcas_esquemas()

            st.markdown("**🚗 Precargar marca / vehículo (sin imagen todavía)**")
            st.caption(
                "Dejá lista la estructura del árbol aunque todavía no tengas ningún esquema para subir — "
                "va a aparecer en 'Explorar por categoría' apenas la guardes."
            )
            cp1, cp2 = st.columns(2)
            if marcas_existentes:
                marca_pre_opcion = cp1.selectbox("Marca", marcas_existentes + ["➕ Nueva marca..."], key="pre_marca_opcion")
                marca_pre = cp1.text_input("Nombre de la nueva marca", key="pre_marca_nueva") \
                    if marca_pre_opcion == "➕ Nueva marca..." else marca_pre_opcion
            else:
                marca_pre = cp1.text_input("Marca", key="pre_marca_sola")
            modelo_pre = cp2.text_input("Vehículo / modelo", placeholder="Ej: Corsa", key="pre_modelo")
            if st.button("➕ Precargar"):
                if not marca_pre.strip() or not modelo_pre.strip():
                    st.warning("Completá marca y modelo.")
                else:
                    agregar_vehiculo_catalogo(marca_pre, modelo_pre)
                    st.success(f"{marca_pre.strip()} {modelo_pre.strip()} precargado.")
                    for k in ["pre_marca_nueva", "pre_marca_sola", "pre_modelo"]:
                        st.session_state.pop(k, None)
                    st.rerun()

            precargados = listar_catalogo_precargado()
            if precargados:
                with st.expander(f"📋 Ver / borrar precargados sin esquema todavía ({len(precargados)})"):
                    for pv in precargados:
                        colp1, colp2 = st.columns([4, 1])
                        colp1.write(f"{pv['marca']} — {pv['modelo']}")
                        colp2.button("🗑️", key=f"del_precarga_{pv['marca']}_{pv['modelo']}",
                                      on_click=eliminar_vehiculo_catalogo, args=(pv["marca"], pv["modelo"]))

            st.markdown("---")
            st.markdown("**➕ Subir un esquema nuevo**")
            marcas_existentes = listar_marcas_esquemas()  # puede haber cambiado si acabás de precargar una
            titulo_esq = st.text_input("Título", placeholder="Ej: Esquema eléctrico bomba de combustible", key="esq_titulo")
            ce1, ce2 = st.columns(2)
            if marcas_existentes:
                marca_opcion = ce1.selectbox("Marca", marcas_existentes + ["➕ Nueva marca..."], key="esq_marca_opcion")
                marca_esq = ce1.text_input("Nombre de la nueva marca", key="esq_marca_nueva") \
                    if marca_opcion == "➕ Nueva marca..." else marca_opcion
            else:
                marca_esq = ce1.text_input("Marca del auto", key="esq_marca_sola")
            modelos_para_marca = listar_modelos_esquemas(marca_esq) if marca_esq else []
            if modelos_para_marca:
                modelo_opcion = ce2.selectbox("Vehículo / modelo", modelos_para_marca + ["➕ Nuevo modelo..."], key="esq_modelo_opcion")
                modelo_esq = ce2.text_input("Nombre del nuevo modelo", key="esq_modelo_nuevo") \
                    if modelo_opcion == "➕ Nuevo modelo..." else modelo_opcion
            else:
                modelo_esq = ce2.text_input("Vehículo / modelo", placeholder="Ej: Gol Trend", key="esq_modelo_solo")
            sistema_opcion = st.selectbox("Sistema / parte:", CATEGORIAS_ESQUEMA, key="esq_sistema_opcion")
            sistema_esq = st.text_input("Especificá el sistema/parte:", key="esq_sistema_nuevo") \
                if sistema_opcion == "Otro" else sistema_opcion
            desc_esq = st.text_input("Descripción (opcional)", key="esq_desc")

            origen_imagen = st.radio(
                "¿De dónde sale la imagen?",
                ["📷 Subir foto real", "🤖 Generar orientativo con IA (sin foto real)"],
                key="esq_origen_imagen"
            )

            archivo_esq = None
            imagen_generada_bytes = None
            if origen_imagen.startswith("📷"):
                archivo_esq = subir_archivo("Imagen del esquema", ["png", "jpg", "jpeg"], "esquema")
                if archivo_esq:
                    archivo_listo(archivo_esq, "imagen")
                    boton_otro_archivo("esquema", "🗑️ Usar otra imagen", key="otra_img_esquema")
            else:
                explicar(
                    "Para cuando no tenés el auto físico enfrente (útil en el mostrador de una casa de "
                    "repuestos):",
                    "la IA arma un dibujo genérico de referencia, **no una foto real de ese vehículo**. "
                    "Sirve para orientar, no para identificar piezas con precisión milimétrica. Usa Gemini, "
                    "con una API key configurada por separado del resto de las funciones."
                )
                motorizacion_ia = st.text_input("Motorización", placeholder="Ej: 1.6 MSI Nafta", key="esq_motorizacion_ia")
                boton_label = "🔄 Generar otra vez" if st.session_state.get("esq_preview_ia") else "🤖 Generar imagen orientativa"
                if st.button(boton_label):
                    if not marca_esq.strip() or not modelo_esq.strip():
                        st.warning("Completá marca y modelo antes de generar.")
                    else:
                        with st.spinner("Generando..."):
                            img_ia, error_ia = generar_esquema_orientativo_ia(
                                marca_esq, modelo_esq, motorizacion_ia,
                                sistema_esq if sistema_opcion == "Otro" else sistema_opcion
                            )
                        if error_ia:
                            st.error(error_ia)
                        else:
                            st.session_state["esq_preview_ia"] = img_ia
                            st.rerun()
                if st.session_state.get("esq_preview_ia"):
                    st.image(st.session_state["esq_preview_ia"], use_container_width=True,
                              caption="Vista previa — orientativo, no es una foto real")
                    imagen_generada_bytes = st.session_state["esq_preview_ia"]

            subir_esq_btn = st.button("📥 Guardar esquema", type="primary")
            if subir_esq_btn:
                imagen_final = archivo_esq.getvalue() if archivo_esq else imagen_generada_bytes
                nombre_final = archivo_esq.name if archivo_esq else "generado_ia.jpg"
                if not titulo_esq.strip() or not imagen_final or not marca_esq.strip() or not modelo_esq.strip():
                    st.warning("Completá título, marca, modelo, y subí o generá una imagen.")
                elif sistema_opcion == "Otro" and not sistema_esq.strip():
                    st.warning("Especificá el sistema/parte.")
                else:
                    guardar_esquema(titulo_esq, marca_esq, modelo_esq, sistema_esq, desc_esq,
                                     imagen_final, nombre_final, generado_ia=(imagen_generada_bytes is not None))
                    st.success("Esquema guardado.")
                    st.session_state.pop("esq_preview_ia", None)
                    for k in ["esq_titulo", "esq_marca_nueva", "esq_modelo_nuevo", "esq_sistema_nuevo",
                              "esq_desc", "esq_motorizacion_ia"]:
                        st.session_state.pop(k, None)
                    olvidar_archivo("esquema")
                    st.rerun()

    # -------- Conversor de unidades --------
    if sub_mec == SUB_MEC[4]:
        st.caption("Conversiones rápidas de unidades que se usan seguido en manuales de taller antiguos o importados.")

        categoria_conv = st.radio("Categoría:", ["Torque", "Presión", "Longitud"], horizontal=True, key="conv_categoria")

        if categoria_conv == "Torque":
            direccion = st.radio("Convertir:", ["lb-ft → Nm", "Nm → lb-ft", "lb-in → Nm", "Nm → lb-in"],
                                  key="conv_torque_dir")
            valor = st.number_input("Valor a convertir:", min_value=0.0, step=0.1, key="conv_torque_valor")
            factores = {
                "lb-ft → Nm": (valor * 1.35582, "Nm"),
                "Nm → lb-ft": (valor / 1.35582, "lb-ft"),
                "lb-in → Nm": (valor * 0.112985, "Nm"),
                "Nm → lb-in": (valor / 0.112985, "lb-in"),
            }
            resultado, unidad = factores[direccion]
            st.metric("Resultado", f"{resultado:.2f} {unidad}")

        elif categoria_conv == "Presión":
            direccion = st.radio("Convertir:", ["PSI → Bar", "Bar → PSI", "PSI → kPa", "kPa → PSI"],
                                  key="conv_presion_dir")
            valor = st.number_input("Valor a convertir:", min_value=0.0, step=0.1, key="conv_presion_valor")
            factores = {
                "PSI → Bar": (valor * 0.0689476, "Bar"),
                "Bar → PSI": (valor / 0.0689476, "PSI"),
                "PSI → kPa": (valor * 6.89476, "kPa"),
                "kPa → PSI": (valor / 6.89476, "PSI"),
            }
            resultado, unidad = factores[direccion]
            st.metric("Resultado", f"{resultado:.2f} {unidad}")

        else:  # Longitud
            direccion = st.radio("Convertir:", ["Pulgadas → mm", "mm → Pulgadas", "Pulgadas → cm", "cm → Pulgadas"],
                                  key="conv_longitud_dir")
            valor = st.number_input("Valor a convertir:", min_value=0.0, step=0.1, key="conv_longitud_valor")
            factores = {
                "Pulgadas → mm": (valor * 25.4, "mm"),
                "mm → Pulgadas": (valor / 25.4, "pulgadas"),
                "Pulgadas → cm": (valor * 2.54, "cm"),
                "cm → Pulgadas": (valor / 2.54, "pulgadas"),
            }
            resultado, unidad = factores[direccion]
            st.metric("Resultado", f"{resultado:.3f} {unidad}")
